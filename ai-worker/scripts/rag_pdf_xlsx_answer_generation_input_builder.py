"""Build diagnostic PDF/XLSX answer-generation inputs.

This script is diagnostic-only. It consumes the row-level answer-shape plan and
review packs, then emits local-LLM input rows under eval/artifacts. It does not
run retrieval, tune retrieval, mutate SearchUnit, update gold CSVs, or create
promotion evidence.
"""

from __future__ import annotations

import argparse
import csv
import hashlib
import json
import re
import shutil
import sys
from collections import Counter
from datetime import datetime, timezone
from pathlib import Path
from typing import Any, Iterable, Mapping


AI_WORKER_ROOT = Path(__file__).resolve().parents[1]
REPO_ROOT = AI_WORKER_ROOT.parent
REPORT_DIR = AI_WORKER_ROOT / "eval" / "reports" / "rag-ingestion"
REVIEW_DIR = AI_WORKER_ROOT / "eval" / "review" / "gold_set_review"
DATASET_DIR = AI_WORKER_ROOT / "eval" / "datasets"
EVAL_RUNS_DIR = AI_WORKER_ROOT / "eval" / "artifacts" / "eval_runs"

DEFAULT_CONTRACT = REPO_ROOT / "docs" / "eval" / "pdf_xlsx_answer_intent_prompt_contract.md"
DEFAULT_PLAN = REPORT_DIR / "rag_pdf_xlsx_answer_shape_diagnostic_plan.json"
DEFAULT_KEYWORD_REVIEW = REPORT_DIR / "rag_pdf_xlsx_keyword_echo_risk_review.csv"
DEFAULT_PDF_REVIEW = REVIEW_DIR / "pdf_gold_review_pack.csv"
DEFAULT_XLSX_REPO_REVIEW = REVIEW_DIR / "xlsx_gold_review_pack.csv"
DEFAULT_XLSX_RETRIEVAL_REPORT = (
    REPORT_DIR / "rag_retrieval_eval_xlsx_v3_positive_reviewed_vector_diagnostic_report.json"
)
DEFAULT_XLSX_FAILURE_BREAKDOWN = REPORT_DIR / "rag_xlsx_v3_after_cleanup_failure_breakdown.json"
DEFAULT_PDF_RETRIEVAL_REPORT = REPORT_DIR / "rag_retrieval_eval_pdf_vector_diagnostic_report.json"
DEFAULT_PDF_C6_REPORT = REPORT_DIR / "rag_pdf_vector_quality_breakdown.json"
DEFAULT_PDF_C7_REPORT = REPORT_DIR / "rag_pdf_gold_policy_review.json"

SCHEMA_VERSION = "rag_pdf_xlsx_answer_generation_inputs_v1"
RUN_PREFIX = "pdf_xlsx_answer_shape_local_llm"
ANSWER_SHAPE_POLICY_PENDING = "NOT_ANSWERABLE_OR_POLICY_PENDING"


def main(argv: list[str] | None = None) -> int:
    args = parse_args(argv)
    report = run_builder(
        contract=Path(args.contract),
        plan=Path(args.plan),
        keyword_review=Path(args.keyword_review),
        pdf_review=Path(args.pdf_review),
        xlsx_external_review=Path(args.xlsx_external_review) if args.xlsx_external_review else None,
        xlsx_repo_review=Path(args.xlsx_repo_review),
        xlsx_retrieval_report=Path(args.xlsx_retrieval_report),
        xlsx_failure_breakdown=Path(args.xlsx_failure_breakdown),
        pdf_retrieval_report=Path(args.pdf_retrieval_report),
        pdf_c6_report=Path(args.pdf_c6_report),
        pdf_c7_report=Path(args.pdf_c7_report),
        output_root=Path(args.output_root),
        run_id=args.run_id,
        max_context_chars=args.max_context_chars,
    )
    print_json(
        {
            "status": report["status"],
            "run_id": report["run_id"],
            "artifact_dir": report["artifact_dir"],
            "manifest": report["manifest_path"],
            "answer_generation_inputs": report["answer_generation_inputs_path"],
            "input_row_count": report["input_row_count"],
            "xlsx_review_pack_used": report["source_review_paths"]["xlsx"],
            "promotion_evidence": report["promotion_evidence"],
            "evidence_role": report["evidence_role"],
        }
    )
    return 0 if report["status"] in {"PASS", "PASS_WITH_WARNINGS"} else 1


def parse_args(argv: list[str] | None) -> argparse.Namespace:
    parser = argparse.ArgumentParser(description=__doc__)
    parser.add_argument("--contract", default=str(DEFAULT_CONTRACT))
    parser.add_argument("--plan", default=str(DEFAULT_PLAN))
    parser.add_argument("--keyword-review", default=str(DEFAULT_KEYWORD_REVIEW))
    parser.add_argument("--pdf-review", default=str(DEFAULT_PDF_REVIEW))
    parser.add_argument(
        "--xlsx-external-review",
        default="",
        help="Optional user-supplied XLSX review pack path. Local paths are redacted from committed artifacts.",
    )
    parser.add_argument("--xlsx-repo-review", default=str(DEFAULT_XLSX_REPO_REVIEW))
    parser.add_argument("--xlsx-retrieval-report", default=str(DEFAULT_XLSX_RETRIEVAL_REPORT))
    parser.add_argument("--xlsx-failure-breakdown", default=str(DEFAULT_XLSX_FAILURE_BREAKDOWN))
    parser.add_argument("--pdf-retrieval-report", default=str(DEFAULT_PDF_RETRIEVAL_REPORT))
    parser.add_argument("--pdf-c6-report", default=str(DEFAULT_PDF_C6_REPORT))
    parser.add_argument("--pdf-c7-report", default=str(DEFAULT_PDF_C7_REPORT))
    parser.add_argument("--output-root", default=str(EVAL_RUNS_DIR))
    parser.add_argument("--run-id", default="")
    parser.add_argument("--max-context-chars", type=int, default=6000)
    return parser.parse_args(argv)


def run_builder(
    *,
    contract: Path,
    plan: Path,
    keyword_review: Path,
    pdf_review: Path,
    xlsx_external_review: Path | None,
    xlsx_repo_review: Path,
    xlsx_retrieval_report: Path,
    xlsx_failure_breakdown: Path,
    pdf_retrieval_report: Path,
    pdf_c6_report: Path,
    pdf_c7_report: Path,
    output_root: Path,
    run_id: str = "",
    max_context_chars: int = 6000,
) -> dict[str, Any]:
    run_id = run_id or utc_run_id()
    generated_at = utc_timestamp()
    artifact_dir = output_root / f"{RUN_PREFIX}_{run_id}"
    manifest_path = artifact_dir / "manifest.json"
    inputs_path = artifact_dir / "answer_generation_inputs.jsonl"

    plan_payload = read_json_object(plan)
    diagnostic_rows = list(plan_payload.get("diagnostic_rows") or []) if plan_payload else []
    if not diagnostic_rows:
        raise SystemExit(f"diagnostic_rows missing in {repo_relative(plan)}")

    warnings: list[str] = []
    xlsx_external_used = bool(xlsx_external_review and xlsx_external_review.exists())
    xlsx_review_used = xlsx_external_review if xlsx_external_used and xlsx_external_review else xlsx_repo_review
    if xlsx_external_review and not xlsx_external_review.exists():
        warnings.append("xlsx external review path was provided but missing; repo-local fallback used")
    pdf_review_rows = keyed_by_query_id(read_csv_rows(pdf_review))
    xlsx_review_rows = keyed_by_query_id(read_csv_rows(xlsx_review_used))
    keyword_rows = keyed_by_query_id(read_csv_rows(keyword_review))
    retrieval_maps = build_retrieval_maps(
        xlsx_retrieval_report=xlsx_retrieval_report,
        xlsx_failure_breakdown=xlsx_failure_breakdown,
        pdf_retrieval_report=pdf_retrieval_report,
        pdf_c6_report=pdf_c6_report,
    )
    dataset_index = build_dataset_index(DATASET_DIR)
    extraction_cache: dict[str, Any] = {}

    input_rows: list[dict[str, Any]] = []
    xlsx_review_snapshot = artifact_dir / "source_xlsx_gold_review_pack_used.csv"
    for row_index, plan_row in enumerate(diagnostic_rows, start=1):
        query_id = clean(plan_row.get("query_id"))
        track = clean(plan_row.get("track")).upper()
        review_row = pdf_review_rows.get(query_id, {}) if track == "PDF" else xlsx_review_rows.get(query_id, {})
        keyword_row = keyword_rows.get(query_id, {})
        merged = merge_row(plan_row, review_row, keyword_row)
        locator = parse_locator(clean(merged.get("expected_current_evidence_location")))
        if not locator:
            locator = locator_from_review_row(merged)
        policy = policy_flags(merged)

        context: dict[str, Any]
        if track == "XLSX":
            context = build_xlsx_context(
                row=merged,
                locator=locator,
                dataset_index=dataset_index,
                cache=extraction_cache,
                retrieval_maps=retrieval_maps,
                policy=policy,
            )
        elif track == "PDF":
            context = build_pdf_context(
                row=merged,
                locator=locator,
                dataset_index=dataset_index,
                cache=extraction_cache,
                retrieval_maps=retrieval_maps,
                policy=policy,
            )
        else:
            context = {
                "context_type": "unknown",
                "context_available": False,
                "context_errors": [f"unknown track {track!r}"],
            }
            warnings.append(f"{query_id}: unknown track {track!r}")

        prompt_context = build_prompt_context(
            row=merged,
            locator=locator,
            context=context,
            policy=policy,
            max_context_chars=max_context_chars,
        )
        input_rows.append(
            {
                "schema_version": SCHEMA_VERSION,
                "run_id": run_id,
                "row_index": row_index,
                "track": track,
                "query_id": query_id,
                "query": clean(merged.get("query")),
                "bucket": clean(merged.get("bucket")),
                "review_group": clean(merged.get("review_group")),
                "expected_answer_shape": clean(merged.get("expected_answer_shape")),
                "future_content_answer_shape_if_unblocked": clean(
                    merged.get("future_content_answer_shape_if_unblocked")
                ),
                "content_target_needed": clean(merged.get("content_target_needed")),
                "expected_evidence_location": locator,
                "expected_current_evidence_location": clean(
                    merged.get("expected_current_evidence_location")
                ),
                "expected_answer_text": clean(merged.get("expected_answer_text")),
                "must_contain_terms": split_terms(merged.get("must_contain_terms")),
                "keyword_echo_risk": clean(merged.get("keyword_echo_risk")),
                "location_only_answer_risk": clean(merged.get("location_only_answer_risk")),
                "answer_eval_allowed": parse_bool(merged.get("answer_eval_allowed")),
                "exclusion_blocker_reason": clean(merged.get("exclusion_blocker_reason")),
                "citation_target_policy": clean(merged.get("citation_target_policy")),
                "source_review_csv": sanitize_source_review_csv(
                    clean(merged.get("source_review_csv")),
                    track=track,
                    xlsx_review_used=xlsx_review_used,
                    xlsx_review_snapshot=xlsx_review_snapshot,
                    xlsx_external_used=xlsx_external_used,
                ),
                "policy": policy,
                "answer_instruction": answer_instruction(merged, policy),
                "context": context,
                "prompt_context": prompt_context,
                "dry_run_preview_used_as_actual_answer": False,
                "local_llm_run": False,
                "external_live_llm_run": False,
                "optional_judge_run": False,
                "promotion_evidence": False,
                "evidence_role": "diagnostic",
            }
        )

    artifact_dir.mkdir(parents=True, exist_ok=True)
    if xlsx_review_used.exists():
        shutil.copy2(xlsx_review_used, xlsx_review_snapshot)
    write_jsonl(inputs_path, input_rows)
    manifest = build_manifest(
        run_id=run_id,
        generated_at=generated_at,
        artifact_dir=artifact_dir,
        manifest_path=manifest_path,
        inputs_path=inputs_path,
        rows=input_rows,
        source_paths={
            "contract": contract,
            "plan": plan,
            "keyword_review": keyword_review,
            "pdf_review": pdf_review,
            "xlsx_review": xlsx_review_snapshot if xlsx_external_used else xlsx_review_used,
            "xlsx_review_snapshot": xlsx_review_snapshot,
            "xlsx_repo_local_fallback": xlsx_repo_review,
            "xlsx_retrieval_report": xlsx_retrieval_report,
            "xlsx_failure_breakdown": xlsx_failure_breakdown,
            "pdf_retrieval_report": pdf_retrieval_report,
            "pdf_c6_report": pdf_c6_report,
            "pdf_c7_report": pdf_c7_report,
        },
        plan_payload=plan_payload,
        warnings=warnings,
        xlsx_external_used=xlsx_external_used,
        xlsx_external_basename=xlsx_external_review.name if xlsx_external_review else "",
    )
    write_json(manifest_path, manifest)
    return manifest


def build_xlsx_context(
    *,
    row: Mapping[str, Any],
    locator: Mapping[str, Any],
    dataset_index: Mapping[str, list[Path]],
    cache: dict[str, Any],
    retrieval_maps: Mapping[str, Mapping[str, Any]],
    policy: Mapping[str, bool],
) -> dict[str, Any]:
    query_id = clean(row.get("query_id"))
    file_name = clean(locator.get("file") or row.get("expected_file_name"))
    sheet_name = clean(locator.get("sheet") or row.get("expected_sheet_name"))
    cell_range = clean(locator.get("range") or row.get("expected_cell_range"))
    table_id = clean(locator.get("table") or row.get("expected_table_id"))
    context: dict[str, Any] = {
        "context_type": "xlsx",
        "file_name": file_name,
        "sheet_name": sheet_name,
        "cell_range": cell_range,
        "table_id": table_id,
        "locator": compact_dict(
            {
                "file": file_name,
                "sheet": sheet_name,
                "range": cell_range,
                "cell": clean(locator.get("cell")),
                "table": table_id,
                "document_version_id": clean(locator.get("docv") or row.get("expected_document_version_id")),
            }
        ),
        "row_label": "",
        "column_label": "",
        "header_context": [],
        "value_context": [],
        "nearby_table_context": [],
        "retrieval_context": retrieval_context_for(query_id, retrieval_maps, "xlsx"),
        "context_available": False,
        "context_has_expected_terms": False,
        "context_errors": [],
    }
    if (
        policy.get("hidden_policy_blocked")
        or policy.get("formula_date_policy_blocked")
        or clean(row.get("expected_answer_shape")) == ANSWER_SHAPE_POLICY_PENDING
    ):
        context["context_errors"].append("content extraction skipped for policy-pending/not-answerable row")
        return context

    source_path = resolve_dataset_file(file_name, dataset_index)
    if source_path is None:
        context["context_errors"].append(f"source workbook not found under {repo_relative(DATASET_DIR)}")
        return context
    context["source_path"] = repo_relative(source_path)
    try:
        extracted = extract_xlsx_context(
            source_path=source_path,
            sheet_name=sheet_name,
            cell_range=cell_range,
            expected_terms=split_terms(row.get("must_contain_terms")) + [clean(row.get("expected_answer_text"))],
            cache=cache,
        )
    except Exception as exc:  # pragma: no cover - diagnostic fallback
        context["context_errors"].append(f"xlsx extraction failed: {type(exc).__name__}: {exc}")
        return context

    context.update(extracted)
    context["context_available"] = bool(
        context.get("nearby_table_context") or context.get("value_context") or context.get("header_context")
    )
    context["context_has_expected_terms"] = has_any_term(
        json.dumps(extracted, ensure_ascii=False), split_terms(row.get("must_contain_terms"))
    )
    return context


def build_pdf_context(
    *,
    row: Mapping[str, Any],
    locator: Mapping[str, Any],
    dataset_index: Mapping[str, list[Path]],
    cache: dict[str, Any],
    retrieval_maps: Mapping[str, Mapping[str, Any]],
    policy: Mapping[str, bool],
) -> dict[str, Any]:
    query_id = clean(row.get("query_id"))
    file_name = clean(locator.get("file") or row.get("expected_file_name"))
    page_no = clean(locator.get("page") or row.get("expected_page_no"))
    bbox = locator.get("bbox") or clean(row.get("expected_bbox"))
    context: dict[str, Any] = {
        "context_type": "pdf",
        "file_name": file_name,
        "page_no": page_no,
        "page_label": clean(locator.get("page_label") or row.get("expected_page_label")),
        "physical_page_index": clean(
            locator.get("physical_page_index") or row.get("expected_physical_page_index")
        ),
        "bbox": bbox,
        "section_id": clean(row.get("expected_section_id")),
        "locator": compact_dict(
            {
                "file": file_name,
                "page": page_no,
                "page_label": clean(locator.get("page_label") or row.get("expected_page_label")),
                "physical_page_index": clean(
                    locator.get("physical_page_index") or row.get("expected_physical_page_index")
                ),
                "bbox": bbox,
                "document_version_id": clean(locator.get("docv") or row.get("expected_document_version_id")),
            }
        ),
        "sentence_context": [],
        "paragraph_context": [],
        "table_or_value_context": [],
        "retrieval_context": retrieval_context_for(query_id, retrieval_maps, "pdf"),
        "context_available": False,
        "context_has_expected_terms": False,
        "context_errors": [],
    }
    if clean(row.get("expected_answer_shape")) == ANSWER_SHAPE_POLICY_PENDING:
        context["context_errors"].append("content extraction skipped for PDF C7 policy-pending row")
        return context

    source_path = resolve_dataset_file(file_name, dataset_index)
    if source_path is None:
        context["context_errors"].append(f"source PDF not found under {repo_relative(DATASET_DIR)}")
        return context
    context["source_path"] = repo_relative(source_path)
    try:
        extracted = extract_pdf_context(
            source_path=source_path,
            page_no=page_no,
            physical_page_index=clean(
                locator.get("physical_page_index") or row.get("expected_physical_page_index")
            ),
            bbox=bbox,
            expected_terms=split_terms(row.get("must_contain_terms")) + [clean(row.get("expected_answer_text"))],
            cache=cache,
        )
    except Exception as exc:  # pragma: no cover - diagnostic fallback
        context["context_errors"].append(f"pdf extraction failed: {type(exc).__name__}: {exc}")
        return context

    context.update(extracted)
    context["context_available"] = bool(
        context.get("sentence_context") or context.get("paragraph_context") or context.get("table_or_value_context")
    )
    context["context_has_expected_terms"] = has_any_term(
        json.dumps(extracted, ensure_ascii=False), split_terms(row.get("must_contain_terms"))
    )
    return context


def extract_xlsx_context(
    *,
    source_path: Path,
    sheet_name: str,
    cell_range: str,
    expected_terms: list[str],
    cache: dict[str, Any],
) -> dict[str, Any]:
    try:
        from openpyxl import load_workbook
        from openpyxl.utils.cell import get_column_letter, range_boundaries
    except Exception as exc:  # pragma: no cover - optional dependency
        return {
            "context_errors": [f"openpyxl unavailable: {type(exc).__name__}: {exc}"],
            "context_available": False,
        }

    cache_key = f"xlsx::{source_path}"
    if cache_key not in cache:
        cache[cache_key] = {
            "formula": load_workbook(source_path, data_only=False, read_only=True),
            "values": load_workbook(source_path, data_only=True, read_only=True),
        }
    workbook_formula = cache[cache_key]["formula"]
    workbook_values = cache[cache_key]["values"]
    if sheet_name not in workbook_formula.sheetnames:
        return {"context_errors": [f"sheet not found: {sheet_name}"]}
    worksheet_formula = workbook_formula[sheet_name]
    worksheet_values = workbook_values[sheet_name]
    if worksheet_formula.sheet_state != "visible":
        return {"context_errors": [f"sheet is not visible: {sheet_name}"]}

    try:
        min_col, min_row, max_col, max_row = range_boundaries(cell_range)
    except Exception:
        min_col, min_row, max_col, max_row = 1, 1, min(worksheet_formula.max_column, 8), min(
            worksheet_formula.max_row, 8
        )

    max_rows = min(max_row, min_row + 7)
    max_cols = min(max_col, min_col + 7)
    header_values = [
        cell_to_text(worksheet_values.cell(row=min_row, column=col).value)
        for col in range(min_col, max_cols + 1)
    ]
    nearby_table: list[str] = []
    value_context: list[dict[str, str]] = []
    row_labels: list[str] = []
    column_labels: list[str] = [value for value in header_values if value]

    for row_index in range(min_row, max_rows + 1):
        if worksheet_formula.row_dimensions[row_index].hidden:
            continue
        row_items: list[str] = []
        first_value = cell_to_text(worksheet_values.cell(row=row_index, column=min_col).value)
        if first_value:
            row_labels.append(first_value)
        for col_index in range(min_col, max_cols + 1):
            column_letter = get_column_letter(col_index)
            if worksheet_formula.column_dimensions[column_letter].hidden:
                continue
            header = header_values[col_index - min_col] or column_letter
            value = cell_to_text(worksheet_values.cell(row=row_index, column=col_index).value)
            formula = cell_to_text(worksheet_formula.cell(row=row_index, column=col_index).value)
            display = value
            if formula.startswith("=") and formula != value:
                display = f"{value} (formula present)" if value else "formula present"
            if display:
                row_items.append(f"{header}: {display}")
                joined = f"{header} {display}"
                if has_any_term(joined, expected_terms):
                    value_context.append(
                        {
                            "cell": f"{column_letter}{row_index}",
                            "row_label": first_value,
                            "column_label": header,
                            "value": value,
                            "formula_present": bool(formula.startswith("=")),
                        }
                    )
        if row_items:
            nearby_table.append(" | ".join(row_items))

    return {
        "row_label": first_nonempty(row_labels),
        "column_label": first_nonempty(column_labels),
        "header_context": column_labels,
        "value_context": value_context[:8],
        "nearby_table_context": nearby_table[:8],
        "xlsx_extraction_policy": "visible_sheets_rows_columns_only_no_raw_formulas",
        "context_errors": [],
    }


def extract_pdf_context(
    *,
    source_path: Path,
    page_no: str,
    physical_page_index: str,
    bbox: object,
    expected_terms: list[str],
    cache: dict[str, Any],
) -> dict[str, Any]:
    try:
        import fitz  # type: ignore
    except Exception as exc:  # pragma: no cover - optional dependency
        return {
            "context_errors": [f"pymupdf unavailable: {type(exc).__name__}: {exc}"],
            "context_available": False,
        }

    cache_key = f"pdf::{source_path}"
    if cache_key not in cache:
        cache[cache_key] = fitz.open(source_path)
    doc = cache[cache_key]
    page_index = int_or_none(physical_page_index)
    if page_index is None:
        page_number = int_or_none(page_no)
        page_index = page_number - 1 if page_number is not None else 0
    if page_index < 0 or page_index >= len(doc):
        return {"context_errors": [f"page index out of bounds: {page_index}"]}

    page = doc[page_index]
    page_text = clean_whitespace(page.get_text("text"))
    bbox_text = ""
    bbox_values = bbox_list(bbox)
    if bbox_values:
        rect = fitz.Rect(*bbox_values)
        bbox_text = clean_whitespace(page.get_text("text", clip=rect))

    term_sentences = select_sentences(page_text, expected_terms)
    table_or_value_context = []
    if bbox_text:
        table_or_value_context.append(bbox_text)
    if not term_sentences and page_text:
        term_sentences = [page_text[:1200]]

    return {
        "sentence_context": term_sentences[:5],
        "paragraph_context": ([bbox_text] if bbox_text else term_sentences[:2]),
        "table_or_value_context": table_or_value_context[:3],
        "page_text_excerpt": page_text[:2000],
        "pdf_extraction_policy": "pymupdf_page_text_and_bbox_clip",
        "context_errors": [],
    }


def build_prompt_context(
    *,
    row: Mapping[str, Any],
    locator: Mapping[str, Any],
    context: Mapping[str, Any],
    policy: Mapping[str, bool],
    max_context_chars: int,
) -> str:
    payload = {
        "task": "diagnostic_pdf_xlsx_answer_shape",
        "query": clean(row.get("query")),
        "expected_answer_shape": clean(row.get("expected_answer_shape")),
        "answer_instruction": answer_instruction(row, policy),
        "content_target_needed": clean(row.get("content_target_needed")),
        "expected_answer_text": clean(row.get("expected_answer_text")),
        "must_contain_terms": split_terms(row.get("must_contain_terms")),
        "expected_evidence_location": locator,
        "citation_policy": clean(row.get("citation_target_policy")),
        "policy": policy,
        "context": context,
    }
    text = json.dumps(payload, ensure_ascii=False, indent=2)
    if max_context_chars > 0 and len(text) > max_context_chars:
        return text[:max_context_chars] + "\n...TRUNCATED_FOR_LOCAL_LLM_INPUT..."
    return text


def answer_instruction(row: Mapping[str, Any], policy: Mapping[str, bool]) -> str:
    shape = clean(row.get("expected_answer_shape"))
    if (
        shape == ANSWER_SHAPE_POLICY_PENDING
        or policy.get("hidden_policy_blocked")
        or policy.get("pdf_c7_policy_pending")
    ):
        return (
            "ABSTAIN_OR_POLICY_PENDING: do not answer with hidden, policy-pending, "
            "or not-answerable content. Fill abstain_reason."
        )
    return (
        "ANSWER_CONTENT_FIRST: provide the content claim/value/summary first; "
        "use page/bbox/sheet/range/cell only as citation evidence."
    )


def build_retrieval_maps(
    *,
    xlsx_retrieval_report: Path,
    xlsx_failure_breakdown: Path,
    pdf_retrieval_report: Path,
    pdf_c6_report: Path,
) -> dict[str, dict[str, Any]]:
    return {
        "xlsx": {
            "query_results": keyed_json_rows(read_json_object(xlsx_retrieval_report), "query_results"),
            "classified_rows": keyed_json_rows(read_json_object(xlsx_failure_breakdown), "classified_query_rows"),
        },
        "pdf": {
            "query_results": keyed_json_rows(read_json_object(pdf_retrieval_report), "query_results"),
            "classified_rows": keyed_json_rows(read_json_object(pdf_c6_report), "classified_query_rows"),
        },
    }


def retrieval_context_for(
    query_id: str,
    retrieval_maps: Mapping[str, Mapping[str, Any]],
    track_key: str,
) -> dict[str, Any]:
    track_maps = retrieval_maps.get(track_key, {})
    query_result = dict(track_maps.get("query_results", {}).get(query_id, {}))
    classified = dict(track_maps.get("classified_rows", {}).get(query_id, {}))
    top_results = list(query_result.get("top_k_results") or [])[:5]
    compact_top = []
    for result in top_results:
        if not isinstance(result, Mapping):
            continue
        location = result.get("location_json") if isinstance(result.get("location_json"), Mapping) else {}
        compact_top.append(
            compact_dict(
                {
                    "rank": result.get("rank"),
                    "score": result.get("score"),
                    "source_file_name": result.get("source_file_name"),
                    "chunk_type": result.get("chunk_type"),
                    "citation_text": result.get("citation_text"),
                    "location_json": location,
                    "match_breakdown": result.get("match_breakdown"),
                }
            )
        )
    return compact_dict(
        {
            "retrieval_query_status": query_result.get("status"),
            "top_k_results": compact_top,
            "failure_or_quality_classification": classified.get("category") or classified.get("classification"),
            "classification_rationale": classified.get("rationale"),
            "top_k_summary": classified.get("top_k_summary"),
            "range_relation_in_top_k": classified.get("range_relation_in_top_k"),
            "supporting_hits": classified.get("supporting_hits"),
        }
    )


def policy_flags(row: Mapping[str, Any]) -> dict[str, bool]:
    blocker = clean(row.get("exclusion_blocker_reason")).upper()
    review_group = clean(row.get("review_group")).lower()
    bucket = clean(row.get("bucket")).lower()
    shape = clean(row.get("expected_answer_shape"))
    return {
        "diagnostic_only": True,
        "pdf_c7_policy_pending": "PDF_C7_POLICY_PENDING" in blocker,
        "pdf_user_review_not_done": "PDF_USER_REVIEW_NOT_DONE" in blocker,
        "xlsx_answer_quality_blocked": "XLSX_ANSWER_QUALITY_BLOCKED" in blocker,
        "hidden_policy_blocked": "hidden_policy" in bucket
        or "HIDDEN_NEGATIVE_POLICY" in blocker
        or "hidden_negative" in review_group,
        "formula_date_policy_blocked": "FORMULA_DATE" in blocker or "formula_date" in review_group,
        "not_answerable_or_policy_pending": shape == ANSWER_SHAPE_POLICY_PENDING,
        "answer_eval_allowed": parse_bool(row.get("answer_eval_allowed")),
        "promotion_evidence": False,
    }


def sanitize_source_review_csv(
    value: str,
    *,
    track: str,
    xlsx_review_used: Path,
    xlsx_review_snapshot: Path,
    xlsx_external_used: bool,
) -> str:
    if track == "XLSX" and xlsx_external_used:
        return repo_relative(xlsx_review_snapshot)
    if not value:
        return ""
    candidate = Path(value)
    if candidate.is_absolute() and not is_under_repo(candidate):
        return f"external_path_redacted:{candidate.name}"
    if candidate == xlsx_review_used and is_under_repo(candidate):
        return repo_relative(candidate)
    return value.replace("\\", "/")


def build_manifest(
    *,
    run_id: str,
    generated_at: str,
    artifact_dir: Path,
    manifest_path: Path,
    inputs_path: Path,
    rows: list[Mapping[str, Any]],
    source_paths: Mapping[str, Path],
    plan_payload: Mapping[str, Any],
    warnings: list[str],
    xlsx_external_used: bool,
    xlsx_external_basename: str,
) -> dict[str, Any]:
    track_counts = Counter(clean(row.get("track")) for row in rows)
    shape_counts = Counter(clean(row.get("expected_answer_shape")) for row in rows)
    context_available_count = sum(1 for row in rows if nested_bool(row, ["context", "context_available"]))
    context_expected_terms_count = sum(
        1 for row in rows if nested_bool(row, ["context", "context_has_expected_terms"])
    )
    source_inventory = {
        key: {
            "path": repo_relative(path),
            "repo_path": repo_relative(path) if is_under_repo(path) else str(path),
            "exists": path.exists(),
            "sha256": sha256_file(path) if path.exists() and path.is_file() else None,
        }
        for key, path in source_paths.items()
    }
    required_false = {
        "promotion_evidence": False,
        "external_live_llm_run": False,
        "optional_judge_run": False,
        "retrieval_tuning_run": False,
        "reranking_run": False,
        "parser_expansion_run": False,
        "threshold_relaxation_run": False,
        "broad_indexing_run": False,
        "db_mutation_run": False,
        "searchunit_mutation_run": False,
        "immutable_baseline_changed": False,
        "candidate_artifact_changed": False,
        "existing_gold_csv_overwritten": False,
    }
    manifest: dict[str, Any] = {
        "schema_version": SCHEMA_VERSION,
        "run_id": run_id,
        "generated_at": generated_at,
        "status": "PASS_WITH_WARNINGS" if warnings else "PASS",
        "artifact_dir": repo_relative(artifact_dir),
        "manifest_path": repo_relative(manifest_path),
        "answer_generation_inputs_path": repo_relative(inputs_path),
        "answer_generation_inputs_sha256": sha256_file(inputs_path),
        "promotion_evidence": False,
        "evidence_role": "diagnostic",
        "local_llm_run": False,
        "external_live_llm_run": False,
        "optional_judge_run": False,
        "dry_run_preview_used_as_actual_answer": False,
        "actual_answer_output_missing": True,
        "input_row_count": len(rows),
        "row_count_by_track": dict(track_counts),
        "expected_answer_shape_counts": dict(shape_counts),
        "context_available_count": context_available_count,
        "context_expected_terms_count": context_expected_terms_count,
        "xlsx_answer_eval_denominator": 0,
        "pdf_answer_eval_denominator": 0,
        "xlsx_retrieval_diagnostic_preserved": True,
        "pdf_policy_pending": True,
        "keyword_echo_only_is_failure": True,
        "location_only_answer_is_failure": True,
        "r8_or_citation_support_blocked_until_answer_shape_alignment": True,
        "source_review_paths": {
            "pdf": repo_relative(source_paths["pdf_review"]),
            "xlsx": repo_relative(source_paths["xlsx_review_snapshot"])
            if xlsx_external_used
            else repo_relative(source_paths["xlsx_review"]),
            "xlsx_snapshot": repo_relative(source_paths["xlsx_review_snapshot"]),
            "xlsx_source_kind": "external_user_supplied_snapshot" if xlsx_external_used else "repo_local",
            "xlsx_external_path_redacted": xlsx_external_used,
            "xlsx_external_basename": xlsx_external_basename,
            "xlsx_repo_local_fallback": repo_relative(source_paths["xlsx_repo_local_fallback"]),
            "xlsx_external_used": xlsx_external_used,
        },
        "source_inventory": source_inventory,
        "plan_source_status": plan_payload.get("status"),
        "plan_actual_answer_output_missing": plan_payload.get("actual_answer_output_missing"),
        "plan_assertions": plan_payload.get("assertions", {}),
        "guardrails": required_false,
        **required_false,
        "warnings": warnings,
    }
    return manifest


def build_dataset_index(root: Path) -> dict[str, list[Path]]:
    index: dict[str, list[Path]] = {}
    if not root.exists():
        return index
    for suffix in ("*.xlsx", "*.xls", "*.pdf"):
        for path in root.rglob(suffix):
            index.setdefault(path.name, []).append(path)
    return index


def resolve_dataset_file(file_name: str, dataset_index: Mapping[str, list[Path]]) -> Path | None:
    if not file_name:
        return None
    candidates = dataset_index.get(file_name) or []
    return candidates[0] if candidates else None


def merge_row(*rows: Mapping[str, Any]) -> dict[str, Any]:
    merged: dict[str, Any] = {}
    for row in rows:
        for key, value in row.items():
            if value not in (None, "") or key not in merged:
                merged[key] = value
    return merged


def parse_locator(text: str) -> dict[str, Any]:
    locator: dict[str, Any] = {}
    for part in text.split(";"):
        if "=" not in part:
            continue
        key, value = part.split("=", 1)
        key = clean(key)
        value = clean(value)
        if key == "bbox":
            locator[key] = bbox_list(value) or value
        else:
            locator[key] = value
    return locator


def locator_from_review_row(row: Mapping[str, Any]) -> dict[str, Any]:
    return compact_dict(
        {
            "file": row.get("expected_file_name"),
            "docv": row.get("expected_document_version_id"),
            "page": row.get("expected_page_no"),
            "physical_page_index": row.get("expected_physical_page_index"),
            "page_label": row.get("expected_page_label"),
            "sheet": row.get("expected_sheet_name"),
            "range": row.get("expected_cell_range"),
            "table": row.get("expected_table_id"),
            "bbox": bbox_list(row.get("expected_bbox")) or clean(row.get("expected_bbox")),
        }
    )


def keyed_by_query_id(rows: Iterable[Mapping[str, Any]]) -> dict[str, dict[str, Any]]:
    return {clean(row.get("query_id")): dict(row) for row in rows if clean(row.get("query_id"))}


def keyed_json_rows(payload: Mapping[str, Any] | None, field: str) -> dict[str, dict[str, Any]]:
    if not payload:
        return {}
    rows = payload.get(field)
    if not isinstance(rows, list):
        return {}
    return keyed_by_query_id(row for row in rows if isinstance(row, Mapping))


def read_json_object(path: Path) -> dict[str, Any] | None:
    if not path.exists():
        return None
    try:
        payload = json.loads(path.read_text(encoding="utf-8"))
    except json.JSONDecodeError:
        return None
    return payload if isinstance(payload, dict) else None


def read_csv_rows(path: Path) -> list[dict[str, str]]:
    if not path.exists():
        return []
    with path.open("r", encoding="utf-8-sig", newline="") as handle:
        return [dict(row) for row in csv.DictReader(handle)]


def write_json(path: Path, payload: Mapping[str, Any]) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    path.write_text(json.dumps(payload, ensure_ascii=False, indent=2) + "\n", encoding="utf-8")


def write_jsonl(path: Path, rows: Iterable[Mapping[str, Any]]) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    with path.open("w", encoding="utf-8", newline="\n") as handle:
        for row in rows:
            handle.write(json.dumps(row, ensure_ascii=False, sort_keys=True) + "\n")


def print_json(payload: Mapping[str, Any]) -> None:
    print(json.dumps(payload, ensure_ascii=False, indent=2))


def sha256_file(path: Path) -> str:
    digest = hashlib.sha256()
    with path.open("rb") as handle:
        for chunk in iter(lambda: handle.read(1024 * 1024), b""):
            digest.update(chunk)
    return digest.hexdigest()


def split_terms(value: object) -> list[str]:
    text = clean(value)
    if not text:
        return []
    terms = [clean(item) for item in re.split(r"[;|]", text)]
    return [term for term in terms if term]


def has_any_term(text: str, terms: Iterable[str]) -> bool:
    folded = normalize_for_match(text)
    return any(normalize_for_match(term) in folded for term in terms if clean(term))


def select_sentences(text: str, terms: Iterable[str]) -> list[str]:
    sentences = [clean(item) for item in re.split(r"(?<=[.!?。！？])\s+|\n+", text) if clean(item)]
    selected = [sentence for sentence in sentences if has_any_term(sentence, terms)]
    if selected:
        return selected
    folded_terms = [normalize_for_match(term) for term in terms if clean(term)]
    fallback = []
    for sentence in sentences:
        folded = normalize_for_match(sentence)
        if any(term and term in folded for term in folded_terms):
            fallback.append(sentence)
    return fallback


def normalize_for_match(value: object) -> str:
    return re.sub(r"\s+", "", clean(value)).lower()


def clean_whitespace(value: object) -> str:
    return re.sub(r"\s+", " ", clean(value)).strip()


def cell_to_text(value: object) -> str:
    if value is None:
        return ""
    return clean(value)


def first_nonempty(values: Iterable[str]) -> str:
    for value in values:
        if clean(value):
            return clean(value)
    return ""


def bbox_list(value: object) -> list[float]:
    if isinstance(value, list):
        try:
            return [float(item) for item in value]
        except (TypeError, ValueError):
            return []
    text = clean(value)
    if not text:
        return []
    numbers = re.findall(r"-?\d+(?:\.\d+)?", text)
    if len(numbers) != 4:
        return []
    return [float(number) for number in numbers]


def int_or_none(value: object) -> int | None:
    text = clean(value)
    if not text:
        return None
    try:
        return int(float(text))
    except ValueError:
        return None


def parse_bool(value: object) -> bool:
    if isinstance(value, bool):
        return value
    return clean(value).lower() in {"1", "true", "yes", "y"}


def compact_dict(payload: Mapping[str, Any]) -> dict[str, Any]:
    return {key: value for key, value in payload.items() if value not in (None, "", [], {})}


def nested_bool(row: Mapping[str, Any], path: list[str]) -> bool:
    current: Any = row
    for key in path:
        if not isinstance(current, Mapping):
            return False
        current = current.get(key)
    return bool(current)


def clean(value: object) -> str:
    if value is None:
        return ""
    if isinstance(value, str):
        return value.strip()
    return str(value).strip()


def is_under_repo(path: Path) -> bool:
    try:
        path.resolve().relative_to(REPO_ROOT.resolve())
        return True
    except ValueError:
        return False


def repo_relative(path: Path) -> str:
    try:
        return str(path.resolve().relative_to(REPO_ROOT.resolve())).replace("\\", "/")
    except ValueError:
        return str(path)


def utc_timestamp() -> str:
    return datetime.now(timezone.utc).isoformat()


def utc_run_id() -> str:
    return datetime.now(timezone.utc).strftime("%Y%m%dT%H%M%SZ")


if __name__ == "__main__":
    sys.exit(main())
