"""Normalize the human-reviewed XLSX gold pack into strict denominator lanes.

The source review pack is treated as human-owned input. This script preserves
the raw labels, adds derived evaluator fields, and emits a separate official
positive subset so mixed diagnostic/excluded rows cannot silently inflate an
official denominator.
"""

from __future__ import annotations

import argparse
import csv
import hashlib
import json
import re
import shutil
import sys
from collections import Counter
from datetime import datetime, timezone
from pathlib import Path
from types import SimpleNamespace
from typing import Any, Iterable, Mapping


AI_WORKER_ROOT = Path(__file__).resolve().parents[1]
REPO_ROOT = AI_WORKER_ROOT.parent
DEFAULT_OUTPUT = AI_WORKER_ROOT / "eval" / "eval_queries" / "gold_queries_xlsx_human_review_normalized_v0.csv"
DEFAULT_OFFICIAL_OUTPUT = (
    AI_WORKER_ROOT / "eval" / "eval_queries" / "gold_queries_xlsx_human_review_official_positive_v0.csv"
)
DEFAULT_OFFICIAL_RETRIEVAL_OUTPUT = (
    AI_WORKER_ROOT
    / "eval"
    / "eval_queries"
    / "gold_queries_xlsx_human_review_official_positive_v0_retrieval.csv"
)
DEFAULT_JSONL_OUTPUT = (
    AI_WORKER_ROOT / "eval" / "eval_queries" / "gold_queries_xlsx_human_review_normalized_v0.jsonl"
)
DEFAULT_REPORT = AI_WORKER_ROOT / "eval" / "reports" / "rag-ingestion" / "rag_xlsx_human_review_gold_normalization_report.json"
DEFAULT_ARTIFACT_ROOT = AI_WORKER_ROOT / "eval" / "artifacts" / "eval_runs"
DEFAULT_DATASET_ROOT = AI_WORKER_ROOT / "eval" / "datasets" / "xlsx"
DEFAULT_REGISTRY = AI_WORKER_ROOT / "eval" / "eval_queries" / "official_denominator_registry.json"

if str(AI_WORKER_ROOT) not in sys.path:
    sys.path.insert(0, str(AI_WORKER_ROOT))

from eval.harness.rag_ingestion_retrieval_eval import (  # noqa: E402
    REQUIRED_COLUMNS as RETRIEVAL_REQUIRED_COLUMNS,
    validate_gold_rows as validate_retrieval_gold_rows,
)

SCHEMA_VERSION = "xlsx_human_review_gold_normalized_v1"

REQUIRED_COLUMNS = [
    "query_id",
    "query",
    "track",
    "sheet",
    "range",
    "citation_locator",
    "evidence_summary",
    "evidence_headers",
    "evidence_row_values",
    "evidence_cell_values",
    "deterministic_compiled_answer",
    "deterministic_compiled_status",
    "expected_answer_text_existing",
    "must_contain_terms_existing",
    "user_answerability_label",
    "user_relevance_label",
    "user_gold_answer_shape",
    "user_expected_answer_text",
    "user_required_citation_policy",
    "user_gold_policy_decision",
    "user_include_in_official_denominator",
]

ANSWERABILITY_LABELS = {
    "ANSWERABLE_CONFIRMED",
    "ANSWERABLE_NEEDS_SOURCE_VERIFICATION",
    "NOT_ANSWERABLE",
}
RELEVANCE_LABELS = {
    "EVIDENCE_RELEVANT",
    "EVIDENCE_PARTIAL",
    "EVIDENCE_MISMATCH",
    "POLICY_EXCLUDED",
}
GOLD_ANSWER_SHAPES = {
    "RANGE_LOCATION_SUMMARY",
    "ROW_SUMMARY",
    "CELL_VALUE",
    "FORMULA_VALUE",
    "NOT_ANSWERABLE_OR_POLICY_PENDING",
    "AGGREGATION_RESULT",
}
CITATION_POLICIES = {
    "SHEET_RANGE_WITH_EXAMPLES",
    "EXACT_ROW",
    "EXACT_CELL",
    "POLICY_EXCLUDED",
    "ROW_GROUP_RANGE",
    "EXACT_CELL, EXACT_ROW",
    "TABLE_RANGE",
}

SOURCE_TEXT_FIELDS = [
    "evidence_summary",
    "evidence_headers",
    "evidence_row_values",
    "evidence_cell_values",
]

DERIVED_COLUMNS = [
    "normalized_schema_version",
    "raw_human_labels_preserved",
    "normalized_answerability_label",
    "normalized_relevance_label",
    "normalized_gold_answer_shape",
    "normalized_expected_answer_text",
    "normalized_expected_answer_source",
    "normalized_must_contain_terms_json",
    "citation_locator_parse_status",
    "citation_locator_sheet",
    "citation_locator_range",
    "citation_locator_document_version_id",
    "citation_locator_file",
    "locator_contract_valid",
    "expected_answer_contract_valid",
    "must_contain_terms_contract_valid",
    "evidence_terms_validation_status",
    "workbook_source_validation_status",
    "source_validation_status",
    "source_validation_basis",
    "llm_answer_used_for_source_validation",
    "denominator_kind",
    "not_answer_generation_denominator",
    "derived_denominator_policy",
    "derived_gold_status",
    "include_in_official_positive_denominator",
    "derived_policy_reasons",
    "official_positive_reason",
]

OFFICIAL_POLICY = "OFFICIAL_POSITIVE"
DIAGNOSTIC_POLICY = "DIAGNOSTIC_ONLY"
EXCLUDED_POLICY = "EXCLUDED"
PENDING_POLICY = "PENDING_SOURCE_VERIFICATION"


def main(argv: list[str] | None = None) -> int:
    args = parse_args(argv)
    report = run_normalization(
        review_pack=Path(args.review_pack),
        output=Path(args.output),
        official_output=Path(args.official_output),
        official_retrieval_output=Path(args.official_retrieval_output),
        jsonl_output=Path(args.jsonl_output),
        report_path=Path(args.report),
        artifact_root=Path(args.artifact_root),
        dataset_root=Path(args.dataset_root),
        registry_path=Path(args.registry),
        expected_row_count=args.expected_row_count,
        run_id=args.run_id or utc_run_id(),
        source_label=args.source_label,
        update_registry=args.update_registry,
    )
    print_json(
        {
            "status": report["status"],
            "normalized_artifact": report["outputs"]["normalized_csv"]["path"],
            "official_positive_artifact": report["outputs"]["official_positive_csv"]["path"],
            "official_positive_retrieval_artifact": report["outputs"]["official_positive_retrieval_csv"]["path"],
            "report": report["outputs"]["report"]["path"],
            "total_rows": report["total_rows"],
            "official_positive_count": report["official_positive_count"],
            "official_positive_retrieval_count": report["official_positive_retrieval_count"],
            "diagnostic_only_count": report["diagnostic_only_count"],
            "pending_source_verification_count": report["pending_source_verification_count"],
            "excluded_count": report["excluded_count"],
            "validation_error_count": len(report["validation_errors"]),
        }
    )
    return 0 if report["status"] == "PASS" else 2


def parse_args(argv: list[str] | None) -> argparse.Namespace:
    parser = argparse.ArgumentParser(description=__doc__)
    parser.add_argument("--review-pack", required=True, help="Human-reviewed XLSX CSV export to normalize.")
    parser.add_argument("--output", default=str(DEFAULT_OUTPUT))
    parser.add_argument("--official-output", default=str(DEFAULT_OFFICIAL_OUTPUT))
    parser.add_argument("--official-retrieval-output", default=str(DEFAULT_OFFICIAL_RETRIEVAL_OUTPUT))
    parser.add_argument("--jsonl-output", default=str(DEFAULT_JSONL_OUTPUT))
    parser.add_argument("--report", default=str(DEFAULT_REPORT))
    parser.add_argument("--artifact-root", default=str(DEFAULT_ARTIFACT_ROOT))
    parser.add_argument("--dataset-root", default=str(DEFAULT_DATASET_ROOT))
    parser.add_argument("--registry", default=str(DEFAULT_REGISTRY))
    parser.add_argument("--expected-row-count", type=int, default=50)
    parser.add_argument("--run-id", default="")
    parser.add_argument("--source-label", default="")
    parser.add_argument(
        "--update-registry",
        action="store_true",
        help="Update official_denominator_registry.json with the normalized XLSX entry.",
    )
    return parser.parse_args(argv)


def run_normalization(
    *,
    review_pack: Path,
    output: Path,
    official_output: Path,
    jsonl_output: Path,
    report_path: Path,
    artifact_root: Path,
    dataset_root: Path,
    registry_path: Path,
    expected_row_count: int,
    run_id: str,
    official_retrieval_output: Path | None = None,
    source_label: str = "",
    update_registry: bool = False,
) -> dict[str, Any]:
    source_rows = read_csv_rows(review_pack)
    source_sha = sha256_file(review_pack)
    source_columns = list(source_rows[0].keys()) if source_rows else []
    validation_errors = validate_source_rows(source_rows, expected_row_count=expected_row_count)
    workbook_cache = WorkbookCache(dataset_root)

    normalized_rows = [normalize_row(row, workbook_cache=workbook_cache) for row in source_rows]
    official_rows = [
        row for row in normalized_rows if row["include_in_official_positive_denominator"].upper() == "TRUE"
    ]
    official_validation_errors = validate_official_rows(official_rows)
    validation_errors.extend(official_validation_errors)
    official_retrieval_output = official_retrieval_output or official_output.with_name(
        f"{official_output.stem}_retrieval.csv"
    )
    official_retrieval_rows = project_official_rows_to_retrieval_gold(official_rows)
    official_retrieval_validation = (
        validate_retrieval_gold_rows(official_retrieval_rows, require_live_bound=False)
        if official_retrieval_rows
        else SimpleNamespace(ok=True, row_count=0, errors=[], row_errors={}, bucket_counts={})
    )
    validation_errors.extend(
        f"official retrieval projection: {error}" for error in official_retrieval_validation.errors
    )

    output_columns = [*source_columns, *[column for column in DERIVED_COLUMNS if column not in source_columns]]

    artifact_dir = artifact_root / f"xlsx_human_review_gold_normalization_{run_id}"
    artifact_dir.mkdir(parents=True, exist_ok=True)
    source_snapshot = artifact_dir / "source_xlsx_gold_human_review_pack_used.csv"
    if review_pack.resolve() != source_snapshot.resolve():
        shutil.copyfile(review_pack, source_snapshot)

    outputs_written = not validation_errors
    if outputs_written:
        write_csv(output, output_columns, normalized_rows)
        write_csv(official_output, output_columns, official_rows)
        write_csv(official_retrieval_output, RETRIEVAL_REQUIRED_COLUMNS, official_retrieval_rows)
        write_jsonl(jsonl_output, normalized_rows)

    report = build_report(
        run_id=run_id,
        source_label=source_label or review_pack.name,
        source_sha=source_sha,
        source_snapshot=source_snapshot,
        normalized_rows=normalized_rows,
        official_rows=official_rows,
        official_retrieval_rows=official_retrieval_rows,
        official_retrieval_validation=official_retrieval_validation,
        output=output,
        official_output=official_output,
        official_retrieval_output=official_retrieval_output,
        jsonl_output=jsonl_output,
        report_path=report_path,
        registry_path=registry_path,
        validation_errors=validation_errors,
        workbook_cache=workbook_cache,
    )
    report["outputs_written"] = outputs_written
    if not outputs_written:
        report["write_skipped_reason"] = "validation_errors"
    if update_registry and report["status"] == "PASS" and outputs_written:
        update_denominator_registry(registry_path=registry_path, report=report)
        report["registry_updated"] = True
        report["outputs"]["registry"]["sha256"] = sha256_file(registry_path)
    write_json(report_path, report)
    return report


def normalize_row(row: Mapping[str, str], *, workbook_cache: "WorkbookCache") -> dict[str, Any]:
    out = dict(row)
    answerability = clean(row.get("user_answerability_label"))
    relevance = clean(row.get("user_relevance_label"))
    shape = clean(row.get("user_gold_answer_shape"))
    citation_policy = clean(row.get("user_required_citation_policy"))
    locator, locator_errors = parse_locator(row.get("citation_locator"))
    sheet = clean(row.get("sheet"))
    cell_range = clean(row.get("range"))
    terms = parse_terms(row.get("must_contain_terms_existing"))
    normalized_expected, expected_source = normalized_expected_answer(row, terms)
    evidence_text = bound_evidence_text(row)

    reasons: list[str] = []
    if locator_errors:
        reasons.extend(locator_errors)
    if not locator:
        reasons.append("invalid_or_empty_citation_locator")
    if not sheet:
        reasons.append("missing_sheet")
    if not cell_range:
        reasons.append("missing_range")
    if locator:
        locator_sheet = clean(locator.get("sheet"))
        locator_range = clean(locator.get("range"))
        if locator_sheet and sheet and locator_sheet != sheet:
            reasons.append("sheet_locator_mismatch")
        if locator_range and cell_range and locator_range != cell_range:
            reasons.append("range_locator_mismatch")
    if not normalized_expected:
        reasons.append("missing_expected_answer")
    if not terms:
        reasons.append("missing_must_contain_terms")

    evidence_missing_terms = [term for term in terms if not term_present(term, evidence_text)]
    evidence_status = "PASS" if terms and not evidence_missing_terms else "FAIL_MISSING_TERMS"
    if not terms:
        evidence_status = "SKIPPED_NO_TERMS"

    workbook_status = workbook_cache.validate_terms(
        file_name=clean(locator.get("file")) if locator else "",
        sheet=sheet,
        cell_range=cell_range,
        terms=terms,
    )
    source_validation_ok = evidence_status == "PASS" or workbook_status == "PASS"
    source_validation_basis = ""
    if evidence_status == "PASS":
        source_validation_basis = "bound_evidence"
    elif workbook_status == "PASS":
        source_validation_basis = "workbook"
    if terms and not source_validation_ok:
        reasons.append("must_contain_terms_not_in_bound_evidence")

    policy, policy_reasons = derive_policy(
        answerability=answerability,
        relevance=relevance,
        base_reasons=reasons,
        source_validation_ok=source_validation_ok,
    )
    all_reasons = dedupe([*reasons, *policy_reasons])

    locator_valid = not any(
        reason
        in {
            "invalid_citation_locator_json",
            "citation_locator_not_object",
            "invalid_or_empty_citation_locator",
            "missing_sheet",
            "missing_range",
            "sheet_locator_mismatch",
            "range_locator_mismatch",
        }
        for reason in all_reasons
    )
    expected_valid = bool(normalized_expected)
    terms_valid = bool(terms)

    out.update(
        {
            "normalized_schema_version": SCHEMA_VERSION,
            "raw_human_labels_preserved": "TRUE",
            "normalized_answerability_label": answerability,
            "normalized_relevance_label": relevance,
            "normalized_gold_answer_shape": shape,
            "normalized_expected_answer_text": normalized_expected,
            "normalized_expected_answer_source": expected_source,
            "normalized_must_contain_terms_json": json.dumps(terms, ensure_ascii=False, separators=(",", ":")),
            "citation_locator_parse_status": "PASS" if not locator_errors else "FAIL",
            "citation_locator_sheet": clean(locator.get("sheet")) if locator else "",
            "citation_locator_range": clean(locator.get("range")) if locator else "",
            "citation_locator_document_version_id": clean(locator.get("document_version_id")) if locator else "",
            "citation_locator_file": clean(locator.get("file")) if locator else "",
            "locator_contract_valid": bool_text(locator_valid),
            "expected_answer_contract_valid": bool_text(expected_valid),
            "must_contain_terms_contract_valid": bool_text(terms_valid),
            "evidence_terms_validation_status": evidence_status,
            "workbook_source_validation_status": workbook_status,
            "source_validation_status": "PASS" if source_validation_ok else "FAIL",
            "source_validation_basis": source_validation_basis,
            "llm_answer_used_for_source_validation": "FALSE",
            "denominator_kind": "xlsx_retrieval_evidence_diagnostic",
            "not_answer_generation_denominator": "TRUE",
            "derived_denominator_policy": policy,
            "derived_gold_status": "gold" if policy == OFFICIAL_POLICY else policy.lower(),
            "include_in_official_positive_denominator": bool_text(policy == OFFICIAL_POLICY),
            "derived_policy_reasons": ";".join(all_reasons),
            "official_positive_reason": official_reason(row, expected_source) if policy == OFFICIAL_POLICY else "",
        }
    )
    out["_citation_policy_for_validation"] = citation_policy
    return out


def derive_policy(
    *,
    answerability: str,
    relevance: str,
    base_reasons: list[str],
    source_validation_ok: bool,
) -> tuple[str, list[str]]:
    if relevance == "EVIDENCE_MISMATCH":
        return EXCLUDED_POLICY, ["human_evidence_mismatch"]
    if relevance == "POLICY_EXCLUDED":
        return EXCLUDED_POLICY, ["human_policy_excluded"]
    if answerability == "NOT_ANSWERABLE":
        return EXCLUDED_POLICY, ["human_not_answerable", "no_abstain_negative_denominator_configured"]
    if answerability == "ANSWERABLE_NEEDS_SOURCE_VERIFICATION":
        if relevance == "EVIDENCE_RELEVANT" and not base_reasons and source_validation_ok:
            return OFFICIAL_POLICY, ["source_verified_from_repo_artifacts"]
        return PENDING_POLICY, ["human_needs_source_verification"]
    if relevance == "EVIDENCE_PARTIAL":
        return DIAGNOSTIC_POLICY, ["human_evidence_partial", "no_relaxed_xlsx_denominator_policy"]
    if answerability == "ANSWERABLE_CONFIRMED" and relevance == "EVIDENCE_RELEVANT":
        if base_reasons or not source_validation_ok:
            return DIAGNOSTIC_POLICY, ["strict_official_contract_not_satisfied"]
        return OFFICIAL_POLICY, ["strict_official_contract_satisfied"]
    return DIAGNOSTIC_POLICY, ["no_official_policy_rule_matched"]


def normalized_expected_answer(row: Mapping[str, str], terms: list[str]) -> tuple[str, str]:
    user_value = clean(row.get("user_expected_answer_text"))
    if user_value:
        return user_value, "user_expected_answer_text"
    existing = clean(row.get("expected_answer_text_existing"))
    if existing and all(term_present(term, existing) for term in terms):
        return existing, "expected_answer_text_existing"
    if terms:
        return " ".join(terms), "must_contain_terms_existing_fallback"
    if existing:
        return existing, "expected_answer_text_existing"
    compiled = clean(row.get("deterministic_compiled_answer"))
    if compiled:
        return compiled, "deterministic_compiled_answer"
    return "", ""


def official_reason(row: Mapping[str, str], expected_source: str) -> str:
    compiled_status = clean(row.get("deterministic_compiled_status"))
    compiled_answer = clean(row.get("deterministic_compiled_answer"))
    if compiled_answer:
        return "human_confirmed_relevant_with_compiled_answer_and_valid_locator"
    return (
        "human_confirmed_relevant_with_valid_locator_and_normalized_expected_answer"
        f";compiled_status={compiled_status};expected_source={expected_source}"
    )


def validate_source_rows(rows: list[dict[str, str]], *, expected_row_count: int) -> list[str]:
    errors: list[str] = []
    if len(rows) != expected_row_count:
        errors.append(f"row count {len(rows)} != expected {expected_row_count}")
    if not rows:
        errors.append("review pack is empty")
        return errors
    columns = set(rows[0].keys())
    missing = [column for column in REQUIRED_COLUMNS if column not in columns]
    if missing:
        errors.append("missing required columns: " + ", ".join(missing))
        return errors

    ids = [clean(row.get("query_id")) for row in rows]
    duplicate_ids = sorted(query_id for query_id, count in Counter(ids).items() if query_id and count > 1)
    if duplicate_ids:
        errors.append("duplicate query_id values: " + ", ".join(duplicate_ids))
    blank_ids = sum(1 for query_id in ids if not query_id)
    if blank_ids:
        errors.append(f"blank query_id count: {blank_ids}")
    non_xlsx = sorted({clean(row.get("track")) for row in rows if clean(row.get("track")).upper() != "XLSX"})
    if non_xlsx:
        errors.append("non-XLSX track values: " + ", ".join(non_xlsx))

    for index, row in enumerate(rows, start=2):
        row_id = clean(row.get("query_id")) or f"row_{index}"
        answerability = clean(row.get("user_answerability_label"))
        relevance = clean(row.get("user_relevance_label"))
        shape = clean(row.get("user_gold_answer_shape"))
        citation_policy = clean(row.get("user_required_citation_policy"))
        _, locator_errors = parse_locator(row.get("citation_locator"))
        if answerability not in ANSWERABILITY_LABELS:
            errors.append(f"{row_id}: invalid user_answerability_label {answerability!r}")
        if relevance not in RELEVANCE_LABELS:
            errors.append(f"{row_id}: invalid user_relevance_label {relevance!r}")
        if shape not in GOLD_ANSWER_SHAPES:
            errors.append(f"{row_id}: invalid user_gold_answer_shape {shape!r}")
        if citation_policy not in CITATION_POLICIES:
            errors.append(f"{row_id}: invalid user_required_citation_policy {citation_policy!r}")
        for locator_error in locator_errors:
            errors.append(f"{row_id}: {locator_error}")
    return errors


def validate_official_rows(rows: list[Mapping[str, Any]]) -> list[str]:
    errors: list[str] = []
    for row in rows:
        query_id = clean(row.get("query_id"))
        if clean(row.get("citation_locator")) in {"", "{}"}:
            errors.append(f"{query_id}: official row has empty citation_locator")
        if clean(row.get("locator_contract_valid")).upper() != "TRUE":
            errors.append(f"{query_id}: official row has invalid locator contract")
        if not clean(row.get("sheet")) or not clean(row.get("range")):
            errors.append(f"{query_id}: official row must have sheet and range")
        if not clean(row.get("normalized_expected_answer_text")):
            errors.append(f"{query_id}: official row must have normalized expected answer")
        if not parse_terms(row.get("normalized_must_contain_terms_json")):
            errors.append(f"{query_id}: official row must have must_contain_terms")
        if clean(row.get("source_validation_status")) != "PASS":
            errors.append(f"{query_id}: official row must pass source validation")
        if clean(row.get("source_validation_basis")) not in {"bound_evidence", "workbook"}:
            errors.append(f"{query_id}: official row must record source_validation_basis")
        if clean(row.get("llm_answer_used_for_source_validation")).upper() != "FALSE":
            errors.append(f"{query_id}: official row cannot use llm_answer for source validation")
        if clean(row.get("denominator_kind")) != "xlsx_retrieval_evidence_diagnostic":
            errors.append(f"{query_id}: official row has unexpected denominator_kind")
        if clean(row.get("not_answer_generation_denominator")).upper() != "TRUE":
            errors.append(f"{query_id}: official row must not enter answer-generation denominator")
        if clean(row.get("user_relevance_label")) in {"EVIDENCE_MISMATCH", "POLICY_EXCLUDED"}:
            errors.append(f"{query_id}: official row cannot have excluded relevance label")
        if clean(row.get("user_answerability_label")) == "NOT_ANSWERABLE":
            errors.append(f"{query_id}: official positive row cannot be NOT_ANSWERABLE")
    return errors


def build_report(
    *,
    run_id: str,
    source_label: str,
    source_sha: str,
    source_snapshot: Path,
    normalized_rows: list[Mapping[str, Any]],
    official_rows: list[Mapping[str, Any]],
    official_retrieval_rows: list[Mapping[str, Any]],
    official_retrieval_validation: Any,
    output: Path,
    official_output: Path,
    official_retrieval_output: Path,
    jsonl_output: Path,
    report_path: Path,
    registry_path: Path,
    validation_errors: list[str],
    workbook_cache: "WorkbookCache",
) -> dict[str, Any]:
    policy_counts = Counter(clean(row.get("derived_denominator_policy")) for row in normalized_rows)
    rows_excluded_despite_relevant = [
        {
            "query_id": row.get("query_id"),
            "derived_denominator_policy": row.get("derived_denominator_policy"),
            "derived_policy_reasons": row.get("derived_policy_reasons"),
            "sheet": row.get("sheet"),
            "range": row.get("range"),
        }
        for row in normalized_rows
        if row.get("user_relevance_label") == "EVIDENCE_RELEVANT"
        and row.get("derived_denominator_policy") != OFFICIAL_POLICY
    ]
    report = {
        "schema_version": SCHEMA_VERSION,
        "status": "FAIL" if validation_errors else "PASS",
        "run_id": run_id,
        "generated_at": utc_timestamp(),
        "source_review_pack": {
            "label": source_label,
            "sha256": source_sha,
            "row_count": len(normalized_rows),
            "snapshot_path": repo_relative(source_snapshot),
        },
        "total_rows": len(normalized_rows),
        "official_positive_count": policy_counts.get(OFFICIAL_POLICY, 0),
        "official_positive_retrieval_count": len(official_retrieval_rows),
        "official_xlsx_answer_generation_denominator": 0,
        "diagnostic_only_count": policy_counts.get(DIAGNOSTIC_POLICY, 0),
        "pending_source_verification_count": policy_counts.get(PENDING_POLICY, 0),
        "excluded_count": policy_counts.get(EXCLUDED_POLICY, 0),
        "answerability_distribution": dict(Counter(row.get("user_answerability_label") for row in normalized_rows)),
        "relevance_distribution": dict(Counter(row.get("user_relevance_label") for row in normalized_rows)),
        "gold_answer_shape_distribution": dict(Counter(row.get("user_gold_answer_shape") for row in normalized_rows)),
        "citation_policy_distribution": dict(Counter(row.get("user_required_citation_policy") for row in normalized_rows)),
        "derived_denominator_policy_distribution": dict(policy_counts),
        "official_positive_query_ids": [row.get("query_id") for row in official_rows],
        "official_positive_retrieval_query_ids": [row.get("query_id") for row in official_retrieval_rows],
        "rows_excluded_despite_human_relevant_labels": rows_excluded_despite_relevant,
        "special_rows": special_row_summary(normalized_rows),
        "official_retrieval_projection_validation": {
            "ok": official_retrieval_validation.ok,
            "row_count": official_retrieval_validation.row_count,
            "error_count": len(official_retrieval_validation.errors),
            "errors": official_retrieval_validation.errors,
            "row_errors": official_retrieval_validation.row_errors,
            "bucket_counts": official_retrieval_validation.bucket_counts,
        },
        "validation_errors": validation_errors,
        "policy": {
            "human_labels_preserved": True,
            "user_gold_policy_decision_was_not_overwritten": True,
            "user_include_in_official_denominator_was_not_overwritten": True,
            "denominator_kind": "xlsx_retrieval_evidence_diagnostic",
            "official_xlsx_answer_generation_denominator": 0,
            "llm_answer_used_for_source_validation": False,
            "retrieval_projection_schema": "eval.harness.rag_ingestion_retrieval_eval.REQUIRED_COLUMNS",
            "empty_user_denominator_fields_are_not_silently_coerced": True,
            "official_positive_rule": (
                "ANSWERABLE_CONFIRMED + EVIDENCE_RELEVANT + valid non-empty citation locator + "
                "sheet/range + normalized expected answer + must_contain_terms + source terms "
                "validated from bound evidence or workbook."
            ),
            "diagnostic_only_rule": (
                "Partial evidence, invalid official contracts, empty locators, missing sheet/range, "
                "or source-term validation failures stay outside the official positive denominator."
            ),
            "excluded_rule": (
                "EVIDENCE_MISMATCH, POLICY_EXCLUDED, and NOT_ANSWERABLE rows are excluded from "
                "the positive denominator because no abstain/negative denominator is configured."
            ),
        },
        "workbook_validation": workbook_cache.report(),
        "outputs": {
            "normalized_csv": output_ref(output),
            "official_positive_csv": output_ref(official_output),
            "official_positive_retrieval_csv": output_ref(official_retrieval_output),
            "normalized_jsonl": output_ref(jsonl_output),
            "report": {"path": repo_relative(report_path), "sha256": None},
            "registry": output_ref(registry_path),
        },
        "registry_updated": False,
    }
    return report


def special_row_summary(rows: list[Mapping[str, Any]]) -> dict[str, Any]:
    wanted = {
        "gq_xlsx_lookup_005",
        "gq_xlsx_lookup_006",
        "gq_xlsx_date_number_format_003",
        "gq_xlsx_aggregation_001",
    }
    return {
        str(row.get("query_id")): {
            "derived_denominator_policy": row.get("derived_denominator_policy"),
            "include_in_official_positive_denominator": row.get("include_in_official_positive_denominator"),
            "normalized_expected_answer_text": row.get("normalized_expected_answer_text"),
            "sheet": row.get("sheet"),
            "range": row.get("range"),
            "source_validation_status": row.get("source_validation_status"),
            "source_validation_basis": row.get("source_validation_basis"),
            "llm_answer_used_for_source_validation": row.get("llm_answer_used_for_source_validation"),
            "denominator_kind": row.get("denominator_kind"),
            "evidence_terms_validation_status": row.get("evidence_terms_validation_status"),
            "workbook_source_validation_status": row.get("workbook_source_validation_status"),
            "derived_policy_reasons": row.get("derived_policy_reasons"),
        }
        for row in rows
        if row.get("query_id") in wanted
    }


def update_denominator_registry(*, registry_path: Path, report: Mapping[str, Any]) -> None:
    registry = json.loads(registry_path.read_text(encoding="utf-8"))
    registry["updated_at"] = str(report.get("generated_at", ""))[:10] or "2026-05-07"
    current_defaults = registry.setdefault("current_defaults", {})
    current_defaults["track_a_xlsx"] = {
        "denominator_key": "track_a_xlsx_human_review_normalized_v0",
        "retrieval_positive_path": report["outputs"]["official_positive_retrieval_csv"]["path"],
        "official_positive_denominator": report["official_positive_count"],
        "official_xlsx_answer_generation_denominator": 0,
        "supersedes": "track_a_xlsx_reviewed_positive",
    }
    denominators = registry.setdefault("official_diagnostic_denominators", {})
    legacy = denominators.get("track_a_xlsx_reviewed_positive")
    if isinstance(legacy, dict):
        legacy.update(
            {
                "current_default": False,
                "superseded_by": "track_a_xlsx_human_review_normalized_v0",
                "retained_for": "legacy_v3_diagnostic_only",
                "not_answer_generation_denominator": True,
                "official_xlsx_answer_generation_denominator": 0,
            }
        )
    denominators["track_a_xlsx_human_review_normalized_v0"] = {
        "path": report["outputs"]["normalized_csv"]["path"],
        "row_count": report["total_rows"],
        "official_positive_denominator": report["official_positive_count"],
        "diagnostic_only_count": report["diagnostic_only_count"],
        "pending_source_verification_count": report["pending_source_verification_count"],
        "excluded_count": report["excluded_count"],
        "sha256": report["outputs"]["normalized_csv"]["sha256"],
        "official_positive_subset_path": report["outputs"]["official_positive_csv"]["path"],
        "official_positive_subset_sha256": report["outputs"]["official_positive_csv"]["sha256"],
        "official_positive_retrieval_subset_path": report["outputs"]["official_positive_retrieval_csv"]["path"],
        "official_positive_retrieval_subset_sha256": report["outputs"]["official_positive_retrieval_csv"]["sha256"],
        "source_review_pack_sha256": report["source_review_pack"]["sha256"],
        "normalization_report": report["outputs"]["report"]["path"],
        "gold_status_policy": (
            "derived_denominator_policy=OFFICIAL_POSITIVE rows are official positive; "
            "DIAGNOSTIC_ONLY, PENDING_SOURCE_VERIFICATION, and EXCLUDED rows are not official positive."
        ),
        "denominator_rule": (
            "human ANSWERABLE_CONFIRMED + EVIDENCE_RELEVANT with strict locator, sheet/range, "
            "expected-answer, must-contain, and source-term validation."
        ),
        "promotion_evidence": False,
        "evidence_role": "diagnostic",
        "denominator_kind": "xlsx_retrieval_evidence_diagnostic",
        "official_xlsx_answer_generation_denominator": 0,
        "not_answer_generation_denominator": True,
        "current_default": True,
        "xlsx_retrieval_wrapper_default": True,
        "replaces_legacy_track_a_xlsx_reviewed_positive_default_for_wrapper": True,
        "legacy_track_a_xlsx_reviewed_positive_artifact_preserved": True,
    }
    registry_path.write_text(json.dumps(registry, ensure_ascii=False, indent=2) + "\n", encoding="utf-8")


def project_official_rows_to_retrieval_gold(rows: list[Mapping[str, Any]]) -> list[dict[str, str]]:
    projected: list[dict[str, str]] = []
    for row in rows:
        locator, _ = parse_locator(row.get("citation_locator"))
        terms = parse_terms(row.get("normalized_must_contain_terms_json"))
        projected.append(
            {
                "query_id": clean(row.get("query_id")),
                "bucket": retrieval_bucket(row),
                "query": clean(row.get("query")),
                "expected_file_name": clean(row.get("citation_locator_file") or locator.get("file")),
                "expected_document_version_id": clean(
                    row.get("citation_locator_document_version_id") or locator.get("document_version_id")
                ),
                "expected_chunk_type": clean(locator.get("chunk_type")) or "row_group",
                "expected_location_type": "xlsx",
                "expected_sheet_name": clean(row.get("sheet") or row.get("citation_locator_sheet") or locator.get("sheet")),
                "expected_cell_range": clean(row.get("range") or row.get("citation_locator_range") or locator.get("range")),
                "expected_table_id": clean(locator.get("table_id")),
                "expected_physical_page_index": "",
                "expected_page_no": "",
                "expected_page_label": "",
                "expected_bbox": "",
                "expected_answer_text": clean(row.get("normalized_expected_answer_text")),
                "must_contain_terms": ";".join(terms),
                "must_not_contain_terms": "",
                "range_match_policy": retrieval_range_policy(row),
                "hidden_policy": "exclude_hidden",
                "requires_formula_value": bool_text(requires_formula_value(row)).lower(),
                "requires_formatted_value": bool_text(requires_formatted_value(row)).lower(),
                "requires_aggregation": bool_text(requires_aggregation(row)).lower(),
                "source_sample_id": "xlsx_human_review_normalized_v0",
                "label_status": "bound",
                "notes": (
                    "projected_from=gold_queries_xlsx_human_review_official_positive_v0;"
                    "denominator_kind=xlsx_retrieval_evidence_diagnostic;"
                    "official_xlsx_answer_generation_denominator=0;"
                    "llm_answer_used_for_source_validation=false"
                ),
            }
        )
    return projected


def retrieval_bucket(row: Mapping[str, Any]) -> str:
    query_id = clean(row.get("query_id")).lower()
    shape = clean(row.get("normalized_gold_answer_shape") or row.get("user_gold_answer_shape")).upper()
    if "hidden" in query_id:
        return "xlsx_hidden_policy"
    if "aggregation" in query_id or shape == "AGGREGATION_RESULT":
        return "xlsx_aggregation"
    if "date_number_format" in query_id:
        return "xlsx_date_number_format"
    if shape == "FORMULA_VALUE":
        return "xlsx_formula_value"
    if "header" in query_id:
        return "xlsx_header_ambiguous"
    return "xlsx_lookup"


def retrieval_range_policy(row: Mapping[str, Any]) -> str:
    policy = clean(row.get("user_required_citation_policy") or row.get("_citation_policy_for_validation")).upper()
    if policy == "POLICY_EXCLUDED":
        return "none"
    return "exact_match"


def requires_formula_value(row: Mapping[str, Any]) -> bool:
    shape = clean(row.get("normalized_gold_answer_shape") or row.get("user_gold_answer_shape")).upper()
    return shape == "FORMULA_VALUE"


def requires_formatted_value(row: Mapping[str, Any]) -> bool:
    query_id = clean(row.get("query_id")).lower()
    shape = clean(row.get("normalized_gold_answer_shape") or row.get("user_gold_answer_shape")).upper()
    return "date_number_format" in query_id or shape == "FORMULA_VALUE"


def requires_aggregation(row: Mapping[str, Any]) -> bool:
    query_id = clean(row.get("query_id")).lower()
    shape = clean(row.get("normalized_gold_answer_shape") or row.get("user_gold_answer_shape")).upper()
    return "aggregation" in query_id or shape == "AGGREGATION_RESULT"


class WorkbookCache:
    def __init__(self, dataset_root: Path) -> None:
        self.dataset_root = dataset_root
        self._paths: dict[str, Path | None] = {}
        self._workbooks: dict[Path, Any] = {}
        self.status_counts: Counter[str] = Counter()
        self.lookups: dict[str, str] = {}

    def validate_terms(self, *, file_name: str, sheet: str, cell_range: str, terms: list[str]) -> str:
        if not terms:
            self.status_counts["SKIPPED_NO_TERMS"] += 1
            return "SKIPPED_NO_TERMS"
        if not file_name:
            self.status_counts["SKIPPED_NO_FILE"] += 1
            return "SKIPPED_NO_FILE"
        path = self.find_file(file_name)
        if path is None:
            self.status_counts["SKIPPED_FILE_NOT_FOUND"] += 1
            return "SKIPPED_FILE_NOT_FOUND"
        if path.suffix.lower() not in {".xlsx", ".xlsm"}:
            self.status_counts["SKIPPED_UNSUPPORTED_EXTENSION"] += 1
            return "SKIPPED_UNSUPPORTED_EXTENSION"
        if not sheet or not cell_range:
            self.status_counts["FAIL_MISSING_SHEET_OR_RANGE"] += 1
            return "FAIL_MISSING_SHEET_OR_RANGE"
        try:
            workbook = self.load_workbook(path)
            if sheet not in workbook.sheetnames:
                self.status_counts["FAIL_SHEET_NOT_FOUND"] += 1
                return "FAIL_SHEET_NOT_FOUND"
            worksheet = workbook[sheet]
            values = []
            for row in worksheet[cell_range]:
                for cell in row:
                    if cell.value is not None:
                        values.append(str(cell.value))
            text = " ".join(values)
        except Exception as exc:  # pragma: no cover - defensive for local workbook drift
            status = f"FAIL_WORKBOOK_READ:{type(exc).__name__}"
            self.status_counts[status] += 1
            return status
        missing = [term for term in terms if not term_present(term, text)]
        if missing:
            self.status_counts["FAIL_TERMS_MISSING"] += 1
            return "FAIL_TERMS_MISSING"
        self.status_counts["PASS"] += 1
        return "PASS"

    def find_file(self, file_name: str) -> Path | None:
        if file_name not in self._paths:
            matches = sorted(self.dataset_root.rglob(file_name)) if self.dataset_root.exists() else []
            self._paths[file_name] = matches[0] if matches else None
            self.lookups[file_name] = repo_relative(matches[0]) if matches else "NOT_FOUND"
        return self._paths[file_name]

    def load_workbook(self, path: Path) -> Any:
        if path not in self._workbooks:
            from openpyxl import load_workbook

            self._workbooks[path] = load_workbook(path, data_only=True, read_only=True)
        return self._workbooks[path]

    def report(self) -> dict[str, Any]:
        return {
            "dataset_root": repo_relative(self.dataset_root),
            "status_counts": dict(self.status_counts),
            "file_lookups": dict(sorted(self.lookups.items())),
        }


def parse_locator(value: Any) -> tuple[dict[str, Any], list[str]]:
    text = clean(value)
    if not text:
        return {}, ["invalid_citation_locator_json"]
    try:
        payload = json.loads(text)
    except json.JSONDecodeError:
        return {}, ["invalid_citation_locator_json"]
    if not isinstance(payload, dict):
        return {}, ["citation_locator_not_object"]
    return payload, []


def parse_terms(value: Any) -> list[str]:
    text = clean(value)
    if not text:
        return []
    try:
        payload = json.loads(text)
        if isinstance(payload, list):
            return [clean(str(item)) for item in payload if clean(str(item))]
    except json.JSONDecodeError:
        pass
    return [clean(part) for part in re.split(r"[;,|]", text) if clean(part)]


def bound_evidence_text(row: Mapping[str, Any]) -> str:
    return " ".join(clean(row.get(field)) for field in SOURCE_TEXT_FIELDS)


def term_present(term: str, text: str) -> bool:
    term_norm = normalize_text(term)
    text_norm = normalize_text(text)
    if term_norm and term_norm in text_norm:
        return True
    parts = [normalize_text(part) for part in clean(term).split() if normalize_text(part)]
    return len(parts) > 1 and all(part in text_norm for part in parts)


def normalize_text(value: Any) -> str:
    return re.sub(r"\s+", "", clean(value)).replace(",", "")


def read_csv_rows(path: Path) -> list[dict[str, str]]:
    with path.open("r", encoding="utf-8-sig", newline="") as handle:
        return list(csv.DictReader(handle))


def write_csv(path: Path, fieldnames: list[str], rows: Iterable[Mapping[str, Any]]) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    cleaned_rows = [{key: value for key, value in row.items() if not key.startswith("_")} for row in rows]
    with path.open("w", encoding="utf-8", newline="") as handle:
        writer = csv.DictWriter(handle, fieldnames=fieldnames)
        writer.writeheader()
        for row in cleaned_rows:
            writer.writerow({field: row.get(field, "") for field in fieldnames})


def write_jsonl(path: Path, rows: Iterable[Mapping[str, Any]]) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    with path.open("w", encoding="utf-8", newline="\n") as handle:
        for row in rows:
            payload = {key: value for key, value in row.items() if not key.startswith("_")}
            handle.write(json.dumps(payload, ensure_ascii=False, separators=(",", ":")) + "\n")


def write_json(path: Path, payload: Mapping[str, Any]) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    path.write_text(json.dumps(payload, ensure_ascii=False, indent=2) + "\n", encoding="utf-8")


def print_json(payload: Mapping[str, Any]) -> None:
    try:
        print(json.dumps(payload, ensure_ascii=False, indent=2))
    except UnicodeEncodeError:
        print(json.dumps(payload, ensure_ascii=True, indent=2))


def output_ref(path: Path) -> dict[str, Any]:
    return {
        "path": repo_relative(path),
        "exists": path.exists(),
        "sha256": sha256_file(path) if path.exists() and path.is_file() else None,
    }


def sha256_file(path: Path) -> str:
    digest = hashlib.sha256()
    with path.open("rb") as handle:
        for chunk in iter(lambda: handle.read(1024 * 1024), b""):
            digest.update(chunk)
    return digest.hexdigest()


def utc_run_id() -> str:
    return datetime.now(timezone.utc).strftime("%Y%m%dT%H%M%SZ")


def utc_timestamp() -> str:
    return datetime.now(timezone.utc).isoformat()


def clean(value: Any) -> str:
    return "" if value is None else str(value).strip()


def bool_text(value: bool) -> str:
    return "TRUE" if value else "FALSE"


def dedupe(values: Iterable[str]) -> list[str]:
    seen: set[str] = set()
    out: list[str] = []
    for value in values:
        if value and value not in seen:
            seen.add(value)
            out.append(value)
    return out


def repo_relative(path: Path) -> str:
    try:
        return path.resolve().relative_to(REPO_ROOT.resolve()).as_posix()
    except ValueError:
        try:
            return path.resolve().relative_to(AI_WORKER_ROOT.resolve()).as_posix()
        except ValueError:
            return path.name


if __name__ == "__main__":  # pragma: no cover
    raise SystemExit(main())
