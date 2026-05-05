"""Build a report for the additional XLSX dataset canary selection.

This is evidence-only: it does not ingest, index, promote, or mutate any
baseline artifact. The hardened manifest is the source of selected sample ids.
"""

from __future__ import annotations

import json
from datetime import datetime, timezone
from pathlib import Path
from typing import Any

from openpyxl import load_workbook


ROOT = Path(__file__).resolve().parents[1]
MANIFEST_PATH = ROOT / "fixtures" / "manifests" / "rag_ingestion_hardened_xlsx_manifest.json"
REPORT_PATH = ROOT / "eval" / "reports" / "rag-ingestion" / "rag_xlsx_additional_dataset_canary_report.json"

DATASETS = {
    "election_turnout_20200415": {
        "requested_name": "중앙선거관리위원회_제21대 국회의원선거 투표율 분석_20200415",
        "root": Path("eval/datasets/xlsx/중앙선거관리위원회_제21대 국회의원선거 투표율 분석_20200415"),
        "domain": "election turnout statistics",
    },
    "employment_sentiment_20201229": {
        "requested_name": "국가데이터처_인공지능 학습을 위한 고용기사 감성지수 라벨링 데이터_20201229",
        "root": Path("eval/datasets/xlsx/국가데이터처_인공지능 학습을 위한 고용기사 감성지수 라벨링 데이터_20201229"),
        "domain": "employment news sentiment labels",
    },
    "surgery_statistics_2024": {
        "requested_name": "2024_주요수술통계연보(한글,_엑셀)",
        "root": Path("eval/datasets/xlsx/2024_주요수술통계연보(한글,_엑셀)"),
        "domain": "healthcare surgery statistics",
    },
}

SELECTED_SAMPLE_IDS = [
    "xlsx_hardened_surgery_major_indicators_001",
    "xlsx_hardened_surgery_laparoscopic_001",
    "xlsx_hardened_election_advance_turnout_001",
    "xlsx_hardened_election_age_gender_001",
    "xlsx_hardened_employment_sentiment_2019_001",
    "xlsx_hardened_employment_sentiment_2020_001",
]


def rel(path: Path) -> str:
    return path.relative_to(ROOT).as_posix()


def bytes_to_mib(size: int) -> float:
    return round(size / (1024 * 1024), 4)


def inventory_dataset(dataset: dict[str, Any]) -> dict[str, Any]:
    root = ROOT / dataset["root"]
    xlsx_files = sorted(root.rglob("*.xlsx"))
    hwp_files = sorted(root.rglob("*.hwp"))
    other_files = sorted(
        path
        for path in root.rglob("*")
        if path.is_file() and path.suffix.lower() not in {".xlsx", ".hwp"}
    )
    xlsx_bytes = sum(path.stat().st_size for path in xlsx_files)
    hwp_bytes = sum(path.stat().st_size for path in hwp_files)
    other_bytes = sum(path.stat().st_size for path in other_files)
    return {
        "requested_name": dataset["requested_name"],
        "domain": dataset["domain"],
        "root": rel(root),
        "exists": root.exists(),
        "xlsx_file_count": len(xlsx_files),
        "xlsx_size_bytes": xlsx_bytes,
        "xlsx_size_mib": bytes_to_mib(xlsx_bytes),
        "hwp_file_count": len(hwp_files),
        "hwp_size_bytes": hwp_bytes,
        "hwp_size_mib": bytes_to_mib(hwp_bytes),
        "other_file_count": len(other_files),
        "other_size_bytes": other_bytes,
        "other_size_mib": bytes_to_mib(other_bytes),
        "xlsx_files_sample": [rel(path) for path in xlsx_files[:8]],
    }


def dataset_id_for_file(file_path: str) -> str | None:
    normalized = file_path.replace("\\", "/")
    for dataset_id, dataset in DATASETS.items():
        root = dataset["root"].as_posix()
        if normalized.startswith(root):
            return dataset_id
    return None


def compact_cell(value: Any) -> str | None:
    if value is None:
        return None
    text = str(value).replace("\n", " ").strip()
    if len(text) > 80:
        return text[:77] + "..."
    return text


def sample_rows(ws: Any, max_rows: int = 4, max_cols: int = 12) -> list[list[str | None]]:
    rows: list[list[str | None]] = []
    for row_idx in range(1, min(ws.max_row, 30) + 1):
        values = [
            compact_cell(ws.cell(row_idx, col_idx).value)
            for col_idx in range(1, min(ws.max_column, max_cols) + 1)
        ]
        while values and values[-1] is None:
            values.pop()
        if values:
            rows.append(values)
        if len(rows) >= max_rows:
            break
    return rows


def workbook_probe(path: Path) -> dict[str, Any]:
    workbook = load_workbook(path, data_only=False, read_only=False)
    sheet_probes = []
    total_hidden_rows = 0
    total_hidden_cols = 0
    total_merged_ranges = 0
    for ws in workbook.worksheets:
        hidden_rows = sum(1 for dim in ws.row_dimensions.values() if dim.hidden)
        hidden_cols = sum(1 for dim in ws.column_dimensions.values() if dim.hidden)
        merged_ranges = len(ws.merged_cells.ranges)
        total_hidden_rows += hidden_rows
        total_hidden_cols += hidden_cols
        total_merged_ranges += merged_ranges
        if len(sheet_probes) < 6:
            sheet_probes.append(
                {
                    "sheet": ws.title,
                    "dimension": ws.calculate_dimension(),
                    "hidden_row_count": hidden_rows,
                    "hidden_col_count": hidden_cols,
                    "merged_range_count": merged_ranges,
                    "sample_rows": sample_rows(ws),
                }
            )
    return {
        "sheet_count": len(workbook.worksheets),
        "sheet_names_sample": workbook.sheetnames[:12],
        "total_hidden_row_count": total_hidden_rows,
        "total_hidden_col_count": total_hidden_cols,
        "total_merged_range_count": total_merged_ranges,
        "sheet_probes": sheet_probes,
    }


def load_manifest() -> dict[str, Any]:
    return json.loads(MANIFEST_PATH.read_text(encoding="utf-8"))


def build_report() -> dict[str, Any]:
    manifest = load_manifest()
    samples = {sample["sample_id"]: sample for sample in manifest["samples"]}
    selected_samples = []
    missing_sample_ids = []
    missing_files = []
    selected_by_dataset: dict[str, int] = {}

    for sample_id in SELECTED_SAMPLE_IDS:
        sample = samples.get(sample_id)
        if sample is None:
            missing_sample_ids.append(sample_id)
            continue
        dataset_id = dataset_id_for_file(sample["file"])
        if dataset_id is not None:
            selected_by_dataset[dataset_id] = selected_by_dataset.get(dataset_id, 0) + 1
        file_path = ROOT / sample["file"]
        exists = file_path.exists()
        if not exists:
            missing_files.append(sample["file"])
        selected_samples.append(
            {
                "sample_id": sample_id,
                "dataset_id": dataset_id,
                "file": sample["file"],
                "exists": exists,
                "priority": sample.get("priority"),
                "diagnostic_only": sample.get("diagnostic_only", False),
                "buckets": sample.get("buckets", []),
                "expected_visible_sheets": sample.get("expected_visible_sheets", []),
                "expected_visible_ranges": sample.get("expected_visible_ranges", []),
                "candidate_query_hints": sample.get("candidate_query_hints", []),
                "workbook_probe": workbook_probe(file_path) if exists else None,
            }
        )

    dataset_inventory = {
        dataset_id: inventory_dataset(dataset) for dataset_id, dataset in DATASETS.items()
    }
    xlsx_file_count = sum(item["xlsx_file_count"] for item in dataset_inventory.values())
    xlsx_size_bytes = sum(item["xlsx_size_bytes"] for item in dataset_inventory.values())
    hwp_file_count = sum(item["hwp_file_count"] for item in dataset_inventory.values())

    has_all_datasets = all(
        selected_by_dataset.get(dataset_id, 0) > 0 for dataset_id in DATASETS.keys()
    )
    has_missing = bool(missing_sample_ids or missing_files)
    sufficiency_status = (
        "SUFFICIENT_FOR_NEXT_CANARY" if has_all_datasets and not has_missing else "NEEDS_FIX"
    )

    return {
        "report_id": "rag_xlsx_additional_dataset_canary_report",
        "generated_at": datetime.now(timezone.utc).isoformat(),
        "manifest_path": rel(MANIFEST_PATH),
        "manifest_version": manifest.get("manifest_version"),
        "scope": "additional_requested_xlsx_datasets",
        "promotion_evidence": False,
        "evidence_role": "dataset_canary_selection",
        "indexing_performed": False,
        "promotion_performed": False,
        "dataset_inventory": dataset_inventory,
        "dataset_totals": {
            "dataset_count": len(DATASETS),
            "xlsx_file_count": xlsx_file_count,
            "xlsx_size_bytes": xlsx_size_bytes,
            "xlsx_size_mib": bytes_to_mib(xlsx_size_bytes),
            "hwp_file_count": hwp_file_count,
        },
        "selection_summary": {
            "selected_sample_count": len(selected_samples),
            "selected_by_dataset": selected_by_dataset,
            "missing_sample_ids": missing_sample_ids,
            "missing_files": missing_files,
        },
        "selected_samples": selected_samples,
        "sufficiency_assessment": {
            "status": sufficiency_status,
            "rationale": [
                "Requested datasets cover election turnout, employment news sentiment labels, and healthcare surgery statistics.",
                "The selection adds multi-sheet merged election tables, text-heavy URL/label matrices, wide healthcare indicator tables, and hidden-row surgery sheets.",
                "The HWP companion files in the surgery dataset are outside the XLSX_EXTRACT scope and are not counted as XLSX canary input.",
            ],
        },
        "recommended_next_steps": [
            "Run manifest-driven XLSX_EXTRACT canary reimport for these six selected samples only.",
            "Create a new candidate namespace/index version if the additional documents are embedded.",
            "Build dataset-specific gold rows only from visible cells after reimport proves document_version/search_unit bindings.",
        ],
    }


def main() -> None:
    report = build_report()
    REPORT_PATH.parent.mkdir(parents=True, exist_ok=True)
    REPORT_PATH.write_text(
        json.dumps(report, ensure_ascii=False, indent=2) + "\n",
        encoding="utf-8",
    )
    print(json.dumps(report["selection_summary"], ensure_ascii=False, indent=2))
    print(f"wrote {rel(REPORT_PATH)}")


if __name__ == "__main__":
    main()
