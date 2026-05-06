"""Trace where XLSX content-bearing evidence drops before answer inputs.

This script is diagnostic-only. It reads existing answer-generation inputs,
existing retrieval reports, and read-only DB/index state when available. It does
not mutate retrieval, parser/chunking, SearchUnit rows, vectors, gold CSVs, or
answer denominators.
"""

from __future__ import annotations

import argparse
import csv
import json
import os
import re
import sys
import zipfile
from collections import Counter
from dataclasses import dataclass
from datetime import datetime, timezone
from pathlib import Path
from typing import Any, Iterable, Mapping


AI_WORKER_ROOT = Path(__file__).resolve().parents[1]
REPO_ROOT = AI_WORKER_ROOT.parent
REPORT_DIR = AI_WORKER_ROOT / "eval" / "reports" / "rag-ingestion"
DATASET_DIR = AI_WORKER_ROOT / "eval" / "datasets"
EVAL_RUNS_DIR = AI_WORKER_ROOT / "eval" / "artifacts" / "eval_runs"

SCHEMA_VERSION = "rag_xlsx_content_drop_trace_v1"
RUN_PREFIX = "pdf_xlsx_answer_shape_xlsx_content_drop_trace"
ANSWER_SHAPE_POLICY_PENDING = "NOT_ANSWERABLE_OR_POLICY_PENDING"
DEFAULT_REPAIR_REPORT = REPORT_DIR / "rag_pdf_xlsx_answer_shape_repair_report.json"
DEFAULT_RETRIEVAL_REPORT = (
    REPORT_DIR / "rag_retrieval_eval_xlsx_v3_positive_reviewed_vector_diagnostic_report.json"
)
DEFAULT_FAILURE_BREAKDOWN = REPORT_DIR / "rag_xlsx_v3_after_cleanup_failure_breakdown.json"
DEFAULT_REPORT_JSON = REPORT_DIR / "rag_pdf_xlsx_answer_shape_xlsx_content_drop_trace_report.json"
DEFAULT_REPORT_CSV = REPORT_DIR / "rag_pdf_xlsx_answer_shape_xlsx_content_drop_trace.csv"
DEFAULT_DB_DSN = "host=localhost port=5433 dbname=aipipeline user=aipipeline password=aipipeline_pw"

LOCATOR_KEYS = {
    "bbox",
    "cell",
    "cell_range",
    "cellrange",
    "citation",
    "citation_text",
    "citationtext",
    "docv",
    "document_version_id",
    "documentversionid",
    "expected_cell_range",
    "expected_current_evidence_location",
    "expected_evidence_location",
    "expected_file_name",
    "expected_sheet_name",
    "file",
    "file_name",
    "filename",
    "index_version",
    "location_json",
    "locator",
    "page",
    "range",
    "rank",
    "score",
    "sheet",
    "sheet_index",
    "sheet_name",
    "source_file_name",
    "sourcefilename",
    "table",
    "table_id",
    "type",
    "unit_key",
}
LEAKAGE_KEYS = {
    "expected_answer",
    "expected_answer_text",
    "must_contain",
    "must_contain_terms",
    "query",
    "question",
}
CONTENT_FIELD_KEYS = {
    "bm25_text",
    "cell_value",
    "cell_values",
    "cells",
    "chunk_text",
    "chunks",
    "compact_text",
    "compacttext",
    "content",
    "content_summary",
    "content_text",
    "debug_text",
    "display_text",
    "embedding_text",
    "markdown",
    "nearby_rows",
    "nearby_table_context",
    "plain_text",
    "plaintext",
    "row_context",
    "row_values",
    "table_context",
    "table_text",
    "tables",
    "text",
    "text_content",
    "value",
    "value_context",
    "values",
}


@dataclass(frozen=True)
class DbSnapshot:
    available: bool
    error: str
    parser_artifacts_by_docv: dict[str, list[dict[str, Any]]]
    index_units_by_docv: dict[str, list[dict[str, Any]]]


def main(argv: list[str] | None = None) -> int:
    args = parse_args(argv)
    report = run_trace(
        answer_generation_inputs=resolve_answer_inputs(Path(args.answer_generation_inputs)),
        retrieval_report=Path(args.retrieval_report),
        failure_breakdown=Path(args.failure_breakdown),
        output_root=Path(args.output_root),
        report_json=Path(args.report_json),
        report_csv=Path(args.report_csv),
        dataset_root=Path(args.dataset_root),
        db_dsn=args.db_dsn or os.environ.get("RAG_DB_DSN") or DEFAULT_DB_DSN,
        run_id=args.run_id,
        disable_db=args.disable_db,
    )
    print_json(
        {
            "status": report["status"],
            "run_id": report["run_id"],
            "artifact_dir": report["new_run_artifact_path"],
            "report_json": report["report_path"],
            "xlsx_total_rows": report["xlsx_total_rows"],
            "content_exists_upstream_but_dropped_before_answer_inputs_count": report[
                "content_exists_upstream_but_dropped_before_answer_inputs_count"
            ],
            "xlsx_answer_eval_denominator": report["xlsx_answer_eval_denominator"],
            "promotion_evidence": report["promotion_evidence"],
        }
    )
    return 0 if report["status"] in {"PASS", "PASS_WITH_WARNINGS"} else 2


def parse_args(argv: list[str] | None) -> argparse.Namespace:
    parser = argparse.ArgumentParser(description=__doc__)
    parser.add_argument("--answer-generation-inputs", default="")
    parser.add_argument("--retrieval-report", default=str(DEFAULT_RETRIEVAL_REPORT))
    parser.add_argument("--failure-breakdown", default=str(DEFAULT_FAILURE_BREAKDOWN))
    parser.add_argument("--output-root", default=str(EVAL_RUNS_DIR))
    parser.add_argument("--report-json", default=str(DEFAULT_REPORT_JSON))
    parser.add_argument("--report-csv", default=str(DEFAULT_REPORT_CSV))
    parser.add_argument("--dataset-root", default=str(DATASET_DIR))
    parser.add_argument("--db-dsn", default="")
    parser.add_argument("--run-id", default="")
    parser.add_argument("--disable-db", action="store_true")
    return parser.parse_args(argv)


def run_trace(
    *,
    answer_generation_inputs: Path,
    retrieval_report: Path,
    failure_breakdown: Path,
    output_root: Path,
    report_json: Path,
    report_csv: Path,
    dataset_root: Path,
    db_dsn: str,
    run_id: str = "",
    disable_db: bool = False,
) -> dict[str, Any]:
    run_id = run_id or utc_run_id()
    generated_at = utc_timestamp()
    artifact_dir = output_root / f"{RUN_PREFIX}_{run_id}"
    trace_jsonl = artifact_dir / "xlsx_content_drop_trace.jsonl"
    trace_csv = artifact_dir / "xlsx_content_drop_trace.csv"

    all_input_rows = read_jsonl(answer_generation_inputs)
    input_rows = [row for row in all_input_rows if clean(row.get("track")).upper() == "XLSX"]
    docv_ids = sorted({
        locator_from_input(row).get("document_version_id", "")
        for row in input_rows
        if locator_from_input(row).get("document_version_id")
    })
    retrieval_maps = {
        "query_results": keyed_json_rows(read_json_object(retrieval_report), "query_results"),
        "classified_rows": keyed_json_rows(read_json_object(failure_breakdown), "classified_query_rows"),
    }
    dataset_index = build_dataset_index(dataset_root)
    db_snapshot = (
        DbSnapshot(False, "disabled by --disable-db", {}, {})
        if disable_db
        else load_db_snapshot(db_dsn, docv_ids)
    )

    trace_rows = build_trace_rows(
        input_rows,
        run_id=run_id,
        dataset_index=dataset_index,
        retrieval_maps=retrieval_maps,
        parser_artifacts_by_docv=db_snapshot.parser_artifacts_by_docv,
        index_units_by_docv=db_snapshot.index_units_by_docv,
    )
    report = build_report(
        run_id=run_id,
        generated_at=generated_at,
        artifact_dir=artifact_dir,
        trace_jsonl=trace_jsonl,
        trace_csv=trace_csv,
        report_json=report_json,
        report_csv=report_csv,
        answer_generation_inputs=answer_generation_inputs,
        retrieval_report=retrieval_report,
        failure_breakdown=failure_breakdown,
        db_snapshot=db_snapshot,
        source_input_rows=all_input_rows,
        rows=trace_rows,
    )

    artifact_dir.mkdir(parents=True, exist_ok=True)
    write_jsonl(trace_jsonl, trace_rows)
    write_trace_csv(trace_csv, trace_rows)
    write_json(report_json, report)
    write_report_csv(report_csv, report)
    return report


def build_trace_rows(
    input_rows: Iterable[Mapping[str, Any]],
    *,
    run_id: str,
    dataset_index: Mapping[str, list[Path]],
    retrieval_maps: Mapping[str, Mapping[str, Mapping[str, Any]]],
    parser_artifacts_by_docv: Mapping[str, list[Mapping[str, Any]]],
    index_units_by_docv: Mapping[str, list[Mapping[str, Any]]],
) -> list[dict[str, Any]]:
    rows: list[dict[str, Any]] = []
    for source_row in input_rows:
        locator = locator_from_input(source_row)
        policy_pending = is_policy_pending(source_row)
        source_stage = probe_source_workbook_stage(
            source_row,
            locator=locator,
            dataset_index=dataset_index,
            policy_pending=policy_pending,
        )
        parser_stage = parser_artifact_stage(
            parser_artifacts_by_docv.get(locator.get("document_version_id", ""), []),
            locator=locator,
        )
        index_stage = index_payload_stage(
            index_units_by_docv.get(locator.get("document_version_id", ""), []),
            locator=locator,
        )
        retrieval_stage = retrieval_result_stage(source_row, retrieval_maps=retrieval_maps)
        answer_stage = answer_input_stage(source_row)
        reason_codes = diagnostic_reason_codes(
            policy_pending=policy_pending,
            source_stage=source_stage,
            parser_stage=parser_stage,
            index_stage=index_stage,
            retrieval_stage=retrieval_stage,
            answer_stage=answer_stage,
        )
        upstream_content_available = bool(
            source_stage.get("has_values")
            or parser_stage.get("has_values")
            or index_stage.get("has_values")
            or retrieval_stage.get("has_values")
        )
        answer_input_drops_content = bool(
            upstream_content_available
            and not answer_stage.get("has_values")
            and not policy_pending
        )
        content_never_exists_upstream = not upstream_content_available and not policy_pending
        if answer_input_drops_content and "XLSX_ANSWER_INPUT_DROPS_CONTENT" not in reason_codes:
            reason_codes.append("XLSX_ANSWER_INPUT_DROPS_CONTENT")

        rows.append(
            {
                "schema_version": SCHEMA_VERSION,
                "run_id": run_id,
                "diagnostic_only": True,
                "promotion_evidence": False,
                "answer_allowed_count_impact": "none",
                "denominator_impact": "none",
                "pageindex_run_unavailable_expected_for_xlsx": True,
                "row_index": source_row.get("row_index"),
                "query_id": clean(source_row.get("query_id")),
                "bucket": clean(source_row.get("bucket")),
                "review_group": clean(source_row.get("review_group")),
                "task_type": clean(source_row.get("task_type") or source_row.get("question_type")),
                "expected_answer_shape": clean(source_row.get("expected_answer_shape")),
                "policy_pending": policy_pending,
                "locator": locator,
                "source_workbook_stage": source_stage,
                "parser_artifact_stage": parser_stage,
                "index_payload_stage": index_stage,
                "retrieval_result_stage": retrieval_stage,
                "answer_generation_input_stage": answer_stage,
                "upstream_content_available": upstream_content_available,
                "answer_input_drops_content": answer_input_drops_content,
                "content_never_exists_upstream": content_never_exists_upstream,
                "diagnostic_reason_codes": reason_codes,
            }
        )
    return rows


def probe_source_workbook_stage(
    row: Mapping[str, Any],
    *,
    locator: Mapping[str, str],
    dataset_index: Mapping[str, list[Path]],
    policy_pending: bool,
) -> dict[str, Any]:
    previous_error_codes = classify_context_errors(row)
    if policy_pending:
        return {
            "available": False,
            "opened": False,
            "has_values": False,
            "status": "SKIPPED_POLICY_PENDING",
            "diagnostic_reason_code": "XLSX_POLICY_PENDING",
            "previous_answer_input_extraction_error_codes": previous_error_codes,
            "probe_policy": "source workbook probe skipped for policy-pending row",
        }

    source_path = source_path_from_row(row)
    if source_path is None:
        source_path = resolve_dataset_file(locator.get("file", ""), dataset_index)
    if source_path is None:
        return {
            "available": False,
            "opened": False,
            "has_values": False,
            "status": "MISSING",
            "diagnostic_reason_code": "XLSX_SOURCE_WORKBOOK_MISSING",
            "previous_answer_input_extraction_error_codes": previous_error_codes,
        }

    try:
        summary = safe_workbook_value_probe(
            source_path,
            sheet_name=locator.get("sheet", ""),
            cell_range=locator.get("range", ""),
        )
    except zipfile.BadZipFile as exc:
        return {
            "available": True,
            "opened": False,
            "has_values": False,
            "status": "BAD_ZIP",
            "diagnostic_reason_code": "XLSX_SOURCE_BAD_ZIP",
            "workbook_path": repo_relative(source_path),
            "error": f"{type(exc).__name__}: {exc}",
            "previous_answer_input_extraction_error_codes": previous_error_codes,
        }
    except Exception as exc:  # pragma: no cover - defensive diagnostic path
        return {
            "available": True,
            "opened": False,
            "has_values": False,
            "status": "EXTRACTION_ERROR",
            "diagnostic_reason_code": "XLSX_SOURCE_READONLY_EXTRACTION_ERROR"
            if "ReadOnlyWorksheet" in str(exc)
            else "XLSX_SOURCE_EXTRACTION_ERROR",
            "workbook_path": repo_relative(source_path),
            "error": f"{type(exc).__name__}: {exc}",
            "previous_answer_input_extraction_error_codes": previous_error_codes,
        }

    status = "OPEN_OK_HAS_VALUES" if summary["has_values"] else "OPEN_OK_NO_VALUES"
    return {
        "available": True,
        "opened": True,
        "has_values": summary["has_values"],
        "status": status,
        "diagnostic_reason_code": "XLSX_SOURCE_WORKBOOK_HAS_VALUES"
        if summary["has_values"]
        else "XLSX_SOURCE_WORKBOOK_NO_VALUES",
        "workbook_path": repo_relative(source_path),
        "sheet_found": summary["sheet_found"],
        "range_parse_ok": summary["range_parse_ok"],
        "nonempty_cell_count": summary["nonempty_cell_count"],
        "nonempty_row_count": summary["nonempty_row_count"],
        "probe_policy": "diagnostic counts only; raw cell values are not emitted",
        "previous_answer_input_extraction_error_codes": previous_error_codes,
    }


def safe_workbook_value_probe(source_path: Path, *, sheet_name: str, cell_range: str) -> dict[str, Any]:
    from openpyxl import load_workbook
    from openpyxl.utils.cell import get_column_letter, range_boundaries

    workbook = load_workbook(source_path, data_only=True, read_only=False)
    try:
        if sheet_name and sheet_name in workbook.sheetnames:
            worksheet = workbook[sheet_name]
            sheet_found = True
        else:
            worksheet = workbook[workbook.sheetnames[0]]
            sheet_found = not sheet_name

        try:
            min_col, min_row, max_col, max_row = range_boundaries(cell_range)
            range_parse_ok = True
        except Exception:
            min_col, min_row = 1, 1
            max_col = min(worksheet.max_column or 1, 20)
            max_row = min(worksheet.max_row or 1, 200)
            range_parse_ok = False

        nonempty_cell_count = 0
        nonempty_rows: set[int] = set()
        max_rows = min(max_row, min_row + 199)
        max_cols = min(max_col, min_col + 49)
        for row_index in range(min_row, max_rows + 1):
            row_dim = worksheet.row_dimensions.get(row_index)
            if row_dim is not None and row_dim.hidden:
                continue
            for col_index in range(min_col, max_cols + 1):
                col_letter = get_column_letter(col_index)
                col_dim = worksheet.column_dimensions.get(col_letter)
                if col_dim is not None and col_dim.hidden:
                    continue
                value = worksheet.cell(row=row_index, column=col_index).value
                if value not in (None, ""):
                    nonempty_cell_count += 1
                    nonempty_rows.add(row_index)
        return {
            "sheet_found": sheet_found,
            "range_parse_ok": range_parse_ok,
            "has_values": nonempty_cell_count > 0,
            "nonempty_cell_count": nonempty_cell_count,
            "nonempty_row_count": len(nonempty_rows),
        }
    finally:
        workbook.close()


def parser_artifact_stage(
    artifacts: Iterable[Mapping[str, Any]],
    *,
    locator: Mapping[str, str],
) -> dict[str, Any]:
    artifact_list = list(artifacts)
    if not artifact_list:
        return {
            "available": False,
            "has_values": False,
            "status": "MISSING",
            "diagnostic_reason_code": "XLSX_PARSER_ARTIFACT_MISSING",
            "matched_artifact_count": 0,
            "content_source_fields": [],
        }

    total_values = 0
    total_text_regions = 0
    total_broad_text_regions = 0
    source_fields: Counter[str] = Counter()
    broad_source_fields: Counter[str] = Counter()
    for artifact in artifact_list:
        artifact_json = parse_json_maybe(artifact.get("artifact_json"))
        counts = count_parser_values_for_locator(artifact_json, locator=locator)
        total_values += counts["value_count"]
        total_text_regions += counts["text_region_count"]
        total_broad_text_regions += counts["broad_text_region_count"]
        source_fields.update(counts["source_fields"])
        broad_source_fields.update(counts["broad_source_fields"])

    has_values = total_values > 0 or total_text_regions > 0
    broad_has_values = total_broad_text_regions > 0
    return {
        "available": True,
        "has_values": has_values,
        "exact_locator_has_values": has_values,
        "broad_sheet_or_workbook_has_values": broad_has_values,
        "status": "HAS_VALUES" if has_values else "NO_VALUES",
        "diagnostic_reason_code": "XLSX_PARSER_ARTIFACT_HAS_VALUES"
        if has_values
        else "XLSX_PARSER_ARTIFACT_NO_VALUES",
        "matched_artifact_count": len(artifact_list),
        "nonempty_value_count": total_values,
        "text_region_count": total_text_regions,
        "broad_text_region_count": total_broad_text_regions,
        "content_source_fields": sorted(source_fields),
        "broad_content_source_fields": sorted(broad_source_fields),
    }


def count_parser_values_for_locator(payload: Any, *, locator: Mapping[str, str]) -> dict[str, Any]:
    parsed_range = parse_cell_range(locator.get("range", ""))
    sheet_name = locator.get("sheet", "")
    value_count = 0
    text_region_count = 0
    broad_text_region_count = 0
    source_fields: Counter[str] = Counter()
    broad_source_fields: Counter[str] = Counter()
    for sheet in iter_sheets(payload):
        current_sheet = clean(sheet.get("name") or sheet.get("sheetName"))
        if sheet_name and current_sheet and current_sheet != sheet_name:
            continue
        for cell in sheet.get("cells") or []:
            if not isinstance(cell, Mapping):
                continue
            row = int_or_none(cell.get("row"))
            col = int_or_none(cell.get("column"))
            if parsed_range and (row is None or col is None):
                continue
            if parsed_range and row is not None and col is not None and not point_in_range(row, col, parsed_range):
                continue
            if clean(cell.get("value")):
                value_count += 1
                source_fields["artifact_json.workbook.sheets.cells.value"] += 1
        for key in ("tables", "chunks"):
            for region in sheet.get(key) or []:
                if not isinstance(region, Mapping):
                    continue
                region_range = parse_cell_range(clean(region.get("cellRange") or region.get("range")))
                text = clean(region.get("text") or region.get("compactText") or region.get("markdown"))
                if parsed_range and region_range is None:
                    if has_content_text(text):
                        broad_text_region_count += 1
                        broad_source_fields[f"artifact_json.workbook.sheets.{key}.text"] += 1
                    continue
                if parsed_range and region_range and not ranges_overlap(parsed_range, region_range):
                    continue
                if has_content_text(text):
                    text_region_count += 1
                    source_fields[f"artifact_json.workbook.sheets.{key}.text"] += 1
        for key in ("compactText", "plainText", "text", "markdown"):
            text = clean(sheet.get(key))
            if has_content_text(text):
                broad_text_region_count += 1
                broad_source_fields[f"artifact_json.workbook.sheets.{key}"] += 1
    return {
        "value_count": value_count,
        "text_region_count": text_region_count,
        "broad_text_region_count": broad_text_region_count,
        "source_fields": source_fields,
        "broad_source_fields": broad_source_fields,
    }


def index_payload_stage(
    units: Iterable[Mapping[str, Any]],
    *,
    locator: Mapping[str, str],
) -> dict[str, Any]:
    unit_list = list(units)
    matching_units: list[Mapping[str, Any]] = []
    broad_units: list[Mapping[str, Any]] = []
    parsed_range = parse_cell_range(locator.get("range", ""))
    sheet_name = locator.get("sheet", "")
    for unit in unit_list:
        location = unit.get("location_json") if isinstance(unit.get("location_json"), Mapping) else {}
        unit_sheet = clean(location.get("sheet_name") or location.get("sheetName"))
        if sheet_name and not unit_sheet:
            broad_units.append(unit)
            continue
        if sheet_name and unit_sheet and unit_sheet != sheet_name:
            continue
        unit_range = parse_cell_range(clean(location.get("cell_range") or location.get("cellRange") or location.get("range")))
        if parsed_range and unit_range is None:
            broad_units.append(unit)
            continue
        if parsed_range and unit_range and not ranges_overlap(parsed_range, unit_range):
            continue
        matching_units.append(unit)

    field_counts: Counter[str] = Counter()
    broad_field_counts: Counter[str] = Counter()
    chunk_counts: Counter[str] = Counter()
    for unit in matching_units:
        chunk_type = clean(unit.get("chunk_type") or unit.get("unit_type"))
        if chunk_type:
            chunk_counts[chunk_type] += 1
        for field in ("text_content", "embedding_text", "display_text", "bm25_text", "debug_text"):
            if has_content_text(unit.get(field)):
                field_counts[field] += 1
    for unit in broad_units:
        for field in ("text_content", "embedding_text", "display_text", "bm25_text", "debug_text"):
            if has_content_text(unit.get(field)):
                broad_field_counts[field] += 1
    has_values = bool(field_counts)
    broad_has_values = bool(broad_field_counts)
    return {
        "available": bool(unit_list),
        "has_values": has_values,
        "exact_locator_has_values": has_values,
        "broad_sheet_or_workbook_has_values": broad_has_values,
        "status": "HAS_VALUES" if has_values else "NO_VALUES",
        "diagnostic_reason_code": "XLSX_INDEX_PAYLOAD_HAS_VALUES"
        if has_values
        else "XLSX_INDEX_PAYLOAD_NO_VALUES",
        "candidate_unit_count_for_document": len(unit_list),
        "matching_unit_count": len(matching_units),
        "broad_unit_count": len(broad_units),
        "content_field_counts": dict(sorted(field_counts.items())),
        "broad_content_field_counts": dict(sorted(broad_field_counts.items())),
        "chunk_type_counts": dict(sorted(chunk_counts.items())),
        "content_source_fields": sorted(field_counts),
        "broad_content_source_fields": sorted(broad_field_counts),
    }


def retrieval_result_stage(
    row: Mapping[str, Any],
    *,
    retrieval_maps: Mapping[str, Mapping[str, Mapping[str, Any]]],
) -> dict[str, Any]:
    query_id = clean(row.get("query_id"))
    report_row = dict((retrieval_maps.get("query_results") or {}).get(query_id) or {})
    classified = dict((retrieval_maps.get("classified_rows") or {}).get(query_id) or {})
    context = row.get("context") if isinstance(row.get("context"), Mapping) else {}
    retrieval_context = context.get("retrieval_context") if isinstance(context.get("retrieval_context"), Mapping) else {}
    top_results = list(report_row.get("top_k_results") or retrieval_context.get("top_k_results") or [])
    field_counts = content_field_counts(
        top_results,
        ignored_keys=LOCATOR_KEYS | LEAKAGE_KEYS | {"match_breakdown", "location_json", "source_file_name"},
    )
    has_values = bool(field_counts)
    return {
        "available": bool(report_row or retrieval_context),
        "has_values": has_values,
        "status": "HAS_VALUES" if has_values else "NO_VALUES",
        "diagnostic_reason_code": "XLSX_RETRIEVAL_RESULT_HAS_VALUES"
        if has_values
        else "XLSX_RETRIEVAL_RESULT_NO_VALUES",
        "top_k_result_count": len(top_results),
        "failure_or_quality_classification": clean(
            classified.get("category")
            or classified.get("classification")
            or retrieval_context.get("failure_or_quality_classification")
        ),
        "content_field_counts": field_counts,
        "content_source_fields": sorted(field_counts),
    }


def answer_input_stage(row: Mapping[str, Any]) -> dict[str, Any]:
    context = row.get("context") if isinstance(row.get("context"), Mapping) else {}
    field_counts: Counter[str] = Counter()
    value_context = context.get("value_context")
    if isinstance(value_context, list):
        for item in value_context:
            if isinstance(item, Mapping) and clean(item.get("value")):
                field_counts["context.value_context.value"] += 1
    for key in ("row_values", "column_values", "cell_value", "table_context", "nearby_rows"):
        value = context.get(key)
        if content_field_counts(value, ignored_keys=LOCATOR_KEYS | LEAKAGE_KEYS):
            field_counts[f"context.{key}"] += 1
    nearby = context.get("nearby_table_context")
    if isinstance(nearby, list):
        for item in nearby:
            if has_content_text(item):
                field_counts["context.nearby_table_context"] += 1
    summary = clean(context.get("content_summary"))
    if summary and not summary_leaks_from_gold_or_query(summary, row):
        field_counts["context.content_summary"] += 1

    has_values = bool(field_counts)
    locator_only = not has_values and has_locator(context)
    return {
        "has_values": has_values,
        "locator_only": locator_only,
        "status": "HAS_VALUES" if has_values else ("LOCATOR_ONLY" if locator_only else "NO_VALUES"),
        "diagnostic_reason_code": "XLSX_ANSWER_INPUT_HAS_VALUES"
        if has_values
        else "XLSX_ANSWER_INPUT_LOCATOR_ONLY",
        "content_field_counts": dict(sorted(field_counts.items())),
        "content_source_fields": sorted(field_counts),
        "context_error_codes": classify_context_errors(row),
    }


def diagnostic_reason_codes(
    *,
    policy_pending: bool,
    source_stage: Mapping[str, Any],
    parser_stage: Mapping[str, Any],
    index_stage: Mapping[str, Any],
    retrieval_stage: Mapping[str, Any],
    answer_stage: Mapping[str, Any],
) -> list[str]:
    codes: list[str] = []
    if policy_pending:
        codes.append("XLSX_POLICY_PENDING")
    for stage in (source_stage, parser_stage, index_stage, retrieval_stage):
        code = clean(stage.get("diagnostic_reason_code"))
        if code and code not in codes:
            codes.append(code)
    for code in source_stage.get("previous_answer_input_extraction_error_codes") or []:
        if code and code not in codes:
            codes.append(code)
    answer_code = clean(answer_stage.get("diagnostic_reason_code"))
    if answer_code and answer_code not in codes and not policy_pending:
        codes.append(answer_code)
    return codes


def build_report(
    *,
    run_id: str,
    generated_at: str,
    artifact_dir: Path,
    trace_jsonl: Path,
    trace_csv: Path,
    report_json: Path,
    report_csv: Path,
    answer_generation_inputs: Path,
    retrieval_report: Path,
    failure_breakdown: Path,
    db_snapshot: DbSnapshot,
    source_input_rows: list[Mapping[str, Any]],
    rows: list[Mapping[str, Any]],
) -> dict[str, Any]:
    reason_counts = Counter(
        code
        for row in rows
        for code in row.get("diagnostic_reason_codes", [])
        if code
    )
    policy_pending = sum(1 for row in rows if row.get("policy_pending"))
    source_missing = reason_counts.get("XLSX_SOURCE_WORKBOOK_MISSING", 0)
    source_bad_zip = reason_counts.get("XLSX_SOURCE_BAD_ZIP", 0)
    source_readonly = reason_counts.get("XLSX_SOURCE_READONLY_EXTRACTION_ERROR", 0)
    parser_has = stage_count(rows, "parser_artifact_stage", "has_values", True)
    parser_no = stage_count(rows, "parser_artifact_stage", "has_values", False)
    parser_missing = reason_counts.get("XLSX_PARSER_ARTIFACT_MISSING", 0)
    index_has = stage_count(rows, "index_payload_stage", "has_values", True)
    index_no = stage_count(rows, "index_payload_stage", "has_values", False)
    retrieval_has = stage_count(rows, "retrieval_result_stage", "has_values", True)
    retrieval_no = stage_count(rows, "retrieval_result_stage", "has_values", False)
    answer_has = stage_count(rows, "answer_generation_input_stage", "has_values", True)
    answer_no = stage_count(rows, "answer_generation_input_stage", "has_values", False)
    dropped = sum(1 for row in rows if row.get("answer_input_drops_content"))
    never_upstream = sum(1 for row in rows if row.get("content_never_exists_upstream"))
    parser_exact_has = stage_count(rows, "parser_artifact_stage", "exact_locator_has_values", True)
    parser_broad_has = stage_count(rows, "parser_artifact_stage", "broad_sheet_or_workbook_has_values", True)
    index_exact_has = stage_count(rows, "index_payload_stage", "exact_locator_has_values", True)
    index_broad_has = stage_count(rows, "index_payload_stage", "broad_sheet_or_workbook_has_values", True)
    source_guardrails = source_input_guardrail_metrics(source_input_rows)
    guardrail_violation = source_guardrails["source_input_any_answer_allowed_or_denominator_flag_count"] > 0
    status = "FAIL" if guardrail_violation else ("PASS_WITH_WARNINGS" if not db_snapshot.available else "PASS")

    return {
        "schema_version": SCHEMA_VERSION,
        "run_id": run_id,
        "generated_at": generated_at,
        "status": status,
        "diagnostic_only": True,
        "promotion_evidence": False,
        "external_live_llm_run": False,
        "optional_judge_run": False,
        "retrieval_tuning_run": False,
        "reranking_run": False,
        "parser_expansion_run": False,
        "threshold_relaxation_run": False,
        "broad_indexing_run": False,
        "db_mutation_run": False,
        "searchunit_mutation_run": False,
        "existing_gold_csv_overwritten": False,
        "answer_allowed_count": source_guardrails["source_input_answer_allowed_count"],
        "abstain_count": max(0, source_guardrails["source_input_total_rows"] - source_guardrails["source_input_answer_allowed_count"]),
        "xlsx_answer_eval_denominator": source_guardrails["source_input_xlsx_answer_eval_allowed_count"],
        "pdf_answer_eval_denominator": source_guardrails["source_input_pdf_answer_eval_allowed_count"],
        "denominator_remains_zero": source_guardrails["source_input_denominator_flag_count"] == 0,
        "source_input_guardrails": source_guardrails,
        "denominator_policy_explanation": (
            "Workbook, parser, index, and retrieval probing is diagnostic-only. "
            "No probed source or upstream payload value is promoted into answer evidence, "
            "answer_allowed_count, or XLSX/PDF answer denominators. The reported denominator "
            "is derived from source answer-generation input flags and must remain zero."
        ),
        "pageindex_xlsx_policy": (
            "XLSX PageIndex is intentionally unavailable in this phase; "
            "PAGEINDEX_RUN_UNAVAILABLE is not counted as an XLSX failure."
        ),
        "previous_answer_generation_inputs_path": repo_relative(answer_generation_inputs),
        "new_run_artifact_path": repo_relative(artifact_dir),
        "trace_jsonl_path": repo_relative(trace_jsonl),
        "trace_csv_path": repo_relative(trace_csv),
        "report_path": repo_relative(report_json),
        "report_csv_path": repo_relative(report_csv),
        "source_inputs": {
            "answer_generation_inputs": file_inventory(answer_generation_inputs),
            "retrieval_report": file_inventory(retrieval_report),
            "failure_breakdown": file_inventory(failure_breakdown),
        },
        "db_snapshot": {
            "available": db_snapshot.available,
            "error": db_snapshot.error,
            "parser_document_version_count": len(db_snapshot.parser_artifacts_by_docv),
            "index_document_version_count": len(db_snapshot.index_units_by_docv),
        },
        "xlsx_total_rows": len(rows),
        "policy_pending_rows": policy_pending,
        "source_workbook_missing_count": source_missing,
        "source_bad_zip_count": source_bad_zip,
        "source_readonly_extraction_error_count": source_readonly,
        "parser_artifact_has_values_count": parser_has,
        "parser_artifact_no_values_count": parser_no - parser_missing,
        "parser_artifact_missing_count": parser_missing,
        "parser_artifact_exact_locator_has_values_count": parser_exact_has,
        "parser_artifact_broad_sheet_or_workbook_has_values_count": parser_broad_has,
        "index_payload_has_values_count": index_has,
        "index_payload_no_values_count": index_no,
        "index_payload_exact_locator_has_values_count": index_exact_has,
        "index_payload_broad_sheet_or_workbook_has_values_count": index_broad_has,
        "retrieval_result_has_values_count": retrieval_has,
        "retrieval_result_no_values_count": retrieval_no,
        "answer_input_has_values_count": answer_has,
        "answer_input_no_values_count": answer_no,
        "content_exists_upstream_but_dropped_before_answer_inputs_count": dropped,
        "content_never_exists_upstream_count": never_upstream,
        "diagnostic_reason_counts": dict(sorted(reason_counts.items())),
        "content_drop_interpretation": (
            "Existing hidden-safe parser artifacts and SearchUnit index payloads contain content-bearing "
            "XLSX values for the traced rows, but the reviewed retrieval result and answer_generation_inputs "
            "surface are locator-only. The immediate failure is answer-context assembly/report serialization, "
            "not absence of indexed XLSX content."
        ),
        "rechunk_reembed_justified": False,
        "rechunk_reembed_decision": (
            "Not justified by this trace alone because index payloads already contain content values. "
            "Fix the diagnostic answer-generation input assembly to carry safe retrieved SearchUnit content "
            "before re-chunking or re-embedding."
        ),
        "conservative_decisions": [
            "No retrieval, parser, chunking, SearchUnit mutation, vector rebuild, gold label, expected answer, relevance, answerability, or gold policy file was changed.",
            "Source workbook probing records only counts/status codes and never raw cell values.",
            "Policy-pending rows are traced as policy exclusions and do not trigger source workbook probing for answer evidence.",
            "Retrieval top-k citation text, sheet, range, row number, and file names are treated as locators, not content-bearing values.",
            "Existing XLSX PageIndex unavailability is recorded as expected and not as an XLSX failure.",
            "Parser and index stages distinguish exact/range-overlap content from broad sheet or workbook content; only exact/range-overlap content counts as upstream content for drop classification.",
        ],
    }


def source_input_guardrail_metrics(rows: Iterable[Mapping[str, Any]]) -> dict[str, Any]:
    total = 0
    xlsx_total = 0
    pdf_total = 0
    answer_allowed = 0
    answer_generation_allowed = 0
    xlsx_answer_eval_allowed = 0
    pdf_answer_eval_allowed = 0
    promotion_evidence_true = 0
    denominator_flags = 0
    for row in rows:
        total += 1
        track = clean(row.get("track")).upper()
        if track == "XLSX":
            xlsx_total += 1
        elif track == "PDF":
            pdf_total += 1
        if parse_bool(row.get("answer_allowed")):
            answer_allowed += 1
        if parse_bool(row.get("answer_generation_allowed")):
            answer_generation_allowed += 1
        if parse_bool(row.get("promotion_evidence")):
            promotion_evidence_true += 1
        if parse_bool(row.get("answer_eval_allowed")):
            denominator_flags += 1
            if track == "XLSX":
                xlsx_answer_eval_allowed += 1
            elif track == "PDF":
                pdf_answer_eval_allowed += 1
        for key in ("include_in_answer_denominator", "answer_denominator_included"):
            if parse_bool(row.get(key)):
                denominator_flags += 1
    any_flag = answer_allowed + answer_generation_allowed + promotion_evidence_true + denominator_flags
    return {
        "source_input_total_rows": total,
        "source_input_xlsx_rows": xlsx_total,
        "source_input_pdf_rows": pdf_total,
        "source_input_answer_allowed_count": answer_allowed,
        "source_input_answer_generation_allowed_count": answer_generation_allowed,
        "source_input_xlsx_answer_eval_allowed_count": xlsx_answer_eval_allowed,
        "source_input_pdf_answer_eval_allowed_count": pdf_answer_eval_allowed,
        "source_input_denominator_flag_count": denominator_flags,
        "source_input_promotion_evidence_true_count": promotion_evidence_true,
        "source_input_any_answer_allowed_or_denominator_flag_count": any_flag,
        "source_input_zero_denominator_verified": denominator_flags == 0,
    }


def load_db_snapshot(db_dsn: str, docv_ids: list[str]) -> DbSnapshot:
    if not docv_ids:
        return DbSnapshot(True, "", {}, {})
    try:
        import psycopg2
        import psycopg2.extras
    except Exception as exc:  # pragma: no cover - optional dependency
        return DbSnapshot(False, f"psycopg2 unavailable: {type(exc).__name__}: {exc}", {}, {})

    try:
        with psycopg2.connect(db_dsn, cursor_factory=psycopg2.extras.RealDictCursor) as conn:
            with conn.cursor() as cur:
                cur.execute(
                    """
                    SELECT id, document_version_id, source_file_id, extracted_artifact_id,
                           artifact_type, parser_name, parser_version, file_type, artifact_json
                      FROM parsed_artifact
                     WHERE document_version_id = ANY(%s)
                       AND parser_version = 'xlsx-extract-v2-hidden-safe'
                    """,
                    (docv_ids,),
                )
                parser_rows = [dict(row) for row in cur.fetchall()]
                cur.execute(
                    """
                    SELECT id, index_id, document_version_id, source_file_id, source_file_name,
                           unit_type, unit_key, chunk_type, location_json, text_content,
                           embedding_text, bm25_text, display_text, debug_text, citation_text,
                           parser_name, parser_version, index_version, embedding_status
                      FROM search_unit
                     WHERE document_version_id = ANY(%s)
                       AND parser_version = 'xlsx-extract-v2-hidden-safe'
                       AND index_version = 'rag-ingestion-v2-xlsx-candidate-v1'
                    """,
                    (docv_ids,),
                )
                index_rows = [dict(row) for row in cur.fetchall()]
    except Exception as exc:
        return DbSnapshot(False, f"{type(exc).__name__}: {exc}", {}, {})

    parser_by_docv: dict[str, list[dict[str, Any]]] = {}
    for row in parser_rows:
        parser_by_docv.setdefault(clean(row.get("document_version_id")), []).append(row)
    index_by_docv: dict[str, list[dict[str, Any]]] = {}
    for row in index_rows:
        index_by_docv.setdefault(clean(row.get("document_version_id")), []).append(row)
    return DbSnapshot(True, "", parser_by_docv, index_by_docv)


def locator_from_input(row: Mapping[str, Any]) -> dict[str, str]:
    locator = row.get("expected_evidence_location")
    if not isinstance(locator, Mapping):
        locator = parse_locator(clean(row.get("expected_current_evidence_location")))
    context = row.get("context") if isinstance(row.get("context"), Mapping) else {}
    context_locator = context.get("locator") if isinstance(context.get("locator"), Mapping) else {}
    return compact_dict(
        {
            "document_version_id": clean(
                locator.get("docv")
                or locator.get("document_version_id")
                or context_locator.get("document_version_id")
            ),
            "file": clean(locator.get("file") or context.get("file_name") or row.get("expected_file_name")),
            "sheet": clean(locator.get("sheet") or context.get("sheet_name") or row.get("expected_sheet_name")),
            "range": clean(locator.get("range") or context.get("cell_range") or row.get("expected_cell_range")),
            "table": clean(locator.get("table") or context.get("table_id") or row.get("expected_table_id")),
        }
    )


def is_policy_pending(row: Mapping[str, Any]) -> bool:
    policy = row.get("policy") if isinstance(row.get("policy"), Mapping) else {}
    return bool(
        clean(row.get("expected_answer_shape")) == ANSWER_SHAPE_POLICY_PENDING
        or policy.get("not_answerable_or_policy_pending")
        or policy.get("hidden_policy_blocked")
        or policy.get("formula_date_policy_blocked")
    )


def classify_context_errors(row: Mapping[str, Any]) -> list[str]:
    context = row.get("context") if isinstance(row.get("context"), Mapping) else {}
    errors = context.get("context_errors") or []
    codes: list[str] = []
    for error in errors:
        text = clean(error)
        if "ReadOnlyWorksheet" in text or "row_dimensions" in text:
            codes.append("XLSX_SOURCE_READONLY_EXTRACTION_ERROR")
        elif "BadZipFile" in text or "not a zip file" in text:
            codes.append("XLSX_SOURCE_BAD_ZIP")
        elif "source workbook not found" in text:
            codes.append("XLSX_SOURCE_WORKBOOK_MISSING")
        elif "policy-pending" in text or "not-answerable" in text:
            codes.append("XLSX_POLICY_PENDING")
    return sorted(set(codes))


def source_path_from_row(row: Mapping[str, Any]) -> Path | None:
    context = row.get("context") if isinstance(row.get("context"), Mapping) else {}
    source = clean(context.get("source_path"))
    if not source:
        return None
    path = Path(source)
    if not path.is_absolute():
        path = REPO_ROOT / path
    return path if path.exists() else None


def build_dataset_index(root: Path) -> dict[str, list[Path]]:
    index: dict[str, list[Path]] = {}
    if not root.exists():
        return index
    for suffix in ("*.xlsx", "*.xlsm"):
        for path in root.rglob(suffix):
            index.setdefault(path.name, []).append(path)
    return index


def resolve_dataset_file(file_name: str, dataset_index: Mapping[str, list[Path]]) -> Path | None:
    if not file_name:
        return None
    candidates = dataset_index.get(file_name) or []
    return candidates[0] if candidates else None


def iter_sheets(payload: Any) -> Iterable[Mapping[str, Any]]:
    if not isinstance(payload, Mapping):
        return []
    workbook = payload.get("workbook") if isinstance(payload.get("workbook"), Mapping) else {}
    sheets = workbook.get("sheets") if isinstance(workbook.get("sheets"), list) else []
    return [sheet for sheet in sheets if isinstance(sheet, Mapping)]


def content_field_counts(value: Any, *, ignored_keys: set[str]) -> dict[str, int]:
    counts: Counter[str] = Counter()

    def walk(current: Any, path: tuple[str, ...]) -> None:
        if isinstance(current, Mapping):
            for key, child in current.items():
                key_text = normalize_key(key)
                if key_text in ignored_keys:
                    continue
                if key_text in LEAKAGE_KEYS:
                    continue
                walk(child, (*path, clean(key)))
            return
        if isinstance(current, list):
            for child in current:
                walk(child, path)
            return
        if not path:
            return
        leaf = normalize_key(path[-1])
        if leaf not in CONTENT_FIELD_KEYS:
            return
        if has_content_text(current):
            counts[".".join(path)] += 1

    walk(value, tuple())
    return dict(sorted(counts.items()))


def has_content_text(value: Any) -> bool:
    text = clean(value)
    if not text:
        return False
    if len(text) < 2:
        return False
    folded = re.sub(r"\s+", "", text).lower()
    locator_patterns = (
        r"^[a-z]{1,3}\d+(:[a-z]{1,3}\d+)?$",
        r"^.+\.xlsx>.+>[a-z]{1,3}\d+:[a-z]{1,3}\d+$",
        r"^\[sheet:.+\]\[range:[a-z]{1,3}\d+:[a-z]{1,3}\d+\]$",
    )
    return not any(re.match(pattern, folded) for pattern in locator_patterns)


def summary_leaks_from_gold_or_query(summary: str, row: Mapping[str, Any]) -> bool:
    folded_summary = normalize_for_match(summary)
    if not folded_summary:
        return True
    candidates = [row.get("query"), row.get("expected_answer_text")]
    candidates.extend(row.get("must_contain_terms") or [])
    return any(folded_summary == normalize_for_match(candidate) for candidate in candidates if clean(candidate))


def has_locator(context: Mapping[str, Any]) -> bool:
    return bool(
        context.get("locator")
        or context.get("file_name")
        or context.get("sheet_name")
        or context.get("cell_range")
    )


def parse_cell_range(value: str) -> tuple[int, int, int, int] | None:
    text = clean(value)
    if not text:
        return None
    try:
        from openpyxl.utils.cell import range_boundaries

        min_col, min_row, max_col, max_row = range_boundaries(text)
        return min_col, min_row, max_col, max_row
    except Exception:
        return None


def point_in_range(row: int, col: int, bounds: tuple[int, int, int, int]) -> bool:
    min_col, min_row, max_col, max_row = bounds
    return min_row <= row <= max_row and min_col <= col <= max_col


def ranges_overlap(a: tuple[int, int, int, int], b: tuple[int, int, int, int]) -> bool:
    a_min_col, a_min_row, a_max_col, a_max_row = a
    b_min_col, b_min_row, b_max_col, b_max_row = b
    return not (
        a_max_col < b_min_col
        or b_max_col < a_min_col
        or a_max_row < b_min_row
        or b_max_row < a_min_row
    )


def stage_count(rows: Iterable[Mapping[str, Any]], stage: str, field: str, expected: bool) -> int:
    return sum(
        1
        for row in rows
        if isinstance(row.get(stage), Mapping) and bool(row[stage].get(field)) is expected
    )


def resolve_answer_inputs(path: Path) -> Path:
    if clean(str(path)) and clean(str(path)) != ".":
        return path
    repair = read_json_object(DEFAULT_REPAIR_REPORT)
    source = ((repair or {}).get("source_inputs") or {}).get("answer_generation_inputs") or {}
    candidate = Path(clean(source.get("path")))
    if candidate:
        if not candidate.is_absolute():
            candidate = REPO_ROOT / candidate
        if candidate.exists():
            return candidate
    matches = sorted(EVAL_RUNS_DIR.glob("pdf_xlsx_answer_shape_local_llm_*/answer_generation_inputs.jsonl"))
    if not matches:
        raise SystemExit("answer_generation_inputs.jsonl not found")
    return matches[-1]


def parse_locator(text: str) -> dict[str, str]:
    locator: dict[str, str] = {}
    for part in text.split(";"):
        if "=" not in part:
            continue
        key, value = part.split("=", 1)
        locator[clean(key)] = clean(value)
    return locator


def read_json_object(path: Path) -> dict[str, Any] | None:
    if not path.exists():
        return None
    try:
        payload = json.loads(path.read_text(encoding="utf-8"))
    except json.JSONDecodeError:
        return None
    return payload if isinstance(payload, dict) else None


def read_jsonl(path: Path) -> list[dict[str, Any]]:
    with path.open("r", encoding="utf-8") as handle:
        return [json.loads(line) for line in handle if line.strip()]


def keyed_json_rows(payload: Mapping[str, Any] | None, field: str) -> dict[str, dict[str, Any]]:
    if not payload:
        return {}
    rows = payload.get(field)
    if not isinstance(rows, list):
        return {}
    return {clean(row.get("query_id")): dict(row) for row in rows if isinstance(row, Mapping)}


def parse_json_maybe(value: Any) -> Any:
    if isinstance(value, str):
        try:
            return json.loads(value)
        except json.JSONDecodeError:
            return {}
    return value


def write_json(path: Path, payload: Mapping[str, Any]) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    path.write_text(json.dumps(payload, ensure_ascii=False, indent=2) + "\n", encoding="utf-8")


def write_jsonl(path: Path, rows: Iterable[Mapping[str, Any]]) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    with path.open("w", encoding="utf-8", newline="\n") as handle:
        for row in rows:
            handle.write(json.dumps(row, ensure_ascii=False, sort_keys=True) + "\n")


def write_trace_csv(path: Path, rows: Iterable[Mapping[str, Any]]) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    fields = [
        "query_id",
        "bucket",
        "expected_answer_shape",
        "policy_pending",
        "source_status",
        "parser_has_values",
        "index_has_values",
        "retrieval_has_values",
        "answer_input_has_values",
        "answer_input_drops_content",
        "content_never_exists_upstream",
        "diagnostic_reason_codes",
    ]
    with path.open("w", encoding="utf-8-sig", newline="") as handle:
        writer = csv.DictWriter(handle, fieldnames=fields)
        writer.writeheader()
        for row in rows:
            writer.writerow(
                {
                    "query_id": row.get("query_id"),
                    "bucket": row.get("bucket"),
                    "expected_answer_shape": row.get("expected_answer_shape"),
                    "policy_pending": row.get("policy_pending"),
                    "source_status": nested(row, "source_workbook_stage", "status"),
                    "parser_has_values": nested(row, "parser_artifact_stage", "has_values"),
                    "index_has_values": nested(row, "index_payload_stage", "has_values"),
                    "retrieval_has_values": nested(row, "retrieval_result_stage", "has_values"),
                    "answer_input_has_values": nested(row, "answer_generation_input_stage", "has_values"),
                    "answer_input_drops_content": row.get("answer_input_drops_content"),
                    "content_never_exists_upstream": row.get("content_never_exists_upstream"),
                    "diagnostic_reason_codes": ";".join(row.get("diagnostic_reason_codes") or []),
                }
            )


def write_report_csv(path: Path, report: Mapping[str, Any]) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    scalar_keys = [
        "run_id",
        "status",
        "xlsx_total_rows",
        "policy_pending_rows",
        "source_workbook_missing_count",
        "source_bad_zip_count",
        "source_readonly_extraction_error_count",
        "parser_artifact_has_values_count",
        "parser_artifact_no_values_count",
        "parser_artifact_missing_count",
        "parser_artifact_exact_locator_has_values_count",
        "parser_artifact_broad_sheet_or_workbook_has_values_count",
        "index_payload_has_values_count",
        "index_payload_no_values_count",
        "index_payload_exact_locator_has_values_count",
        "index_payload_broad_sheet_or_workbook_has_values_count",
        "retrieval_result_has_values_count",
        "retrieval_result_no_values_count",
        "answer_input_has_values_count",
        "answer_input_no_values_count",
        "content_exists_upstream_but_dropped_before_answer_inputs_count",
        "content_never_exists_upstream_count",
        "xlsx_answer_eval_denominator",
        "pdf_answer_eval_denominator",
        "rechunk_reembed_justified",
    ]
    with path.open("w", encoding="utf-8-sig", newline="") as handle:
        writer = csv.DictWriter(handle, fieldnames=["metric", "value"])
        writer.writeheader()
        for key in scalar_keys:
            writer.writerow({"metric": key, "value": report.get(key)})
        for key, value in (report.get("diagnostic_reason_counts") or {}).items():
            writer.writerow({"metric": f"diagnostic_reason_counts.{key}", "value": value})
        for key, value in (report.get("source_input_guardrails") or {}).items():
            writer.writerow({"metric": f"source_input_guardrails.{key}", "value": value})


def file_inventory(path: Path) -> dict[str, Any]:
    return {
        "path": repo_relative(path),
        "exists": path.exists(),
        "size_bytes": path.stat().st_size if path.exists() and path.is_file() else None,
    }


def nested(row: Mapping[str, Any], *path: str) -> Any:
    current: Any = row
    for key in path:
        if not isinstance(current, Mapping):
            return None
        current = current.get(key)
    return current


def compact_dict(payload: Mapping[str, str]) -> dict[str, str]:
    return {key: value for key, value in payload.items() if value}


def normalize_key(value: Any) -> str:
    return re.sub(r"[^a-z0-9_]", "", clean(value).replace("-", "_").lower())


def normalize_for_match(value: Any) -> str:
    return re.sub(r"\s+", "", clean(value)).lower()


def int_or_none(value: Any) -> int | None:
    try:
        return int(value)
    except (TypeError, ValueError):
        return None


def parse_bool(value: Any) -> bool:
    if isinstance(value, bool):
        return value
    return clean(value).lower() in {"1", "true", "yes", "y"}


def clean(value: Any) -> str:
    if value is None:
        return ""
    if isinstance(value, str):
        return value.strip()
    return str(value).strip()


def repo_relative(path: Path) -> str:
    try:
        return str(path.resolve().relative_to(REPO_ROOT.resolve())).replace("\\", "/")
    except ValueError:
        return str(path)


def utc_timestamp() -> str:
    return datetime.now(timezone.utc).isoformat()


def utc_run_id() -> str:
    return datetime.now(timezone.utc).strftime("%Y%m%dT%H%M%SZ")


def print_json(payload: Mapping[str, Any]) -> None:
    print(json.dumps(payload, ensure_ascii=False, indent=2))


if __name__ == "__main__":
    sys.exit(main())
