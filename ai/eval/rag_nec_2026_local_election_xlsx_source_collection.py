from __future__ import annotations

import csv
import hashlib
import json
import os
import re
from collections import Counter
from pathlib import Path
from typing import Any, Iterable, Mapping, Sequence

from openpyxl import load_workbook
from openpyxl.utils import get_column_letter

from ai.eval import rag_eval_registry as registry
from ai.eval import rag_v5_diagnostic_common as common


LOGICAL_RUN_KEY = "nec_2026_local_election_xlsx"
SHORT_RUN_ID = "nec_2026_local_election_xlsx_source_collection_diagnostic_nonprod"
CANONICAL_LONG_RUN_ID = SHORT_RUN_ID
STATUS = "NEC_2026_LOCAL_ELECTION_XLSX_SOURCE_COLLECTION_DIAGNOSTIC_NONPROD_READY"

REPORT_ROOT = Path("ai/eval/reports/rag-ingestion")
RUN_ROOT = REPORT_ROOT / "runs" / LOGICAL_RUN_KEY
SHORT_REPORT_PATH = RUN_ROOT / "report.json"
STATUS_JSONL_PATH = REPORT_ROOT / "status.jsonl"
SOURCE_MANIFEST_JSONL_PATH = RUN_ROOT / "source_manifest.jsonl"
WORKBOOK_ARTIFACT_PREVIEW_JSONL_PATH = RUN_ROOT / "workbook_artifact_preview.jsonl"
SEARCH_UNIT_PREVIEW_JSONL_PATH = RUN_ROOT / "search_unit_preview.jsonl"
SOURCE_ATOMS_JSONL_PATH = RUN_ROOT / "source_atoms.jsonl"
SEARCH_VIEWS_JSONL_PATH = RUN_ROOT / "search_views.jsonl"

SOURCE_COLLECTION_ID = "source_collection_20260605_nec_election_results"
SOURCE_COLLECTION_ENV_VAR = "RAG_NEC_2026_LOCAL_ELECTION_SOURCE_COLLECTION_ROOT"
DEFAULT_SOURCE_COLLECTION_ROOT = (
    Path("D:/_external_runtime_artifacts")
    / "async-ocr-rag-multimodal-pipeline"
    / SOURCE_COLLECTION_ID
)
KST_DOC_DATE = "2026-06-05"
CURRENT_ALIAS_STAYS_ON = "v5_6"

REQUIRED_SHEETS = ("source_requests", "raw_display_rows", "parsed_votes", "national_summary")
FORBIDDEN_FALSE_KEYS = (
    "official_metric",
    "official_metric_denominator_usage_allowed",
    "official_metric_dry_run_opened",
    "gold_mutation",
    "qrels_mutation",
    "label_mutation",
    "expected_answer_mutation",
    "supporting_evidence_mutation",
    "denominator_mutation",
    "official_qrels_created",
    "official_relevance_labels_created",
    "official_answerability_labels_created",
    "official_gold_labels_created",
    "training_dataset_created",
    "training_manifest_jsonl_created",
    "training_job_created",
    "fine_tuning_dataset_export_created",
    "fine_tuning",
    "fine_tuning_started",
    "fine_tuning_executed",
    "ft_a_execution",
    "promotion_evidence",
    "product_success_evidence_allowed",
    "live_db_index_cache_readiness",
    "production_db_mutated",
    "source_registry_mutated",
    "silver_mutation",
    "index_rebuilt",
    "cache_mutated",
    "answer_generation_attempted",
    "raw_prompt_payload_written",
    "raw_response_payload_written",
    "raw_xlsx_query_time_parsing",
    "direct_normalized_answer_value_matching",
    "formula_evaluation",
    "formula_text_exposure",
    "source_file_title_shortcut_used",
    "workbook_or_source_title_shortcut_used",
    "target_or_gold_locator_used_for_candidate_construction",
    "query_id_case_id_hack_used",
)
RAW_PAYLOAD_FORBIDDEN_KEYS = {
    "raw_prompt",
    "prompt_payload",
    "raw_response",
    "raw_llm_response",
    "expected_answer",
    "supporting_evidence",
    "gold_locator",
    "target_locator",
}

utc_now_iso = common.utc_now_iso
read_jsonl = common.read_jsonl
write_json = common.write_json
write_jsonl = common.write_jsonl
sha256_file = common.sha256_file


def _clean(value: Any) -> str:
    return "" if value is None else str(value).strip()


def _int(value: Any, default: int = 0) -> int:
    try:
        return int(float(_clean(value).replace(",", "")))
    except (TypeError, ValueError):
        return default


def _json_clone(payload: Mapping[str, Any]) -> dict[str, Any]:
    return common.json_clone(payload)


def _sha256_text(value: str) -> str:
    return hashlib.sha256(value.encode("utf-8")).hexdigest()


def source_collection_root_from_env(env: Mapping[str, str] | None = None) -> Path:
    current_env = os.environ if env is None else env
    configured = _clean(current_env.get(SOURCE_COLLECTION_ENV_VAR))
    return Path(configured) if configured else DEFAULT_SOURCE_COLLECTION_ROOT


def _read_manifest_rows(collection_root: Path) -> list[dict[str, str]]:
    csv_path = collection_root / "manifest.csv"
    if csv_path.exists():
        with csv_path.open(encoding="utf-8-sig", newline="") as handle:
            return [dict(row) for row in csv.DictReader(handle)]

    json_path = collection_root / "manifest.json"
    if json_path.exists():
        payload = json.loads(json_path.read_text(encoding="utf-8"))
        for key in ("files", "rows", "items"):
            rows = payload.get(key)
            if isinstance(rows, list):
                return [dict(row) for row in rows if isinstance(row, Mapping)]

    raise FileNotFoundError(f"NEC source collection manifest not found under {collection_root}")


def _load_verification_items(collection_root: Path) -> dict[str, dict[str, Any]]:
    path = collection_root / "verification.json"
    if not path.exists():
        return {}
    payload = json.loads(path.read_text(encoding="utf-8"))
    items = payload.get("items") or []
    return {str(item.get("file") or ""): dict(item) for item in items if isinstance(item, Mapping)}


def _resolve_workbook_path(collection_root: Path, row: Mapping[str, Any]) -> Path:
    configured = Path(_clean(row.get("xlsx_path")))
    if configured and configured.exists():
        return configured
    filename = configured.name or _clean(row.get("file")) or _clean(row.get("filename"))
    candidate = collection_root / "xlsx" / filename
    if candidate.exists():
        return candidate
    raise FileNotFoundError(f"NEC workbook not found for manifest row: {filename or row!r}")


def _relative_to_collection(collection_root: Path, path: Path) -> str:
    try:
        return path.resolve().relative_to(collection_root.resolve()).as_posix()
    except ValueError:
        return f"xlsx/{path.name}"


def _used_range(row_count: int, column_count: int) -> str:
    if row_count <= 0 or column_count <= 0:
        return "A1:A1"
    return f"A1:{get_column_letter(column_count)}{row_count}"


def _cell_range(start_row: int, end_row: int, column_count: int) -> str:
    return f"A{start_row}:{get_column_letter(max(column_count, 1))}{end_row}"


def _sheet_rows(sheet: Any) -> list[list[Any]]:
    return [list(row) for row in sheet.iter_rows(values_only=True)]


def _headers(rows: Sequence[Sequence[Any]]) -> list[str]:
    if not rows:
        return []
    return [_clean(value) for value in rows[0]]


def _max_width(rows: Sequence[Sequence[Any]]) -> int:
    return max((len(row) for row in rows), default=1)


def _row_records(rows: Sequence[Sequence[Any]]) -> list[dict[str, Any]]:
    headers = _headers(rows)
    records: list[dict[str, Any]] = []
    for offset, row in enumerate(rows[1:], start=2):
        record = {headers[index]: value for index, value in enumerate(row) if index < len(headers)}
        record["_row_number"] = offset
        records.append(record)
    return records


def _contiguous_groups(
    records: Sequence[Mapping[str, Any]],
    key_fields: Sequence[str],
) -> list[tuple[tuple[str, ...], int, int, list[Mapping[str, Any]]]]:
    groups: list[tuple[tuple[str, ...], int, int, list[Mapping[str, Any]]]] = []
    current_key: tuple[str, ...] | None = None
    current_rows: list[Mapping[str, Any]] = []
    for record in records:
        key = tuple(_clean(record.get(field)) for field in key_fields)
        if not any(key):
            continue
        if current_key is not None and key != current_key:
            groups.append(
                (
                    current_key,
                    int(current_rows[0]["_row_number"]),
                    int(current_rows[-1]["_row_number"]),
                    list(current_rows),
                )
            )
            current_rows = []
        current_key = key
        current_rows.append(record)
    if current_key is not None and current_rows:
        groups.append(
            (
                current_key,
                int(current_rows[0]["_row_number"]),
                int(current_rows[-1]["_row_number"]),
                list(current_rows),
            )
        )
    return groups


def _sample_row_text(rows: Sequence[Mapping[str, Any]], fields: Sequence[str]) -> str:
    parts: list[str] = []
    for row in rows[:3]:
        values = [f"{field}={_clean(row.get(field))}" for field in fields if _clean(row.get(field))]
        if values:
            parts.append("; ".join(values))
    return " | ".join(parts)


def _display_value(value: Any) -> str:
    if isinstance(value, bool):
        return "true" if value else "false"
    if isinstance(value, int):
        return f"{value:,}"
    if isinstance(value, float):
        if value.is_integer():
            return f"{int(value):,}"
        return str(value)
    return _clean(value)


def _unit_id(workbook_sha256: str, sheet: str, cell_range: str, unit_type: str, key: str) -> str:
    digest = _sha256_text("|".join((workbook_sha256, sheet, cell_range, unit_type, key)))
    return f"nec2026:{digest[:24]}"


def _unit(
    *,
    workbook_filename: str,
    workbook_sha256: str,
    election_code: str,
    election_label: str,
    sheet: str,
    cell_range: str,
    unit_type: str,
    key: str,
    embedding_text: str,
) -> dict[str, Any]:
    search_unit_id = _unit_id(workbook_sha256, sheet, cell_range, unit_type, key)
    return {
        "schema_version": f"{SHORT_RUN_ID}_search_unit_preview_v1",
        "source_collection_id": SOURCE_COLLECTION_ID,
        "source_family": "XLSX",
        "route_key": LOGICAL_RUN_KEY,
        "election_code": election_code,
        "election_label": election_label,
        "workbook": workbook_filename,
        "workbook_sha256": workbook_sha256,
        "sheet": sheet,
        "range": cell_range,
        "unit_type": unit_type,
        "unit_key": key,
        "search_unit_id": search_unit_id,
        "locator_fingerprint": _sha256_text(f"{workbook_sha256}|{sheet}|{cell_range}|{key}"),
        "embedding_text": embedding_text,
        "citation": {
            "source_family": "XLSX",
            "workbook": workbook_filename,
            "sheet": sheet,
            "range": cell_range,
            "search_unit_id": search_unit_id,
        },
    }


def _source_atom_id(workbook_sha256: str, sheet: str, cell_range: str, cell: str, row_label: str, target_column: str) -> str:
    digest = _sha256_text("|".join((workbook_sha256, sheet, cell_range, cell, row_label, target_column)))
    return f"nec2026_atom:{digest[:24]}"


def _source_identity(workbook_filename: str) -> str:
    return f"{SOURCE_COLLECTION_ID}/{workbook_filename}"


def _column_cell(headers: Sequence[str], column_name: str, row_number: int) -> str:
    try:
        index = headers.index(column_name) + 1
    except ValueError:
        index = max(len(headers), 1)
    return f"{get_column_letter(index)}{row_number}"


def _source_atom(
    *,
    workbook_filename: str,
    workbook_sha256: str,
    election_code: str,
    election_label: str,
    sheet: str,
    cell_range: str,
    cell: str,
    row_label: str,
    column_label: str,
    target_column: str,
    raw_value: Any,
    extra_fields: Mapping[str, Any],
) -> dict[str, Any]:
    normalized_value = _clean(raw_value)
    display_value = _display_value(raw_value)
    identity = _source_identity(workbook_filename)
    atom_id = _source_atom_id(workbook_sha256, sheet, cell_range, cell, row_label, target_column)
    search_unit_id = _unit_id(workbook_sha256, sheet, cell_range, "source_atom_value", atom_id)
    locator_fingerprint = _sha256_text(f"{identity}|{sheet}|{cell_range}|{cell}|{row_label}|{target_column}")
    raw_locator = {
        "workbook": workbook_filename,
        "sheet": sheet,
        "range": cell_range,
        "cell": cell,
        "row_label": row_label,
        "column_label": column_label,
        "target_column": target_column,
        "value_locator": f"{cell}={display_value}",
    }
    citation_payload = {
        "source_family": "XLSX",
        "source_identity": identity,
        "document_version_id": workbook_sha256,
        "locator_fingerprint": locator_fingerprint,
        "search_unit_id": search_unit_id,
        "workbook": workbook_filename,
        "sheet": sheet,
        "range": cell_range,
        "cell": cell,
        "row_label": row_label,
        "target_column": target_column,
        "normalized_value": normalized_value,
    }
    display_metadata = {
        "display_value": display_value,
        "raw_value": raw_value,
        "normalized_value": normalized_value,
        "number_format": "",
        "value_type": type(raw_value).__name__,
    }
    return {
        "schema_version": f"{SHORT_RUN_ID}_source_atom_v1",
        "source_atom_id": atom_id,
        "source_family": "XLSX",
        "source_identity": identity,
        "content_hash": _sha256_text(json.dumps({"locator": raw_locator, "value": normalized_value}, ensure_ascii=False, sort_keys=True)),
        "extraction_version": "nec_2026_local_election_xlsx_preview_v1",
        "workbook_id": workbook_sha256,
        "workbook_version_id": workbook_sha256,
        "document_version_id": workbook_sha256,
        "source_collection_id": SOURCE_COLLECTION_ID,
        "route_key": LOGICAL_RUN_KEY,
        "election_code": election_code,
        "election_label": election_label,
        "workbook": workbook_filename,
        "sheet": sheet,
        "range": cell_range,
        "cell": cell,
        "row_label": row_label,
        "column_label": column_label,
        "target_column": target_column,
        "normalized_text_or_value_snapshot": normalized_value,
        "raw_locator": raw_locator,
        "canonical_citation_payload": citation_payload,
        "xlsx_display_contract": "display_value_raw_value_normalized_value_preview_v1",
        "xlsx_display_metadata": display_metadata,
        "metadata": dict(extra_fields),
    }


def _search_view_for_atom(atom: Mapping[str, Any]) -> dict[str, Any]:
    locator = atom["raw_locator"]
    metadata = atom.get("metadata") or {}
    field_lines = [
        f"workbook={atom['workbook']}",
        f"sheet={atom['sheet']}",
        f"range={atom['range']}",
        f"cell={atom['cell']}",
        f"row_label={locator['row_label']}",
        f"column_label={locator['column_label']}",
        f"target_column={locator['target_column']}",
        f"normalized_value={atom['normalized_text_or_value_snapshot']}",
        f"display_value={atom['xlsx_display_metadata']['display_value']}",
    ]
    for key in ("candidate_name", "party_name", "party_group", "votes", "vote_share_pct", "rank", "metric"):
        value = _clean(metadata.get(key))
        if value:
            field_lines.append(f"{key}={value}")
    text = "\n".join(field_lines)
    return {
        "schema_version": f"{SHORT_RUN_ID}_search_view_v1",
        "search_view_id": f"nec2026_view:{_sha256_text(atom['source_atom_id'])[:24]}",
        "source_family": "XLSX",
        "source_atom_ids": [atom["source_atom_id"]],
        "source_atom_id": atom["source_atom_id"],
        "search_unit_id": atom["canonical_citation_payload"]["search_unit_id"],
        "workbook": atom["workbook"],
        "sheet": atom["sheet"],
        "range": atom["range"],
        "cell": atom["cell"],
        "bm25_text": text,
        "embedding_text": text,
        "display_text": text,
        "retrieval_default": True,
        "candidate_only": True,
    }


def _workbook_atoms_and_views(
    *,
    workbook_filename: str,
    workbook_sha256: str,
    election_code: str,
    election_label: str,
    sheet_payloads: Mapping[str, Sequence[Sequence[Any]]],
) -> tuple[list[dict[str, Any]], list[dict[str, Any]]]:
    atoms: list[dict[str, Any]] = []

    parsed_rows = sheet_payloads.get("parsed_votes") or []
    parsed_headers = _headers(parsed_rows)
    for record in _row_records(parsed_rows):
        row_number = int(record["_row_number"])
        row_range = _cell_range(row_number, row_number, _max_width(parsed_rows))
        row_label = _clean(record.get("contest_label")) or _clean(record.get("district_name"))
        candidate = _clean(record.get("candidate_name")) or _clean(record.get("label"))
        column_label = candidate or row_label
        atom = _source_atom(
            workbook_filename=workbook_filename,
            workbook_sha256=workbook_sha256,
            election_code=election_code,
            election_label=election_label,
            sheet="parsed_votes",
            cell_range=row_range,
            cell=_column_cell(parsed_headers, "votes", row_number),
            row_label=row_label,
            column_label=column_label,
            target_column="votes",
            raw_value=record.get("votes"),
            extra_fields={
                "request_id": _clean(record.get("request_id")),
                "contest_label": row_label,
                "candidate_name": candidate,
                "party_name": _clean(record.get("party_name")),
                "party_group": _clean(record.get("party_group")),
                "votes": _clean(record.get("votes")),
                "vote_share_pct": _clean(record.get("vote_share_pct")),
                "rank": _clean(record.get("rank")),
            },
        )
        atoms.append(atom)

    summary_rows = sheet_payloads.get("national_summary") or []
    summary_headers = _headers(summary_rows)
    for record in _row_records(summary_rows):
        row_number = int(record["_row_number"])
        metric = _clean(record.get("metric"))
        if not metric:
            continue
        row_range = _cell_range(row_number, row_number, _max_width(summary_rows))
        atoms.append(
            _source_atom(
                workbook_filename=workbook_filename,
                workbook_sha256=workbook_sha256,
                election_code=election_code,
                election_label=election_label,
                sheet="national_summary",
                cell_range=row_range,
                cell=_column_cell(summary_headers, "value", row_number),
                row_label=metric,
                column_label="value",
                target_column="value",
                raw_value=record.get("value"),
                extra_fields={"metric": metric, "value": _clean(record.get("value"))},
            )
        )

    return atoms, [_search_view_for_atom(atom) for atom in atoms]


def _workbook_units(
    *,
    workbook_filename: str,
    workbook_sha256: str,
    election_code: str,
    election_label: str,
    sheet_payloads: Mapping[str, Sequence[Sequence[Any]]],
) -> list[dict[str, Any]]:
    units: list[dict[str, Any]] = []

    source_rows = sheet_payloads.get("source_requests") or []
    source_records = _row_records(source_rows)
    for (request_id,), start, end, rows in _contiguous_groups(source_records, ("request_id",)):
        cell_range = _cell_range(start, end, _max_width(source_rows))
        text = (
            f"Election {election_code} {election_label}\n"
            f"Workbook {workbook_filename}\nSheet source_requests\nRange {cell_range}\n"
            f"Request {request_id}\n{_sample_row_text(rows, ('city_name', 'city_code', 'election_label'))}"
        )
        units.append(
            _unit(
                workbook_filename=workbook_filename,
                workbook_sha256=workbook_sha256,
                election_code=election_code,
                election_label=election_label,
                sheet="source_requests",
                cell_range=cell_range,
                unit_type="source_request",
                key=request_id,
                embedding_text=text,
            )
        )

    raw_rows = sheet_payloads.get("raw_display_rows") or []
    raw_records = _row_records(raw_rows)
    for (request_id,), start, end, rows in _contiguous_groups(raw_records, ("request_id",)):
        cell_range = _cell_range(start, end, _max_width(raw_rows))
        text = (
            f"Election {election_code} {election_label}\n"
            f"Workbook {workbook_filename}\nSheet raw_display_rows\nRange {cell_range}\n"
            f"Request {request_id}\n{_sample_row_text(rows, tuple(_headers(raw_rows)[1:5]))}"
        )
        units.append(
            _unit(
                workbook_filename=workbook_filename,
                workbook_sha256=workbook_sha256,
                election_code=election_code,
                election_label=election_label,
                sheet="raw_display_rows",
                cell_range=cell_range,
                unit_type="raw_display_request_block",
                key=request_id,
                embedding_text=text,
            )
        )

    parsed_rows = sheet_payloads.get("parsed_votes") or []
    parsed_records = _row_records(parsed_rows)
    for key_tuple, start, end, rows in _contiguous_groups(parsed_records, ("request_id", "contest_label")):
        request_id, contest_label = key_tuple
        cell_range = _cell_range(start, end, _max_width(parsed_rows))
        text = (
            f"Election {election_code} {election_label}\n"
            f"Workbook {workbook_filename}\nSheet parsed_votes\nRange {cell_range}\n"
            f"Request {request_id}\nContest {contest_label}\n"
            f"{_sample_row_text(rows, ('candidate_name', 'party_name', 'party_group', 'votes', 'vote_share_pct', 'rank'))}"
        )
        units.append(
            _unit(
                workbook_filename=workbook_filename,
                workbook_sha256=workbook_sha256,
                election_code=election_code,
                election_label=election_label,
                sheet="parsed_votes",
                cell_range=cell_range,
                unit_type="parsed_votes_contest_span",
                key=f"{request_id}|{contest_label}",
                embedding_text=text,
            )
        )

    summary_rows = sheet_payloads.get("national_summary") or []
    if len(summary_rows) > 1:
        cell_range = _used_range(len(summary_rows), _max_width(summary_rows))
        records = _row_records(summary_rows)
        text = (
            f"Election {election_code} {election_label}\n"
            f"Workbook {workbook_filename}\nSheet national_summary\nRange {cell_range}\n"
            f"{_sample_row_text(records, ('metric', 'value'))}"
        )
        units.append(
            _unit(
                workbook_filename=workbook_filename,
                workbook_sha256=workbook_sha256,
                election_code=election_code,
                election_label=election_label,
                sheet="national_summary",
                cell_range=cell_range,
                unit_type="national_summary_table",
                key="national_summary",
                embedding_text=text,
            )
        )
    return units


def _inspect_workbook(
    collection_root: Path,
    row: Mapping[str, Any],
) -> tuple[dict[str, Any], dict[str, Any], list[dict[str, Any]], list[dict[str, Any]], list[dict[str, Any]]]:
    workbook_path = _resolve_workbook_path(collection_root, row)
    workbook_filename = workbook_path.name
    actual_sha = sha256_file(workbook_path)
    expected_sha = _clean(row.get("xlsx_sha256"))
    if expected_sha and actual_sha != expected_sha:
        raise ValueError(f"NEC workbook SHA-256 mismatch: {workbook_filename}")

    workbook = load_workbook(workbook_path, data_only=True, read_only=False)
    sheet_payloads: dict[str, list[list[Any]]] = {}
    sheet_summaries: list[dict[str, Any]] = []
    hidden_sheet_count = 0
    native_table_count = 0
    try:
        for sheet_name in workbook.sheetnames:
            sheet = workbook[sheet_name]
            if sheet.sheet_state != "visible":
                hidden_sheet_count += 1
                continue
            rows = _sheet_rows(sheet)
            sheet_payloads[sheet_name] = rows
            table_count = len(getattr(sheet, "tables", {}) or {})
            native_table_count += table_count
            sheet_summaries.append(
                {
                    "sheet": sheet_name,
                    "visible": True,
                    "row_count": max(len(rows) - 1, 0),
                    "column_count": _max_width(rows),
                    "used_range": _used_range(len(rows), _max_width(rows)),
                    "native_excel_table_count": table_count,
                }
            )
    finally:
        workbook.close()

    missing_sheets = [sheet for sheet in REQUIRED_SHEETS if sheet not in sheet_payloads]
    if missing_sheets:
        raise ValueError(f"NEC workbook missing required sheets {missing_sheets}: {workbook_filename}")

    election_code = _clean(row.get("election_code"))
    election_label = _clean(row.get("election_label"))
    units = _workbook_units(
        workbook_filename=workbook_filename,
        workbook_sha256=actual_sha,
        election_code=election_code,
        election_label=election_label,
        sheet_payloads=sheet_payloads,
    )
    atoms, search_views = _workbook_atoms_and_views(
        workbook_filename=workbook_filename,
        workbook_sha256=actual_sha,
        election_code=election_code,
        election_label=election_label,
        sheet_payloads=sheet_payloads,
    )
    source_manifest_row = {
        "schema_version": f"{SHORT_RUN_ID}_source_manifest_v1",
        "source_collection_id": SOURCE_COLLECTION_ID,
        "source_family": "XLSX",
        "election_code": election_code,
        "election_label": election_label,
        "workbook": workbook_filename,
        "relative_path": _relative_to_collection(collection_root, workbook_path),
        "xlsx_sha256": actual_sha,
        "xlsx_sha256_verified": True,
        "source_request_count": _int(row.get("source_request_count"), len(sheet_payloads["source_requests"]) - 1),
        "parsed_vote_rows": _int(row.get("parsed_vote_rows"), len(sheet_payloads["parsed_votes"]) - 1),
        "contest_count": _int(row.get("contest_count")),
        "summed_voters": _int(row.get("summed_voters")),
        "summed_ballots": _int(row.get("summed_ballots")),
    }
    workbook_artifact = {
        "schema_version": f"{SHORT_RUN_ID}_workbook_artifact_preview_v1",
        "source_collection_id": SOURCE_COLLECTION_ID,
        "fileType": "xlsx",
        "pipelineVersion": "xlsx-extract-v2-hidden-safe-synthetic-chunks-preview",
        "workbook": workbook_filename,
        "xlsx_sha256": actual_sha,
        "election_code": election_code,
        "election_label": election_label,
        "sheets": sheet_summaries,
        "hidden_sheet_count": hidden_sheet_count,
        "native_excel_table_count": native_table_count,
        "synthetic_search_unit_count": len(units),
    }
    return source_manifest_row, workbook_artifact, units, atoms, search_views


def _code4_provenance_warnings(
    manifest_rows: Sequence[Mapping[str, Any]],
    verification_items: Mapping[str, Mapping[str, Any]],
) -> list[dict[str, Any]]:
    warnings: list[dict[str, Any]] = []
    for row in manifest_rows:
        if _clean(row.get("election_code")) != "4":
            continue
        file_name = Path(_clean(row.get("xlsx_path"))).name
        verification = verification_items.get(file_name) or {}
        manifest_matched = _int(row.get("detailed_map_matched_contest_count"), -1)
        verified_matched = _int(
            verification.get("actual_district_rendered_parsed_vote_unit_count")
            or verification.get("detailed_map_matched_parsed_vote_contest_count"),
            -1,
        )
        if manifest_matched >= 0 and verified_matched >= 0 and manifest_matched != verified_matched:
            warnings.append(
                {
                    "file": file_name,
                    "field": "detailed_map_matched_contest_count",
                    "manifest_value": manifest_matched,
                    "verification_value": verified_matched,
                    "policy": "trust_verification_json_and_code4_sidecars_for_actual_district_rendering",
                }
            )
    return warnings


def _collection_summary(
    *,
    collection_root: Path,
    manifest_rows: Sequence[Mapping[str, Any]],
    source_rows: Sequence[Mapping[str, Any]],
    workbook_artifacts: Sequence[Mapping[str, Any]],
    verification_items: Mapping[str, Mapping[str, Any]],
) -> dict[str, Any]:
    manifest_csv = collection_root / "manifest.csv"
    manifest_json = collection_root / "manifest.json"
    visible_sheet_count = sum(len(artifact["sheets"]) for artifact in workbook_artifacts)
    native_table_count = sum(int(artifact["native_excel_table_count"]) for artifact in workbook_artifacts)
    hidden_sheet_count = sum(int(artifact["hidden_sheet_count"]) for artifact in workbook_artifacts)
    return {
        "schema_version": f"{SHORT_RUN_ID}_source_collection_summary_v1",
        "source_collection_id": SOURCE_COLLECTION_ID,
        "source_collection_root_redacted": True,
        "manifest_csv_available": manifest_csv.exists(),
        "manifest_csv_sha256": sha256_file(manifest_csv) if manifest_csv.exists() else "",
        "manifest_json_available": manifest_json.exists(),
        "manifest_json_sha256": sha256_file(manifest_json) if manifest_json.exists() else "",
        "verification_json_available": bool(verification_items),
        "workbook_count": len(source_rows),
        "manifest_row_count": len(manifest_rows),
        "verified_xlsx_count": sum(1 for row in source_rows if row.get("xlsx_sha256_verified") is True),
        "visible_sheet_count": visible_sheet_count,
        "hidden_sheet_count": hidden_sheet_count,
        "native_excel_table_count": native_table_count,
        "required_sheets": list(REQUIRED_SHEETS),
        "code4_provenance_warnings": _code4_provenance_warnings(manifest_rows, verification_items),
        "workbooks": list(source_rows),
    }


def _chunk_summary(units: Sequence[Mapping[str, Any]]) -> dict[str, Any]:
    counts = Counter(_clean(unit.get("unit_type")) for unit in units)
    return {
        "schema_version": f"{SHORT_RUN_ID}_synthetic_chunk_summary_v1",
        "source_request_chunk_count": counts["source_request"],
        "raw_display_request_block_count": counts["raw_display_request_block"],
        "parsed_votes_contest_span_count": counts["parsed_votes_contest_span"],
        "national_summary_table_count": counts["national_summary_table"],
        "search_unit_preview_rows": len(units),
        "chunking_policy": "source_requests_by_request_id_raw_display_by_request_block_parsed_votes_by_request_id_and_contest_label",
    }


def _source_atom_summary(atoms: Sequence[Mapping[str, Any]]) -> dict[str, Any]:
    counts = Counter(_clean(atom.get("sheet")) for atom in atoms)
    return {
        "schema_version": f"{SHORT_RUN_ID}_source_atom_summary_v1",
        "source_atom_rows": len(atoms),
        "parsed_votes_atom_rows": counts["parsed_votes"],
        "national_summary_atom_rows": counts["national_summary"],
        "retrieval_default_included_sheets": ["parsed_votes", "national_summary"],
        "retrieval_default_excluded_sheets": ["source_requests", "raw_display_rows"],
        "source_atom_policy": "value_granular_atoms_from_parsed_votes_and_national_summary_only",
    }


def _search_view_summary(search_views: Sequence[Mapping[str, Any]]) -> dict[str, Any]:
    return {
        "schema_version": f"{SHORT_RUN_ID}_search_view_summary_v1",
        "search_view_rows": len(search_views),
        "candidate_only": True,
        "default_source_sheets": ["parsed_votes", "national_summary"],
        "bm25_embedding_text_policy": (
            "fielded_locator_text_with_workbook_sheet_range_cell_row_label_target_column_normalized_and_display_value"
        ),
    }


def build_report(
    *,
    root: Path,
    source_collection_root: Path | str | None = None,
    generated_at: str | None = None,
    check: bool = True,
) -> dict[str, Any]:
    collection_root = Path(source_collection_root) if source_collection_root is not None else source_collection_root_from_env()
    manifest_rows = _read_manifest_rows(collection_root)
    verification_items = _load_verification_items(collection_root)
    source_rows: list[dict[str, Any]] = []
    workbook_artifacts: list[dict[str, Any]] = []
    search_units: list[dict[str, Any]] = []
    source_atoms: list[dict[str, Any]] = []
    search_views: list[dict[str, Any]] = []
    for manifest_row in manifest_rows:
        source_row, artifact, units, atoms, views = _inspect_workbook(collection_root, manifest_row)
        source_rows.append(source_row)
        workbook_artifacts.append(artifact)
        search_units.extend(units)
        source_atoms.extend(atoms)
        search_views.extend(views)

    report: dict[str, Any] = {
        "schema_version": f"{SHORT_RUN_ID}_report_v1",
        "logical_run_key": LOGICAL_RUN_KEY,
        "run_id": SHORT_RUN_ID,
        "short_run_id": SHORT_RUN_ID,
        "canonical_long_run_id": CANONICAL_LONG_RUN_ID,
        "status": STATUS,
        "generated_at": generated_at or utc_now_iso(),
        "artifact_paths": {
            "report_json": SHORT_REPORT_PATH.as_posix(),
            "status_jsonl": STATUS_JSONL_PATH.as_posix(),
            "source_manifest_jsonl": SOURCE_MANIFEST_JSONL_PATH.as_posix(),
            "workbook_artifact_preview_jsonl": WORKBOOK_ARTIFACT_PREVIEW_JSONL_PATH.as_posix(),
            "search_unit_preview_jsonl": SEARCH_UNIT_PREVIEW_JSONL_PATH.as_posix(),
            "source_atoms_jsonl": SOURCE_ATOMS_JSONL_PATH.as_posix(),
            "search_views_jsonl": SEARCH_VIEWS_JSONL_PATH.as_posix(),
        },
        "artifact_sha256": {},
        "current_resolves_to": CURRENT_ALIAS_STAYS_ON,
        "route_direct_only": True,
        "diagnostic_only": True,
        "non_production": True,
        "source_collection_registration_only": True,
        "xlsx_source_collection_preview": True,
        "source_collection": _collection_summary(
            collection_root=collection_root,
            manifest_rows=manifest_rows,
            source_rows=source_rows,
            workbook_artifacts=workbook_artifacts,
            verification_items=verification_items,
        ),
        "synthetic_chunk_summary": _chunk_summary(search_units),
        "source_atom_summary": _source_atom_summary(source_atoms),
        "search_view_summary": _search_view_summary(search_views),
        "search_unit_preview_sample": search_units[:25],
        "source_atom_preview_sample": source_atoms[:25],
        "search_view_preview_sample": search_views[:25],
        "workbook_artifact_preview_sample": workbook_artifacts[:10],
        "official_metric": False,
        "official_metric_denominator_usage_allowed": False,
        "official_metric_input_rows": 0,
        "official_metric_input_rows_created": 0,
        "official_metric_input_rows_scope": "nec_2026_local_election_diagnostic_route_created_rows_only",
        "official_metric_dry_run_opened": False,
        "silver_official_metric_input_rows": 0,
        "silver_promoted_to_gold_count": 0,
        "gold_mutation": False,
        "qrels_mutation": False,
        "label_mutation": False,
        "expected_answer_mutation": False,
        "supporting_evidence_mutation": False,
        "denominator_mutation": False,
        "official_qrels_created": False,
        "official_relevance_labels_created": False,
        "official_answerability_labels_created": False,
        "official_gold_labels_created": False,
        "training_dataset_created": False,
        "training_manifest_jsonl_created": False,
        "training_job_created": False,
        "fine_tuning_dataset_export_created": False,
        "fine_tuning": False,
        "fine_tuning_started": False,
        "fine_tuning_executed": False,
        "ft_a_execution": False,
        "promotion_evidence": False,
        "product_success_evidence_allowed": False,
        "live_db_index_cache_readiness": False,
        "production_db_mutated": False,
        "source_registry_mutated": False,
        "silver_mutation": False,
        "index_rebuilt": False,
        "cache_mutated": False,
        "protected_namespaces_touched": [],
        "SearchView_vector_payload_role": "candidate_preview_only",
        "SourceAtom_EvidenceBundle_role": "source_collection_preview_not_evidence_truth_promotion",
        "answer_generation_attempted": False,
        "generated_response_count": 0,
        "raw_prompt_payload_written": False,
        "raw_response_payload_written": False,
        "raw_xlsx_query_time_parsing": False,
        "direct_normalized_answer_value_matching": False,
        "formula_evaluation": False,
        "formula_text_exposure": False,
        "source_file_title_shortcut_used": False,
        "workbook_or_source_title_shortcut_used": False,
        "target_or_gold_locator_used_for_candidate_construction": False,
        "query_id_case_id_hack_used": False,
        "decision_policy": {
            "route_policy": "new_direct_diagnostic_route_current_alias_unchanged",
            "path_policy": "external_absolute_paths_redacted_from_report_and_status",
            "chunk_policy": "synthetic_chunks_required_because_nec_workbooks_have_no_native_excel_tables",
            "code4_policy": "verification_json_and_code4_sidecars_override_stale_manifest_aggregate_fields",
        },
        "residual_risks": [
            "license/use policy remains metadata-only; this route does not assert commercial or publication rights",
            "search units are preview artifacts only and are not written to live indexes",
            "code-4 map aggregate fields in manifest.csv/json include stale counters; verification.json sidecars are authoritative",
        ],
        "next_recommendations": [
            "run this route with --write to materialize ignored preview artifacts",
            "review search_unit_preview.jsonl before opening any nonprod vector-index build",
            "keep official metric, gold/qrels, denominator, training, promotion, and live readiness gates closed",
        ],
        "counters": {
            "current_resolves_to": CURRENT_ALIAS_STAYS_ON,
            "official_metric_input_rows": 0,
            "official_metric_input_rows_created": 0,
            "workbook_count": len(source_rows),
            "verified_xlsx_count": sum(1 for row in source_rows if row.get("xlsx_sha256_verified") is True),
            "search_unit_preview_rows": len(search_units),
            "source_atom_rows": len(source_atoms),
            "search_view_rows": len(search_views),
            "parsed_votes_contest_span_count": Counter(_clean(unit.get("unit_type")) for unit in search_units)[
                "parsed_votes_contest_span"
            ],
        },
        "_source_manifest_rows": source_rows,
        "_workbook_artifact_rows": workbook_artifacts,
        "_search_unit_preview_rows": search_units,
        "_source_atom_rows": source_atoms,
        "_search_view_rows": search_views,
    }
    if check:
        check_report(report)
    return report


def _public_report(report: Mapping[str, Any], artifact_hashes: Mapping[str, str] | None = None) -> dict[str, Any]:
    payload = {key: value for key, value in report.items() if not key.startswith("_")}
    if artifact_hashes is not None:
        payload["artifact_sha256"] = dict(artifact_hashes)
    return payload


def write_report_bundle(root: Path | str, report: Mapping[str, Any]) -> tuple[dict[str, Any], dict[str, str]]:
    repo_root = Path(root)
    source_rows = list(report.get("_source_manifest_rows") or [])
    artifact_rows = list(report.get("_workbook_artifact_rows") or [])
    search_units = list(report.get("_search_unit_preview_rows") or [])
    source_atoms = list(report.get("_source_atom_rows") or [])
    search_views = list(report.get("_search_view_rows") or [])
    write_jsonl(repo_root / SOURCE_MANIFEST_JSONL_PATH, source_rows)
    write_jsonl(repo_root / WORKBOOK_ARTIFACT_PREVIEW_JSONL_PATH, artifact_rows)
    write_jsonl(repo_root / SEARCH_UNIT_PREVIEW_JSONL_PATH, search_units)
    write_jsonl(repo_root / SOURCE_ATOMS_JSONL_PATH, source_atoms)
    write_jsonl(repo_root / SEARCH_VIEWS_JSONL_PATH, search_views)

    hashes = {
        "source_manifest_jsonl_sha256": sha256_file(repo_root / SOURCE_MANIFEST_JSONL_PATH),
        "workbook_artifact_preview_jsonl_sha256": sha256_file(repo_root / WORKBOOK_ARTIFACT_PREVIEW_JSONL_PATH),
        "search_unit_preview_jsonl_sha256": sha256_file(repo_root / SEARCH_UNIT_PREVIEW_JSONL_PATH),
        "source_atoms_jsonl_sha256": sha256_file(repo_root / SOURCE_ATOMS_JSONL_PATH),
        "search_views_jsonl_sha256": sha256_file(repo_root / SEARCH_VIEWS_JSONL_PATH),
    }
    public = _public_report(report, artifact_hashes=hashes)
    write_json(repo_root / SHORT_REPORT_PATH, public)
    hashes["report_json_sha256"] = sha256_file(repo_root / SHORT_REPORT_PATH)
    public["artifact_sha256"] = hashes
    write_json(repo_root / SHORT_REPORT_PATH, public)
    return public, hashes


def status_event(report: Mapping[str, Any], *, artifact_hashes: Mapping[str, str]) -> dict[str, Any]:
    source = report["source_collection"]
    chunks = report["synthetic_chunk_summary"]
    atoms = report["source_atom_summary"]
    views = report["search_view_summary"]
    return {
        "schema_version": f"{SHORT_RUN_ID}_status_event_v1",
        "event_type": "diagnostic_nec_2026_local_election_xlsx_source_collection_nonprod",
        "run_id": SHORT_RUN_ID,
        "logical_run_key": LOGICAL_RUN_KEY,
        "short_run_id": SHORT_RUN_ID,
        "canonical_long_run_id": CANONICAL_LONG_RUN_ID,
        "status": STATUS,
        "generated_at": report["generated_at"],
        "artifact_paths": dict(report["artifact_paths"]),
        "artifact_sha256": dict(artifact_hashes),
        "current_resolves_to": CURRENT_ALIAS_STAYS_ON,
        "diagnostic_only": True,
        "non_production": True,
        "route_direct_only": True,
        "source_collection_root_redacted": True,
        "workbook_count": source["workbook_count"],
        "verified_xlsx_count": source["verified_xlsx_count"],
        "visible_sheet_count": source["visible_sheet_count"],
        "native_excel_table_count": source["native_excel_table_count"],
        "source_request_chunk_count": chunks["source_request_chunk_count"],
        "raw_display_request_block_count": chunks["raw_display_request_block_count"],
        "parsed_votes_contest_span_count": chunks["parsed_votes_contest_span_count"],
        "search_unit_preview_rows": chunks["search_unit_preview_rows"],
        "source_atom_rows": atoms["source_atom_rows"],
        "search_view_rows": views["search_view_rows"],
        "retrieval_default_included_sheets": atoms["retrieval_default_included_sheets"],
        "retrieval_default_excluded_sheets": atoms["retrieval_default_excluded_sheets"],
        "code4_provenance_warning_count": len(source["code4_provenance_warnings"]),
        "official_metric": False,
        "official_metric_input_rows": 0,
        "official_metric_input_rows_created": 0,
        "gold_mutation": False,
        "qrels_mutation": False,
        "label_mutation": False,
        "expected_answer_mutation": False,
        "supporting_evidence_mutation": False,
        "denominator_mutation": False,
        "training_dataset_created": False,
        "training_manifest_jsonl_created": False,
        "training_job_created": False,
        "fine_tuning_dataset_export_created": False,
        "fine_tuning": False,
        "fine_tuning_started": False,
        "fine_tuning_executed": False,
        "ft_a_execution": False,
        "promotion_evidence": False,
        "product_success_evidence_allowed": False,
        "live_db_index_cache_readiness": False,
        "protected_namespaces_touched": [],
        "source_registry_mutated": False,
        "index_rebuilt": False,
        "raw_xlsx_query_time_parsing": False,
        "direct_normalized_answer_value_matching": False,
    }


def append_status(root: Path | str, report: Mapping[str, Any], *, artifact_hashes: Mapping[str, str]) -> None:
    repo_root = Path(root)
    path = repo_root / STATUS_JSONL_PATH
    rows = read_jsonl(path) if path.exists() else []
    event_type = "diagnostic_nec_2026_local_election_xlsx_source_collection_nonprod"
    rows = [
        row
        for row in rows
        if row.get("run_id") not in {SHORT_RUN_ID, CANONICAL_LONG_RUN_ID}
        and row.get("canonical_long_run_id") != CANONICAL_LONG_RUN_ID
        and row.get("event_type") != event_type
    ]
    rows.append(status_event(report, artifact_hashes=artifact_hashes))
    write_jsonl(path, rows)


def _upsert_block_at_top(text: str, *, start_marker: str, end_marker: str, block: str) -> str:
    return common.upsert_block_at_top(text, start_marker=start_marker, end_marker=end_marker, block=block)


def _sync_last_updated(text: str) -> str:
    return common.sync_last_updated(text, KST_DOC_DATE)


def update_docs(root: Path | str, report: Mapping[str, Any]) -> None:
    repo_root = Path(root)
    progress = repo_root / "docs/rag-ingestion-progress.md"
    measurements = repo_root / "docs/rag-ingestion-measurements.md"
    triage = repo_root / "docs/rag-ingestion-triage.md"
    scripts_readme = repo_root / "ai/scripts/README.md"
    source = report["source_collection"]
    chunks = report["synthetic_chunk_summary"]
    atoms = report["source_atom_summary"]
    views = report["search_view_summary"]

    progress_block = (
        f"- Overall status: `{STATUS}`; {SHORT_RUN_ID} opens a direct 2026 NEC local-election XLSX diagnostic "
        f"route. Artifact: `{SHORT_REPORT_PATH.as_posix()}`. Route key: `{LOGICAL_RUN_KEY}`; `current` remains "
        f"on `{CURRENT_ALIAS_STAYS_ON}` and is not moved by this route. The route verifies "
        f"{source['verified_xlsx_count']}/{source['workbook_count']} workbooks from the external source collection, "
        f"materializes preview-only source manifest/workbook/search-unit/source-atom/search-view artifacts, and creates "
        f"{chunks['search_unit_preview_rows']} synthetic XLSX search-unit preview rows, including "
        f"{chunks['parsed_votes_contest_span_count']} parsed-votes contest spans. Retrieval-default source atoms/search "
        f"views are {atoms['source_atom_rows']}/{views['search_view_rows']} rows from parsed_votes/national_summary only. "
        f"Native Excel table count is "
        f"{source['native_excel_table_count']}; synthetic chunks are required for contest-level retrieval. "
        "official_metric_input_rows=0, official_metric_input_rows_created=0, no gold/qrels/label/expected/"
        "supporting/denominator/training/fine-tuning/FT-A/promotion/product-success/live-readiness gates are opened."
    )
    progress_text = _upsert_block_at_top(
        progress.read_text(encoding="utf-8"),
        start_marker=f"<!-- {SHORT_RUN_ID}:progress-entry:start -->",
        end_marker=f"<!-- {SHORT_RUN_ID}:progress-entry:end -->",
        block=progress_block,
    )
    progress.write_text(_sync_last_updated(progress_text), encoding="utf-8")

    measurements_block = f"""## NEC 2026 local-election XLSX source collection route

- Run key: `{LOGICAL_RUN_KEY}`
- Primary artifact: `{SHORT_REPORT_PATH.as_posix()}`
- Interpretation: direct diagnostic source-collection route only; `current` remains `{CURRENT_ALIAS_STAYS_ON}`.

| counter | value |
| --- | --- |
| status | {STATUS} |
| workbook_count | {source['workbook_count']} |
| verified_xlsx_count | {source['verified_xlsx_count']} |
| visible_sheet_count | {source['visible_sheet_count']} |
| native_excel_table_count | {source['native_excel_table_count']} |
| source_request_chunk_count | {chunks['source_request_chunk_count']} |
| raw_display_request_block_count | {chunks['raw_display_request_block_count']} |
| parsed_votes_contest_span_count | {chunks['parsed_votes_contest_span_count']} |
| search_unit_preview_rows | {chunks['search_unit_preview_rows']} |
| source_atom_rows | {atoms['source_atom_rows']} |
| search_view_rows | {views['search_view_rows']} |
| retrieval_default_included_sheets | {json.dumps(atoms['retrieval_default_included_sheets'], ensure_ascii=False)} |
| retrieval_default_excluded_sheets | {json.dumps(atoms['retrieval_default_excluded_sheets'], ensure_ascii=False)} |
| code4_provenance_warning_count | {len(source['code4_provenance_warnings'])} |
| official_metric_input_rows | 0 |
| official_metric_input_rows_created | 0 |
| source_registry_mutated | false |
| index_rebuilt | false |
| live_db_index_cache_readiness | false |
"""
    measurements_text = _upsert_block_at_top(
        measurements.read_text(encoding="utf-8"),
        start_marker=f"<!-- {SHORT_RUN_ID}:measurements-entry:start -->",
        end_marker=f"<!-- {SHORT_RUN_ID}:measurements-entry:end -->",
        block=measurements_block,
    )
    measurements.write_text(_sync_last_updated(measurements_text), encoding="utf-8")

    triage_block = (
        "### NEC 2026 local-election XLSX source collection route\n\n"
        f"- Scope: `{LOGICAL_RUN_KEY}` is a direct diagnostic route for the newly downloaded NEC XLSX workbooks. "
        f"It does not move `current` from `{CURRENT_ALIAS_STAYS_ON}`.\n"
        f"- Retrieval shape: parsed votes are chunked by contiguous `(request_id, contest_label)` spans "
        f"({chunks['parsed_votes_contest_span_count']} preview spans), raw display rows by request block, and "
        "source requests by request id. Default candidate search views use parsed_votes/national_summary atoms only; "
        "source_requests and raw_display_rows are excluded from retrieval-default views.\n"
        f"- Code-4 caution: {len(source['code4_provenance_warnings'])} stale manifest/verification counter warning(s) "
        "are recorded; verification.json and code-4 sidecars are the authority for actual district rendering.\n"
        "- Fail-closed status: official_metric_input_rows=0, source_registry_mutated=false, index_rebuilt=false, "
        "training_dataset_created=false, fine_tuning_dataset_export_created=false, protected_namespaces_touched=[]."
    )
    triage_text = _upsert_block_at_top(
        triage.read_text(encoding="utf-8"),
        start_marker=f"<!-- {SHORT_RUN_ID}:triage-entry:start -->",
        end_marker=f"<!-- {SHORT_RUN_ID}:triage-entry:end -->",
        block=triage_block,
    )
    triage.write_text(_sync_last_updated(triage_text), encoding="utf-8")

    scripts_text = scripts_readme.read_text(encoding="utf-8")
    row = (
        "| `rag_eval.py nec_2026_local_election_xlsx` | Direct 2026 NEC local-election XLSX diagnostic route; "
        "verifies the external source collection and writes preview-only source manifest, workbook artifact, and "
        "synthetic search-unit/source-atom/search-view artifacts while leaving `current`, official/gold/qrels/denominator/training/"
        "promotion/live gates closed. |"
    )
    route_row_pattern = r"\| `rag_eval\.py nec_2026_local_election_xlsx` \|.*?\|\n"
    if re.search(route_row_pattern, scripts_text):
        scripts_text = re.sub(route_row_pattern, row + "\n", scripts_text, count=1)
    else:
        scripts_text = re.sub(r"(\| `rag_eval.py` \|.*?\|\n)", r"\1" + row + "\n", scripts_text, count=1)
    scripts_readme.write_text(scripts_text, encoding="utf-8")


def _assert_no_raw_payload_keys(value: Any) -> None:
    common.assert_no_raw_payload_keys(value, RAW_PAYLOAD_FORBIDDEN_KEYS, context=LOGICAL_RUN_KEY)


def _assert_no_absolute_external_path(value: Any) -> None:
    serialized = json.dumps(value, ensure_ascii=False)
    if re.search(r"[A-Za-z]:[\\/]", serialized):
        raise ValueError("NEC route report/status must redact absolute local paths")


def check_report(report: Mapping[str, Any]) -> None:
    public = _public_report(report)
    if public.get("short_run_id") != SHORT_RUN_ID:
        raise ValueError("NEC route short_run_id mismatch")
    if public.get("logical_run_key") != LOGICAL_RUN_KEY:
        raise ValueError("NEC route logical run key mismatch")
    if public.get("status") != STATUS:
        raise ValueError("NEC route status mismatch")
    if public.get("current_resolves_to") != CURRENT_ALIAS_STAYS_ON:
        raise ValueError("NEC route must not move current")
    if registry.resolve_run("current").logical_key != CURRENT_ALIAS_STAYS_ON:
        raise ValueError("registry current alias drifted during NEC route")
    if public.get("diagnostic_only") is not True or public.get("non_production") is not True:
        raise ValueError("NEC route must remain diagnostic-only and non-production")
    for key in FORBIDDEN_FALSE_KEYS:
        if public.get(key) is not False:
            raise ValueError(f"NEC route opened forbidden gate: {key}")
    if public.get("official_metric_input_rows") != 0 or public.get("official_metric_input_rows_created") != 0:
        raise ValueError("NEC route opened official metric rows")
    if public.get("protected_namespaces_touched") != []:
        raise ValueError("NEC route touched protected namespaces")

    source = public.get("source_collection") or {}
    chunks = public.get("synthetic_chunk_summary") or {}
    if source.get("source_collection_root_redacted") is not True:
        raise ValueError("NEC route source collection root must stay redacted")
    if source.get("workbook_count", 0) <= 0:
        raise ValueError("NEC route requires at least one workbook")
    if source.get("verified_xlsx_count") != source.get("workbook_count"):
        raise ValueError("NEC route requires all XLSX hashes to verify")
    if source.get("hidden_sheet_count") != 0:
        raise ValueError("NEC route encountered hidden sheets")
    if not set(REQUIRED_SHEETS).issubset(set(source.get("required_sheets") or [])):
        raise ValueError("NEC route required sheet contract drift")
    if chunks.get("parsed_votes_contest_span_count", 0) <= 0:
        raise ValueError("NEC route missing parsed-votes contest span chunks")
    if chunks.get("search_unit_preview_rows", 0) <= 0:
        raise ValueError("NEC route missing search-unit preview rows")
    atoms = public.get("source_atom_summary") or {}
    views = public.get("search_view_summary") or {}
    if atoms.get("source_atom_rows", 0) <= 0:
        raise ValueError("NEC route missing source atoms")
    if views.get("search_view_rows") != atoms.get("source_atom_rows"):
        raise ValueError("NEC route source atoms/search views row-count mismatch")
    if atoms.get("retrieval_default_included_sheets") != ["parsed_votes", "national_summary"]:
        raise ValueError("NEC route retrieval-default included sheets drift")
    if atoms.get("retrieval_default_excluded_sheets") != ["source_requests", "raw_display_rows"]:
        raise ValueError("NEC route retrieval-default excluded sheets drift")
    sample = public.get("search_unit_preview_sample") or []
    if not sample:
        raise ValueError("NEC route search-unit preview sample missing")
    for unit in sample:
        citation = unit.get("citation") or {}
        for key in ("source_family", "workbook", "sheet", "range", "search_unit_id"):
            if not _clean(citation.get(key)):
                raise ValueError(f"NEC route sample citation missing {key}")
        if citation.get("source_family") != "XLSX":
            raise ValueError("NEC route sample citation source family drift")
    for atom in public.get("source_atom_preview_sample") or []:
        citation = atom.get("canonical_citation_payload") or {}
        locator = atom.get("raw_locator") or {}
        for key in ("source_identity", "locator_fingerprint", "search_unit_id", "workbook", "sheet", "range"):
            if not _clean(citation.get(key)):
                raise ValueError(f"NEC route source atom citation missing {key}")
        for key in ("row_label", "target_column", "value_locator"):
            if not _clean(locator.get(key)):
                raise ValueError(f"NEC route source atom locator missing {key}")
    for view in public.get("search_view_preview_sample") or []:
        text = f"{view.get('bm25_text', '')}\n{view.get('embedding_text', '')}"
        for term in ("workbook=", "sheet=", "range=", "row_label=", "target_column=", "normalized_value="):
            if term not in text:
                raise ValueError(f"NEC route search view missing fielded locator text: {term}")
    _assert_no_raw_payload_keys(public)
    _assert_no_absolute_external_path(public)
