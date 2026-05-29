from __future__ import annotations

import argparse
import csv
import hashlib
import json
import re
import zipfile
from collections import Counter, defaultdict
from dataclasses import dataclass
from datetime import date, datetime
from html import escape
from pathlib import Path
from typing import Any, Iterable, Mapping, Sequence

from openpyxl import load_workbook
from openpyxl.utils import get_column_letter

from rag_local_llm_expected_answer_generation_v1 import (
    DEFAULT_BACKEND as DEFAULT_LLM_BACKEND,
    DEFAULT_MODEL as DEFAULT_LLM_MODEL,
    call_local_llm_strict_json,
    local_llm_entry_blockers,
    resolve_base_url as resolve_llm_base_url,
)
import rag_v4_7_preofficial_external_holdout_candidate_manifest_registration_nonprod as v47


ROOT = v47.ROOT
REPORT_DIR = v47.REPORT_DIR
STATUS_JSONL = v47.STATUS_JSONL
PROGRESS_DOC = v47.PROGRESS_DOC
MEASUREMENTS_DOC = v47.MEASUREMENTS_DOC
TRIAGE_DOC = v47.TRIAGE_DOC
README = v47.README
EVAL_README = v47.EVAL_README
SCRIPTS_README = ROOT / "ai" / "scripts" / "README.md"

V4_NAME = v47.V4_NAME
V4_RUN_FAMILY = v47.V4_RUN_FAMILY
RUN_ID = "official_answer_citation_agentic_loop_run_v4_7_2_source_grounded_korean_query_review_packet_hydration_nonprod"
EVENT_TYPE = "diagnostic_v4_7_2_source_grounded_korean_query_review_packet_hydration_nonprod"
STATUS = "DIAGNOSTIC_V4_7_2_SOURCE_GROUNDED_KOREAN_QUERY_REVIEW_PACKET_HYDRATION_NONPROD_READY"
REPORT_SCHEMA_VERSION = "rag_v4_7_2_source_grounded_korean_query_review_packet_hydration_report_v1"
OUTPUT_DIR = REPORT_DIR / "quality" / RUN_ID

EXTERNAL_SOURCE_COLLECTION_DIR = Path(
    r"D:\_external_runtime_artifacts\async-ocr-rag-multimodal-pipeline\source_collection_20260510"
)
DEFAULT_CANDIDATE_MANIFEST = (
    EXTERNAL_SOURCE_COLLECTION_DIR / "v4_7_preofficial_external_holdout_candidate_manifest_registration_input.jsonl"
)
SOURCE_COLLECTION_MANIFEST_CSV = EXTERNAL_SOURCE_COLLECTION_DIR / "manifest.csv"
V4_7_1_RUN_ID = "official_answer_citation_agentic_loop_run_v4_7_1_korean_review_packet_and_readme_status_snapshot_nonprod"
V4_7_1_PACKET_CSV = REPORT_DIR / "quality" / V4_7_1_RUN_ID / "review_packet_ko.csv"

EXPECTED_V4_7_MANIFEST_SHA256 = "15b2f5f61a03bf588bf49d74a95a11259e2a6a83c0a32a727625344cae7af58c"
V4_7_SOURCE_RUN_ID = v47.RUN_ID

REPORT_JSON = OUTPUT_DIR / "report.json"
REVIEW_PACKET_XLSX = OUTPUT_DIR / "review_packet_ko_hydrated.xlsx"
REVIEW_PACKET_CSV = OUTPUT_DIR / "review_packet_ko_hydrated.csv"
REVIEW_PACKET_JSONL = OUTPUT_DIR / "review_packet_ko_hydrated.jsonl"
REVIEW_GUIDELINES = OUTPUT_DIR / "review_guidelines_ko.md"
REVIEW_SUMMARY_JSON = OUTPUT_DIR / "review_summary_ko.json"

ALLOWED_ARTIFACT_NAMES = {
    "report.json",
    "review_packet_ko_hydrated.xlsx",
    "review_packet_ko_hydrated.csv",
    "review_packet_ko_hydrated.jsonl",
    "review_guidelines_ko.md",
    "review_summary_ko.json",
}

HUMAN_REVIEW_COLUMNS = [
    "검수상태",
    "소스계열",
    "후보ID",
    "질의ID",
    "질의문",
    "질의자연성",
    "질의승인",
    "질의보존성",
    "관련성라벨",
    "답변가능성라벨",
    "기대답변_한국어",
    "근거판단_한국어",
    "근거위치_확인",
    "공식분모포함판단",
    "제외사유",
    "정책메모",
    "검수자",
    "검수일시",
    "재검수필요",
]

MACHINE_HELPER_COLUMNS = [
    "질의생성방식",
    "질의생성근거",
    "질의생성신뢰도",
    "질의생성경고",
    "근거후보_스니펫",
    "근거후보_위치",
    "근거후보_충분성_기계판단",
    "기대답변_초안_비공식",
    "source_family",
    "candidate_id_hash",
    "query_id_hash",
    "document_or_workbook_identity_hash",
    "source_identity_kind",
    "source_disjointness_gate",
    "query_fidelity_included",
    "leakage_bucket",
    "prior_identity_collision",
    "manifest_sha256",
    "source_manifest_title",
    "source_manifest_lane",
    "source_manifest_subtype",
    "source_manifest_role",
    "source_preview_redacted",
    "evidence_preview_redacted",
    "locator_preview_redacted",
    "page_or_sheet_locator_redacted",
    "source_report_run_id",
]

XLSX_COLUMNS = [
    "워크북명_표시",
    "시트명_표시",
    "근거후보_범위",
    "근거후보_표시값_미리보기",
    "근거후보_헤더_미리보기",
    "근거후보_행열축_미리보기",
]

PDF_COLUMNS = [
    "문서명_표시",
    "페이지_후보",
    "섹션_후보",
    "문단_후보",
]

REVIEW_COLUMNS = HUMAN_REVIEW_COLUMNS + MACHINE_HELPER_COLUMNS + XLSX_COLUMNS + PDF_COLUMNS + ["machine_notes"]

EXCLUSION_REASONS = [
    "질의의도불명확",
    "근거부족",
    "관련성부족",
    "답변불가",
    "중복질의",
    "문서정체성불명확",
    "워크북정체성불명확",
    "누출위험",
    "경로노출위험",
    "소스불일치",
    "범위과대",
    "정책판단필요",
    "기타",
]

SHEET_NAMES = ["검수_대상_전체", "PDF_검수", "XLSX_검수", "추출실패_검수", "라벨_가이드", "제외_사유_가이드", "요약"]
QUERY_FIELD_NAMES = (
    "query_text",
    "actual_user_query",
    "user_query",
    "natural_query",
    "korean_query",
    "question",
    "generated_query",
    "candidate_query",
    "질의문",
)
FORBIDDEN_TEXT_PATTERNS = (
    r"\bD:[\\/]",
    r"\bD:\\",
    r"v4_7_external_pdf_document_sha256_",
    r"v4_7_external_xlsx_workbook_sha256_",
    r"source_identity_key",
    r"target_locator",
    r"gold_locator",
    r"expected_answer",
    r"supporting_evidence",
    r"official_metric_input_rows\.jsonl",
    r"prompt_payload",
    r"raw_llm_response",
    r"checkpoint",
)


@dataclass(frozen=True)
class SourceContext:
    kind: str
    snippet: str
    locator: str
    title: str
    topic: str
    confidence: str
    sufficiency: str
    generation_basis: str
    warnings: tuple[str, ...] = ()
    page: str = ""
    section: str = ""
    paragraph: str = ""
    workbook: str = ""
    sheet: str = ""
    cell_range: str = ""
    value_preview: str = ""
    header_preview: str = ""
    axis_preview: str = ""


@dataclass(frozen=True)
class LLMQueryDraft:
    query: str
    confidence: str
    warning: str
    response_sha256: str


def clean(value: Any) -> str:
    return v47.clean(value)


def utc_now() -> str:
    return v47.utc_now()


def sha256_file(path: Path) -> str:
    return v47.sha256_file(path)


def repo_relative(path: Path) -> str:
    return v47.repo_relative(path)


def read_json(path: Path) -> dict[str, Any]:
    return json.loads(path.read_text(encoding="utf-8"))


def read_jsonl(path: Path) -> list[dict[str, Any]]:
    return v47.read_jsonl(path)


def write_json(path: Path, payload: Mapping[str, Any]) -> None:
    v47.write_json(path, payload)


def write_jsonl(path: Path, rows: Sequence[Mapping[str, Any]]) -> None:
    v47.write_jsonl(path, rows)


def _hash(value: str) -> str:
    return hashlib.sha256(value.encode("utf-8")).hexdigest()


def _bounded(value: Any, *, limit: int = 420) -> str:
    text = _sanitize_packet_text(value)
    if len(text) <= limit:
        return text
    return text[: limit - 3].rstrip() + "..."


def _sanitize_packet_text(value: Any) -> str:
    text = clean(value)
    if not text:
        return ""
    text = re.sub(r"\s+", " ", text).strip()
    text = text.replace("\\", "/")
    text = re.sub(r"\b[A-Za-z]:/[^\s,;]+", "__local_path_redacted__", text)
    text = re.sub(r"v4_7_external_(?:pdf_document|xlsx_workbook)_sha256_[0-9a-f]{64}", "__source_identity_redacted__", text)
    text = text.replace("source_identity_key", "source_identity_redacted")
    return text


def _display_title(metadata: Mapping[str, str]) -> str:
    title = _sanitize_packet_text(metadata.get("title"))
    if title.startswith("Extracted ") and " from " in title:
        title = title.split(" from ", 1)[0].removeprefix("Extracted ").strip()
    return _bounded(title or metadata.get("role") or "source", limit=150)


def _manifest_sha_from_identity(identity: str) -> str:
    match = re.search(r"([0-9a-f]{64})$", clean(identity))
    return match.group(1) if match else ""


def _source_identity(row: Mapping[str, Any]) -> tuple[str, str]:
    family = clean(row.get("source_family")).upper()
    if family == "PDF":
        return "PDF_source_document", clean(row.get("source_document_id") or row.get("document_id"))
    if family == "XLSX":
        return "XLSX_workbook", clean(row.get("workbook_id") or row.get("source_workbook_id"))
    return "unsupported", ""


def read_source_collection_manifest(path: Path) -> dict[str, dict[str, str]]:
    if not path.exists():
        return {}
    with path.open(encoding="utf-8-sig", newline="") as handle:
        rows = list(csv.DictReader(handle))
    return {
        clean(row.get("sha256")): {key: clean(value) for key, value in row.items()}
        for row in rows
        if clean(row.get("sha256"))
    }


def _source_path(base_dir: Path, metadata: Mapping[str, str]) -> Path:
    relative = clean(metadata.get("relative_path"))
    return base_dir / relative


def _source_preview(metadata: Mapping[str, str]) -> str:
    title = _display_title(metadata)
    lane = _sanitize_packet_text(metadata.get("lane"))
    subtype = _sanitize_packet_text(metadata.get("subtype"))
    role = _sanitize_packet_text(metadata.get("role"))
    return _bounded(f"title={title}; lane={lane}; subtype={subtype}; role={role}", limit=360)


def _candidate_manifest_hash(path: Path) -> str:
    return sha256_file(path) if path.exists() else ""


def _is_placeholder_query(text: str) -> bool:
    return bool(re.fullmatch(r"v4_7_(?:pdf|xlsx)_query_\d+_\d+", clean(text)))


def _existing_query_from_row(row: Mapping[str, Any], prior_row: Mapping[str, str] | None) -> str:
    for field_name in QUERY_FIELD_NAMES:
        value = clean(row.get(field_name))
        if value and not _is_placeholder_query(value):
            return _bounded(value, limit=220)
    if prior_row:
        for field_name in QUERY_FIELD_NAMES:
            value = clean(prior_row.get(field_name))
            if value and not _is_placeholder_query(value):
                return _bounded(value, limit=220)
    return ""


def read_prior_packet_by_candidate(path: Path = V4_7_1_PACKET_CSV) -> dict[str, dict[str, str]]:
    if not path.exists():
        return {}
    with path.open(encoding="utf-8-sig", newline="") as handle:
        rows = list(csv.DictReader(handle))
    return {clean(row.get("후보ID")): row for row in rows if clean(row.get("후보ID"))}


def prior_packet_counters(path: Path = V4_7_1_PACKET_CSV) -> dict[str, int]:
    if not path.exists():
        return {"prior_packet_row_count": 0, "prior_packet_non_empty_query_count": 0}
    with path.open(encoding="utf-8-sig", newline="") as handle:
        rows = list(csv.DictReader(handle))
    return {
        "prior_packet_row_count": len(rows),
        "prior_packet_non_empty_query_count": sum(1 for row in rows if clean(row.get("질의문"))),
    }


def _topic_from_text(text: str, *, fallback: str) -> str:
    normalized = _sanitize_packet_text(text)
    normalized = re.sub(r"[·•●■□▶※]+", " ", normalized)
    clauses = [part.strip(" -:;,.") for part in re.split(r"[.。:：\n\r]| {2,}", normalized) if part.strip()]
    for clause in clauses:
        if len(clause) >= 8 and not re.fullmatch(r"[0-9\s/%.,()-]+", clause):
            words = clause.split()
            if len(words) > 14:
                clause = " ".join(words[:14])
            return _bounded(clause, limit=90)
    return _bounded(fallback, limit=90)


def _text_quality(text: str) -> int:
    if not text:
        return 0
    korean = len(re.findall(r"[가-힣]", text))
    latin = len(re.findall(r"[A-Za-z]", text))
    digits = len(re.findall(r"\d", text))
    return korean * 3 + latin + digits


def _pdf_page_indexes(page_count: int) -> list[int]:
    indexes = list(range(min(page_count, 12)))
    for extra in (page_count // 2, page_count - 1):
        if 0 <= extra < page_count and extra not in indexes:
            indexes.append(extra)
    return indexes[:16]


def extract_pdf_contexts(
    *,
    source_path: Path,
    metadata: Mapping[str, str],
    contexts_needed: int = 5,
) -> list[SourceContext]:
    title = _display_title(metadata)
    contexts: list[SourceContext] = []
    warnings: list[str] = []
    try:
        import fitz  # type: ignore

        document = fitz.open(str(source_path))
        for page_index in _pdf_page_indexes(document.page_count):
            page = document[page_index]
            blocks = page.get_text("blocks") or []
            for block_index, block in enumerate(blocks):
                text = _bounded(block[4] if len(block) >= 5 else "", limit=520)
                if len(text) < 24 or _text_quality(text) < 20:
                    continue
                topic = _topic_from_text(text, fallback=title)
                page_no = page_index + 1
                contexts.append(
                    SourceContext(
                        kind="pdf_text",
                        snippet=text,
                        locator=f"page={page_no}; block={block_index}; extraction=pymupdf_text",
                        title=title,
                        topic=topic,
                        confidence="high" if len(text) >= 80 else "medium",
                        sufficiency="충분후보" if len(text) >= 80 else "부분후보",
                        generation_basis=f"pymupdf bounded text block from page {page_no}",
                        page=str(page_no),
                        section=topic,
                        paragraph=f"block={block_index}",
                    )
                )
        document.close()
    except Exception as exc:  # pragma: no cover - exercised by missing local parser states.
        warnings.append(f"PDF_TEXT_EXTRACTION_WARNING:{type(exc).__name__}")

    contexts.sort(key=lambda context: (-_text_quality(context.snippet), context.locator))
    deduped: list[SourceContext] = []
    seen_topics: set[str] = set()
    for context in contexts:
        topic_key = context.topic[:40]
        if topic_key in seen_topics:
            continue
        seen_topics.add(topic_key)
        deduped.append(context)
        if len(deduped) >= contexts_needed:
            break
    if deduped:
        return deduped

    notes = _sanitize_packet_text(metadata.get("notes"))
    snippet = _bounded(f"{_source_preview(metadata)}; notes={notes}", limit=520)
    return [
        SourceContext(
            kind="pdf_metadata_fallback",
            snippet=snippet,
            locator="source_collection_manifest_metadata; page=metadata",
            title=title,
            topic=_topic_from_text(f"{title} {notes}", fallback=title),
            confidence="low",
            sufficiency="부족후보",
            generation_basis="source_collection manifest metadata fallback because bounded PDF text was unavailable",
            warnings=tuple(warnings + ["PDF_TEXT_EXTRACTION_UNAVAILABLE"]),
            page="metadata",
            section="metadata",
            paragraph="metadata",
        )
    ]


def _safe_cell_value(value: Any) -> str:
    if value is None:
        return ""
    if isinstance(value, (datetime, date)):
        return value.isoformat(sep=" ")
    text = _sanitize_packet_text(value)
    if text.startswith("="):
        return "__formula_value_redacted__"
    return text


def _non_empty_cells(row: Sequence[str]) -> list[tuple[int, str]]:
    return [(index, value) for index, value in enumerate(row, start=1) if clean(value)]


def _sheet_matrix(ws: Any, *, max_rows: int = 36, max_cols: int = 14) -> list[list[str]]:
    row_limit = min(int(ws.max_row or 1), max_rows)
    col_limit = min(int(ws.max_column or 1), max_cols)
    matrix: list[list[str]] = []
    for raw_row in ws.iter_rows(min_row=1, max_row=row_limit, max_col=col_limit, values_only=True):
        matrix.append([_safe_cell_value(value) for value in raw_row])
    return matrix


def _first_meaningful_row(matrix: Sequence[Sequence[str]]) -> int:
    for index, row in enumerate(matrix):
        cells = [value for value in row if clean(value)]
        if len(cells) >= 2 and _text_quality(" ".join(cells)) >= 10:
            return index
    return 0


def _preview_rows(matrix: Sequence[Sequence[str]], *, limit_rows: int = 5) -> str:
    rows: list[str] = []
    for row in matrix:
        cells = [value for value in row if clean(value)]
        if cells:
            rows.append(" | ".join(cells[:6]))
        if len(rows) >= limit_rows:
            break
    return _bounded(" / ".join(rows), limit=520)


def extract_xlsx_contexts(
    *,
    source_path: Path,
    metadata: Mapping[str, str],
    contexts_needed: int = 13,
) -> list[SourceContext]:
    workbook_title = _display_title(metadata)
    contexts: list[SourceContext] = []
    warnings: list[str] = []
    try:
        workbook = load_workbook(source_path, read_only=True, data_only=True)
        for ws in workbook.worksheets:
            matrix = _sheet_matrix(ws)
            if not matrix or not any(clean(value) for row in matrix for value in row):
                continue
            header_index = _first_meaningful_row(matrix)
            headers = list(matrix[header_index])
            header_preview = _bounded(" | ".join(value for value in headers if clean(value)), limit=300)
            if not header_preview:
                header_preview = _bounded(ws.title, limit=120)
            sheet_title = _bounded(ws.title, limit=120)
            range_end_row = min(int(ws.max_row or len(matrix) or 1), header_index + 10)
            range_end_col = min(int(ws.max_column or len(headers) or 1), max(len(headers), 1), 8)
            range_ref = f"A{header_index + 1}:{get_column_letter(max(range_end_col, 1))}{max(range_end_row, header_index + 1)}"
            preview = _preview_rows(matrix[header_index : header_index + 8])
            data_rows = matrix[header_index + 1 :] or matrix[header_index : header_index + 1]
            for row_offset, row in enumerate(data_rows, start=header_index + 2):
                cells = _non_empty_cells(row)
                if len(cells) < 2:
                    continue
                label_index, row_label = cells[0]
                value_index, _value = cells[min(1, len(cells) - 1)]
                col_header = clean(headers[value_index - 1]) if value_index - 1 < len(headers) else ""
                if not col_header or col_header == row_label:
                    col_header = clean(headers[label_index]) if label_index < len(headers) else ""
                if not col_header:
                    col_header = f"{get_column_letter(value_index)}열"
                axis_preview = _bounded(f"행축={row_label}; 열축={col_header}", limit=220)
                topic = _topic_from_text(f"{sheet_title} {row_label} {col_header}", fallback=sheet_title)
                contexts.append(
                    SourceContext(
                        kind="xlsx_table_preview",
                        snippet=preview,
                        locator=f"sheet={sheet_title}; range={range_ref}; extraction=openpyxl_data_only",
                        title=workbook_title,
                        topic=topic,
                        confidence="high" if preview and header_preview else "medium",
                        sufficiency="충분후보" if preview and header_preview else "부분후보",
                        generation_basis="openpyxl read_only data_only sheet/range preview",
                        workbook=workbook_title,
                        sheet=sheet_title,
                        cell_range=range_ref,
                        value_preview=preview,
                        header_preview=header_preview,
                        axis_preview=axis_preview,
                    )
                )
                if len(contexts) >= contexts_needed * 3:
                    break
            if len(contexts) >= contexts_needed * 3:
                break
        workbook.close()
    except Exception as exc:  # pragma: no cover - depends on local workbook/parser state.
        warnings.append(f"XLSX_EXTRACTION_WARNING:{type(exc).__name__}")

    if contexts:
        return contexts[:contexts_needed]

    title_topic = _topic_from_text(_source_preview(metadata), fallback=workbook_title)
    return [
        SourceContext(
            kind="xlsx_extraction_failed",
            snippet=_bounded(_source_preview(metadata), limit=420),
            locator="source_collection_manifest_metadata; workbook_structure_unavailable",
            title=workbook_title,
            topic=title_topic,
            confidence="low",
            sufficiency="부족후보",
            generation_basis="source_collection manifest metadata fallback because workbook structure was unavailable",
            warnings=tuple(warnings + ["XLSX_STRUCTURE_EXTRACTION_UNAVAILABLE"]),
            workbook=workbook_title,
            sheet="",
            cell_range="",
        )
    ]


def _pdf_query(title: str, context: SourceContext, ordinal: int) -> str:
    if context.kind == "pdf_metadata_fallback":
        templates = [
            "{title} 문서가 어떤 자료인지 확인할 수 있는 식별 정보는 무엇인가요?",
            "{title} 문서의 수집 메타데이터에서 확인되는 자료 유형은 무엇인가요?",
            "{title} 문서가 어떤 기관 또는 업무 맥락의 자료로 등록되어 있나요?",
            "{title} 문서의 메타데이터상 역할과 분류는 무엇으로 정리되어 있나요?",
            "{title} 문서에 대해 사람 검수자가 우선 확인해야 할 source 단서는 무엇인가요?",
        ]
        return templates[(ordinal - 1) % len(templates)].format(title=title)
    templates = [
        "{title}에서 {topic}과 관련해 확인할 수 있는 핵심 내용은 무엇인가요?",
        "{title}의 {section} 부분에서 제시된 주요 항목은 무엇인가요?",
        "{title}에서 {topic}에 대한 근거 후보는 어떤 내용을 담고 있나요?",
        "{title}의 {page}쪽 근거 후보에서 {topic}은 어떻게 설명되어 있나요?",
        "{title}에서 {topic}과 관련된 검토 또는 결과 내용은 무엇인가요?",
    ]
    return templates[(ordinal - 1) % len(templates)].format(
        title=title,
        topic=context.topic,
        section=context.section or context.topic,
        page=f"{context.page}쪽" if context.page.isdigit() else "해당 위치",
    )


def _xlsx_query(context: SourceContext, ordinal: int) -> str:
    row_axis = context.axis_preview
    row_label = ""
    col_label = ""
    if "행축=" in row_axis and "; 열축=" in row_axis:
        row_label = row_axis.split("행축=", 1)[1].split("; 열축=", 1)[0]
        col_label = row_axis.split("; 열축=", 1)[1]
    row_label = _bounded(row_label or "해당 행", limit=80)
    col_label = _bounded(col_label or "해당 지표", limit=80)
    templates = [
        "{workbook}의 {sheet} 시트에서 {row_label} 항목의 {col_label} 값은 무엇인가요?",
        "{workbook}의 {sheet} 표에서 {col_label} 지표는 어떤 항목별로 정리되어 있나요?",
        "{sheet} 시트에서 {row_label} 행은 어떤 값으로 정리되어 있나요?",
        "{workbook}의 {sheet} 시트에서 {row_label}과 {col_label}이 만나는 값을 어떻게 확인할 수 있나요?",
        "{sheet} 표에서 {row_label} 항목과 관련된 주요 열 값은 무엇인가요?",
    ]
    return templates[(ordinal - 1) % len(templates)].format(
        workbook=context.workbook,
        sheet=context.sheet or "해당",
        row_label=row_label,
        col_label=col_label,
    )


def _llm_context_item(context: SourceContext, ordinal: int) -> dict[str, str]:
    item = {
        "ordinal": str(ordinal),
        "context_kind": context.kind,
        "display_title": _bounded(context.title or context.workbook, limit=140),
        "topic_hint": _bounded(context.topic, limit=140),
        "evidence_preview": _bounded(context.snippet, limit=520),
        "locator_preview": _bounded(context.locator, limit=220),
        "page_or_sheet": _bounded(context.page or context.sheet, limit=80),
        "range": _bounded(context.cell_range, limit=80),
        "header_preview": _bounded(context.header_preview, limit=180),
        "axis_preview": _bounded(context.axis_preview, limit=180),
        "value_preview": _bounded(context.value_preview, limit=180),
    }
    return {key: value for key, value in item.items() if clean(value)}


def _llm_prompt(*, family: str, title: str, contexts: Sequence[SourceContext]) -> str:
    items = [_llm_context_item(context, index) for index, context in enumerate(contexts, start=1)]
    family_instruction = (
        "PDF page/block preview를 보고 문서 안에서 답을 찾을 수 있는 자연스러운 한국어 질문을 작성하세요."
        if family == "PDF"
        else "XLSX sheet/range/header/axis/value preview를 보고 사용자가 물을 법한 자연스러운 한국어 질문을 작성하세요."
    )
    return json.dumps(
        {
            "task": "source_grounded_korean_review_query_generation",
            "family": family,
            "display_title": _bounded(title, limit=180),
            "instruction": family_instruction,
            "hard_rules": [
                "Return exactly one JSON object.",
                "Return a queries array with exactly one item per context.",
                "Each item must include ordinal, query, confidence, and warnings.",
                "Queries must be Korean, concrete, source-specific, and human-reviewable.",
                "Do not write expected answers, final labels, qrels, gold evidence, or official metric decisions.",
                "Do not mention raw file paths, hashes, source identity ids, target/gold locators, prompt payloads, or formula text.",
                "Do not ask only for a cell address such as A1 unless the source itself is explicitly a cell-location test.",
                "Do not copy the answer value into the question as the thing to verify.",
            ],
            "allowed_question_styles": [
                "문서/워크북명과 page/sheet context를 함께 언급하는 질문",
                "표의 행/열 축 또는 섹션 주제를 이용한 질문",
                "감사의견, 정책 항목, 통계 지표, 재무/환경/인구/시설 항목처럼 source preview에 드러난 주제 질문",
            ],
            "contexts": items,
            "output_schema": {
                "queries": [
                    {
                        "ordinal": 1,
                        "query": "한국어 질의문",
                        "confidence": "high|medium|low",
                        "warnings": [],
                    }
                ]
            },
        },
        ensure_ascii=False,
    )


def _single_llm_prompt(*, family: str, title: str, context: SourceContext, ordinal: int) -> str:
    prompt = json.loads(_llm_prompt(family=family, title=title, contexts=[context]))
    prompt["contexts"][0]["ordinal"] = str(ordinal)
    prompt["output_schema"] = {
        "query": "한국어 질의문",
        "confidence": "high|medium|low",
        "warnings": [],
    }
    return json.dumps(prompt, ensure_ascii=False)


def _normalize_llm_query(text: Any) -> str:
    query = _sanitize_packet_text(clean(text))
    query = re.sub(r"\s+", " ", query).strip()
    if not query:
        raise ValueError("local LLM query is blank")
    if _is_placeholder_query(query):
        raise ValueError("local LLM query is a placeholder id")
    if len(query) < 8:
        raise ValueError("local LLM query is too short")
    return query


def _confidence(value: Any, fallback: str) -> str:
    confidence = clean(value).lower()
    if confidence in {"high", "medium", "low"}:
        return confidence
    if fallback in {"high", "medium", "low"}:
        return fallback
    return "medium"


def _warnings(value: Any) -> str:
    if isinstance(value, list):
        return "; ".join(_bounded(_sanitize_packet_text(item), limit=80) for item in value if clean(item))
    return _bounded(_sanitize_packet_text(value), limit=240)


def _query_value(payload: Mapping[str, Any]) -> Any:
    for key in ("query", "질의문", "question", "korean_query", "generated_query", "candidate_query"):
        if clean(payload.get(key)):
            return payload.get(key)
    return ""


def _query_items(payload: Mapping[str, Any]) -> list[Mapping[str, Any]]:
    for key in ("queries", "질의문목록", "questions", "items"):
        value = payload.get(key)
        if isinstance(value, list):
            return [item for item in value if isinstance(item, Mapping)]
    return []


def _drafts_from_group_payload(
    *,
    payload: Mapping[str, Any],
    meta: Mapping[str, Any],
    contexts: Sequence[SourceContext],
) -> dict[int, LLMQueryDraft]:
    raw_queries = _query_items(payload)
    if not raw_queries:
        raise ValueError("local LLM group output must include queries array")
    response_sha = clean(meta.get("raw_response_sha256") or meta.get("response_sha256"))
    drafts: dict[int, LLMQueryDraft] = {}
    for item in raw_queries:
        if not isinstance(item, Mapping):
            continue
        ordinal = int(clean(item.get("ordinal")) or "0")
        if ordinal < 1 or ordinal > len(contexts):
            continue
        context = contexts[ordinal - 1]
        drafts[ordinal] = LLMQueryDraft(
            query=_normalize_llm_query(_query_value(item)),
            confidence=_confidence(item.get("confidence"), context.confidence),
            warning=_warnings(item.get("warnings")),
            response_sha256=response_sha,
        )
    expected = set(range(1, len(contexts) + 1))
    if set(drafts) != expected:
        missing = sorted(expected - set(drafts))
        raise ValueError(f"local LLM group output missing ordinals: {missing}")
    return drafts


def _call_group_llm(
    *,
    family: str,
    title: str,
    contexts: Sequence[SourceContext],
    backend: str,
    base_url: str,
    model: str,
    timeout_seconds: int,
    max_tokens: int,
) -> dict[int, LLMQueryDraft]:
    parsed, meta = call_local_llm_strict_json(
        backend=backend,
        base_url=base_url,
        model=model,
        prompt=_llm_prompt(family=family, title=title, contexts=contexts),
        temperature=0.0,
        max_tokens=max_tokens,
        timeout_seconds=timeout_seconds,
    )
    return _drafts_from_group_payload(payload=parsed, meta=meta, contexts=contexts)


def _call_single_llm(
    *,
    family: str,
    title: str,
    context: SourceContext,
    ordinal: int,
    backend: str,
    base_url: str,
    model: str,
    timeout_seconds: int,
    max_tokens: int,
) -> LLMQueryDraft:
    parsed, meta = call_local_llm_strict_json(
        backend=backend,
        base_url=base_url,
        model=model,
        prompt=_single_llm_prompt(family=family, title=title, context=context, ordinal=ordinal),
        temperature=0.0,
        max_tokens=max_tokens,
        timeout_seconds=timeout_seconds,
    )
    query_value = _query_value(parsed)
    if not clean(query_value):
        query_items = _query_items(parsed)
        query_value = _query_value(query_items[0]) if query_items else ""
    return LLMQueryDraft(
        query=_normalize_llm_query(query_value),
        confidence=_confidence(parsed.get("confidence"), context.confidence),
        warning=_warnings(parsed.get("warnings")),
        response_sha256=clean(meta.get("raw_response_sha256") or meta.get("response_sha256")),
    )


def _llm_query_drafts_for_source(
    *,
    family: str,
    title: str,
    contexts: Sequence[SourceContext],
    backend: str,
    base_url: str,
    model: str,
    timeout_seconds: int,
    max_tokens: int,
) -> dict[int, LLMQueryDraft]:
    try:
        return _call_group_llm(
            family=family,
            title=title,
            contexts=contexts,
            backend=backend,
            base_url=base_url,
            model=model,
            timeout_seconds=timeout_seconds,
            max_tokens=max_tokens,
        )
    except Exception:
        drafts: dict[int, LLMQueryDraft] = {}
        for index, context in enumerate(contexts, start=1):
            drafts[index] = _call_single_llm(
                family=family,
                title=title,
                context=context,
                ordinal=index,
                backend=backend,
                base_url=base_url,
                model=model,
                timeout_seconds=timeout_seconds,
                max_tokens=max(320, max_tokens // max(1, len(contexts))),
            )
        return drafts


def _row_context_sequence(
    *,
    candidate_rows: Sequence[Mapping[str, Any]],
    source_manifest_rows: Mapping[str, Mapping[str, str]],
    source_contexts: Mapping[str, Sequence[SourceContext]],
) -> dict[str, list[tuple[Mapping[str, Any], SourceContext, int]]]:
    per_source_ordinals: defaultdict[str, int] = defaultdict(int)
    groups: dict[str, list[tuple[Mapping[str, Any], SourceContext, int]]] = defaultdict(list)
    for row in candidate_rows:
        _identity_kind, identity = _source_identity(row)
        source_sha = _manifest_sha_from_identity(identity)
        if not source_sha or not source_manifest_rows.get(source_sha):
            continue
        contexts = list(source_contexts.get(source_sha) or [])
        if not contexts or contexts[0].kind == "xlsx_extraction_failed":
            continue
        per_source_ordinals[source_sha] += 1
        ordinal = per_source_ordinals[source_sha]
        context = contexts[(ordinal - 1) % len(contexts)]
        groups[source_sha].append((row, context, ordinal))
    return groups


def _build_llm_query_cache(
    *,
    candidate_rows: Sequence[Mapping[str, Any]],
    source_manifest_rows: Mapping[str, Mapping[str, str]],
    source_contexts: Mapping[str, Sequence[SourceContext]],
    backend: str,
    base_url: str,
    model: str,
    timeout_seconds: int,
    max_tokens: int,
) -> dict[tuple[str, int], LLMQueryDraft]:
    blockers = local_llm_entry_blockers(
        backend=backend,
        base_url=base_url,
        model=model,
        check_endpoint=True,
        timeout_seconds=min(timeout_seconds, 5),
    )
    if blockers:
        raise RuntimeError("LOCAL_LLM_UNAVAILABLE_FAIL_CLOSED: " + "; ".join(blockers))
    cache: dict[tuple[str, int], LLMQueryDraft] = {}
    for source_sha, items in _row_context_sequence(
        candidate_rows=candidate_rows,
        source_manifest_rows=source_manifest_rows,
        source_contexts=source_contexts,
    ).items():
        metadata = source_manifest_rows.get(source_sha, {})
        family = clean(items[0][0].get("source_family")).upper()
        row_contexts = [context for _row, context, _ordinal in items]
        drafts = _llm_query_drafts_for_source(
            family=family,
            title=_display_title(metadata),
            contexts=row_contexts,
            backend=backend,
            base_url=base_url,
            model=model,
            timeout_seconds=timeout_seconds,
            max_tokens=max_tokens,
        )
        for output_ordinal, (_row, _context, row_ordinal) in enumerate(items, start=1):
            cache[(source_sha, row_ordinal)] = drafts[output_ordinal]
    return cache


def _draft_answer(context: SourceContext) -> str:
    return "비공식 기계초안(검수 필요): " + _bounded(context.snippet, limit=220)


def _base_human_columns(row: Mapping[str, Any], query: str) -> dict[str, str]:
    family = clean(row.get("source_family")).upper()
    return {
        "검수상태": "미검수",
        "소스계열": family,
        "후보ID": clean(row.get("candidate_id")),
        "질의ID": clean(row.get("query_id")),
        "질의문": query,
        "질의자연성": "보류",
        "질의승인": "보류",
        "질의보존성": "보류",
        "관련성라벨": "보류",
        "답변가능성라벨": "보류",
        "기대답변_한국어": "",
        "근거판단_한국어": "",
        "근거위치_확인": "보류",
        "공식분모포함판단": "보류",
        "제외사유": "",
        "정책메모": "",
        "검수자": "",
        "검수일시": "",
        "재검수필요": "보류",
    }


def _machine_columns(
    row: Mapping[str, Any],
    *,
    metadata: Mapping[str, str],
    context: SourceContext,
    query_method: str,
    query_basis: str,
    manifest_sha256: str,
    query_confidence: str | None = None,
    query_warning: str = "",
    llm_response_sha256: str = "",
) -> dict[str, str]:
    family = clean(row.get("source_family")).upper()
    identity_kind, identity = _source_identity(row)
    warning = "; ".join(item for item in (*context.warnings, query_warning) if clean(item))
    return {
        "질의생성방식": query_method,
        "질의생성근거": _bounded(query_basis, limit=360),
        "질의생성신뢰도": query_confidence or context.confidence,
        "질의생성경고": _bounded(warning, limit=240),
        "근거후보_스니펫": _bounded(context.snippet, limit=520),
        "근거후보_위치": _bounded(context.locator, limit=220),
        "근거후보_충분성_기계판단": context.sufficiency,
        "기대답변_초안_비공식": _draft_answer(context),
        "source_family": family,
        "candidate_id_hash": _hash(clean(row.get("candidate_id"))),
        "query_id_hash": _hash(clean(row.get("query_id"))),
        "document_or_workbook_identity_hash": _hash(f"{family}:{identity}"),
        "source_identity_kind": identity_kind,
        "source_disjointness_gate": "pass" if row.get("disjoint_from_prior") is True else "review",
        "query_fidelity_included": "true" if row.get("query_fidelity_included") is True else "false",
        "leakage_bucket": "none",
        "prior_identity_collision": "false",
        "manifest_sha256": manifest_sha256,
        "source_manifest_title": _display_title(metadata),
        "source_manifest_lane": _sanitize_packet_text(metadata.get("lane")),
        "source_manifest_subtype": _sanitize_packet_text(metadata.get("subtype")),
        "source_manifest_role": _sanitize_packet_text(metadata.get("role")),
        "source_preview_redacted": _source_preview(metadata),
        "evidence_preview_redacted": _bounded(context.snippet, limit=520),
        "locator_preview_redacted": _bounded(context.locator, limit=220),
        "page_or_sheet_locator_redacted": _bounded(
            context.page or context.sheet or context.cell_range or context.locator,
            limit=220,
        ),
        "source_report_run_id": V4_7_SOURCE_RUN_ID,
        "machine_notes": (
            "v4_7_2 source-grounded hydration generated a human-review-only Korean query candidate "
            "with local LLM strict JSON from bounded source evidence. "
            f"LLM response hash={llm_response_sha256[:12] or 'unrecorded'}. "
            "User-owned gold/label/denominator fields remain unchanged."
        ),
    }


def _family_columns(context: SourceContext, family: str) -> dict[str, str]:
    if family == "PDF":
        return {
            **{column: "" for column in XLSX_COLUMNS},
            "문서명_표시": context.title,
            "페이지_후보": context.page or "metadata",
            "섹션_후보": context.section or context.topic,
            "문단_후보": context.paragraph or context.locator,
        }
    return {
        "워크북명_표시": context.workbook,
        "시트명_표시": context.sheet,
        "근거후보_범위": context.cell_range,
        "근거후보_표시값_미리보기": context.value_preview,
        "근거후보_헤더_미리보기": context.header_preview,
        "근거후보_행열축_미리보기": context.axis_preview,
        **{column: "" for column in PDF_COLUMNS},
    }


def _context_groups(
    *,
    candidate_rows: Sequence[Mapping[str, Any]],
    source_manifest_rows: Mapping[str, Mapping[str, str]],
    source_collection_dir: Path,
) -> dict[str, list[SourceContext]]:
    source_contexts: dict[str, list[SourceContext]] = {}
    for row in candidate_rows:
        _identity_kind, identity = _source_identity(row)
        source_sha = _manifest_sha_from_identity(identity)
        if not source_sha or source_sha in source_contexts:
            continue
        metadata = source_manifest_rows.get(source_sha, {})
        if not metadata:
            source_contexts[source_sha] = [
                SourceContext(
                    kind="manifest_missing",
                    snippet="source_collection manifest row missing",
                    locator="source_collection_manifest_missing",
                    title="source manifest missing",
                    topic="source manifest missing",
                    confidence="low",
                    sufficiency="부족후보",
                    generation_basis="source manifest match missing",
                    warnings=("SOURCE_MANIFEST_MATCH_MISSING",),
                )
            ]
            continue
        source_path = _source_path(source_collection_dir, metadata)
        family = clean(row.get("source_family")).upper()
        if family == "PDF":
            source_contexts[source_sha] = extract_pdf_contexts(source_path=source_path, metadata=metadata)
        elif family == "XLSX":
            source_contexts[source_sha] = extract_xlsx_contexts(source_path=source_path, metadata=metadata)
    return source_contexts


def build_review_rows(
    *,
    candidate_manifest_path: Path = DEFAULT_CANDIDATE_MANIFEST,
    source_collection_manifest_path: Path = SOURCE_COLLECTION_MANIFEST_CSV,
    source_collection_dir: Path = EXTERNAL_SOURCE_COLLECTION_DIR,
    llm_backend: str = DEFAULT_LLM_BACKEND,
    llm_base_url: str = "",
    llm_model: str = DEFAULT_LLM_MODEL,
    llm_timeout_seconds: int = 120,
    llm_max_tokens: int = 1800,
) -> tuple[list[dict[str, str]], list[dict[str, str]], dict[str, int]]:
    candidate_rows = [
        row for row in read_jsonl(candidate_manifest_path) if clean(row.get("source_family")).upper() in {"PDF", "XLSX"}
    ]
    prior_rows = read_prior_packet_by_candidate()
    source_manifest_rows = read_source_collection_manifest(source_collection_manifest_path)
    source_contexts = _context_groups(
        candidate_rows=candidate_rows,
        source_manifest_rows=source_manifest_rows,
        source_collection_dir=source_collection_dir,
    )
    resolved_llm_base_url = resolve_llm_base_url(llm_backend, llm_base_url)
    llm_query_cache = _build_llm_query_cache(
        candidate_rows=candidate_rows,
        source_manifest_rows=source_manifest_rows,
        source_contexts=source_contexts,
        backend=llm_backend,
        base_url=resolved_llm_base_url,
        model=llm_model,
        timeout_seconds=llm_timeout_seconds,
        max_tokens=llm_max_tokens,
    )
    manifest_sha256 = _candidate_manifest_hash(candidate_manifest_path)
    per_source_ordinals: defaultdict[str, int] = defaultdict(int)
    method_counts: Counter[str] = Counter()
    review_rows: list[dict[str, str]] = []
    extraction_failed_rows: list[dict[str, str]] = []

    for row in candidate_rows:
        family = clean(row.get("source_family")).upper()
        candidate_id = clean(row.get("candidate_id"))
        _identity_kind, identity = _source_identity(row)
        source_sha = _manifest_sha_from_identity(identity)
        metadata = source_manifest_rows.get(source_sha, {})
        contexts = source_contexts.get(source_sha) or []
        if not metadata:
            extraction_failed_rows.append(
                {
                    "소스계열": family,
                    "후보ID": candidate_id,
                    "질의ID": clean(row.get("query_id")),
                    "failure_reason": "SOURCE_MANIFEST_MATCH_MISSING",
                }
            )
            continue
        if not contexts or contexts[0].kind == "xlsx_extraction_failed":
            extraction_failed_rows.append(
                {
                    "소스계열": family,
                    "후보ID": candidate_id,
                    "질의ID": clean(row.get("query_id")),
                    "failure_reason": "XLSX_STRUCTURE_EXTRACTION_UNAVAILABLE" if family == "XLSX" else "SOURCE_CONTEXT_UNAVAILABLE",
                }
            )
            continue
        per_source_ordinals[source_sha] += 1
        ordinal = per_source_ordinals[source_sha]
        context = contexts[(ordinal - 1) % len(contexts)]
        existing_query = _existing_query_from_row(row, prior_rows.get(candidate_id))
        if existing_query:
            query = existing_query
            query_method = "existing_artifact_query_reused"
            query_basis = f"existing linked query field reused; evidence hydrated from {context.generation_basis}"
            query_confidence = context.confidence
            query_warning = ""
            llm_response_sha256 = ""
        else:
            query_method = "local_llm_source_grounded_draft"
            draft = llm_query_cache.get((source_sha, ordinal))
            if draft is None:
                raise RuntimeError(f"missing local LLM query draft for source {source_sha} ordinal {ordinal}")
            query = draft.query
            query_basis = f"local LLM strict JSON draft from bounded source context; {context.generation_basis}"
            query_confidence = draft.confidence
            query_warning = draft.warning
            llm_response_sha256 = draft.response_sha256
        method_counts[query_method] += 1
        combined = {
            **_base_human_columns(row, query),
            **_machine_columns(
                row,
                metadata=metadata,
                context=context,
                query_method=query_method,
                query_basis=query_basis,
                manifest_sha256=manifest_sha256,
                query_confidence=query_confidence,
                query_warning=query_warning,
                llm_response_sha256=llm_response_sha256,
            ),
            **_family_columns(context, family),
        }
        review_rows.append({column: str(combined.get(column, "")) for column in REVIEW_COLUMNS})

    review_rows.sort(key=lambda item: (item["소스계열"], item["후보ID"], item["질의ID"]))
    counters = {
        "existing_artifact_query_reused": method_counts["existing_artifact_query_reused"],
        "deterministic_source_text_template": method_counts["deterministic_source_text_template"],
        "local_llm_source_grounded_draft": method_counts["local_llm_source_grounded_draft"],
    }
    return review_rows, extraction_failed_rows, counters


def _count_by_family(rows: Sequence[Mapping[str, str]]) -> dict[str, int]:
    counts = Counter(row["소스계열"] for row in rows)
    return {"PDF": counts.get("PDF", 0), "XLSX": counts.get("XLSX", 0), "TEXT": counts.get("TEXT", 0)}


def _artifact_paths(output_dir: Path) -> dict[str, str]:
    paths = {
        "report_json": output_dir / "report.json",
        "review_packet_ko_hydrated_xlsx": output_dir / "review_packet_ko_hydrated.xlsx",
        "review_packet_ko_hydrated_csv": output_dir / "review_packet_ko_hydrated.csv",
        "review_packet_ko_hydrated_jsonl": output_dir / "review_packet_ko_hydrated.jsonl",
        "review_guidelines_ko_md": output_dir / "review_guidelines_ko.md",
        "review_summary_ko_json": output_dir / "review_summary_ko.json",
    }
    if output_dir == OUTPUT_DIR:
        return {key: repo_relative(path) for key, path in paths.items()}
    return {key: path.as_posix() for key, path in paths.items()}


def _review_summary(
    *,
    review_rows: Sequence[Mapping[str, str]],
    extraction_failed_rows: Sequence[Mapping[str, str]],
    method_counts: Mapping[str, int],
) -> dict[str, Any]:
    counts = _count_by_family(review_rows)
    return {
        "schema_version": f"{RUN_ID}_review_summary_v1",
        "run_id": RUN_ID,
        "status": STATUS,
        "human_review_only": True,
        "diagnostic_only": True,
        "hydrated_packet_row_count": len(review_rows),
        "hydrated_packet_counts_by_family": counts,
        "hydrated_packet_non_empty_query_count": sum(1 for row in review_rows if clean(row.get("질의문"))),
        "extraction_failed_row_count": len(extraction_failed_rows),
        "existing_query_reused_count": int(method_counts.get("existing_artifact_query_reused", 0)),
        "deterministic_query_generated_count": int(method_counts.get("deterministic_source_text_template", 0)),
        "local_llm_query_generated_count": int(method_counts.get("local_llm_source_grounded_draft", 0)),
        "official_metric": False,
        "official_metric_input_rows": 0,
        "v4_7_official_metric_gate_opened": False,
        "product_success_evidence_allowed": False,
        "promotion_evidence": False,
        "ft_a_execution": False,
        "fine_tuning": False,
        "live_db_index_cache_readiness": False,
    }


def build_artifacts(
    *,
    candidate_manifest_path: Path = DEFAULT_CANDIDATE_MANIFEST,
    source_collection_manifest_path: Path = SOURCE_COLLECTION_MANIFEST_CSV,
    source_collection_dir: Path = EXTERNAL_SOURCE_COLLECTION_DIR,
    output_dir: Path = OUTPUT_DIR,
    llm_backend: str = DEFAULT_LLM_BACKEND,
    llm_base_url: str = "",
    llm_model: str = DEFAULT_LLM_MODEL,
    llm_timeout_seconds: int = 120,
    llm_max_tokens: int = 1800,
) -> dict[str, Any]:
    v4_7_artifacts = v47.build_artifacts(candidate_manifest_path=candidate_manifest_path)
    v47.check_report(v4_7_artifacts["report"])
    review_rows, extraction_failed_rows, method_counts = build_review_rows(
        candidate_manifest_path=candidate_manifest_path,
        source_collection_manifest_path=source_collection_manifest_path,
        source_collection_dir=source_collection_dir,
        llm_backend=llm_backend,
        llm_base_url=llm_base_url,
        llm_model=llm_model,
        llm_timeout_seconds=llm_timeout_seconds,
        llm_max_tokens=llm_max_tokens,
    )
    family_counts = _count_by_family(review_rows)
    prior = prior_packet_counters()
    summary = _review_summary(
        review_rows=review_rows,
        extraction_failed_rows=extraction_failed_rows,
        method_counts=method_counts,
    )
    registration = v4_7_artifacts["report"]["preofficial_external_holdout_candidate_manifest_registration"]
    report = {
        "schema_version": REPORT_SCHEMA_VERSION,
        "run_id": RUN_ID,
        "event_type": EVENT_TYPE,
        "status": STATUS,
        "v4_name": V4_NAME,
        "run_family": V4_RUN_FAMILY,
        "generated_at": utc_now(),
        "diagnostic_only": True,
        "non_production": True,
        "human_review_only": True,
        "source_grounded_query_review_packet_hydration_only": True,
        "not_official_metric": True,
        "not_gold_mutation": True,
        "not_qrels_mutation": True,
        "not_label_mutation": True,
        "not_training_dataset": True,
        "official_metric": False,
        "official_metric_input_rows": 0,
        "v4_7_official_metric_gate_opened": False,
        "product_success_evidence_allowed": False,
        "promotion_evidence": False,
        "live_db_index_cache_readiness": False,
        "ft_a_execution": False,
        "fine_tuning": False,
        "candidate_manifest_sha256": _candidate_manifest_hash(candidate_manifest_path),
        "candidate_manifest_sha256_verified_against_expected": _candidate_manifest_hash(candidate_manifest_path)
        == EXPECTED_V4_7_MANIFEST_SHA256,
        "candidate_manifest_path_redacted": True,
        "source_collection_manifest_available": source_collection_manifest_path.exists(),
        "source_collection_manifest_path_redacted": True,
        "source_collection_manifest_sha256": sha256_file(source_collection_manifest_path)
        if source_collection_manifest_path.exists()
        else "",
        "source_report_run_id": V4_7_SOURCE_RUN_ID,
        "source_report_status": clean(v4_7_artifacts["report"]["status"]),
        "source_report_artifact": repo_relative(v47.REPORT_JSON),
        "source_report_registration_gate_passed": bool(v4_7_artifacts["report"]["registration_gate_passed"]),
        "v4_7_registration_snapshot": {
            "status": clean(v4_7_artifacts["report"]["status"]),
            "candidate_rows_registered": int(registration.get("candidate_rows_registered") or 0),
            "candidate_counts_by_family": dict(registration.get("candidate_counts_by_family") or {}),
            "accepted_candidate_counts_by_family": dict(registration.get("accepted_candidate_counts_by_family") or {}),
            "accepted_pdf_holdout_candidates": int(registration.get("accepted_pdf_holdout_candidates") or 0),
            "accepted_xlsx_holdout_candidates": int(registration.get("accepted_xlsx_holdout_candidates") or 0),
            "rejected_candidate_count": int(registration.get("rejected_candidate_count") or 0),
            "source_identity_collision_count": int(registration.get("source_identity_collision_count") or 0),
            "real_query_fidelity_included_counts": dict(registration.get("real_query_fidelity_included_counts") or {}),
            "real_holdout_sufficient": False,
            "official_metric": False,
            "official_metric_input_rows": 0,
            "v4_7_official_metric_gate_opened": False,
            "promotion_evidence": False,
            "product_success_evidence_allowed": False,
            "live_db_index_cache_readiness": False,
        },
        **prior,
        "hydrated_packet_row_count": len(review_rows),
        "hydrated_pdf_row_count": family_counts["PDF"],
        "hydrated_xlsx_row_count": family_counts["XLSX"],
        "hydrated_text_row_count": family_counts["TEXT"],
        "hydrated_packet_non_empty_query_count": sum(1 for row in review_rows if clean(row.get("질의문"))),
        "extraction_failed_row_count": len(extraction_failed_rows),
        "existing_query_reused_count": int(method_counts.get("existing_artifact_query_reused", 0)),
        "deterministic_query_generated_count": int(method_counts.get("deterministic_source_text_template", 0)),
        "local_llm_query_generated_count": int(method_counts.get("local_llm_source_grounded_draft", 0)),
        "query_generation_strategy": {
            "searched_existing_query_fields": list(QUERY_FIELD_NAMES),
            "existing_query_reuse_sources_checked": [
                "v4_7_candidate_manifest_rows",
                "v4_7_1_review_packet_rows",
                "linked_row_fields_when_present",
            ],
            "fallback": "fail_closed_no_deterministic_query_generation",
            "deterministic_template_fallback_enabled": False,
            "local_llm_used": True,
            "local_llm_backend": llm_backend,
            "local_llm_model": clean(llm_model),
            "local_llm_base_url": resolve_llm_base_url(llm_backend, llm_base_url),
            "prompt_body_persisted": False,
            "model_response_body_persisted": False,
        },
        "artifact_paths": _artifact_paths(output_dir),
        "user_owned_columns_kept_blank_or_default": [
            "기대답변_한국어",
            "근거판단_한국어",
            "관련성라벨",
            "답변가능성라벨",
            "공식분모포함판단",
            "제외사유",
            "정책메모",
            "검수자",
            "검수일시",
        ],
        "qrels_mutation": False,
        "gold_mutation": False,
        "label_mutation": False,
        "training_dataset_created": False,
        "candidate_manifest_jsonl_created": False,
        "qrels_jsonl_created": False,
        "gold_jsonl_created": False,
        "labels_jsonl_created": False,
        "answer_key_jsonl_created": False,
        "evidence_key_jsonl_created": False,
        "training_manifest_jsonl_created": False,
        "prompt_manifest_jsonl_created": False,
        "raw_response_payload_jsonl_created": False,
        "review_summary": summary,
        "guardrails": {
            "raw_local_paths_absent": True,
            "raw_source_identities_absent": True,
            "protected_oracle_fields_absent": True,
            "formula_text_exposed": False,
            "formula_evaluation_at_query_time": False,
            "gold_qrels_expected_supporting_labels_mutated": False,
            "official_metric_input_rows_created": 0,
            "training_dataset_created": False,
            "production_db_index_cache_namespaces_touched": False,
        },
        "remaining_user_owned_actions": [
            "human reviewers must approve, revise, or exclude each generated Korean query candidate",
            "gold/qrels, expected answers, supporting evidence, relevance, answerability, official denominator inclusion, and promotion policy remain user-owned",
        ],
        "readiness_decision": "hydrated_korean_review_packet_ready_official_metric_closed",
    }
    return {
        "report": report,
        "review_rows": review_rows,
        "extraction_failed_rows": extraction_failed_rows,
        "review_summary": summary,
    }


def _write_csv(path: Path, rows: Sequence[Mapping[str, str]]) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    with path.open("w", encoding="utf-8-sig", newline="") as handle:
        writer = csv.DictWriter(handle, fieldnames=REVIEW_COLUMNS)
        writer.writeheader()
        writer.writerows(rows)


def _cell_ref(row_index: int, col_index: int) -> str:
    letters = ""
    col = col_index
    while col:
        col, rem = divmod(col - 1, 26)
        letters = chr(65 + rem) + letters
    return f"{letters}{row_index}"


def _worksheet_xml(rows: Sequence[Sequence[Any]]) -> str:
    sheet_rows: list[str] = []
    for row_index, row in enumerate(rows, start=1):
        cells: list[str] = []
        for col_index, value in enumerate(row, start=1):
            text = "" if value is None else str(value)
            ref = _cell_ref(row_index, col_index)
            cells.append(f'<c r="{ref}" t="inlineStr"><is><t xml:space="preserve">{escape(text)}</t></is></c>')
        sheet_rows.append(f'<row r="{row_index}">' + "".join(cells) + "</row>")
    max_col = max((len(row) for row in rows), default=1)
    max_row = max(len(rows), 1)
    ref = f"A1:{_cell_ref(max_row, max_col)}"
    return (
        '<?xml version="1.0" encoding="UTF-8" standalone="yes"?>'
        '<worksheet xmlns="http://schemas.openxmlformats.org/spreadsheetml/2006/main" '
        'xmlns:r="http://schemas.openxmlformats.org/officeDocument/2006/relationships">'
        '<sheetViews><sheetView workbookViewId="0"><pane ySplit="1" topLeftCell="A2" '
        'activePane="bottomLeft" state="frozen"/></sheetView></sheetViews>'
        "<sheetData>"
        + "".join(sheet_rows)
        + f'</sheetData><autoFilter ref="{ref}"/></worksheet>'
    )


def _workbook_xml() -> str:
    sheets = "".join(
        f'<sheet name="{escape(name)}" sheetId="{index}" r:id="rId{index}"/>'
        for index, name in enumerate(SHEET_NAMES, start=1)
    )
    return (
        '<?xml version="1.0" encoding="UTF-8" standalone="yes"?>'
        '<workbook xmlns="http://schemas.openxmlformats.org/spreadsheetml/2006/main" '
        'xmlns:r="http://schemas.openxmlformats.org/officeDocument/2006/relationships">'
        f"<sheets>{sheets}</sheets></workbook>"
    )


def _content_types_xml() -> str:
    worksheets = "".join(
        f'<Override PartName="/xl/worksheets/sheet{index}.xml" '
        'ContentType="application/vnd.openxmlformats-officedocument.spreadsheetml.worksheet+xml"/>'
        for index in range(1, len(SHEET_NAMES) + 1)
    )
    return (
        '<?xml version="1.0" encoding="UTF-8" standalone="yes"?>'
        '<Types xmlns="http://schemas.openxmlformats.org/package/2006/content-types">'
        '<Default Extension="rels" ContentType="application/vnd.openxmlformats-package.relationships+xml"/>'
        '<Default Extension="xml" ContentType="application/xml"/>'
        '<Override PartName="/xl/workbook.xml" '
        'ContentType="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet.main+xml"/>'
        f"{worksheets}</Types>"
    )


def _root_rels_xml() -> str:
    return (
        '<?xml version="1.0" encoding="UTF-8" standalone="yes"?>'
        '<Relationships xmlns="http://schemas.openxmlformats.org/package/2006/relationships">'
        '<Relationship Id="rId1" Type="http://schemas.openxmlformats.org/officeDocument/2006/relationships/officeDocument" '
        'Target="xl/workbook.xml"/></Relationships>'
    )


def _workbook_rels_xml() -> str:
    rels = "".join(
        f'<Relationship Id="rId{index}" Type="http://schemas.openxmlformats.org/officeDocument/2006/relationships/worksheet" '
        f'Target="worksheets/sheet{index}.xml"/>'
        for index in range(1, len(SHEET_NAMES) + 1)
    )
    return (
        '<?xml version="1.0" encoding="UTF-8" standalone="yes"?>'
        f'<Relationships xmlns="http://schemas.openxmlformats.org/package/2006/relationships">{rels}</Relationships>'
    )


def _xlsx_sheet_rows(
    *,
    review_rows: Sequence[Mapping[str, str]],
    extraction_failed_rows: Sequence[Mapping[str, str]],
    summary: Mapping[str, Any],
) -> dict[str, list[list[Any]]]:
    all_rows = [REVIEW_COLUMNS] + [[row.get(column, "") for column in REVIEW_COLUMNS] for row in review_rows]
    pdf_rows = [REVIEW_COLUMNS] + [
        [row.get(column, "") for column in REVIEW_COLUMNS] for row in review_rows if row.get("소스계열") == "PDF"
    ]
    xlsx_rows = [REVIEW_COLUMNS] + [
        [row.get(column, "") for column in REVIEW_COLUMNS] for row in review_rows if row.get("소스계열") == "XLSX"
    ]
    failed_columns = ["소스계열", "후보ID", "질의ID", "failure_reason"]
    failed_rows = [failed_columns] + [[row.get(column, "") for column in failed_columns] for row in extraction_failed_rows]
    labels = [
        ["구분", "라벨", "설명"],
        ["질의자연성", "승인", "검수자가 자연스러운 실제 사용자 질의로 판단했다."],
        ["질의자연성", "보류", "검수 전 기본값이다."],
        ["질의승인", "승인", "후보 질의를 그대로 사용할 수 있다."],
        ["질의승인", "수정필요", "검수자가 질의문을 수정해야 한다."],
        ["질의승인", "제외", "공식 후보로 사용하지 않는다."],
        ["관련성라벨", "정확근거", "근거 후보만으로 답을 직접 산출할 수 있다."],
        ["관련성라벨", "보조근거", "답변 맥락을 보조하지만 단독 충분성은 낮다."],
        ["관련성라벨", "무관", "질의와 관련성이 없다."],
        ["답변가능성라벨", "완전답변가능", "근거 후보만으로 답변 가능하다."],
        ["답변가능성라벨", "부분답변가능", "일부 답변만 가능하다."],
        ["공식분모포함판단", "보류", "공식 denominator 여부는 아직 사용자 판단 전이다."],
    ]
    exclusions = [["제외사유", "사용 기준"]] + [[reason, "제외로 판단한 경우에만 선택"] for reason in EXCLUSION_REASONS]
    summary_rows = [
        ["항목", "값"],
        ["status", summary["status"]],
        ["human_review_only", "true"],
        ["hydrated_packet_row_count", summary["hydrated_packet_row_count"]],
        ["PDF rows", summary["hydrated_packet_counts_by_family"]["PDF"]],
        ["XLSX rows", summary["hydrated_packet_counts_by_family"]["XLSX"]],
        ["non_empty_query_count", summary["hydrated_packet_non_empty_query_count"]],
        ["extraction_failed_row_count", summary["extraction_failed_row_count"]],
        ["existing_query_reused_count", summary["existing_query_reused_count"]],
        ["deterministic_query_generated_count", summary["deterministic_query_generated_count"]],
        ["local_llm_query_generated_count", summary["local_llm_query_generated_count"]],
        ["official_metric_input_rows", 0],
        ["qrels_mutation", "false"],
        ["gold_mutation", "false"],
        ["label_mutation", "false"],
        ["training_dataset_created", "false"],
    ]
    return {
        "검수_대상_전체": all_rows,
        "PDF_검수": pdf_rows,
        "XLSX_검수": xlsx_rows,
        "추출실패_검수": failed_rows,
        "라벨_가이드": labels,
        "제외_사유_가이드": exclusions,
        "요약": summary_rows,
    }


def _write_xlsx(
    path: Path,
    *,
    review_rows: Sequence[Mapping[str, str]],
    extraction_failed_rows: Sequence[Mapping[str, str]],
    summary: Mapping[str, Any],
) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    sheet_payloads = _xlsx_sheet_rows(
        review_rows=review_rows,
        extraction_failed_rows=extraction_failed_rows,
        summary=summary,
    )
    with zipfile.ZipFile(path, "w", compression=zipfile.ZIP_DEFLATED) as archive:
        archive.writestr("[Content_Types].xml", _content_types_xml())
        archive.writestr("_rels/.rels", _root_rels_xml())
        archive.writestr("xl/workbook.xml", _workbook_xml())
        archive.writestr("xl/_rels/workbook.xml.rels", _workbook_rels_xml())
        for index, sheet_name in enumerate(SHEET_NAMES, start=1):
            archive.writestr(f"xl/worksheets/sheet{index}.xml", _worksheet_xml(sheet_payloads[sheet_name]))


def _guidelines_markdown(summary: Mapping[str, Any]) -> str:
    return f"""# v4_7_2 한국어 질의 후보 검수 가이드

이 packet은 v4_7 pre-official PDF/XLSX 후보를 사람이 검수할 수 있도록 source-grounded Korean query 후보와 bounded evidence preview를 붙인 자료입니다.
모든 행은 human-review-only이며, gold, qrels, relevance label, answerability label, official denominator 포함 여부, promotion policy는 사람이 결정합니다.

## 검수 방법

1. `review_packet_ko_hydrated.xlsx`의 `검수_대상_전체`, `PDF_검수`, `XLSX_검수` sheet에서 후보를 확인합니다.
2. `질의문`, `근거후보_스니펫`, `근거후보_위치`, locator preview를 함께 보고 query를 승인, 수정, 제외합니다.
3. `기대답변_초안_비공식`은 machine hint입니다. gold expected answer가 아니며 qrels, official metric, training data로 사용하지 않습니다.
4. `기대답변_한국어`, `근거판단_한국어`, `관련성라벨`, `답변가능성라벨`, `공식분모포함판단`은 user-owned review field입니다.

## 경계

- official_metric=false
- official_metric_input_rows=0
- v4_7_official_metric_gate_opened=false
- product_success_evidence_allowed=false
- promotion_evidence=false
- FT-A execution=false
- fine_tuning=false
- live_db_index_cache_readiness=false
- training_dataset_created=false

요약: hydrated rows {summary["hydrated_packet_row_count"]}, PDF {summary["hydrated_packet_counts_by_family"]["PDF"]}, XLSX {summary["hydrated_packet_counts_by_family"]["XLSX"]}, extraction_failed {summary["extraction_failed_row_count"]}.
"""


def write_artifacts(artifacts: Mapping[str, Any], *, output_dir: Path = OUTPUT_DIR) -> dict[str, Any]:
    output_dir.mkdir(parents=True, exist_ok=True)
    unexpected = {path.name for path in output_dir.iterdir() if path.is_file()} - ALLOWED_ARTIFACT_NAMES
    if unexpected:
        raise RuntimeError(f"unexpected v4_7_2 artifacts present: {sorted(unexpected)}")
    report = dict(artifacts["report"])
    report["artifact_paths"] = _artifact_paths(output_dir)
    review_rows = list(artifacts["review_rows"])
    extraction_failed_rows = list(artifacts["extraction_failed_rows"])
    summary = dict(artifacts["review_summary"])
    write_json(output_dir / "report.json", report)
    _write_csv(output_dir / "review_packet_ko_hydrated.csv", review_rows)
    write_jsonl(output_dir / "review_packet_ko_hydrated.jsonl", review_rows)
    write_json(output_dir / "review_summary_ko.json", summary)
    (output_dir / "review_guidelines_ko.md").write_text(_guidelines_markdown(summary), encoding="utf-8")
    _write_xlsx(
        output_dir / "review_packet_ko_hydrated.xlsx",
        review_rows=review_rows,
        extraction_failed_rows=extraction_failed_rows,
        summary=summary,
    )
    return report


def _serialized_artifact_text(output_dir: Path) -> str:
    paths = [
        output_dir / "report.json",
        output_dir / "review_packet_ko_hydrated.csv",
        output_dir / "review_packet_ko_hydrated.jsonl",
        output_dir / "review_summary_ko.json",
        output_dir / "review_guidelines_ko.md",
    ]
    return "\n".join(path.read_text(encoding="utf-8-sig" if path.suffix == ".csv" else "utf-8") for path in paths)


def _contains_forbidden_text(payload: str) -> bool:
    return any(re.search(pattern, payload) for pattern in FORBIDDEN_TEXT_PATTERNS)


def check_report(report: Mapping[str, Any]) -> None:
    if report.get("schema_version") != REPORT_SCHEMA_VERSION:
        raise AssertionError("unexpected v4_7_2 report schema")
    if report.get("status") != STATUS:
        raise AssertionError("unexpected v4_7_2 status")
    for field in (
        "diagnostic_only",
        "human_review_only",
        "source_grounded_query_review_packet_hydration_only",
        "not_official_metric",
        "not_gold_mutation",
        "not_qrels_mutation",
        "not_label_mutation",
        "not_training_dataset",
    ):
        if report.get(field) is not True:
            raise AssertionError(f"{field} must remain true")
    for field in (
        "official_metric",
        "v4_7_official_metric_gate_opened",
        "product_success_evidence_allowed",
        "promotion_evidence",
        "live_db_index_cache_readiness",
        "ft_a_execution",
        "fine_tuning",
        "qrels_mutation",
        "gold_mutation",
        "label_mutation",
        "training_dataset_created",
        "candidate_manifest_jsonl_created",
        "qrels_jsonl_created",
        "gold_jsonl_created",
        "labels_jsonl_created",
        "answer_key_jsonl_created",
        "evidence_key_jsonl_created",
        "training_manifest_jsonl_created",
        "prompt_manifest_jsonl_created",
        "raw_response_payload_jsonl_created",
    ):
        if report.get(field) is not False:
            raise AssertionError(f"{field} must remain false")
    if int(report.get("official_metric_input_rows") or 0) != 0:
        raise AssertionError("official_metric_input_rows must remain 0")
    if report.get("candidate_manifest_sha256") != EXPECTED_V4_7_MANIFEST_SHA256:
        raise AssertionError("candidate manifest sha drifted")
    if report.get("hydrated_packet_row_count") != 204:
        raise AssertionError("hydrated packet must cover 204 rows")
    if report.get("hydrated_pdf_row_count") != 100:
        raise AssertionError("hydrated PDF row count must be 100")
    if report.get("hydrated_xlsx_row_count") != 104:
        raise AssertionError("hydrated XLSX row count must be 104")
    if report.get("hydrated_packet_non_empty_query_count") != report.get("hydrated_packet_row_count"):
        raise AssertionError("all hydrated reviewable rows must have query text")
    if int(report.get("extraction_failed_row_count") or 0) != 0:
        raise AssertionError("default v4_7_2 packet should not have extraction-failed rows")
    if int(report.get("deterministic_query_generated_count") or 0) != 0:
        raise AssertionError("deterministic query generation must stay disabled for v4_7_2 LLM hydration")
    if int(report.get("local_llm_query_generated_count") or 0) + int(report.get("existing_query_reused_count") or 0) != 204:
        raise AssertionError("query generation accounting drifted")
    strategy = report.get("query_generation_strategy") if isinstance(report.get("query_generation_strategy"), Mapping) else {}
    if strategy.get("local_llm_used") is not True or strategy.get("deterministic_template_fallback_enabled") is not False:
        raise AssertionError("v4_7_2 query generation must use local LLM and keep deterministic fallback disabled")
    payload = json.dumps(report, ensure_ascii=False, sort_keys=True)
    if _contains_forbidden_text(payload):
        raise AssertionError("forbidden raw path, raw identity, oracle, qrels, prompt, training, or checkpoint text leaked")


def check_written_artifacts(output_dir: Path = OUTPUT_DIR) -> None:
    files = {path.name for path in output_dir.iterdir() if path.is_file()}
    if files != ALLOWED_ARTIFACT_NAMES:
        raise AssertionError(f"unexpected v4_7_2 artifact set: {sorted(files)}")
    report = read_json(output_dir / "report.json")
    check_report(report)
    rows = read_jsonl(output_dir / "review_packet_ko_hydrated.jsonl")
    if len(rows) != 204:
        raise AssertionError("review_packet_ko_hydrated.jsonl row count drifted")
    if any(set(row) != set(REVIEW_COLUMNS) for row in rows):
        raise AssertionError("review packet columns drifted")
    if any(not clean(row.get("질의문")) for row in rows):
        raise AssertionError("all reviewable rows must have non-empty 질의문")
    if any(_is_placeholder_query(clean(row.get("질의문"))) for row in rows):
        raise AssertionError("질의문 must not be only placeholder ids")
    if any(not clean(row.get("근거후보_스니펫") or row.get("evidence_preview_redacted")) for row in rows):
        raise AssertionError("all rows need bounded evidence previews")
    if any(not clean(row.get("근거후보_위치") or row.get("locator_preview_redacted")) for row in rows):
        raise AssertionError("all rows need locator previews")
    if any(clean(row.get("기대답변_한국어")) or clean(row.get("근거판단_한국어")) for row in rows):
        raise AssertionError("user-owned gold answer/evidence fields must stay blank")
    if any(row.get("공식분모포함판단") != "보류" for row in rows):
        raise AssertionError("official denominator decision must stay 보류")
    if any(row.get("소스계열") == "XLSX" and "=" in clean(row.get("근거후보_표시값_미리보기")) for row in rows):
        raise AssertionError("formula text must not appear in XLSX value previews")
    if _contains_forbidden_text(_serialized_artifact_text(output_dir)):
        raise AssertionError("forbidden text leaked into generated packet artifacts")
    with zipfile.ZipFile(output_dir / "review_packet_ko_hydrated.xlsx") as archive:
        workbook = archive.read("xl/workbook.xml").decode("utf-8")
        for sheet_name in SHEET_NAMES:
            if sheet_name not in workbook:
                raise AssertionError(f"missing sheet {sheet_name}")


def update_status(report: Mapping[str, Any]) -> None:
    artifact_hashes = {
        "report_json_sha256": sha256_file(REPORT_JSON),
        "review_packet_ko_hydrated_xlsx_sha256": sha256_file(REVIEW_PACKET_XLSX),
        "review_packet_ko_hydrated_csv_sha256": sha256_file(REVIEW_PACKET_CSV),
        "review_packet_ko_hydrated_jsonl_sha256": sha256_file(REVIEW_PACKET_JSONL),
        "review_guidelines_ko_md_sha256": sha256_file(REVIEW_GUIDELINES),
        "review_summary_ko_json_sha256": sha256_file(REVIEW_SUMMARY_JSON),
    }
    event = {
        "schema_version": f"{RUN_ID}_status_event_v1",
        "run_id": RUN_ID,
        "event_type": EVENT_TYPE,
        "status": STATUS,
        "v4_name": V4_NAME,
        "run_family": V4_RUN_FAMILY,
        "artifact_paths": dict(report["artifact_paths"]),
        "artifact_sha256": artifact_hashes,
        "diagnostic_only": True,
        "human_review_only": True,
        "source_grounded_query_review_packet_hydration_only": True,
        "prior_packet_row_count": report["prior_packet_row_count"],
        "prior_packet_non_empty_query_count": report["prior_packet_non_empty_query_count"],
        "hydrated_packet_row_count": report["hydrated_packet_row_count"],
        "hydrated_packet_non_empty_query_count": report["hydrated_packet_non_empty_query_count"],
        "hydrated_pdf_row_count": report["hydrated_pdf_row_count"],
        "hydrated_xlsx_row_count": report["hydrated_xlsx_row_count"],
        "extraction_failed_row_count": report["extraction_failed_row_count"],
        "existing_query_reused_count": report["existing_query_reused_count"],
        "deterministic_query_generated_count": report["deterministic_query_generated_count"],
        "local_llm_query_generated_count": report["local_llm_query_generated_count"],
        "official_metric": False,
        "official_metric_input_rows": 0,
        "v4_7_official_metric_gate_opened": False,
        "product_success_evidence_allowed": False,
        "promotion_evidence": False,
        "live_db_index_cache_readiness": False,
        "ft_a_execution": False,
        "fine_tuning": False,
        "qrels_mutation": False,
        "gold_mutation": False,
        "label_mutation": False,
        "training_dataset_created": False,
        "source_report_run_id": V4_7_SOURCE_RUN_ID,
    }
    rows = [
        row
        for row in read_jsonl(STATUS_JSONL)
        if not (row.get("run_id") == RUN_ID and row.get("event_type") == EVENT_TYPE)
    ]
    rows.append(event)
    write_jsonl(STATUS_JSONL, rows)


def update_root_readme(report: Mapping[str, Any]) -> None:
    text = README.read_text(encoding="utf-8")
    snapshot = f"""## Current RAG Diagnostic Status

- Current RAG status: `{STATUS}`.
- Phase: v4_7 remains pre-official. v4_7_2 supersedes the abstract v4_7_1 Korean review packet with a source-grounded hydrated Korean query review packet for the registered v4_7 PDF/XLSX candidates.
- v4_7_2 packet counters: prior packet rows {report["prior_packet_row_count"]}, prior non-empty `질의문` {report["prior_packet_non_empty_query_count"]}; hydrated rows {report["hydrated_packet_row_count"]}, PDF {report["hydrated_pdf_row_count"]}, XLSX {report["hydrated_xlsx_row_count"]}, non-empty `질의문` {report["hydrated_packet_non_empty_query_count"]}, extraction_failed {report["extraction_failed_row_count"]}.
- Hydration strategy: existing query reuse {report["existing_query_reused_count"]}, local LLM source-grounded query generation {report["local_llm_query_generated_count"]}, deterministic template generation {report["deterministic_query_generated_count"]}. Bounded evidence and locator previews are machine hints for human review only.
- Rolling evidence docs: `docs/rag-ingestion-progress.md`, `docs/rag-ingestion-measurements.md`, and `docs/rag-ingestion-triage.md` remain the canonical human-readable status ledgers; this packet is not production promotion evidence.
- Hard boundary: not official metric, not product-success evidence, not promotion evidence, not FT-A execution, not fine-tuning, not training data, and not live DB/index/cache readiness. Locked flags remain `official_metric=false`, `official_metric_input_rows=0`, `v4_7_official_metric_gate_opened=false`, `product_success_evidence_allowed=false`, `promotion_evidence=false`, `ft_a_execution=false`, `fine_tuning=false`, `fine_tuning_executed=false`, and `live_db_index_cache_readiness=false`.
"""
    if "## Current RAG Diagnostic Status" in text:
        text = re.sub(
            r"## Current RAG Diagnostic Status\n.*?(?=\n## (?:Recent Focus:|전체 구조|구성 요소|폴더 구조))",
            snapshot.rstrip() + "\n",
            text,
            count=1,
            flags=re.S,
        )
    elif "\n## 전체 구조" in text:
        text = text.replace("\n## 전체 구조", "\n" + snapshot.rstrip() + "\n\n## 전체 구조", 1)
    else:
        text = text.rstrip() + "\n\n" + snapshot.rstrip() + "\n"

    rolling_scripts = [
        "rag_v4_3_pdf_file_identity_confidence_and_evidence_window_split_nonprod.py",
        "rag_v4_4_real_blind_ood_holdout_and_leakage_audit_nonprod.py",
        "rag_v4_5_finetune_readiness_packet_nonprod.py",
        "rag_v4_5_1_holdout_candidate_intake_gate_nonprod.py",
        "rag_v4_5_2_external_holdout_candidate_source_identity_audit_nonprod.py",
        "rag_v4_5_3_external_holdout_prior_source_identity_ledger_summary_nonprod.py",
        "rag_v4_6_ft_route_policy_dry_run_preflight_nonprod.py",
        "rag_v4_6_1_holdout_candidate_manifest_identity_contract_bridge_nonprod.py",
        "rag_v4_6_2_ft_route_policy_fixture_contract_nonprod.py",
        "rag_v4_6_3_ft_a_prompt_policy_baseline_schema_nonprod.py",
        "rag_v4_6_4_ft_a_dry_run_input_manifest_validator_nonprod.py",
        "rag_v4_6_5_ft_a_dry_run_execution_plan_gate_nonprod.py",
        "rag_v4_6_6_holdout_gap_and_dry_run_blocker_ledger_nonprod.py",
        "rag_v4_6_7_holdout_candidate_runtime_gate_parity_bridge_nonprod.py",
        "rag_v4_6_8_runtime_readiness_dependency_freshness_gate_nonprod.py",
        "rag_v4_6_9_holdout_candidate_duplicate_hygiene_gate_nonprod.py",
        "rag_v4_6_10_external_holdout_candidate_manifest_gate_replay_nonprod.py",
        "rag_v4_6_11_ft_a_runtime_input_validation_route_parity_nonprod.py",
        "rag_v4_6_12_external_holdout_runtime_replay_route_parity_nonprod.py",
        "rag_v4_7_preofficial_external_holdout_candidate_manifest_registration_nonprod.py",
        "rag_v4_7_1_korean_review_packet_and_readme_status_snapshot_nonprod.py",
        "rag_v4_7_2_source_grounded_korean_query_review_packet_hydration_nonprod.py",
    ]
    verify_lines = [
        "진단 산출물 재검증은 원본 외부 manifest 경로를 README에 노출하지 않고 다음 명령으로 수행합니다.",
        "",
        "```powershell",
        *[f"python -X utf8 -m py_compile ai\\scripts\\{script_name}" for script_name in rolling_scripts],
        "python -X utf8 ai\\scripts\\rag_v4_7_preofficial_external_holdout_candidate_manifest_registration_nonprod.py --check --candidate-manifest <external-candidate-manifest-jsonl>",
        *[
            f"python -X utf8 ai\\scripts\\{script_name} --check"
            for script_name in rolling_scripts
            if script_name != "rag_v4_7_preofficial_external_holdout_candidate_manifest_registration_nonprod.py"
        ],
        "```",
    ]
    verify_block = "\n".join(verify_lines)
    if "## How To Verify Locally" in text and "## Repo Map" in text:
        start = text.index("## How To Verify Locally")
        end = text.index("## Repo Map")
        section = text[start:end]
        if "rag_v4_7_2_source_grounded_korean_query_review_packet_hydration_nonprod.py --check" not in section:
            section = section.rstrip() + "\n\n" + verify_block + "\n"
        text = text[:start] + section + "\n" + text[end:]
    elif "## 로컬 실행 메모" in text and "## 라이선스와 외부 데이터" in text:
        start = text.index("## 로컬 실행 메모")
        end = text.index("## 라이선스와 외부 데이터")
        section = text[start:end]
        section = re.sub(
            r"진단 산출물 재검증은 원본 외부 manifest 경로를 README에 노출하지 않고 다음 명령으로 수행합니다\.\n\n```powershell\n.*?```\n(?:\n```powershell\n.*?```\n)?",
            lambda _match: verify_block,
            section,
            count=1,
            flags=re.S,
        )
        if "rag_v4_7_2_source_grounded_korean_query_review_packet_hydration_nonprod.py --check" not in section:
            section = section.rstrip() + "\n\n" + verify_block + "\n"
        text = text[:start] + section + "\n" + text[end:]
    README.write_text(text, encoding="utf-8")


def update_eval_readme(report: Mapping[str, Any]) -> None:
    text = EVAL_README.read_text(encoding="utf-8")
    text = re.sub(r"- Current RAG status: `[^`]+`", f"- Current RAG status: `{STATUS}`", text, count=1)
    section = f"""## Korean human review packet

The previous v4_7_1 Korean review packet was abstract because v4_7 registration contained source-disjoint candidate identities but not query text. v4_7_2 supersedes it with `ai/eval/reports/rag-ingestion/quality/{RUN_ID}/review_packet_ko_hydrated.xlsx`, plus CSV/JSONL equivalents.

The hydrated packet contains actual Korean query candidates, bounded evidence previews, and locator previews for PDF {report["hydrated_pdf_row_count"]} rows and XLSX {report["hydrated_xlsx_row_count"]} rows. User-owned fields remain blank/default: expected answers, evidence judgment, relevance labels, answerability labels, official denominator inclusion, exclusion reasons, policy memo, reviewer, and review timestamp.

This is human-review-only and diagnostic-only. It is not official metric, not gold/qrels, not training data, not product-success evidence, not promotion evidence, not FT-A execution, not fine-tuning, and not live DB/index/cache readiness. Locked flags include `official_metric_input_rows=0`, `promotion_evidence=false`, `product_success_evidence_allowed=false`, `fine_tuning_executed=false`, and `live_db_index_cache_readiness=false`.
"""
    if "## Korean human review packet" in text:
        text = re.sub(
            r"## Korean human review packet\n.*?(?=\n## (?:Evaluation Boundary|평가 경계))",
            section.rstrip() + "\n",
            text,
            count=1,
            flags=re.S,
        )
    elif "\n## 평가 경계" in text:
        text = text.replace("\n## 평가 경계", "\n" + section.rstrip() + "\n\n## 평가 경계", 1)
    else:
        text = text.rstrip() + "\n\n" + section.rstrip() + "\n"
    EVAL_README.write_text(text, encoding="utf-8")


def update_scripts_readme() -> None:
    text = SCRIPTS_README.read_text(encoding="utf-8")
    row = (
        "| `rag_v4_7_2_source_grounded_korean_query_review_packet_hydration_nonprod.py` | "
        "Hydrates the v4_7 Korean human-review packet with local-LLM source-grounded Korean query candidates, bounded evidence previews, "
        "and locator previews while keeping official metrics, FT-A execution, fine-tuning, promotion, product-success evidence, "
        "training data, and live readiness closed. |"
    )
    pattern = r"\n?\| `rag_v4_7_2_source_grounded_korean_query_review_packet_hydration_nonprod\.py` \| .*?\|"
    text = re.sub(pattern, "", text)
    text = text.replace(
        "\n\nv4 scripts remain diagnostic/non-production",
        f"\n{row}\n\nv4 scripts remain diagnostic/non-production",
        1,
    )
    SCRIPTS_README.write_text(text, encoding="utf-8")


def update_progress_doc(report: Mapping[str, Any]) -> None:
    entry = (
        f"- v4_7_2 source-grounded Korean query review packet hydration (`{RUN_ID}`) is {STATUS}. "
        "It fixed the Korean review packet by hydrating or generating source-grounded Korean queries and bounded evidence "
        "previews for the registered v4_7 PDF/XLSX candidates. It supersedes the abstract v4_7_1 packet that had blank "
        "`질의문` rows. It remains human-review-only and does not create official metric rows, gold/qrels, labels, training "
        "data, product-success evidence, promotion evidence, or live readiness."
    )
    v47.v4610.v469.v467.replace_marked_entry(PROGRESS_DOC, RUN_ID, entry)
    text = PROGRESS_DOC.read_text(encoding="utf-8")
    text = re.sub(r"Overall status: `[^`]+`;", f"Overall status: `{STATUS}`;", text, count=1)
    PROGRESS_DOC.write_text(text, encoding="utf-8")


def update_measurements_doc(report: Mapping[str, Any]) -> None:
    entry = f"""### v4_7_2 Source-Grounded Korean Query Review Packet Hydration

- Run: `{RUN_ID}`
- Primary artifacts: `report.json`, `review_packet_ko_hydrated.xlsx`, `review_packet_ko_hydrated.csv`, `review_packet_ko_hydrated.jsonl`, `review_guidelines_ko.md`, `review_summary_ko.json` under `{repo_relative(OUTPUT_DIR)}`.
- Interpretation: hydrated Korean query/evidence/locator previews are machine-owned review hints only. They are not gold/qrels, labels, official metric input, training data, FT-A execution, promotion evidence, product-success evidence, or live readiness evidence.

| Counter | Value |
|---|---:|
| prior_packet_row_count | {report["prior_packet_row_count"]} |
| prior_packet_non_empty_query_count | {report["prior_packet_non_empty_query_count"]} |
| hydrated_packet_row_count | {report["hydrated_packet_row_count"]} |
| hydrated_packet_non_empty_query_count | {report["hydrated_packet_non_empty_query_count"]} |
| hydrated_pdf_row_count | {report["hydrated_pdf_row_count"]} |
| hydrated_xlsx_row_count | {report["hydrated_xlsx_row_count"]} |
| extraction_failed_row_count | {report["extraction_failed_row_count"]} |
| existing_query_reused_count | {report["existing_query_reused_count"]} |
| deterministic_query_generated_count | {report["deterministic_query_generated_count"]} |
| local_llm_query_generated_count | {report["local_llm_query_generated_count"]} |
| official_metric_input_rows | 0 |
| qrels_mutation | false |
| gold_mutation | false |
| label_mutation | false |
| training_dataset_created | false |
"""
    v47.v4610.v469.v467.replace_marked_entry(MEASUREMENTS_DOC, RUN_ID, entry)


def update_triage_doc(report: Mapping[str, Any]) -> None:
    entry = f"""### v4_7_2 Source-Grounded Korean Query Review Packet Hydration Triage

- Run: `{RUN_ID}`
- Root cause: v4_7_1 treated missing query text in the registration artifacts as a reason to leave every review query blank, even though source-disjoint candidate sources were registered and extractable source content was available.
- Fix: v4_7_2 searches for linked query fields first, then uses the local LLM in strict-JSON mode to draft source-grounded Korean query candidates from bounded PDF text and XLSX workbook structure. Deterministic query templates stay disabled and unavailable as a fallback.
- Reviewable rows: {report["hydrated_packet_row_count"]}; PDF {report["hydrated_pdf_row_count"]}; XLSX {report["hydrated_xlsx_row_count"]}; extraction_failed {report["extraction_failed_row_count"]}; non-empty `질의문` {report["hydrated_packet_non_empty_query_count"]}.
- User-owned fields remain blank/default. Machine draft answer/evidence columns are non-official hints and require review.
- It is not official metric, not product-success evidence, not promotion evidence, not FT-A execution, not fine-tuning, not training data, and not live DB/index/cache readiness.
"""
    v47.v4610.v469.v467.replace_marked_entry(TRIAGE_DOC, RUN_ID, entry)


def update_human_docs(report: Mapping[str, Any]) -> None:
    update_root_readme(report)
    update_eval_readme(report)
    update_scripts_readme()
    update_progress_doc(report)
    update_measurements_doc(report)
    update_triage_doc(report)


def run_write(
    *,
    candidate_manifest_path: Path = DEFAULT_CANDIDATE_MANIFEST,
    source_collection_manifest_path: Path = SOURCE_COLLECTION_MANIFEST_CSV,
    source_collection_dir: Path = EXTERNAL_SOURCE_COLLECTION_DIR,
    output_dir: Path = OUTPUT_DIR,
    update_docs: bool = True,
    llm_backend: str = DEFAULT_LLM_BACKEND,
    llm_base_url: str = "",
    llm_model: str = DEFAULT_LLM_MODEL,
    llm_timeout_seconds: int = 120,
    llm_max_tokens: int = 1800,
) -> dict[str, Any]:
    artifacts = build_artifacts(
        candidate_manifest_path=candidate_manifest_path,
        source_collection_manifest_path=source_collection_manifest_path,
        source_collection_dir=source_collection_dir,
        output_dir=output_dir,
        llm_backend=llm_backend,
        llm_base_url=llm_base_url,
        llm_model=llm_model,
        llm_timeout_seconds=llm_timeout_seconds,
        llm_max_tokens=llm_max_tokens,
    )
    check_report(artifacts["report"])
    report = write_artifacts(artifacts, output_dir=output_dir)
    check_written_artifacts(output_dir)
    if update_docs and output_dir == OUTPUT_DIR:
        update_status(report)
        update_human_docs(report)
    return report


def main(argv: list[str] | None = None) -> int:
    parser = argparse.ArgumentParser(description=__doc__)
    parser.add_argument("--check", action="store_true")
    parser.add_argument("--candidate-manifest", type=Path, default=DEFAULT_CANDIDATE_MANIFEST)
    parser.add_argument("--source-collection-manifest", type=Path, default=SOURCE_COLLECTION_MANIFEST_CSV)
    parser.add_argument("--source-collection-dir", type=Path, default=EXTERNAL_SOURCE_COLLECTION_DIR)
    parser.add_argument("--output-dir", type=Path, default=OUTPUT_DIR)
    parser.add_argument("--llm-backend", default=DEFAULT_LLM_BACKEND, choices=["llamacpp", "openai-compatible", "ollama"])
    parser.add_argument("--llm-base-url", default="")
    parser.add_argument("--llm-model", default=DEFAULT_LLM_MODEL)
    parser.add_argument("--llm-timeout-seconds", type=int, default=120)
    parser.add_argument("--llm-max-tokens", type=int, default=1800)
    args = parser.parse_args(argv)
    if args.check:
        check_written_artifacts(args.output_dir)
        report = read_json(args.output_dir / "report.json")
        print(
            json.dumps(
                {
                    "run_id": RUN_ID,
                    "status": report["status"],
                    "hydrated_packet_row_count": report["hydrated_packet_row_count"],
                    "hydrated_packet_non_empty_query_count": report["hydrated_packet_non_empty_query_count"],
                    "hydrated_pdf_row_count": report["hydrated_pdf_row_count"],
                    "hydrated_xlsx_row_count": report["hydrated_xlsx_row_count"],
                    "extraction_failed_row_count": report["extraction_failed_row_count"],
                    "official_metric_input_rows": 0,
                    "v4_7_official_metric_gate_opened": False,
                    "promotion_evidence": False,
                    "product_success_evidence_allowed": False,
                    "ft_a_execution": False,
                    "fine_tuning": False,
                    "live_db_index_cache_readiness": False,
                },
                ensure_ascii=False,
                sort_keys=True,
            )
        )
        return 0
    report = run_write(
        candidate_manifest_path=args.candidate_manifest,
        source_collection_manifest_path=args.source_collection_manifest,
        source_collection_dir=args.source_collection_dir,
        output_dir=args.output_dir,
        update_docs=True,
        llm_backend=args.llm_backend,
        llm_base_url=args.llm_base_url,
        llm_model=args.llm_model,
        llm_timeout_seconds=args.llm_timeout_seconds,
        llm_max_tokens=args.llm_max_tokens,
    )
    print(json.dumps({"run_id": RUN_ID, "status": report["status"], "report": report["artifact_paths"]["report_json"]}, sort_keys=True))
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
