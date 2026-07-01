"""Build approved index-time XLSX workbook snapshots for SourceAtom indexing."""

from __future__ import annotations

import argparse
import hashlib
import json
import re
from datetime import date, datetime, time
from pathlib import Path
from typing import Any

from openpyxl import load_workbook
from openpyxl.utils import get_column_letter, range_boundaries

XLSX_PIPELINE_VERSION = "xlsx-extract-v2-hidden-safe"
SNAPSHOT_APPROVAL_POLICY = "source_owned_index_time_exact_hash_row_cells_v1"


def _sha256_bytes(value: bytes) -> str:
    return hashlib.sha256(value).hexdigest()


def _sha256_text(value: str) -> str:
    return hashlib.sha256(value.encode("utf-8")).hexdigest()


def _canonical_sha256(value: Any) -> str:
    text = _clean(value)
    if text.casefold().startswith("sha256:"):
        text = text.split(":", 1)[1]
    if not re.fullmatch(r"[0-9a-fA-F]{64}", text):
        return ""
    return text.lower()


def _clean(value: Any) -> str:
    return re.sub(r"\s+", " ", str(value or "")).strip()


def _display_value(value: Any) -> str:
    if value is None:
        return ""
    if isinstance(value, datetime):
        return value.isoformat(sep=" ", timespec="seconds")
    if isinstance(value, date):
        return value.isoformat()
    if isinstance(value, time):
        return value.isoformat(timespec="seconds")
    if isinstance(value, bool):
        return "true" if value else "false"
    return _clean(value)


def _hidden_columns(worksheet: Any) -> set[int]:
    hidden: set[int] = set()
    for column_key, dimension in worksheet.column_dimensions.items():
        if not getattr(dimension, "hidden", False):
            continue
        start = getattr(dimension, "min", None)
        end = getattr(dimension, "max", None)
        if isinstance(start, int) and isinstance(end, int) and start > 0 and end >= start:
            hidden.update(range(start, end + 1))
            continue
        try:
            column_index = range_boundaries(f"{column_key}1:{column_key}1")[0]
        except ValueError:
            continue
        hidden.add(column_index)
    return hidden


def _source_atom_id_for_cell(
    *,
    raw_xlsx_sha256: str,
    document_version_id: str,
    sheet_name: str,
    cell_range: str,
    cell_ref: str,
    header: str,
) -> str:
    basis = json.dumps(
        {
            "raw_xlsx_sha256": raw_xlsx_sha256,
            "document_version_id": document_version_id,
            "sheet": sheet_name,
            "cell_range": cell_range,
            "cell": cell_ref,
            "header": header,
        },
        ensure_ascii=False,
        sort_keys=True,
    )
    return f"srcatom_xlsx_snapshot_cell_{_sha256_text(basis)[:24]}"


def build_approved_xlsx_workbook_snapshot(
    workbook_path: Path | str,
    *,
    expected_sha256: str,
    document_version_id: str,
    sheet_name: str,
    cell_range: str,
    row_index_1based: int,
    header_row_index: int = 1,
) -> dict[str, Any]:
    """Read a raw XLSX at index time and return a bounded, approved row snapshot."""

    path = Path(workbook_path)
    content = path.read_bytes()
    actual_sha256 = _sha256_bytes(content)
    expected_canonical_sha256 = _canonical_sha256(expected_sha256)
    if not expected_canonical_sha256:
        raise ValueError("expected_sha256 must be a canonical SHA-256 hex digest")
    if actual_sha256 != expected_canonical_sha256:
        raise ValueError(
            "raw XLSX SHA-256 mismatch: "
            f"expected={expected_canonical_sha256} actual={actual_sha256}"
        )
    doc_id = _clean(document_version_id)
    if not doc_id:
        raise ValueError("document_version_id is required")

    min_col, min_row, max_col, max_row = range_boundaries(cell_range)
    if row_index_1based < min_row or row_index_1based > max_row:
        raise ValueError("row_index_1based must be inside cell_range")

    formulas_workbook = load_workbook(
        path,
        read_only=False,
        data_only=False,
        keep_links=False,
        keep_vba=False,
    )
    values_workbook = load_workbook(
        path,
        read_only=True,
        data_only=True,
        keep_links=False,
        keep_vba=False,
    )
    try:
        if sheet_name not in formulas_workbook.sheetnames or sheet_name not in values_workbook.sheetnames:
            raise ValueError(f"sheet not found: {sheet_name}")
        formula_sheet = formulas_workbook[sheet_name]
        value_sheet = values_workbook[sheet_name]
        if formula_sheet.sheet_state != "visible":
            raise ValueError(f"sheet is hidden: {sheet_name}")
        if getattr(formula_sheet.row_dimensions[row_index_1based], "hidden", False):
            raise ValueError(f"target row is hidden: {row_index_1based}")
        if getattr(formula_sheet.row_dimensions[header_row_index], "hidden", False):
            raise ValueError(f"header row is hidden: {header_row_index}")

        hidden_columns = _hidden_columns(formula_sheet)
        headers: list[str] = []
        for column in range(min_col, max_col + 1):
            if column in hidden_columns:
                headers.append("")
                continue
            header_formula_cell = formula_sheet.cell(header_row_index, column)
            if header_formula_cell.data_type == "f" or _clean(header_formula_cell.value).startswith("="):
                raise ValueError(f"header cell contains a formula: {get_column_letter(column)}{header_row_index}")
            headers.append(_display_value(value_sheet.cell(header_row_index, column).value))

        cells: list[dict[str, Any]] = []
        hidden_column_letters = [
            get_column_letter(column)
            for column in sorted(hidden_columns)
            if min_col <= column <= max_col
        ]
        for column in range(min_col, max_col + 1):
            cell_ref = f"{get_column_letter(column)}{row_index_1based}"
            if column in hidden_columns:
                continue
            formula_cell = formula_sheet.cell(row_index_1based, column)
            if formula_cell.data_type == "f" or _clean(formula_cell.value).startswith("="):
                continue
            value = _display_value(value_sheet.cell(row_index_1based, column).value)
            if not value:
                continue
            header = headers[column - min_col] if column - min_col < len(headers) else ""
            cells.append(
                {
                    "cell": cell_ref,
                    "cellRef": cell_ref,
                    "row": row_index_1based,
                    "column": column,
                    "columnLetter": get_column_letter(column),
                    "header": header,
                    "value": value,
                    "formattedValue": value,
                    "hiddenRow": False,
                    "hiddenColumn": False,
                    "sourceAtomId": _source_atom_id_for_cell(
                        raw_xlsx_sha256=actual_sha256,
                        document_version_id=doc_id,
                        sheet_name=sheet_name,
                        cell_range=cell_range,
                        cell_ref=cell_ref,
                        header=header,
                    ),
                }
            )
        if not cells:
            raise ValueError("approved snapshot would contain no visible non-formula cells")

        return {
            "fileType": "xlsx",
            "pipelineVersion": XLSX_PIPELINE_VERSION,
            "snapshotApprovalPolicy": SNAPSHOT_APPROVAL_POLICY,
            "rawXlsxSha256": actual_sha256,
            "documentVersionId": doc_id,
            "hiddenPolicy": "exclude_hidden",
            "hiddenPolicyVersion": "exclude-hidden-v1",
            "sanitizerVersion": "exclude-hidden-v1",
            "extractor": "openpyxl-index-time-target-row",
            "security": {
                "keepLinks": False,
                "macrosExecuted": False,
            },
            "warnings": [],
            "workbook": {
                "role": "workbook",
                "sheetCount": 1,
                "visibleSheetCount": 1,
                "sheets": [
                    {
                        "role": "sheet",
                        "name": sheet_name,
                        "sheetName": sheet_name,
                        "index": formula_sheet.parent.worksheets.index(formula_sheet),
                        "sheetIndex": formula_sheet.parent.worksheets.index(formula_sheet),
                        "hidden": False,
                        "usedRange": f"{get_column_letter(min_col)}{header_row_index}:{get_column_letter(max_col)}{max_row}",
                        "cellRange": f"{get_column_letter(min_col)}{header_row_index}:{get_column_letter(max_col)}{max_row}",
                        "rowStart": header_row_index,
                        "rowEnd": max_row,
                        "columnStart": min_col,
                        "columnEnd": max_col,
                        "hiddenRows": [],
                        "hiddenColumns": hidden_column_letters,
                        "hiddenColumnIndexes": sorted(
                            column for column in hidden_columns if min_col <= column <= max_col
                        ),
                        "hiddenCells": [],
                        "warnings": [],
                        "indexable": True,
                        "tables": [
                            {
                                "role": "table",
                                "type": "approved_row_window",
                                "name": "ApprovedSourceOwnedRowWindow",
                                "tableId": "ApprovedSourceOwnedRowWindow",
                                "range": f"{get_column_letter(min_col)}{header_row_index}:{get_column_letter(max_col)}{max_row}",
                                "cellRange": f"{get_column_letter(min_col)}{header_row_index}:{get_column_letter(max_col)}{max_row}",
                                "rowStart": header_row_index,
                                "rowEnd": max_row,
                                "columnStart": min_col,
                                "columnEnd": max_col,
                                "headers": headers,
                            }
                        ],
                        "cells": cells,
                    }
                ],
            },
        }
    finally:
        formulas_workbook.close()
        values_workbook.close()


def build_parser() -> argparse.ArgumentParser:
    parser = argparse.ArgumentParser(description=__doc__)
    parser.add_argument("--input-xlsx", required=True, type=Path)
    parser.add_argument("--expected-sha256", required=True)
    parser.add_argument("--document-version-id", required=True)
    parser.add_argument("--sheet", required=True)
    parser.add_argument("--cell-range", required=True)
    parser.add_argument("--row-index-1based", required=True, type=int)
    parser.add_argument("--header-row-index", type=int, default=1)
    parser.add_argument("--output", required=True, type=Path)
    return parser


def main(argv: list[str] | None = None) -> int:
    args = build_parser().parse_args(argv)
    payload = build_approved_xlsx_workbook_snapshot(
        args.input_xlsx,
        expected_sha256=args.expected_sha256,
        document_version_id=args.document_version_id,
        sheet_name=args.sheet,
        cell_range=args.cell_range,
        row_index_1based=args.row_index_1based,
        header_row_index=args.header_row_index,
    )
    args.output.parent.mkdir(parents=True, exist_ok=True)
    args.output.write_text(json.dumps(payload, ensure_ascii=False, indent=2), encoding="utf-8")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
