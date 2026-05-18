"""Fail-closed readiness gate for official-denominator SearchUnit indexing.

This module prepares the non-production official answer/citation denominator
index path. It deliberately does not synthesize source content from gold
answers or report-only candidate rows. If source-bound SearchUnit locator
fields are incomplete, it writes a compact diagnostic artifact and refuses the
build.
"""

from __future__ import annotations

import argparse
import hashlib
import json
import re
import sys
from collections import Counter
from datetime import datetime, timezone
from pathlib import Path
from typing import Any, Mapping, Sequence

from openpyxl import load_workbook
from openpyxl.utils.cell import coordinate_to_tuple, get_column_letter


AI_WORKER_ROOT = Path(__file__).resolve().parents[1]
REPO_ROOT = AI_WORKER_ROOT.parent
REPORT_DIR = AI_WORKER_ROOT / "eval" / "reports" / "rag-ingestion"
DEFAULT_METRIC_INPUT_CONFIG = REPORT_DIR / "metric_input_v1.json"
DEFAULT_READINESS_JSON = REPORT_DIR / "source_bound_readiness_v1.json"
DEFAULT_STATUS_JSONL = REPORT_DIR / "status.jsonl"

SOURCE_BOUND_INDEX_VERSION = (
    "official-answer-citation-agentic-loop-v1-nonprod-official-denominator-source-bound"
)
TARGET_INDEX_PATH = Path("ai/eval/indexes/rag-data-official-denominator-v1")
SEARCH_UNIT_MANIFEST_FILE = "search_unit_manifest.jsonl"
INGEST_MANIFEST_FILE = "ingest_manifest.json"
OFFICIAL_SOURCE_BOUND_EMBEDDING_TEXT_VARIANT = "official_denominator_source_bound_search_unit_v1"
OFFICIAL_SOURCE_BOUND_BUILDER_VERSION = "official-denominator-source-bound-v1"

SCHEMA_VERSION = "official_answer_citation_source_bound_index_build_readiness_v1"
BLOCKER_SOURCE_FIELDS_MISSING = "SOURCE_BOUND_OFFICIAL_DENOMINATOR_SOURCE_FIELDS_MISSING"
BLOCKER_SOURCE_FILES_MISSING = "SOURCE_BOUND_OFFICIAL_DENOMINATOR_SOURCE_FILES_MISSING"
BLOCKER_DENOMINATOR_CONTRACT_INVALID = "SOURCE_BOUND_OFFICIAL_DENOMINATOR_CONTRACT_INVALID"
BLOCKER_INDEX_BUILD_FAILED = "SOURCE_BOUND_OFFICIAL_DENOMINATOR_INDEX_BUILD_FAILED"
BLOCKER_INDEX_LOAD_CHECK_FAILED = "SOURCE_BOUND_OFFICIAL_DENOMINATOR_INDEX_LOAD_CHECK_FAILED"

EXPECTED_OFFICIAL_ROWS_BY_TRACK = {
    "pdf_business_ocr_mm": 4,
    "text_namu_v2_1": 6,
    "xlsx_business_structured": 19,
}
EXTERNAL_REPO_ROOT_NAMES = (
    "_external_runtime_artifacts",
    "_external_workspace_archive",
)
PDF_SOURCE_LOCATOR_MANIFEST_NAMES = (
    "pdf_answer_citation_diagnostic_review_input.jsonl",
)

REQUIRED_FIELDS_BY_TRACK: dict[str, tuple[str, ...]] = {
    "text_namu_v2_1": (
        "document_id",
        "document_version_id",
        "search_unit_id",
        "text_locator",
    ),
    "xlsx_business_structured": (
        "workbook",
        "sheet",
        "range",
        "cell",
        "row_label",
        "target_column",
        "normalized_value",
        "search_unit_id",
        "document_version_id",
    ),
    "pdf_business_ocr_mm": (
        "source_pdf_path",
        "page",
        "physical_page_index",
        "bbox",
        "region_type",
        "row_label",
        "target_column",
        "search_unit_id",
        "document_version_id",
    ),
}


def main(argv: list[str] | None = None) -> int:
    args = parse_args(argv)
    report = build_readiness_report(
        metric_input_config_path=Path(args.metric_input_config),
        output_index=Path(args.output_index),
        index_version=args.index_version,
        source_roots=[Path(item) for item in args.source_root],
    )
    if report["build_ready"]:
        report = build_and_load_check_ready_index(
            report,
            output_index=Path(args.output_index),
            index_version=args.index_version,
        )
    write_json(Path(args.readiness_json), report)
    append_status_event(Path(args.status_jsonl), report)
    print(json.dumps({
        "status": report["status"],
        "blocker_category": report["blocker_category"],
        "build_ready": report["build_ready"],
        "target_index_built": report["target_index_built"],
        "load_check_passed": report["load_check_passed"],
        "rerun_allowed": report["rerun_allowed"],
        "readiness_json": repo_relative(Path(args.readiness_json)),
    }, ensure_ascii=False, sort_keys=True))
    return 0 if report["rerun_allowed"] else 2


def parse_args(argv: list[str] | None = None) -> argparse.Namespace:
    parser = argparse.ArgumentParser(description=__doc__)
    parser.add_argument("--metric-input-config", default=str(DEFAULT_METRIC_INPUT_CONFIG))
    parser.add_argument("--output-index", default=str(TARGET_INDEX_PATH))
    parser.add_argument("--index-version", default=SOURCE_BOUND_INDEX_VERSION)
    parser.add_argument("--readiness-json", default=str(DEFAULT_READINESS_JSON))
    parser.add_argument("--status-jsonl", default=str(DEFAULT_STATUS_JSONL))
    parser.add_argument(
        "--source-root",
        action="append",
        default=[],
        help="Repo-local source root to check for referenced XLSX/PDF files. Repeatable.",
    )
    return parser.parse_args(argv)


def build_readiness_report(
    *,
    metric_input_config_path: Path,
    output_index: Path,
    index_version: str,
    source_roots: Sequence[Path] | None = None,
) -> dict[str, Any]:
    config = read_json(metric_input_config_path)
    rows = list(config.get("candidate_manifest") or [])
    source_roots = list(source_roots or default_source_roots())
    resolver = SourceBoundResolver(source_roots)
    missing_fields_by_query_id: dict[str, list[str]] = {}
    normalized_locator_by_query_id: dict[str, dict[str, Any]] = {}
    missing_source_files_by_query_id: dict[str, list[str]] = {}
    source_file_inventory_by_query_id: dict[str, list[dict[str, Any]]] = {}
    rows_by_track = Counter()

    for row in rows:
        query_id = clean(row.get("query_id"))
        track = clean(row.get("track"))
        rows_by_track[track] += 1
        locator = mapping(row.get("citation_locator"))
        normalized = normalize_locator(track, locator)
        normalized = resolver.resolve(track=track, normalized=normalized, raw_locator=locator)
        normalized["query_id"] = query_id
        normalized["track"] = track
        normalized_locator_by_query_id[query_id] = normalized
        source_file_inventory_by_query_id[query_id] = resolver.inventory_for(
            track=track,
            normalized=normalized,
            raw_locator=locator,
        )
        missing_fields = [
            field
            for field in REQUIRED_FIELDS_BY_TRACK.get(track, ())
            if not has_required_value(normalized.get(field), field=field)
        ]
        if missing_fields:
            missing_fields_by_query_id[query_id] = missing_fields

        missing_sources = missing_source_files(normalized, track=track, source_roots=source_roots)
        if not missing_sources:
            missing_sources = resolver.missing_source_references(
                track=track,
                normalized=normalized,
                raw_locator=locator,
            )
        if missing_sources:
            missing_source_files_by_query_id[query_id] = missing_sources

    blocked_query_ids = [
        clean(row.get("query_id"))
        for row in rows
        if clean(row.get("query_id")) in missing_fields_by_query_id
        or clean(row.get("query_id")) in missing_source_files_by_query_id
    ]
    denominator_errors = denominator_contract_errors(rows=rows, rows_by_track=rows_by_track)
    build_ready = not blocked_query_ids and not denominator_errors
    if missing_fields_by_query_id:
        blocker_category = BLOCKER_SOURCE_FIELDS_MISSING
    elif missing_source_files_by_query_id:
        blocker_category = BLOCKER_SOURCE_FILES_MISSING
    elif denominator_errors:
        blocker_category = BLOCKER_DENOMINATOR_CONTRACT_INVALID
    else:
        blocker_category = None

    return {
        "schema_version": SCHEMA_VERSION,
        "generated_at": utc_timestamp(),
        "entrypoint_implemented": True,
        "status": "BUILD_READY" if build_ready else "FAIL_CLOSED",
        "blocker_category": blocker_category,
        "target_index_path": repo_relative(resolve_repo_path(output_index)),
        "index_version": index_version,
        "non_production_only": True,
        "official_denominator_rows": len(rows),
        "official_rows_by_track": dict(sorted(rows_by_track.items())),
        "expected_official_rows_by_track": dict(sorted(EXPECTED_OFFICIAL_ROWS_BY_TRACK.items())),
        "denominator_contract_errors": denominator_errors,
        "required_fields_by_track": {
            track: list(fields) for track, fields in REQUIRED_FIELDS_BY_TRACK.items()
        },
        "missing_fields_by_query_id": missing_fields_by_query_id,
        "missing_source_files_by_query_id": missing_source_files_by_query_id,
        "blocked_query_ids": blocked_query_ids,
        "source_bound_locators_by_query_id": normalized_locator_by_query_id,
        "source_file_inventory_by_query_id": source_file_inventory_by_query_id,
        "normalized_source_locator_sample": {
            query_id: normalized_locator_by_query_id[query_id]
            for query_id in list(normalized_locator_by_query_id)[:5]
        },
        "source_roots_checked": [repo_relative(resolve_repo_path(path)) for path in source_roots],
        "build_ready": build_ready,
        "target_index_built": False,
        "load_check_passed": False,
        "rerun_allowed": False,
        "production_index_path_used": False,
        "candidate_index_path_used": False,
        "candidate_artifacts_as_generation_source": False,
        "generation_used_expected_answer": False,
        "generation_used_supporting_evidence": False,
        "generation_used_gold_fields": False,
        "promotion_evidence": False,
        "baseline_overwrite": False,
        "gold_mutation": False,
        "denominator_mutation": False,
        "human_label_mutation": False,
        "decision_rationale": (
            "The official source-bound index build may proceed only when every "
            "official denominator row has source identity and track-specific "
            "SearchUnit locator fields. Gold expected/supporting fields are not "
            "used as generation content."
        ),
    }


def build_and_load_check_ready_index(
    report: Mapping[str, Any],
    *,
    output_index: Path,
    index_version: str,
) -> dict[str, Any]:
    updated = dict(report)
    try:
        settings = runtime_worker_settings()
        embedder = runtime_embedder(settings)
        build_result = build_source_bound_index_from_readiness(
            updated,
            output_index=output_index,
            index_version=index_version,
            embedder=embedder,
            build_device=clean(getattr(settings, "rag_faiss_build_device", "cpu")) or "cpu",
            max_seq_length=getattr(embedder, "max_seq_length", None),
        )
    except Exception as exc:  # noqa: BLE001 - fail-closed report detail is intentional.
        updated.update(
            {
                "status": "FAIL_CLOSED",
                "blocker_category": BLOCKER_INDEX_BUILD_FAILED,
                "target_index_built": False,
                "load_check_passed": False,
                "rerun_allowed": False,
                "index_build_error": f"{type(exc).__name__}: {exc}",
            }
        )
        return updated

    load_check = load_check_source_bound_index(
        output_index,
        readiness_report=updated,
        runtime_embedding_model=build_result.get("embedding_model"),
    )
    updated["index_build_result"] = build_result
    updated["index_load_check"] = load_check
    updated["target_index_built"] = bool(build_result.get("built"))
    updated["load_check_passed"] = bool(load_check.get("passed"))
    updated["rerun_allowed"] = bool(load_check.get("passed"))
    if load_check.get("passed"):
        updated["status"] = "BUILD_READY_LOAD_CHECK_PASSED"
        updated["blocker_category"] = None
    else:
        updated["status"] = "FAIL_CLOSED"
        updated["blocker_category"] = BLOCKER_INDEX_LOAD_CHECK_FAILED
    return updated


def build_source_bound_index_from_readiness(
    readiness_report: Mapping[str, Any],
    *,
    output_index: Path,
    index_version: str,
    embedder: Any,
    build_device: str = "cpu",
    max_seq_length: Any = None,
) -> dict[str, Any]:
    if readiness_report.get("build_ready") is not True:
        raise RuntimeError("source-bound official denominator readiness is not build_ready")
    manifest_rows = build_search_unit_manifest_rows(readiness_report)
    if len(manifest_rows) != 29:
        raise RuntimeError(f"expected 29 source-bound SearchUnit rows, got {len(manifest_rows)}")
    search_unit_ids = [clean(row.get("search_unit_id")) for row in manifest_rows]
    if len(search_unit_ids) != len(set(search_unit_ids)):
        raise RuntimeError("source-bound SearchUnit manifest contains duplicate search_unit_id values")
    texts_to_embed = [clean(row.get("embedding_text")) for row in manifest_rows]
    if any(not text for text in texts_to_embed):
        raise RuntimeError("source-bound SearchUnit manifest contains empty embedding text")

    ensure_ai_worker_on_path()
    from app.capabilities.rag.faiss_index import FaissIndex

    index_dir = resolve_repo_path(output_index)
    index = FaissIndex(index_dir, build_device=build_device)
    vectors = embedder.embed_passages(texts_to_embed)
    if vectors.shape[0] != len(manifest_rows):
        raise RuntimeError(
            f"embedder returned {vectors.shape[0]} vectors for {len(manifest_rows)} SearchUnits"
        )
    info, stage_dir = index.build_staged(
        vectors,
        index_version=index_version,
        embedding_model=embedder.model_name,
    )
    try:
        write_jsonl(stage_dir / SEARCH_UNIT_MANIFEST_FILE, manifest_rows)
        ingest_manifest = official_ingest_manifest(
            readiness_report=readiness_report,
            manifest_rows=manifest_rows,
            info=info,
            max_seq_length=max_seq_length,
            output_index=index_dir,
        )
        write_json(stage_dir / INGEST_MANIFEST_FILE, ingest_manifest)
        enrich_build_json(
            stage_dir / "build.json",
            readiness_report=readiness_report,
            manifest_rows=manifest_rows,
            search_unit_manifest_file=SEARCH_UNIT_MANIFEST_FILE,
        )
        index.promote_staged(
            stage_dir,
            info,
            extra_files=(INGEST_MANIFEST_FILE, SEARCH_UNIT_MANIFEST_FILE),
        )
    except Exception:
        index.discard_staged(stage_dir)
        raise

    track_counts = dict(sorted(Counter(row["track"] for row in manifest_rows).items()))
    return {
        "built": True,
        "target_index_path": repo_relative(index_dir),
        "index_version": info.index_version,
        "embedding_model": info.embedding_model,
        "dimension": info.dimension,
        "official_denominator_rows": len(manifest_rows),
        "track_counts": track_counts,
        "search_unit_manifest_file": SEARCH_UNIT_MANIFEST_FILE,
        "ingest_manifest_file": INGEST_MANIFEST_FILE,
        "non_production_only": True,
        "official_denominator_source_bound": True,
        "candidate_artifacts_as_generation_source": False,
        "production_index_path_used": False,
    }


def load_check_source_bound_index(
    output_index: Path,
    *,
    readiness_report: Mapping[str, Any],
    runtime_embedding_model: str | None = None,
) -> dict[str, Any]:
    ensure_ai_worker_on_path()
    from app.capabilities.rag.faiss_index import FaissIndex

    index_dir = resolve_repo_path(output_index)
    blockers: list[str] = []
    required_files = ("faiss.index", "build.json", INGEST_MANIFEST_FILE, SEARCH_UNIT_MANIFEST_FILE)
    missing_files = [name for name in required_files if not (index_dir / name).exists()]
    if missing_files:
        blockers.append("missing_files:" + ",".join(missing_files))
    try:
        build_payload = read_json(index_dir / "build.json") if (index_dir / "build.json").exists() else {}
    except Exception as exc:  # noqa: BLE001
        build_payload = {}
        blockers.append(f"build_json_unreadable:{type(exc).__name__}")
    try:
        ingest_manifest = (
            read_json(index_dir / INGEST_MANIFEST_FILE)
            if (index_dir / INGEST_MANIFEST_FILE).exists()
            else {}
        )
    except Exception as exc:  # noqa: BLE001
        ingest_manifest = {}
        blockers.append(f"ingest_manifest_unreadable:{type(exc).__name__}")
    try:
        manifest_rows = read_jsonl(index_dir / SEARCH_UNIT_MANIFEST_FILE)
    except Exception as exc:  # noqa: BLE001
        manifest_rows = []
        if (index_dir / SEARCH_UNIT_MANIFEST_FILE).exists():
            blockers.append(f"search_unit_manifest_unreadable:{type(exc).__name__}")

    info = None
    if not missing_files:
        try:
            info = FaissIndex(index_dir).load()
        except Exception as exc:  # noqa: BLE001
            blockers.append(f"faiss_load_failed:{type(exc).__name__}:{exc}")

    provenance = mapping(ingest_manifest.get("official_denominator_source_bound_provenance"))
    if build_payload.get("index_version") != SOURCE_BOUND_INDEX_VERSION:
        blockers.append("build_json_index_version_mismatch")
    if ingest_manifest.get("index_version") != SOURCE_BOUND_INDEX_VERSION:
        blockers.append("ingest_manifest_index_version_mismatch")
    if runtime_embedding_model:
        if clean(build_payload.get("embedding_model")) != runtime_embedding_model:
            blockers.append("build_json_runtime_model_mismatch")
        if clean(ingest_manifest.get("embedding_model")) != runtime_embedding_model:
            blockers.append("ingest_manifest_runtime_model_mismatch")
    if build_payload.get("official_denominator_source_bound") is not True:
        blockers.append("build_json_source_bound_provenance_missing")
    if provenance.get("official_denominator_source_bound") is not True:
        blockers.append("ingest_manifest_source_bound_provenance_missing")
    if provenance.get("non_production_only") is not True:
        blockers.append("ingest_manifest_non_production_flag_missing")

    track_counts = Counter(clean(row.get("track")) for row in manifest_rows)
    row_count = len(manifest_rows)
    if info is not None and int(info.chunk_count) != row_count:
        blockers.append(f"faiss_chunk_count_mismatch:{info.chunk_count}!={row_count}")
    if row_count != 29:
        blockers.append(f"search_unit_manifest_rows_expected_29_actual_{row_count}")
    if dict(track_counts) != EXPECTED_OFFICIAL_ROWS_BY_TRACK:
        blockers.append("search_unit_manifest_track_counts_mismatch")
    if clean(provenance.get("official_denominator_rows")) and int(provenance.get("official_denominator_rows")) != 29:
        blockers.append("provenance_official_denominator_rows_mismatch")

    schema_missing = source_bound_manifest_schema_missing(manifest_rows)
    if schema_missing:
        blockers.append("search_unit_manifest_required_fields_missing")
    duplicate_search_units = duplicate_values(clean(row.get("search_unit_id")) for row in manifest_rows)
    if duplicate_search_units:
        blockers.append("search_unit_manifest_duplicate_search_unit_id")
    if readiness_report.get("candidate_artifacts_as_generation_source") is not False:
        blockers.append("readiness_candidate_generation_source_guard_failed")
    if readiness_report.get("generation_used_expected_answer") is not False:
        blockers.append("readiness_expected_answer_generation_guard_failed")
    if readiness_report.get("generation_used_supporting_evidence") is not False:
        blockers.append("readiness_supporting_evidence_generation_guard_failed")
    if readiness_report.get("generation_used_gold_fields") is not False:
        blockers.append("readiness_gold_generation_guard_failed")

    return {
        "passed": not blockers,
        "target_index_path": repo_relative(index_dir),
        "required_files": list(required_files),
        "missing_files": missing_files,
        "blockers": blockers,
        "index_version": build_payload.get("index_version"),
        "runtime_embedding_model": runtime_embedding_model,
        "build_embedding_model": build_payload.get("embedding_model"),
        "ingest_manifest_embedding_model": ingest_manifest.get("embedding_model"),
        "official_denominator_rows": row_count,
        "track_counts": dict(sorted(track_counts.items())),
        "expected_track_counts": dict(sorted(EXPECTED_OFFICIAL_ROWS_BY_TRACK.items())),
        "required_locator_schema_coverage_passed": not schema_missing,
        "missing_fields_by_manifest_row": schema_missing,
        "source_bound_provenance_passed": (
            provenance.get("official_denominator_source_bound") is True
            and provenance.get("non_production_only") is True
        ),
    }


def build_search_unit_manifest_rows(readiness_report: Mapping[str, Any]) -> list[dict[str, Any]]:
    locators = mapping(readiness_report.get("source_bound_locators_by_query_id"))
    rows: list[dict[str, Any]] = []
    for faiss_row_id, query_id in enumerate(sorted(locators)):
        locator = mapping(locators[query_id])
        track = clean(locator.get("track")) or infer_track_from_locator(locator)
        if track not in REQUIRED_FIELDS_BY_TRACK:
            raise RuntimeError(f"unknown source-bound track for query_id={query_id}: {track!r}")
        missing = [
            field
            for field in REQUIRED_FIELDS_BY_TRACK[track]
            if not has_required_value(locator.get(field), field=field)
        ]
        if missing:
            raise RuntimeError(f"query_id={query_id} missing required source fields: {missing}")
        source_locator = source_locator_for_manifest(track=track, locator=locator)
        embedding_text, display_text, bm25_text = source_bound_text_surfaces(
            query_id=query_id,
            track=track,
            locator=locator,
        )
        citation_payload = canonical_citation_payload(track=track, locator=locator)
        content_sha256 = clean(locator.get("source_content_sha256")) or sha256_text(display_text)
        rows.append(
            {
                "faiss_row_id": faiss_row_id,
                "query_id": query_id,
                "track": track,
                "search_unit_id": clean(locator.get("search_unit_id")),
                "document_version_id": clean(locator.get("document_version_id")),
                "source_identity": source_identity_for_locator(track=track, locator=locator),
                "source_content_sha256": content_sha256,
                "embedding_text": embedding_text,
                "embedding_text_sha256": sha256_text(embedding_text),
                "bm25_text": bm25_text,
                "display_text": display_text,
                "locator": source_locator,
                "canonical_citation_payload": citation_payload,
                "source_bound_official_denominator": True,
                "non_production_only": True,
                "promotion_evidence": False,
                "candidate_artifact_generation_source": False,
            }
        )
    return rows


def source_bound_text_surfaces(
    *,
    query_id: str,
    track: str,
    locator: Mapping[str, Any],
) -> tuple[str, str, str]:
    if track == "text_namu_v2_1":
        text = clean(locator.get("source_text"))
        if not text:
            raise RuntimeError(f"text source content missing for query_id={query_id}")
        text_locator = mapping(locator.get("text_locator"))
        title = clean(text_locator.get("title"))
        section_path = " > ".join(clean(item) for item in (text_locator.get("section_path") or []) if clean(item))
        header = " | ".join(part for part in (title, section_path) if part)
        display = text
        embedding = "\n".join(part for part in (header, text) if part)
        return embedding, display, text
    if track == "xlsx_business_structured":
        display = (
            f"{clean(locator.get('workbook'))} / {clean(locator.get('sheet'))} "
            f"{clean(locator.get('cell'))}: {clean(locator.get('row_label'))} | "
            f"{clean(locator.get('target_column'))}={clean(locator.get('normalized_value'))}"
        )
        embedding = "\n".join(
            [
                f"Workbook: {clean(locator.get('workbook'))}",
                f"Sheet: {clean(locator.get('sheet'))}",
                f"Range: {clean(locator.get('range'))}",
                f"Cell: {clean(locator.get('cell'))}",
                f"Row: {clean(locator.get('row_label'))}",
                f"Column: {clean(locator.get('target_column'))}",
                f"Value: {clean(locator.get('normalized_value'))}",
            ]
        )
        return embedding, display, display
    if track == "pdf_business_ocr_mm":
        source_lines = [
            clean(mapping(line).get("text"))
            for line in mapping(locator.get("pdf_source_text_locator")).get("source_lines", [])
        ]
        source_text = " | ".join(line for line in source_lines if line)
        locator_text = (
            f"{clean(locator.get('source_pdf_path'))} page {clean(locator.get('page'))}: "
            f"{clean(locator.get('row_label'))} | {clean(locator.get('target_column'))}"
        )
        display = (
            f"{source_text} ({locator_text})"
            if source_text
            else locator_text
        )
        embedding = "\n".join(
            part
            for part in (
                f"PDF: {clean(locator.get('source_pdf_path'))}",
                f"Page: {clean(locator.get('page'))}",
                f"Region: {clean(locator.get('region_type'))}",
                f"Row: {clean(locator.get('row_label'))}",
                f"Column: {clean(locator.get('target_column'))}",
                f"Source text: {source_text}" if source_text else "",
            )
            if part
        )
        return embedding, display, source_text or locator_text
    raise RuntimeError(f"unsupported source-bound track: {track}")


def official_ingest_manifest(
    *,
    readiness_report: Mapping[str, Any],
    manifest_rows: Sequence[Mapping[str, Any]],
    info: Any,
    max_seq_length: Any,
    output_index: Path,
) -> dict[str, Any]:
    texts = [clean(row.get("embedding_text")) for row in manifest_rows]
    track_counts = Counter(clean(row.get("track")) for row in manifest_rows)
    return {
        "embedding_text_variant": OFFICIAL_SOURCE_BOUND_EMBEDDING_TEXT_VARIANT,
        "embedding_text_builder_version": OFFICIAL_SOURCE_BOUND_BUILDER_VERSION,
        "embedding_model": info.embedding_model,
        "max_seq_length": int(max_seq_length) if max_seq_length else None,
        "chunk_count": int(info.chunk_count),
        "document_count": len({clean(row.get("document_version_id")) for row in manifest_rows}),
        "dimension": int(info.dimension),
        "index_version": info.index_version,
        "corpus_path": "official-denominator-source-bound-readiness",
        "embed_text_sha256": digest_texts(texts),
        "embed_text_samples": [
            {
                "row": index,
                "query_id": row.get("query_id"),
                "search_unit_id": row.get("search_unit_id"),
                "track": row.get("track"),
                "sha256": row.get("embedding_text_sha256"),
                "preview": clean(row.get("embedding_text"))[:240],
            }
            for index, row in enumerate(manifest_rows[:5])
        ],
        "official_denominator_source_bound_provenance": {
            "official_denominator_source_bound": True,
            "non_production_only": True,
            "source_bound_locator_schema_covered": True,
            "target_index_path": repo_relative(output_index),
            "schema_version": SCHEMA_VERSION,
            "index_version": info.index_version,
            "official_denominator_rows": len(manifest_rows),
            "track_counts": dict(sorted(track_counts.items())),
            "expected_track_counts": dict(sorted(EXPECTED_OFFICIAL_ROWS_BY_TRACK.items())),
            "readiness_report_sha256": sha256_text(
                json.dumps(readiness_report, ensure_ascii=False, sort_keys=True, default=str)
            ),
            "required_fields_by_track": {
                track: list(fields) for track, fields in REQUIRED_FIELDS_BY_TRACK.items()
            },
            "search_unit_manifest_file": SEARCH_UNIT_MANIFEST_FILE,
            "candidate_artifacts_as_generation_source": False,
            "generation_used_expected_answer": False,
            "generation_used_supporting_evidence": False,
            "generation_used_gold_fields": False,
            "production_index_path_used": False,
            "candidate_index_path_used": False,
            "promotion_evidence": False,
            "baseline_overwrite": False,
        },
        "official_denominator_source_bound": True,
        "non_production_only": True,
        "source_bound_locator_schema_covered": True,
        "official_denominator_rows": len(manifest_rows),
        "track_counts": dict(sorted(track_counts.items())),
        "candidate_artifacts_as_generation_source": False,
        "generation_used_expected_answer": False,
        "generation_used_supporting_evidence": False,
        "generation_used_gold_fields": False,
        "production_index_path_used": False,
        "candidate_index_path_used": False,
        "promotion_evidence": False,
        "baseline_overwrite": False,
    }


def enrich_build_json(
    build_json_path: Path,
    *,
    readiness_report: Mapping[str, Any],
    manifest_rows: Sequence[Mapping[str, Any]],
    search_unit_manifest_file: str,
) -> None:
    payload = read_json(build_json_path)
    track_counts = Counter(clean(row.get("track")) for row in manifest_rows)
    payload.update(
        {
            "official_denominator_source_bound": True,
            "non_production_only": True,
            "official_denominator_rows": len(manifest_rows),
            "official_rows_by_track": dict(sorted(track_counts.items())),
            "expected_official_rows_by_track": dict(sorted(EXPECTED_OFFICIAL_ROWS_BY_TRACK.items())),
            "source_bound_locator_schema_covered": True,
            "search_unit_manifest_file": search_unit_manifest_file,
            "ingest_manifest_file": INGEST_MANIFEST_FILE,
            "readiness_schema_version": readiness_report.get("schema_version"),
            "candidate_artifacts_as_generation_source": False,
            "generation_used_expected_answer": False,
            "generation_used_supporting_evidence": False,
            "generation_used_gold_fields": False,
            "production_index_path_used": False,
            "candidate_index_path_used": False,
            "promotion_evidence": False,
            "baseline_overwrite": False,
        }
    )
    write_json(build_json_path, payload)


def source_locator_for_manifest(*, track: str, locator: Mapping[str, Any]) -> dict[str, Any]:
    fields = list(REQUIRED_FIELDS_BY_TRACK[track])
    extra_fields = {
        "text_namu_v2_1": ("source_corpus_path", "source_content_sha256"),
        "xlsx_business_structured": ("source_file_path", "source_content_sha256"),
        "pdf_business_ocr_mm": (
            "source_pdf_filename",
            "source_file_id",
            "source_content_sha256",
            "pdf_source_text_locator",
        ),
    }[track]
    out = {field: locator.get(field) for field in [*fields, *extra_fields] if locator.get(field) is not None}
    if track == "text_namu_v2_1":
        text_locator = mapping(locator.get("text_locator"))
        if text_locator.get("source_corpus_path"):
            out["source_corpus_path"] = text_locator.get("source_corpus_path")
    return out


def canonical_citation_payload(*, track: str, locator: Mapping[str, Any]) -> dict[str, Any]:
    payload = {
        "searchUnitId": clean(locator.get("search_unit_id")),
        "search_unit_id": clean(locator.get("search_unit_id")),
        "document_version_id": clean(locator.get("document_version_id")),
        "track": track,
        "source_bound_official_denominator": True,
    }
    if track == "text_namu_v2_1":
        payload.update(
            {
                "document_id": clean(locator.get("document_id")),
                "text_locator": locator.get("text_locator"),
            }
        )
    elif track == "xlsx_business_structured":
        for field in (
            "workbook",
            "sheet",
            "range",
            "cell",
            "row_label",
            "target_column",
            "normalized_value",
        ):
            payload[field] = locator.get(field)
    elif track == "pdf_business_ocr_mm":
        for field in (
            "source_pdf_path",
            "page",
            "physical_page_index",
            "bbox",
            "region_type",
            "row_label",
            "target_column",
        ):
            payload[field] = locator.get(field)
    return payload


def source_identity_for_locator(*, track: str, locator: Mapping[str, Any]) -> str:
    document_version_id = clean(locator.get("document_version_id"))
    if track == "text_namu_v2_1":
        return f"{document_version_id}:{clean(locator.get('document_id'))}"
    if track == "xlsx_business_structured":
        return (
            f"{document_version_id}:{clean(locator.get('workbook'))}:"
            f"{clean(locator.get('sheet'))}:{clean(locator.get('range'))}:{clean(locator.get('cell'))}"
        )
    if track == "pdf_business_ocr_mm":
        return (
            f"{document_version_id}:{clean(locator.get('source_pdf_path'))}:"
            f"{clean(locator.get('page'))}:{json.dumps(locator.get('bbox'), sort_keys=True)}"
        )
    return document_version_id


def source_bound_manifest_schema_missing(rows: Sequence[Mapping[str, Any]]) -> dict[str, list[str]]:
    missing_by_query_id: dict[str, list[str]] = {}
    for row in rows:
        track = clean(row.get("track"))
        locator = mapping(row.get("locator"))
        if track not in REQUIRED_FIELDS_BY_TRACK:
            missing_by_query_id[clean(row.get("query_id"))] = ["track"]
            continue
        missing = [
            field
            for field in REQUIRED_FIELDS_BY_TRACK[track]
            if not has_required_value(locator.get(field), field=field)
        ]
        if missing:
            missing_by_query_id[clean(row.get("query_id"))] = missing
    return missing_by_query_id


def append_status_event(path: Path, report: Mapping[str, Any]) -> None:
    event = {
        "event_type": "official_answer_citation_agentic_loop_source_bound_index_preparation",
        "run_id": "official_answer_citation_source_bound_index_build_readiness_v1",
        "generated_at": report.get("generated_at") or utc_timestamp(),
        "status": report.get("status"),
        "source_bound_official_denominator_index_design": compact_index_design(report),
        "search_unit_citation_payload_wired": True,
        "xlsx_source_bound_adapter_opt_in_wired": True,
        "pdf_source_bound_adapter_opt_in_wired": True,
        "guardrails": {
            "gold_mutation": False,
            "denominator_mutation": False,
            "human_label_mutation": False,
            "production_mutation": False,
            "candidate_artifacts_as_generation_source": False,
            "generation_used_expected_answer": False,
            "generation_used_supporting_evidence": False,
            "generation_used_gold_fields": False,
            "promotion_evidence": False,
        },
    }
    path.parent.mkdir(parents=True, exist_ok=True)
    with path.open("a", encoding="utf-8") as handle:
        handle.write(json.dumps(event, ensure_ascii=False, sort_keys=True) + "\n")


def compact_index_design(report: Mapping[str, Any]) -> dict[str, Any]:
    locators = mapping(report.get("source_bound_locators_by_query_id"))
    required_complete = Counter()
    source_identity_resolved = Counter()
    for locator in locators.values():
        row = mapping(locator)
        track = clean(row.get("track")) or infer_track_from_locator(row)
        if not track:
            continue
        if not [
            field
            for field in REQUIRED_FIELDS_BY_TRACK.get(track, ())
            if not has_required_value(row.get(field), field=field)
        ]:
            required_complete[track] += 1
        if source_identity_resolved_for_track(track=track, locator=row):
            source_identity_resolved[track] += 1
    return {
        "entrypoint_implemented": True,
        "status": report.get("status"),
        "blocker_category": report.get("blocker_category"),
        "target_index_path": report.get("target_index_path"),
        "index_version": report.get("index_version"),
        "non_production_only": True,
        "official_denominator_rows": report.get("official_denominator_rows"),
        "official_rows_by_track": report.get("official_rows_by_track"),
        "blocked_query_count": len(report.get("blocked_query_ids") or []),
        "blocked_query_ids": report.get("blocked_query_ids") or [],
        "required_field_complete_counts": {
            track: int(required_complete.get(track, 0))
            for track in sorted(EXPECTED_OFFICIAL_ROWS_BY_TRACK)
        },
        "source_identity_resolved_counts": {
            track: int(source_identity_resolved.get(track, 0))
            for track in sorted(EXPECTED_OFFICIAL_ROWS_BY_TRACK)
        },
        "missing_fields_by_query_id": report.get("missing_fields_by_query_id") or {},
        "missing_source_files_by_query_id": report.get("missing_source_files_by_query_id") or {},
        "build_ready": bool(report.get("build_ready")),
        "target_index_built": bool(report.get("target_index_built")),
        "load_check_passed": bool(report.get("load_check_passed")),
        "rerun_allowed": bool(report.get("rerun_allowed")),
        "production_index_path_used": False,
        "candidate_index_path_used": False,
        "candidate_artifacts_as_generation_source": False,
        "source_bound_artifact_contract_ok": bool(
            mapping(report.get("index_load_check")).get("source_bound_provenance_passed")
        ),
    }


def source_identity_resolved_for_track(*, track: str, locator: Mapping[str, Any]) -> bool:
    if not has_required_value(locator.get("search_unit_id"), field="search_unit_id"):
        return False
    if not has_required_value(locator.get("document_version_id"), field="document_version_id"):
        return False
    if track == "text_namu_v2_1":
        return has_required_value(locator.get("document_id"), field="document_id")
    if track == "xlsx_business_structured":
        return has_required_value(locator.get("source_file_path"), field="source_file_path")
    if track == "pdf_business_ocr_mm":
        return has_required_value(locator.get("source_pdf_path"), field="source_pdf_path")
    return False


def infer_track_from_locator(locator: Mapping[str, Any]) -> str:
    if locator.get("source_pdf_path"):
        return "pdf_business_ocr_mm"
    if locator.get("workbook"):
        return "xlsx_business_structured"
    if locator.get("text_locator"):
        return "text_namu_v2_1"
    return ""


def runtime_worker_settings() -> Any:
    ensure_ai_worker_on_path()
    from app.core.config import WorkerSettings

    return WorkerSettings()


def runtime_embedder(settings: Any) -> Any:
    ensure_ai_worker_on_path()
    from app.capabilities.rag.embeddings import SentenceTransformerEmbedder, resolve_max_seq_length

    max_seq_length = resolve_max_seq_length(getattr(settings, "rag_embedding_max_seq_length", None))
    return SentenceTransformerEmbedder(
        model_name=settings.rag_embedding_model,
        query_prefix=settings.rag_embedding_prefix_query,
        passage_prefix=settings.rag_embedding_prefix_passage,
        max_seq_length=max_seq_length,
        batch_size=int(settings.rag_embedding_batch_size),
        cuda_alloc_conf=settings.rag_embedding_cuda_alloc_conf or None,
    )


def ensure_ai_worker_on_path() -> None:
    ai_root = str(AI_WORKER_ROOT)
    if ai_root not in sys.path:
        sys.path.insert(0, ai_root)


def normalize_locator(track: str, locator: Mapping[str, Any]) -> dict[str, Any]:
    if track == "text_namu_v2_1":
        return {
            "document_id": first_present(locator, "document_id", "doc_id"),
            "document_version_id": first_present(locator, "document_version_id", "documentVersionId"),
            "search_unit_id": first_present(locator, "search_unit_id", "searchUnitId"),
            "text_locator": first_present(locator, "text_locator", "textLocator"),
        }
    if track == "xlsx_business_structured":
        return {
            "workbook": first_present(locator, "workbook", "source_workbook", "file"),
            "sheet": first_present(locator, "sheet", "sheet_name", "sheetName"),
            "range": first_present(locator, "range", "cellRange", "cell_range"),
            "cell": first_present(locator, "cell") or first_list_value(locator.get("matched_cells")),
            "row_label": first_present(locator, "row_label", "rowLabel"),
            "target_column": first_present(locator, "target_column", "targetColumn"),
            "normalized_value": first_present(locator, "normalized_value", "normalizedValue"),
            "search_unit_id": first_present(locator, "search_unit_id", "searchUnitId"),
            "document_version_id": first_present(locator, "document_version_id", "documentVersionId"),
        }
    if track == "pdf_business_ocr_mm":
        return {
            "source_pdf_path": first_present(locator, "source_pdf_path", "sourcePdfPath"),
            "page": first_present(locator, "page", "page_no", "pageNo"),
            "physical_page_index": first_present(
                locator, "physical_page_index", "physicalPageIndex"
            ),
            "bbox": first_present(locator, "bbox", "boundingBox"),
            "region_type": first_present(locator, "region_type", "regionType"),
            "row_label": first_present(locator, "row_label", "rowLabel"),
            "target_column": first_present(locator, "target_column", "targetColumn"),
            "search_unit_id": first_present(locator, "search_unit_id", "searchUnitId"),
            "document_version_id": first_present(locator, "document_version_id", "documentVersionId"),
        }
    return dict(locator)


class SourceBoundResolver:
    def __init__(self, source_roots: Sequence[Path]) -> None:
        self.source_roots = list(source_roots)
        self._sha256_cache: dict[Path, str] = {}
        self._text_chunks: dict[str, dict[str, Any]] | None = None
        self._pdf_locator_rows: list[dict[str, Any]] | None = None
        self._candidates_cache: dict[tuple[str, tuple[str, ...]], list[Path]] = {}

    def resolve(
        self,
        *,
        track: str,
        normalized: Mapping[str, Any],
        raw_locator: Mapping[str, Any],
    ) -> dict[str, Any]:
        if track == "text_namu_v2_1":
            return self.resolve_text(normalized=normalized, raw_locator=raw_locator)
        if track == "xlsx_business_structured":
            return self.resolve_xlsx(normalized=normalized, raw_locator=raw_locator)
        if track == "pdf_business_ocr_mm":
            return self.resolve_pdf(normalized=normalized, raw_locator=raw_locator)
        return dict(normalized)

    def resolve_text(
        self,
        *,
        normalized: Mapping[str, Any],
        raw_locator: Mapping[str, Any],
    ) -> dict[str, Any]:
        out = dict(normalized)
        chunk_id = clean(out.get("search_unit_id")) or clean(first_list_value(raw_locator.get("cited_chunk_ids")))
        if not chunk_id:
            return out
        chunk_record = self.text_chunks().get(chunk_id)
        if not chunk_record:
            return out
        chunk = mapping(chunk_record.get("chunk"))
        source_path = Path(clean(chunk_record.get("source_path")))
        document_id = clean(chunk.get("doc_id"))
        chunk_text = clean(chunk.get("chunk_text") or chunk.get("embedding_text"))
        text_locator = {
            "source_corpus_path": repo_relative(source_path),
            "line_number": chunk_record.get("line_number"),
            "chunk_id": chunk_id,
            "section_id": clean(chunk.get("section_id")),
            "section_path": chunk.get("section_path") or [],
            "section_type": clean(chunk.get("section_type")),
            "title": clean(chunk.get("title") or chunk.get("display_title") or chunk.get("retrieval_title")),
        }
        out["document_id"] = out.get("document_id") or document_id
        out["document_version_id"] = out.get("document_version_id") or document_version_id_for_file(
            source_path,
            sha256_cache=self._sha256_cache,
        )
        out["text_locator"] = out.get("text_locator") or text_locator
        out["search_unit_id"] = deterministic_search_unit_id(
            track="text_namu_v2_1",
            source_identity=f"{out.get('document_version_id')}:{document_id}",
            locator=text_locator,
            normalized_value=sha256_text(chunk_text),
        )
        out["source_text"] = chunk_text
        out["source_content_sha256"] = sha256_text(chunk_text)
        return out

    def resolve_xlsx(
        self,
        *,
        normalized: Mapping[str, Any],
        raw_locator: Mapping[str, Any],
    ) -> dict[str, Any]:
        out = dict(normalized)
        workbook_name = clean(out.get("workbook"))
        sheet_name = clean(out.get("sheet"))
        range_name = clean(out.get("range"))
        cell_name = clean(out.get("cell"))
        if not workbook_name or not sheet_name or not cell_name:
            return out
        for candidate in self.find_candidates(workbook_name, suffixes=(".xlsx", ".xlsm")):
            resolved = self.resolve_xlsx_from_workbook(
                candidate,
                workbook_name=workbook_name,
                sheet_name=sheet_name,
                range_name=range_name,
                cell_name=cell_name,
                raw_locator=raw_locator,
            )
            if resolved:
                out.update(resolved)
                break
        return out

    def resolve_xlsx_from_workbook(
        self,
        workbook_path: Path,
        *,
        workbook_name: str,
        sheet_name: str,
        range_name: str,
        cell_name: str,
        raw_locator: Mapping[str, Any],
    ) -> dict[str, Any] | None:
        try:
            workbook = load_workbook(workbook_path, read_only=True, data_only=True)
        except Exception:
            return None
        try:
            if sheet_name not in workbook.sheetnames:
                return None
            sheet = workbook[sheet_name]
            row_index, column_index = coordinate_to_tuple(cell_name)
            target_cell_value = sheet.cell(row=row_index, column=column_index).value
            if target_cell_value in (None, ""):
                return None
            header_value = sheet.cell(row=1, column=column_index).value
            row_values = [
                sheet.cell(row=row_index, column=index).value
                for index in range(1, max(column_index, sheet.max_column) + 1)
            ]
            headers = [
                sheet.cell(row=1, column=index).value
                for index in range(1, max(column_index, sheet.max_column) + 1)
            ]
        finally:
            workbook.close()

        document_version_id = document_version_id_for_file(
            workbook_path,
            sha256_cache=self._sha256_cache,
        )
        locator = {
            "workbook": workbook_name,
            "source_file_path": repo_relative(workbook_path),
            "sheet": sheet_name,
            "range": range_name,
            "cell": cell_name,
        }
        normalized_value = normalize_cell_value(target_cell_value)
        row_label = source_row_label(
            headers=headers,
            row_values=row_values,
            target_column_index=column_index,
        )
        target_column = clean(header_value) or get_column_letter(column_index)
        return {
            **locator,
            "row_label": row_label,
            "target_column": target_column,
            "normalized_value": normalized_value,
            "document_version_id": document_version_id,
            "search_unit_id": deterministic_search_unit_id(
                track="xlsx_business_structured",
                source_identity=document_version_id,
                locator=locator,
                normalized_value=normalized_value,
            ),
            "source_content_sha256": self.sha256_for(workbook_path),
        }

    def resolve_pdf(
        self,
        *,
        normalized: Mapping[str, Any],
        raw_locator: Mapping[str, Any],
    ) -> dict[str, Any]:
        out = dict(normalized)
        locator_row = self.find_pdf_locator_row(normalized=out, raw_locator=raw_locator)
        if locator_row:
            source_locator = mapping(locator_row.get("citation_locator"))
            for field in (
                "page",
                "physical_page_index",
                "bbox",
                "region_type",
                "search_unit_id",
                "document_version_id",
            ):
                out[field] = out.get(field) or locator_row.get(field) or source_locator.get(field)
            for field in ("row_label", "target_column"):
                out[field] = out.get(field) or locator_row.get(field) or source_locator.get(field)
            source_file_id = clean(locator_row.get("source_file_id")) or clean(source_locator.get("file"))
            if source_file_id:
                out["source_file_id"] = source_file_id
            source_pdf_name = clean(
                first_present(source_locator, "source_pdf_path", "sourcePdfPath")
                or source_locator.get("file")
                or locator_row.get("source_pdf_path")
                or locator_row.get("source_pdf_filename")
            )
            if source_pdf_name.lower().endswith(".pdf"):
                out["source_pdf_filename"] = source_pdf_name
                source_candidates = self.find_candidates(source_pdf_name, suffixes=(".pdf",))
                if source_candidates:
                    source_path = source_candidates[0]
                    out["source_pdf_path"] = repo_relative(source_path)
                    out["source_content_sha256"] = self.sha256_for(source_path)
                    self.enrich_pdf_source_text_locator(out, source_path)
            source_manifest_path = clean(locator_row.get("_source_locator_manifest_path"))
            if source_manifest_path:
                out["source_locator_manifest_path"] = repo_relative(Path(source_manifest_path))
                out["source_locator_manifest_kind"] = "pdf_answer_citation_diagnostic_review_input"

        candidate_name = clean(out.get("source_pdf_path")) or clean(raw_locator.get("source_pdf_path"))
        if not candidate_name:
            return out
        candidates = self.find_candidates(candidate_name, suffixes=(".pdf",))
        if not candidates:
            return out
        source_path = candidates[0]
        out["source_pdf_path"] = repo_relative(source_path)
        out["document_version_id"] = out.get("document_version_id") or document_version_id_for_file(
            source_path,
            sha256_cache=self._sha256_cache,
        )
        self.enrich_pdf_source_text_locator(out, source_path)
        return out

    def enrich_pdf_source_text_locator(self, locator: dict[str, Any], source_path: Path) -> None:
        if has_required_value(locator.get("row_label"), field="row_label") and has_required_value(
            locator.get("target_column"), field="target_column"
        ):
            return
        derived = derive_pdf_text_locator_from_source(
            source_path=source_path,
            page_index=locator.get("physical_page_index"),
            page_number=locator.get("page"),
            bbox=locator.get("bbox"),
            region_type=clean(locator.get("region_type")),
        )
        if not derived:
            return
        if not has_required_value(locator.get("row_label"), field="row_label"):
            locator["row_label"] = derived.get("row_label")
        if not has_required_value(locator.get("target_column"), field="target_column"):
            locator["target_column"] = derived.get("target_column")
        locator["pdf_source_text_locator"] = {
            **mapping(derived.get("pdf_source_text_locator")),
            "source_pdf_path": repo_relative(source_path),
        }

    def inventory_for(
        self,
        *,
        track: str,
        normalized: Mapping[str, Any],
        raw_locator: Mapping[str, Any],
    ) -> list[dict[str, Any]]:
        if track == "text_namu_v2_1":
            chunk_id = clean(first_list_value(raw_locator.get("cited_chunk_ids"))) or clean(
                normalized.get("search_unit_id")
            )
            chunk_record = self.text_chunks().get(chunk_id) if chunk_id else None
            if not chunk_id:
                return []
            if not chunk_record:
                return [{"reference": chunk_id, "exists": False, "kind": "text_chunk"}]
            source_path = Path(clean(chunk_record.get("source_path")))
            return [
                {
                    "reference": chunk_id,
                    "exists": True,
                    "kind": "text_chunk",
                    "source_path": repo_relative(source_path),
                    "line_number": chunk_record.get("line_number"),
                    "document_version_id": document_version_id_for_file(
                        source_path,
                        sha256_cache=self._sha256_cache,
                    ),
                }
            ]
        if track == "xlsx_business_structured":
            name = clean(normalized.get("workbook")) or clean(raw_locator.get("file"))
            return self.inventory_candidates(name, suffixes=(".xlsx", ".xlsm"))
        if track == "pdf_business_ocr_mm":
            name = clean(normalized.get("source_pdf_path")) or clean(raw_locator.get("source_pdf_path")) or clean(
                raw_locator.get("file")
            )
            locator_row = self.find_pdf_locator_row(normalized=normalized, raw_locator=raw_locator)
            if locator_row:
                source_locator = mapping(locator_row.get("citation_locator"))
                source_manifest_path = clean(locator_row.get("_source_locator_manifest_path"))
                return [
                    {
                        "reference": clean(locator_row.get("source_file_id"))
                        or clean(source_locator.get("file"))
                        or name,
                        "exists": True,
                        "kind": "pdf_locator_manifest",
                        "source_path": repo_relative(Path(source_manifest_path))
                        if source_manifest_path
                        else "",
                        "document_version_id": clean(locator_row.get("document_version_id"))
                        or clean(source_locator.get("document_version_id")),
                        "search_unit_id": clean(locator_row.get("search_unit_id"))
                        or clean(source_locator.get("search_unit_id")),
                        "source_pdf_filename": clean(source_locator.get("file")),
                        "source_pdf_path_resolved": bool(normalized.get("source_pdf_path")),
                    }
                ]
            if clean(normalized.get("source_pdf_path")):
                return self.inventory_candidates(name, suffixes=(".pdf",))
            return [{"reference": name, "exists": False, "kind": "pdf_source"}] if name else []
        return []

    def missing_source_references(
        self,
        *,
        track: str,
        normalized: Mapping[str, Any],
        raw_locator: Mapping[str, Any],
    ) -> list[str]:
        if track == "text_namu_v2_1":
            if has_required_value(normalized.get("text_locator"), field="text_locator"):
                return []
            chunk_id = clean(first_list_value(raw_locator.get("cited_chunk_ids")))
            return [f"text_chunk:{chunk_id}"] if chunk_id else []
        if track == "xlsx_business_structured":
            if has_required_value(normalized.get("source_file_path"), field="source_file_path"):
                return []
            name = clean(normalized.get("workbook")) or clean(raw_locator.get("file"))
            return [name] if name else []
        if track == "pdf_business_ocr_mm":
            if has_required_value(normalized.get("source_pdf_path"), field="source_pdf_path"):
                return []
            name = clean(raw_locator.get("source_pdf_path")) or clean(raw_locator.get("file"))
            return [name] if name else []
        return []

    def inventory_candidates(self, name: str, *, suffixes: tuple[str, ...]) -> list[dict[str, Any]]:
        if not name:
            return []
        candidates = self.find_candidates(name, suffixes=suffixes)
        if not candidates:
            return [{"reference": name, "exists": False, "kind": "source_file"}]
        out = []
        for index, path in enumerate(candidates[:10]):
            out.append(
                {
                    "reference": name,
                    "exists": True,
                    "kind": "source_file",
                    "selected": index == 0,
                    "source_path": repo_relative(path),
                    "sha256": self.sha256_for(path),
                }
            )
        return out

    def find_candidates(self, name: str, *, suffixes: tuple[str, ...]) -> list[Path]:
        key = (name, suffixes)
        if key in self._candidates_cache:
            return self._candidates_cache[key]
        wanted_path = Path(name)
        wanted_name = wanted_path.name
        wanted_lower = wanted_name.lower()
        wanted_tokens = source_match_tokens(wanted_name)
        candidates: list[Path] = []
        seen: set[Path] = set()
        for root in self.source_roots:
            resolved_root = resolve_repo_path(root)
            if not resolved_root.exists():
                continue
            direct = resolved_root / name
            if direct.is_file() and direct.suffix.lower() in suffixes:
                seen.add(direct.resolve())
                candidates.append(direct)
            for path in resolved_root.rglob("*"):
                if not path.is_file() or path.suffix.lower() not in suffixes:
                    continue
                path_resolved = path.resolve()
                if path_resolved in seen:
                    continue
                path_name_lower = path.name.lower()
                if path_name_lower == wanted_lower or (
                    wanted_tokens and wanted_tokens.issubset(source_match_tokens(path.name))
                ):
                    seen.add(path_resolved)
                    candidates.append(path)
        candidates = sorted(candidates, key=lambda item: repo_relative(item))
        self._candidates_cache[key] = candidates
        return candidates

    def text_chunks(self) -> dict[str, dict[str, Any]]:
        if self._text_chunks is not None:
            return self._text_chunks
        chunks: dict[str, dict[str, Any]] = {}
        for root in self.source_roots:
            resolved_root = resolve_repo_path(root)
            if not resolved_root.exists():
                continue
            for path in resolved_root.rglob("rag_chunks.jsonl"):
                try:
                    with path.open("r", encoding="utf-8") as handle:
                        for line_number, line in enumerate(handle, start=1):
                            if not line.strip():
                                continue
                            try:
                                chunk = json.loads(line)
                            except json.JSONDecodeError:
                                continue
                            chunk_id = clean(mapping(chunk).get("chunk_id"))
                            if chunk_id and chunk_id not in chunks:
                                chunks[chunk_id] = {
                                    "source_path": str(path),
                                    "line_number": line_number,
                                    "chunk": chunk,
                                }
                except OSError:
                    continue
        self._text_chunks = chunks
        return chunks

    def find_pdf_locator_row(
        self,
        *,
        normalized: Mapping[str, Any],
        raw_locator: Mapping[str, Any],
    ) -> dict[str, Any] | None:
        search_unit_id = clean(normalized.get("search_unit_id")) or clean(raw_locator.get("search_unit_id"))
        source_file_id = clean(raw_locator.get("file")) or clean(raw_locator.get("source_file_id"))
        page = clean(normalized.get("page")) or clean(raw_locator.get("page"))
        bbox = normalized.get("bbox") or raw_locator.get("bbox")
        for row in self.pdf_locator_rows():
            source_locator = mapping(row.get("citation_locator"))
            row_search_unit_id = clean(row.get("search_unit_id")) or clean(source_locator.get("search_unit_id"))
            if search_unit_id and row_search_unit_id == search_unit_id:
                return row
            row_source_file_id = clean(row.get("source_file_id")) or clean(source_locator.get("file"))
            row_page = clean(row.get("page")) or clean(source_locator.get("page"))
            row_bbox = row.get("bbox") or source_locator.get("bbox")
            if (
                source_file_id
                and row_source_file_id == source_file_id
                and page
                and row_page == page
                and bbox
                and row_bbox == bbox
            ):
                return row
        return None

    def pdf_locator_rows(self) -> list[dict[str, Any]]:
        if self._pdf_locator_rows is not None:
            return self._pdf_locator_rows
        rows: list[dict[str, Any]] = []
        seen: set[tuple[str, int]] = set()
        for root in self.source_roots:
            resolved_root = resolve_repo_path(root)
            if not resolved_root.exists():
                continue
            for manifest_name in PDF_SOURCE_LOCATOR_MANIFEST_NAMES:
                for path in resolved_root.rglob(manifest_name):
                    try:
                        with path.open("r", encoding="utf-8") as handle:
                            for line_number, line in enumerate(handle, start=1):
                                if not line.strip():
                                    continue
                                try:
                                    row = json.loads(line)
                                except json.JSONDecodeError:
                                    continue
                                if not is_safe_pdf_source_locator_row(row):
                                    continue
                                key = (str(path.resolve()), line_number)
                                if key in seen:
                                    continue
                                seen.add(key)
                                row["_source_locator_manifest_path"] = str(path)
                                row["_source_locator_manifest_line"] = line_number
                                rows.append(row)
                    except OSError:
                        continue
        self._pdf_locator_rows = rows
        return rows

    def sha256_for(self, path: Path) -> str:
        resolved = path.resolve()
        if resolved not in self._sha256_cache:
            self._sha256_cache[resolved] = sha256_file(resolved)
        return self._sha256_cache[resolved]


def missing_source_files(
    normalized: Mapping[str, Any],
    *,
    track: str,
    source_roots: Sequence[Path],
) -> list[str]:
    if track == "text_namu_v2_1" and normalized.get("text_locator"):
        return []
    if track == "xlsx_business_structured" and normalized.get("source_file_path"):
        return []
    if track == "pdf_business_ocr_mm" and normalized.get("source_pdf_path"):
        return []
    if track == "xlsx_business_structured":
        names = [clean(normalized.get("workbook"))]
    elif track == "pdf_business_ocr_mm":
        names = [clean(normalized.get("source_pdf_path"))]
    else:
        names = []
    missing: list[str] = []
    for name in names:
        if not name:
            continue
        if not source_file_exists(name, source_roots):
            missing.append(name)
    return missing


def derive_pdf_text_locator_from_source(
    *,
    source_path: Path,
    page_index: Any,
    page_number: Any,
    bbox: Any,
    region_type: str,
) -> dict[str, Any]:
    try:
        import fitz  # type: ignore[import-untyped]
    except Exception:
        return {}

    page_index_int = int_or_none(page_index)
    if page_index_int is None:
        page_number_int = int_or_none(page_number)
        if page_number_int is None:
            return {}
        page_index_int = page_number_int - 1
    rect = pdf_rect_or_none(bbox)
    if rect is None:
        return {}
    try:
        doc = fitz.open(source_path)
    except Exception:
        return {}
    try:
        if page_index_int < 0 or page_index_int >= len(doc):
            return {}
        page = doc[page_index_int]
        lines = extract_pdf_text_lines(page)
    finally:
        doc.close()
    if not lines:
        return {}

    selected = pdf_lines_intersecting(lines, rect)
    if not selected:
        selected = pdf_lines_intersecting(lines, rect + (-20, -15, 220, 45))
    if not selected:
        return {}

    if region_type == "table_body":
        row_label, target_column, used_lines = derive_table_axis_from_pdf_lines(
            lines=lines,
            selected_lines=selected,
        )
    else:
        row_label = selected[0]["text"]
        target_column = f"{region_type or 'pdf_region'}_text"
        used_lines = selected[:1]

    if not row_label or not target_column:
        return {}
    return {
        "row_label": row_label,
        "target_column": target_column,
        "pdf_source_text_locator": {
            "method": "pymupdf_source_pdf_text",
            "page": int_or_none(page_number),
            "physical_page_index": page_index_int,
            "bbox": [round(float(value), 4) for value in rect],
            "region_type": region_type,
            "source_lines": [
                {"text": line["text"], "bbox": [round(float(value), 4) for value in line["bbox"]]}
                for line in used_lines
            ],
        },
    }


def extract_pdf_text_lines(page: Any) -> list[dict[str, Any]]:
    raw = page.get_text("dict")
    lines: list[dict[str, Any]] = []
    for block in raw.get("blocks") or []:
        if block.get("type") not in (None, 0):
            continue
        for line in block.get("lines") or []:
            text = normalize_pdf_text(" ".join(clean(span.get("text")) for span in line.get("spans") or []))
            if not text:
                continue
            bbox = line.get("bbox")
            if not bbox or len(bbox) != 4:
                continue
            lines.append({"text": text, "bbox": [float(value) for value in bbox]})
    return sorted(lines, key=lambda item: (item["bbox"][1], item["bbox"][0]))


def pdf_lines_intersecting(lines: Sequence[Mapping[str, Any]], rect: Any) -> list[dict[str, Any]]:
    try:
        import fitz  # type: ignore[import-untyped]
    except Exception:
        return []
    selected = []
    for line in lines:
        line_rect = fitz.Rect(line["bbox"])
        if line_rect.intersects(rect) or (
            vertical_overlap(line_rect, rect) and horizontal_overlap(line_rect, rect)
        ):
            selected.append(dict(line))
    return selected


def derive_table_axis_from_pdf_lines(
    *,
    lines: Sequence[Mapping[str, Any]],
    selected_lines: Sequence[Mapping[str, Any]],
) -> tuple[str, str, list[dict[str, Any]]]:
    row_line = first_table_label_line(selected_lines) or dict(selected_lines[0])
    try:
        row_index = next(index for index, line in enumerate(lines) if line["bbox"] == row_line["bbox"])
    except StopIteration:
        row_index = 0
    header_lines: list[dict[str, Any]] = []
    row_bottom = float(row_line["bbox"][3])
    for line in lines[row_index + 1 :]:
        if float(line["bbox"][1]) - row_bottom > 80 and header_lines:
            break
        text = clean(line.get("text"))
        if not text or text == row_line["text"] or is_pdf_unit_line(text):
            continue
        if is_pdf_table_data_line(text):
            if header_lines:
                break
            continue
        header_lines.append(dict(line))
        if len(header_lines) >= 4:
            break
    target_column = " | ".join(line["text"] for line in header_lines)
    if not target_column:
        target_column = "table_body"
    body_lines = collect_pdf_table_body_lines(
        lines=lines,
        start_index=row_index + 1,
        max_rows=5,
    )
    return row_line["text"], target_column, unique_pdf_lines([dict(row_line), *header_lines, *body_lines])


def collect_pdf_table_body_lines(
    *,
    lines: Sequence[Mapping[str, Any]],
    start_index: int,
    max_rows: int,
) -> list[dict[str, Any]]:
    body_lines: list[dict[str, Any]] = []
    anchor_bboxes: list[Sequence[float]] = []
    for index, line in enumerate(lines[start_index:], start=start_index):
        text = clean(line.get("text"))
        if not text or is_pdf_unit_line(text) or not is_pdf_table_data_line(text):
            continue
        bbox = line.get("bbox")
        if not valid_bbox(bbox):
            continue
        if any(bbox_vertical_overlap(bbox, anchor) for anchor in anchor_bboxes):
            continue
        row_lines = [
            dict(candidate)
            for candidate in lines[index:]
            if valid_bbox(candidate.get("bbox"))
            and bbox_vertical_overlap(candidate.get("bbox"), bbox)
            and clean(candidate.get("text"))
            and not is_pdf_unit_line(clean(candidate.get("text")))
        ]
        if not row_lines:
            continue
        body_lines.extend(sorted(row_lines, key=lambda item: float(item["bbox"][0])))
        anchor_bboxes.append(bbox)
        if len(anchor_bboxes) >= max_rows:
            break
    return body_lines


def unique_pdf_lines(lines: Sequence[Mapping[str, Any]]) -> list[dict[str, Any]]:
    out: list[dict[str, Any]] = []
    seen: set[tuple[str, tuple[float, ...]]] = set()
    for line in lines:
        text = clean(line.get("text"))
        bbox = line.get("bbox")
        if not text or not valid_bbox(bbox):
            continue
        key = (text, tuple(round(float(value), 4) for value in bbox))
        if key in seen:
            continue
        seen.add(key)
        out.append(dict(line))
    return out


def first_table_label_line(lines: Sequence[Mapping[str, Any]]) -> dict[str, Any] | None:
    for line in lines:
        text = clean(line.get("text"))
        if text and not is_pdf_unit_line(text) and not is_pdf_table_data_line(text):
            return dict(line)
    return None


def pdf_rect_or_none(bbox: Any) -> Any:
    try:
        import fitz  # type: ignore[import-untyped]
    except Exception:
        return None
    if isinstance(bbox, (str, bytes)):
        return None
    try:
        values = list(bbox)
    except TypeError:
        return None
    if len(values) != 4:
        return None
    try:
        return fitz.Rect([float(value) for value in values])
    except Exception:
        return None


def normalize_pdf_text(text: str) -> str:
    return re.sub(r"\s+", " ", text).strip()


def vertical_overlap(left: Any, right: Any) -> bool:
    return min(left.y1, right.y1) >= max(left.y0, right.y0)


def horizontal_overlap(left: Any, right: Any) -> bool:
    return min(left.x1, right.x1) >= max(left.x0, right.x0)


def is_pdf_unit_line(text: str) -> bool:
    stripped = text.strip()
    return stripped.startswith("(") and stripped.endswith(")")


def is_pdf_table_data_line(text: str) -> bool:
    first = text.split()[0] if text.split() else ""
    return bool(
        re.match(
            r"^(?:\d{4}(?:\.\s*(?:\d{1,2}|[ⅠⅡⅢⅣ]+))?|[ⅠⅡⅢⅣ]+|\d{1,2}|\d[\d,]*(?:\.\d+)?)$",
            first,
        )
    )


def valid_bbox(value: Any) -> bool:
    return isinstance(value, Sequence) and not isinstance(value, (str, bytes)) and len(value) == 4


def bbox_vertical_overlap(left: Any, right: Any) -> bool:
    if not valid_bbox(left) or not valid_bbox(right):
        return False
    return min(float(left[3]), float(right[3])) >= max(float(left[1]), float(right[1]))


def int_or_none(value: Any) -> int | None:
    try:
        return int(value)
    except (TypeError, ValueError):
        return None


def source_file_exists(name: str, source_roots: Sequence[Path]) -> bool:
    normalized_name = Path(name).name
    for root in source_roots:
        resolved_root = resolve_repo_path(root)
        if not resolved_root.exists():
            continue
        direct = resolved_root / name
        if direct.exists():
            return True
        for candidate in resolved_root.rglob(normalized_name):
            if candidate.is_file():
                return True
    return False


def default_source_roots() -> list[Path]:
    roots = [
        AI_WORKER_ROOT / "eval" / "datasets",
        AI_WORKER_ROOT / "eval" / "corpora",
        REPO_ROOT / "datasets",
        REPO_ROOT / "local-storage",
        AI_WORKER_ROOT / "eval" / "artifacts",
    ]
    roots.extend(external_source_roots())
    return roots


def external_source_roots() -> list[Path]:
    roots: list[Path] = []
    drive_root = Path(REPO_ROOT.anchor) if REPO_ROOT.anchor else REPO_ROOT.parent
    for root_name in EXTERNAL_REPO_ROOT_NAMES:
        external_repo_root = drive_root / root_name / REPO_ROOT.name
        candidates = (
            external_repo_root / "source_collection_20260510",
            external_repo_root / "rag-ingestion" / "hard-cleanup-20260517" / "reports",
            external_repo_root
            / "2026-05-14-rag-ingestion-report-compaction"
            / "ai-worker"
            / "eval"
            / "reports"
            / "rag-ingestion",
            external_repo_root / "20260507_214525" / "local-storage",
        )
        roots.extend(candidate for candidate in candidates if candidate.exists())
    return roots


def is_safe_pdf_source_locator_row(row: Mapping[str, Any]) -> bool:
    if row.get("track") not in (None, "", "pdf_business_ocr_mm"):
        return False
    if row.get("official_metric_input") is not False:
        return False
    if row.get("promotion_evidence") is not False:
        return False
    if row.get("diagnostic_only") is not True:
        return False
    source_locator = mapping(row.get("citation_locator"))
    return bool(clean(row.get("search_unit_id")) or clean(source_locator.get("search_unit_id")))


def mapping(value: Any) -> dict[str, Any]:
    if isinstance(value, Mapping):
        return dict(value)
    if isinstance(value, str):
        try:
            parsed = json.loads(value)
        except json.JSONDecodeError:
            return {}
        return dict(parsed) if isinstance(parsed, Mapping) else {}
    return {}


def first_present(values: Mapping[str, Any], *keys: str) -> Any:
    for key in keys:
        value = values.get(key)
        if value not in (None, "", []):
            return value
    return None


def first_list_value(value: Any) -> Any:
    if isinstance(value, Sequence) and not isinstance(value, (str, bytes)) and value:
        return value[0]
    return None


def has_required_value(value: Any, *, field: str) -> bool:
    if field == "bbox":
        return isinstance(value, Sequence) and not isinstance(value, (str, bytes)) and len(value) == 4
    if isinstance(value, str):
        return bool(value.strip())
    if isinstance(value, Sequence) and not isinstance(value, (str, bytes)):
        return bool(value)
    return value is not None


def denominator_contract_errors(
    *,
    rows: Sequence[Mapping[str, Any]],
    rows_by_track: Counter,
) -> list[str]:
    errors: list[str] = []
    query_ids = [clean(row.get("query_id")) for row in rows]
    if len(rows) != 29:
        errors.append(f"official_denominator_rows_expected_29_actual_{len(rows)}")
    if len(set(query_ids)) != len(query_ids):
        errors.append("official_denominator_query_ids_not_unique")
    if any(not query_id for query_id in query_ids):
        errors.append("official_denominator_query_id_missing")
    actual_track_counts = dict(rows_by_track)
    if actual_track_counts != EXPECTED_OFFICIAL_ROWS_BY_TRACK:
        errors.append(
            "official_denominator_track_counts_expected_"
            + json.dumps(EXPECTED_OFFICIAL_ROWS_BY_TRACK, ensure_ascii=False, sort_keys=True)
            + "_actual_"
            + json.dumps(actual_track_counts, ensure_ascii=False, sort_keys=True)
        )
    unknown_tracks = sorted(set(actual_track_counts) - set(REQUIRED_FIELDS_BY_TRACK))
    if unknown_tracks:
        errors.append("official_denominator_unknown_tracks_" + ",".join(unknown_tracks))
    return errors


def deterministic_search_unit_id(
    *,
    track: str,
    source_identity: str,
    locator: Mapping[str, Any],
    normalized_value: Any,
) -> str:
    payload = {
        "track": track,
        "source_identity": source_identity,
        "locator": locator,
        "normalized_value": normalized_value,
    }
    encoded = json.dumps(payload, ensure_ascii=False, sort_keys=True, separators=(",", ":"))
    return "su_" + hashlib.sha256(encoded.encode("utf-8")).hexdigest()[:32]


def document_version_id_for_file(path: Path, *, sha256_cache: dict[Path, str] | None = None) -> str:
    resolved = path.resolve()
    if sha256_cache is not None:
        if resolved not in sha256_cache:
            sha256_cache[resolved] = sha256_file(resolved)
        digest = sha256_cache[resolved]
    else:
        digest = sha256_file(resolved)
    return "docv_sha256_" + digest[:16]


def sha256_file(path: Path) -> str:
    digest = hashlib.sha256()
    with path.open("rb") as handle:
        for chunk in iter(lambda: handle.read(1024 * 1024), b""):
            digest.update(chunk)
    return digest.hexdigest()


def sha256_text(value: str) -> str:
    return hashlib.sha256(value.encode("utf-8")).hexdigest()


def normalize_cell_value(value: Any) -> str:
    if value is None:
        return ""
    if isinstance(value, bool):
        return "true" if value else "false"
    if isinstance(value, int):
        return str(value)
    if isinstance(value, float):
        if value.is_integer():
            return str(int(value))
        return format(value, "g")
    if hasattr(value, "isoformat") and not isinstance(value, str):
        return value.isoformat()
    return str(value).strip()


def source_row_label(
    *,
    headers: Sequence[Any],
    row_values: Sequence[Any],
    target_column_index: int,
) -> str:
    parts: list[str] = []
    for index, value in enumerate(row_values, start=1):
        if index == target_column_index:
            continue
        normalized_value = normalize_cell_value(value)
        if not normalized_value:
            continue
        header = clean(headers[index - 1] if index - 1 < len(headers) else None) or get_column_letter(index)
        parts.append(f"{header}={normalized_value}")
        if len(parts) >= 3:
            break
    return " | ".join(parts)


def source_match_tokens(name: str) -> set[str]:
    stem = Path(name).stem.lower()
    tokens = set(re.findall(r"\d{8}|\d{4}[._]\d{1,2}|\d{4}|\d+", stem))
    return {token.rstrip("._") for token in tokens if len(token.rstrip("._")) >= 4}


def read_json(path: Path) -> dict[str, Any]:
    return json.loads(path.read_text(encoding="utf-8"))


def write_json(path: Path, payload: Mapping[str, Any]) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    path.write_text(
        json.dumps(payload, ensure_ascii=False, indent=2, sort_keys=True) + "\n",
        encoding="utf-8",
    )


def write_jsonl(path: Path, rows: Sequence[Mapping[str, Any]]) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    path.write_text(
        "".join(json.dumps(row, ensure_ascii=False, sort_keys=True) + "\n" for row in rows),
        encoding="utf-8",
    )


def read_jsonl(path: Path) -> list[dict[str, Any]]:
    rows: list[dict[str, Any]] = []
    with path.open("r", encoding="utf-8") as handle:
        for line in handle:
            if not line.strip():
                continue
            row = json.loads(line)
            if isinstance(row, Mapping):
                rows.append(dict(row))
    return rows


def digest_texts(texts: Sequence[str]) -> str:
    digest = hashlib.sha256()
    for text in texts:
        digest.update(text.encode("utf-8"))
        digest.update(b"\n")
    return digest.hexdigest()


def duplicate_values(values: Sequence[str] | Any) -> list[str]:
    counts = Counter(clean(value) for value in values if clean(value))
    return sorted(value for value, count in counts.items() if count > 1)


def resolve_repo_path(path: Path) -> Path:
    if path.is_absolute():
        return path
    parts = path.parts
    if parts and parts[0] == "eval":
        return AI_WORKER_ROOT / path
    return REPO_ROOT / path


def repo_relative(path: Path) -> str:
    resolved = path.resolve()
    try:
        return resolved.relative_to(REPO_ROOT.resolve()).as_posix()
    except ValueError:
        return resolved.as_posix()


def clean(value: Any) -> str:
    if value is None:
        return ""
    return str(value).strip()


def utc_timestamp() -> str:
    return datetime.now(timezone.utc).isoformat()


if __name__ == "__main__":  # pragma: no cover
    raise SystemExit(main())
