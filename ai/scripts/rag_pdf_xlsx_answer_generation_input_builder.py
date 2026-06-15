"""Build diagnostic PDF/XLSX answer-generation inputs.

This script is diagnostic-only. It consumes the row-level answer-shape plan and
review packs, then emits local-LLM input rows under eval/artifacts. It does not
run retrieval, tune retrieval, mutate SearchUnit, update gold CSVs, or create
promotion evidence.
"""

from __future__ import annotations

import argparse
import csv
import hashlib
import json
import os
import re
import shutil
import sys
from collections import Counter
from datetime import datetime, timezone
from pathlib import Path
from typing import Any, Iterable, Mapping


AI_WORKER_ROOT = Path(__file__).resolve().parents[1]
REPO_ROOT = AI_WORKER_ROOT.parent
REPORT_DIR = AI_WORKER_ROOT.parent / "reports" / "rag_eval" / "rag-ingestion"
REVIEW_DIR = AI_WORKER_ROOT / "eval" / "review" / "gold_set_review"
DATASET_DIR = AI_WORKER_ROOT / "eval" / "datasets"
EVAL_RUNS_DIR = AI_WORKER_ROOT / "eval" / "artifacts" / "eval_runs"

DEFAULT_CONTRACT = REPO_ROOT / "docs" / "eval" / "pdf_xlsx_answer_intent_prompt_contract.md"
DEFAULT_PLAN = REPORT_DIR / "rag_pdf_xlsx_answer_shape_diagnostic_plan.json"
DEFAULT_KEYWORD_REVIEW = REPORT_DIR / "rag_pdf_xlsx_keyword_echo_risk_review.csv"
DEFAULT_PDF_REVIEW = REVIEW_DIR / "pdf_gold_review_pack.csv"
DEFAULT_XLSX_REPO_REVIEW = REVIEW_DIR / "xlsx_gold_review_pack.csv"
DEFAULT_XLSX_RETRIEVAL_REPORT = (
    REPORT_DIR / "rag_retrieval_eval_xlsx_v3_positive_reviewed_vector_diagnostic_report.json"
)
DEFAULT_XLSX_FAILURE_BREAKDOWN = REPORT_DIR / "rag_xlsx_v3_after_cleanup_failure_breakdown.json"
DEFAULT_PDF_RETRIEVAL_REPORT = REPORT_DIR / "rag_retrieval_eval_pdf_vector_diagnostic_report.json"
DEFAULT_PDF_C6_REPORT = REPORT_DIR / "rag_pdf_vector_quality_breakdown.json"
DEFAULT_PDF_C7_REPORT = REPORT_DIR / "rag_pdf_gold_policy_review.json"

SCHEMA_VERSION = "rag_pdf_xlsx_answer_generation_inputs_v1"
RUN_PREFIX = "pdf_xlsx_answer_shape_local_llm"
DEFAULT_DB_DSN = os.environ.get(
    "RAG_DB_DSN",
    "host=localhost port=5433 dbname=aipipeline user=aipipeline password=aipipeline_pw",
)
ANSWER_SHAPE_POLICY_PENDING = "NOT_ANSWERABLE_OR_POLICY_PENDING"
XLSX_ALLOWED_EVIDENCE_CHUNK_TYPES = {
    "cell",
    "key_value",
    "row",
    "row_group",
    "table",
    "table_region",
}
XLSX_BROAD_EVIDENCE_CHUNK_TYPES = {"sheet_summary", "workbook_summary"}


def main(argv: list[str] | None = None) -> int:
    args = parse_args(argv)
    report = run_builder(
        contract=Path(args.contract),
        plan=Path(args.plan),
        keyword_review=Path(args.keyword_review),
        pdf_review=Path(args.pdf_review),
        xlsx_external_review=Path(args.xlsx_external_review) if args.xlsx_external_review else None,
        xlsx_repo_review=Path(args.xlsx_repo_review),
        xlsx_retrieval_report=Path(args.xlsx_retrieval_report),
        xlsx_failure_breakdown=Path(args.xlsx_failure_breakdown),
        pdf_retrieval_report=Path(args.pdf_retrieval_report),
        pdf_c6_report=Path(args.pdf_c6_report),
        pdf_c7_report=Path(args.pdf_c7_report),
        output_root=Path(args.output_root),
        run_id=args.run_id,
        run_prefix=args.run_prefix,
        max_context_chars=args.max_context_chars,
        db_dsn=args.db_dsn,
        xlsx_searchunit_content_join=not args.disable_xlsx_searchunit_content_join,
        enriched_inputs_alias=Path(args.enriched_inputs_alias).name if args.enriched_inputs_alias else "",
    )
    print_json(
        {
            "status": report["status"],
            "run_id": report["run_id"],
            "artifact_dir": report["artifact_dir"],
            "manifest": report["manifest_path"],
            "answer_generation_inputs": report["answer_generation_inputs_path"],
            "input_row_count": report["input_row_count"],
            "xlsx_review_pack_used": report["source_review_paths"]["xlsx"],
            "promotion_evidence": report["promotion_evidence"],
            "evidence_role": report["evidence_role"],
        }
    )
    return 0 if report["status"] in {"PASS", "PASS_WITH_WARNINGS"} else 1


def parse_args(argv: list[str] | None) -> argparse.Namespace:
    parser = argparse.ArgumentParser(description=__doc__)
    parser.add_argument("--contract", default=str(DEFAULT_CONTRACT))
    parser.add_argument("--plan", default=str(DEFAULT_PLAN))
    parser.add_argument("--keyword-review", default=str(DEFAULT_KEYWORD_REVIEW))
    parser.add_argument("--pdf-review", default=str(DEFAULT_PDF_REVIEW))
    parser.add_argument(
        "--xlsx-external-review",
        default="",
        help="Optional user-supplied XLSX review pack path. Local paths are redacted from committed artifacts.",
    )
    parser.add_argument("--xlsx-repo-review", default=str(DEFAULT_XLSX_REPO_REVIEW))
    parser.add_argument("--xlsx-retrieval-report", default=str(DEFAULT_XLSX_RETRIEVAL_REPORT))
    parser.add_argument("--xlsx-failure-breakdown", default=str(DEFAULT_XLSX_FAILURE_BREAKDOWN))
    parser.add_argument("--pdf-retrieval-report", default=str(DEFAULT_PDF_RETRIEVAL_REPORT))
    parser.add_argument("--pdf-c6-report", default=str(DEFAULT_PDF_C6_REPORT))
    parser.add_argument("--pdf-c7-report", default=str(DEFAULT_PDF_C7_REPORT))
    parser.add_argument("--output-root", default=str(EVAL_RUNS_DIR))
    parser.add_argument("--run-id", default="")
    parser.add_argument("--run-prefix", default=RUN_PREFIX)
    parser.add_argument("--max-context-chars", type=int, default=6000)
    parser.add_argument("--db-dsn", default=DEFAULT_DB_DSN)
    parser.add_argument(
        "--disable-xlsx-searchunit-content-join",
        action="store_true",
        help="Disable the read-only join from selected XLSX retrieval hits to SearchUnit payload text.",
    )
    parser.add_argument(
        "--enriched-inputs-alias",
        default="",
        help="Optional extra filename in the artifact dir for the same answer-generation inputs.",
    )
    return parser.parse_args(argv)


def run_builder(
    *,
    contract: Path,
    plan: Path,
    keyword_review: Path,
    pdf_review: Path,
    xlsx_external_review: Path | None,
    xlsx_repo_review: Path,
    xlsx_retrieval_report: Path,
    xlsx_failure_breakdown: Path,
    pdf_retrieval_report: Path,
    pdf_c6_report: Path,
    pdf_c7_report: Path,
    output_root: Path,
    run_id: str = "",
    run_prefix: str = RUN_PREFIX,
    max_context_chars: int = 6000,
    db_dsn: str = DEFAULT_DB_DSN,
    xlsx_searchunit_content_join: bool = True,
    enriched_inputs_alias: str = "",
) -> dict[str, Any]:
    run_id = run_id or utc_run_id()
    generated_at = utc_timestamp()
    artifact_dir = output_root / f"{run_prefix}_{run_id}"
    manifest_path = artifact_dir / "manifest.json"
    inputs_path = artifact_dir / "answer_generation_inputs.jsonl"
    enriched_inputs_alias_path = artifact_dir / enriched_inputs_alias if enriched_inputs_alias else None

    plan_payload = read_json_object(plan)
    diagnostic_rows = list(plan_payload.get("diagnostic_rows") or []) if plan_payload else []
    if not diagnostic_rows:
        raise SystemExit(f"diagnostic_rows missing in {repo_relative(plan)}")

    warnings: list[str] = []
    xlsx_external_used = bool(xlsx_external_review and xlsx_external_review.exists())
    xlsx_review_used = xlsx_external_review if xlsx_external_used and xlsx_external_review else xlsx_repo_review
    if xlsx_external_review and not xlsx_external_review.exists():
        warnings.append("xlsx external review path was provided but missing; repo-local fallback used")
    pdf_review_rows = keyed_by_query_id(read_csv_rows(pdf_review))
    xlsx_review_rows = keyed_by_query_id(read_csv_rows(xlsx_review_used))
    keyword_rows = keyed_by_query_id(read_csv_rows(keyword_review))
    retrieval_maps = build_retrieval_maps(
        xlsx_retrieval_report=xlsx_retrieval_report,
        xlsx_failure_breakdown=xlsx_failure_breakdown,
        pdf_retrieval_report=pdf_retrieval_report,
        pdf_c6_report=pdf_c6_report,
    )
    xlsx_searchunit_index = (
        load_xlsx_searchunit_content_index(retrieval_maps["xlsx"], db_dsn=db_dsn)
        if xlsx_searchunit_content_join
        else empty_xlsx_searchunit_content_index("disabled_by_cli")
    )
    if xlsx_searchunit_index.get("error"):
        warnings.append(f"xlsx SearchUnit content join unavailable: {xlsx_searchunit_index['error']}")
    dataset_index = build_dataset_index(DATASET_DIR)
    extraction_cache: dict[str, Any] = {}

    input_rows: list[dict[str, Any]] = []
    xlsx_review_snapshot = artifact_dir / "source_xlsx_gold_review_pack_used.csv"
    for row_index, plan_row in enumerate(diagnostic_rows, start=1):
        query_id = clean(plan_row.get("query_id"))
        track = clean(plan_row.get("track")).upper()
        review_row = pdf_review_rows.get(query_id, {}) if track == "PDF" else xlsx_review_rows.get(query_id, {})
        keyword_row = keyword_rows.get(query_id, {})
        merged = merge_row(plan_row, review_row, keyword_row)
        locator = parse_locator(clean(merged.get("expected_current_evidence_location")))
        if not locator:
            locator = locator_from_review_row(merged)
        policy = policy_flags(merged)

        context: dict[str, Any]
        if track == "XLSX":
            context = build_xlsx_context(
                row=merged,
                locator=locator,
                dataset_index=dataset_index,
                cache=extraction_cache,
                retrieval_maps=retrieval_maps,
                xlsx_searchunit_index=xlsx_searchunit_index,
                policy=policy,
            )
        elif track == "PDF":
            context = build_pdf_context(
                row=merged,
                locator=locator,
                dataset_index=dataset_index,
                cache=extraction_cache,
                retrieval_maps=retrieval_maps,
                policy=policy,
            )
        else:
            context = {
                "context_type": "unknown",
                "context_available": False,
                "context_errors": [f"unknown track {track!r}"],
            }
            warnings.append(f"{query_id}: unknown track {track!r}")

        prompt_context = build_prompt_context(
            row=merged,
            locator=locator,
            context=context,
            policy=policy,
            max_context_chars=max_context_chars,
        )
        input_rows.append(
            {
                "schema_version": SCHEMA_VERSION,
                "run_id": run_id,
                "row_index": row_index,
                "track": track,
                "query_id": query_id,
                "query": clean(merged.get("query")),
                "bucket": clean(merged.get("bucket")),
                "review_group": clean(merged.get("review_group")),
                "expected_answer_shape": clean(merged.get("expected_answer_shape")),
                "future_content_answer_shape_if_unblocked": clean(
                    merged.get("future_content_answer_shape_if_unblocked")
                ),
                "content_target_needed": clean(merged.get("content_target_needed")),
                "expected_evidence_location": locator,
                "expected_current_evidence_location": clean(
                    merged.get("expected_current_evidence_location")
                ),
                "expected_answer_text": clean(merged.get("expected_answer_text")),
                "must_contain_terms": split_terms(merged.get("must_contain_terms")),
                "keyword_echo_risk": clean(merged.get("keyword_echo_risk")),
                "location_only_answer_risk": clean(merged.get("location_only_answer_risk")),
                "answer_eval_allowed": parse_bool(merged.get("answer_eval_allowed")),
                "exclusion_blocker_reason": clean(merged.get("exclusion_blocker_reason")),
                "citation_target_policy": clean(merged.get("citation_target_policy")),
                "source_review_csv": sanitize_source_review_csv(
                    clean(merged.get("source_review_csv")),
                    track=track,
                    xlsx_review_used=xlsx_review_used,
                    xlsx_review_snapshot=xlsx_review_snapshot,
                    xlsx_external_used=xlsx_external_used,
                ),
                "policy": policy,
                "answer_instruction": answer_instruction(merged, policy),
                "context": context,
                "prompt_context": prompt_context,
                "dry_run_preview_used_as_actual_answer": False,
                "local_llm_run": False,
                "external_live_llm_run": False,
                "optional_judge_run": False,
                "promotion_evidence": False,
                "evidence_role": "diagnostic",
            }
        )

    artifact_dir.mkdir(parents=True, exist_ok=True)
    if xlsx_review_used.exists():
        shutil.copy2(xlsx_review_used, xlsx_review_snapshot)
    write_jsonl(inputs_path, input_rows)
    if enriched_inputs_alias_path is not None:
        write_jsonl(enriched_inputs_alias_path, input_rows)
    manifest = build_manifest(
        run_id=run_id,
        generated_at=generated_at,
        artifact_dir=artifact_dir,
        manifest_path=manifest_path,
        inputs_path=inputs_path,
        enriched_inputs_alias_path=enriched_inputs_alias_path,
        rows=input_rows,
        source_paths={
            "contract": contract,
            "plan": plan,
            "keyword_review": keyword_review,
            "pdf_review": pdf_review,
            "xlsx_review": xlsx_review_snapshot if xlsx_external_used else xlsx_review_used,
            "xlsx_review_snapshot": xlsx_review_snapshot,
            "xlsx_repo_local_fallback": xlsx_repo_review,
            "xlsx_retrieval_report": xlsx_retrieval_report,
            "xlsx_failure_breakdown": xlsx_failure_breakdown,
            "pdf_retrieval_report": pdf_retrieval_report,
            "pdf_c6_report": pdf_c6_report,
            "pdf_c7_report": pdf_c7_report,
        },
        plan_payload=plan_payload,
        warnings=warnings,
        xlsx_external_used=xlsx_external_used,
        xlsx_external_basename=xlsx_external_review.name if xlsx_external_review else "",
        xlsx_searchunit_index=xlsx_searchunit_index,
    )
    write_json(manifest_path, manifest)
    return manifest


def build_xlsx_context(
    *,
    row: Mapping[str, Any],
    locator: Mapping[str, Any],
    dataset_index: Mapping[str, list[Path]],
    cache: dict[str, Any],
    retrieval_maps: Mapping[str, Mapping[str, Any]],
    xlsx_searchunit_index: Mapping[str, Any],
    policy: Mapping[str, bool],
) -> dict[str, Any]:
    query_id = clean(row.get("query_id"))
    file_name = clean(locator.get("file") or row.get("expected_file_name"))
    sheet_name = clean(locator.get("sheet") or row.get("expected_sheet_name"))
    cell_range = clean(locator.get("range") or row.get("expected_cell_range"))
    table_id = clean(locator.get("table") or row.get("expected_table_id"))
    expected_locator_diagnostic_only = compact_dict(
        {
            "file": file_name,
            "sheet": sheet_name,
            "range": cell_range,
            "cell": clean(locator.get("cell")),
            "table": table_id,
            "document_version_id": clean(locator.get("docv") or row.get("expected_document_version_id")),
        }
    )
    context: dict[str, Any] = {
        "context_type": "xlsx",
        "file_name": file_name,
        "sheet_name": "",
        "cell_range": "",
        "table_id": "",
        "locator": {},
        "expected_evidence_locator_diagnostic_only": expected_locator_diagnostic_only,
        "row_label": "",
        "column_label": "",
        "header_context": [],
        "value_context": [],
        "nearby_table_context": [],
        "retrieval_context": retrieval_context_for(query_id, retrieval_maps, "xlsx"),
        "context_available": False,
        "context_has_expected_terms": False,
        "context_errors": [],
    }
    if (
        policy.get("hidden_policy_blocked")
        or policy.get("formula_date_policy_blocked")
        or clean(row.get("expected_answer_shape")) == ANSWER_SHAPE_POLICY_PENDING
    ):
        context["context_errors"].append("content extraction skipped for policy-pending/not-answerable row")
        context["xlsx_searchunit_content_join"] = {
            "status": "SKIPPED_POLICY_PENDING",
            "answer_evidence_used": False,
        }
        return context

    join_context = xlsx_searchunit_answer_context(
        retrieval_context=context["retrieval_context"],
        searchunit_index=xlsx_searchunit_index,
    )
    if join_context.get("context_available"):
        context.update(join_context)
        context["context_has_expected_terms"] = False
        return context
    context["xlsx_searchunit_content_join"] = join_context.get(
        "xlsx_searchunit_content_join",
        {"status": "NO_JOIN", "answer_evidence_used": False},
    )

    source_path = resolve_dataset_file(file_name, dataset_index)
    if source_path is None:
        context["context_errors"].append(f"source workbook not found under {repo_relative(DATASET_DIR)}")
    else:
        context["source_path"] = repo_relative(source_path)
    context["context_errors"].append(
        "no selected SearchUnit content joined; source workbook probing is diagnostic-only and not promoted"
    )
    return context


def build_pdf_context(
    *,
    row: Mapping[str, Any],
    locator: Mapping[str, Any],
    dataset_index: Mapping[str, list[Path]],
    cache: dict[str, Any],
    retrieval_maps: Mapping[str, Mapping[str, Any]],
    policy: Mapping[str, bool],
) -> dict[str, Any]:
    query_id = clean(row.get("query_id"))
    file_name = clean(locator.get("file") or row.get("expected_file_name"))
    page_no = clean(locator.get("page") or row.get("expected_page_no"))
    bbox = locator.get("bbox") or clean(row.get("expected_bbox"))
    context: dict[str, Any] = {
        "context_type": "pdf",
        "file_name": file_name,
        "page_no": page_no,
        "page_label": clean(locator.get("page_label") or row.get("expected_page_label")),
        "physical_page_index": clean(
            locator.get("physical_page_index") or row.get("expected_physical_page_index")
        ),
        "bbox": bbox,
        "section_id": clean(row.get("expected_section_id")),
        "locator": compact_dict(
            {
                "file": file_name,
                "page": page_no,
                "page_label": clean(locator.get("page_label") or row.get("expected_page_label")),
                "physical_page_index": clean(
                    locator.get("physical_page_index") or row.get("expected_physical_page_index")
                ),
                "bbox": bbox,
                "document_version_id": clean(locator.get("docv") or row.get("expected_document_version_id")),
            }
        ),
        "sentence_context": [],
        "paragraph_context": [],
        "table_or_value_context": [],
        "retrieval_context": retrieval_context_for(query_id, retrieval_maps, "pdf"),
        "context_available": False,
        "context_has_expected_terms": False,
        "context_errors": [],
    }
    if clean(row.get("expected_answer_shape")) == ANSWER_SHAPE_POLICY_PENDING:
        context["context_errors"].append("content extraction skipped for PDF C7 policy-pending row")
        return context

    source_path = resolve_dataset_file(file_name, dataset_index)
    if source_path is None:
        context["context_errors"].append(f"source PDF not found under {repo_relative(DATASET_DIR)}")
        return context
    context["source_path"] = repo_relative(source_path)
    try:
        extracted = extract_pdf_context(
            source_path=source_path,
            page_no=page_no,
            physical_page_index=clean(
                locator.get("physical_page_index") or row.get("expected_physical_page_index")
            ),
            bbox=bbox,
            expected_terms=split_terms(row.get("must_contain_terms")) + [clean(row.get("expected_answer_text"))],
            cache=cache,
        )
    except Exception as exc:  # pragma: no cover - diagnostic fallback
        context["context_errors"].append(f"pdf extraction failed: {type(exc).__name__}: {exc}")
        return context

    context.update(extracted)
    context["context_available"] = bool(
        context.get("sentence_context") or context.get("paragraph_context") or context.get("table_or_value_context")
    )
    context["context_has_expected_terms"] = has_any_term(
        json.dumps(extracted, ensure_ascii=False), split_terms(row.get("must_contain_terms"))
    )
    return context


def extract_xlsx_context(
    *,
    source_path: Path,
    sheet_name: str,
    cell_range: str,
    expected_terms: list[str],
    cache: dict[str, Any],
) -> dict[str, Any]:
    try:
        from openpyxl import load_workbook
        from openpyxl.utils.cell import get_column_letter, range_boundaries
    except Exception as exc:  # pragma: no cover - optional dependency
        return {
            "context_errors": [f"openpyxl unavailable: {type(exc).__name__}: {exc}"],
            "context_available": False,
        }

    cache_key = f"xlsx::{source_path}"
    if cache_key not in cache:
        cache[cache_key] = {
            "formula": load_workbook(source_path, data_only=False, read_only=True),
            "values": load_workbook(source_path, data_only=True, read_only=True),
        }
    workbook_formula = cache[cache_key]["formula"]
    workbook_values = cache[cache_key]["values"]
    if sheet_name not in workbook_formula.sheetnames:
        return {"context_errors": [f"sheet not found: {sheet_name}"]}
    worksheet_formula = workbook_formula[sheet_name]
    worksheet_values = workbook_values[sheet_name]
    if worksheet_formula.sheet_state != "visible":
        return {"context_errors": [f"sheet is not visible: {sheet_name}"]}

    try:
        min_col, min_row, max_col, max_row = range_boundaries(cell_range)
    except Exception:
        min_col, min_row, max_col, max_row = 1, 1, min(worksheet_formula.max_column, 8), min(
            worksheet_formula.max_row, 8
        )

    max_rows = min(max_row, min_row + 7)
    max_cols = min(max_col, min_col + 7)
    header_values = [
        cell_to_text(worksheet_values.cell(row=min_row, column=col).value)
        for col in range(min_col, max_cols + 1)
    ]
    nearby_table: list[str] = []
    value_context: list[dict[str, str]] = []
    row_labels: list[str] = []
    column_labels: list[str] = [value for value in header_values if value]

    for row_index in range(min_row, max_rows + 1):
        if worksheet_formula.row_dimensions[row_index].hidden:
            continue
        row_items: list[str] = []
        first_value = cell_to_text(worksheet_values.cell(row=row_index, column=min_col).value)
        if first_value:
            row_labels.append(first_value)
        for col_index in range(min_col, max_cols + 1):
            column_letter = get_column_letter(col_index)
            if worksheet_formula.column_dimensions[column_letter].hidden:
                continue
            header = header_values[col_index - min_col] or column_letter
            value = cell_to_text(worksheet_values.cell(row=row_index, column=col_index).value)
            formula = cell_to_text(worksheet_formula.cell(row=row_index, column=col_index).value)
            display = value
            if formula.startswith("=") and formula != value:
                display = f"{value} (formula present)" if value else "formula present"
            if display:
                row_items.append(f"{header}: {display}")
                joined = f"{header} {display}"
                if has_any_term(joined, expected_terms):
                    value_context.append(
                        {
                            "cell": f"{column_letter}{row_index}",
                            "row_label": first_value,
                            "column_label": header,
                            "value": value,
                            "formula_present": bool(formula.startswith("=")),
                        }
                    )
        if row_items:
            nearby_table.append(" | ".join(row_items))

    return {
        "row_label": first_nonempty(row_labels),
        "column_label": first_nonempty(column_labels),
        "header_context": column_labels,
        "value_context": value_context[:8],
        "nearby_table_context": nearby_table[:8],
        "xlsx_extraction_policy": "visible_sheets_rows_columns_only_no_raw_formulas",
        "context_errors": [],
    }


def extract_pdf_context(
    *,
    source_path: Path,
    page_no: str,
    physical_page_index: str,
    bbox: object,
    expected_terms: list[str],
    cache: dict[str, Any],
) -> dict[str, Any]:
    try:
        import fitz  # type: ignore
    except Exception as exc:  # pragma: no cover - optional dependency
        return {
            "context_errors": [f"pymupdf unavailable: {type(exc).__name__}: {exc}"],
            "context_available": False,
        }

    cache_key = f"pdf::{source_path}"
    if cache_key not in cache:
        cache[cache_key] = fitz.open(source_path)
    doc = cache[cache_key]
    page_index = int_or_none(physical_page_index)
    if page_index is None:
        page_number = int_or_none(page_no)
        page_index = page_number - 1 if page_number is not None else 0
    if page_index < 0 or page_index >= len(doc):
        return {"context_errors": [f"page index out of bounds: {page_index}"]}

    page = doc[page_index]
    page_text = clean_whitespace(page.get_text("text"))
    bbox_text = ""
    bbox_values = bbox_list(bbox)
    paragraph_window = ""
    if bbox_values:
        rect = fitz.Rect(*bbox_values)
        bbox_text = clean_whitespace(page.get_text("text", clip=rect))
        paragraph_window = pdf_adjacent_block_window_from_blocks(
            page.get_text("blocks"),
            bbox=bbox_values,
            max_chars=1200,
        )

    term_sentences = select_sentences(page_text, expected_terms)
    table_or_value_context = []
    if bbox_text:
        table_or_value_context.append(bbox_text)
    if not term_sentences and page_text:
        term_sentences = [page_text[:1200]]

    return {
        "sentence_context": term_sentences[:5],
        "paragraph_context": ([bbox_text] if bbox_text else term_sentences[:2]),
        "paragraph_window": paragraph_window,
        "table_or_value_context": table_or_value_context[:3],
        "page_text_excerpt": page_text[:2000],
        "pdf_extraction_policy": "pymupdf_page_text_and_bbox_clip",
        "context_errors": [],
    }


def pdf_adjacent_block_window_from_blocks(
    blocks: Iterable[object],
    *,
    bbox: object,
    max_chars: int,
    sibling_window: int = 1,
) -> str:
    bbox_values = bbox_list(bbox)
    if not bbox_values:
        return ""
    text_blocks = normalized_pdf_text_blocks(blocks)
    if not text_blocks:
        return ""
    target_index = pdf_target_block_index(text_blocks, bbox_values)
    if target_index is None:
        return ""
    width = max(0, sibling_window)
    start = max(0, target_index - width)
    end = min(len(text_blocks), target_index + width + 1)
    window = " ".join(block["text"] for block in text_blocks[start:end])
    return truncate(clean_whitespace(window), max_chars)


def normalized_pdf_text_blocks(blocks: Iterable[object]) -> list[dict[str, Any]]:
    normalized: list[dict[str, Any]] = []
    for order, block in enumerate(blocks):
        if not isinstance(block, (list, tuple)) or len(block) < 5:
            continue
        block_type = int_or_none(block[6]) if len(block) > 6 else 0
        if block_type not in (None, 0):
            continue
        text = clean_whitespace(block[4])
        if not text:
            continue
        try:
            rect = [float(block[0]), float(block[1]), float(block[2]), float(block[3])]
        except (TypeError, ValueError):
            continue
        normalized.append({"order": order, "rect": rect, "text": text})
    return sorted(normalized, key=lambda item: (item["rect"][1], item["rect"][0], item["order"]))


def pdf_target_block_index(blocks: list[dict[str, Any]], bbox: list[float]) -> int | None:
    scored = [
        (pdf_rect_overlap_area(block["rect"], bbox), index)
        for index, block in enumerate(blocks)
    ]
    scored = [(score, index) for score, index in scored if score > 0]
    if scored:
        return max(scored, key=lambda item: item[0])[1]

    center_x = (bbox[0] + bbox[2]) / 2
    center_y = (bbox[1] + bbox[3]) / 2
    for index, block in enumerate(blocks):
        left, top, right, bottom = block["rect"]
        if left <= center_x <= right and top <= center_y <= bottom:
            return index
    return None


def pdf_rect_overlap_area(left: list[float], right: list[float]) -> float:
    x0 = max(left[0], right[0])
    y0 = max(left[1], right[1])
    x1 = min(left[2], right[2])
    y1 = min(left[3], right[3])
    if x1 <= x0 or y1 <= y0:
        return 0.0
    return (x1 - x0) * (y1 - y0)


def build_prompt_context(
    *,
    row: Mapping[str, Any],
    locator: Mapping[str, Any],
    context: Mapping[str, Any],
    policy: Mapping[str, bool],
    max_context_chars: int,
) -> str:
    payload = {
        "task": "diagnostic_pdf_xlsx_answer_shape",
        "query": clean(row.get("query")),
        "expected_answer_shape": clean(row.get("expected_answer_shape")),
        "answer_instruction": answer_instruction(row, policy),
        "citation_policy": "cite_only_bound_context_locator",
        "policy": policy,
        "context": context,
    }
    text = json.dumps(payload, ensure_ascii=False, indent=2)
    if max_context_chars > 0 and len(text) > max_context_chars:
        return text[:max_context_chars] + "\n...TRUNCATED_FOR_LOCAL_LLM_INPUT..."
    return text


def answer_instruction(row: Mapping[str, Any], policy: Mapping[str, bool]) -> str:
    shape = clean(row.get("expected_answer_shape"))
    if (
        shape == ANSWER_SHAPE_POLICY_PENDING
        or policy.get("hidden_policy_blocked")
        or policy.get("pdf_c7_policy_pending")
    ):
        return (
            "ABSTAIN_OR_POLICY_PENDING: do not answer with hidden, policy-pending, "
            "or not-answerable content. Fill abstain_reason."
        )
    return (
        "ANSWER_CONTENT_FIRST: provide the content claim/value/summary first; "
        "use page/bbox/sheet/range/cell only as citation evidence."
    )


def build_retrieval_maps(
    *,
    xlsx_retrieval_report: Path,
    xlsx_failure_breakdown: Path,
    pdf_retrieval_report: Path,
    pdf_c6_report: Path,
) -> dict[str, dict[str, Any]]:
    return {
        "xlsx": {
            "query_results": keyed_json_rows(read_json_object(xlsx_retrieval_report), "query_results"),
            "classified_rows": keyed_json_rows(read_json_object(xlsx_failure_breakdown), "classified_query_rows"),
        },
        "pdf": {
            "query_results": keyed_json_rows(read_json_object(pdf_retrieval_report), "query_results"),
            "classified_rows": keyed_json_rows(read_json_object(pdf_c6_report), "classified_query_rows"),
        },
    }


def retrieval_context_for(
    query_id: str,
    retrieval_maps: Mapping[str, Mapping[str, Any]],
    track_key: str,
) -> dict[str, Any]:
    track_maps = retrieval_maps.get(track_key, {})
    query_result = dict(track_maps.get("query_results", {}).get(query_id, {}))
    classified = dict(track_maps.get("classified_rows", {}).get(query_id, {}))
    top_results = list(query_result.get("top_k_results") or [])[:5]
    compact_top = []
    for result in top_results:
        if not isinstance(result, Mapping):
            continue
        location = result.get("location_json") if isinstance(result.get("location_json"), Mapping) else {}
        compact_top.append(
            compact_dict(
                {
                    "rank": result.get("rank"),
                    "score": result.get("score"),
                    "search_unit_id": result.get("search_unit_id") or result.get("searchUnitId"),
                    "node_id": result.get("node_id") or result.get("nodeId"),
                    "chunk_id": result.get("chunk_id") or result.get("chunkId"),
                    "source_file_name": result.get("source_file_name"),
                    "source_file_type": result.get("source_file_type"),
                    "chunk_type": result.get("chunk_type"),
                    "citation_text": result.get("citation_text"),
                    "location_json": location,
                    "parser_name": result.get("parser_name"),
                    "parser_version": result.get("parser_version"),
                    "index_version": result.get("index_version"),
                    "embedding_status": result.get("embedding_status"),
                    "match_breakdown": result.get("match_breakdown"),
                }
            )
        )
    return compact_dict(
        {
            "retrieval_query_status": query_result.get("status"),
            "top_k_results": compact_top,
            "failure_or_quality_classification": classified.get("category") or classified.get("classification"),
            "classification_rationale": classified.get("rationale"),
            "top_k_summary": classified.get("top_k_summary"),
            "range_relation_in_top_k": classified.get("range_relation_in_top_k"),
            "supporting_hits": classified.get("supporting_hits"),
        }
    )


def empty_xlsx_searchunit_content_index(error: str = "") -> dict[str, Any]:
    return {
        "enabled": False,
        "status": "DISABLED" if error else "EMPTY",
        "error": error,
        "by_id": {},
        "by_docv": {},
        "selected_search_unit_id_count": 0,
        "loaded_search_unit_count": 0,
    }


def load_xlsx_searchunit_content_index(
    retrieval_maps: Mapping[str, Mapping[str, Any]],
    *,
    db_dsn: str,
) -> dict[str, Any]:
    query_results = retrieval_maps.get("query_results") or {}
    selected_ids: set[str] = set()
    docv_ids: set[str] = set()
    for row in query_results.values():
        if not isinstance(row, Mapping):
            continue
        for hit in row.get("top_k_results") or []:
            if not isinstance(hit, Mapping):
                continue
            for selected_id in retrieval_hit_identity_candidates(hit):
                selected_ids.add(selected_id)
            location = hit.get("location_json") if isinstance(hit.get("location_json"), Mapping) else {}
            docv = clean(location.get("document_version_id") or hit.get("document_version_id"))
            if docv:
                docv_ids.add(docv)
    if not selected_ids and not docv_ids:
        return empty_xlsx_searchunit_content_index("no_selected_xlsx_searchunit_or_docv_ids")
    try:
        import psycopg2
        import psycopg2.extras
    except Exception as exc:  # pragma: no cover - optional dependency
        return empty_xlsx_searchunit_content_index(f"psycopg2 unavailable: {type(exc).__name__}: {exc}")

    try:
        with psycopg2.connect(db_dsn, cursor_factory=psycopg2.extras.RealDictCursor) as conn:
            conn.set_session(readonly=True, autocommit=True)
            with conn.cursor() as cur:
                cur.execute(
                    """
                    SELECT id, index_id, document_version_id, source_file_id, source_file_name,
                           unit_type, unit_key, chunk_type, location_json, text_content,
                           embedding_text, bm25_text, display_text, debug_text, citation_text,
                           parser_name, parser_version, index_version, embedding_status
                      FROM search_unit
                     WHERE (
                            id = ANY(%s)
                            OR index_id = ANY(%s)
                            OR document_version_id = ANY(%s)
                           )
                       AND parser_version = 'xlsx-extract-v2-hidden-safe'
                       AND index_version = 'rag-ingestion-v2-xlsx-candidate-v1'
                    """,
                    (list(selected_ids), list(selected_ids), list(docv_ids)),
                )
                rows = [dict(row) for row in cur.fetchall()]
    except Exception as exc:
        return empty_xlsx_searchunit_content_index(f"{type(exc).__name__}: {exc}")

    by_id: dict[str, dict[str, Any]] = {}
    by_docv: dict[str, list[dict[str, Any]]] = {}
    for row in rows:
        unit_id = clean(row.get("id"))
        if unit_id:
            by_id[unit_id] = row
        index_id = clean(row.get("index_id"))
        if index_id:
            by_id[index_id] = row
        by_docv.setdefault(clean(row.get("document_version_id")), []).append(row)
    return {
        "enabled": True,
        "status": "LOADED",
        "error": "",
        "by_id": by_id,
        "by_docv": by_docv,
        "selected_search_unit_id_count": len(selected_ids),
        "selected_document_version_id_count": len(docv_ids),
        "loaded_search_unit_count": len(rows),
    }


def xlsx_searchunit_answer_context(
    *,
    retrieval_context: Mapping[str, Any],
    searchunit_index: Mapping[str, Any],
) -> dict[str, Any]:
    join_meta: dict[str, Any] = {
        "status": "NO_JOIN",
        "answer_evidence_used": False,
        "join_type": "",
        "rank": None,
        "search_unit_id": "",
        "failed_join_reasons": [],
    }
    if not parse_bool(searchunit_index.get("enabled")):
        join_meta["status"] = "JOIN_UNAVAILABLE"
        join_meta["failed_join_reasons"].append(clean(searchunit_index.get("error")) or "join disabled")
        return {"context_available": False, "xlsx_searchunit_content_join": join_meta}

    for hit in retrieval_context.get("top_k_results") or []:
        if not isinstance(hit, Mapping):
            continue
        unit, join_type, reason = find_joined_xlsx_searchunit(hit, searchunit_index)
        if not unit:
            if reason:
                join_meta["failed_join_reasons"].append(reason)
            continue
        content = context_from_joined_xlsx_searchunit(unit, hit, join_type=join_type)
        if content.get("context_available"):
            return content
        join_meta["failed_join_reasons"].append(
            f"rank {clean(hit.get('rank')) or '?'} joined but had no concrete content"
        )
    return {"context_available": False, "xlsx_searchunit_content_join": join_meta}


def find_joined_xlsx_searchunit(
    hit: Mapping[str, Any],
    searchunit_index: Mapping[str, Any],
) -> tuple[dict[str, Any] | None, str, str]:
    by_id = searchunit_index.get("by_id") if isinstance(searchunit_index.get("by_id"), Mapping) else {}
    by_docv = searchunit_index.get("by_docv") if isinstance(searchunit_index.get("by_docv"), Mapping) else {}
    for hit_id in retrieval_hit_identity_candidates(hit):
        unit = by_id.get(hit_id)
        if not isinstance(unit, Mapping):
            continue
        join_type = safe_xlsx_join_type(hit, unit)
        if not join_type:
            return None, "", f"rank {clean(hit.get('rank')) or '?'} selected SearchUnit id was not safe answer evidence"
        return dict(unit), join_type, ""
    if retrieval_hit_identity_candidates(hit):
        return None, "", f"rank {clean(hit.get('rank')) or '?'} selected SearchUnit id not loaded"

    hit_loc = xlsx_location(hit.get("location_json") if isinstance(hit.get("location_json"), Mapping) else {})
    docv = hit_loc.get("document_version_id")
    candidates = by_docv.get(docv) if docv else []
    if not isinstance(candidates, list):
        return None, "", f"rank {clean(hit.get('rank')) or '?'} has no document_version candidates"

    exact_matches: list[tuple[dict[str, Any], str]] = []
    overlap_matches: list[tuple[dict[str, Any], str]] = []
    for candidate in candidates:
        if not isinstance(candidate, Mapping):
            continue
        join_type = safe_xlsx_join_type(hit, candidate)
        if join_type == "exact_locator":
            exact_matches.append((dict(candidate), join_type))
        elif join_type == "safe_range_overlap":
            overlap_matches.append((dict(candidate), join_type))
    if exact_matches:
        return exact_matches[0][0], exact_matches[0][1], ""
    if overlap_matches:
        return overlap_matches[0][0], overlap_matches[0][1], ""
    return None, "", f"rank {clean(hit.get('rank')) or '?'} had no exact or safe range-overlap SearchUnit"


def safe_xlsx_join_type(hit: Mapping[str, Any], unit: Mapping[str, Any]) -> str:
    chunk_type = clean(unit.get("chunk_type") or hit.get("chunk_type") or unit.get("unit_type")).lower()
    if chunk_type in XLSX_BROAD_EVIDENCE_CHUNK_TYPES:
        return ""
    if chunk_type and chunk_type not in XLSX_ALLOWED_EVIDENCE_CHUNK_TYPES:
        return ""
    hit_loc = xlsx_location(hit.get("location_json") if isinstance(hit.get("location_json"), Mapping) else {})
    unit_loc = xlsx_location(unit.get("location_json") if isinstance(unit.get("location_json"), Mapping) else {})
    hit_docv = hit_loc.get("document_version_id")
    unit_docv = clean(unit.get("document_version_id") or unit_loc.get("document_version_id"))
    if hit_docv and unit_docv and hit_docv != unit_docv:
        return ""
    if not hit_loc.get("sheet") or not unit_loc.get("sheet"):
        return ""
    if hit_loc["sheet"] != unit_loc["sheet"]:
        return ""
    hit_range = parse_cell_range(hit_loc.get("range", ""))
    unit_range = parse_cell_range(unit_loc.get("range", ""))
    if not hit_range or not unit_range:
        return ""
    if hit_range == unit_range:
        return "exact_locator"
    if ranges_overlap(hit_range, unit_range):
        return "safe_range_overlap"
    return ""


def retrieval_hit_identity_candidates(hit: Mapping[str, Any]) -> list[str]:
    candidates = [
        hit.get("search_unit_id"),
        hit.get("searchUnitId"),
        hit.get("node_id"),
        hit.get("nodeId"),
        hit.get("chunk_id"),
        hit.get("chunkId"),
        hit.get("index_id"),
        hit.get("indexId"),
        hit.get("id"),
    ]
    seen: set[str] = set()
    identities: list[str] = []
    for value in candidates:
        identity = clean(value)
        if not identity or identity in seen:
            continue
        seen.add(identity)
        identities.append(identity)
    return identities


def context_from_joined_xlsx_searchunit(
    unit: Mapping[str, Any],
    hit: Mapping[str, Any],
    *,
    join_type: str,
) -> dict[str, Any]:
    content_field, text = first_searchunit_content(unit)
    if not xlsx_text_has_content(text):
        return {"context_available": False}
    location = xlsx_location(unit.get("location_json") if isinstance(unit.get("location_json"), Mapping) else {})
    parsed = parse_xlsx_payload_text(text, source_field=f"search_unit.{content_field}")
    if not (parsed["row_values"] or parsed["cell_values"] or parsed["table_context"]):
        return {"context_available": False}
    locator = compact_dict(
        {
            "file": clean(unit.get("source_file_name") or hit.get("source_file_name")),
            "sheet": location.get("sheet"),
            "range": location.get("range"),
            "document_version_id": clean(unit.get("document_version_id") or location.get("document_version_id")),
            "search_unit_id": clean(unit.get("id")),
            "chunk_type": clean(unit.get("chunk_type") or hit.get("chunk_type")),
            "join_type": join_type,
            "rank": hit.get("rank"),
        }
    )
    first_row_value = parsed["row_values"][0] if parsed["row_values"] else {}
    context = {
        "file_name": locator.get("file", ""),
        "sheet_name": locator.get("sheet", ""),
        "cell_range": locator.get("range", ""),
        "table_id": locator.get("table", ""),
        "selected_search_unit_id": locator.get("search_unit_id", ""),
        "selected_searchunit_locator": locator,
        "content_source_locator": locator,
        "citation_locator": locator,
        "retrieval_text": truncate(clean_whitespace(text), 2400),
        "table_context": parsed["table_context"],
        "nearby_rows": parsed["nearby_rows"],
        "nearby_table_context": parsed["table_context"],
        "header_context": parsed["headers"],
        "headers": parsed["headers"],
        "row_values": parsed["row_values"],
        "cell_values": parsed["cell_values"],
        "column_values": parsed["column_values"],
        "value_context": parsed["row_values"][:8],
        "row_label": clean(first_row_value.get("row_label")),
        "column_label": clean(first_row_value.get("column_label")),
        "value": clean(first_row_value.get("value")),
        "content_summary": parsed["content_summary"],
        "citation_locator": locator,
        "locator": locator,
        "retrieval_context_joined_hit": compact_dict(
            {
                "rank": hit.get("rank"),
                "score": hit.get("score"),
                "search_unit_id": clean(unit.get("id")),
                "join_type": join_type,
                "content_source_field": f"search_unit.{content_field}",
            }
        ),
        "xlsx_searchunit_content_join": {
            "status": "JOINED",
            "answer_evidence_used": True,
            "join_type": join_type,
            "rank": hit.get("rank"),
            "search_unit_id": clean(unit.get("id")),
            "content_source_field": f"search_unit.{content_field}",
            "chunk_type": clean(unit.get("chunk_type") or hit.get("chunk_type")),
            "source_workbook_promoted_evidence": False,
            "gold_leakage": False,
            "broad_fallback_promoted": False,
        },
        "context_available": True,
        "context_errors": [],
    }
    return compact_dict(context)


def first_searchunit_content(unit: Mapping[str, Any]) -> tuple[str, str]:
    for field in ("display_text", "text_content", "bm25_text", "debug_text", "embedding_text"):
        text = clean(unit.get(field))
        if xlsx_text_has_content(text):
            return field, text
    return "", ""


def parse_xlsx_payload_text(text: str, *, source_field: str) -> dict[str, Any]:
    lines = normalized_xlsx_content_lines(text)
    rows: list[str] = []
    headers: list[str] = []
    row_values: list[dict[str, str]] = []
    cell_values: list[dict[str, str]] = []
    column_values: list[dict[str, str]] = []
    for line in lines:
        pairs = key_value_pairs_from_line(line)
        if not pairs:
            continue
        rows.append(line)
        row_label = clean(pairs[0][1])
        ordered_pairs = pairs[1:] if len(pairs) > 1 else pairs
        for header, _value in pairs:
            if header and header not in headers:
                headers.append(header)
        for header, value in ordered_pairs:
            if header and header not in headers:
                headers.append(header)
            if not xlsx_text_has_content(value):
                continue
            item = compact_dict(
                {
                    "row_label": row_label,
                    "column_label": header,
                    "value": value,
                    "row_text": line,
                    "source_field": source_field,
                }
            )
            row_values.append(item)
            cell_values.append(item)
            column_values.append(
                compact_dict(
                    {
                        "column_label": header,
                        "value": value,
                        "column_text": f"{header}: {value}",
                        "source_field": source_field,
                    }
                )
            )
    summary = f"Retrieved SearchUnit content: {rows[0]}" if rows else ""
    return {
        "headers": headers[:16],
        "nearby_rows": rows[:8],
        "table_context": rows[:8],
        "row_values": row_values[:24],
        "cell_values": cell_values[:24],
        "column_values": column_values[:16],
        "content_summary": truncate(summary, 1200),
    }


def normalized_xlsx_content_lines(text: str) -> list[str]:
    normalized: list[str] = []
    for raw_line in re.split(r"[\r\n]+", clean(text)):
        line = clean(raw_line)
        if not line:
            continue
        folded = line.lower()
        if folded.startswith(("[sheet:", "[range:", "source:", "content:")):
            continue
        if "|" not in line and line.count(":") < 1:
            continue
        if xlsx_text_has_content(line):
            normalized.append(line)
    if not normalized and "|" in text:
        line = clean_whitespace(text)
        if xlsx_text_has_content(line):
            normalized.append(line)
    return normalized[:16]


def key_value_pairs_from_line(line: str) -> list[tuple[str, str]]:
    pairs: list[tuple[str, str]] = []
    for segment in line.split("|"):
        if ":" not in segment:
            continue
        key, value = segment.split(":", 1)
        key = clean(key)
        value = clean(value)
        if not key or not value:
            continue
        if looks_like_locator_only(key) or looks_like_locator_only(value):
            continue
        pairs.append((key, value))
    return pairs


def xlsx_location(value: Mapping[str, Any]) -> dict[str, str]:
    return compact_dict(
        {
            "document_version_id": clean(value.get("document_version_id")),
            "sheet": clean(value.get("sheet_name") or value.get("sheetName") or value.get("sheet")),
            "range": clean(value.get("cell_range") or value.get("cellRange") or value.get("range")),
            "cell": clean(value.get("cell") or value.get("cell_ref") or value.get("cellRef")),
        }
    )


def parse_cell_range(value: str) -> tuple[int, int, int, int] | None:
    text = clean(value)
    if not text:
        return None
    try:
        from openpyxl.utils.cell import range_boundaries

        min_col, min_row, max_col, max_row = range_boundaries(text)
        return min_col, min_row, max_col, max_row
    except Exception:
        return None


def ranges_overlap(a: tuple[int, int, int, int], b: tuple[int, int, int, int]) -> bool:
    a_min_col, a_min_row, a_max_col, a_max_row = a
    b_min_col, b_min_row, b_max_col, b_max_row = b
    return not (
        a_max_col < b_min_col
        or b_max_col < a_min_col
        or a_max_row < b_min_row
        or b_max_row < a_min_row
    )


def xlsx_text_has_content(value: object) -> bool:
    text = clean(value)
    if len(text) < 2:
        return False
    if looks_like_locator_only(text):
        return False
    return bool(re.search(r"[A-Za-z0-9가-힣]", text))


def looks_like_locator_only(value: object) -> bool:
    folded = re.sub(r"\s+", "", clean(value)).lower()
    if not folded:
        return True
    patterns = (
        r"^[a-z]{1,3}\d+(:[a-z]{1,3}\d+)?$",
        r"^sheet\d*$",
        r"^.+\.xlsx$",
        r"^docv_[0-9a-f]+$",
    )
    return any(re.match(pattern, folded) for pattern in patterns)


def truncate(value: str, limit: int) -> str:
    text = clean(value)
    if limit <= 0 or len(text) <= limit:
        return text
    return text[: max(limit - 3, 0)] + "..."


def policy_flags(row: Mapping[str, Any]) -> dict[str, bool]:
    blocker = clean(row.get("exclusion_blocker_reason")).upper()
    review_group = clean(row.get("review_group")).lower()
    bucket = clean(row.get("bucket")).lower()
    shape = clean(row.get("expected_answer_shape"))
    return {
        "diagnostic_only": True,
        "pdf_c7_policy_pending": "PDF_C7_POLICY_PENDING" in blocker,
        "pdf_user_review_not_done": "PDF_USER_REVIEW_NOT_DONE" in blocker,
        "xlsx_answer_quality_blocked": "XLSX_ANSWER_QUALITY_BLOCKED" in blocker,
        "hidden_policy_blocked": "hidden_policy" in bucket
        or "HIDDEN_NEGATIVE_POLICY" in blocker
        or "hidden_negative" in review_group,
        "formula_date_policy_blocked": "FORMULA_DATE" in blocker or "formula_date" in review_group,
        "not_answerable_or_policy_pending": shape == ANSWER_SHAPE_POLICY_PENDING,
        "answer_eval_allowed": parse_bool(row.get("answer_eval_allowed")),
        "promotion_evidence": False,
    }


def sanitize_source_review_csv(
    value: str,
    *,
    track: str,
    xlsx_review_used: Path,
    xlsx_review_snapshot: Path,
    xlsx_external_used: bool,
) -> str:
    if track == "XLSX" and xlsx_external_used:
        return repo_relative(xlsx_review_snapshot)
    if not value:
        return ""
    candidate = Path(value)
    if candidate.is_absolute() and not is_under_repo(candidate):
        return f"external_path_redacted:{candidate.name}"
    if candidate == xlsx_review_used and is_under_repo(candidate):
        return repo_relative(candidate)
    return value.replace("\\", "/")


def build_manifest(
    *,
    run_id: str,
    generated_at: str,
    artifact_dir: Path,
    manifest_path: Path,
    inputs_path: Path,
    enriched_inputs_alias_path: Path | None,
    rows: list[Mapping[str, Any]],
    source_paths: Mapping[str, Path],
    plan_payload: Mapping[str, Any],
    warnings: list[str],
    xlsx_external_used: bool,
    xlsx_external_basename: str,
    xlsx_searchunit_index: Mapping[str, Any],
) -> dict[str, Any]:
    track_counts = Counter(clean(row.get("track")) for row in rows)
    shape_counts = Counter(clean(row.get("expected_answer_shape")) for row in rows)
    context_available_count = sum(1 for row in rows if nested_bool(row, ["context", "context_available"]))
    context_expected_terms_count = sum(
        1 for row in rows if nested_bool(row, ["context", "context_has_expected_terms"])
    )
    source_inventory = {
        key: {
            "path": repo_relative(path),
            "repo_path": repo_relative(path) if is_under_repo(path) else str(path),
            "exists": path.exists(),
            "sha256": sha256_file(path) if path.exists() and path.is_file() else None,
        }
        for key, path in source_paths.items()
    }
    required_false = {
        "promotion_evidence": False,
        "external_live_llm_run": False,
        "optional_judge_run": False,
        "retrieval_tuning_run": False,
        "reranking_run": False,
        "parser_expansion_run": False,
        "threshold_relaxation_run": False,
        "broad_indexing_run": False,
        "db_mutation_run": False,
        "searchunit_mutation_run": False,
        "immutable_baseline_changed": False,
        "candidate_artifact_changed": False,
        "existing_gold_csv_overwritten": False,
    }
    manifest: dict[str, Any] = {
        "schema_version": SCHEMA_VERSION,
        "run_id": run_id,
        "generated_at": generated_at,
        "status": "PASS_WITH_WARNINGS" if warnings else "PASS",
        "artifact_dir": repo_relative(artifact_dir),
        "manifest_path": repo_relative(manifest_path),
        "answer_generation_inputs_path": repo_relative(inputs_path),
        "answer_generation_inputs_with_xlsx_content_path": repo_relative(enriched_inputs_alias_path)
        if enriched_inputs_alias_path
        else "",
        "answer_generation_inputs_sha256": sha256_file(inputs_path),
        "answer_generation_inputs_with_xlsx_content_sha256": sha256_file(enriched_inputs_alias_path)
        if enriched_inputs_alias_path and enriched_inputs_alias_path.exists()
        else "",
        "promotion_evidence": False,
        "evidence_role": "diagnostic",
        "local_llm_run": False,
        "external_live_llm_run": False,
        "optional_judge_run": False,
        "dry_run_preview_used_as_actual_answer": False,
        "actual_answer_output_missing": True,
        "input_row_count": len(rows),
        "row_count_by_track": dict(track_counts),
        "expected_answer_shape_counts": dict(shape_counts),
        "context_available_count": context_available_count,
        "context_expected_terms_count": context_expected_terms_count,
        "xlsx_answer_eval_denominator": 0,
        "pdf_answer_eval_denominator": 0,
        "xlsx_retrieval_diagnostic_preserved": True,
        "pdf_policy_pending": True,
        "keyword_echo_only_is_failure": True,
        "location_only_answer_is_failure": True,
        "r8_or_citation_support_blocked_until_answer_shape_alignment": True,
        "xlsx_searchunit_content_join": {
            "enabled": parse_bool(xlsx_searchunit_index.get("enabled")),
            "status": clean(xlsx_searchunit_index.get("status")),
            "error": clean(xlsx_searchunit_index.get("error")),
            "selected_search_unit_id_count": xlsx_searchunit_index.get("selected_search_unit_id_count", 0),
            "loaded_search_unit_count": xlsx_searchunit_index.get("loaded_search_unit_count", 0),
            "join_policy": "selected_hit_exact_id_or_safe_sheet_range_overlap_no_sheet_or_workbook_broad_fallback",
            "source_workbook_promoted_evidence_count": 0,
            "gold_leakage_count": 0,
            "broad_fallback_promoted_evidence_count": 0,
        },
        "source_review_paths": {
            "pdf": repo_relative(source_paths["pdf_review"]),
            "xlsx": repo_relative(source_paths["xlsx_review_snapshot"])
            if xlsx_external_used
            else repo_relative(source_paths["xlsx_review"]),
            "xlsx_snapshot": repo_relative(source_paths["xlsx_review_snapshot"]),
            "xlsx_source_kind": "external_user_supplied_snapshot" if xlsx_external_used else "repo_local",
            "xlsx_external_path_redacted": xlsx_external_used,
            "xlsx_external_basename": xlsx_external_basename,
            "xlsx_repo_local_fallback": repo_relative(source_paths["xlsx_repo_local_fallback"]),
            "xlsx_external_used": xlsx_external_used,
        },
        "source_inventory": source_inventory,
        "plan_source_status": plan_payload.get("status"),
        "plan_actual_answer_output_missing": plan_payload.get("actual_answer_output_missing"),
        "plan_assertions": plan_payload.get("assertions", {}),
        "guardrails": required_false,
        **required_false,
        "warnings": warnings,
    }
    return manifest


def build_dataset_index(root: Path) -> dict[str, list[Path]]:
    index: dict[str, list[Path]] = {}
    if not root.exists():
        return index
    for suffix in ("*.xlsx", "*.xls", "*.pdf"):
        for path in root.rglob(suffix):
            index.setdefault(path.name, []).append(path)
    return index


def resolve_dataset_file(file_name: str, dataset_index: Mapping[str, list[Path]]) -> Path | None:
    if not file_name:
        return None
    candidates = dataset_index.get(file_name) or []
    return candidates[0] if candidates else None


def merge_row(*rows: Mapping[str, Any]) -> dict[str, Any]:
    merged: dict[str, Any] = {}
    for row in rows:
        for key, value in row.items():
            if value not in (None, "") or key not in merged:
                merged[key] = value
    return merged


def parse_locator(text: str) -> dict[str, Any]:
    locator: dict[str, Any] = {}
    for part in text.split(";"):
        if "=" not in part:
            continue
        key, value = part.split("=", 1)
        key = clean(key)
        value = clean(value)
        if key == "bbox":
            locator[key] = bbox_list(value) or value
        else:
            locator[key] = value
    return locator


def locator_from_review_row(row: Mapping[str, Any]) -> dict[str, Any]:
    return compact_dict(
        {
            "file": row.get("expected_file_name"),
            "docv": row.get("expected_document_version_id"),
            "page": row.get("expected_page_no"),
            "physical_page_index": row.get("expected_physical_page_index"),
            "page_label": row.get("expected_page_label"),
            "sheet": row.get("expected_sheet_name"),
            "range": row.get("expected_cell_range"),
            "table": row.get("expected_table_id"),
            "bbox": bbox_list(row.get("expected_bbox")) or clean(row.get("expected_bbox")),
        }
    )


def keyed_by_query_id(rows: Iterable[Mapping[str, Any]]) -> dict[str, dict[str, Any]]:
    return {clean(row.get("query_id")): dict(row) for row in rows if clean(row.get("query_id"))}


def keyed_json_rows(payload: Mapping[str, Any] | None, field: str) -> dict[str, dict[str, Any]]:
    if not payload:
        return {}
    rows = payload.get(field)
    if not isinstance(rows, list):
        return {}
    return keyed_by_query_id(row for row in rows if isinstance(row, Mapping))


def read_json_object(path: Path) -> dict[str, Any] | None:
    if not path.exists():
        return None
    try:
        payload = json.loads(path.read_text(encoding="utf-8"))
    except json.JSONDecodeError:
        return None
    return payload if isinstance(payload, dict) else None


def read_csv_rows(path: Path) -> list[dict[str, str]]:
    if not path.exists():
        return []
    with path.open("r", encoding="utf-8-sig", newline="") as handle:
        return [dict(row) for row in csv.DictReader(handle)]


def write_json(path: Path, payload: Mapping[str, Any]) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    path.write_text(json.dumps(payload, ensure_ascii=False, indent=2) + "\n", encoding="utf-8")


def write_jsonl(path: Path, rows: Iterable[Mapping[str, Any]]) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    with path.open("w", encoding="utf-8", newline="\n") as handle:
        for row in rows:
            handle.write(json.dumps(row, ensure_ascii=False, sort_keys=True) + "\n")


def print_json(payload: Mapping[str, Any]) -> None:
    print(json.dumps(payload, ensure_ascii=False, indent=2))


def sha256_file(path: Path) -> str:
    digest = hashlib.sha256()
    with path.open("rb") as handle:
        for chunk in iter(lambda: handle.read(1024 * 1024), b""):
            digest.update(chunk)
    return digest.hexdigest()


def split_terms(value: object) -> list[str]:
    text = clean(value)
    if not text:
        return []
    terms = [clean(item) for item in re.split(r"[;|]", text)]
    return [term for term in terms if term]


def has_any_term(text: str, terms: Iterable[str]) -> bool:
    folded = normalize_for_match(text)
    return any(normalize_for_match(term) in folded for term in terms if clean(term))


def select_sentences(text: str, terms: Iterable[str]) -> list[str]:
    sentences = [clean(item) for item in re.split(r"(?<=[.!?。！？])\s+|\n+", text) if clean(item)]
    selected = [sentence for sentence in sentences if has_any_term(sentence, terms)]
    if selected:
        return selected
    folded_terms = [normalize_for_match(term) for term in terms if clean(term)]
    fallback = []
    for sentence in sentences:
        folded = normalize_for_match(sentence)
        if any(term and term in folded for term in folded_terms):
            fallback.append(sentence)
    return fallback


def normalize_for_match(value: object) -> str:
    return re.sub(r"\s+", "", clean(value)).lower()


def clean_whitespace(value: object) -> str:
    return re.sub(r"\s+", " ", clean(value)).strip()


def cell_to_text(value: object) -> str:
    if value is None:
        return ""
    return clean(value)


def first_nonempty(values: Iterable[str]) -> str:
    for value in values:
        if clean(value):
            return clean(value)
    return ""


def bbox_list(value: object) -> list[float]:
    if isinstance(value, list):
        try:
            return [float(item) for item in value]
        except (TypeError, ValueError):
            return []
    text = clean(value)
    if not text:
        return []
    numbers = re.findall(r"-?\d+(?:\.\d+)?", text)
    if len(numbers) != 4:
        return []
    return [float(number) for number in numbers]


def int_or_none(value: object) -> int | None:
    text = clean(value)
    if not text:
        return None
    try:
        return int(float(text))
    except ValueError:
        return None


def parse_bool(value: object) -> bool:
    if isinstance(value, bool):
        return value
    return clean(value).lower() in {"1", "true", "yes", "y"}


def compact_dict(payload: Mapping[str, Any]) -> dict[str, Any]:
    return {key: value for key, value in payload.items() if value not in (None, "", [], {})}


def nested_bool(row: Mapping[str, Any], path: list[str]) -> bool:
    current: Any = row
    for key in path:
        if not isinstance(current, Mapping):
            return False
        current = current.get(key)
    return bool(current)


def clean(value: object) -> str:
    if value is None:
        return ""
    if isinstance(value, str):
        return value.strip()
    return str(value).strip()


def is_under_repo(path: Path) -> bool:
    try:
        path.resolve().relative_to(REPO_ROOT.resolve())
        return True
    except ValueError:
        return False


def repo_relative(path: Path) -> str:
    try:
        return str(path.resolve().relative_to(REPO_ROOT.resolve())).replace("\\", "/")
    except ValueError:
        return str(path)


def utc_timestamp() -> str:
    return datetime.now(timezone.utc).isoformat()


def utc_run_id() -> str:
    return datetime.now(timezone.utc).strftime("%Y%m%dT%H%M%SZ")


if __name__ == "__main__":
    sys.exit(main())
