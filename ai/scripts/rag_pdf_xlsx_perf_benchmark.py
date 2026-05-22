"""Synthetic PDF/XLSX ingestion and SearchUnit indexing performance smoke.

The fixtures are generated in memory so this command never reads or mutates
gold, silver, qrels, labels, namespaces, or production indexes.
"""

from __future__ import annotations

import argparse
import gc
import json
import platform
import statistics
import sys
import time
import tracemalloc
from dataclasses import replace
from datetime import datetime, timezone
from io import BytesIO
from pathlib import Path
from types import SimpleNamespace
from typing import Any, Callable

import numpy as np
from openpyxl import Workbook

AI_ROOT = Path(__file__).resolve().parents[1]
REPO_ROOT = AI_ROOT.parent
if str(AI_ROOT) not in sys.path:
    sys.path.insert(0, str(AI_ROOT))

from app.capabilities.base import CapabilityInput, CapabilityInputArtifact
from app.capabilities.pdf.service import PdfExtractService
from app.capabilities.rag.faiss_index import IndexBuildInfo
from app.capabilities.rag.metadata_store import ChunkRow
from app.capabilities.rag.search_unit_indexing import (
    SearchUnitIndexDocument,
    SearchUnitVectorIndexer,
    build_search_unit_embedding_text,
    to_chunk_row,
)
from app.capabilities.xlsx.service import XlsxExtractService

REPORT_DIR = AI_ROOT / "eval" / "reports" / "rag-ingestion"
DEFAULT_OUTPUT = REPORT_DIR / "perf" / "pdf_xlsx_perf_benchmark_current.json"


class FakeOcrProvider:
    def extract(
        self,
        image_bytes: bytes,
        *,
        source_record_id: str,
        pipeline_version: str,
        content_type: str,
        filename: str,
    ) -> SimpleNamespace:
        del image_bytes, source_record_id, pipeline_version, content_type, filename
        block = SimpleNamespace(
            text="synthetic OCR text",
            bbox=[10.0, 10.0, 200.0, 40.0],
            confidence=0.93,
        )
        return SimpleNamespace(pages=[SimpleNamespace(blocks=[block])])


class FakeEmbedder:
    model_name = "synthetic-embedding-v1"
    dimension = 8

    def __init__(self) -> None:
        self.calls: list[int] = []

    def embed_passages(self, texts: list[str]) -> np.ndarray:
        self.calls.append(len(texts))
        if not texts:
            return np.empty((0, self.dimension), dtype=np.float32)
        vectors = np.ones((len(texts), self.dimension), dtype=np.float32)
        vectors /= np.linalg.norm(vectors, axis=1, keepdims=True)
        return vectors

    def embed_queries(self, texts: list[str]) -> np.ndarray:
        return self.embed_passages(texts)


class FakeMetadataStore:
    def __init__(self, chunks: list[ChunkRow]) -> None:
        self._chunks = chunks

    def list_chunks(self, index_version: str) -> list[ChunkRow]:
        return [chunk for chunk in self._chunks if chunk.index_version == index_version]

    def upsert_index_rows(self, *, documents: list[Any], chunks: list[ChunkRow]) -> None:
        del documents, chunks

    def record_index_build(self, **kwargs: Any) -> None:
        del kwargs


class FakeIndex:
    def __init__(self, *, index_version: str, chunks: list[ChunkRow], dimension: int) -> None:
        self._info = IndexBuildInfo(
            index_version=index_version,
            embedding_model=FakeEmbedder.model_name,
            dimension=dimension,
            chunk_count=len(chunks),
        )
        self._vectors = np.ones((len(chunks), dimension), dtype=np.float32)
        self.index_dir = REPORT_DIR / "_synthetic_index_not_written"

    @property
    def info(self) -> IndexBuildInfo:
        return self._info

    def load(self) -> IndexBuildInfo:
        return self._info

    def vectors(self) -> np.ndarray:
        return self._vectors

    def build_staged(self, vectors: np.ndarray, *, index_version: str, embedding_model: str) -> tuple[IndexBuildInfo, Path]:
        raise AssertionError("duplicate benchmark should not rebuild the FAISS index")

    def promote_staged(self, stage_dir: Path, info: IndexBuildInfo, *, extra_files: tuple[str, ...] = ()) -> None:
        del stage_dir, info, extra_files

    def discard_staged(self, stage_dir: Path) -> None:
        del stage_dir


def synthetic_text_pdf(*, pages: int, blocks_per_page: int) -> bytes:
    import fitz

    document = fitz.open()
    try:
        for page_index in range(pages):
            page = document.new_page(width=595, height=842)
            for block_index in range(blocks_per_page):
                x = 45 + (block_index % 3) * 165
                y = 45 + (block_index // 3) * 22
                page.insert_text(
                    (x, y),
                    f"plain paragraph p{page_index + 1}-{block_index + 1} value {block_index}",
                    fontsize=8,
                )
        return document.tobytes()
    finally:
        document.close()


def synthetic_blank_pdf(*, pages: int) -> bytes:
    import fitz

    document = fitz.open()
    try:
        for _ in range(pages):
            document.new_page(width=595, height=842)
        return document.tobytes()
    finally:
        document.close()


def synthetic_xlsx_large_merged_range(*, rows: int, columns: int) -> bytes:
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "MergedRange"
    sheet.cell(row=1, column=1, value="merged heading")
    sheet.merge_cells(start_row=1, start_column=1, end_row=rows, end_column=columns)
    for row in range(rows + 1, rows + 41):
        sheet.cell(row=row, column=1, value=f"item-{row}")
        sheet.cell(row=row, column=2, value=row)
        sheet.cell(row=row, column=3, value=row * 1.5)
    stream = BytesIO()
    workbook.save(stream)
    workbook.close()
    return stream.getvalue()


def capability_input(
    *,
    artifact_id: str,
    capability: str,
    content: bytes,
    filename: str,
    content_type: str,
) -> CapabilityInput:
    return CapabilityInput(
        job_id=f"benchmark-{artifact_id}",
        capability=capability,
        attempt_no=1,
        inputs=[
            CapabilityInputArtifact(
                artifact_id=artifact_id,
                type="INPUT_FILE",
                content=content,
                content_type=content_type,
                filename=filename,
                source_file_id=f"source-{artifact_id}",
            )
        ],
    )


def synthetic_search_unit_docs(count: int) -> list[SearchUnitIndexDocument]:
    docs: list[SearchUnitIndexDocument] = []
    for index in range(count):
        source_file_id = f"source-{index % 17:02d}"
        unit_type = "ROW_GROUP" if index % 2 else "PAGE"
        file_type = "xlsx" if unit_type == "ROW_GROUP" else "pdf"
        docs.append(
            SearchUnitIndexDocument(
                search_unit_id=f"su-{index:05d}",
                claim_token=f"claim-{index:05d}",
                index_id=f"source_file:{source_file_id}:unit:{unit_type}:bench-{index:05d}",
                source_file_id=source_file_id,
                source_file_name=f"benchmark.{file_type}",
                extracted_artifact_id=f"artifact-{index:05d}",
                artifact_type="XLSX_WORKBOOK_JSON" if file_type == "xlsx" else "PDF_PARSED_JSON",
                unit_type=unit_type,
                unit_key=f"bench-{index:05d}",
                title=f"Benchmark {index}",
                section_path="Sheet1" if file_type == "xlsx" else "Page",
                page_start=index + 1 if file_type == "pdf" else None,
                page_end=index + 1 if file_type == "pdf" else None,
                text_content=f"benchmark content {index}",
                content_sha256=f"sha-{index:05d}",
                metadata_json={
                    "fileType": file_type,
                    "sheetName": "Sheet1" if file_type == "xlsx" else None,
                    "cellRange": f"A{index + 1}:D{index + 1}" if file_type == "xlsx" else None,
                    "sourceAtomId": f"atom-{index:05d}",
                    "sourceRegistryVersion": "synthetic-v1",
                },
                index_metadata={
                    "diagnostic_only": True,
                    "namespace": "synthetic-nonprod",
                },
            )
        )
    return docs


def existing_chunks_for_docs(
    docs: list[SearchUnitIndexDocument],
    *,
    embedder: FakeEmbedder,
    index_version: str,
) -> list[ChunkRow]:
    chunks: list[ChunkRow] = []
    for row_id, doc in enumerate(docs):
        embedding_text = build_search_unit_embedding_text(doc)
        chunks.append(
            to_chunk_row(
                doc,
                faiss_row_id=row_id,
                index_version=index_version,
                embedding_model=embedder.model_name,
                embedding_text=embedding_text,
            )
        )
    return chunks


def bench_pdf_native_text(pdf_bytes: bytes) -> dict[str, Any]:
    service = PdfExtractService()
    output = service.run(
        capability_input(
            artifact_id="pdf-native-text",
            capability="PDF_EXTRACT",
            content=pdf_bytes,
            filename="synthetic-text.pdf",
            content_type="application/pdf",
        )
    )
    return {"outputs": len(output.outputs)}


def bench_pdf_ocr_fallback(pdf_bytes: bytes) -> dict[str, Any]:
    service = PdfExtractService(
        ocr_fallback_enabled=True,
        ocr_pdf_dpi=110,
        min_native_chars=1,
        ocr_provider=FakeOcrProvider(),
    )
    output = service.run(
        capability_input(
            artifact_id="pdf-ocr-fallback",
            capability="PDF_EXTRACT",
            content=pdf_bytes,
            filename="synthetic-blank.pdf",
            content_type="application/pdf",
        )
    )
    return {"outputs": len(output.outputs)}


def bench_xlsx_large_merged_range(xlsx_bytes: bytes) -> dict[str, Any]:
    service = XlsxExtractService()
    output = service.run(
        capability_input(
            artifact_id="xlsx-large-merged-range",
            capability="XLSX_EXTRACT",
            content=xlsx_bytes,
            filename="synthetic-merged.xlsx",
            content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )
    )
    return {"outputs": len(output.outputs)}


def bench_search_unit_duplicate_skip(docs: list[SearchUnitIndexDocument], chunks: list[ChunkRow]) -> dict[str, Any]:
    embedder = FakeEmbedder()
    index_version = chunks[0].index_version if chunks else "synthetic-search-unit-v1"
    indexer = SearchUnitVectorIndexer(
        embedder=embedder,
        metadata_store=FakeMetadataStore(chunks),  # type: ignore[arg-type]
        index=FakeIndex(index_version=index_version, chunks=chunks, dimension=embedder.dimension),  # type: ignore[arg-type]
        index_version=index_version,
    )
    result = indexer.index_documents([replace(doc) for doc in docs])
    return {
        "indexed": len(result.indexed),
        "embedder_calls": list(embedder.calls),
    }


def measure_case(
    name: str,
    func: Callable[[], dict[str, Any]],
    *,
    warmups: int,
    iterations: int,
) -> dict[str, Any]:
    for _ in range(warmups):
        func()
    samples: list[dict[str, Any]] = []
    for _ in range(iterations):
        gc.collect()
        tracemalloc.start()
        start = time.perf_counter()
        extra = func()
        elapsed_ms = (time.perf_counter() - start) * 1000.0
        _current, peak = tracemalloc.get_traced_memory()
        tracemalloc.stop()
        samples.append({
            "elapsed_ms": round(elapsed_ms, 3),
            "peak_tracemalloc_kib": round(peak / 1024, 1),
            "extra": extra,
        })
    elapsed = [sample["elapsed_ms"] for sample in samples]
    peaks = [sample["peak_tracemalloc_kib"] for sample in samples]
    return {
        "case": name,
        "iterations": iterations,
        "warmups": warmups,
        "elapsed_ms": {
            "median": round(statistics.median(elapsed), 3),
            "min": round(min(elapsed), 3),
            "max": round(max(elapsed), 3),
        },
        "peak_tracemalloc_kib": {
            "median": round(statistics.median(peaks), 1),
            "min": round(min(peaks), 1),
            "max": round(max(peaks), 1),
        },
        "samples": samples,
    }


def run_benchmark(*, label: str, warmups: int, iterations: int) -> dict[str, Any]:
    pdf_text = synthetic_text_pdf(pages=8, blocks_per_page=72)
    pdf_blank = synthetic_blank_pdf(pages=6)
    xlsx_merged = synthetic_xlsx_large_merged_range(rows=1000, columns=1000)
    docs = synthetic_search_unit_docs(count=1800)
    existing_embedder = FakeEmbedder()
    chunks = existing_chunks_for_docs(
        docs,
        embedder=existing_embedder,
        index_version="synthetic-search-unit-v1",
    )

    cases = [
        measure_case(
            "pdf_native_text_no_supported_tables",
            lambda: bench_pdf_native_text(pdf_text),
            warmups=warmups,
            iterations=iterations,
        ),
        measure_case(
            "pdf_ocr_fallback_blank_pages",
            lambda: bench_pdf_ocr_fallback(pdf_blank),
            warmups=warmups,
            iterations=iterations,
        ),
        measure_case(
            "xlsx_large_merged_range",
            lambda: bench_xlsx_large_merged_range(xlsx_merged),
            warmups=warmups,
            iterations=iterations,
        ),
        measure_case(
            "search_unit_duplicate_skip",
            lambda: bench_search_unit_duplicate_skip(docs, chunks),
            warmups=warmups,
            iterations=iterations,
        ),
    ]
    return {
        "schema_version": "rag_pdf_xlsx_perf_benchmark_v1",
        "label": label,
        "generated_at": datetime.now(timezone.utc).isoformat(),
        "environment": {
            "python": sys.version,
            "platform": platform.platform(),
            "cwd": str(Path.cwd()),
        },
        "inputs": {
            "pdf_native_text": {
                "bytes": len(pdf_text),
                "pages": 8,
                "blocks_per_page": 72,
            },
            "pdf_ocr_fallback": {
                "bytes": len(pdf_blank),
                "pages": 6,
                "dpi": 110,
                "ocr_provider": "FakeOcrProvider",
            },
            "xlsx_large_merged_range": {
                "bytes": len(xlsx_merged),
                "sheets": 1,
                "merged_range": "A1:ALL1000",
                "rows_with_values_after_merge": 40,
            },
            "search_unit_duplicate_skip": {
                "documents": len(docs),
                "existing_chunks": len(chunks),
                "embedding_model": FakeEmbedder.model_name,
                "namespace": "synthetic-nonprod",
                "diagnostic_only": True,
            },
        },
        "policy": {
            "gold_or_qrels_read": False,
            "gold_or_qrels_written": False,
            "official_denominator_mutated": False,
            "production_index_written": False,
            "synthetic_inputs_only": True,
            "diagnostic_only": True,
        },
        "cases": cases,
    }


def parse_args(argv: list[str] | None = None) -> argparse.Namespace:
    parser = argparse.ArgumentParser(description=__doc__)
    parser.add_argument("--label", default="current", help="Label to store in the report.")
    parser.add_argument("--warmups", type=int, default=1)
    parser.add_argument("--iterations", type=int, default=3)
    parser.add_argument("--output", type=Path, default=DEFAULT_OUTPUT)
    return parser.parse_args(argv)


def main(argv: list[str] | None = None) -> int:
    args = parse_args(argv)
    if args.iterations < 1:
        raise SystemExit("--iterations must be >= 1")
    if args.warmups < 0:
        raise SystemExit("--warmups must be >= 0")
    report = run_benchmark(
        label=str(args.label),
        warmups=int(args.warmups),
        iterations=int(args.iterations),
    )
    output = Path(args.output)
    output.parent.mkdir(parents=True, exist_ok=True)
    output.write_text(json.dumps(report, ensure_ascii=False, indent=2), encoding="utf-8")
    print(json.dumps({
        "label": report["label"],
        "output": str(output),
        "cases": {
            case["case"]: case["elapsed_ms"]["median"]
            for case in report["cases"]
        },
    }, ensure_ascii=False, indent=2))
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
