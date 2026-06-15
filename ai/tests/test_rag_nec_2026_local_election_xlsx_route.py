from __future__ import annotations

import csv
import hashlib
import json
from pathlib import Path

import pytest
from openpyxl import Workbook

from ai.eval import rag_eval_registry as registry
from ai.eval import rag_nec_2026_local_election_xlsx_source_collection as nec
import ai.scripts.rag_eval as runner


ROOT = Path(__file__).resolve().parents[2]


def _sha256_file(path: Path) -> str:
    hasher = hashlib.sha256()
    with path.open("rb") as handle:
        for chunk in iter(lambda: handle.read(1024 * 1024), b""):
            hasher.update(chunk)
    return hasher.hexdigest()


def _write_fixture_collection(tmp_path: Path) -> Path:
    collection = tmp_path / "source_collection_20260605_nec_election_results"
    xlsx_dir = collection / "xlsx"
    xlsx_dir.mkdir(parents=True)
    workbook_path = xlsx_dir / "nec_0020260603_VCCP09_4_구시군의_장선거_nationwide_results.xlsx"

    wb = Workbook()
    ws = wb.active
    ws.title = "source_requests"
    ws.append(["request_id", "election_code", "election_label", "city_code", "city_name"])
    ws.append(["4_city_1100", "4", "구시군의 장선거", "1100", "서울특별시"])
    ws.append(["4_city_2600", "4", "구시군의 장선거", "2600", "부산광역시"])

    raw = wb.create_sheet("raw_display_rows")
    raw.append(["request_id", "c01", "c02", "c03"])
    raw.append(["4_city_1100", "종로구", "더불어민주당", "100"])
    raw.append(["4_city_1100", "종로구", "국민의힘", "90"])
    raw.append(["4_city_2600", "중구", "국민의힘", "80"])
    raw.append(["4_city_2600", "중구", "더불어민주당", "70"])

    parsed = wb.create_sheet("parsed_votes")
    parsed.append(
        [
            "request_id",
            "election_code",
            "election_label",
            "city_code",
            "city_name",
            "district_code",
            "district_name",
            "contest_label",
            "candidate_no",
            "candidate_name",
            "party_name",
            "party_group",
            "votes",
            "vote_share_pct",
            "rank",
            "is_winner",
            "source_row_index",
            "raw_label",
            "raw_party",
            "raw_votes",
            "raw_vote_share_pct",
            "note",
        ]
    )
    parsed.append(["4_city_1100", "4", "구시군의 장선거", "1100", "서울", "1111", "종로구", "종로구 / 종로구", "1", "김A", "더불어민주당", "더불어민주당", 100, 52.6, 1, True, 2, "", "", "", "", ""])
    parsed.append(["4_city_1100", "4", "구시군의 장선거", "1100", "서울", "1111", "종로구", "종로구 / 종로구", "2", "박B", "국민의힘", "국민의힘", 90, 47.4, 2, False, 3, "", "", "", "", ""])
    parsed.append(["4_city_2600", "4", "구시군의 장선거", "2600", "부산", "2611", "중구", "중구 / 중구", "1", "이C", "국민의힘", "국민의힘", 80, 53.3, 1, True, 4, "", "", "", "", ""])
    parsed.append(["4_city_2600", "4", "구시군의 장선거", "2600", "부산", "2611", "중구", "중구 / 중구", "2", "최D", "더불어민주당", "더불어민주당", 70, 46.7, 2, False, 5, "", "", "", "", ""])

    summary = wb.create_sheet("national_summary")
    summary.append(["metric", "value"])
    summary.append(["parsed_vote_rows", 4])
    summary.append(["contest_count", 2])
    wb.save(workbook_path)

    digest = _sha256_file(workbook_path)
    manifest_rows = [
        {
            "election_code": "4",
            "election_label": "구시군의 장선거",
            "xlsx_path": str(workbook_path),
            "xlsx_sha256": digest,
            "source_request_count": "2",
            "parsed_vote_rows": "4",
            "contest_count": "2",
        }
    ]
    with (collection / "manifest.csv").open("w", encoding="utf-8-sig", newline="") as handle:
        writer = csv.DictWriter(handle, fieldnames=list(manifest_rows[0]))
        writer.writeheader()
        writer.writerows(manifest_rows)
    (collection / "manifest.json").write_text(
        json.dumps({"files": manifest_rows}, ensure_ascii=False, indent=2),
        encoding="utf-8",
    )
    return collection


def test_nec_2026_local_election_route_builds_xlsx_search_unit_preview(tmp_path: Path) -> None:
    collection = _write_fixture_collection(tmp_path)

    report = nec.build_report(
        root=ROOT,
        source_collection_root=collection,
        generated_at="2026-06-05T00:00:00Z",
    )

    assert report["logical_run_key"] == "nec_2026_local_election_xlsx"
    assert report["status"] == nec.STATUS
    assert report["diagnostic_only"] is True
    assert report["official_metric_input_rows"] == 0
    assert report["denominator_mutation"] is False
    assert report["live_db_index_cache_readiness"] is False
    assert report["source_collection"]["workbook_count"] == 1
    assert report["source_collection"]["verified_xlsx_count"] == 1
    assert report["source_collection"]["visible_sheet_count"] == 4
    assert report["synthetic_chunk_summary"]["parsed_votes_contest_span_count"] == 2
    assert report["synthetic_chunk_summary"]["raw_display_request_block_count"] == 2
    assert report["synthetic_chunk_summary"]["source_request_chunk_count"] == 2
    assert report["synthetic_chunk_summary"]["search_unit_preview_rows"] == 7
    assert report["source_atom_summary"]["source_atom_rows"] == 6
    assert report["source_atom_summary"]["parsed_votes_atom_rows"] == 4
    assert report["source_atom_summary"]["national_summary_atom_rows"] == 2
    assert report["source_atom_summary"]["retrieval_default_excluded_sheets"] == ["source_requests", "raw_display_rows"]
    assert report["search_view_summary"]["search_view_rows"] == 6

    serialized = json.dumps(report, ensure_ascii=False)
    assert str(collection) not in serialized
    assert "xlsx_path" not in report["source_collection"]["workbooks"][0]

    parsed_units = [
        unit
        for unit in report["search_unit_preview_sample"]
        if unit["unit_type"] == "parsed_votes_contest_span"
    ]
    assert parsed_units
    assert parsed_units[0]["citation"] == {
        "source_family": "XLSX",
        "workbook": "nec_0020260603_VCCP09_4_구시군의_장선거_nationwide_results.xlsx",
        "sheet": "parsed_votes",
        "range": "A2:V3",
        "search_unit_id": parsed_units[0]["search_unit_id"],
    }
    assert "종로구 / 종로구" in parsed_units[0]["embedding_text"]

    atom = report["source_atom_preview_sample"][0]
    assert atom["sheet"] == "parsed_votes"
    assert atom["raw_locator"]["row_label"] == "종로구 / 종로구"
    assert atom["raw_locator"]["target_column"] == "votes"
    assert atom["canonical_citation_payload"]["workbook"] == "nec_0020260603_VCCP09_4_구시군의_장선거_nationwide_results.xlsx"
    assert atom["canonical_citation_payload"]["sheet"] == "parsed_votes"
    assert atom["canonical_citation_payload"]["search_unit_id"]
    assert atom["xlsx_display_metadata"]["display_value"] == "100"

    view = report["search_view_preview_sample"][0]
    assert view["source_atom_ids"] == [atom["source_atom_id"]]
    assert "row_label=종로구 / 종로구" in view["bm25_text"]
    assert "target_column=votes" in view["embedding_text"]


def test_nec_2026_local_election_route_is_direct_and_does_not_move_current(tmp_path: Path, monkeypatch: pytest.MonkeyPatch) -> None:
    collection = _write_fixture_collection(tmp_path)
    monkeypatch.setenv("RAG_NEC_2026_LOCAL_ELECTION_SOURCE_COLLECTION_ROOT", str(collection))

    resolved = registry.resolve_run("nec_2026_local_election_xlsx", root=ROOT)
    assert resolved.report_path == ROOT / "reports/rag_eval/rag-ingestion/runs/nec_2026_local_election_xlsx/report.json"
    assert registry.resolve_run("current", root=ROOT).logical_key == "v5_6"
    assert runner.DEFAULT_RUN_KEY == "v5_6"

    report = runner.check_run("nec_2026_local_election_xlsx")
    assert report["logical_run_key"] == "nec_2026_local_election_xlsx"
    assert report["current_resolves_to"] == "v5_6"
    assert report["official_metric_input_rows"] == 0
    assert report["protected_namespaces_touched"] == []


@pytest.mark.parametrize("flag", nec.FORBIDDEN_FALSE_KEYS)
def test_nec_2026_local_election_route_rejects_protected_gate_drift(tmp_path: Path, flag: str) -> None:
    report = nec.build_report(root=ROOT, source_collection_root=_write_fixture_collection(tmp_path))
    bad = dict(report)
    bad[flag] = True

    with pytest.raises(ValueError, match=f"opened forbidden gate: {flag}"):
        nec.check_report(bad)
