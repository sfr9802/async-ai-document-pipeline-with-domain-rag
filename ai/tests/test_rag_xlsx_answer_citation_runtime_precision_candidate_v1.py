from __future__ import annotations

import importlib.util
import json
import sys
from io import BytesIO
from pathlib import Path

from openpyxl import Workbook

ROOT = Path(__file__).resolve().parents[2]
sys.path.insert(0, str(ROOT / "ai"))

from app.capabilities.base import CapabilityInput, CapabilityInputArtifact
from app.capabilities.xlsx import artifact_builder as xlsx_artifacts
from app.capabilities.xlsx import service as xlsx_service

SCRIPT_PATH = ROOT / "ai" / "scripts" / "rag_xlsx_answer_citation_runtime_precision_candidate_v1.py"
REPORT_DIR = ROOT / "reports" / "rag_eval" / "rag-ingestion"
REPORT_ARCHIVE_DIR = REPORT_DIR / "_archive" / "legacy"


def windows_long_path(path: Path) -> Path:
    if sys.platform != "win32":
        return path
    path_text = str(path)
    if path_text.startswith("\\\\?\\"):
        return path
    if path.is_absolute():
        return Path("\\\\?\\" + path_text)
    return path


EXTERNAL_REPORT_ARCHIVE_DIRS = (
    windows_long_path(Path(
        "D:/_external_runtime_artifacts/async-ocr-rag-multimodal-pipeline/"
        "rag-ingestion/repo-wide-cleanup-20260521/reports/rag-ingestion-legacy"
    )),
    windows_long_path(Path(
        "D:/_external_runtime_artifacts/async-ocr-rag-multimodal-pipeline/"
        "rag-ingestion/repo-wide-cleanup-20260519/reports/rag-ingestion-legacy"
    )),
)


def resolve_report_artifact_path(path: Path) -> Path:
    if path.exists():
        return path
    if path.parent == REPORT_DIR:
        for archive_dir in EXTERNAL_REPORT_ARCHIVE_DIRS:
            archived = archive_dir / path.name
            if archived.exists():
                return archived
        legacy = REPORT_ARCHIVE_DIR / path.name
        if legacy.exists():
            return legacy
    return path


def load_module():
    spec = importlib.util.spec_from_file_location("rag_xlsx_answer_citation_runtime_precision_candidate_v1_for_tests", SCRIPT_PATH)
    assert spec is not None
    module = importlib.util.module_from_spec(spec)
    assert spec.loader is not None
    sys.modules[spec.name] = module
    spec.loader.exec_module(module)
    return module


def test_xlsx_extract_uses_ooxml_metadata_for_large_merged_ranges(monkeypatch) -> None:
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "MergedRange"
    sheet.cell(row=1, column=1, value="merged heading")
    sheet.merge_cells(start_row=1, start_column=1, end_row=1000, end_column=1000)
    sheet.cell(row=1001, column=1, value="visible row")
    sheet.cell(row=1001, column=2, value=42)
    stream = BytesIO()
    workbook.save(stream)
    workbook.close()

    original_load_workbook = xlsx_service.load_workbook

    def guarded_load_workbook(*args, **kwargs):
        if kwargs.get("read_only") is False:
            raise AssertionError("full metadata workbook load should be skipped")
        return original_load_workbook(*args, **kwargs)

    monkeypatch.setattr(xlsx_service, "load_workbook", guarded_load_workbook)

    output = xlsx_service.XlsxExtractService().run(
        CapabilityInput(
            job_id="xlsx-large-merged-range",
            capability="XLSX_EXTRACT",
            attempt_no=1,
            inputs=[
                CapabilityInputArtifact(
                    artifact_id="xlsx-large-merged-range",
                    type="INPUT_FILE",
                    content=stream.getvalue(),
                    content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                    filename="large-merged.xlsx",
                    source_file_id="source-large-merged",
                )
            ],
        )
    )

    workbook_json = next(
        artifact
        for artifact in output.outputs
        if artifact.type == xlsx_artifacts.XLSX_WORKBOOK_JSON
    )
    payload = json.loads(workbook_json.content.decode("utf-8"))
    sheet_payload = payload["workbook"]["sheets"][0]

    assert sheet_payload["indexable"] is True
    assert "A1:ALL1000" in sheet_payload["mergedCells"]
    assert sheet_payload["rowCount"] == 2
    assert sheet_payload["cells"][0]["mergedCell"] is True
    assert not any("hidden row/column/cell metadata" in warning for warning in sheet_payload["warnings"])


def test_runtime_candidate_generates_single_row_locator_and_target_answer_from_question_conditions():
    module = load_module()

    candidate = module.generate_xlsx_runtime_candidate(
        runtime_input(
            query_id="gq_auto_012",
            question="2019년 2월 5호선의 승차총승객수는 몇 명입니까?",
            citation_text=(
                "대중교통구분: 지하철 | 노선명: 5호선 | 년월: 201902 | 승차총승객수: 15,446,522; "
                "대중교통구분: 지하철 | 노선명: 5호선 | 년월: 201903 | 승차총승객수: 18,834,177"
            ),
            locator=locator("A352:D401", [352, 353], ["A", "B", "C", "D"]),
        )
    )

    assert candidate["generation_applied"] is True
    assert candidate["generated_answer"] == "15,446,522입니다."
    assert candidate["generated_citation"]["locator"]["range"] == "A352:D352"
    assert candidate["row_selection_basis"] == "question_condition_match_citation_segment"
    assert candidate["target_column_selection_basis"] == "question_target_column_dictionary_normalized_match"
    assert candidate["gold_fields_used_for_generation"] is False
    assert candidate["expected_answer_seen_by_generation"] is False
    assert candidate["supporting_evidence_seen_by_generation"] is False


def test_runtime_candidate_selects_non_first_target_row_from_question_condition():
    module = load_module()

    candidate = module.generate_xlsx_runtime_candidate(
        runtime_input(
            query_id="non_first",
            question="2019년 3월 5호선의 승차총승객수는 몇 명입니까?",
            citation_text=(
                "대중교통구분: 지하철 | 노선명: 5호선 | 년월: 201902 | 승차총승객수: 15,446,522; "
                "대중교통구분: 지하철 | 노선명: 5호선 | 년월: 201903 | 승차총승객수: 18,834,177"
            ),
            locator=locator("A352:D401", [352, 353], ["A", "B", "C", "D"]),
        )
    )

    assert candidate["generation_applied"] is True
    assert candidate["generated_answer"] == "18,834,177입니다."
    assert candidate["generated_citation"]["locator"]["range"] == "A353:D353"
    assert candidate["selected_target_row"] == 353
    assert candidate["selected_citation_segment_index"] == 1


def test_runtime_candidate_selects_same_facility_by_date_when_answer_row_is_not_first():
    module = load_module()

    candidate = module.generate_xlsx_runtime_candidate(
        runtime_input(
            query_id="facility_non_first",
            question="2022년 6월에 지정된 해오름요양원의 설치신고일자는 언제입니까?",
            citation_text=(
                "장기요양기관이름: 해오름요양원 | 지정일자: 2022-05-01 | 설치신고일자: 2022-05-03 | 우편번호: 11111; "
                "장기요양기관이름: 해오름요양원 | 지정일자: 2022-06-15 | 설치신고일자: 2022-06-20 | 우편번호: 22222"
            ),
            locator=locator("A1102:J1151", [1102, 1103], ["A", "B", "C", "D", "E", "F", "G", "H", "I", "J"]),
        )
    )

    assert candidate["generation_applied"] is True
    assert candidate["generated_answer"] == "2022-06-20입니다."
    assert candidate["selected_citation_segment_index"] == 1
    assert candidate["selected_target_row"] == 1103
    assert candidate["generated_citation"]["locator"]["range"] == "A1103:J1103"


def test_runtime_candidate_fails_closed_for_duplicate_matching_rows():
    module = load_module()

    candidate = module.generate_xlsx_runtime_candidate(
        runtime_input(
            query_id="duplicate_winner",
            question="2022년 6월에 지정된 해오름요양원의 우편번호는 무엇입니까?",
            citation_text=(
                "장기요양기관이름: 해오름요양원 | 지정일자: 2022-06-01 | 우편번호: 11111; "
                "장기요양기관이름: 해오름요양원 | 지정일자: 2022-06-15 | 우편번호: 22222"
            ),
            locator=locator("A1102:J1151", [1102, 1103], ["A", "B", "C", "D", "E", "F", "G", "H", "I", "J"]),
        )
    )

    assert candidate["generation_applied"] is False
    assert candidate["failure_reason"] == "ambiguous_row_selection"


def test_runtime_candidate_fails_closed_when_target_column_missing_from_selected_row():
    module = load_module()

    candidate = module.generate_xlsx_runtime_candidate(
        runtime_input(
            query_id="target_missing",
            question="2022년 6월에 지정된 해오름요양원의 기관별 상세주소는 무엇입니까?",
            citation_text="장기요양기관이름: 해오름요양원 | 지정일자: 2022-06-15 | 우편번호: 22222",
            locator=locator("A1102:J1151", [1102], ["A", "B", "C", "D", "E", "F", "G", "H", "I", "J"]),
        )
    )

    assert candidate["generation_applied"] is False
    assert candidate["failure_reason"] == "target_column_missing_from_selected_row"


def test_runtime_candidate_fails_closed_when_selected_segment_has_no_target_row():
    module = load_module()

    candidate = module.generate_xlsx_runtime_candidate(
        runtime_input(
            query_id="target_row_missing",
            question="2022년 6월에 지정된 해오름요양원의 우편번호는 무엇입니까?",
            citation_text=(
                "장기요양기관이름: 해오름요양원 | 지정일자: 2022-05-01 | 우편번호: 11111; "
                "장기요양기관이름: 해오름요양원 | 지정일자: 2022-06-15 | 우편번호: 22222"
            ),
            locator=locator("A1102:J1151", [1102], ["A", "B", "C", "D", "E", "F", "G", "H", "I", "J"]),
        )
    )

    assert candidate["generation_applied"] is False
    assert candidate["failure_reason"] == "selected_segment_has_no_target_row"


def test_runtime_candidate_extracts_facility_address_and_installation_date():
    module = load_module()
    citation_text = (
        "장기요양기관코드: 12,820,000,827 | 장기요양기관이름: 인천은빛요양원 | 우편번호: 21540 | "
        "시도코드: 28 | 시군구코드: 200 | 법정동코드: 103 | 시도 시군구 법정동명: 인천광역시 남동구 만수동 | "
        "지정일자: 2022-05-01 | 설치신고일자: 2022-05-01 | 기관별 상세주소: 인천광역시 남동구 하촌로 26 7층701 702호 (만수동 거신빌딩)"
    )

    address = module.generate_xlsx_runtime_candidate(
        runtime_input(
            query_id="address",
            question="2022년 5월에 지정된 인천은빛요양원의 기관별 상세주소는 무엇입니까?",
            citation_text=citation_text,
            locator=locator("A1102:J1151", [1102], ["A", "B", "C", "D", "E", "F", "G", "H", "I", "J"]),
        )
    )
    install_date = module.generate_xlsx_runtime_candidate(
        runtime_input(
            query_id="install",
            question="2022년 5월에 지정된 인천은빛요양원의 설치신고일자는 언제입니까?",
            citation_text=citation_text,
            locator=locator("A1102:J1151", [1102], ["A", "B", "C", "D", "E", "F", "G", "H", "I", "J"]),
        )
    )

    assert address["generated_answer"] == "인천광역시 남동구 하촌로 26 7층701 702호 (만수동 거신빌딩)입니다."
    assert install_date["generated_answer"] == "2022-05-01입니다."
    assert address["generated_citation"]["locator"]["range"] == "A1102:J1102"
    assert install_date["generated_citation"]["locator"]["range"] == "A1102:J1102"


def test_runtime_candidate_fails_closed_for_ambiguous_multi_row_without_unique_condition():
    module = load_module()

    candidate = module.generate_xlsx_runtime_candidate(
        runtime_input(
            query_id="ambiguous",
            question="승차총승객수는 몇 명입니까?",
            citation_text=(
                "노선명: 5호선 | 년월: 201902 | 승차총승객수: 15,446,522; "
                "노선명: 5호선 | 년월: 201903 | 승차총승객수: 18,834,177"
            ),
            locator=locator("A352:D401", [352, 353], ["A", "B", "C", "D"]),
        )
    )

    assert candidate["generation_applied"] is False
    assert candidate["repair_confidence"] == "failed"
    assert candidate["failure_reason"] == "ambiguous_row_selection"


def test_runtime_candidate_hidden_leakage_and_gold_poison_fail_or_do_not_change_generation():
    module = load_module()
    base = runtime_input(
        query_id="poison",
        question="2020년 11월에 지정된 하얀민들레노인요양원의 우편번호는 무엇입니까?",
        citation_text=(
            "장기요양기관이름: 하얀민들레노인요양원 | 우편번호: 41786 | 지정일자: 2020-11-26 | "
            "설치신고일자: 2020-11-26 | 기관별 상세주소: 대구광역시 서구 통학로38길 13 (평리동)"
        ),
        locator=locator("A702:J751", [702], ["A", "B", "C", "D", "E", "F", "G", "H", "I", "J"]),
    )
    with_poison = {**base, "expected_answer": "POISON", "supporting_evidence": "Z999"}

    normal = module.generate_xlsx_runtime_candidate(base)
    poison = module.generate_xlsx_runtime_candidate(with_poison)
    hidden = module.generate_xlsx_runtime_candidate({**base, "hidden_excluded_leakage_count": 1})

    assert normal["generated_answer"] == poison["generated_answer"] == "41786입니다."
    assert normal["generated_citation"]["locator"]["range"] == poison["generated_citation"]["locator"]["range"]
    assert poison["expected_answer_seen_by_generation"] is False
    assert poison["supporting_evidence_seen_by_generation"] is False
    assert hidden["generation_applied"] is False
    assert hidden["failure_reason"] == "hidden_excluded_leakage"


def test_runtime_candidate_run_emits_report_only_artifacts_without_mutating_baseline(tmp_path: Path):
    module = load_module()
    results_path = tmp_path / "runtime.jsonl"
    status_path = tmp_path / "status.md"
    baseline_path = resolve_report_artifact_path(REPORT_DIR / "baseline_v1.json")
    baseline_before = baseline_path.read_bytes()

    report = module.run_candidate(
        baseline_report_path=baseline_path,
        report_only_repair_candidate_path=None,
        scorer_results_path=resolve_report_artifact_path(REPORT_DIR / "scorer_v1.jsonl"),
        xlsx_gold_csv_path=ROOT / "ai" / "eval" / "eval_queries" / "gold_queries_xlsx_question_gold_v2.csv",
        output_report=None,
        output_md=None,
        output_results_jsonl=results_path,
        runtime_environment_report_path=None,
        status_md=status_path,
    )

    rows = [json.loads(line) for line in results_path.read_text(encoding="utf-8").splitlines() if line.strip()]
    status_text = status_path.read_text(encoding="utf-8")
    assert baseline_path.read_bytes() == baseline_before
    assert len(rows) == 29
    assert len({row["query_id"] for row in rows}) == 29
    assert "## xlsx_runtime_candidate" in status_text
    assert "`xlsx_pass=19`" in status_text
    assert report["runtime_candidate_failure_category_counts"] == {"PARTIAL_OR_UNSUPPORTED": 3, "PASS": 26}
    assert report["xlsx_summary"]["runtime_candidate_pass_count"] == 19
    assert report["xlsx_summary"]["remaining_xlsx_failures"] == []
    assert report["remaining_failures_by_track"] == {"pdf_business_ocr_mm": ["gq_auto_010", "gq_auto_030", "gq_pdf_section_question_001"]}
    assert report["runtime_generation_trace_counts"]["question_condition_match_citation_segment"] >= 18
    assert report["runtime_failure_reason_counts"] == {}
    assert sum(report["selected_segment_index_histogram"].values()) == 19
    assert "0" in report["selected_segment_index_histogram"]
    assert report["row_selection_condition_counts"]["년월=question_year_month"] >= 1
    assert report["ambiguous_selection_query_ids"] == []
    assert report["guardrails"]["promotion_evidence"] is False
    assert report["guardrails"]["production_mutation"] is False
    assert report["guardrails"]["expected_answer_used_for_generation"] is False
    assert report["guardrails"]["supporting_evidence_used_for_generation"] is False
    assert report["local_llm_gpu_usage"]["used"] is False
    assert "gq_auto_030" in {row["query_id"] for row in report["pdf_remaining_failure_analysis"]}


def runtime_input(
    *,
    query_id: str,
    question: str,
    citation_text: str,
    locator: dict,
    hidden_excluded_leakage_count: int = 0,
) -> dict:
    return {
        "query_id": query_id,
        "question": question,
        "generated_answer": "",
        "generated_citation": {"citation_text": citation_text, "locator": locator},
        "hidden_excluded_leakage_count": hidden_excluded_leakage_count,
    }


def locator(range_ref: str, rows: list[int], columns: list[str]) -> dict:
    return {
        "file": "sample.xlsx",
        "sheet": "Sheet1",
        "range": range_ref,
        "matched_cells": [],
        "target_rows": rows,
        "target_columns": columns,
        "document_version_id": "docv",
        "search_unit_id": "su",
    }
