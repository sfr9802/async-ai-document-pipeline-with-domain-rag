from __future__ import annotations

import csv
import hashlib
import json
import re
import subprocess
import sys
from pathlib import Path

from ai.eval import rag_v560_official_metric_scored_execution_and_failure_attribution_nonprod as v560


ROOT = Path(__file__).resolve().parents[2]
V4_7_7_SHORT_KEY = "v4_7_7"
V4_7_7_SHORT_RUN_ID = "v4_7_7_v3_legacy_archive_and_runner_consolidation"
V4_7_7_LONG_RUN_ID = (
    "official_answer_citation_agentic_loop_run_v4_7_7_"
    "v3_legacy_artifact_archive_and_diagnostic_runner_consolidation_nonprod"
)
V4_7_7_STATUS = "V4_7_7_V3_LEGACY_ARCHIVE_RUNNER_CONSOLIDATION_NONPROD_READY"
V4_7_7_REPORT = ROOT / "ai" / "eval" / "reports" / "rag-ingestion" / "runs" / "v4_7_7" / "report.json"
V4_7_8_SHORT_KEY = "v4_7_8"
V4_7_8_SHORT_RUN_ID = "v4_7_8_test_doc_dependency_decoupling_runner_alias_expansion"
V4_7_8_LONG_RUN_ID = (
    "official_answer_citation_agentic_loop_run_v4_7_8_"
    "test_doc_dependency_decoupling_and_legacy_runner_alias_expansion_nonprod"
)
V4_7_8_STATUS = "V4_7_8_TEST_DOC_DEPENDENCY_DECOUPLING_RUNNER_ALIAS_EXPANSION_NONPROD_READY"
V4_7_8_REPORT = ROOT / "ai" / "eval" / "reports" / "rag-ingestion" / "runs" / "v4_7_8" / "report.json"
V4_7_9_SHORT_KEY = "v4_7_9"
V4_7_9_SHORT_RUN_ID = "v4_7_9_pdf_evidence_residual_answer_quality_replay"
V4_7_9_LONG_RUN_ID = (
    "official_answer_citation_agentic_loop_run_v4_7_9_"
    "pdf_evidence_residual_answer_quality_replay_nonprod"
)
V4_7_9_STATUS = "V4_7_9_PDF_EVIDENCE_RESIDUAL_ANSWER_QUALITY_REPLAY_NONPROD_READY"
V4_7_9_REPORT = ROOT / "ai" / "eval" / "reports" / "rag-ingestion" / "runs" / "v4_7_9" / "report.json"
V4_7_10_SHORT_KEY = "v4_7_10"
V4_7_10_SHORT_RUN_ID = "v4_7_10_pdf_korean_evidence_normalization_and_answer_replay_readiness"
V4_7_10_LONG_RUN_ID = (
    "official_answer_citation_agentic_loop_run_v4_7_10_"
    "pdf_korean_evidence_normalization_and_answer_replay_readiness_nonprod"
)
V4_7_10_STATUS = "V4_7_10_PDF_KOREAN_EVIDENCE_NORMALIZATION_AND_ANSWER_REPLAY_READINESS_NONPROD_READY"
V4_7_10_REPORT = ROOT / "ai" / "eval" / "reports" / "rag-ingestion" / "runs" / "v4_7_10" / "report.json"
V4_7_11_SHORT_KEY = "v4_7_11"
V4_7_11_SHORT_RUN_ID = "v4_7_11_actual_llm_answer_replay_and_silver_diagnostic_smoke"
V4_7_11_LONG_RUN_ID = (
    "official_answer_citation_agentic_loop_run_v4_7_11_"
    "actual_llm_answer_replay_and_silver_diagnostic_smoke_nonprod"
)
V4_7_11_STATUS = "V4_7_11_ACTUAL_LLM_ANSWER_REPLAY_AND_SILVER_DIAGNOSTIC_SMOKE_NONPROD_READY"
V4_7_11_REPORT = ROOT / "ai" / "eval" / "reports" / "rag-ingestion" / "runs" / "v4_7_11" / "report.json"
V4_7_12_SHORT_KEY = "v4_7_12"
V4_7_12_SHORT_RUN_ID = "v4_7_12_layered_retrieval_generalization_and_overfit_audit"
V4_7_12_ACTIVE_GOAL_ALIAS = "v4_7_12_answer_policy_calibration_and_silver_manifest_reconnect"
V4_7_12_LONG_RUN_ID = (
    "official_answer_citation_agentic_loop_run_v4_7_12_"
    "layered_retrieval_generalization_and_overfit_audit_nonprod"
)
V4_7_12_STATUS = "V4_7_12_LAYERED_RETRIEVAL_GENERALIZATION_AND_OVERFIT_AUDIT_NONPROD_READY"
V4_7_12_REPORT = ROOT / "ai" / "eval" / "reports" / "rag-ingestion" / "runs" / "v4_7_12" / "report.json"
V4_7_13_SHORT_KEY = "v4_7_13"
V4_7_13_SHORT_RUN_ID = "v4_7_13_live_retrieval_answerability_and_full_pdf_replay"
V4_7_13_LONG_RUN_ID = (
    "official_answer_citation_agentic_loop_run_v4_7_13_"
    "live_retrieval_answerability_and_full_pdf_replay_nonprod"
)
V4_7_13_STATUS = "V4_7_13_LIVE_RETRIEVAL_ANSWERABILITY_AND_FULL_PDF_REPLAY_NONPROD_READY"
V4_7_13_REPORT = ROOT / "ai" / "eval" / "reports" / "rag-ingestion" / "runs" / "v4_7_13" / "report.json"
V4_7_14_SHORT_KEY = "v4_7_14"
V4_7_14_SHORT_RUN_ID = "v4_7_14_diagnostic_precondition_hardening"
V4_7_14_LONG_RUN_ID = (
    "official_answer_citation_agentic_loop_run_v4_7_14_"
    "diagnostic_precondition_hardening_nonprod"
)
V4_7_14_STATUS = "V4_7_14_DIAGNOSTIC_PRECONDITION_HARDENING_NONPROD_READY"
V4_7_14_REPORT = ROOT / "ai" / "eval" / "reports" / "rag-ingestion" / "runs" / "v4_7_14" / "report.json"
V4_7_15_SHORT_KEY = "v4_7_15"
V4_7_15_SHORT_RUN_ID = "v4_7_15_read_only_searchindex_replay_projection"
V4_7_15_LONG_RUN_ID = (
    "official_answer_citation_agentic_loop_run_v4_7_15_"
    "read_only_searchindex_replay_projection_nonprod"
)
V4_7_15_STATUS = "V4_7_15_READ_ONLY_SEARCHINDEX_REPLAY_PROJECTION_NONPROD_READY"
V4_7_15_REPORT = ROOT / "ai" / "eval" / "reports" / "rag-ingestion" / "runs" / "v4_7_15" / "report.json"
V4_7_16_SHORT_KEY = "v4_7_16"
V4_7_16_SHORT_RUN_ID = "v4_7_16_target_recall_repair_prototype"
V4_7_16_LONG_RUN_ID = (
    "official_answer_citation_agentic_loop_run_v4_7_16_"
    "target_recall_repair_prototype_nonprod"
)
V4_7_16_STATUS = "V4_7_16_TARGET_RECALL_REPAIR_PROTOTYPE_NONPROD_READY"
V4_7_16_REPORT = ROOT / "ai" / "eval" / "reports" / "rag-ingestion" / "runs" / "v4_7_16" / "report.json"
V4_7_17_SHORT_KEY = "v4_7_17"
V4_7_17_SHORT_RUN_ID = "v4_7_17_candidate_only_generalization_validation_and_xlsx_table_axis_repair_audit"
V4_7_17_LONG_RUN_ID = (
    "official_answer_citation_agentic_loop_run_v4_7_17_"
    "candidate_only_generalization_validation_and_xlsx_table_axis_repair_audit_nonprod"
)
V4_7_17_STATUS = (
    "V4_7_17_CANDIDATE_ONLY_GENERALIZATION_VALIDATION_AND_XLSX_TABLE_AXIS_REPAIR_AUDIT_NONPROD_READY"
)
V4_7_17_REPORT = ROOT / "ai" / "eval" / "reports" / "rag-ingestion" / "runs" / "v4_7_17" / "report.json"
V4_7_18_SHORT_KEY = "v4_7_18"
V4_7_18_SHORT_RUN_ID = "v4_7_18_xlsx_candidate_only_materialization_repair_and_lineage_reproducibility"
V4_7_18_LONG_RUN_ID = (
    "official_answer_citation_agentic_loop_run_v4_7_18_"
    "xlsx_candidate_only_materialization_repair_and_lineage_reproducibility_nonprod"
)
V4_7_18_STATUS = (
    "V4_7_18_XLSX_CANDIDATE_ONLY_MATERIALIZATION_REPAIR_AND_LINEAGE_REPRODUCIBILITY_NONPROD_READY"
)
V4_7_18_REPORT = ROOT / "ai" / "eval" / "reports" / "rag-ingestion" / "runs" / "v4_7_18" / "report.json"
V5_0_SHORT_KEY = "v5_0"
V5_0_SHORT_RUN_ID = "v5_0_v4_closeout_and_v5_gate_plan"
V5_0_LONG_RUN_ID = "official_answer_citation_agentic_loop_run_v5_0_v4_closeout_and_v5_gate_plan_nonprod"
V5_0_STATUS = "V5_0_V4_CLOSEOUT_AND_V5_GATE_PLAN_DIAGNOSTIC_NONPROD_READY"
V5_0_REPORT = ROOT / "ai" / "eval" / "reports" / "rag-ingestion" / "runs" / "v5_0" / "report.json"
V5_1_SHORT_KEY = "v5_1"
V5_1_SHORT_RUN_ID = "v5_1_official_eval_gate_scaffolding"
V5_1_LONG_RUN_ID = "official_answer_citation_agentic_loop_run_v5_1_official_eval_gate_scaffolding_nonprod"
V5_1_STATUS = "V5_1_OFFICIAL_EVAL_GATE_SCAFFOLDING_DIAGNOSTIC_NONPROD_READY"
V5_1_REPORT = ROOT / "ai" / "eval" / "reports" / "rag-ingestion" / "runs" / "v5_1" / "report.json"
V5_2_SHORT_KEY = "v5_2"
V5_2_SHORT_RUN_ID = "v5_2_xlsx_residual_candidate_only_retrieval_engineering"
V5_2_LONG_RUN_ID = (
    "official_answer_citation_agentic_loop_run_v5_2_"
    "xlsx_residual_candidate_only_retrieval_engineering_nonprod"
)
V5_2_STATUS = "V5_2_XLSX_RESIDUAL_CANDIDATE_ONLY_RETRIEVAL_ENGINEERING_DIAGNOSTIC_NONPROD_READY"
V5_2_REPORT = ROOT / "ai" / "eval" / "reports" / "rag-ingestion" / "runs" / "v5_2" / "report.json"
V5_3_SHORT_KEY = "v5_3"
V5_3_SHORT_RUN_ID = "v5_3_pdf_text_residual_retrieval_evidence_hardening"
V5_3_LONG_RUN_ID = (
    "official_answer_citation_agentic_loop_run_v5_3_"
    "pdf_text_residual_retrieval_evidence_hardening_nonprod"
)
V5_3_STATUS = "V5_3_PDF_TEXT_RESIDUAL_RETRIEVAL_EVIDENCE_HARDENING_DIAGNOSTIC_NONPROD_READY"
V5_3_REPORT = ROOT / "ai" / "eval" / "reports" / "rag-ingestion" / "runs" / "v5_3" / "report.json"
V5_4_SHORT_KEY = "v5_4"
V5_4_SHORT_RUN_ID = "v5_4_user_owned_official_eval_approval_packet"
V5_4_LONG_RUN_ID = (
    "official_answer_citation_agentic_loop_run_v5_4_"
    "user_owned_official_eval_approval_packet_nonprod"
)
V5_4_STATUS = "V5_4_USER_OWNED_OFFICIAL_EVAL_APPROVAL_PACKET_NONPROD_READY"
V5_4_RUN_DIR = ROOT / "ai" / "eval" / "reports" / "rag-ingestion" / "runs" / "v5_4"
V5_4_REPORT = V5_4_RUN_DIR / "report.json"
V5_4_SCHEMA = V5_4_RUN_DIR / "user_owned_approval_schema.json"
V5_4_POLICY_TEMPLATE = V5_4_RUN_DIR / "user_owned_policy_template.json"
V5_4_PACKET_JSONL = V5_4_RUN_DIR / "user_review_packet.jsonl"
V5_4_PACKET_CSV = V5_4_RUN_DIR / "user_review_packet.csv"
V5_4_PACKET_XLSX = V5_4_RUN_DIR / "user_review_packet.xlsx"
V5_4_USER_OWNED_FIELDS = (
    "include_in_official_denominator",
    "relevance_label",
    "answerability_label",
    "expected_answer_ko",
    "supporting_evidence_ids",
    "supporting_evidence_note",
    "gold_status",
    "policy_note",
    "reviewer",
    "reviewed_at",
)
V5_5_SHORT_KEY = "v5_5"
V5_5_SHORT_RUN_ID = "v5_5_user_approved_gold_packet_ingestion_and_official_metric_dry_run"
V5_5_LONG_RUN_ID = (
    "official_answer_citation_agentic_loop_run_v5_5_"
    "user_approved_gold_packet_ingestion_and_official_metric_dry_run_nonprod"
)
V5_5_STATUS = "V5_5_USER_APPROVED_GOLD_PACKET_INGESTION_AND_OFFICIAL_METRIC_DRY_RUN_NONPROD_READY"
V5_5_RUN_DIR = ROOT / "ai" / "eval" / "reports" / "rag-ingestion" / "runs" / "v5_5"
V5_5_REPORT = V5_5_RUN_DIR / "report.json"
V5_5_APPROVED_GOLD_PACKET = V5_5_RUN_DIR / "user_approved_gold_packet.jsonl"
V5_5_DENOMINATOR = V5_5_RUN_DIR / "user_approved_denominator.jsonl"
V5_5_QRELS = V5_5_RUN_DIR / "user_approved_qrels.jsonl"
V5_5_EXPECTED_ANSWERS = V5_5_RUN_DIR / "user_approved_expected_answers.jsonl"
V5_5_OFFICIAL_METRIC_INPUT = V5_5_RUN_DIR / "official_metric_input.jsonl"
V5_5_DRY_RUN_RESULT = V5_5_RUN_DIR / "official_metric_dry_run_result.json"
V5_6_SHORT_KEY = "v5_6"
V5_6_SHORT_RUN_ID = "v5_6_official_metric_scored_execution_and_failure_attribution_nonprod"
V5_6_LONG_RUN_ID = (
    "official_answer_citation_agentic_loop_run_v5_6_"
    "official_metric_scored_execution_and_failure_attribution_nonprod"
)
V5_6_STATUS = "V5_6_OFFICIAL_METRIC_SCORED_EXECUTION_BACKEND_UNAVAILABLE_FAIL_CLOSED_NONPROD_READY"
V5_6_RUN_DIR = ROOT / "ai" / "eval" / "reports" / "rag-ingestion" / "runs" / "v5_6"
V5_6_REPORT = V5_6_RUN_DIR / "report.json"
V5_6_SCORED_RESULT = V5_6_RUN_DIR / "official_metric_scored_result.json"
V5_6_FAILURE_ATTRIBUTION = V5_6_RUN_DIR / "failure_attribution.jsonl"
V5_6_2_SHORT_KEY = "v5_6_2"
V5_6_2_SHORT_RUN_ID = "v5_6_2_official_metric_backend_enabled_preflight_scored_rerun_nonprod"
V5_6_2_LONG_RUN_ID = (
    "official_answer_citation_agentic_loop_run_v5_6_2_"
    "official_metric_backend_enabled_preflight_scored_rerun_nonprod"
)
V5_6_2_STATUS = "V5_6_2_OFFICIAL_METRIC_BACKEND_ENABLED_PREFLIGHT_FAIL_CLOSED_NONPROD_READY"
V5_6_2_SCORED_STATUS = "V5_6_2_OFFICIAL_METRIC_BACKEND_ENABLED_PREFLIGHT_SCORED_RERUN_NONPROD_READY"
V5_6_2_RUN_DIR = ROOT / "ai" / "eval" / "reports" / "rag-ingestion" / "runs" / "v5_6_2"
V5_6_2_REPORT = V5_6_2_RUN_DIR / "report.json"
V5_6_2_SCORED_RESULT = V5_6_2_RUN_DIR / "official_metric_scored_result.json"
V5_6_2_FAILURE_ATTRIBUTION = V5_6_2_RUN_DIR / "failure_attribution.jsonl"
V5_6_2_BACKEND_PREFLIGHT_RESULT = V5_6_2_RUN_DIR / "backend_preflight_result.json"
V5_6_3_SHORT_KEY = "v5_6_3"
V5_6_3_SHORT_RUN_ID = "v5_6_3_official_metric_backend_probe_and_scored_execution_nonprod"
V5_6_3_LONG_RUN_ID = (
    "official_answer_citation_agentic_loop_run_v5_6_3_"
    "official_metric_backend_probe_and_scored_execution_nonprod"
)
V5_6_3_STATUS = "V5_6_3_OFFICIAL_METRIC_BACKEND_PROBE_FAIL_CLOSED_NONPROD_READY"
V5_6_3_SCORED_STATUS = "V5_6_3_OFFICIAL_METRIC_BACKEND_PROBE_SCORED_EXECUTION_NONPROD_READY"
V5_6_3_RUN_DIR = ROOT / "ai" / "eval" / "reports" / "rag-ingestion" / "runs" / "v5_6_3"
V5_6_3_REPORT = V5_6_3_RUN_DIR / "report.json"
V5_6_3_SCORED_RESULT = V5_6_3_RUN_DIR / "official_metric_scored_result.json"
V5_6_3_FAILURE_ATTRIBUTION = V5_6_3_RUN_DIR / "failure_attribution.jsonl"
V5_6_3_BACKEND_PREFLIGHT_RESULT = V5_6_3_RUN_DIR / "backend_preflight_result.json"
V5_6_REFACTOR_COMPARISON_SHORT_KEY = "v5_6_refactor_comparison"
V5_6_REFACTOR_COMPARISON_SHORT_RUN_ID = "v5_6_refactor_route_comparison_packet_diagnostic_nonprod"
V5_6_REFACTOR_COMPARISON_LONG_RUN_ID = V5_6_REFACTOR_COMPARISON_SHORT_RUN_ID
V5_6_REFACTOR_COMPARISON_STATUS = "V5_6_REFACTOR_ROUTE_COMPARISON_PACKET_DIAGNOSTIC_NONPROD_READY"
V5_6_REFACTOR_COMPARISON_RUN_DIR = (
    ROOT / "ai" / "eval" / "reports" / "rag-ingestion" / "runs" / V5_6_REFACTOR_COMPARISON_SHORT_KEY
)
V5_6_REFACTOR_COMPARISON_REPORT = V5_6_REFACTOR_COMPARISON_RUN_DIR / "report.json"
REPORT_ROOT = ROOT / "ai" / "eval" / "reports" / "rag-ingestion"
STATUS_JSONL = REPORT_ROOT / "status.jsonl"
PROGRESS_DOC = ROOT / "docs" / "rag-ingestion-progress.md"
MEASUREMENTS_DOC = ROOT / "docs" / "rag-ingestion-measurements.md"
TRIAGE_DOC = ROOT / "docs" / "rag-ingestion-triage.md"


def _read_json(path: Path) -> dict[str, object]:
    return json.loads(path.read_text(encoding="utf-8"))


def _read_jsonl(path: Path) -> list[dict[str, object]]:
    return [json.loads(line) for line in path.read_text(encoding="utf-8").splitlines() if line.strip()]


def _sha256_file(path: Path) -> str:
    return hashlib.sha256(path.read_bytes()).hexdigest()


def _optional_doc_section(text: str, heading: str) -> str:
    if heading not in text:
        return ""
    return text.split(heading, 1)[1].split("\n### ", 1)[0]


def _assert_status_recorded_in_progress_and_report(
    progress: str,
    report: dict[str, object],
    *,
    short_run_id: str,
    status: str,
) -> None:
    assert report["short_run_id"] == short_run_id
    assert report["status"] == status
    assert f"Overall status: `{status}`" in progress


def _load_v4710_report() -> dict[str, object]:
    if V4_7_10_REPORT.exists():
        return _read_json(V4_7_10_REPORT)
    import ai.scripts.rag_eval as runner

    return runner.check_run(V4_7_10_SHORT_KEY)


def _fake_strict_korean_answer(prompt: str) -> str:
    payload = json.loads(prompt)
    evidence = str(payload.get("bounded_evidence_excerpt") or payload.get("evidence") or "")
    terms = [token for token in re.findall(r"[가-힣A-Za-z0-9]{2,}", evidence) if token]
    head = " ".join(terms[:2]) or "근거"
    return json.dumps(
        {
            "final_answer": f"근거에 따르면 {head} 관련 내용입니다.",
            "abstain": False,
            "citations": ["evidence_1"],
            "answer_plan": "제공된 근거 문장만 사용해 한 문장으로 답변합니다.",
            "unsupported_claim_risk": False,
            "evidence_underuse_flag": False,
            "context_understanding_miss": False,
            "over_abstain_candidate": False,
        },
        ensure_ascii=False,
        sort_keys=True,
    )


def test_v5_diagnostic_common_helpers_preserve_write_doc_and_payload_semantics(tmp_path: Path) -> None:
    from ai.eval import rag_v5_diagnostic_common as common

    original = {"short_run_id": "v5_helper_probe", "nested": {"value": 1}}
    report, hashes = common.write_report_bundle(
        tmp_path,
        Path("ai/eval/reports/rag-ingestion/runs/v5_helper_probe/report.json"),
        original,
    )

    original["nested"]["value"] = 2
    report_path = tmp_path / "ai/eval/reports/rag-ingestion/runs/v5_helper_probe/report.json"
    assert report == {"short_run_id": "v5_helper_probe", "nested": {"value": 1}}
    assert json.loads(report_path.read_text(encoding="utf-8")) == report
    assert hashes == {"report_json_sha256": _sha256_file(report_path)}

    text = "Last updated: 2026-05-31 KST.\n\nbody\n"
    text = common.upsert_block_at_top(text, start_marker="<!-- start -->", end_marker="<!-- end -->", block="new")
    assert text.startswith("<!-- start -->\nnew\n<!-- end -->\n\n")
    assert common.sync_last_updated(text, "2026-06-01").startswith("<!-- start -->")
    assert "Last updated: 2026-06-01 KST." in common.sync_last_updated(text, "2026-06-01")

    replaced = common.replace_summary_block(
        "<!-- v5_2_summary_start -->\nold\n<!-- v5_2_summary_end -->\nbody",
        start_marker="<!-- v5_3_summary_start -->",
        end_marker="<!-- v5_3_summary_end -->",
        block="summary",
        marker_pattern=r"<!-- v5_[0-9]+_summary_start -->.*?<!-- v5_[0-9]+_summary_end -->",
    )
    assert replaced.startswith("<!-- v5_3_summary_start -->\nsummary\n<!-- v5_3_summary_end -->")

    try:
        common.assert_no_raw_payload_keys(
            {"nested": [{"raw_prompt": "do not serialize"}]},
            {"raw_prompt"},
            context="v5_helper_probe",
        )
    except ValueError as exc:
        assert "v5_helper_probe" in str(exc)
        assert "raw_prompt" in str(exc)
    else:
        raise AssertionError("common helper accepted a raw payload key")


def test_v477_registry_resolves_current_and_previous_short_keys() -> None:
    from ai.eval import rag_eval_registry as registry
    import ai.scripts.rag_eval as runner

    expected = {
        "v4_7_preofficial": "ai/eval/reports/rag-ingestion/runs/v4_7_preofficial/report.json",
        "v4_7_2": "ai/eval/reports/rag-ingestion/runs/v4_7_2/report.json",
        "v4_7_3": "ai/eval/reports/rag-ingestion/runs/v4_7_3/report.json",
        "v4_7_4": "ai/eval/reports/rag-ingestion/runs/v4_7_4/report.json",
        "v4_7_5": "ai/eval/reports/rag-ingestion/runs/v4_7_5/report.json",
        "v4_7_6": "ai/eval/reports/rag-ingestion/runs/v4_7_6/report.json",
        "v4_7_7": "ai/eval/reports/rag-ingestion/runs/v4_7_7/report.json",
        "v4_7_8": "ai/eval/reports/rag-ingestion/runs/v4_7_8/report.json",
        "v4_7_9": "ai/eval/reports/rag-ingestion/runs/v4_7_9/report.json",
        "v4_7_10": "ai/eval/reports/rag-ingestion/runs/v4_7_10/report.json",
        "v4_7_11": "ai/eval/reports/rag-ingestion/runs/v4_7_11/report.json",
        "v4_7_12": "ai/eval/reports/rag-ingestion/runs/v4_7_12/report.json",
        V4_7_12_ACTIVE_GOAL_ALIAS: "ai/eval/reports/rag-ingestion/runs/v4_7_12/report.json",
        "v4_7_13": "ai/eval/reports/rag-ingestion/runs/v4_7_13/report.json",
        "v4_7_14": "ai/eval/reports/rag-ingestion/runs/v4_7_14/report.json",
        "v4_7_15": "ai/eval/reports/rag-ingestion/runs/v4_7_15/report.json",
        "v4_7_16": "ai/eval/reports/rag-ingestion/runs/v4_7_16/report.json",
        "v4_7_17": "ai/eval/reports/rag-ingestion/runs/v4_7_17/report.json",
        "v4_7_18": "ai/eval/reports/rag-ingestion/runs/v4_7_18/report.json",
        "v5_0": "ai/eval/reports/rag-ingestion/runs/v5_0/report.json",
        "v5_1": "ai/eval/reports/rag-ingestion/runs/v5_1/report.json",
        "v5_2": "ai/eval/reports/rag-ingestion/runs/v5_2/report.json",
        "v5_3": "ai/eval/reports/rag-ingestion/runs/v5_3/report.json",
        "v5_4": "ai/eval/reports/rag-ingestion/runs/v5_4/report.json",
        "v5_5": "ai/eval/reports/rag-ingestion/runs/v5_5/report.json",
        "v5_6": "ai/eval/reports/rag-ingestion/runs/v5_6/report.json",
        "v5_6_2": "ai/eval/reports/rag-ingestion/runs/v5_6_2/report.json",
        "v5_6_3": "ai/eval/reports/rag-ingestion/runs/v5_6_3/report.json",
        "current": "ai/eval/reports/rag-ingestion/runs/v5_6/report.json",
    }
    for key, rel_path in expected.items():
        resolved = registry.resolve_run(key, root=ROOT)
        assert resolved.report_path == ROOT / rel_path
        in_memory_keys = {
            V4_7_11_SHORT_KEY,
            V4_7_12_SHORT_KEY,
            V4_7_12_ACTIVE_GOAL_ALIAS,
            V4_7_13_SHORT_KEY,
            V4_7_14_SHORT_KEY,
            V4_7_15_SHORT_KEY,
            V4_7_16_SHORT_KEY,
            V4_7_17_SHORT_KEY,
            V4_7_18_SHORT_KEY,
            V5_0_SHORT_KEY,
            V5_1_SHORT_KEY,
            V5_2_SHORT_KEY,
            V5_3_SHORT_KEY,
            V5_4_SHORT_KEY,
            V5_5_SHORT_KEY,
            V5_6_SHORT_KEY,
            V5_6_2_SHORT_KEY,
            V5_6_3_SHORT_KEY,
            "current",
        }
        if key in in_memory_keys and not resolved.report_path.exists():
            built = runner.check_run(key)
            assert built["artifact_paths"]["report_json"] == rel_path
        else:
            assert resolved.report_path.exists(), key

    prior = _load_v4710_report()
    assert prior["short_run_id"] == V4_7_10_SHORT_RUN_ID
    assert prior["canonical_long_run_id"] == V4_7_10_LONG_RUN_ID
    assert prior["status"] == V4_7_10_STATUS
    current = runner.check_run("current")
    assert current["short_run_id"] == V5_6_SHORT_RUN_ID
    assert current["canonical_long_run_id"] == V5_6_LONG_RUN_ID
    assert current["status"] == V5_6_STATUS
    explicit_v562 = runner.check_run("v5_6_2")
    assert explicit_v562["short_run_id"] == V5_6_2_SHORT_RUN_ID
    assert explicit_v562["canonical_long_run_id"] == V5_6_2_LONG_RUN_ID
    assert explicit_v562["status"] == V5_6_2_STATUS
    explicit_v563 = runner.check_run("v5_6_3")
    assert explicit_v563["short_run_id"] == V5_6_3_SHORT_RUN_ID
    assert explicit_v563["canonical_long_run_id"] == V5_6_3_LONG_RUN_ID
    assert explicit_v563["status"] == V5_6_3_STATUS
    explicit_v550 = runner.check_run("v5_5")
    assert explicit_v550["short_run_id"] == V5_5_SHORT_RUN_ID
    assert explicit_v550["status"] == V5_5_STATUS
    explicit_v540 = runner.check_run("v5_4")
    assert explicit_v540["short_run_id"] == V5_4_SHORT_RUN_ID
    assert explicit_v540["status"] == V5_4_STATUS
    explicit_v530 = runner.check_run("v5_3")
    assert explicit_v530["short_run_id"] == V5_3_SHORT_RUN_ID
    assert explicit_v530["status"] == V5_3_STATUS
    explicit_v520 = runner.check_run("v5_2")
    assert explicit_v520["short_run_id"] == V5_2_SHORT_RUN_ID
    assert explicit_v520["status"] == V5_2_STATUS
    explicit_v510 = runner.check_run("v5_1")
    assert explicit_v510["short_run_id"] == V5_1_SHORT_RUN_ID
    assert explicit_v510["status"] == V5_1_STATUS
    explicit_v500 = runner.check_run("v5_0")
    assert explicit_v500["short_run_id"] == V5_0_SHORT_RUN_ID
    assert explicit_v500["status"] == V5_0_STATUS
    explicit_v4718 = runner.check_run("v4_7_18")
    assert explicit_v4718["short_run_id"] == V4_7_18_SHORT_RUN_ID
    assert explicit_v4718["status"] == V4_7_18_STATUS
    explicit_v4717 = runner.check_run("v4_7_17")
    assert explicit_v4717["short_run_id"] == V4_7_17_SHORT_RUN_ID
    assert explicit_v4717["status"] == V4_7_17_STATUS
    previous_current = runner.check_run("v4_7_13")
    assert previous_current["short_run_id"] == V4_7_13_SHORT_RUN_ID
    assert previous_current["status"] == V4_7_13_STATUS
    explicit_prior = runner.check_run("v4_7_11")
    assert explicit_prior["short_run_id"] == V4_7_11_SHORT_RUN_ID
    assert explicit_prior["status"] == V4_7_11_STATUS


def test_v477_runner_dispatches_current_previous_and_safe_legacy_checks() -> None:
    import ai.scripts.rag_eval as runner

    assert runner.DEFAULT_RUN_KEY == V5_6_SHORT_KEY
    assert "v3_18" in runner.SAFE_LEGACY_CHECK_ALIASES
    assert "v3_19" in runner.SAFE_LEGACY_CHECK_ALIASES
    assert "v3_20" in runner.SAFE_LEGACY_CHECK_ALIASES
    assert "v3_22" in runner.SAFE_LEGACY_CHECK_ALIASES
    assert "v3_21" in runner.SAFE_LEGACY_CHECK_ALIASES
    assert "v3_16" not in runner.SAFE_LEGACY_CHECK_ALIASES

    for args, expected_key, expected_status in (
        (["--check"], V5_6_SHORT_KEY, V5_6_STATUS),
        (["current", "--check"], V5_6_SHORT_KEY, V5_6_STATUS),
        (["v5_6", "--check"], V5_6_SHORT_KEY, V5_6_STATUS),
        (["v5_6_2", "--check"], V5_6_2_SHORT_KEY, V5_6_2_STATUS),
        (["v5_6_3", "--check"], V5_6_3_SHORT_KEY, V5_6_3_STATUS),
        (["v5_5", "--check"], V5_5_SHORT_KEY, V5_5_STATUS),
        (["v5_4", "--check"], V5_4_SHORT_KEY, V5_4_STATUS),
        (["v5_3", "--check"], V5_3_SHORT_KEY, V5_3_STATUS),
        (["v5_2", "--check"], V5_2_SHORT_KEY, V5_2_STATUS),
        (["v5_1", "--check"], V5_1_SHORT_KEY, V5_1_STATUS),
        (["v5_0", "--check"], V5_0_SHORT_KEY, V5_0_STATUS),
        (["v4_7_18", "--check"], V4_7_18_SHORT_KEY, V4_7_18_STATUS),
        (["v4_7_17", "--check"], V4_7_17_SHORT_KEY, V4_7_17_STATUS),
        (["v4_7_16", "--check"], V4_7_16_SHORT_KEY, V4_7_16_STATUS),
        (["v4_7_15", "--check"], V4_7_15_SHORT_KEY, V4_7_15_STATUS),
        (["v4_7_14", "--check"], V4_7_14_SHORT_KEY, V4_7_14_STATUS),
        (["v4_7_13", "--check"], V4_7_13_SHORT_KEY, V4_7_13_STATUS),
        (["v4_7_12", "--check"], V4_7_12_SHORT_KEY, V4_7_12_STATUS),
        ([V4_7_12_ACTIVE_GOAL_ALIAS, "--check"], V4_7_12_SHORT_KEY, V4_7_12_STATUS),
        (["v4_7_11", "--check"], V4_7_11_SHORT_KEY, V4_7_11_STATUS),
        (["v4_7_10", "--check"], V4_7_10_SHORT_KEY, V4_7_10_STATUS),
        (["v4_7_9", "--check"], V4_7_9_SHORT_KEY, V4_7_9_STATUS),
        (["v4_7_8", "--check"], V4_7_8_SHORT_KEY, V4_7_8_STATUS),
        (["v4_7_7", "--check"], V4_7_7_SHORT_KEY, V4_7_7_STATUS),
        (["v4_7_6", "--check"], "v4_7_6", "V4_7_6_EVAL_ARTIFACT_ARCHIVE_PURGE_NONPROD_READY"),
    ):
        result = subprocess.run(
            [sys.executable, "-X", "utf8", "ai/scripts/rag_eval.py", *args],
            cwd=ROOT,
            text=True,
            capture_output=True,
            check=True,
        )
        payload = json.loads(result.stdout)
        assert payload["run_key"] == expected_key
        assert payload["status"] == expected_status


def test_v4712_explicit_check_builds_in_memory_and_current_uses_v560_with_v550_v540_v530_v520_v510_v500_v4718_explicit(
    monkeypatch,
) -> None:
    import ai.scripts.rag_eval as runner
    from ai.eval import rag_v4712_layered_retrieval_generalization_and_overfit_audit as v4712
    from ai.eval import rag_v4713_live_retrieval_answerability_and_full_pdf_replay as v4713
    from ai.eval import rag_v4718_xlsx_candidate_only_materialization_repair_and_lineage_reproducibility as v4718
    from ai.eval import rag_v500_v4_closeout_and_v5_gate_plan as v500
    from ai.eval import rag_v510_official_eval_gate_scaffolding as v510
    from ai.eval import rag_v520_xlsx_residual_candidate_only_retrieval_engineering as v520
    from ai.eval import rag_v530_pdf_text_residual_retrieval_evidence_hardening as v530
    from ai.eval import rag_v540_user_owned_official_eval_approval_packet as v540
    from ai.eval import rag_v550_user_approved_gold_packet_ingestion_and_official_metric_dry_run as v550
    from ai.eval import rag_v560_official_metric_scored_execution_and_failure_attribution_nonprod as v560
    from ai.eval import rag_v562_official_metric_backend_enabled_preflight_scored_rerun_nonprod as v562

    missing_report = Path("ai/eval/reports/rag-ingestion/runs/v4_7_12_missing_for_test/report.json")
    monkeypatch.setattr(v4712, "SHORT_REPORT_PATH", missing_report)

    current = runner.check_run("current")
    explicit_v562 = runner.check_run("v5_6_2")
    explicit_v550 = runner.check_run("v5_5")
    explicit_v540 = runner.check_run("v5_4")
    explicit_v530 = runner.check_run("v5_3")
    explicit_v520 = runner.check_run("v5_2")
    explicit_v500 = runner.check_run("v5_0")
    explicit_v4718 = runner.check_run("v4_7_18")
    long_alias = runner.check_run(V4_7_12_LONG_RUN_ID)

    v560.check_report(current)
    assert current["short_run_id"] == V5_6_SHORT_RUN_ID
    v562.check_report(explicit_v562)
    assert explicit_v562["short_run_id"] == V5_6_2_SHORT_RUN_ID
    v550.check_report(explicit_v550)
    assert explicit_v550["short_run_id"] == V5_5_SHORT_RUN_ID
    v540.check_report(explicit_v540)
    assert explicit_v540["short_run_id"] == V5_4_SHORT_RUN_ID
    v530.check_report(explicit_v530)
    assert explicit_v530["short_run_id"] == V5_3_SHORT_RUN_ID
    v520.check_report(explicit_v520)
    assert explicit_v520["short_run_id"] == V5_2_SHORT_RUN_ID
    explicit_v510 = runner.check_run("v5_1")
    v510.check_report(explicit_v510)
    assert explicit_v510["short_run_id"] == V5_1_SHORT_RUN_ID
    v500.check_report(explicit_v500)
    assert explicit_v500["short_run_id"] == V5_0_SHORT_RUN_ID
    v4718.check_report(explicit_v4718)
    assert explicit_v4718["short_run_id"] == V4_7_18_SHORT_RUN_ID
    v4712.check_report(long_alias)
    assert long_alias["short_run_id"] == V4_7_12_SHORT_RUN_ID
    assert long_alias["artifact_paths"]["report_json"] == missing_report.as_posix()
    assert not (ROOT / missing_report).exists()
    explicit_v4713 = runner.check_run("v4_7_13")
    v4713.check_report(explicit_v4713)
    assert explicit_v4713["short_run_id"] == V4_7_13_SHORT_RUN_ID


def test_v4712_layered_retrieval_audit_preserves_architecture_and_is_not_canary_limited() -> None:
    from ai.eval import rag_v4712_layered_retrieval_generalization_and_overfit_audit as v4712

    source_packet = ROOT / "ai" / "eval" / "reports" / "rag-ingestion" / "runs" / "v4_7_11" / "answer_review_packet_ko.jsonl"
    source_packet_sha_before = _sha256_file(source_packet)
    report = v4712.build_report(
        root=ROOT,
        execute=False,
        sync_surfaces=False,
        env={},
        generated_at="2026-05-31T00:00:00Z",
    )
    v4712.check_report(report)
    counters = report["counters"]

    assert report["short_run_id"] == V4_7_12_SHORT_RUN_ID
    assert report["source_run_id"] == V4_7_11_SHORT_RUN_ID
    assert report["source_pdf_surface_run_id"] == V4_7_10_SHORT_RUN_ID
    assert report["diagnostic_only"] is True
    assert report["non_production"] is True
    assert report["official_metric"] is False
    assert report["official_metric_input_rows"] == 0
    assert report["promotion_evidence"] is False
    assert report["product_success_evidence_allowed"] is False
    assert report["live_db_index_cache_readiness"] is False
    assert report["protected_namespaces_touched"] == []
    assert counters["v4_7_11_canary_row_count"] == 9
    assert counters["pdf_survivor_row_count"] == 58
    assert counters["pdf_answer_ready_evidencebundle_count"] == 57
    assert counters["pdf_full_replay_eligible_count"] == 57
    assert counters["layered_retrieval_audit_row_count"] >= 57
    assert counters["layered_retrieval_audit_row_count"] > counters["v4_7_11_canary_row_count"]
    assert counters["searchview_vector_payload_candidate_only_count"] >= counters["layered_retrieval_audit_row_count"]
    assert counters["sourceatom_evidencebundle_truth_count"] >= counters["layered_retrieval_audit_row_count"]
    assert counters["vector_payload_used_as_evidence_truth_violation_count"] == 0
    assert counters["raw_pdf_query_time_parsing_attempt_count"] == 0
    assert counters["raw_xlsx_query_time_parsing_attempt_count"] == 0
    assert counters["broad_source_atom_scan_attempt_count"] == 0
    assert counters["hidden_target_locator_used_count"] == 0
    assert counters["expected_or_supporting_gold_text_used_count"] == 0
    assert counters["source_file_title_shortcut_used_count"] == 0
    assert counters["direct_answer_value_matching_used_count"] == 0
    assert counters["full_page_dump_used_count"] == 0
    assert counters["agent_tool_layer_policy_violation_count"] == 0
    assert counters["family_router_invoked_count"] >= counters["layered_retrieval_audit_row_count"]
    assert counters["sourceatom_hydration_tool_invoked_count"] >= counters["layered_retrieval_audit_row_count"]
    assert counters["evidencebundle_builder_invoked_count"] >= counters["layered_retrieval_audit_row_count"]
    assert counters["citation_renderer_invoked_count"] >= counters["layered_retrieval_audit_row_count"]
    assert report["architecture_compliance_audit"]["layered_retrieval_architecture_preserved"] is True
    assert report["agent_tooling_audit"]["unsafe_shortcut_blocked_count"] >= 0
    assert counters["official_metric_input_rows"] == 0
    assert counters["silver_official_metric_input_rows"] == 0
    assert _sha256_file(source_packet) == source_packet_sha_before
    event = v4712.status_event(report, artifact_hashes={"report_json_sha256": "0" * 64})
    assert event["schema_version"] == f"{V4_7_12_SHORT_RUN_ID}_status_event_v1"
    assert event["logical_run_key"] == V4_7_12_SHORT_KEY
    assert event["short_run_id"] == V4_7_12_SHORT_RUN_ID
    assert event["non_production"] is True
    assert event["silver_topk_found"] == counters["silver_topk_found"]


def test_v4712_silver_reconnect_runs_retrieval_only_audit_or_fails_closed_without_promotion() -> None:
    from ai.eval import rag_v4712_layered_retrieval_generalization_and_overfit_audit as v4712

    report = v4712.build_report(
        root=ROOT,
        execute=False,
        sync_surfaces=False,
        env={},
        generated_at="2026-05-31T00:00:00Z",
    )
    v4712.check_report(report)
    audit = report["silver_layered_retrieval_audit"]
    counters = report["counters"]

    assert audit["diagnostic_silver_only"] is True
    assert audit["silver_regenerated"] is False
    assert counters["silver_promoted_to_gold_count"] == 0
    assert counters["silver_official_metric_input_rows"] == 0
    assert counters["official_metric_input_rows"] == 0
    if counters["silver_manifest_found"]:
        assert counters["silver_total_row_count"] == 1000
        assert counters["silver_topk_found"] is True
        assert counters["silver_unique_id_count"] == 1000
        assert counters["silver_unique_query_hash_count"] == 1000
        assert counters["silver_text_count"] == 350
        assert counters["silver_pdf_count"] == 325
        assert counters["silver_xlsx_count"] == 325
        assert counters["silver_core_count"] == 665
        assert counters["silver_review_only_count"] == 335
        assert counters["silver_quarantine_count"] == 0
        assert counters["silver_retrieval_audit_row_count"] == 1000
        assert counters["silver_query_hash_unique_count"] == 1000
        assert counters["silver_duplicate_query_hash_count"] == 0
        assert counters["silver_likely_unanswerable_count"] == 0
        assert audit["status"] == "SILVER_LAYERED_RETRIEVAL_AUDIT_COMPLETED_DIAGNOSTIC_ONLY"
        assert audit["audit_rows_total"] == 1000
        assert len(audit["audit_rows"]) == 1000
        assert all("too_broad_query" in row for row in audit["audit_rows"])
        assert all("likely_unanswerable" in row for row in audit["audit_rows"])
        assert not any(
            row["likely_unanswerable"]
            for row in audit["audit_rows"]
            if row.get("weak_answerability_status") == "auto_weak_silver_likely_answerable"
        )
        assert counters["silver_family_route_selected_count_by_family"]["TEXT"] == 350
        assert counters["silver_family_route_selected_count_by_family"]["PDF"] == 325
        assert counters["silver_family_route_selected_count_by_family"]["XLSX"] == 325
        assert counters["silver_sourceatom_hydration_success_count_by_family"]["PDF"] >= 0
        assert counters["silver_evidencebundle_created_count_by_family"]["XLSX"] >= 0
        assert counters["silver_citation_render_success_count_by_family"]["TEXT"] >= 0
        assert counters["silver_manifest_sha256"]
    else:
        assert audit["status"] == "SILVER_SOURCE_ARTIFACTS_UNAVAILABLE_FAIL_CLOSED"
        assert counters["silver_total_row_count"] == 0
        assert audit["artifact_resolution_evidence"]["searched_paths"]
        assert audit["artifact_resolution_evidence"]["archive_manifest_hints"]


def test_v4712_silver_reconnect_fails_closed_when_topk_artifact_is_missing(monkeypatch) -> None:
    from ai.eval import rag_v4712_layered_retrieval_generalization_and_overfit_audit as v4712

    def missing_topk(root: Path) -> tuple[list[dict[str, object]], dict[str, object]]:
        logical_path = v4712.V3_7_2_TOPK_ROWS.as_posix()
        return [], {
            "found": False,
            "logical_path": logical_path,
            "sha256": "",
            "artifact_resolution_evidence": {
                "searched_paths": [{"path": logical_path, "resolved_exists": False}],
                "archive_manifest_hints": [],
            },
        }

    monkeypatch.setattr(v4712, "_load_v3_7_2_topk", missing_topk)
    report = v4712.build_report(
        root=ROOT,
        execute=False,
        sync_surfaces=False,
        env={},
        generated_at="2026-05-31T00:00:00Z",
    )
    v4712.check_report(report)
    audit = report["silver_layered_retrieval_audit"]
    counters = report["counters"]

    assert counters["silver_manifest_found"] is True
    assert counters["silver_topk_found"] is False
    assert counters["silver_total_row_count"] == 1000
    assert counters["silver_retrieval_audit_row_count"] == 0
    assert counters["layered_retrieval_audit_row_count"] == counters["pdf_full_replay_eligible_count"]
    assert audit["status"] == "SILVER_TOPK_ARTIFACT_UNAVAILABLE_FAIL_CLOSED"
    assert audit["blocked_reason"] == "exact v3_7_2 row-level retrieval top-k artifact unavailable or sha verification failed"
    assert report["completion_branch"] == "B_silver_unavailable_layered_retrieval_audit_fail_closed"


def test_v4712_full_pdf_replay_and_silver_smoke_are_env_gated_without_fake_answers() -> None:
    from ai.eval import rag_v4712_layered_retrieval_generalization_and_overfit_audit as v4712

    disabled = v4712.build_report(
        root=ROOT,
        execute=False,
        sync_surfaces=False,
        env={},
        generated_at="2026-05-31T00:00:00Z",
    )
    v4712.check_report(disabled)
    counters = disabled["counters"]
    assert counters["pdf_full_replay_env_enabled"] is False
    assert counters["pdf_llm_invoked_count"] == 0
    assert counters["pdf_generated_response_count"] == 0
    assert counters["silver_llm_smoke_env_enabled"] is False
    assert counters["silver_llm_invoked_count"] == 0
    assert counters["silver_generated_response_count"] == 0
    assert disabled["full_pdf_llm_replay"]["status"] == "FULL_PDF_LLM_REPLAY_DISABLED_FAIL_CLOSED"
    assert disabled["silver_answer_smoke"]["status"] in {
        "SILVER_LLM_SMOKE_DISABLED_FAIL_CLOSED",
        "SILVER_LLM_SMOKE_SOURCE_UNAVAILABLE_FAIL_CLOSED",
    }

    mutated = json.loads(json.dumps(disabled))
    mutated["full_pdf_llm_replay"]["rows"] = [{"final_answer": "fake deterministic answer"}]
    mutated["counters"]["pdf_generated_response_count"] = 1
    try:
        v4712.check_report(mutated)
    except ValueError as exc:
        assert "full PDF replay counted answers while replay was disabled" in str(exc)
    else:
        raise AssertionError("v4_7_12 accepted fake full PDF LLM answers")


def test_v4712_silver_llm_smoke_runs_bounded_balanced_when_enabled(monkeypatch) -> None:
    from ai.eval import rag_v4712_layered_retrieval_generalization_and_overfit_audit as v4712

    def fake_probe(*, execute: bool, env: object) -> dict[str, object]:
        return {
            "available": bool(execute),
            "status": "LOCAL_LLM_AVAILABLE_DIAGNOSTIC_ONLY",
            "backend": "injected-test-client",
            "base_url_redacted": "injected",
            "model": "injected",
            "blockers": [],
        }

    def fake_client(prompt: str) -> str:
        payload = json.loads(prompt)
        evidence = str(payload.get("evidence") or "")
        citation_id = str(payload.get("citation_id") or "")
        terms = [token for token in re.findall(r"[가-힣A-Za-z0-9]{2,}", evidence) if token]
        head = " ".join(terms[:2]) or "근거 내용"
        return json.dumps(
            {
                "final_answer": f"근거에 따르면 {head} 관련 내용입니다.",
                "answer_type": "answer",
                "citations": [citation_id],
            },
            ensure_ascii=False,
            sort_keys=True,
        )

    monkeypatch.setattr(v4712, "_local_llm_probe", fake_probe)
    report = v4712.build_report(
        root=ROOT,
        execute=True,
        sync_surfaces=False,
        env={"RAG_V4_7_12_ENABLE_SILVER_LLM_SMOKE": "1"},
        llm_client=fake_client,
        generated_at="2026-05-31T00:00:00Z",
    )
    v4712.check_report(report)
    counters = report["counters"]

    assert counters["silver_llm_smoke_env_enabled"] is True
    assert counters["silver_llm_smoke_sample_count"] == 90
    assert counters["silver_llm_smoke_text_count"] == 30
    assert counters["silver_llm_smoke_pdf_count"] == 30
    assert counters["silver_llm_smoke_xlsx_count"] == 30
    assert counters["silver_llm_invoked_count"] == 90
    assert counters["silver_generated_response_count"] == 90
    assert counters["silver_citation_rendered_count"] == 90
    assert counters["silver_official_metric_input_rows"] == 0
    assert counters["silver_promoted_to_gold_count"] == 0
    assert len(report["silver_answer_smoke"]["rows"]) == 90
    assert all(row["raw_response_sha256"] for row in report["silver_answer_smoke"]["rows"])


def test_v4716_status_docs_do_not_leave_stale_current_alias_text() -> None:
    from ai.eval import rag_eval_registry as registry
    import ai.scripts.rag_eval as runner

    assert runner.DEFAULT_RUN_KEY == V5_6_SHORT_KEY
    assert registry.resolve_run("current", root=ROOT).logical_key == V5_6_SHORT_KEY
    for key in (
        V5_6_2_SHORT_KEY,
        V5_6_SHORT_KEY,
        V5_5_SHORT_KEY,
        V5_4_SHORT_KEY,
        V5_3_SHORT_KEY,
        V5_2_SHORT_KEY,
        V5_1_SHORT_KEY,
        V5_0_SHORT_KEY,
        V4_7_18_SHORT_KEY,
        V4_7_17_SHORT_KEY,
        V4_7_16_SHORT_KEY,
    ):
        assert registry.resolve_run(key, root=ROOT).logical_key == key


def test_v4716_summary_replacement_removes_prior_v47_current_blocks() -> None:
    from ai.eval import rag_v4716_target_recall_repair_prototype as v4716

    prior_summary = (
        "<!-- v4_7_15_summary_start -->\n"
        "## Current RAG Diagnostic Status\n"
        f"Current RAG status: `{V4_7_15_STATUS}`.\n"
        "`current` resolves to `v4_7_15`.\n"
        "<!-- v4_7_15_summary_end -->\n"
        "# Existing README Body\n"
    )
    replacement = (
        "## Current RAG Diagnostic Status\n"
        f"Current RAG status: `{V4_7_16_STATUS}`.\n"
        "`current` resolves to `v4_7_16`."
    )

    updated = v4716._replace_summary_block(prior_summary, block=replacement)

    assert "<!-- v4_7_16_summary_start -->" in updated
    assert f"Current RAG status: `{V4_7_16_STATUS}`" in updated
    assert f"Current RAG status: `{V4_7_15_STATUS}`" not in updated
    assert "<!-- v4_7_15_summary_start -->" not in updated
    assert "`current` resolves to `v4_7_15`" not in updated
    assert "# Existing README Body" in updated


def test_v4713_live_retrieval_disabled_fails_closed_and_preserves_v4712_artifacts() -> None:
    from ai.eval import rag_v4713_live_retrieval_answerability_and_full_pdf_replay as v4713

    v4712_artifacts = [
        ROOT / "ai/eval/reports/rag-ingestion/runs/v4_7_12/report.json",
        ROOT / "ai/eval/reports/rag-ingestion/runs/v4_7_12/silver_layered_retrieval_audit.json",
        ROOT / "ai/eval/reports/rag-ingestion/runs/v4_7_12/silver_answer_smoke_ko.jsonl",
    ]
    before = {path: _sha256_file(path) for path in v4712_artifacts}
    report = v4713.build_report(
        root=ROOT,
        execute=False,
        sync_surfaces=False,
        env={},
        generated_at="2026-05-31T00:00:00Z",
    )
    v4713.check_report(report)
    counters = report["counters"]

    assert report["short_run_id"] == V4_7_13_SHORT_RUN_ID
    assert report["source_run_id"] == V4_7_12_SHORT_RUN_ID
    assert report["diagnostic_only"] is True
    assert report["non_production"] is True
    assert report["official_metric"] is False
    assert counters["current_resolves_to"] == V4_7_13_SHORT_KEY
    assert counters["official_metric_input_rows"] == 0
    assert counters["silver_official_metric_input_rows"] == 0
    assert counters["silver_promoted_to_gold_count"] == 0
    assert counters["protected_namespaces_touched"] == []
    assert counters["live_silver_retrieval_env_enabled"] is False
    assert counters["live_silver_retrieval_row_count"] == 0
    assert report["live_silver_retrieval_replay"]["status"] == "LIVE_SILVER_RETRIEVAL_REPLAY_DISABLED_FAIL_CLOSED"
    assert counters["live_vector_payload_evidence_truth_violation_count"] == 0
    assert counters["live_raw_pdf_query_time_parsing_attempt_count"] == 0
    assert counters["live_raw_xlsx_query_time_parsing_attempt_count"] == 0
    assert counters["live_source_title_shortcut_used_count"] == 0
    assert counters["live_direct_answer_value_matching_used_count"] == 0
    assert counters["live_hidden_target_locator_used_count"] == 0
    assert counters["live_expected_or_supporting_gold_text_used_count"] == 0
    assert counters["pdf_full_replay_env_enabled"] is False
    assert counters["pdf_generated_response_count"] == 0
    assert report["full_pdf_llm_replay"]["status"] == "FULL_PDF_LLM_REPLAY_DISABLED_FAIL_CLOSED"
    assert {path: _sha256_file(path) for path in v4712_artifacts} == before


def test_v4713_full_pdf_replay_runs_all_answer_ready_rows_with_injected_llm(monkeypatch) -> None:
    from ai.eval import rag_v4713_live_retrieval_answerability_and_full_pdf_replay as v4713

    def fake_probe(*, execute: bool, env: object) -> dict[str, object]:
        return {
            "available": bool(execute),
            "status": "LOCAL_LLM_AVAILABLE_DIAGNOSTIC_ONLY",
            "backend": "injected-test-client",
            "base_url_redacted": "injected",
            "model": "injected",
            "blockers": [],
        }

    prompts: list[str] = []

    def fake_client(prompt: str) -> str:
        prompts.append(prompt)
        payload = json.loads(prompt)
        citation_id = str(payload.get("citation_id") or "")
        evidence = str(payload.get("bounded_evidence_excerpt") or "")
        terms = [token for token in re.findall(r"[가-힣A-Za-z0-9]{2,}", evidence) if token]
        head = " ".join(terms[:2]) or "근거 내용"
        return json.dumps(
            {
                "final_answer": f"근거에 따르면 {head} 관련 내용입니다.",
                "answer_type": "answer",
                "citations": [citation_id],
                "unsupported_claim_risk": False,
                "evidence_underuse_flag": False,
            },
            ensure_ascii=False,
            sort_keys=True,
        )

    monkeypatch.setattr(v4713, "_local_llm_probe", fake_probe)
    report = v4713.build_report(
        root=ROOT,
        execute=True,
        sync_surfaces=False,
        env={"RAG_V4_7_13_ENABLE_FULL_PDF_LLM_REPLAY": "1"},
        llm_client=fake_client,
        generated_at="2026-05-31T00:00:00Z",
    )
    v4713.check_report(report)
    counters = report["counters"]

    assert counters["pdf_full_replay_env_enabled"] is True
    assert counters["pdf_full_replay_eligible_count"] == 57
    assert counters["pdf_full_replay_excluded_weak_residual_count"] == 1
    assert counters["pdf_llm_invoked_count"] == 57
    assert counters["pdf_generated_response_count"] == 57
    assert counters["pdf_parsed_final_answer_present_count"] == 57
    assert counters["pdf_citation_rendered_count"] == 57
    assert counters["pdf_citation_grounded_to_evidence_count"] == 57
    assert counters["pdf_claim_support_pass_count"] == 57
    assert counters["pdf_parser_fail_count"] == 0
    assert len(report["full_pdf_llm_replay"]["rows"]) == 57
    assert len(prompts) == 57
    assert all("gold" not in prompt.lower() and "expected" not in prompt.lower() for prompt in prompts)


def test_v4713_silver_answerability_overlay_explains_prior_smoke_without_promoting_silver() -> None:
    from ai.eval import rag_v4713_live_retrieval_answerability_and_full_pdf_replay as v4713

    report = v4713.build_report(
        root=ROOT,
        execute=False,
        sync_surfaces=False,
        env={},
        generated_at="2026-05-31T00:00:00Z",
    )
    v4713.check_report(report)
    counters = report["counters"]
    overlay = report["silver_answerability_overlay"]

    assert overlay["diagnostic_silver_only"] is True
    assert overlay["silver_regenerated"] is False
    assert counters["silver_answerability_overlay_row_count"] == counters["silver_prior_smoke_row_count"]
    assert counters["silver_prior_smoke_text_count"] == 30
    assert counters["silver_prior_smoke_pdf_count"] == 30
    assert counters["silver_prior_smoke_xlsx_count"] == 30
    assert counters["silver_prior_claim_support_pass_count"] == 60
    assert counters["silver_prior_claim_support_fail_count"] == 30
    assert counters["silver_likely_answerable_but_answer_failed_count_by_family"]["TEXT"] >= 1
    assert counters["silver_answer_parser_fail_count_by_family"]["PDF"] == 1
    assert counters["silver_official_metric_input_rows"] == 0
    assert counters["silver_promoted_to_gold_count"] == 0
    assert len(overlay["rows"]) == counters["silver_answerability_overlay_row_count"]
    assert all(row["query_id"] and row["source_family"] in {"TEXT", "PDF", "XLSX"} for row in overlay["rows"])


def test_v4713_tooling_counters_split_retrieval_and_answer_generation_scopes() -> None:
    from ai.eval import rag_v4713_live_retrieval_answerability_and_full_pdf_replay as v4713

    report = v4713.build_report(
        root=ROOT,
        execute=False,
        sync_surfaces=False,
        env={},
        generated_at="2026-05-31T00:00:00Z",
    )
    v4713.check_report(report)
    counters = report["counters"]

    assert report["retrieval_tooling_audit"]["scope"] == "retrieval_only"
    assert report["answer_generation_tooling_audit"]["scope"] == "answer_generation_only"
    assert counters["retrieval_tooling_family_router_invoked_count"] >= 1000
    assert counters["retrieval_tooling_sourceatom_hydration_invoked_count"] >= 1000
    assert counters["retrieval_tooling_evidencebundle_builder_invoked_count"] >= 1000
    assert counters["retrieval_tooling_citation_renderer_invoked_count"] >= 1000
    assert counters["answer_generation_tooling_llm_invoked_count"] == counters["pdf_llm_invoked_count"]
    assert counters["answer_generation_tooling_parser_invoked_count"] == counters["pdf_parsed_final_answer_present_count"]
    assert counters["answer_generation_tooling_claim_verifier_invoked_count"] == (
        counters["pdf_claim_support_pass_count"] + counters["pdf_claim_support_fail_count"]
    )
    assert counters["tooling_counter_scope_mismatch_count"] == 0


def test_v4713_enabled_live_retrieval_without_safe_runtime_fails_closed() -> None:
    from ai.eval import rag_v4713_live_retrieval_answerability_and_full_pdf_replay as v4713

    report = v4713.build_report(
        root=ROOT,
        execute=True,
        sync_surfaces=False,
        env={"RAG_V4_7_13_ENABLE_LIVE_SILVER_RETRIEVAL_REPLAY": "1"},
        generated_at="2026-05-31T00:00:00Z",
    )
    v4713.check_report(report)
    counters = report["counters"]
    live = report["live_silver_retrieval_replay"]

    assert counters["live_silver_retrieval_env_enabled"] is True
    assert counters["live_silver_retrieval_row_count"] == 0
    assert live["status"] == "LIVE_SILVER_RETRIEVAL_REPLAY_UNAVAILABLE_FAIL_CLOSED"
    assert "no configured read-only live SearchIndexContract" in live["blocked_reason"]
    assert live["protected_namespaces_touched"] == []
    assert live["index_rebuilt"] is False
    assert live["source_registry_mutated"] is False
    assert live["cache_mutated"] is False


def test_v4713_check_report_rejects_promotional_or_mutating_flags() -> None:
    from ai.eval import rag_v4713_live_retrieval_answerability_and_full_pdf_replay as v4713

    report = v4713.build_report(
        root=ROOT,
        execute=False,
        sync_surfaces=False,
        env={},
        generated_at="2026-05-31T00:00:00Z",
    )
    v4713.check_report(report)

    mutations = [
        ("official_metric", True, "opened forbidden gate"),
        ("promotion_evidence", True, "opened forbidden gate"),
        ("gold_mutation", True, "opened forbidden gate"),
    ]
    for key, value, expected in mutations:
        mutated = json.loads(json.dumps(report))
        mutated[key] = value
        try:
            v4713.check_report(mutated)
        except ValueError as exc:
            assert expected in str(exc)
        else:
            raise AssertionError(f"v4_7_13 check_report accepted {key}={value}")

    counter_mutated = json.loads(json.dumps(report))
    counter_mutated["counters"]["silver_promoted_to_gold_count"] = 1
    try:
        v4713.check_report(counter_mutated)
    except ValueError as exc:
        assert "promoted silver" in str(exc)
    else:
        raise AssertionError("v4_7_13 check_report accepted silver promotion")


def test_v4713_full_pdf_replay_fail_closed_lanes_reject_fake_answer_payloads() -> None:
    from ai.eval import rag_v4713_live_retrieval_answerability_and_full_pdf_replay as v4713

    report = v4713.build_report(
        root=ROOT,
        execute=False,
        sync_surfaces=False,
        env={},
        generated_at="2026-05-31T00:00:00Z",
    )
    v4713.check_report(report)
    mutated = json.loads(json.dumps(report))
    mutated["full_pdf_llm_replay"]["rows"] = [
        {
            "query_id": "fake",
            "final_answer": "가짜 답변",
            "citations": ["evidence_1_fake"],
            "diagnostic_only": True,
        }
    ]
    try:
        v4713.check_report(mutated)
    except ValueError as exc:
        assert "disabled full PDF replay carried answer rows" in str(exc)
    else:
        raise AssertionError("v4_7_13 check_report accepted disabled fake full-PDF rows")

    live_mutated = json.loads(json.dumps(report))
    live_mutated["live_silver_retrieval_replay"]["rows"] = [{"query_id": "fake"}]
    try:
        v4713.check_report(live_mutated)
    except ValueError as exc:
        assert "fail-closed live retrieval replay carried rows" in str(exc)
    else:
        raise AssertionError("v4_7_13 check_report accepted fail-closed live retrieval rows")


def test_v4713_silver_answerability_overlay_is_diagnostic_not_label_or_qrels() -> None:
    from ai.eval import rag_v4713_live_retrieval_answerability_and_full_pdf_replay as v4713

    report = v4713.build_report(
        root=ROOT,
        execute=False,
        sync_surfaces=False,
        env={},
        generated_at="2026-05-31T00:00:00Z",
    )
    v4713.check_report(report)
    forbidden_keys = {
        "gold",
        "qrels",
        "label",
        "expected_answer",
        "supporting_evidence",
        "source_file_title",
        "source_path",
    }

    overlay = report["silver_answerability_overlay"]
    assert overlay["diagnostic_silver_only"] is True
    assert overlay["official_metric_input_rows"] == 0
    assert overlay["silver_promoted_to_gold_count"] == 0
    assert all(not (forbidden_keys & set(row)) for row in overlay["rows"])

    mutated = json.loads(json.dumps(report))
    mutated["silver_answerability_overlay"]["silver_regenerated"] = True
    try:
        v4713.check_report(mutated)
    except ValueError as exc:
        assert "diagnostic silver only" in str(exc)
    else:
        raise AssertionError("v4_7_13 check_report accepted regenerated silver overlay")


def test_v4713_check_report_rejects_tooling_counter_scope_mismatch() -> None:
    from ai.eval import rag_v4713_live_retrieval_answerability_and_full_pdf_replay as v4713

    report = v4713.build_report(
        root=ROOT,
        execute=False,
        sync_surfaces=False,
        env={},
        generated_at="2026-05-31T00:00:00Z",
    )
    v4713.check_report(report)

    counter_mutated = json.loads(json.dumps(report))
    counter_mutated["counters"]["tooling_counter_scope_mismatch_count"] = 1
    try:
        v4713.check_report(counter_mutated)
    except ValueError as exc:
        assert "unsafe or inconsistent counter nonzero" in str(exc)
    else:
        raise AssertionError("v4_7_13 check_report accepted a tooling mismatch counter")

    scope_mutated = json.loads(json.dumps(report))
    scope_mutated["answer_generation_tooling_audit"]["scope"] = "retrieval_only"
    try:
        v4713.check_report(scope_mutated)
    except ValueError as exc:
        assert "answer tooling scope mismatch" in str(exc)
    else:
        raise AssertionError("v4_7_13 check_report accepted an answer tooling scope mismatch")


def test_v4714_reclassifies_unavailable_runtime_preconditions_without_quality_failures() -> None:
    from ai.eval import rag_v4714_diagnostic_precondition_hardening as v4714

    report = v4714.build_report(root=ROOT, generated_at="2026-05-31T00:00:00Z")
    v4714.check_report(report)
    counters = report["counters"]
    live = report["live_retrieval_preflight"]
    llm = report["local_llm_preflight"]

    assert report["short_run_id"] == V4_7_14_SHORT_RUN_ID
    assert report["source_run_id"] == V4_7_13_SHORT_RUN_ID
    assert counters["current_resolves_to"] == V4_7_14_SHORT_KEY
    assert live["status"] == "LIVE_RETRIEVAL_PRECONDITION_UNAVAILABLE_FAIL_CLOSED"
    assert live["read_only_search_index_contract_available"] is False
    assert live["source_candidate_row_count"] == 1000
    assert live["attempted_row_count"] == 0
    assert live["quality_evaluated_row_count"] == 0
    assert live["row_count"] == 0
    assert live["retrieval_quality_failure_count"] == 0
    assert live["precondition_unavailable_count"] == 1
    assert live["not_evaluated_count_by_reason"]["read_only_live_SearchIndexContract_unavailable"] == 1000
    assert live["blocked_source_denominator_by_family"] == {"TEXT": 350, "PDF": 325, "XLSX": 325}
    assert counters["live_retrieval_precondition_unavailable_count"] == 1
    assert counters["live_retrieval_quality_failure_count"] == 0
    assert counters["live_retrieval_quality_failure_count_by_family"] == {"TEXT": 0, "PDF": 0, "XLSX": 0}
    assert live["production_db_mutated"] is False
    assert live["cache_mutated"] is False
    assert live["source_registry_mutated"] is False
    assert live["silver_mutated"] is False
    assert live["index_rebuilt"] is False

    assert llm["status"] == "LOCAL_LLM_UNAVAILABLE_GENERATION_NOT_ATTEMPTED_FAIL_CLOSED"
    assert llm["eligible_count"] == 57
    assert llm["source_candidate_row_count"] == 57
    assert llm["attempted_row_count"] == 0
    assert llm["quality_evaluated_row_count"] == 0
    assert llm["llm_unavailable_skip_count"] == 57
    assert llm["generated_response_count"] == 0
    assert llm["parser_failure_count"] == 0
    assert llm["claim_support_fail_count"] == 0
    assert llm["citation_failure_count"] == 0
    assert llm["unsupported_answer_count"] == 0
    assert llm["claim_support_not_evaluated_due_to_no_generation_count"] == 57
    assert llm["parser_not_evaluated_count"] == 57
    assert llm["citation_not_evaluated_count"] == 57
    assert llm["noop_or_extractive_fallback_answer_count"] == 0
    assert llm["raw_prompt_payload_written"] is False
    assert llm["raw_response_payload_written"] is False
    assert counters["llm_unavailable_skip_count"] == 57
    assert counters["claim_support_not_evaluated_due_to_no_generation_count"] == 57
    assert counters["claim_support_fail_count"] == 0
    assert counters["parser_failure_count"] == 0
    assert counters["citation_failure_count"] == 0
    assert counters["unsupported_answer_count"] == 0
    assert counters["generated_response_count"] == 0
    assert counters["noop_or_extractive_fallback_answer_count"] == 0
    assert report["full_pdf_generation_rows"] == []


def test_v4714_silver_overlay_root_cause_queues_are_diagnostic_only() -> None:
    from ai.eval import rag_v4714_diagnostic_precondition_hardening as v4714

    report = v4714.build_report(root=ROOT, generated_at="2026-05-31T00:00:00Z")
    v4714.check_report(report)
    queues = report["silver_answerability_root_cause_queues"]

    assert queues["status"] == "SILVER_ANSWERABILITY_ROOT_CAUSE_QUEUES_READY_DIAGNOSTIC_ONLY"
    assert queues["row_count"] == 90
    assert queues["diagnostic_silver_only"] is True
    assert queues["silver_mutation"] is False
    assert queues["gold_mutation"] is False
    assert queues["qrels_mutation"] is False
    assert queues["label_mutation"] is False
    assert queues["expected_answer_mutation"] is False
    assert queues["supporting_evidence_mutation"] is False
    assert queues["denominator_mutation"] is False
    assert queues["silver_promoted_to_gold_count"] == 0
    assert queues["official_metric_input_rows"] == 0

    text = queues["queues_by_family"]["TEXT"]["root_cause_counts"]
    xlsx = queues["queues_by_family"]["XLSX"]["root_cause_counts"]
    pdf = queues["queues_by_family"]["PDF"]["root_cause_counts"]
    assert text["target_not_in_topk"] == 28
    assert text["evidence_mismatch_after_family_route"] == 30
    assert xlsx["target_not_in_topk"] == 28
    assert xlsx["repeated_prefix_cluster"] == 22
    assert pdf["evidence_window_insufficient"] == 16
    assert pdf["source_family_route_ok_but_evidence_mismatch"] == 17
    assert pdf["query_too_broad"] == 5
    assert queues["mutation_policy"] == [
        "do_not_modify_silver_rows",
        "do_not_modify_gold_qrels_labels_expected_or_supporting_evidence",
        "do_not_modify_denominator_rows",
    ]


def test_v4714_written_report_status_docs_and_no_raw_payload_leakage() -> None:
    from ai.eval import rag_eval_registry as registry
    from ai.eval import rag_v4714_diagnostic_precondition_hardening as v4714

    report = registry.load_report("v4_7_14", root=ROOT)
    v4714.check_report(report)
    latest = next(row for row in reversed(_read_jsonl(STATUS_JSONL)) if row.get("short_run_id") == V4_7_14_SHORT_RUN_ID)
    progress = PROGRESS_DOC.read_text(encoding="utf-8")
    measurements = MEASUREMENTS_DOC.read_text(encoding="utf-8")
    triage = TRIAGE_DOC.read_text(encoding="utf-8")
    current_progress = progress.split("## Short History", 1)[0]

    assert V4_7_14_REPORT.exists()
    assert latest["status"] == V4_7_14_STATUS
    assert latest["artifact_paths"]["report_json"] == "ai/eval/reports/rag-ingestion/runs/v4_7_14/report.json"
    assert latest["artifact_sha256"]["report_json_sha256"] == _sha256_file(V4_7_14_REPORT)
    assert latest["live_retrieval_precondition_unavailable_count"] == 1
    assert latest["live_retrieval_quality_failure_count"] == 0
    assert latest["llm_unavailable_skip_count"] == 57
    assert latest["claim_support_fail_count"] == 0
    assert latest["parser_failure_count"] == 0
    assert latest["citation_failure_count"] == 0
    assert latest["unsupported_answer_count"] == 0

    assert V4_7_14_SHORT_RUN_ID in progress
    assert V4_7_14_SHORT_RUN_ID in measurements
    assert V4_7_14_SHORT_RUN_ID in triage
    assert "artifact-ready / fail-closed diagnostic-ready" in progress
    assert "LIVE_RETRIEVAL_PRECONDITION_UNAVAILABLE_FAIL_CLOSED" in measurements
    assert "LOCAL_LLM_UNAVAILABLE_GENERATION_NOT_ATTEMPTED_FAIL_CLOSED" in measurements
    assert "target_not_in_topk 28" in triage
    assert "repeated_prefix_cluster 22" in triage
    assert "evidence_window_insufficient 16" in triage
    _assert_status_recorded_in_progress_and_report(
        progress,
        report,
        short_run_id=V4_7_14_SHORT_RUN_ID,
        status=V4_7_14_STATUS,
    )
    assert registry.resolve_run("current", root=ROOT).logical_key == V5_6_SHORT_KEY
    assert registry.resolve_run(V4_7_18_SHORT_KEY, root=ROOT).logical_key == V4_7_18_SHORT_KEY
    assert registry.resolve_run(V4_7_17_SHORT_KEY, root=ROOT).logical_key == V4_7_17_SHORT_KEY
    assert registry.resolve_run(V4_7_16_SHORT_KEY, root=ROOT).logical_key == V4_7_16_SHORT_KEY

    def assert_no_raw_payload_keys(value: object) -> None:
        if isinstance(value, dict):
            forbidden = {"prompt", "raw_prompt", "raw_response", "response", "raw_llm_response", "final_answer"}
            assert not (forbidden & set(value)), forbidden & set(value)
            for child in value.values():
                assert_no_raw_payload_keys(child)
        elif isinstance(value, list):
            for child in value:
                assert_no_raw_payload_keys(child)

    assert_no_raw_payload_keys(report)
    generated_text = "\n".join(
        [
            json.dumps(report, ensure_ascii=False),
            json.dumps(latest, ensure_ascii=False),
            progress.split("<!-- v4_7_14_diagnostic_precondition_hardening:progress-entry:start -->", 1)[1].split(
                "<!-- v4_7_14_diagnostic_precondition_hardening:progress-entry:end -->",
                1,
            )[0],
            measurements.split("<!-- v4_7_14_measurements_start -->", 1)[1].split(
                "<!-- v4_7_14_measurements_end -->",
                1,
            )[0],
            triage.split("<!-- v4_7_14_triage_start -->", 1)[1].split(
                "<!-- v4_7_14_triage_end -->",
                1,
            )[0],
        ]
    )
    for pattern in (
        r"\b[A-Z]:[\\/]",
        r"prompt_payload\":\s*[{[]",
        r"raw_response_payload\":\s*[{[]",
        r"promotion-ready",
        r"product-success",
    ):
        assert not re.search(pattern, generated_text), pattern


def test_v4714_check_report_rejects_precondition_misclassified_as_quality_failure() -> None:
    from ai.eval import rag_v4714_diagnostic_precondition_hardening as v4714

    report = v4714.build_report(root=ROOT, generated_at="2026-05-31T00:00:00Z")
    v4714.check_report(report)

    for key in (
        "live_retrieval_quality_failure_count",
        "claim_support_fail_count",
        "parser_failure_count",
        "citation_failure_count",
        "unsupported_answer_count",
        "generated_response_count",
        "noop_or_extractive_fallback_answer_count",
    ):
        mutated = json.loads(json.dumps(report))
        mutated["counters"][key] = 1
        try:
            v4714.check_report(mutated)
        except ValueError as exc:
            assert "precondition unavailable states must not be counted as quality failures" in str(exc)
        else:
            raise AssertionError(f"v4_7_14 accepted misclassified counter {key}")

    mutated_live = json.loads(json.dumps(report))
    mutated_live["live_retrieval_preflight"]["row_count"] = 1
    try:
        v4714.check_report(mutated_live)
    except ValueError as exc:
        assert "live precondition unavailable row_count must stay 0" in str(exc)
    else:
        raise AssertionError("v4_7_14 accepted live rows for unavailable SearchIndexContract")


def test_v4714_check_report_rejects_raw_payload_and_not_evaluated_counter_drift() -> None:
    from ai.eval import rag_v4714_diagnostic_precondition_hardening as v4714

    report = v4714.build_report(root=ROOT, generated_at="2026-05-31T00:00:00Z")
    v4714.check_report(report)

    for surface, key in (
        ("local_llm_preflight", "raw_prompt_payload_written"),
        ("local_llm_preflight", "raw_response_payload_written"),
        ("counters", "raw_prompt_payload_written"),
        ("counters", "raw_response_payload_written"),
    ):
        mutated = json.loads(json.dumps(report))
        mutated[surface][key] = True
        try:
            v4714.check_report(mutated)
        except ValueError as exc:
            assert "raw prompt/response payload" in str(exc)
        else:
            raise AssertionError(f"v4_7_14 accepted {surface}.{key}=True")

    drift_cases = [
        ("live_retrieval_preflight", "attempted_row_count", 1, "live precondition attempted/evaluated counts"),
        ("live_retrieval_preflight", "quality_evaluated_row_count", 1, "live precondition attempted/evaluated counts"),
        (
            "local_llm_preflight",
            "attempted_row_count",
            1,
            "local LLM unavailable attempted/evaluated counts",
        ),
        (
            "local_llm_preflight",
            "quality_evaluated_row_count",
            1,
            "local LLM unavailable attempted/evaluated counts",
        ),
        ("local_llm_preflight", "parser_not_evaluated_count", 0, "not-evaluated count must match eligible rows"),
        ("local_llm_preflight", "citation_not_evaluated_count", 0, "not-evaluated count must match eligible rows"),
    ]
    for surface, key, value, expected in drift_cases:
        mutated = json.loads(json.dumps(report))
        mutated[surface][key] = value
        try:
            v4714.check_report(mutated)
        except ValueError as exc:
            assert expected in str(exc)
        else:
            raise AssertionError(f"v4_7_14 accepted {surface}.{key}={value}")

    mutated = json.loads(json.dumps(report))
    mutated["live_retrieval_preflight"]["not_evaluated_count_by_reason"][
        "read_only_live_SearchIndexContract_unavailable"
    ] = 0
    try:
        v4714.check_report(mutated)
    except ValueError as exc:
        assert "live precondition not-evaluated count" in str(exc)
    else:
        raise AssertionError("v4_7_14 accepted live not-evaluated count drift")


def test_v4715_unblocks_archived_read_only_searchindex_replay_without_live_readiness() -> None:
    from ai.eval import rag_v4715_read_only_searchindex_replay_projection as v4715

    report = v4715.build_report(root=ROOT, generated_at="2026-05-31T00:00:00Z")
    v4715.check_report(report)
    replay = report["read_only_searchindexcontract_replay"]
    counters = report["counters"]

    assert report["short_run_id"] == V4_7_15_SHORT_RUN_ID
    assert report["source_run_id"] == V4_7_14_SHORT_RUN_ID
    assert report["diagnostic_only"] is True
    assert report["official_metric"] is False
    assert report["official_metric_input_rows"] == 0
    assert report["promotion_evidence"] is False
    assert report["product_success_evidence_allowed"] is False
    assert report["live_db_index_cache_readiness"] is False
    assert replay["status"] == "READ_ONLY_SEARCHINDEXCONTRACT_REPLAY_UNBLOCKED_ARCHIVED_TOPK_DIAGNOSTIC_ONLY"
    assert replay["source_topk_sha256_verified"] is True
    assert replay["source_topk_resolved_via_archive"] is True
    assert replay["v3_7_0_source_registry_manifest_record_count"] == 5
    assert replay["v3_7_1_index_manifest_record_count"] == 5
    assert replay["replay_input_row_count"] == 1000
    assert replay["replay_counts_by_family"] == {"TEXT": 350, "PDF": 325, "XLSX": 325}
    assert replay["topk_envelope_count"] == 5000
    assert replay["sourceatom_hydration_success_envelope_count"] == 5000
    assert replay["evidencebundle_renderable_envelope_count"] == 5000
    assert replay["citation_renderable_envelope_count"] == 5000
    assert replay["vector_payload_evidence_truth_violation_count"] == 0
    assert replay["live_runtime_adapter_invoked"] is False
    assert replay["index_rebuilt"] is False
    assert replay["source_registry_mutated"] is False
    assert replay["silver_mutated"] is False
    assert replay["production_db_mutated"] is False
    assert counters["read_only_replay_row_count"] == 1000
    assert counters["read_only_replay_topk_envelope_count"] == 5000
    assert counters["live_retrieval_quality_failure_count"] == 0


def test_v4715_repair_projection_uses_target_first_counts_and_overlap_matrix() -> None:
    from ai.eval import rag_v4715_read_only_searchindex_replay_projection as v4715

    report = v4715.build_report(root=ROOT, generated_at="2026-05-31T00:00:00Z")
    v4715.check_report(report)
    projection = report["diagnostic_retrieval_evidence_repair_projection"]
    primary = projection["primary_projection_counts"]
    overlap = projection["root_cause_overlap_matrix_by_family"]

    assert projection["status"] == "SILVER_RETRIEVAL_EVIDENCE_REPAIR_PROJECTION_READY_DIAGNOSTIC_ONLY"
    assert projection["projection_input_row_count"] == 90
    assert projection["projection_counts_by_family"] == {"TEXT": 30, "PDF": 30, "XLSX": 30}
    assert projection["projection_source_audit_row_count"] == 1000
    assert projection["projection_source_audit_counts_by_family"] == {"TEXT": 350, "PDF": 325, "XLSX": 325}
    assert projection["overlay_rows_missing_from_audit_count"] == 0
    assert primary["retrieval_target_not_in_topk"]["row_count"] == 68
    assert primary["retrieval_target_not_in_topk"]["counts_by_family"] == {"TEXT": 28, "PDF": 12, "XLSX": 28}
    assert primary["target_hit_evidence_context_repair"]["row_count"] == 14
    assert primary["target_hit_evidence_context_repair"]["counts_by_family"] == {"TEXT": 2, "PDF": 10, "XLSX": 2}
    assert primary["query_specificity_fixture_review"]["row_count"] == 3
    assert primary["query_specificity_fixture_review"]["counts_by_family"] == {"TEXT": 0, "PDF": 3, "XLSX": 0}
    assert primary["no_repair_projection"]["row_count"] == 5
    assert primary["no_repair_projection"]["counts_by_family"] == {"TEXT": 0, "PDF": 5, "XLSX": 0}
    assert projection["target_not_in_topk_count_by_family"] == {"TEXT": 28, "PDF": 12, "XLSX": 28}
    assert overlap["TEXT"]["target_not_in_topk_and_evidence_window_insufficient"] == 28
    assert overlap["XLSX"]["repeated_prefix_cluster_total"] == 22
    assert overlap["XLSX"]["repeated_prefix_cluster_overlap_with_target_miss"] == 20
    assert overlap["XLSX"]["repeated_prefix_cluster_target_hit"] == 2
    assert overlap["PDF"]["evidence_window_insufficient_total"] == 16
    assert overlap["PDF"]["evidence_window_insufficient_target_hit"] == 10
    assert overlap["PDF"]["source_family_route_ok_but_evidence_mismatch_target_hit"] == 10
    assert overlap["PDF"]["query_too_broad_total"] == 5
    assert overlap["PDF"]["query_too_broad_primary_review"] == 3


def test_v4715_written_report_status_docs_and_guardrails() -> None:
    from ai.eval import rag_eval_registry as registry
    from ai.eval import rag_v4715_read_only_searchindex_replay_projection as v4715

    report = registry.load_report("v4_7_15", root=ROOT)
    v4715.check_report(report)
    latest = next(row for row in reversed(_read_jsonl(STATUS_JSONL)) if row.get("short_run_id") == V4_7_15_SHORT_RUN_ID)
    progress = PROGRESS_DOC.read_text(encoding="utf-8")
    measurements = MEASUREMENTS_DOC.read_text(encoding="utf-8")
    triage = TRIAGE_DOC.read_text(encoding="utf-8")

    assert V4_7_15_REPORT.exists()
    assert latest["status"] == V4_7_15_STATUS
    assert latest["artifact_paths"]["report_json"] == "ai/eval/reports/rag-ingestion/runs/v4_7_15/report.json"
    assert latest["artifact_sha256"]["report_json_sha256"] == _sha256_file(V4_7_15_REPORT)
    assert latest["read_only_replay_row_count"] == 1000
    assert latest["projection_input_row_count"] == 90
    assert latest["retrieval_target_not_in_topk_projection_count"] == 68
    assert latest["target_hit_evidence_context_repair_projection_count"] == 14
    assert latest["official_metric_input_rows"] == 0
    assert report["protected_namespaces_touched"] == []
    assert latest["promotion_evidence"] is False
    assert latest["live_db_index_cache_readiness"] is False
    assert V4_7_15_SHORT_RUN_ID in progress
    assert V4_7_15_SHORT_RUN_ID in measurements
    assert V4_7_15_SHORT_RUN_ID in triage
    assert "READ_ONLY_SEARCHINDEXCONTRACT_REPLAY_UNBLOCKED_ARCHIVED_TOPK_DIAGNOSTIC_ONLY" in measurements
    assert "retrieval target not in top-k 68" in triage
    _assert_status_recorded_in_progress_and_report(
        progress,
        report,
        short_run_id=V4_7_15_SHORT_RUN_ID,
        status=V4_7_15_STATUS,
    )
    assert registry.resolve_run("current", root=ROOT).logical_key == V5_6_SHORT_KEY
    assert registry.resolve_run(V4_7_18_SHORT_KEY, root=ROOT).logical_key == V4_7_18_SHORT_KEY
    assert registry.resolve_run(V4_7_17_SHORT_KEY, root=ROOT).logical_key == V4_7_17_SHORT_KEY
    assert registry.resolve_run(V4_7_16_SHORT_KEY, root=ROOT).logical_key == V4_7_16_SHORT_KEY
    assert "promotion-ready" not in json.dumps(report, ensure_ascii=False)
    assert report["protected_namespaces_touched"] == []


def test_v4715_check_report_rejects_projection_drift_or_opened_gates() -> None:
    from ai.eval import rag_v4715_read_only_searchindex_replay_projection as v4715

    report = v4715.build_report(root=ROOT, generated_at="2026-05-31T00:00:00Z")
    v4715.check_report(report)

    mutations = [
        ("official_metric", True, "opened forbidden gate"),
        ("promotion_evidence", True, "opened forbidden gate"),
        ("live_db_index_cache_readiness", True, "opened forbidden gate"),
    ]
    for key, value, expected in mutations:
        mutated = json.loads(json.dumps(report))
        mutated[key] = value
        try:
            v4715.check_report(mutated)
        except ValueError as exc:
            assert expected in str(exc)
        else:
            raise AssertionError(f"v4_7_15 accepted {key}={value}")

    mutated = json.loads(json.dumps(report))
    mutated["read_only_searchindexcontract_replay"]["vector_payload_evidence_truth_violation_count"] = 1
    try:
        v4715.check_report(mutated)
    except ValueError as exc:
        assert "vector payload" in str(exc)
    else:
        raise AssertionError("v4_7_15 accepted vector payload evidence-truth violation")

    for key, value, expected in (
        ("read_only", False, "read-only"),
        ("diagnostic_only", False, "diagnostic-only"),
        ("canonical_payload_source_registry_envelope_count", 0, "source registry"),
    ):
        mutated = json.loads(json.dumps(report))
        mutated["read_only_searchindexcontract_replay"][key] = value
        try:
            v4715.check_report(mutated)
        except ValueError as exc:
            assert expected in str(exc)
        else:
            raise AssertionError(f"v4_7_15 accepted replay {key}={value}")

    mutated = json.loads(json.dumps(report))
    mutated["diagnostic_retrieval_evidence_repair_projection"]["diagnostic_silver_only"] = False
    try:
        v4715.check_report(mutated)
    except ValueError as exc:
        assert "diagnostic silver" in str(exc)
    else:
        raise AssertionError("v4_7_15 accepted diagnostic_silver_only=false")

    mutated = json.loads(json.dumps(report))
    mutated["diagnostic_retrieval_evidence_repair_projection"]["overlay_rows_missing_from_audit_count"] = 1
    try:
        v4715.check_report(mutated)
    except ValueError as exc:
        assert "overlay/audit join drift" in str(exc)
    else:
        raise AssertionError("v4_7_15 accepted overlay/audit join drift")


def test_v4716_candidate_only_target_recall_prototype_uses_no_oracle_inputs() -> None:
    from ai.eval import rag_v4716_target_recall_repair_prototype as v4716

    report = v4716.build_report(root=ROOT, generated_at="2026-05-31T00:00:00Z")
    v4716.check_report(report)
    prototype = report["target_recall_repair_prototype"]
    archive = prototype["archive_1000_candidate_only_target_recall"]
    families = archive["families"]
    construction = prototype["candidate_construction"]
    guardrails = report["anti_overfit_guardrails"]

    assert report["short_run_id"] == V4_7_16_SHORT_RUN_ID
    assert report["source_run_id"] == V4_7_15_SHORT_RUN_ID
    assert prototype["status"] == "TEXT_XLSX_TARGET_RECALL_REPAIR_PROTOTYPE_READY_DIAGNOSTIC_ONLY"
    assert prototype["candidate_budget_per_query"] == 5
    assert construction["diagnostic_target_labels_used_for_candidate_construction"] is False
    assert construction["diagnostic_target_labels_used_for_candidate_scoring"] is False
    assert construction["diagnostic_target_labels_used_for_after_the_fact_evaluation"] is True
    assert construction["raw_xlsx_query_time_parsing"] is False
    assert construction["direct_normalized_answer_value_matching"] is False
    assert construction["source_file_title_shortcut_used"] is False
    assert construction["hidden_locator_or_gold_field_use_count"] == 0
    assert "query_text" in construction["allowed_candidate_construction_fields"]["query"]
    assert "source_family" in construction["allowed_candidate_construction_fields"]["query"]
    assert "normalized_text_or_value_snapshot" in construction["allowed_candidate_construction_fields"]["source_registry"]["TEXT"]
    assert "raw_locator.row_label" in construction["allowed_candidate_construction_fields"]["source_registry"]["XLSX"]
    forbidden = json.dumps(construction["forbidden_candidate_construction_fields"], ensure_ascii=False)
    for token in (
        "target_source_atom_ids",
        "question_gold_locator_target",
        "official_manifest_target",
        "supporting_evidence",
        "expected_answer",
        "query_id",
        "case_id",
        "source_file_path",
        "source_path",
        "workbook",
        "normalized_value",
    ):
        assert token in forbidden

    assert archive["baseline_target_hit_count"] == 300
    assert archive["combined_target_hit_count"] == 514
    assert archive["baseline_miss_to_hit_count"] == 214
    assert archive["baseline_hit_to_miss_count"] == 0
    assert families["PDF"]["prototype_attempted_row_count"] == 0
    assert families["PDF"]["combined_target_hit_count"] == 265
    assert families["PDF"]["target_hit_regression_count"] == 0
    assert families["TEXT"]["baseline_target_hit_count"] == 20
    assert families["TEXT"]["prototype_candidate_count"] == 1714
    assert families["TEXT"]["combined_target_hit_count"] == 232
    assert families["TEXT"]["baseline_miss_to_hit_count"] == 212
    assert families["XLSX"]["baseline_target_hit_count"] == 15
    assert families["XLSX"]["prototype_candidate_count"] == 133
    assert families["XLSX"]["combined_target_hit_count"] == 17
    assert families["XLSX"]["baseline_miss_to_hit_count"] == 2

    assert prototype["repair_idea_decisions"]["accepted"][0]["idea_id"] == "TEXT_SAFE_LEXICAL_SEARCHUNIT_SEARCHVIEW_REPAIR"
    assert prototype["repair_idea_decisions"]["inconclusive"][0]["idea_id"] == "XLSX_SAFE_TABLE_AXIS_SEARCHUNIT_SEARCHVIEW_REPAIR"
    rejected = json.dumps(prototype["repair_idea_decisions"]["rejected"], ensure_ascii=False)
    assert "DIRECT_NORMALIZED_VALUE_MATCHING" in rejected
    assert "RAW_XLSX_QUERY_TIME_PARSING" in rejected
    assert prototype["overlay_90_root_cause_summary"]["primary_projection_counts"]["retrieval_target_not_in_topk"][
        "row_count"
    ] == 68
    assert prototype["overlay_90_root_cause_summary"]["primary_projection_counts"]["retrieval_target_not_in_topk"][
        "counts_by_family"
    ] == {"TEXT": 28, "PDF": 12, "XLSX": 28}
    for key, value in guardrails.items():
        if key.endswith("_allowed") or key.endswith("_used") or key.endswith("_created"):
            assert value is False, key
    assert guardrails["protected_namespaces_touched"] == []
    assert guardrails["official_metric_input_rows"] == 0


def test_v4716_candidate_set_digest_ignores_poisoned_target_fields() -> None:
    from ai.eval import rag_v4716_target_recall_repair_prototype as v4716

    rows, _resolution = v4716._load_silver_topk_rows(ROOT)
    baseline = v4716.build_candidate_only_repair_prototype(root=ROOT, silver_topk_rows=rows)
    poisoned = json.loads(json.dumps(rows, ensure_ascii=False))
    for row in poisoned:
        row["target_source_atom_ids"] = ["poisoned_target_atom"]
        row["target_hit_at_k"] = not bool(row.get("target_hit_at_k"))
        row["target_hit_in_topk"] = not bool(row.get("target_hit_in_topk"))
        row["question_gold_locator_target"] = {"poison": True}
        row["official_manifest_target"] = {"poison": True}
        row["query_id"] = f"poisoned-{row.get('query_id')}"
        row["case_id"] = "poisoned-case"
    mutated = v4716.build_candidate_only_repair_prototype(root=ROOT, silver_topk_rows=poisoned)

    assert baseline["candidate_set_sha256"] == mutated["candidate_set_sha256"]
    assert baseline["archive_1000_candidate_only_target_recall"]["candidate_set_sha256"] == mutated[
        "archive_1000_candidate_only_target_recall"
    ]["candidate_set_sha256"]
    assert baseline["archive_1000_candidate_only_target_recall"]["combined_target_hit_count"] != mutated[
        "archive_1000_candidate_only_target_recall"
    ]["combined_target_hit_count"]


def test_v4716_written_report_status_docs_and_guardrails() -> None:
    from ai.eval import rag_eval_registry as registry
    from ai.eval import rag_v4716_target_recall_repair_prototype as v4716

    report = registry.load_report("v4_7_16", root=ROOT)
    v4716.check_report(report)
    latest = next(row for row in reversed(_read_jsonl(STATUS_JSONL)) if row.get("short_run_id") == V4_7_16_SHORT_RUN_ID)
    progress = PROGRESS_DOC.read_text(encoding="utf-8")
    measurements = MEASUREMENTS_DOC.read_text(encoding="utf-8")
    triage = TRIAGE_DOC.read_text(encoding="utf-8")

    assert V4_7_16_REPORT.exists()
    assert latest["status"] == V4_7_16_STATUS
    assert latest["artifact_paths"]["report_json"] == "ai/eval/reports/rag-ingestion/runs/v4_7_16/report.json"
    assert latest["artifact_sha256"]["report_json_sha256"] == _sha256_file(V4_7_16_REPORT)
    assert latest["baseline_target_hit_count"] == 300
    assert latest["combined_target_hit_count"] == 514
    assert latest["text_baseline_miss_to_hit_count"] == 212
    assert latest["xlsx_baseline_miss_to_hit_count"] == 2
    assert latest["pdf_target_hit_regression_count"] == 0
    assert latest["official_metric_input_rows"] == 0
    assert latest["promotion_evidence"] is False
    assert latest["live_db_index_cache_readiness"] is False
    assert V4_7_16_SHORT_RUN_ID in progress
    assert V4_7_16_SHORT_RUN_ID in measurements
    assert V4_7_16_SHORT_RUN_ID in triage
    assert "TEXT_SAFE_LEXICAL_SEARCHUNIT_SEARCHVIEW_REPAIR" in triage
    assert "DIRECT_NORMALIZED_VALUE_MATCHING" in triage
    assert "combined_target_hit_count | 514" in measurements
    _assert_status_recorded_in_progress_and_report(
        progress,
        report,
        short_run_id=V4_7_16_SHORT_RUN_ID,
        status=V4_7_16_STATUS,
    )
    assert registry.resolve_run("current", root=ROOT).logical_key == V5_6_SHORT_KEY
    assert registry.resolve_run(V4_7_18_SHORT_KEY, root=ROOT).logical_key == V4_7_18_SHORT_KEY
    assert registry.resolve_run(V4_7_17_SHORT_KEY, root=ROOT).logical_key == V4_7_17_SHORT_KEY
    assert registry.resolve_run(V4_7_16_SHORT_KEY, root=ROOT).logical_key == V4_7_16_SHORT_KEY
    assert "promotion-ready" not in json.dumps(report, ensure_ascii=False)
    report_json = json.dumps(report, ensure_ascii=False)
    assert '"raw_prompt":' not in report_json
    assert '"raw_response":' not in report_json
    assert report["raw_prompt_payload_written"] is False
    assert report["raw_response_payload_written"] is False
    assert report["protected_namespaces_touched"] == []


def test_v4716_check_report_rejects_oracle_shortcuts_metric_drift_and_raw_payloads() -> None:
    from ai.eval import rag_v4716_target_recall_repair_prototype as v4716

    report = v4716.build_report(root=ROOT, generated_at="2026-05-31T00:00:00Z")
    v4716.check_report(report)

    for key, value, expected in (
        ("official_metric", True, "opened forbidden gate"),
        ("promotion_evidence", True, "opened forbidden gate"),
        ("live_db_index_cache_readiness", True, "opened forbidden gate"),
    ):
        mutated = json.loads(json.dumps(report))
        mutated[key] = value
        try:
            v4716.check_report(mutated)
        except ValueError as exc:
            assert expected in str(exc)
        else:
            raise AssertionError(f"v4_7_16 accepted {key}={value}")

    prototype_mutations = [
        ("diagnostic_target_labels_used_for_candidate_construction", True, "target labels"),
        ("diagnostic_target_labels_used_for_candidate_scoring", True, "target labels"),
        ("direct_normalized_answer_value_matching", True, "normalized value"),
        ("raw_xlsx_query_time_parsing", True, "raw XLSX"),
        ("source_file_title_shortcut_used", True, "title shortcut"),
        ("threshold_tuning_used", True, "threshold tuning"),
    ]
    for key, value, expected in prototype_mutations:
        mutated = json.loads(json.dumps(report))
        mutated["target_recall_repair_prototype"]["candidate_construction"][key] = value
        try:
            v4716.check_report(mutated)
        except ValueError as exc:
            assert expected in str(exc)
        else:
            raise AssertionError(f"v4_7_16 accepted candidate construction {key}={value}")

    for key, value, expected in (
        ("combined_target_hit_count", 513, "combined target"),
        ("baseline_hit_to_miss_count", 1, "regression"),
    ):
        mutated = json.loads(json.dumps(report))
        mutated["target_recall_repair_prototype"]["archive_1000_candidate_only_target_recall"][key] = value
        try:
            v4716.check_report(mutated)
        except ValueError as exc:
            assert expected in str(exc)
        else:
            raise AssertionError(f"v4_7_16 accepted metric drift {key}={value}")

    mutated = json.loads(json.dumps(report))
    mutated["target_recall_repair_prototype"]["archive_1000_candidate_only_target_recall"]["families"]["PDF"][
        "target_hit_regression_count"
    ] = 1
    try:
        v4716.check_report(mutated)
    except ValueError as exc:
        assert "PDF" in str(exc)
    else:
        raise AssertionError("v4_7_16 accepted PDF target-hit regression")

    mutated = json.loads(json.dumps(report))
    mutated["target_recall_repair_prototype"]["raw_prompt"] = "forbidden"
    try:
        v4716.check_report(mutated)
    except ValueError as exc:
        assert "raw prompt/response" in str(exc)
    else:
        raise AssertionError("v4_7_16 accepted raw prompt payload")


def test_v4717_candidate_only_generalization_validation_preserves_candidate_digest_and_guardrails() -> None:
    from ai.eval import rag_v4717_candidate_only_generalization_validation_and_xlsx_table_axis_repair_audit as v4717

    report = v4717.build_report(root=ROOT, generated_at="2026-05-31T00:00:00Z")
    v4717.check_report(report)
    validation = report["candidate_only_generalization_validation"]
    source_replay = validation["source_v4_7_16_candidate_replay"]
    field_scope = validation["candidate_construction_field_scope"]
    guardrails = report["anti_overfit_guardrails"]

    assert report["short_run_id"] == V4_7_17_SHORT_RUN_ID
    assert report["source_run_id"] == V4_7_16_SHORT_RUN_ID
    assert report["SearchView_vector_payload_role"] == "candidate_only"
    assert report["SourceAtom_EvidenceBundle_role"] == "evidence_truth"
    assert validation["status"] == "CANDIDATE_ONLY_GENERALIZATION_VALIDATED_DIAGNOSTIC_ONLY"
    assert source_replay["row_count"] == 1000
    assert source_replay["baseline_target_hit_count"] == 300
    assert source_replay["combined_target_hit_count"] == 514
    assert source_replay["source_candidate_set_sha256_matches_recomputed"] is True
    assert source_replay["source_topk_sha256_verified"] is True
    assert source_replay["source_topk_resolved_via_archive"] is True
    assert validation["poisoned_oracle_field_digest_stable"] is True
    assert validation["poisoned_oracle_field_evaluation_changed"] is True
    assert validation["per_query_candidates_written"] is False
    assert field_scope["allowed_field_count_by_scope"]["query"] == 2
    assert field_scope["allowed_field_count_by_scope"]["source_registry_TEXT"] == 3
    assert field_scope["allowed_field_count_by_scope"]["source_registry_XLSX"] == 6
    assert "target_source_atom_ids" in field_scope["forbidden_fields"]
    assert "expected_answer" in field_scope["forbidden_fields"]
    assert "raw_locator.normalized_value" in field_scope["forbidden_fields"]
    assert "raw_locator.cell" in field_scope["forbidden_fields"]
    assert validation["diagnostic_target_labels_used_for_after_the_fact_evaluation"] is True
    assert validation["diagnostic_target_labels_used_for_candidate_construction"] is False
    assert validation["diagnostic_target_labels_used_for_candidate_scoring"] is False
    for key, value in guardrails.items():
        if key.endswith("_allowed") or key.endswith("_used") or key.endswith("_created"):
            assert value is False, key
    assert guardrails["protected_namespaces_touched"] == []
    assert guardrails["official_metric_input_rows"] == 0


def test_v4717_xlsx_table_axis_repair_audit_keeps_low_gain_inconclusive_without_value_shortcuts() -> None:
    from ai.eval import rag_v4717_candidate_only_generalization_validation_and_xlsx_table_axis_repair_audit as v4717

    report = v4717.build_report(root=ROOT, generated_at="2026-05-31T00:00:00Z")
    v4717.check_report(report)
    audit = report["xlsx_table_axis_repair_audit"]
    family = audit["archive_1000_xlsx_family_recall"]
    overlay = audit["overlay_90_xlsx_queue_context"]
    axis_presence = audit["source_registry_xlsx_axis_presence_summary"]
    boundary = audit["future_repair_boundary"]

    assert audit["status"] == "XLSX_TABLE_AXIS_REPAIR_AUDIT_INCONCLUSIVE_DIAGNOSTIC_ONLY"
    assert audit["decision"] == "keep_inconclusive_low_gain_candidate_only"
    assert audit["safe_table_axis_fields"] == [
        "raw_locator.sheet",
        "raw_locator.row_label",
        "raw_locator.column_label",
        "raw_locator.range",
    ]
    assert family["row_count"] == 325
    assert family["baseline_target_hit_count"] == 15
    assert family["combined_target_hit_count"] == 17
    assert family["baseline_miss_to_hit_count"] == 2
    assert family["target_hit_regression_count"] == 0
    assert family["prototype_candidate_count"] == 133
    assert audit["safe_table_axis_target_hit_gain_count"] == 2
    assert audit["safe_table_axis_candidate_count"] == 133
    assert audit["safe_table_axis_gain_rate_per_baseline_miss"] == "2/310"
    assert overlay["target_not_in_topk_total"] == 28
    assert overlay["repeated_prefix_cluster_total"] == 22
    assert overlay["repeated_prefix_cluster_overlap_with_target_miss"] == 20
    assert overlay["target_hit_evidence_context_repair_count"] == 2
    assert axis_presence["source_atoms_audited"] == 343
    assert axis_presence["row_label_present_count"] == 19
    assert axis_presence["column_label_present_count"] == 19
    assert axis_presence["target_column_present_count"] == 19
    assert axis_presence["cell_level_locator_count"] == 96
    assert axis_presence["range_only_locator_count"] == 247
    assert axis_presence["normalized_value_present_but_forbidden_count"] == 19
    assert audit["direct_normalized_value_matching_used_count"] == 0
    assert audit["raw_xlsx_query_time_parsing_used_count"] == 0
    assert audit["workbook_or_source_title_shortcut_used_count"] == 0
    assert audit["formula_text_exposure_used_count"] == 0
    assert boundary["raw_xlsx_query_time_parsing"] is False
    assert boundary["direct_normalized_answer_value_matching"] is False
    assert boundary["source_file_title_shortcut_used"] is False
    assert boundary["candidate_only_searchview_materialization_required"] is True


def test_v4717_written_report_status_docs_and_current_alias() -> None:
    from ai.eval import rag_eval_registry as registry
    from ai.eval import rag_v4717_candidate_only_generalization_validation_and_xlsx_table_axis_repair_audit as v4717
    import ai.scripts.rag_eval as runner

    report = registry.load_report("v4_7_17", root=ROOT)
    current = runner.check_run("v4_7_17")
    v4717.check_report(report)
    v4717.check_report(current)
    latest = next(row for row in reversed(_read_jsonl(STATUS_JSONL)) if row.get("short_run_id") == V4_7_17_SHORT_RUN_ID)
    progress = PROGRESS_DOC.read_text(encoding="utf-8")
    measurements = MEASUREMENTS_DOC.read_text(encoding="utf-8")
    triage = TRIAGE_DOC.read_text(encoding="utf-8")

    assert runner.DEFAULT_RUN_KEY == V5_6_SHORT_KEY
    assert registry.resolve_run("current", root=ROOT).logical_key == V5_6_SHORT_KEY
    assert current["short_run_id"] == V4_7_17_SHORT_RUN_ID
    assert V4_7_17_REPORT.exists()
    assert latest["status"] == V4_7_17_STATUS
    assert latest["artifact_paths"]["report_json"] == "ai/eval/reports/rag-ingestion/runs/v4_7_17/report.json"
    assert latest["artifact_sha256"]["report_json_sha256"] == _sha256_file(V4_7_17_REPORT)
    assert latest["candidate_only_generalization_validated"] is True
    assert latest["xlsx_table_axis_repair_decision"] == "keep_inconclusive_low_gain_candidate_only"
    assert latest["xlsx_table_axis_target_hit_gain_count"] == 2
    assert latest["official_metric_input_rows"] == 0
    assert latest["promotion_evidence"] is False
    assert latest["live_db_index_cache_readiness"] is False
    assert V4_7_17_SHORT_RUN_ID in progress
    assert V4_7_17_SHORT_RUN_ID in measurements
    assert V4_7_17_SHORT_RUN_ID in triage
    assert "CANDIDATE_ONLY_GENERALIZATION_VALIDATED_DIAGNOSTIC_ONLY" in measurements
    assert "XLSX_TABLE_AXIS_REPAIR_AUDIT_INCONCLUSIVE_DIAGNOSTIC_ONLY" in measurements
    assert "keep_inconclusive_low_gain_candidate_only" in triage
    _assert_status_recorded_in_progress_and_report(
        progress,
        report,
        short_run_id=V4_7_17_SHORT_RUN_ID,
        status=V4_7_17_STATUS,
    )
    assert registry.resolve_run(V4_7_18_SHORT_KEY, root=ROOT).logical_key == V4_7_18_SHORT_KEY
    assert registry.resolve_run(V4_7_17_SHORT_KEY, root=ROOT).logical_key == V4_7_17_SHORT_KEY
    assert registry.resolve_run(V4_7_16_SHORT_KEY, root=ROOT).logical_key == V4_7_16_SHORT_KEY
    assert report["protected_namespaces_touched"] == []
    assert "promotion-ready" not in json.dumps(report, ensure_ascii=False)


def test_v4717_check_report_rejects_generalization_drift_xlsx_shortcuts_and_raw_payloads() -> None:
    from ai.eval import rag_v4717_candidate_only_generalization_validation_and_xlsx_table_axis_repair_audit as v4717

    report = v4717.build_report(root=ROOT, generated_at="2026-05-31T00:00:00Z")
    v4717.check_report(report)

    for key, value, expected in (
        ("official_metric", True, "opened forbidden gate"),
        ("promotion_evidence", True, "opened forbidden gate"),
        ("live_db_index_cache_readiness", True, "opened forbidden gate"),
    ):
        mutated = json.loads(json.dumps(report))
        mutated[key] = value
        try:
            v4717.check_report(mutated)
        except ValueError as exc:
            assert expected in str(exc)
        else:
            raise AssertionError(f"v4_7_17 accepted {key}={value}")

    validation_mutations = [
        ("poisoned_oracle_field_digest_stable", False, "candidate digest"),
        ("diagnostic_target_labels_used_for_candidate_construction", True, "target labels"),
        ("diagnostic_target_labels_used_for_candidate_scoring", True, "target labels"),
        ("per_query_candidates_written", True, "per-query candidates"),
    ]
    for key, value, expected in validation_mutations:
        mutated = json.loads(json.dumps(report))
        mutated["candidate_only_generalization_validation"][key] = value
        try:
            v4717.check_report(mutated)
        except ValueError as exc:
            assert expected in str(exc)
        else:
            raise AssertionError(f"v4_7_17 accepted validation drift {key}={value}")

    xlsx_mutations = [
        ("direct_normalized_value_matching_used_count", 1, "normalized value"),
        ("raw_xlsx_query_time_parsing_used_count", 1, "raw XLSX"),
        ("workbook_or_source_title_shortcut_used_count", 1, "source title"),
        ("formula_text_exposure_used_count", 1, "formula"),
        ("decision", "accepted_for_promotion", "XLSX"),
    ]
    for key, value, expected in xlsx_mutations:
        mutated = json.loads(json.dumps(report))
        mutated["xlsx_table_axis_repair_audit"][key] = value
        try:
            v4717.check_report(mutated)
        except ValueError as exc:
            assert expected in str(exc)
        else:
            raise AssertionError(f"v4_7_17 accepted XLSX shortcut/drift {key}={value}")

    mutated = json.loads(json.dumps(report))
    mutated["candidate_only_generalization_validation"]["raw_response"] = "forbidden"
    try:
        v4717.check_report(mutated)
    except ValueError as exc:
        assert "raw prompt/response" in str(exc)
    else:
        raise AssertionError("v4_7_17 accepted raw response payload")

    for key in ("prompt_payload", "raw_prompt_payload", "raw_response_payload"):
        mutated = json.loads(json.dumps(report))
        mutated["candidate_only_generalization_validation"][key] = "forbidden"
        try:
            v4717.check_report(mutated)
        except ValueError as exc:
            assert "raw prompt/response" in str(exc)
        else:
            raise AssertionError(f"v4_7_17 accepted payload alias {key}")

    for key in ("silver_mutation", "source_registry_mutated", "cache_mutated", "production_db_mutated", "index_rebuilt"):
        mutated = json.loads(json.dumps(report))
        mutated["anti_overfit_guardrails"][key] = True
        try:
            v4717.check_report(mutated)
        except ValueError as exc:
            assert "guardrail opened" in str(exc)
        else:
            raise AssertionError(f"v4_7_17 accepted guardrail mutation alias {key}")


def test_v4717_write_path_synthesizes_source_report_when_prior_ignored_report_is_missing(monkeypatch) -> None:
    import ai.scripts.rag_eval as runner
    from ai.eval import rag_v4716_target_recall_repair_prototype as v4716
    from ai.eval import rag_v4717_candidate_only_generalization_validation_and_xlsx_table_axis_repair_audit as v4717

    source_report = v4716.build_report(root=ROOT, generated_at="2026-05-31T00:00:00Z")
    source_report["sentinel_from_check_run"] = True
    observed: dict[str, object] = {}

    def fake_check_run(key: str) -> dict[str, object]:
        assert key == "v4_7_16"
        return source_report

    def fake_build_report(*, root: Path, source_report: dict[str, object] | None = None, **_: object) -> dict[str, object]:
        assert root == ROOT
        assert source_report is not None
        assert source_report["sentinel_from_check_run"] is True
        observed["used_source_report"] = True
        return {
            "status": V4_7_17_STATUS,
            "artifact_paths": {"report_json": "ai/eval/reports/rag-ingestion/runs/v4_7_17/report.json"},
            "official_metric_input_rows": 0,
            "counters": {},
        }

    monkeypatch.setattr(runner, "check_run", fake_check_run)
    monkeypatch.setattr(v4717, "build_report", fake_build_report)
    monkeypatch.setattr(v4717, "write_report_bundle", lambda root, report: (report, {"report_json_sha256": "0" * 64}))
    monkeypatch.setattr(v4717, "check_report", lambda report: None)
    monkeypatch.setattr(v4717, "update_docs", lambda root, report: None)
    monkeypatch.setattr(v4717, "append_status", lambda root, report, *, artifact_hashes: None)
    monkeypatch.setattr(runner, "sha256_file", lambda path: "0" * 64)

    assert runner.main(["v4_7_17", "--write"]) == 0
    assert observed["used_source_report"] is True


def test_v4718_reproduces_v4716_v4717_counter_chain_and_archive_denominator_trace() -> None:
    from ai.eval import rag_v4718_xlsx_candidate_only_materialization_repair_and_lineage_reproducibility as v4718

    report = v4718.build_report(root=ROOT, generated_at="2026-05-31T00:00:00Z")
    v4718.check_report(report)
    reproduction = report["source_counter_reproduction"]
    denominator = report["archive_denominator_trace"]
    validation = report["candidate_only_generalization_validation_reproduction"]
    gate_plan = report["v4_closeout_and_v5_gate_plan"]

    assert report["short_run_id"] == V4_7_18_SHORT_RUN_ID
    assert report["source_run_id"] == V4_7_17_SHORT_RUN_ID
    assert report["SearchView_vector_payload_role"] == "candidate_only"
    assert report["SourceAtom_EvidenceBundle_role"] == "evidence_truth"
    assert report["completion_branch"] == "v5_0_v4_closeout_and_v5_gate_plan"
    assert gate_plan["v4_closeout_basis_short_run_id"] == V4_7_18_SHORT_RUN_ID
    assert gate_plan["v4_closeout_status"] == "V4_CLOSED_DIAGNOSTIC_ONLY_SOURCE_FIRST_CANDIDATE_ONLY_LINEAGE_REPRODUCIBLE"
    assert gate_plan["ambiguous_non_gold_choices_remain_diagnostic_only"] is True
    assert gate_plan["official_metric_opening_preconditions_documented"] is True
    assert gate_plan["official_metric_opening_preconditions_satisfied"] is False
    assert gate_plan["live_readiness_promotion_preconditions_documented"] is True
    assert gate_plan["live_readiness_promotion_preconditions_satisfied"] is False
    assert gate_plan["still_closed"]["official_metric_input_rows"] == 0
    assert gate_plan["still_closed"]["denominator_mutation"] is False
    assert reproduction["v4_7_16"]["baseline_target_hit_count"] == 300
    assert reproduction["v4_7_16"]["combined_target_hit_count"] == 514
    assert reproduction["v4_7_16"]["baseline_miss_to_hit_count"] == 214
    assert reproduction["v4_7_16"]["baseline_hit_to_miss_count"] == 0
    assert reproduction["v4_7_16"]["families"]["TEXT"]["baseline_miss_to_hit_count"] == 212
    assert reproduction["v4_7_16"]["families"]["XLSX"]["baseline_miss_to_hit_count"] == 2
    assert reproduction["v4_7_16"]["families"]["PDF"]["target_hit_regression_count"] == 0
    assert reproduction["v4_7_17"]["xlsx_baseline_target_hit_count"] == 15
    assert reproduction["v4_7_17"]["xlsx_combined_target_hit_count"] == 17
    assert reproduction["v4_7_17"]["xlsx_table_axis_candidate_count"] == 133
    assert reproduction["v4_7_17"]["xlsx_table_axis_repair_decision"] == "keep_inconclusive_low_gain_candidate_only"
    assert validation["source_candidate_set_sha256"] == "b388d4fec10886142f8d3cee25db2eb771e7f4236e311b91c4ea175325a1bc5d"
    assert validation["source_candidate_set_sha256_recomputed"] == "b388d4fec10886142f8d3cee25db2eb771e7f4236e311b91c4ea175325a1bc5d"
    assert validation["source_candidate_set_sha256_source_report_claimed_match"] is True
    assert validation["source_candidate_set_sha256_matches_recomputed"] is True
    assert validation["poisoned_oracle_field_digest_stable"] is True
    assert validation["poisoned_oracle_field_evaluation_changed"] is True
    assert denominator["source_topk_sha256"] == "3a14a4908972a118606b5d2967544c278d60dd99590af3004676271a6e9ad7b3"
    assert denominator["topk_artifact_row_count"] == 1029
    assert denominator["filtered_replay_row_count"] == 1000
    assert denominator["excluded_row_count"] == 29
    assert denominator["excluded_family_counts"] == {"PDF": 4, "TEXT": 6, "XLSX": 19}
    assert denominator["filtered_family_counts"] == {"PDF": 325, "TEXT": 350, "XLSX": 325}
    assert denominator["duplicate_query_id_count"] == 0
    assert denominator["missing_query_id_count"] == 0
    assert denominator["topk_envelope_length_distribution"] == {"5": 1000}


def test_v4718_xlsx_materialization_overlay_improves_candidate_only_recall_without_shortcuts() -> None:
    from ai.eval import rag_v4718_xlsx_candidate_only_materialization_repair_and_lineage_reproducibility as v4718

    report = v4718.build_report(root=ROOT, generated_at="2026-05-31T00:00:00Z")
    v4718.check_report(report)
    repair = report["xlsx_candidate_only_materialization_repair"]
    family = repair["archive_1000_xlsx_family_recall"]
    overlay = repair["overlay_90_xlsx_projection"]
    budget = repair["candidate_budget_summary"]
    decisions = repair["rule_decisions"]

    assert repair["status"] == "XLSX_CANDIDATE_ONLY_MATERIALIZATION_REPAIR_ACCEPTED_DIAGNOSTIC_ONLY"
    assert repair["decision"] == "accept_materialized_axis_value_overlay_diagnostic_only"
    assert family["row_count"] == 325
    assert family["baseline_target_hit_count"] == 15
    assert family["v4_7_17_combined_target_hit_count"] == 17
    assert family["v4_7_18_combined_target_hit_count"] == 26
    assert family["v4_7_18_derived_overlay_target_hit_count"] == 12
    assert family["v4_7_18_gain_over_v4_7_17_count"] == 9
    assert family["v4_7_18_gain_rate_per_v4_7_17_miss"] == "9/308"
    assert family["target_hit_regression_count"] == 0
    assert budget["candidate_budget_per_query"] == 5
    assert budget["XLSX"]["attempted_row_count"] == 325
    assert budget["XLSX"]["candidate_count"] == 881
    assert budget["XLSX"]["zero_candidate_row_count"] == 78
    assert budget["XLSX"]["at_budget_row_count"] == 143
    assert budget["XLSX"]["candidate_budget_exhaustion_count"] == 109
    assert budget["XLSX"]["candidate_count_distribution"] == {"0": 78, "1": 58, "2": 31, "3": 14, "4": 1, "5": 143}
    assert overlay["xlsx_overlay_row_count"] == 30
    assert overlay["target_not_in_topk_total"] == 28
    assert overlay["repeated_prefix_cluster_total"] == 22
    assert overlay["repeated_prefix_cluster_overlap_with_target_miss"] == 20
    assert overlay["v4_7_18_gain_over_v4_7_17_count"] == 1
    assert overlay["v4_7_18_gain_repeated_prefix_count"] == 1
    assert decisions["accepted"] == ["materialized_axis_value_searchunit_overlay"]
    assert "direct_normalized_value_matching" in decisions["rejected"]
    assert "raw_xlsx_query_time_parsing" in decisions["rejected"]
    assert "source_title_or_workbook_shortcut" in decisions["rejected"]
    assert "query_id_case_id_hack" in decisions["rejected"]
    assert repair["direct_normalized_answer_value_matching"] is False
    assert repair["raw_xlsx_query_time_parsing"] is False
    assert repair["formula_evaluation"] is False
    assert repair["formula_text_exposure"] is False
    assert repair["source_file_title_shortcut_used"] is False
    assert repair["target_or_gold_locator_used_for_candidate_construction"] is False
    assert repair["diagnostic_target_labels_used_for_after_the_fact_evaluation"] is True
    assert repair["diagnostic_target_labels_used_for_candidate_construction"] is False
    assert repair["diagnostic_target_labels_used_for_candidate_scoring"] is False


def test_v4718_lineage_reproducibility_tracks_required_runner_modules_and_regression_guards() -> None:
    from ai.eval import rag_v4718_xlsx_candidate_only_materialization_repair_and_lineage_reproducibility as v4718

    report = v4718.build_report(root=ROOT, generated_at="2026-05-31T00:00:00Z")
    v4718.check_report(report)
    lineage = report["lineage_reproducibility"]
    guards = report["regression_guards"]
    modules = {entry["logical_key"]: entry for entry in lineage["required_runner_modules"]}

    assert lineage["status"] == "LINEAGE_REPRODUCIBILITY_HARDENED_DIAGNOSTIC_ONLY"
    assert lineage["clean_checkout_risk_status"] == "NO_REQUIRED_RUNNER_MODULE_RISK_DETECTED"
    assert lineage["required_runner_module_tracking_status"] == "REQUIRED_RUNNER_MODULES_TRACKED_AND_NOT_IGNORED"
    assert lineage["compile_check_mode"] == "source_text_compile_no_bytecode_write"
    assert lineage["bytecode_written_by_lineage_check"] is False
    for key in ("v4_7_13", "v4_7_14", "v4_7_15", "v4_7_16", "v4_7_17", "v4_7_18"):
        module = modules[key]
        assert module["exists"] is True, key
        assert module["tracked"] is True, key
        assert module["ignored"] is False, key
        assert module["py_compile_ok"] is True, key
        assert module["resolver_checkable"] is True, key
    assert lineage["generated_report_artifacts_ignored"] is True
    assert guards["TEXT"]["v4_7_17_combined_target_hit_count"] == 232
    assert guards["TEXT"]["v4_7_18_combined_target_hit_count"] == 232
    assert guards["TEXT"]["target_hit_regression_count"] == 0
    assert guards["PDF"]["v4_7_17_combined_target_hit_count"] == 265
    assert guards["PDF"]["v4_7_18_combined_target_hit_count"] == 265
    assert guards["PDF"]["target_hit_regression_count"] == 0
    assert guards["XLSX"]["v4_7_17_combined_target_hit_count"] == 17
    assert guards["XLSX"]["v4_7_18_combined_target_hit_count"] == 26
    assert guards["XLSX"]["target_hit_regression_count"] == 0


def test_v4718_written_report_status_docs_current_alias_and_explicit_historical_aliases() -> None:
    from ai.eval import rag_eval_registry as registry
    from ai.eval import rag_v4718_xlsx_candidate_only_materialization_repair_and_lineage_reproducibility as v4718
    import ai.scripts.rag_eval as runner

    report = registry.load_report("v4_7_18", root=ROOT)
    explicit_v4718 = runner.check_run("v4_7_18")
    current = runner.check_run("current")
    v4718.check_report(report)
    v4718.check_report(explicit_v4718)
    v560.check_report(current)
    latest = next(row for row in reversed(_read_jsonl(STATUS_JSONL)) if row.get("short_run_id") == V4_7_18_SHORT_RUN_ID)
    progress = PROGRESS_DOC.read_text(encoding="utf-8")
    measurements = MEASUREMENTS_DOC.read_text(encoding="utf-8")
    triage = TRIAGE_DOC.read_text(encoding="utf-8")

    assert runner.DEFAULT_RUN_KEY == V5_6_SHORT_KEY
    assert registry.resolve_run("current", root=ROOT).logical_key == V5_6_SHORT_KEY
    assert registry.resolve_run("v5_4", root=ROOT).logical_key == V5_4_SHORT_KEY
    assert registry.resolve_run("v5_3", root=ROOT).logical_key == V5_3_SHORT_KEY
    assert registry.resolve_run("v5_1", root=ROOT).logical_key == V5_1_SHORT_KEY
    assert registry.resolve_run("v5_0", root=ROOT).logical_key == V5_0_SHORT_KEY
    assert registry.resolve_run("v4_7_18", root=ROOT).logical_key == V4_7_18_SHORT_KEY
    assert explicit_v4718["short_run_id"] == V4_7_18_SHORT_RUN_ID
    assert V4_7_18_REPORT.exists()
    assert latest["status"] == V4_7_18_STATUS
    assert latest["artifact_paths"]["report_json"] == "ai/eval/reports/rag-ingestion/runs/v4_7_18/report.json"
    assert latest["artifact_sha256"]["report_json_sha256"] == _sha256_file(V4_7_18_REPORT)
    assert latest["lineage_reproducibility_status"] == "LINEAGE_REPRODUCIBILITY_HARDENED_DIAGNOSTIC_ONLY"
    assert latest["xlsx_materialization_repair_decision"] == "accept_materialized_axis_value_overlay_diagnostic_only"
    assert latest["xlsx_v4_7_18_combined_target_hit_count"] == 26
    assert latest["official_metric_input_rows"] == 0
    assert latest["promotion_evidence"] is False
    assert latest["live_db_index_cache_readiness"] is False
    assert latest["v5_gate_plan_id"] == "v5_0_v4_closeout_and_v5_gate_plan"
    assert latest["v4_closeout_basis_short_run_id"] == V4_7_18_SHORT_RUN_ID
    assert latest["ambiguous_non_gold_choices_remain_diagnostic_only"] is True
    assert latest["official_metric_opening_preconditions_documented"] is True
    assert latest["official_metric_opening_preconditions_satisfied"] is False
    assert latest["live_readiness_promotion_preconditions_documented"] is True
    assert latest["live_readiness_promotion_preconditions_satisfied"] is False
    assert V4_7_18_SHORT_RUN_ID in progress
    assert V4_7_18_SHORT_RUN_ID in measurements
    assert V4_7_18_SHORT_RUN_ID in triage
    assert "v5_0_v4_closeout_and_v5_gate_plan" in progress
    assert "v5_0 v4 closeout and v5 gate plan" in measurements
    assert "v5_0 v4 closeout and v5 gate plan" in triage
    assert f"<!-- {V5_5_SHORT_RUN_ID}:measurements-entry:start -->" in measurements
    assert f"<!-- {V5_5_SHORT_RUN_ID}:triage-entry:start -->" in triage
    assert "Overall status: `V4_7_12_LAYERED_RETRIEVAL_GENERALIZATION_AND_OVERFIT_AUDIT_NONPROD_READY`;" not in progress
    assert "frozen v4 closeout basis" in progress
    assert "user-owned official-eval approval packet" in progress
    assert "User-owned decisions" in triage
    assert "Codex-owned work" in triage
    assert "XLSX residual backlog" in triage
    assert "Official metric opening preconditions" in triage
    assert "Live-readiness and promotion preconditions" in triage
    assert "| official_metric_opening_preconditions_satisfied | false |" in measurements
    assert "| live_readiness_promotion_preconditions_satisfied | false |" in measurements
    assert "v5_0_v4_closeout_and_v5_gate_plan" in progress
    assert "v4_7_18 remains the frozen v4 basis consumed by v5_0" in progress
    assert "LINEAGE_REPRODUCIBILITY_HARDENED_DIAGNOSTIC_ONLY" in measurements
    assert "XLSX_CANDIDATE_ONLY_MATERIALIZATION_REPAIR_ACCEPTED_DIAGNOSTIC_ONLY" in measurements
    assert "accept_materialized_axis_value_overlay_diagnostic_only" in triage
    _assert_status_recorded_in_progress_and_report(
        progress,
        current,
        short_run_id=V5_6_SHORT_RUN_ID,
        status=V5_6_STATUS,
    )
    assert registry.resolve_run(V4_7_18_SHORT_KEY, root=ROOT).logical_key == V4_7_18_SHORT_KEY
    assert registry.resolve_run(V5_0_SHORT_KEY, root=ROOT).logical_key == V5_0_SHORT_KEY
    assert report["protected_namespaces_touched"] == []
    assert "promotion-ready" not in json.dumps(report, ensure_ascii=False)


def test_v500_current_profile_checks_frozen_v4718_basis_guardrails_without_recomputing() -> None:
    from ai.eval import rag_v4718_xlsx_candidate_only_materialization_repair_and_lineage_reproducibility as v4718

    report = _read_json(V4_7_18_REPORT)
    v4718.check_report(report)

    assert report["short_run_id"] == V4_7_18_SHORT_RUN_ID
    assert report["counters"]["current_resolves_to"] == V4_7_18_SHORT_KEY
    assert report["official_metric_input_rows"] == 0
    assert report["silver_promoted_to_gold_count"] == 0
    assert report["protected_namespaces_touched"] == []
    assert report["lineage_reproducibility"]["status"] == "LINEAGE_REPRODUCIBILITY_HARDENED_DIAGNOSTIC_ONLY"
    assert report["xlsx_candidate_only_materialization_repair"]["archive_1000_xlsx_family_recall"][
        "v4_7_18_combined_target_hit_count"
    ] == 26

    drift_cases = [
        (("official_metric",), True, "opened forbidden gate"),
        (("promotion_evidence",), True, "opened forbidden gate"),
        (("live_db_index_cache_readiness",), True, "opened forbidden gate"),
        (("training_dataset_created",), True, "opened forbidden gate"),
        (("fine_tuning",), True, "opened forbidden gate"),
        (("protected_namespaces_touched",), ["ai/eval/eval_queries"], "touched protected namespaces"),
        (("source_counter_reproduction", "v4_7_16", "baseline_target_hit_count"), 299, "source v4_7_16"),
        (
            (
                "xlsx_candidate_only_materialization_repair",
                "archive_1000_xlsx_family_recall",
                "v4_7_18_combined_target_hit_count",
            ),
            25,
            "XLSX family",
        ),
        (("lineage_reproducibility", "required_runner_modules", 0, "tracked"), False, "required runner module"),
        (
            ("v4_closeout_and_v5_gate_plan", "official_metric_opening_preconditions_satisfied"),
            True,
            "official metric opening",
        ),
        (
            ("v4_closeout_and_v5_gate_plan", "live_readiness_promotion_preconditions_satisfied"),
            True,
            "live-readiness/promotion",
        ),
        (
            ("candidate_only_generalization_validation_reproduction", "source_candidate_set_sha256_recomputed"),
            "0" * 64,
            "recomputed source candidate digest",
        ),
        (
            ("candidate_only_generalization_validation_reproduction", "source_candidate_set_sha256_matches_recomputed"),
            False,
            "did not match recomputation",
        ),
        (
            ("lineage_reproducibility", "compile_check_mode"),
            "py_compile_default_bytecode_write",
            "avoid bytecode",
        ),
        (("lineage_reproducibility", "bytecode_written_by_lineage_check"), True, "wrote bytecode"),
        (("xlsx_candidate_only_materialization_repair", "prompt_payload"), "forbidden", "raw prompt/response"),
    ]
    for key, expected in (
        ("direct_normalized_answer_value_matching", "normalized"),
        ("raw_xlsx_query_time_parsing", "raw XLSX"),
        ("formula_evaluation", "formula"),
        ("formula_text_exposure", "formula"),
        ("source_file_title_shortcut_used", "source title"),
        ("target_or_gold_locator_used_for_candidate_construction", "target/gold"),
        ("diagnostic_target_labels_used_for_candidate_construction", "target labels"),
        ("diagnostic_target_labels_used_for_candidate_scoring", "target labels"),
    ):
        drift_cases.append((("xlsx_candidate_only_materialization_repair", key), True, expected))
    for family in ("TEXT", "PDF", "XLSX"):
        drift_cases.append((("regression_guards", family, "target_hit_regression_count"), 1, "regression"))

    for path, value, expected in drift_cases:
        mutated = json.loads(json.dumps(report))
        cursor = mutated
        for part in path[:-1]:
            cursor = cursor[part]
        cursor[path[-1]] = value
        try:
            v4718.check_report(mutated)
        except ValueError as exc:
            assert expected in str(exc)
        else:
            raise AssertionError(f"v4_7_18 accepted frozen-basis drift at {path}")


def test_v4718_check_report_rejects_shortcuts_opened_gates_raw_payloads_and_regression_drift() -> None:
    from ai.eval import rag_v4718_xlsx_candidate_only_materialization_repair_and_lineage_reproducibility as v4718

    report = v4718.build_report(root=ROOT, generated_at="2026-05-31T00:00:00Z")
    v4718.check_report(report)

    for key, value, expected in (
        ("official_metric", True, "opened forbidden gate"),
        ("promotion_evidence", True, "opened forbidden gate"),
        ("live_db_index_cache_readiness", True, "opened forbidden gate"),
        ("training_dataset_created", True, "opened forbidden gate"),
        ("fine_tuning", True, "opened forbidden gate"),
    ):
        mutated = json.loads(json.dumps(report))
        mutated[key] = value
        try:
            v4718.check_report(mutated)
        except ValueError as exc:
            assert expected in str(exc)
        else:
            raise AssertionError(f"v4_7_18 accepted {key}={value}")

    repair_mutations = [
        ("direct_normalized_answer_value_matching", True, "normalized"),
        ("raw_xlsx_query_time_parsing", True, "raw XLSX"),
        ("formula_evaluation", True, "formula"),
        ("formula_text_exposure", True, "formula"),
        ("source_file_title_shortcut_used", True, "source title"),
        ("target_or_gold_locator_used_for_candidate_construction", True, "target/gold"),
        ("diagnostic_target_labels_used_for_candidate_scoring", True, "target labels"),
        ("diagnostic_target_labels_used_for_candidate_construction", True, "target labels"),
    ]
    for key, value, expected in repair_mutations:
        mutated = json.loads(json.dumps(report))
        mutated["xlsx_candidate_only_materialization_repair"][key] = value
        try:
            v4718.check_report(mutated)
        except ValueError as exc:
            assert expected in str(exc)
        else:
            raise AssertionError(f"v4_7_18 accepted repair shortcut {key}={value}")

    for family in ("TEXT", "PDF", "XLSX"):
        mutated = json.loads(json.dumps(report))
        mutated["regression_guards"][family]["target_hit_regression_count"] = 1
        try:
            v4718.check_report(mutated)
        except ValueError as exc:
            assert "regression" in str(exc)
        else:
            raise AssertionError(f"v4_7_18 accepted {family} regression")

    for key in ("prompt_payload", "raw_prompt_payload", "raw_response_payload", "raw_response"):
        mutated = json.loads(json.dumps(report))
        mutated["xlsx_candidate_only_materialization_repair"][key] = "forbidden"
        try:
            v4718.check_report(mutated)
        except ValueError as exc:
            assert "raw prompt/response" in str(exc)
        else:
            raise AssertionError(f"v4_7_18 accepted raw payload {key}")

    gate_plan_mutations = [
        ("ambiguous_non_gold_choices_remain_diagnostic_only", False, "ambiguous non-gold"),
        ("official_metric_opening_preconditions_satisfied", True, "official metric opening preconditions"),
        ("live_readiness_promotion_preconditions_satisfied", True, "live-readiness/promotion preconditions"),
        ("user_owned_decisions", [], "missing section"),
        ("official_metric_opening_preconditions", [], "missing section"),
    ]
    for key, value, expected in gate_plan_mutations:
        mutated = json.loads(json.dumps(report))
        mutated["v4_closeout_and_v5_gate_plan"][key] = value
        try:
            v4718.check_report(mutated)
        except ValueError as exc:
            assert expected in str(exc)
        else:
            raise AssertionError(f"v4_7_18 accepted v5 gate plan drift {key}={value}")

    validation_mutations = [
        ("source_candidate_set_sha256_recomputed", "0" * 64, "recomputed source candidate digest"),
        ("source_candidate_set_sha256_source_report_claimed_match", False, "source report did not claim"),
        ("source_candidate_set_sha256_matches_recomputed", False, "did not match recomputation"),
    ]
    for key, value, expected in validation_mutations:
        mutated = json.loads(json.dumps(report))
        mutated["candidate_only_generalization_validation_reproduction"][key] = value
        try:
            v4718.check_report(mutated)
        except ValueError as exc:
            assert expected in str(exc)
        else:
            raise AssertionError(f"v4_7_18 accepted validation drift {key}={value}")

    lineage_mutations = [
        ("compile_check_mode", "py_compile_default_bytecode_write", "avoid bytecode"),
        ("bytecode_written_by_lineage_check", True, "wrote bytecode"),
    ]
    for key, value, expected in lineage_mutations:
        mutated = json.loads(json.dumps(report))
        mutated["lineage_reproducibility"][key] = value
        try:
            v4718.check_report(mutated)
        except ValueError as exc:
            assert expected in str(exc)
        else:
            raise AssertionError(f"v4_7_18 accepted lineage side-effect drift {key}={value}")


def test_v4718_rejects_recomputed_v4716_candidate_digest_drift(monkeypatch) -> None:
    from ai.eval import rag_v4718_xlsx_candidate_only_materialization_repair_and_lineage_reproducibility as v4718

    original = v4718._base_v4716_candidate_sets

    def fake_base_v4716_candidate_sets(**kwargs: object) -> tuple[list[list[str]], str, dict[str, object]]:
        candidate_sets, digest, index = original(**kwargs)
        assert digest == "b388d4fec10886142f8d3cee25db2eb771e7f4236e311b91c4ea175325a1bc5d"
        return candidate_sets, "0" * 64, index

    monkeypatch.setattr(v4718, "_base_v4716_candidate_sets", fake_base_v4716_candidate_sets)
    report = v4718.build_report(root=ROOT, generated_at="2026-05-31T00:00:00Z", check=False)
    validation = report["candidate_only_generalization_validation_reproduction"]

    assert validation["source_candidate_set_sha256"] == "b388d4fec10886142f8d3cee25db2eb771e7f4236e311b91c4ea175325a1bc5d"
    assert validation["source_candidate_set_sha256_recomputed"] == "0" * 64
    assert validation["source_candidate_set_sha256_matches_recomputed"] is False
    try:
        v4718.check_report(report)
    except ValueError as exc:
        assert "recomputed source candidate digest" in str(exc)
    else:
        raise AssertionError("v4_7_18 accepted recomputed source candidate digest drift")


def test_v4718_write_path_synthesizes_v4717_source_report_when_prior_ignored_report_is_missing(monkeypatch) -> None:
    import ai.scripts.rag_eval as runner
    from ai.eval import rag_v4717_candidate_only_generalization_validation_and_xlsx_table_axis_repair_audit as v4717
    from ai.eval import rag_v4718_xlsx_candidate_only_materialization_repair_and_lineage_reproducibility as v4718

    source_report = v4717.build_report(root=ROOT, generated_at="2026-05-31T00:00:00Z")
    source_report["sentinel_from_check_run"] = True
    observed: dict[str, object] = {}

    def fake_check_run(key: str) -> dict[str, object]:
        assert key == "v4_7_17"
        return source_report

    def fake_build_report(*, root: Path, source_report: dict[str, object] | None = None, **_: object) -> dict[str, object]:
        assert root == ROOT
        assert source_report is not None
        assert source_report["sentinel_from_check_run"] is True
        observed["used_source_report"] = True
        return {
            "status": V4_7_18_STATUS,
            "artifact_paths": {"report_json": "ai/eval/reports/rag-ingestion/runs/v4_7_18/report.json"},
            "official_metric_input_rows": 0,
            "counters": {},
        }

    monkeypatch.setattr(runner, "check_run", fake_check_run)
    monkeypatch.setattr(v4718, "build_report", fake_build_report)
    monkeypatch.setattr(v4718, "write_report_bundle", lambda root, report: (report, {"report_json_sha256": "0" * 64}))
    monkeypatch.setattr(v4718, "check_report", lambda report: None)
    monkeypatch.setattr(v4718, "update_docs", lambda root, report: None)
    monkeypatch.setattr(v4718, "append_status", lambda root, report, *, artifact_hashes: None)
    monkeypatch.setattr(runner, "sha256_file", lambda path: "0" * 64)

    assert runner.main(["v4_7_18", "--write"]) == 0
    assert observed["used_source_report"] is True


def test_v500_closeout_report_freezes_v4718_basis_and_keeps_all_gates_closed() -> None:
    from ai.eval import rag_eval_registry as registry
    from ai.eval import rag_v500_v4_closeout_and_v5_gate_plan as v500

    source_report = registry.load_report("v4_7_18", root=ROOT)
    report = v500.build_report(root=ROOT, source_report=source_report, generated_at="2026-05-31T00:00:00Z")
    v500.check_report(report)

    assert report["schema_version"] == f"{V5_0_SHORT_RUN_ID}_report_v1"
    assert report["logical_run_key"] == V5_0_SHORT_KEY
    assert report["run_id"] == V5_0_SHORT_RUN_ID
    assert report["short_run_id"] == V5_0_SHORT_RUN_ID
    assert report["canonical_long_run_id"] == V5_0_LONG_RUN_ID
    assert report["status"] == V5_0_STATUS
    assert report["source_run_id"] == V4_7_18_SHORT_RUN_ID
    assert report["source_report_status"] == V4_7_18_STATUS
    assert report["source_report_sha256"] == _sha256_file(V4_7_18_REPORT)
    assert report["v4_closeout_basis"] == V4_7_18_SHORT_KEY
    assert report["v4_closeout_basis_short_run_id"] == V4_7_18_SHORT_RUN_ID
    assert report["current_resolves_to"] == V5_0_SHORT_KEY
    assert report["SearchView_vector_payload_role"] == "candidate_only"
    assert report["SourceAtom_EvidenceBundle_role"] == "evidence_truth"
    assert report["protected_namespaces_touched"] == []
    assert report["official_metric_input_rows"] == 0
    assert report["silver_official_metric_input_rows"] == 0
    assert report["silver_promoted_to_gold_count"] == 0

    for key in (
        "official_metric",
        "gold_mutation",
        "qrels_mutation",
        "label_mutation",
        "expected_answer_mutation",
        "supporting_evidence_mutation",
        "denominator_mutation",
        "training_dataset_created",
        "fine_tuning",
        "ft_a_execution",
        "promotion_evidence",
        "product_success_evidence_allowed",
        "live_db_index_cache_readiness",
        "production_db_mutated",
        "source_registry_mutated",
        "index_rebuilt",
        "cache_mutated",
    ):
        assert report[key] is False, key

    summary = report["v4_7_18_summary_counters"]
    assert summary["TEXT"] == {"hit": 232, "total": 350, "miss": 118}
    assert summary["PDF"] == {"hit": 265, "total": 325, "miss": 60}
    assert summary["XLSX"] == {"hit": 26, "total": 325, "miss": 299}
    assert summary["xlsx_zero_candidate_row_count"] == 78
    assert summary["xlsx_candidate_budget_exhaustion_count"] == 109
    assert summary["family_target_hit_regression_count"] == {"TEXT": 0, "PDF": 0, "XLSX": 0}

    gate_plan = report["v5_gate_plan"]
    assert set(gate_plan) == {"A", "B", "C", "D", "E"}
    assert gate_plan["A"]["owner"] == "user"
    assert gate_plan["B"]["owner"] == "codex"
    assert gate_plan["D"]["status"] == "closed_pending_user_owned_policy"
    assert gate_plan["E"]["status"] == "closed_pending_official_metric_and_promotion_policy"
    assert "XLSX 299 misses" in " ".join(report["residual_risks"])
    assert "78 zero-candidate" in " ".join(report["residual_risks"])
    assert "109 budget-exhausted" in " ".join(report["next_recommendations"])


def test_v500_written_report_status_docs_current_alias_and_ignored_artifacts() -> None:
    from ai.eval import rag_eval_registry as registry
    from ai.eval import rag_v4718_xlsx_candidate_only_materialization_repair_and_lineage_reproducibility as v4718
    from ai.eval import rag_v500_v4_closeout_and_v5_gate_plan as v500
    from ai.eval import rag_v520_xlsx_residual_candidate_only_retrieval_engineering as v520
    from ai.eval import rag_v530_pdf_text_residual_retrieval_evidence_hardening as v530
    from ai.eval import rag_v540_user_owned_official_eval_approval_packet as v540
    from ai.eval import rag_v550_user_approved_gold_packet_ingestion_and_official_metric_dry_run as v550
    import ai.scripts.rag_eval as runner

    report = registry.load_report("v5_0", root=ROOT)
    current = runner.check_run("current")
    explicit_v530 = runner.check_run("v5_3")
    explicit_v500 = runner.check_run("v5_0")
    explicit_v4718 = runner.check_run("v4_7_18")
    v500.check_report(report)
    v500.check_report(explicit_v500)
    v560.check_report(current)
    v530.check_report(explicit_v530)
    v4718.check_report(explicit_v4718)
    latest = next(row for row in reversed(_read_jsonl(STATUS_JSONL)) if row.get("short_run_id") == V5_0_SHORT_RUN_ID)
    progress = PROGRESS_DOC.read_text(encoding="utf-8")
    current_progress = progress.split("## Short History", 1)[0]
    measurements = MEASUREMENTS_DOC.read_text(encoding="utf-8")
    triage = TRIAGE_DOC.read_text(encoding="utf-8")

    assert runner.DEFAULT_RUN_KEY == V5_6_SHORT_KEY
    assert registry.resolve_run("current", root=ROOT).logical_key == V5_6_SHORT_KEY
    assert registry.resolve_run("v5_4", root=ROOT).logical_key == V5_4_SHORT_KEY
    assert registry.resolve_run("v5_3", root=ROOT).logical_key == V5_3_SHORT_KEY
    assert registry.resolve_run("v5_1", root=ROOT).logical_key == V5_1_SHORT_KEY
    assert registry.resolve_run("v5_0", root=ROOT).logical_key == V5_0_SHORT_KEY
    assert registry.resolve_run("v4_7_18", root=ROOT).logical_key == V4_7_18_SHORT_KEY
    assert current["short_run_id"] == V5_6_SHORT_RUN_ID
    assert explicit_v530["short_run_id"] == V5_3_SHORT_RUN_ID
    assert explicit_v500["short_run_id"] == V5_0_SHORT_RUN_ID
    assert explicit_v4718["short_run_id"] == V4_7_18_SHORT_RUN_ID
    assert V5_0_REPORT.exists()
    assert latest["status"] == V5_0_STATUS
    assert latest["artifact_paths"]["report_json"] == "ai/eval/reports/rag-ingestion/runs/v5_0/report.json"
    assert latest["artifact_sha256"]["report_json_sha256"] == _sha256_file(V5_0_REPORT)
    assert latest["source_run_id"] == V4_7_18_SHORT_RUN_ID
    assert latest["v4_closeout_basis"] == V4_7_18_SHORT_KEY
    assert latest["current_resolves_to"] == V5_0_SHORT_KEY
    assert latest["official_metric_input_rows"] == 0
    assert latest["promotion_evidence"] is False
    assert latest["live_db_index_cache_readiness"] is False

    assert V5_0_SHORT_RUN_ID in progress
    assert V5_0_SHORT_RUN_ID in measurements
    assert V5_0_SHORT_RUN_ID in triage
    assert f"<!-- {V5_6_3_SHORT_RUN_ID}:progress-entry:start -->" in progress
    assert f"<!-- {V5_5_SHORT_RUN_ID}:measurements-entry:start -->" in measurements
    assert f"<!-- {V5_5_SHORT_RUN_ID}:triage-entry:start -->" in triage
    assert f"Overall status: `{V5_0_STATUS}`;" in progress
    assert "v4 closeout basis: `v4_7_18`" in progress
    assert "Overall status: `V4_7_12_LAYERED_RETRIEVAL_GENERALIZATION_AND_OVERFIT_AUDIT_NONPROD_READY`;" not in current_progress
    for stale_current in (
        "`current` resolves to `v4_7_18`",
        "`current` resolves to `v4_7_17`",
        "`current` resolves to `v4_7_16`",
        "`current` resolves to `v4_7_15`",
        "`current` resolves to `v4_7_14`",
        "`current` resolves to `v4_7_13`",
        "`current` resolves to `v4_7_12`",
    ):
        assert stale_current not in current_progress
    assert "XLSX 299 misses" in progress
    assert "| official_metric_input_rows | 0 |" in measurements
    assert "| xlsx_miss_count | 299 |" in measurements
    assert "current_source_of_truth | v4_7_18" not in measurements
    assert "| v4_closeout_source_of_truth | v4_7_18 |" in measurements
    assert "User-owned decisions" in triage
    assert "Codex-owned work" in triage
    assert "Official metric opening preconditions" in triage
    assert "Live-readiness and promotion preconditions" in triage
    _assert_status_recorded_in_progress_and_report(
        progress,
        current,
        short_run_id=V5_6_SHORT_RUN_ID,
        status=V5_6_STATUS,
    )
    assert registry.resolve_run(V5_3_SHORT_KEY, root=ROOT).logical_key == V5_3_SHORT_KEY
    assert registry.resolve_run(V5_2_SHORT_KEY, root=ROOT).logical_key == V5_2_SHORT_KEY
    assert registry.resolve_run(V5_0_SHORT_KEY, root=ROOT).logical_key == V5_0_SHORT_KEY
    assert registry.resolve_run(V4_7_18_SHORT_KEY, root=ROOT).logical_key == V4_7_18_SHORT_KEY

    for rel_path in (
        "ai/eval/reports/rag-ingestion/runs/v5_0/report.json",
        "ai/eval/reports/rag-ingestion/status.jsonl",
    ):
        assert subprocess.run(["git", "check-ignore", "-q", rel_path], cwd=ROOT).returncode == 0


def test_v500_check_report_rejects_opened_gates_source_drift_raw_payloads_and_counter_drift() -> None:
    from ai.eval import rag_v500_v4_closeout_and_v5_gate_plan as v500

    report = v500.build_report(root=ROOT, generated_at="2026-05-31T00:00:00Z")
    v500.check_report(report)

    for key in (
        "official_metric",
        "gold_mutation",
        "qrels_mutation",
        "label_mutation",
        "expected_answer_mutation",
        "supporting_evidence_mutation",
        "denominator_mutation",
        "training_dataset_created",
        "fine_tuning",
        "ft_a_execution",
        "promotion_evidence",
        "product_success_evidence_allowed",
        "live_db_index_cache_readiness",
        "production_db_mutated",
        "source_registry_mutated",
        "silver_mutation",
        "index_rebuilt",
        "cache_mutated",
    ):
        mutated = json.loads(json.dumps(report))
        mutated[key] = True
        try:
            v500.check_report(mutated)
        except ValueError as exc:
            assert "opened forbidden gate" in str(exc)
            assert key in str(exc)
        else:
            raise AssertionError(f"v5_0 accepted {key}=True")

    for path, value, expected in (
        (("answer_generation_attempted",), True, "answer generation"),
        (("generated_response_count",), 1, "generated response"),
        (("counters", "official_metric_input_rows"), 1, "official_metric_input_rows"),
        (("counters", "silver_official_metric_input_rows"), 1, "silver_official_metric_input_rows"),
        (("counters", "silver_promoted_to_gold_count"), 1, "silver_promoted_to_gold_count"),
        (("counters", "generated_response_count"), 1, "generated response"),
    ):
        mutated = json.loads(json.dumps(report))
        cursor = mutated
        for part in path[:-1]:
            cursor = cursor[part]
        cursor[path[-1]] = value
        try:
            v500.check_report(mutated)
        except ValueError as exc:
            assert expected in str(exc)
        else:
            raise AssertionError(f"v5_0 accepted generation drift at {path}")

    for key in ("prompt_payload", "raw_prompt_payload", "raw_response_payload", "raw_response"):
        mutated = json.loads(json.dumps(report))
        mutated["v5_gate_plan"][key] = "forbidden"
        try:
            v500.check_report(mutated)
        except ValueError as exc:
            assert "raw prompt/response" in str(exc)
        else:
            raise AssertionError(f"v5_0 accepted raw payload {key}")

    for path, value, expected in (
        (("source_run_id",), "v4_7_17_candidate_only_generalization_validation_and_xlsx_table_axis_repair_audit", "source run"),
        (("v4_7_18_summary_counters", "XLSX", "hit"), 25, "XLSX"),
        (("v4_7_18_summary_counters", "xlsx_zero_candidate_row_count"), 77, "zero-candidate"),
        (("v4_7_18_summary_counters", "family_target_hit_regression_count", "PDF"), 1, "regression"),
        (("v5_gate_plan", "D", "status"), "open", "gate plan"),
    ):
        mutated = json.loads(json.dumps(report))
        cursor = mutated
        for part in path[:-1]:
            cursor = cursor[part]
        cursor[path[-1]] = value
        try:
            v500.check_report(mutated)
        except ValueError as exc:
            assert expected in str(exc)
        else:
            raise AssertionError(f"v5_0 accepted drift at {path}")


def test_v500_write_path_synthesizes_v4718_source_report_when_prior_ignored_report_is_missing(monkeypatch) -> None:
    import ai.scripts.rag_eval as runner
    from ai.eval import rag_v500_v4_closeout_and_v5_gate_plan as v500

    source_report = {
        "short_run_id": V4_7_18_SHORT_RUN_ID,
        "canonical_long_run_id": V4_7_18_LONG_RUN_ID,
        "status": V4_7_18_STATUS,
        "sentinel_from_check_run": True,
    }
    observed: dict[str, object] = {}

    def fake_check_run(key: str) -> dict[str, object]:
        assert key == "v4_7_18"
        return source_report

    def fake_build_report(*, root: Path, source_report: dict[str, object] | None = None, **_: object) -> dict[str, object]:
        assert root == ROOT
        assert source_report is not None
        assert source_report["sentinel_from_check_run"] is True
        observed["used_source_report"] = True
        return {
            "status": V5_0_STATUS,
            "artifact_paths": {"report_json": "ai/eval/reports/rag-ingestion/runs/v5_0/report.json"},
            "official_metric_input_rows": 0,
            "counters": {},
        }

    monkeypatch.setattr(runner, "check_run", fake_check_run)
    monkeypatch.setattr(v500, "build_report", fake_build_report)
    monkeypatch.setattr(v500, "write_report_bundle", lambda root, report: (report, {"report_json_sha256": "0" * 64}))
    monkeypatch.setattr(v500, "check_report", lambda report: None)
    monkeypatch.setattr(v500, "update_docs", lambda root, report: None)
    monkeypatch.setattr(v500, "append_status", lambda root, report, *, artifact_hashes: None)

    assert runner.main(["v5_0", "--write"]) == 0
    assert observed["used_source_report"] is True


def test_v510_official_eval_gate_scaffold_represents_user_owned_inputs_and_zero_rows() -> None:
    from ai.eval import rag_eval_registry as registry
    from ai.eval import rag_v510_official_eval_gate_scaffolding as v510

    source_report = registry.load_report("v5_0", root=ROOT)
    report = v510.build_report(root=ROOT, source_report=source_report, generated_at="2026-06-01T00:00:00Z")
    v510.check_report(report)

    assert report["schema_version"] == f"{V5_1_SHORT_RUN_ID}_report_v1"
    assert report["logical_run_key"] == V5_1_SHORT_KEY
    assert report["run_id"] == V5_1_SHORT_RUN_ID
    assert report["short_run_id"] == V5_1_SHORT_RUN_ID
    assert report["canonical_long_run_id"] == V5_1_LONG_RUN_ID
    assert report["status"] == V5_1_STATUS
    assert report["source_run_id"] == V5_0_SHORT_RUN_ID
    assert report["source_report_status"] == V5_0_STATUS
    assert report["v4_closeout_basis"] == V4_7_18_SHORT_KEY
    assert report["current_resolves_to"] == V5_1_SHORT_KEY
    assert report["official_eval_scaffold_created"] is True
    assert report["official_eval_user_gate_ready"] is False
    assert report["official_eval_approval_artifact_found"] is False
    assert report["official_metric"] is False
    assert report["official_metric_input_rows"] == 0
    assert report["official_metric_input_rows_created"] == 0
    assert report["official_metric_input_rows_scope"] == "v5_1_scaffold_created_rows_only"
    assert report["existing_registry_backed_official_metric_input_rows_snapshot"] == 29
    assert report["existing_registry_backed_official_metric_input_rows_by_track_snapshot"] == {
        "text_namu_v2_1": 6,
        "xlsx_business_structured": 19,
        "pdf_business_ocr_mm": 4,
    }
    assert report["official_metric_dry_run_opened"] is False
    assert report["official_metric_denominator_usage_allowed"] is False
    assert report["training_dataset_created"] is False
    assert report["training_manifest_jsonl_created"] is False
    assert report["training_job_created"] is False
    assert report["fine_tuning_dataset_export_created"] is False
    assert report["fine_tuning"] is False
    assert report["fine_tuning_started"] is False
    assert report["fine_tuning_executed"] is False
    assert report["protected_namespaces_touched"] == []

    gate = report["official_eval_gate_scaffold"]
    required = gate["required_user_owned_approval_artifacts"]
    assert set(required) == {
        "gold_set_creation_review",
        "expected_answer",
        "expected_supporting_evidence",
        "relevance_label",
        "answerability_label",
        "official_denominator_policy",
        "gold_qrels_policy",
        "promotion_policy",
    }
    for key, requirement in required.items():
        assert requirement["owner"] == "user", key
        assert requirement["status"] == "missing_required_external_input", key
        assert requirement["codex_may_infer"] is False, key

    assert gate["default_behavior"] == "fail_closed_zero_official_rows"
    assert gate["official_metric_input_rows_by_default"] == 0
    assert gate["future_official_metric_builder_enabled"] is False
    assert gate["blocked_by_user_owned_gold_qrels_or_denominator_gate"] is True
    assert "missing_user_owned_gold_qrels_or_denominator_approval" in gate["blocked_reasons"]
    assert "approval_artifact_schema_validator" in gate["validators"]
    assert "official_metric_input_row_builder_disabled_by_default" in gate["validators"]
    assert report["counters"]["missing_user_owned_approval_artifact_count"] == len(required)

    readiness = report["ft_readiness_compatibility"]
    assert readiness["status"] == "schema_compatible_no_dataset_export"
    assert readiness["training_dataset_created"] is False
    assert readiness["fine_tuning_dataset_export_created"] is False
    assert readiness["blocked_by_eval"] is True
    assert readiness["blocked_by_user_gate"] is True
    assert "blocked_missing_gold_or_expected_evidence" in readiness["training_example_source_classifications"].values()


def test_v510_written_report_status_docs_current_alias_and_ignored_artifacts() -> None:
    from ai.eval import rag_eval_registry as registry
    from ai.eval import rag_v500_v4_closeout_and_v5_gate_plan as v500
    from ai.eval import rag_v510_official_eval_gate_scaffolding as v510
    from ai.eval import rag_v520_xlsx_residual_candidate_only_retrieval_engineering as v520
    from ai.eval import rag_v530_pdf_text_residual_retrieval_evidence_hardening as v530
    from ai.eval import rag_v540_user_owned_official_eval_approval_packet as v540
    from ai.eval import rag_v550_user_approved_gold_packet_ingestion_and_official_metric_dry_run as v550
    import ai.scripts.rag_eval as runner

    report = registry.load_report("v5_1", root=ROOT)
    current = runner.check_run("current")
    explicit_v530 = runner.check_run("v5_3")
    explicit_v510 = runner.check_run("v5_1")
    explicit_v500 = runner.check_run("v5_0")
    v510.check_report(report)
    v510.check_report(explicit_v510)
    v560.check_report(current)
    v530.check_report(explicit_v530)
    v500.check_report(explicit_v500)
    latest = next(row for row in reversed(_read_jsonl(STATUS_JSONL)) if row.get("short_run_id") == V5_1_SHORT_RUN_ID)
    progress = PROGRESS_DOC.read_text(encoding="utf-8")
    current_progress = progress.split("## Short History", 1)[0]
    current_status_block = progress.split("## Current Status", 1)[1].split("## Short History", 1)[0]
    measurements = MEASUREMENTS_DOC.read_text(encoding="utf-8")
    triage = TRIAGE_DOC.read_text(encoding="utf-8")

    assert runner.DEFAULT_RUN_KEY == V5_6_SHORT_KEY
    assert registry.resolve_run("current", root=ROOT).logical_key == V5_6_SHORT_KEY
    assert registry.resolve_run("v5_4", root=ROOT).logical_key == V5_4_SHORT_KEY
    assert registry.resolve_run("v5_3", root=ROOT).logical_key == V5_3_SHORT_KEY
    assert registry.resolve_run("v5_1", root=ROOT).logical_key == V5_1_SHORT_KEY
    assert registry.resolve_run("v5_0", root=ROOT).logical_key == V5_0_SHORT_KEY
    assert current["short_run_id"] == V5_6_SHORT_RUN_ID
    assert explicit_v530["short_run_id"] == V5_3_SHORT_RUN_ID
    assert explicit_v510["short_run_id"] == V5_1_SHORT_RUN_ID
    assert explicit_v500["short_run_id"] == V5_0_SHORT_RUN_ID
    assert V5_1_REPORT.exists()
    assert latest["status"] == V5_1_STATUS
    assert latest["artifact_paths"]["report_json"] == "ai/eval/reports/rag-ingestion/runs/v5_1/report.json"
    assert latest["artifact_sha256"]["report_json_sha256"] == _sha256_file(V5_1_REPORT)
    assert latest["source_run_id"] == V5_0_SHORT_RUN_ID
    assert latest["current_resolves_to"] == V5_1_SHORT_KEY
    assert latest["v4_closeout_basis"] == V4_7_18_SHORT_KEY
    assert latest["official_metric_input_rows"] == 0
    assert latest["official_metric_input_rows_created"] == 0
    assert latest["official_metric_input_rows_scope"] == "v5_1_scaffold_created_rows_only"
    assert latest["existing_registry_backed_official_metric_input_rows_snapshot"] == 29
    assert latest["missing_user_owned_approval_artifact_count"] == 8
    assert latest["blocked_by_user_owned_gold_qrels_or_denominator_gate"] is True
    assert latest["training_dataset_created"] is False
    assert latest["training_manifest_jsonl_created"] is False
    assert latest["training_job_created"] is False
    assert latest["fine_tuning_dataset_export_created"] is False

    assert V5_1_SHORT_RUN_ID in progress
    assert V5_1_SHORT_RUN_ID in measurements
    assert V5_1_SHORT_RUN_ID in triage
    assert f"<!-- {V5_6_3_SHORT_RUN_ID}:progress-entry:start -->" in progress
    assert f"<!-- {V5_5_SHORT_RUN_ID}:measurements-entry:start -->" in measurements
    assert f"<!-- {V5_5_SHORT_RUN_ID}:triage-entry:start -->" in triage
    assert f"Overall status: `{V5_1_STATUS}`;" in progress
    assert "user-owned approval packet" in current_status_block
    assert "official_metric_input_rows=29" in current_status_block
    assert "v4_7_12" not in current_status_block
    assert "| official_metric_input_rows | 0 |" in measurements
    assert "| official_metric_input_rows_created | 0 |" in measurements
    assert "| official_metric_input_rows_scope | v5_1_scaffold_created_rows_only |" in measurements
    assert "| existing_registry_backed_official_metric_input_rows_snapshot | 29 |" in measurements
    assert "| blocked_by_user_owned_gold_qrels_or_denominator_gate | true |" in measurements
    assert "User-owned approval artifacts" in triage
    assert "Codex-owned validator-name placeholders" in triage
    assert "FT readiness compatibility" in triage
    _assert_status_recorded_in_progress_and_report(
        progress,
        current,
        short_run_id=V5_6_SHORT_RUN_ID,
        status=V5_6_STATUS,
    )
    assert registry.resolve_run(V5_3_SHORT_KEY, root=ROOT).logical_key == V5_3_SHORT_KEY
    assert registry.resolve_run(V5_0_SHORT_KEY, root=ROOT).logical_key == V5_0_SHORT_KEY

    for rel_path in (
        "ai/eval/reports/rag-ingestion/runs/v5_1/report.json",
        "ai/eval/reports/rag-ingestion/status.jsonl",
    ):
        assert subprocess.run(["git", "check-ignore", "-q", rel_path], cwd=ROOT).returncode == 0


def test_v510_check_report_rejects_opened_user_gates_official_rows_training_and_raw_payloads() -> None:
    from ai.eval import rag_v510_official_eval_gate_scaffolding as v510

    report = v510.build_report(root=ROOT, generated_at="2026-06-01T00:00:00Z")
    v510.check_report(report)

    for key in (
        "official_metric",
        "official_metric_denominator_usage_allowed",
        "gold_mutation",
        "qrels_mutation",
        "label_mutation",
        "expected_answer_mutation",
        "supporting_evidence_mutation",
        "denominator_mutation",
        "official_qrels_created",
        "official_relevance_labels_created",
        "official_answerability_labels_created",
        "official_gold_labels_created",
        "training_dataset_created",
        "training_manifest_jsonl_created",
        "training_job_created",
        "fine_tuning_dataset_export_created",
        "fine_tuning",
        "fine_tuning_started",
        "fine_tuning_executed",
        "ft_a_execution",
        "promotion_evidence",
        "product_success_evidence_allowed",
        "live_db_index_cache_readiness",
        "raw_xlsx_query_time_parsing",
        "direct_normalized_answer_value_matching",
        "formula_evaluation",
        "formula_text_exposure",
        "source_file_title_shortcut_used",
        "workbook_or_source_title_shortcut_used",
        "target_or_gold_locator_used_for_candidate_construction",
        "query_id_case_id_hack_used",
    ):
        mutated = json.loads(json.dumps(report))
        mutated[key] = True
        try:
            v510.check_report(mutated)
        except ValueError as exc:
            assert "opened forbidden gate" in str(exc)
            assert key in str(exc)
        else:
            raise AssertionError(f"v5_1 accepted {key}=True")

    for path, value, expected in (
        (("official_metric_input_rows",), 1, "official metric rows"),
        (("official_metric_input_rows_created",), 1, "official metric rows"),
        (("official_metric_input_rows_scope",), "all_repo_rows", "scope"),
        (("counters", "official_metric_input_rows"), 1, "official_metric_input_rows"),
        (("counters", "official_metric_input_rows_created"), 1, "official_metric_input_rows_created"),
        (("counters", "missing_user_owned_approval_artifact_count"), 7, "approval count"),
        (("official_eval_user_gate_ready",), True, "user gate"),
        (("official_eval_approval_artifact_found",), True, "approval artifact"),
        (("official_eval_gate_scaffold", "future_official_metric_builder_enabled"), True, "official metric builder"),
        (("official_eval_gate_scaffold", "blocked_by_user_owned_gold_qrels_or_denominator_gate"), False, "blocked"),
        (("ft_readiness_compatibility", "fine_tuning_dataset_export_created"), True, "dataset export"),
        (("source_run_id",), V4_7_18_SHORT_RUN_ID, "source run"),
    ):
        mutated = json.loads(json.dumps(report))
        cursor = mutated
        for part in path[:-1]:
            cursor = cursor[part]
        cursor[path[-1]] = value
        try:
            v510.check_report(mutated)
        except ValueError as exc:
            assert expected in str(exc)
        else:
            raise AssertionError(f"v5_1 accepted drift at {path}")

    mutated = json.loads(json.dumps(report))
    mutated["official_eval_gate_scaffold"]["required_user_owned_approval_artifacts"]["expected_answer"][
        "codex_may_infer"
    ] = True
    try:
        v510.check_report(mutated)
    except ValueError as exc:
        assert "user-owned approval" in str(exc)
    else:
        raise AssertionError("v5_1 accepted Codex-inferred expected answer approval")

    for key in ("prompt_payload", "raw_prompt_payload", "raw_response_payload", "raw_response"):
        mutated = json.loads(json.dumps(report))
        mutated["official_eval_gate_scaffold"][key] = "forbidden"
        try:
            v510.check_report(mutated)
        except ValueError as exc:
            assert "raw prompt/response" in str(exc)
        else:
            raise AssertionError(f"v5_1 accepted raw payload {key}")


def test_v510_write_path_synthesizes_v500_source_report_when_prior_ignored_report_is_missing(monkeypatch) -> None:
    import ai.scripts.rag_eval as runner
    from ai.eval import rag_v510_official_eval_gate_scaffolding as v510

    source_report = {
        "short_run_id": V5_0_SHORT_RUN_ID,
        "canonical_long_run_id": V5_0_LONG_RUN_ID,
        "status": V5_0_STATUS,
        "sentinel_from_check_run": True,
    }
    observed: dict[str, object] = {}

    def fake_check_run(key: str) -> dict[str, object]:
        assert key == "v5_0"
        return source_report

    def fake_build_report(*, root: Path, source_report: dict[str, object] | None = None, **_: object) -> dict[str, object]:
        assert root == ROOT
        assert source_report is not None
        assert source_report["sentinel_from_check_run"] is True
        observed["used_source_report"] = True
        return {
            "status": V5_1_STATUS,
            "artifact_paths": {"report_json": "ai/eval/reports/rag-ingestion/runs/v5_1/report.json"},
            "official_metric_input_rows": 0,
            "counters": {},
        }

    monkeypatch.setattr(runner, "check_run", fake_check_run)
    monkeypatch.setattr(v510, "build_report", fake_build_report)
    monkeypatch.setattr(v510, "write_report_bundle", lambda root, report: (report, {"report_json_sha256": "0" * 64}))
    monkeypatch.setattr(v510, "check_report", lambda report: None)
    monkeypatch.setattr(v510, "update_docs", lambda root, report: None)
    monkeypatch.setattr(v510, "append_status", lambda root, report, *, artifact_hashes: None)

    assert runner.main(["v5_1", "--write"]) == 0
    assert observed["used_source_report"] is True


def test_v520_xlsx_residual_candidate_state_taxonomy_keeps_residual_overlap_fail_closed() -> None:
    from ai.eval import rag_v520_xlsx_residual_candidate_only_retrieval_engineering as v520

    report = v520.build_report(root=ROOT, generated_at="2026-06-01T00:00:00Z")
    v520.check_report(report)

    taxonomy = report["xlsx_residual_candidate_state_taxonomy"]
    residual = report["xlsx_residual_basis"]
    candidate_state = taxonomy["candidate_state_buckets"]

    assert report["schema_version"] == f"{V5_2_SHORT_RUN_ID}_report_v1"
    assert report["logical_run_key"] == V5_2_SHORT_KEY
    assert report["short_run_id"] == V5_2_SHORT_RUN_ID
    assert report["canonical_long_run_id"] == V5_2_LONG_RUN_ID
    assert report["status"] == V5_2_STATUS
    assert report["source_run_id"] == V5_1_SHORT_RUN_ID
    assert report["source_logical_run_key"] == V5_1_SHORT_KEY
    assert report["current_resolves_to"] == V5_2_SHORT_KEY
    assert residual["source_short_run_id"] == V4_7_18_SHORT_RUN_ID
    assert residual["xlsx_row_count"] == 325
    assert residual["xlsx_v4_7_18_combined_target_hit_count"] == 26
    assert residual["xlsx_v4_7_18_combined_target_miss_count"] == 299
    assert residual["residual_overlap_counts_available"] is False
    assert residual["residual_overlap_counts_reason"] == "v4_7_18_report_exposes_aggregate_residuals_not_safe_row_level_residual_mask"
    assert taxonomy["source"] == "v4_7_18_candidate_budget_summary_read_only"
    assert taxonomy["candidate_state_bucket_count_sum"] == 325
    assert candidate_state["zero_candidate_structural_gap"]["count"] == 78
    assert candidate_state["budget_exhausted_diversity_gap"]["count"] == 109
    assert candidate_state["bounded_candidate_rank_gap"]["upper_bound_count"] == 138
    assert candidate_state["unclassified_residual_overlap"]["aggregate_count"] == 299
    assert candidate_state["value_only_or_forbidden_required"]["count_status"] == "intentionally_not_counted"
    assert taxonomy["candidate_count_distribution"] == {"0": 78, "1": 58, "2": 31, "3": 14, "4": 1, "5": 143}
    assert report["family_target_hit_regression_count"] == {"TEXT": 0, "PDF": 0, "XLSX": 0}
    assert report["safe_repair_applied"] is False
    assert report["safe_gain_claimed"] is False
    assert report["official_metric_input_rows"] == 0
    assert report["official_metric_input_rows_created"] == 0
    assert report["training_dataset_created"] is False


def test_v520_written_report_status_docs_current_alias_and_ignored_artifacts() -> None:
    from ai.eval import rag_eval_registry as registry
    from ai.eval import rag_v510_official_eval_gate_scaffolding as v510
    from ai.eval import rag_v520_xlsx_residual_candidate_only_retrieval_engineering as v520
    from ai.eval import rag_v530_pdf_text_residual_retrieval_evidence_hardening as v530
    from ai.eval import rag_v540_user_owned_official_eval_approval_packet as v540
    from ai.eval import rag_v550_user_approved_gold_packet_ingestion_and_official_metric_dry_run as v550
    import ai.scripts.rag_eval as runner

    report = registry.load_report("v5_2", root=ROOT)
    current = runner.check_run("current")
    explicit_v530 = runner.check_run("v5_3")
    explicit_v510 = runner.check_run("v5_1")
    v520.check_report(report)
    v560.check_report(current)
    v530.check_report(explicit_v530)
    v510.check_report(explicit_v510)
    latest = next(row for row in reversed(_read_jsonl(STATUS_JSONL)) if row.get("short_run_id") == V5_2_SHORT_RUN_ID)
    progress = PROGRESS_DOC.read_text(encoding="utf-8")
    current_status_block = progress.split("## Current Status", 1)[1].split("## Short History", 1)[0]
    measurements = MEASUREMENTS_DOC.read_text(encoding="utf-8")
    triage = TRIAGE_DOC.read_text(encoding="utf-8")

    assert runner.DEFAULT_RUN_KEY == V5_6_SHORT_KEY
    assert registry.resolve_run("current", root=ROOT).logical_key == V5_6_SHORT_KEY
    assert registry.resolve_run("v5_4", root=ROOT).logical_key == V5_4_SHORT_KEY
    assert registry.resolve_run("v5_3", root=ROOT).logical_key == V5_3_SHORT_KEY
    assert registry.resolve_run("v5_1", root=ROOT).logical_key == V5_1_SHORT_KEY
    assert current["short_run_id"] == V5_6_SHORT_RUN_ID
    assert explicit_v530["short_run_id"] == V5_3_SHORT_RUN_ID
    assert explicit_v510["short_run_id"] == V5_1_SHORT_RUN_ID
    assert V5_2_REPORT.exists()
    assert latest["status"] == V5_2_STATUS
    assert latest["artifact_paths"]["report_json"] == "ai/eval/reports/rag-ingestion/runs/v5_2/report.json"
    assert latest["artifact_sha256"]["report_json_sha256"] == _sha256_file(V5_2_REPORT)
    assert latest["source_run_id"] == V5_1_SHORT_RUN_ID
    assert latest["current_resolves_to"] == V5_2_SHORT_KEY
    assert latest["v4_closeout_basis"] == V4_7_18_SHORT_KEY
    assert latest["xlsx_v4_7_18_combined_target_miss_count"] == 299
    assert latest["zero_candidate_row_count"] == 78
    assert latest["candidate_budget_exhaustion_count"] == 109
    assert latest["bounded_candidate_not_budget_exhausted_row_count"] == 138
    assert latest["residual_overlap_counts_available"] is False
    assert latest["safe_repair_applied"] is False
    assert latest["safe_gain_claimed"] is False
    assert latest["official_metric_input_rows"] == 0
    assert latest["official_metric_input_rows_created"] == 0
    assert latest["training_dataset_created"] is False
    assert latest["fine_tuning_dataset_export_created"] is False

    assert V5_2_SHORT_RUN_ID in progress
    assert V5_2_SHORT_RUN_ID in measurements
    assert V5_2_SHORT_RUN_ID in triage
    assert f"<!-- {V5_6_3_SHORT_RUN_ID}:progress-entry:start -->" in progress
    assert f"<!-- {V5_5_SHORT_RUN_ID}:measurements-entry:start -->" in measurements
    assert f"<!-- {V5_5_SHORT_RUN_ID}:triage-entry:start -->" in triage
    assert f"Overall status: `{V5_2_STATUS}`;" in progress
    assert "candidate-state taxonomy" in progress
    assert "residual_overlap_counts_available=false" in progress
    assert "| zero_candidate_structural_gap | 78 |" in measurements
    assert "| budget_exhausted_diversity_gap | 109 |" in measurements
    assert "| bounded_candidate_rank_gap_upper_bound | 138 |" in measurements
    assert "row-level residual mask" in triage
    _assert_status_recorded_in_progress_and_report(
        progress,
        current,
        short_run_id=V5_6_SHORT_RUN_ID,
        status=V5_6_STATUS,
    )
    assert registry.resolve_run(V5_3_SHORT_KEY, root=ROOT).logical_key == V5_3_SHORT_KEY
    assert registry.resolve_run(V5_1_SHORT_KEY, root=ROOT).logical_key == V5_1_SHORT_KEY

    for rel_path in (
        "ai/eval/reports/rag-ingestion/runs/v5_2/report.json",
        "ai/eval/reports/rag-ingestion/status.jsonl",
    ):
        assert subprocess.run(["git", "check-ignore", "-q", rel_path], cwd=ROOT).returncode == 0


def test_v520_check_report_rejects_row_level_overlap_shortcuts_official_rows_and_training() -> None:
    from ai.eval import rag_v520_xlsx_residual_candidate_only_retrieval_engineering as v520

    report = v520.build_report(root=ROOT, generated_at="2026-06-01T00:00:00Z")
    v520.check_report(report)

    for key in (
        "official_metric",
        "official_metric_denominator_usage_allowed",
        "official_metric_dry_run_opened",
        "gold_mutation",
        "qrels_mutation",
        "training_dataset_created",
        "training_manifest_jsonl_created",
        "fine_tuning_dataset_export_created",
        "fine_tuning_started",
        "fine_tuning_executed",
        "safe_repair_applied",
        "safe_gain_claimed",
        "residual_overlap_recomputed",
        "row_level_residual_mask_created",
        "raw_xlsx_query_time_parsing",
        "direct_normalized_answer_value_matching",
        "formula_evaluation",
        "formula_text_exposure",
        "workbook_or_source_title_shortcut_used",
        "target_or_gold_locator_used_for_candidate_construction",
        "query_id_case_id_hack_used",
    ):
        mutated = json.loads(json.dumps(report))
        mutated[key] = True
        try:
            v520.check_report(mutated)
        except ValueError as exc:
            assert "v5_2" in str(exc)
        else:
            raise AssertionError(f"v5_2 accepted {key}=True")

    mutated = json.loads(json.dumps(report))
    mutated["official_metric_input_rows"] = 1
    try:
        v520.check_report(mutated)
    except ValueError as exc:
        assert "official metric rows" in str(exc)
    else:
        raise AssertionError("v5_2 accepted official metric rows")

    mutated = json.loads(json.dumps(report))
    mutated["xlsx_residual_basis"]["residual_overlap_counts_available"] = True
    try:
        v520.check_report(mutated)
    except ValueError as exc:
        assert "residual basis" in str(exc)
    else:
        raise AssertionError("v5_2 accepted row-level residual overlap")

    mutated = json.loads(json.dumps(report))
    mutated["xlsx_residual_candidate_state_taxonomy"]["candidate_state_buckets"]["value_only_or_forbidden_required"][
        "count_status"
    ] = "counted"
    try:
        v520.check_report(mutated)
    except ValueError as exc:
        assert "forbidden-required" in str(exc)
    else:
        raise AssertionError("v5_2 accepted counted value-only bucket")


def test_v520_write_path_synthesizes_v510_source_report_when_prior_ignored_report_is_missing(monkeypatch) -> None:
    import ai.scripts.rag_eval as runner
    from ai.eval import rag_v520_xlsx_residual_candidate_only_retrieval_engineering as v520

    source_report = {
        "short_run_id": V5_1_SHORT_RUN_ID,
        "canonical_long_run_id": V5_1_LONG_RUN_ID,
        "status": V5_1_STATUS,
        "sentinel_from_check_run": True,
    }
    observed: dict[str, object] = {}

    def fake_check_run(key: str) -> dict[str, object]:
        assert key == "v5_1"
        return source_report

    def fake_build_report(*, root: Path, source_report: dict[str, object] | None = None, **_: object) -> dict[str, object]:
        assert root == ROOT
        assert source_report is not None
        assert source_report["sentinel_from_check_run"] is True
        observed["used_source_report"] = True
        return {
            "status": V5_2_STATUS,
            "artifact_paths": {"report_json": "ai/eval/reports/rag-ingestion/runs/v5_2/report.json"},
            "official_metric_input_rows": 0,
            "counters": {},
        }

    monkeypatch.setattr(runner, "check_run", fake_check_run)
    monkeypatch.setattr(v520, "build_report", fake_build_report)
    monkeypatch.setattr(v520, "write_report_bundle", lambda root, report: (report, {"report_json_sha256": "0" * 64}))
    monkeypatch.setattr(v520, "check_report", lambda report: None)
    monkeypatch.setattr(v520, "update_docs", lambda root, report: None)
    monkeypatch.setattr(v520, "append_status", lambda root, report, *, artifact_hashes: None)

    assert runner.main(["v5_2", "--write"]) == 0
    assert observed["used_source_report"] is True


def test_v530_pdf_text_residual_retrieval_evidence_hardening_records_scope_and_boundaries() -> None:
    from ai.eval import rag_v530_pdf_text_residual_retrieval_evidence_hardening as v530

    report = v530.build_report(root=ROOT, generated_at="2026-06-01T00:00:00Z")
    v530.check_report(report)

    pdf = report["pdf_residual_taxonomy"]
    text = report["text_residual_taxonomy"]
    overlay = report["pdf_text_overlay_90_sample_taxonomy"]
    unavailable = report["unavailable_metrics"]

    assert report["schema_version"] == f"{V5_3_SHORT_RUN_ID}_report_v1"
    assert report["logical_run_key"] == V5_3_SHORT_KEY
    assert report["run_id"] == V5_3_SHORT_RUN_ID
    assert report["short_run_id"] == V5_3_SHORT_RUN_ID
    assert report["canonical_long_run_id"] == V5_3_LONG_RUN_ID
    assert report["status"] == V5_3_STATUS
    assert report["source_run_id"] == V5_2_SHORT_RUN_ID
    assert report["source_logical_run_key"] == V5_2_SHORT_KEY
    assert report["current_resolves_to"] == V5_3_SHORT_KEY
    handoff_note = report["cleanup_handoff_integrity_note"]
    assert handoff_note["cleanup_run_created"] is False
    assert handoff_note["current_remains"] == V5_3_SHORT_KEY
    assert handoff_note["status_progress_only_handoff_note_recorded"] is True
    assert handoff_note["v5_4_created"] is False
    assert handoff_note["v5_4_blocked_by_user_owned_approval_artifacts"] is True
    assert report["v4_closeout_short_run_id"] == V4_7_18_SHORT_RUN_ID
    assert report["v4_7_16_short_run_id"] == V4_7_16_SHORT_RUN_ID

    assert pdf["row_count"] == 325
    assert pdf["v4_7_18_combined_target_hit_count"] == 265
    assert pdf["aggregate_residual_count"] == 60
    assert pdf["target_hit_regression_count"] == 0
    assert pdf["pdf_candidate_overlay_attempted"] is False
    assert pdf["candidate_state_counts_available"] is False
    assert pdf["candidate_state_unavailable_reason"] == "pdf_candidate_overlay_not_attempted_in_v4_7_18"
    assert pdf["overlay_90_sample"]["target_not_in_topk_sample"] == 12
    assert pdf["overlay_90_sample"]["evidence_window_insufficient_sample"] == 16
    assert pdf["overlay_90_sample"]["source_family_route_ok_but_evidence_mismatch_sample"] == 17
    assert pdf["overlay_90_sample"]["query_too_broad_sample"] == 5
    assert pdf["overlay_90_sample"]["target_hit_evidence_context_repair_sample"] == 10

    assert text["row_count"] == 350
    assert text["v4_7_18_combined_target_hit_count"] == 232
    assert text["aggregate_residual_count"] == 118
    assert text["target_hit_regression_count"] == 0
    assert text["candidate_state_counts_available"] is True
    assert text["candidate_count"] == 1714
    assert text["zero_candidate_row_count"] == 2
    assert text["at_budget_row_count"] == 336
    assert text["candidate_budget_exhaustion_count"] == 336
    assert text["bounded_candidate_nonzero_not_at_budget_row_count"] == 12
    assert text["overlay_90_sample"]["target_not_in_topk_sample"] == 28
    assert text["overlay_90_sample"]["evidence_window_insufficient_sample"] == 30
    assert text["overlay_90_sample"]["source_family_route_ok_but_evidence_mismatch_sample"] == 30
    assert text["overlay_90_sample"]["target_hit_evidence_context_repair_sample"] == 2

    assert overlay["scope"] == "overlay_90_sample_not_full_pdf_text_denominator"
    assert unavailable["row_level_pdf_residual_mask"] == "unavailable_not_created"
    assert unavailable["row_level_text_residual_mask"] == "unavailable_not_created"
    assert unavailable["pdf_candidate_budget_taxonomy"] == "unavailable_pdf_candidate_overlay_not_attempted_in_v4_7_18"
    assert unavailable["official_hit_mrr_ndcg"] == "blocked_no_user_approved_qrels_denominator"
    assert report["family_target_hit_regression_count"] == {"TEXT": 0, "PDF": 0, "XLSX": 0}
    assert report["SearchView_vector_payload_role"] == "candidate_only"
    assert report["SourceAtom_EvidenceBundle_role"] == "evidence_truth"
    assert report["safe_repair_applied"] is False
    assert report["safe_gain_claimed"] is False
    assert report["official_metric_input_rows"] == 0
    assert report["official_metric_input_rows_created"] == 0
    assert report["training_dataset_created"] is False
    assert report["fine_tuning_dataset_export_created"] is False


def test_v530_written_report_status_docs_current_alias_and_ignored_artifacts() -> None:
    from ai.eval import rag_eval_registry as registry
    from ai.eval import rag_v520_xlsx_residual_candidate_only_retrieval_engineering as v520
    from ai.eval import rag_v530_pdf_text_residual_retrieval_evidence_hardening as v530
    from ai.eval import rag_v540_user_owned_official_eval_approval_packet as v540
    from ai.eval import rag_v550_user_approved_gold_packet_ingestion_and_official_metric_dry_run as v550
    import ai.scripts.rag_eval as runner

    report = registry.load_report("v5_3", root=ROOT)
    current = runner.check_run("current")
    explicit_v520 = runner.check_run("v5_2")
    v530.check_report(report)
    v560.check_report(current)
    v520.check_report(explicit_v520)
    latest = next(row for row in reversed(_read_jsonl(STATUS_JSONL)) if row.get("short_run_id") == V5_3_SHORT_RUN_ID)
    progress = PROGRESS_DOC.read_text(encoding="utf-8")
    current_status_block = progress.split("## Current Status", 1)[1].split("## Short History", 1)[0]
    measurements = MEASUREMENTS_DOC.read_text(encoding="utf-8")
    triage = TRIAGE_DOC.read_text(encoding="utf-8")

    assert runner.DEFAULT_RUN_KEY == V5_6_SHORT_KEY
    assert registry.resolve_run("current", root=ROOT).logical_key == V5_6_SHORT_KEY
    assert registry.resolve_run("v5_4", root=ROOT).logical_key == V5_4_SHORT_KEY
    assert registry.resolve_run("v5_3", root=ROOT).logical_key == V5_3_SHORT_KEY
    assert registry.resolve_run("v5_2", root=ROOT).logical_key == V5_2_SHORT_KEY
    assert registry.resolve_run("v5_1", root=ROOT).logical_key == V5_1_SHORT_KEY
    assert registry.resolve_run("v5_0", root=ROOT).logical_key == V5_0_SHORT_KEY
    assert registry.resolve_run("v4_7_18", root=ROOT).logical_key == V4_7_18_SHORT_KEY
    assert current["short_run_id"] == V5_6_SHORT_RUN_ID
    assert explicit_v520["short_run_id"] == V5_2_SHORT_RUN_ID
    assert V5_3_REPORT.exists()
    assert latest["schema_version"] == f"{V5_3_SHORT_RUN_ID}_status_event_v1"
    assert latest["status"] == V5_3_STATUS
    assert latest["run_id"] == V5_3_SHORT_RUN_ID
    assert latest["cleanup_run_created"] is False
    assert latest["current_remains"] == V5_3_SHORT_KEY
    assert latest["status_progress_only_handoff_note_recorded"] is True
    assert latest["v5_4_created"] is False
    assert latest["v5_4_blocked_by_user_owned_approval_artifacts"] is True
    assert latest["artifact_paths"]["report_json"] == "ai/eval/reports/rag-ingestion/runs/v5_3/report.json"
    assert latest["artifact_sha256"]["report_json_sha256"] == _sha256_file(V5_3_REPORT)
    assert latest["source_run_id"] == V5_2_SHORT_RUN_ID
    assert latest["current_resolves_to"] == V5_3_SHORT_KEY
    assert latest["v4_closeout_basis"] == V4_7_18_SHORT_KEY
    assert latest["text_v4_7_18_combined_target_miss_count"] == 118
    assert latest["pdf_v4_7_18_combined_target_miss_count"] == 60
    assert latest["pdf_text_residual_aggregate_count"] == 178
    assert latest["overlay_90_text_target_not_in_topk_total"] == 28
    assert latest["overlay_90_pdf_target_not_in_topk_total"] == 12
    assert latest["safe_repair_applied"] is False
    assert latest["safe_gain_claimed"] is False
    for key in (
        "pdf_text_repair_applied",
        "pdf_text_safe_gain_claimed",
        "official_metric",
        "official_metric_denominator_usage_allowed",
        "official_metric_dry_run_opened",
        "official_qrels_created",
        "official_relevance_labels_created",
        "official_answerability_labels_created",
        "official_gold_labels_created",
        "gold_mutation",
        "qrels_mutation",
        "label_mutation",
        "expected_answer_mutation",
        "supporting_evidence_mutation",
        "denominator_mutation",
        "training_manifest_jsonl_created",
        "training_job_created",
        "fine_tuning",
        "fine_tuning_started",
        "fine_tuning_executed",
        "ft_a_execution",
        "product_success_evidence_allowed",
        "production_db_mutated",
        "source_registry_mutated",
        "silver_mutation",
        "index_rebuilt",
        "cache_mutated",
        "residual_overlap_recomputed",
        "row_level_residual_mask_created",
        "per_query_candidates_written",
        "raw_pdf_query_time_parsing",
        "raw_xlsx_query_time_parsing",
        "broad_pdf_scan_or_full_page_dump",
        "direct_normalized_answer_value_matching",
        "formula_evaluation",
        "formula_text_exposure",
        "workbook_or_source_title_shortcut_used",
        "source_file_title_shortcut_used",
        "target_or_gold_locator_used_for_candidate_construction",
        "query_id_case_id_hack_used",
        "expected_or_supporting_gold_text_used",
        "raw_prompt_payload_written",
        "raw_response_payload_written",
    ):
        assert latest[key] is False
    assert latest["official_metric_input_rows"] == 0
    assert latest["official_metric_input_rows_created"] == 0
    assert latest["training_dataset_created"] is False
    assert latest["fine_tuning_dataset_export_created"] is False

    assert f"<!-- {V5_6_3_SHORT_RUN_ID}:progress-entry:start -->" in progress
    assert f"<!-- {V5_5_SHORT_RUN_ID}:measurements-entry:start -->" in measurements
    assert f"<!-- {V5_5_SHORT_RUN_ID}:triage-entry:start -->" in triage
    assert V5_5_SHORT_RUN_ID in current_status_block
    assert "`current` resolves to `v5_6`" in current_status_block
    assert "`v5_6_2`, `v5_5`, `v5_4`, `v5_3`, `v5_2`, `v5_1`, `v5_0`, and `v4_7_18` remain directly checkable" in current_status_block
    assert "user-approved gold packet" in current_status_block
    assert "official_metric_dry_run_opened=true" in current_status_block
    assert "| text_v4_7_18_combined_target_miss_count | 118 |" in measurements
    assert "| pdf_v4_7_18_combined_target_miss_count | 60 |" in measurements
    assert "| overlay_90_sample_scope | overlay_90_sample_not_full_pdf_text_denominator |" in measurements
    assert "raw PDF query-time parsing" in triage
    _assert_status_recorded_in_progress_and_report(
        progress,
        current,
        short_run_id=V5_6_SHORT_RUN_ID,
        status=V5_6_STATUS,
    )
    assert registry.resolve_run(V5_3_SHORT_KEY, root=ROOT).logical_key == V5_3_SHORT_KEY
    assert registry.resolve_run(V5_2_SHORT_KEY, root=ROOT).logical_key == V5_2_SHORT_KEY

    for rel_path in (
        "ai/eval/reports/rag-ingestion/runs/v5_3/report.json",
        "ai/eval/reports/rag-ingestion/status.jsonl",
    ):
        assert subprocess.run(["git", "check-ignore", "-q", rel_path], cwd=ROOT).returncode == 0


def test_v530_check_report_rejects_shortcuts_official_rows_training_and_residual_drift() -> None:
    from ai.eval import rag_v530_pdf_text_residual_retrieval_evidence_hardening as v530

    report = v530.build_report(root=ROOT, generated_at="2026-06-01T00:00:00Z")
    v530.check_report(report)

    for key in (
        "official_metric",
        "official_metric_denominator_usage_allowed",
        "official_metric_dry_run_opened",
        "gold_mutation",
        "qrels_mutation",
        "label_mutation",
        "expected_answer_mutation",
        "supporting_evidence_mutation",
        "denominator_mutation",
        "training_dataset_created",
        "training_manifest_jsonl_created",
        "training_job_created",
        "fine_tuning_dataset_export_created",
        "fine_tuning_started",
        "fine_tuning_executed",
        "promotion_evidence",
        "product_success_evidence_allowed",
        "live_db_index_cache_readiness",
        "safe_repair_applied",
        "safe_gain_claimed",
        "pdf_text_repair_applied",
        "pdf_text_safe_gain_claimed",
        "raw_pdf_query_time_parsing",
        "broad_pdf_scan_or_full_page_dump",
        "expected_or_supporting_gold_text_used",
        "target_or_gold_locator_used_for_candidate_construction",
        "workbook_or_source_title_shortcut_used",
        "source_file_title_shortcut_used",
        "query_id_case_id_hack_used",
    ):
        mutated = json.loads(json.dumps(report))
        mutated[key] = True
        try:
            v530.check_report(mutated)
        except ValueError as exc:
            assert "v5_3" in str(exc)
        else:
            raise AssertionError(f"v5_3 accepted {key}=True")

    for path, value, message in (
        (("source_run_id",), V5_1_SHORT_RUN_ID, "source run"),
        (("run_id",), "v5_3_drift", "run_id"),
        (("official_metric_input_rows",), 1, "official metric rows"),
        (("pdf_text_residual_basis", "TEXT", "v4_7_18_combined_target_miss_count"), 117, "residual basis"),
        (("pdf_text_residual_basis", "PDF", "v4_7_18_combined_target_miss_count"), 61, "residual basis"),
        (("family_target_hit_regression_count", "PDF"), 1, "family regression"),
        (("pdf_residual_taxonomy", "candidate_state_counts_available"), True, "PDF residual candidate state"),
        (("text_residual_taxonomy", "residual_overlap_with_candidate_state_available"), True, "TEXT residual overlap"),
        (("pdf_text_overlay_90_sample_taxonomy", "scope"), "full_denominator", "overlay scope"),
    ):
        mutated = json.loads(json.dumps(report))
        cursor = mutated
        for key in path[:-1]:
            cursor = cursor[key]
        cursor[path[-1]] = value
        try:
            v530.check_report(mutated)
        except ValueError as exc:
            assert message in str(exc)
        else:
            raise AssertionError(f"v5_3 accepted drift at {path}")


def test_v530_write_path_validates_report_before_writing_and_synthesizes_v520_source_report(monkeypatch) -> None:
    import ai.scripts.rag_eval as runner
    from ai.eval import rag_v530_pdf_text_residual_retrieval_evidence_hardening as v530

    source_report = {
        "short_run_id": V5_2_SHORT_RUN_ID,
        "canonical_long_run_id": V5_2_LONG_RUN_ID,
        "status": V5_2_STATUS,
        "sentinel_from_check_run": True,
    }
    observed: dict[str, object] = {}
    call_order: list[str] = []

    def fake_check_run(key: str) -> dict[str, object]:
        assert key == "v5_2"
        return source_report

    def fake_build_report(*, root: Path, source_report: dict[str, object] | None = None, **_: object) -> dict[str, object]:
        assert root == ROOT
        assert source_report is not None
        assert source_report["sentinel_from_check_run"] is True
        observed["used_source_report"] = True
        call_order.append("build")
        return {
            "status": V5_3_STATUS,
            "artifact_paths": {"report_json": "ai/eval/reports/rag-ingestion/runs/v5_3/report.json"},
            "official_metric_input_rows": 0,
            "counters": {},
        }

    def fake_check_report(report: dict[str, object]) -> None:
        call_order.append("check")

    def fake_write_report_bundle(root: Path, report: dict[str, object]) -> tuple[dict[str, object], dict[str, str]]:
        assert "check" in call_order
        call_order.append("write")
        return report, {"report_json_sha256": "0" * 64}

    monkeypatch.setattr(runner, "check_run", fake_check_run)
    monkeypatch.setattr(v530, "build_report", fake_build_report)
    monkeypatch.setattr(v530, "write_report_bundle", fake_write_report_bundle)
    monkeypatch.setattr(v530, "check_report", fake_check_report)
    monkeypatch.setattr(v530, "update_docs", lambda root, report: None)
    monkeypatch.setattr(v530, "append_status", lambda root, report, *, artifact_hashes: None)

    assert runner.main(["v5_3", "--write"]) == 0
    assert observed["used_source_report"] is True
    assert call_order == ["build", "check", "write", "check"]


def _assert_v540_user_owned_fields_blank(row: dict[str, object]) -> None:
    for field in V5_4_USER_OWNED_FIELDS:
        assert field in row
        assert row[field] in ("", None, "pending_user_review", [])


def test_v540_user_owned_approval_packet_materializes_blank_user_fields_and_closes_metric_gate() -> None:
    import ai.scripts.rag_eval as runner
    from ai.eval import rag_v540_user_owned_official_eval_approval_packet as v540

    source_report = runner.check_run("v5_3")
    report = v540.build_report(root=ROOT, source_report=source_report, generated_at="2026-06-01T00:00:00Z")
    v540.check_report(report)

    assert report["logical_run_key"] == V5_4_SHORT_KEY
    assert report["run_id"] == V5_4_SHORT_RUN_ID
    assert report["short_run_id"] == V5_4_SHORT_RUN_ID
    assert report["canonical_long_run_id"] == V5_4_LONG_RUN_ID
    assert report["status"] == V5_4_STATUS
    assert report["source_run_id"] == V5_3_SHORT_RUN_ID
    assert report["current_resolves_to"] == V5_4_SHORT_KEY
    assert report["review_surface_source"] == "existing_registry_backed_29_official_snapshot"
    assert report["review_packet_row_count"] == 29
    assert report["user_approval_packet_created"] is True
    assert report["user_policy_template_created"] is True
    assert report["user_review_packet_created"] is True
    assert report["user_review_packet_xlsx_created"] is True
    assert report["user_owned_final_fields_filled_by_codex"] is False
    assert report["official_metric_input_rows"] == 0
    assert report["official_metric_input_rows_created"] == 0
    assert report["official_metric_dry_run_opened"] is False
    assert report["official_eval_user_gate_ready"] is False
    assert report["protected_namespaces_touched"] == []
    assert report["artifact_paths"] == {
        "report_json": "ai/eval/reports/rag-ingestion/runs/v5_4/report.json",
        "status_jsonl": "ai/eval/reports/rag-ingestion/status.jsonl",
        "source_report_json": "ai/eval/reports/rag-ingestion/runs/v5_3/report.json",
        "user_owned_approval_schema_json": "ai/eval/reports/rag-ingestion/runs/v5_4/user_owned_approval_schema.json",
        "user_owned_policy_template_json": "ai/eval/reports/rag-ingestion/runs/v5_4/user_owned_policy_template.json",
        "user_review_packet_jsonl": "ai/eval/reports/rag-ingestion/runs/v5_4/user_review_packet.jsonl",
        "user_review_packet_csv": "ai/eval/reports/rag-ingestion/runs/v5_4/user_review_packet.csv",
        "user_review_packet_xlsx": "ai/eval/reports/rag-ingestion/runs/v5_4/user_review_packet.xlsx",
    }

    schema = report["user_owned_approval_schema"]
    assert tuple(schema["final_user_owned_fields"]) == V5_4_USER_OWNED_FIELDS
    for field in V5_4_USER_OWNED_FIELDS:
        field_policy = schema["field_policies"][field]
        assert field_policy["owner"] == "user"
        assert field_policy["codex_may_fill"] is False
        assert field_policy["required_before_official_metric"] is True

    policy_template = report["user_owned_policy_template"]
    assert policy_template["owner"] == "user"
    assert policy_template["codex_may_fill_user_owned_fields"] is False
    assert policy_template["official_metric_dry_run_requested"] is False
    assert policy_template["status"] == "pending_user_review"

    preview = report["user_review_packet_preview"]
    assert len(preview) == 3
    for row in preview:
        _assert_v540_user_owned_fields_blank(row)
        assert all(key.startswith("machine_") or key in V5_4_USER_OWNED_FIELDS for key in row)
    assert {row["machine_track"] for row in report["user_review_packet_rows"]} == {
        "pdf_business_ocr_mm",
        "text_namu_v2_1",
        "xlsx_business_structured",
    }


def test_v540_written_report_status_docs_current_alias_and_packet_artifacts() -> None:
    import openpyxl
    import ai.scripts.rag_eval as runner
    from ai.eval import rag_eval_registry as registry
    from ai.eval import rag_v530_pdf_text_residual_retrieval_evidence_hardening as v530
    from ai.eval import rag_v540_user_owned_official_eval_approval_packet as v540
    from ai.eval import rag_v550_user_approved_gold_packet_ingestion_and_official_metric_dry_run as v550

    report = registry.load_report("v5_4", root=ROOT)
    current = runner.check_run("current")
    explicit_v540 = runner.check_run("v5_4")
    explicit_v530 = runner.check_run("v5_3")
    v540.check_report(report)
    v560.check_report(current)
    v540.check_report(explicit_v540)
    v530.check_report(explicit_v530)
    latest = next(row for row in reversed(_read_jsonl(STATUS_JSONL)) if row.get("short_run_id") == V5_4_SHORT_RUN_ID)
    progress = PROGRESS_DOC.read_text(encoding="utf-8")
    current_status_block = progress.split("## Current Status", 1)[1].split("## Short History", 1)[0]
    measurements = MEASUREMENTS_DOC.read_text(encoding="utf-8")
    triage = TRIAGE_DOC.read_text(encoding="utf-8")

    assert runner.DEFAULT_RUN_KEY == V5_6_SHORT_KEY
    assert registry.resolve_run("current", root=ROOT).logical_key == V5_6_SHORT_KEY
    assert registry.resolve_run("v5_4", root=ROOT).logical_key == V5_4_SHORT_KEY
    assert registry.resolve_run("v5_3", root=ROOT).logical_key == V5_3_SHORT_KEY
    assert current["short_run_id"] == V5_6_SHORT_RUN_ID
    assert explicit_v540["short_run_id"] == V5_4_SHORT_RUN_ID
    assert explicit_v530["short_run_id"] == V5_3_SHORT_RUN_ID

    for path in (V5_4_REPORT, V5_4_SCHEMA, V5_4_POLICY_TEMPLATE, V5_4_PACKET_JSONL, V5_4_PACKET_CSV, V5_4_PACKET_XLSX):
        assert path.exists(), path

    packet_rows = _read_jsonl(V5_4_PACKET_JSONL)
    assert len(packet_rows) == 29
    for row in packet_rows:
        _assert_v540_user_owned_fields_blank(row)
        assert all(key.startswith("machine_") or key in V5_4_USER_OWNED_FIELDS for key in row)

    with V5_4_PACKET_CSV.open("r", encoding="utf-8-sig", newline="") as handle:
        csv_rows = list(csv.DictReader(handle))
    assert len(csv_rows) == 29
    assert "검수_안내" in csv_rows[0]
    assert "질문_확인" in csv_rows[0]
    assert "기존_답변_참고" in csv_rows[0]
    assert csv_rows[0]["검수_안내"].startswith("왼쪽 10개 user-owned 필드")
    assert csv_rows[0]["질문_확인"] == csv_rows[0]["machine_question_ko_hint"]
    for row in csv_rows:
        _assert_v540_user_owned_fields_blank(row)

    workbook = openpyxl.load_workbook(V5_4_PACKET_XLSX, read_only=True)
    sheet = workbook.active
    assert sheet.max_row == 30
    xlsx_header = [cell.value for cell in next(sheet.iter_rows(min_row=1, max_row=1))]
    assert xlsx_header[: len(V5_4_USER_OWNED_FIELDS)] == list(V5_4_USER_OWNED_FIELDS)
    assert "검수_안내" in xlsx_header
    assert "기존_근거_참고" in xlsx_header
    workbook.close()

    assert latest["schema_version"] == f"{V5_4_SHORT_RUN_ID}_status_event_v1"
    assert latest["event_type"] == "diagnostic_v5_4_user_owned_official_eval_approval_packet_nonprod"
    assert latest["status"] == V5_4_STATUS
    assert latest["source_run_id"] == V5_3_SHORT_RUN_ID
    assert latest["current_resolves_to"] == V5_4_SHORT_KEY
    assert latest["user_approval_packet_created"] is True
    assert latest["user_policy_template_created"] is True
    assert latest["user_review_packet_created"] is True
    assert latest["user_owned_final_fields_filled_by_codex"] is False
    assert latest["official_metric_input_rows"] == 0
    assert latest["official_metric_input_rows_created"] == 0
    assert latest["official_metric_dry_run_opened"] is False
    assert latest["official_eval_user_gate_ready"] is False
    assert latest["protected_namespaces_touched"] == []
    assert latest["artifact_paths"]["report_json"] == "ai/eval/reports/rag-ingestion/runs/v5_4/report.json"
    assert latest["artifact_paths"]["user_review_packet_csv"] == "ai/eval/reports/rag-ingestion/runs/v5_4/user_review_packet.csv"
    assert latest["artifact_sha256"]["user_review_packet_jsonl_sha256"] == _sha256_file(V5_4_PACKET_JSONL)
    assert latest["artifact_sha256"]["user_review_packet_csv_sha256"] == _sha256_file(V5_4_PACKET_CSV)
    assert latest["artifact_sha256"]["user_review_packet_xlsx_sha256"] == _sha256_file(V5_4_PACKET_XLSX)

    assert f"<!-- {V5_4_SHORT_RUN_ID}:progress-entry:start -->" in progress
    assert f"<!-- {V5_4_SHORT_RUN_ID}:measurements-entry:start -->" in measurements
    assert f"<!-- {V5_4_SHORT_RUN_ID}:triage-entry:start -->" in triage
    assert V5_5_SHORT_RUN_ID in current_status_block
    assert "`current` resolves to `v5_6`" in current_status_block
    assert "`v5_6_2`, `v5_5`, `v5_4`, `v5_3`, `v5_2`, `v5_1`, `v5_0`, and `v4_7_18` remain directly checkable" in current_status_block
    assert "user-approved gold packet ingestion" in current_status_block
    assert "| user_review_packet_row_count | 29 |" in measurements
    assert "| official_metric_dry_run_opened | false |" in measurements
    assert "Do not fill expected answers, supporting evidence, relevance, answerability, denominator" in triage
    _assert_status_recorded_in_progress_and_report(
        progress,
        current,
        short_run_id=V5_6_SHORT_RUN_ID,
        status=V5_6_STATUS,
    )
    assert registry.resolve_run("current", root=ROOT).logical_key == V5_6_SHORT_KEY

    for rel_path in (
        "ai/eval/reports/rag-ingestion/runs/v5_4/report.json",
        "ai/eval/reports/rag-ingestion/runs/v5_4/user_owned_approval_schema.json",
        "ai/eval/reports/rag-ingestion/runs/v5_4/user_owned_policy_template.json",
        "ai/eval/reports/rag-ingestion/runs/v5_4/user_review_packet.jsonl",
        "ai/eval/reports/rag-ingestion/runs/v5_4/user_review_packet.csv",
        "ai/eval/reports/rag-ingestion/runs/v5_4/user_review_packet.xlsx",
        "ai/eval/reports/rag-ingestion/status.jsonl",
    ):
        assert subprocess.run(["git", "check-ignore", "-q", rel_path], cwd=ROOT).returncode == 0, rel_path


def test_v540_check_report_rejects_filled_user_fields_official_rows_training_and_dry_run() -> None:
    from ai.eval import rag_v540_user_owned_official_eval_approval_packet as v540

    report = v540.build_report(root=ROOT, generated_at="2026-06-01T00:00:00Z")
    v540.check_report(report)

    for key in (
        "official_metric",
        "official_metric_denominator_usage_allowed",
        "official_metric_dry_run_opened",
        "gold_mutation",
        "qrels_mutation",
        "label_mutation",
        "expected_answer_mutation",
        "supporting_evidence_mutation",
        "denominator_mutation",
        "official_qrels_created",
        "official_relevance_labels_created",
        "official_answerability_labels_created",
        "official_gold_labels_created",
        "training_dataset_created",
        "training_manifest_jsonl_created",
        "training_job_created",
        "fine_tuning_dataset_export_created",
        "fine_tuning_started",
        "fine_tuning_executed",
        "promotion_evidence",
        "product_success_evidence_allowed",
        "live_db_index_cache_readiness",
        "raw_prompt_payload_written",
        "raw_response_payload_written",
    ):
        mutated = json.loads(json.dumps(report))
        mutated[key] = True
        try:
            v540.check_report(mutated)
        except ValueError as exc:
            assert "v5_4" in str(exc)
        else:
            raise AssertionError(f"v5_4 accepted {key}=True")

    for path, value, message in (
        (("source_run_id",), V5_2_SHORT_RUN_ID, "source run"),
        (("current_resolves_to",), V5_3_SHORT_KEY, "current"),
        (("official_metric_input_rows",), 1, "official metric rows"),
        (("official_metric_input_rows_created",), 1, "official metric rows"),
        (("protected_namespaces_touched",), ["ai/eval/eval_queries"], "protected"),
        (("review_surface_source",), "all_silver_residual_rows", "review surface"),
        (("review_packet_row_count",), 1000, "review packet"),
        (("user_owned_final_fields_filled_by_codex",), True, "user-owned"),
        (("user_review_packet_preview", 0, "expected_answer_ko"), "Codex-filled answer", "user-owned"),
        (("user_review_packet_rows", 0, "supporting_evidence_ids"), ["source_1"], "user-owned"),
    ):
        mutated = json.loads(json.dumps(report))
        cursor = mutated
        for key in path[:-1]:
            cursor = cursor[key]
        cursor[path[-1]] = value
        try:
            v540.check_report(mutated)
        except ValueError as exc:
            assert message in str(exc)
        else:
            raise AssertionError(f"v5_4 accepted drift at {path}")


def test_v540_write_path_validates_report_before_writing_and_synthesizes_v530_source_report(monkeypatch) -> None:
    import ai.scripts.rag_eval as runner
    from ai.eval import rag_v540_user_owned_official_eval_approval_packet as v540

    source_report = {
        "short_run_id": V5_3_SHORT_RUN_ID,
        "canonical_long_run_id": V5_3_LONG_RUN_ID,
        "status": V5_3_STATUS,
        "sentinel_from_check_run": True,
    }
    observed: dict[str, object] = {}
    call_order: list[str] = []

    def fake_check_run(key: str) -> dict[str, object]:
        assert key == "v5_3"
        return source_report

    def fake_build_report(*, root: Path, source_report: dict[str, object] | None = None, **_: object) -> dict[str, object]:
        assert root == ROOT
        assert source_report is not None
        assert source_report["sentinel_from_check_run"] is True
        observed["used_source_report"] = True
        call_order.append("build")
        return {
            "status": V5_4_STATUS,
            "artifact_paths": {"report_json": "ai/eval/reports/rag-ingestion/runs/v5_4/report.json"},
            "official_metric_input_rows": 0,
            "counters": {},
        }

    def fake_check_report(report: dict[str, object]) -> None:
        call_order.append("check")

    def fake_write_report_bundle(root: Path, report: dict[str, object]) -> tuple[dict[str, object], dict[str, str]]:
        assert "check" in call_order
        call_order.append("write")
        return report, {"report_json_sha256": "0" * 64}

    monkeypatch.setattr(runner, "check_run", fake_check_run)
    monkeypatch.setattr(v540, "build_report", fake_build_report)
    monkeypatch.setattr(v540, "write_report_bundle", fake_write_report_bundle)
    monkeypatch.setattr(v540, "check_report", fake_check_report)
    monkeypatch.setattr(v540, "update_docs", lambda root, report: None)
    monkeypatch.setattr(v540, "append_status", lambda root, report, *, artifact_hashes: None)

    assert runner.main(["v5_4", "--write"]) == 0
    assert observed["used_source_report"] is True
    assert call_order == ["build", "check", "write", "check"]


def test_v550_user_approved_gold_packet_ingests_only_v540_rows_and_builds_official_inputs() -> None:
    import ai.scripts.rag_eval as runner
    from ai.eval import rag_v550_user_approved_gold_packet_ingestion_and_official_metric_dry_run as v550

    report = v550.build_report(root=ROOT, generated_at="2026-06-01T06:30:00Z", source_report=runner.check_run("v5_4"))
    v550.check_report(report)

    assert report["short_run_id"] == V5_5_SHORT_RUN_ID
    assert report["canonical_long_run_id"] == V5_5_LONG_RUN_ID
    assert report["status"] == V5_5_STATUS
    assert report["source_run_id"] == V5_4_SHORT_RUN_ID
    assert report["source_logical_run_key"] == V5_4_SHORT_KEY
    assert report["current_resolves_to"] == V5_5_SHORT_KEY
    assert report["review_packet_source_row_count"] == 29
    assert report["review_packet_rows_by_track"] == {
        "text_namu_v2_1": 6,
        "xlsx_business_structured": 19,
        "pdf_business_ocr_mm": 4,
    }
    assert report["approval_scope"] == {
        "source_run_id": V5_4_SHORT_RUN_ID,
        "source_packet_paths": [
            "ai/eval/reports/rag-ingestion/runs/v5_4/user_review_packet.csv",
            "ai/eval/reports/rag-ingestion/runs/v5_4/user_review_packet.jsonl",
            "ai/eval/reports/rag-ingestion/runs/v5_4/user_review_packet.xlsx",
        ],
        "row_count": 29,
        "scope_policy": "exact_v5_4_user_review_packet_rows_only",
        "excluded_scopes": [
            "all_1000_silver_rows",
            "v5_2_or_v5_3_residual_rows",
            "overlay_90_sample",
            "xlsx_candidate_state_or_pdf_text_residual_taxonomy_denominators",
        ],
    }
    assert report["artifact_paths"] == {
        "report_json": "ai/eval/reports/rag-ingestion/runs/v5_5/report.json",
        "status_jsonl": "ai/eval/reports/rag-ingestion/status.jsonl",
        "source_report_json": "ai/eval/reports/rag-ingestion/runs/v5_4/report.json",
        "user_approved_gold_packet_jsonl": "ai/eval/reports/rag-ingestion/runs/v5_5/user_approved_gold_packet.jsonl",
        "user_approved_denominator_jsonl": "ai/eval/reports/rag-ingestion/runs/v5_5/user_approved_denominator.jsonl",
        "user_approved_qrels_jsonl": "ai/eval/reports/rag-ingestion/runs/v5_5/user_approved_qrels.jsonl",
        "user_approved_expected_answers_jsonl": "ai/eval/reports/rag-ingestion/runs/v5_5/user_approved_expected_answers.jsonl",
        "official_metric_input_jsonl": "ai/eval/reports/rag-ingestion/runs/v5_5/official_metric_input.jsonl",
        "official_metric_dry_run_result_json": "ai/eval/reports/rag-ingestion/runs/v5_5/official_metric_dry_run_result.json",
    }

    assert report["user_approved_gold_packet_created"] is True
    assert report["user_approved_denominator_created"] is True
    assert report["user_approved_qrels_created"] is True
    assert report["user_approved_expected_answers_created"] is True
    assert report["official_metric_input_rows"] == 29
    assert report["official_metric_input_rows_created"] == 29
    assert report["official_metric_dry_run_opened"] is True
    assert report["official_metric_dry_run_executed"] is True
    assert report["official_eval_user_gate_ready"] is True
    assert report["official_metric_finalized"] is False
    assert report["training_dataset_created"] is False
    assert report["fine_tuning"] is False
    assert report["promotion_evidence"] is False
    assert report["product_success_evidence_allowed"] is False
    assert report["live_db_index_cache_readiness"] is False
    assert report["protected_namespaces_touched"] == []

    approved = report["user_approved_gold_packet_rows"]
    denominator = report["user_approved_denominator_rows"]
    qrels = report["user_approved_qrels_rows"]
    expected_answers = report["user_approved_expected_answer_rows"]
    metric_input = report["official_metric_input_rows_payload"]
    assert len(approved) == len(denominator) == len(qrels) == len(expected_answers) == len(metric_input) == 29
    assert {row["source_v5_4_review_row_id"] for row in approved} == {
        row["source_v5_4_review_row_id"] for row in metric_input
    }

    source_rows = {row["machine_review_row_id"]: row for row in report["source_review_packet_rows"]}
    for row in approved:
        source = source_rows[row["source_v5_4_review_row_id"]]
        assert not any(key.startswith("machine_") for key in row)
        assert row["include_in_official_denominator"] == "INCLUDE"
        assert row["gold_status"] == "APPROVED"
        assert row["relevance_label"] == 3
        assert row["answerability_label"] == 3
        assert row["expected_answer_ko"] == source["machine_existing_expected_answer_ko_hint"]
        assert row["supporting_evidence_note"] == source["machine_existing_supporting_evidence_hint"]
        assert row["supporting_evidence_ids"]
        assert row["reviewer"] == "user-approved-bulk-review"
        assert row["reviewed_at"] == "2026-06-01T06:30:00Z"
        assert row["policy_note"] == (
            "user bulk-approved existing registry-backed human-audit-approved 29-row gold snapshot"
        )

    by_review_id = {row["source_v5_4_review_row_id"]: row for row in approved}
    assert by_review_id["v5_4_review_001"]["supporting_evidence_ids"] == ["a648c3a062d55aa3"]
    assert by_review_id["v5_4_review_007"]["supporting_evidence_ids"] == [
        "fde8e2c3-6faf-4c93-8c37-1b3a0b6789dd"
    ]
    assert by_review_id["v5_4_review_026"]["supporting_evidence_ids"] == [
        "7bf516bf-2a17-4303-86d8-3cffaa04846e"
    ]

    dry_run = report["official_metric_dry_run_result"]
    assert dry_run["status"] == "OFFICIAL_METRIC_DRY_RUN_EXECUTED_USER_APPROVED_PACKET_ONLY"
    assert dry_run["official_metric_input_rows"] == 29
    assert dry_run["row_count_by_track"] == report["review_packet_rows_by_track"]
    assert dry_run["contract_validation_passed"] is True
    assert dry_run["answer_quality_metric_computed"] is False
    assert dry_run["promotion_evidence"] is False


def test_v550_written_report_status_docs_current_alias_and_official_artifacts() -> None:
    import ai.scripts.rag_eval as runner
    from ai.eval import rag_eval_registry as registry
    from ai.eval import rag_v540_user_owned_official_eval_approval_packet as v540
    from ai.eval import rag_v550_user_approved_gold_packet_ingestion_and_official_metric_dry_run as v550

    report = registry.load_report("v5_5", root=ROOT)
    current = runner.check_run("current")
    explicit_v540 = runner.check_run("v5_4")
    v550.check_report(report)
    v560.check_report(current)
    v540.check_report(explicit_v540)
    latest = next(row for row in reversed(_read_jsonl(STATUS_JSONL)) if row.get("short_run_id") == V5_5_SHORT_RUN_ID)
    progress = PROGRESS_DOC.read_text(encoding="utf-8")
    current_status_block = progress.split("## Current Status", 1)[1].split("## Short History", 1)[0]
    measurements = MEASUREMENTS_DOC.read_text(encoding="utf-8")
    triage = TRIAGE_DOC.read_text(encoding="utf-8")

    assert runner.DEFAULT_RUN_KEY == V5_6_SHORT_KEY
    assert registry.resolve_run("current", root=ROOT).logical_key == V5_6_SHORT_KEY
    assert registry.resolve_run("v5_4", root=ROOT).logical_key == V5_4_SHORT_KEY
    assert current["short_run_id"] == V5_6_SHORT_RUN_ID
    assert explicit_v540["short_run_id"] == V5_4_SHORT_RUN_ID

    for path in (
        V5_5_REPORT,
        V5_5_APPROVED_GOLD_PACKET,
        V5_5_DENOMINATOR,
        V5_5_QRELS,
        V5_5_EXPECTED_ANSWERS,
        V5_5_OFFICIAL_METRIC_INPUT,
        V5_5_DRY_RUN_RESULT,
    ):
        assert path.exists(), path

    assert len(_read_jsonl(V5_5_APPROVED_GOLD_PACKET)) == 29
    assert len(_read_jsonl(V5_5_DENOMINATOR)) == 29
    assert len(_read_jsonl(V5_5_QRELS)) == 29
    assert len(_read_jsonl(V5_5_EXPECTED_ANSWERS)) == 29
    assert len(_read_jsonl(V5_5_OFFICIAL_METRIC_INPUT)) == 29
    assert _read_json(V5_5_DRY_RUN_RESULT)["official_metric_input_rows"] == 29

    assert latest["schema_version"] == f"{V5_5_SHORT_RUN_ID}_status_event_v1"
    assert latest["event_type"] == "v5_5_user_approved_gold_packet_ingestion_and_official_metric_dry_run_nonprod"
    assert latest["status"] == V5_5_STATUS
    assert latest["source_run_id"] == V5_4_SHORT_RUN_ID
    assert latest["current_resolves_to"] == V5_5_SHORT_KEY
    assert latest["official_metric_input_rows"] == 29
    assert latest["official_metric_input_rows_created"] == 29
    assert latest["official_metric_dry_run_opened"] is True
    assert latest["official_metric_dry_run_executed"] is True
    assert latest["official_eval_user_gate_ready"] is True
    assert latest["training_dataset_created"] is False
    assert latest["promotion_evidence"] is False
    assert latest["live_db_index_cache_readiness"] is False
    assert latest["protected_namespaces_touched"] == []
    assert latest["artifact_sha256"]["official_metric_input_jsonl_sha256"] == _sha256_file(V5_5_OFFICIAL_METRIC_INPUT)
    assert latest["artifact_sha256"]["official_metric_dry_run_result_json_sha256"] == _sha256_file(V5_5_DRY_RUN_RESULT)

    assert f"<!-- {V5_6_3_SHORT_RUN_ID}:progress-entry:start -->" in progress
    assert f"<!-- {V5_5_SHORT_RUN_ID}:measurements-entry:start -->" in measurements
    assert f"<!-- {V5_5_SHORT_RUN_ID}:triage-entry:start -->" in triage
    assert V5_5_SHORT_RUN_ID in current_status_block
    assert "`current` resolves to `v5_6`" in current_status_block
    assert "`v5_6_2`, `v5_5`, `v5_4`, `v5_3`, `v5_2`, `v5_1`, `v5_0`, and `v4_7_18` remain directly checkable" in current_status_block
    assert "scored_answer_rows=0" in current_status_block
    assert "backend_unavailable=true" in current_status_block
    assert "| official_metric_input_rows | 29 |" in measurements
    assert "| user_approved_qrels_rows | 29 |" in measurements
    assert "overlay-90" in triage
    assert registry.resolve_run("current", root=ROOT).logical_key == V5_6_SHORT_KEY

    for rel_path in (
        "ai/eval/reports/rag-ingestion/runs/v5_5/report.json",
        "ai/eval/reports/rag-ingestion/runs/v5_5/user_approved_gold_packet.jsonl",
        "ai/eval/reports/rag-ingestion/runs/v5_5/user_approved_denominator.jsonl",
        "ai/eval/reports/rag-ingestion/runs/v5_5/user_approved_qrels.jsonl",
        "ai/eval/reports/rag-ingestion/runs/v5_5/user_approved_expected_answers.jsonl",
        "ai/eval/reports/rag-ingestion/runs/v5_5/official_metric_input.jsonl",
        "ai/eval/reports/rag-ingestion/runs/v5_5/official_metric_dry_run_result.json",
        "ai/eval/reports/rag-ingestion/status.jsonl",
    ):
        assert subprocess.run(["git", "check-ignore", "-q", rel_path], cwd=ROOT).returncode == 0, rel_path


def test_v550_check_report_rejects_scope_expansion_missing_user_approval_and_closed_surface_drift() -> None:
    from ai.eval import rag_v550_user_approved_gold_packet_ingestion_and_official_metric_dry_run as v550

    report = v550.build_report(root=ROOT, generated_at="2026-06-01T06:30:00Z")
    v550.check_report(report)

    for path, value, message in (
        (("source_run_id",), V5_3_SHORT_RUN_ID, "source run"),
        (("current_resolves_to",), V5_4_SHORT_KEY, "current"),
        (("review_packet_source_row_count",), 30, "row count"),
        (("official_metric_input_rows",), 30, "official metric rows"),
        (("official_metric_input_rows_created",), 30, "official metric rows"),
        (("protected_namespaces_touched",), ["ai/eval/eval_queries"], "protected"),
        (("approval_scope", "row_count"), 1000, "approval scope"),
        (("user_approved_gold_packet_rows", 0, "relevance_label"), 2, "relevance"),
        (("user_approved_gold_packet_rows", 0, "answerability_label"), 2, "answerability"),
        (("user_approved_gold_packet_rows", 0, "expected_answer_ko"), "", "expected"),
        (("user_approved_gold_packet_rows", 0, "supporting_evidence_ids"), [], "supporting evidence"),
        (("official_metric_input_rows_payload", 0, "query_id"), "outside_v5_4_packet", "source"),
        (("official_metric_input_rows_payload", 0, "relevance_label"), 2, "projection"),
        (("user_approved_qrels_rows", 0, "supporting_evidence_ids"), ["wrong-evidence"], "projection"),
        (("user_approved_expected_answer_rows", 0, "expected_answer_ko"), "오염된 정답", "projection"),
        (("user_approved_denominator_rows", 0, "include_in_official_denominator"), "EXCLUDE", "projection"),
        (("official_metric_dry_run_result", "official_metric_input_rows"), 30, "dry run"),
    ):
        mutated = json.loads(json.dumps(report))
        cursor = mutated
        for key in path[:-1]:
            cursor = cursor[key]
        cursor[path[-1]] = value
        try:
            v550.check_report(mutated)
        except ValueError as exc:
            assert message in str(exc)
        else:
            raise AssertionError(f"v5_5 accepted drift at {path}")

    for key in (
        "training_dataset_created",
        "training_manifest_jsonl_created",
        "training_job_created",
        "fine_tuning_dataset_export_created",
        "fine_tuning_started",
        "fine_tuning_executed",
        "promotion_evidence",
        "product_success_evidence_allowed",
        "live_db_index_cache_readiness",
        "raw_prompt_payload_written",
        "raw_response_payload_written",
    ):
        mutated = json.loads(json.dumps(report))
        mutated[key] = True
        try:
            v550.check_report(mutated)
        except ValueError as exc:
            assert key in str(exc) or "v5_5" in str(exc)
        else:
            raise AssertionError(f"v5_5 accepted {key}=True")


def test_v550_check_report_rejects_written_child_artifact_hash_and_payload_drift() -> None:
    from ai.eval import rag_v550_user_approved_gold_packet_ingestion_and_official_metric_dry_run as v550

    report = _read_json(V5_5_REPORT)
    v550.check_report(report, root=ROOT)

    mutated_hash = json.loads(json.dumps(report))
    mutated_hash["artifact_sha256"]["official_metric_input_jsonl_sha256"] = "0" * 64
    try:
        v550.check_report(mutated_hash, root=ROOT)
    except ValueError as exc:
        assert "artifact hash" in str(exc)
    else:
        raise AssertionError("v5_5 accepted official metric input artifact hash drift")

    mutated_payload = json.loads(json.dumps(report))
    mutated_payload["official_metric_input_rows_payload"][0]["expected_answer_ko"] = "오염된 정답"
    try:
        v550.check_report(mutated_payload, root=ROOT)
    except ValueError as exc:
        assert "projection" in str(exc) or "artifact payload" in str(exc)
    else:
        raise AssertionError("v5_5 accepted embedded official metric payload drift")


def test_v550_write_path_validates_report_before_writing_and_synthesizes_v540_source_report(monkeypatch) -> None:
    import ai.scripts.rag_eval as runner
    from ai.eval import rag_v550_user_approved_gold_packet_ingestion_and_official_metric_dry_run as v550

    source_report = {
        "short_run_id": V5_4_SHORT_RUN_ID,
        "canonical_long_run_id": V5_4_LONG_RUN_ID,
        "status": V5_4_STATUS,
        "sentinel_from_check_run": True,
    }
    observed: dict[str, object] = {}
    call_order: list[str] = []

    def fake_check_run(key: str) -> dict[str, object]:
        assert key == "v5_4"
        return source_report

    def fake_build_report(*, root: Path, source_report: dict[str, object] | None = None, **_: object) -> dict[str, object]:
        assert root == ROOT
        assert source_report is not None
        assert source_report["sentinel_from_check_run"] is True
        observed["used_source_report"] = True
        call_order.append("build")
        return {
            "status": V5_5_STATUS,
            "artifact_paths": {"report_json": "ai/eval/reports/rag-ingestion/runs/v5_5/report.json"},
            "official_metric_input_rows": 29,
            "counters": {},
        }

    def fake_check_report(report: dict[str, object], **_: object) -> None:
        call_order.append("check")

    def fake_write_report_bundle(root: Path, report: dict[str, object]) -> tuple[dict[str, object], dict[str, str]]:
        assert "check" in call_order
        call_order.append("write")
        return report, {"report_json_sha256": "0" * 64}

    monkeypatch.setattr(runner, "check_run", fake_check_run)
    monkeypatch.setattr(v550, "build_report", fake_build_report)
    monkeypatch.setattr(v550, "write_report_bundle", fake_write_report_bundle)
    monkeypatch.setattr(v550, "check_report", fake_check_report)
    monkeypatch.setattr(v550, "update_docs", lambda root, report: None)
    monkeypatch.setattr(v550, "append_status", lambda root, report, *, artifact_hashes: None)

    assert runner.main(["v5_5", "--write"]) == 0
    assert observed["used_source_report"] is True
    assert call_order == ["build", "check", "write", "check"]


def test_v560_fail_closed_consumes_only_v550_official_metric_input_and_records_duplicate_policy() -> None:
    from ai.eval import rag_v560_official_metric_scored_execution_and_failure_attribution_nonprod as v560

    report = v560.build_report(
        root=ROOT,
        execute=True,
        env={
            "RAG_V5_6_ENABLE_OFFICIAL_METRIC_EXECUTION": "1",
            "RAG_V5_6_LOCAL_LLM_BASE_URL": "http://127.0.0.1:9/v1",
        },
        backend_timeout_seconds=1,
        generated_at="2026-06-04T00:00:00Z",
    )
    v560.check_report(report)
    result = report["official_metric_scored_result"]

    assert report["short_run_id"] == V5_6_SHORT_RUN_ID
    assert report["source_run_id"] == V5_5_SHORT_RUN_ID
    assert report["current_resolves_to"] == V5_6_SHORT_KEY
    assert report["non_production"] is True
    assert report["diagnostic_only"] is False
    assert report["approval_scope"] == {
        "source_run_id": V5_5_SHORT_RUN_ID,
        "source_artifact_path": "ai/eval/reports/rag-ingestion/runs/v5_5/official_metric_input.jsonl",
        "row_count": 29,
        "scope_policy": "exact_v5_5_official_metric_input_rows_only",
        "excluded_scopes": [
            "silver_rows",
            "v5_2_or_v5_3_residual_rows",
            "overlay_90_rows",
            "xlsx_candidate_state_buckets",
            "pdf_text_residual_taxonomy_denominators",
        ],
    }
    assert report["official_metric_input_rows"] == 29
    assert report["official_metric_input_rows_consumed"] == 29
    assert report["row_count_by_track"] == {
        "text_namu_v2_1": 6,
        "xlsx_business_structured": 19,
        "pdf_business_ocr_mm": 4,
    }
    assert report["source_artifact_validation"]["path_matches_v5_5_report"] is True
    assert report["source_artifact_validation"]["sha256_matches_v5_5_report"] is True
    assert result["status"] == "backend_unavailable"
    assert result["backend_unavailable"] is True
    assert result["scored_answer_rows"] == 0
    assert result["answer_quality_metric_computed"] is False
    assert result["official_metric_finalized"] is False
    assert result["failure_category_counts"] == {"backend_unavailable": 29}
    assert report["failure_attribution_row_count"] == 29
    assert {row["failure_category"] for row in report["failure_attribution_rows"]} == {"backend_unavailable"}
    assert report["duplicate_supporting_evidence_policy"] == (
        "recorded_for_locator_precision_audit; row-level citation_locator remains authoritative"
    )
    assert report["duplicate_supporting_evidence_id_count"] >= 1
    assert report["protected_namespaces_touched"] == []
    for key in (
        "gold_mutation",
        "qrels_mutation",
        "expected_answer_mutation",
        "supporting_evidence_mutation",
        "denominator_mutation",
        "training_dataset_created",
        "fine_tuning",
        "ft_a_execution",
        "promotion_evidence",
        "product_success_evidence_allowed",
        "live_db_index_cache_readiness",
        "production_db_mutated",
        "index_rebuilt",
        "cache_mutated",
        "raw_prompt_payload_written",
        "raw_response_payload_written",
    ):
        assert report[key] is False, key
    generated_text = json.dumps(report, ensure_ascii=False)
    for forbidden in ('"raw_prompt_payload":', '"raw_response_payload":', '"raw_llm_response":'):
        assert forbidden not in generated_text


def test_v560_injected_answer_and_scorer_backends_score_all_29_rows_without_raw_payloads() -> None:
    from ai.eval import rag_v560_official_metric_scored_execution_and_failure_attribution_nonprod as v560

    calls = {"answer": 0, "scorer": 0}

    def answer_client(prompt: str) -> str:
        calls["answer"] += 1
        payload = json.loads(prompt)
        context = payload["official_metric_row"]
        return json.dumps(
            {
                "final_answer": context["supporting_evidence_note"],
                "citations": context["supporting_evidence_ids"][:1],
                "abstain": False,
            },
            ensure_ascii=False,
            sort_keys=True,
        )

    def scorer_client(prompt: str) -> str:
        calls["scorer"] += 1
        payload = json.loads(prompt)
        cited = bool(payload["generated_citations"])
        passed = cited and bool(payload["generated_answer"])
        return json.dumps(
            {
                "passed": passed,
                "answer_score": 1.0,
                "citation_support_score": 1.0,
                "failure_category": "pass" if passed else "answer_wrong",
                "failure_detail": "generated answer exactly matches expected answer and cites approved evidence id",
            },
            ensure_ascii=False,
            sort_keys=True,
        )

    report = v560.build_report(
        root=ROOT,
        execute=True,
        env={"RAG_V5_6_ENABLE_OFFICIAL_METRIC_EXECUTION": "1"},
        answer_client=answer_client,
        scorer_client=scorer_client,
        generated_at="2026-06-04T00:00:00Z",
    )
    v560.check_report(report)
    result = report["official_metric_scored_result"]

    assert calls == {"answer": 29, "scorer": 29}
    assert result["status"] == "scored"
    assert result["backend_unavailable"] is False
    assert result["answer_quality_metric_computed"] is True
    assert result["official_metric_finalized"] is True
    assert result["scored_answer_rows"] == 29
    assert result["pass_count"] == 29
    assert result["fail_count"] == 0
    assert result["failure_category_counts"] == {"pass": 29}
    assert len(result["row_results"]) == 29
    assert len(report["failure_attribution_rows"]) == 29
    assert all(row["failure_category"] == "pass" for row in report["failure_attribution_rows"])
    assert all(row["prompt_sha256"] for row in result["row_results"])
    assert all(row["scorer_prompt_sha256"] for row in result["row_results"])
    assert all(row["raw_answer_response_sha256"] for row in result["row_results"])
    assert all(row["raw_scorer_response_sha256"] for row in result["row_results"])
    generated_text = json.dumps(report, ensure_ascii=False)
    for forbidden in ('"raw_prompt_payload":', '"raw_response_payload":', '"raw_llm_response":'):
        assert forbidden not in generated_text


def test_v560_check_report_rejects_scope_expansion_fake_noop_metrics_and_protected_drift() -> None:
    from ai.eval import rag_v560_official_metric_scored_execution_and_failure_attribution_nonprod as v560

    report = v560.build_report(root=ROOT, execute=False, generated_at="2026-06-04T00:00:00Z")
    v560.check_report(report)

    for path, value, message in (
        (("source_run_id",), V5_4_SHORT_RUN_ID, "source run"),
        (("current_resolves_to",), V5_5_SHORT_KEY, "current"),
        (("approval_scope", "row_count"), 30, "approval scope"),
        (("official_metric_input_rows",), 30, "official metric input"),
        (("official_metric_input_rows_consumed",), 30, "official metric input"),
        (("row_count_by_track", "text_namu_v2_1"), 7, "row count"),
        (("official_metric_scored_result", "scored_answer_rows"), 29, "scored rows"),
        (("official_metric_scored_result", "answer_quality_metric_computed"), True, "fake metric"),
        (("official_metric_scored_result", "official_metric_finalized"), True, "fake metric"),
        (("official_metric_scored_result", "failure_category_counts", "backend_unavailable"), 28, "failure"),
        (("failure_attribution_rows", 0, "failure_category"), "answer_wrong", "backend unavailable"),
        (("backend_preflight", "noop_backend_used"), True, "noop"),
        (("protected_namespaces_touched",), ["ai/eval/eval_queries"], "protected"),
        (("duplicate_supporting_evidence_policy",), "collapse duplicate evidence ids", "duplicate"),
    ):
        mutated = json.loads(json.dumps(report))
        cursor = mutated
        for key in path[:-1]:
            cursor = cursor[key]
        cursor[path[-1]] = value
        try:
            v560.check_report(mutated)
        except ValueError as exc:
            assert message in str(exc)
        else:
            raise AssertionError(f"v5_6 accepted drift at {path}")

    for key in (
        "gold_mutation",
        "qrels_mutation",
        "expected_answer_mutation",
        "supporting_evidence_mutation",
        "denominator_mutation",
        "training_dataset_created",
        "fine_tuning",
        "ft_a_execution",
        "promotion_evidence",
        "product_success_evidence_allowed",
        "live_db_index_cache_readiness",
        "production_db_mutated",
        "index_rebuilt",
        "cache_mutated",
        "raw_prompt_payload_written",
        "raw_response_payload_written",
    ):
        mutated = json.loads(json.dumps(report))
        mutated[key] = True
        try:
            v560.check_report(mutated)
        except ValueError as exc:
            assert key in str(exc) or "closed" in str(exc)
        else:
            raise AssertionError(f"v5_6 accepted {key}=True")


def test_v560_write_path_writes_scored_result_failure_attribution_status_and_ignored_artifacts(tmp_path: Path) -> None:
    from ai.eval import rag_v560_official_metric_scored_execution_and_failure_attribution_nonprod as v560

    report = v560.build_report(root=ROOT, execute=False, generated_at="2026-06-04T00:00:00Z")
    written, artifact_hashes = v560.write_report_bundle(tmp_path, report)
    v560.check_report(written, root=tmp_path)
    v560.append_status(tmp_path, written, artifact_hashes=artifact_hashes)

    paths = written["artifact_paths"]
    assert (tmp_path / paths["report_json"]).exists()
    assert (tmp_path / paths["official_metric_scored_result_json"]).exists()
    assert (tmp_path / paths["failure_attribution_jsonl"]).exists()
    assert _read_json(tmp_path / paths["official_metric_scored_result_json"]) == written["official_metric_scored_result"]
    assert _read_jsonl(tmp_path / paths["failure_attribution_jsonl"]) == written["failure_attribution_rows"]
    assert artifact_hashes["official_metric_scored_result_json_sha256"] == _sha256_file(
        tmp_path / paths["official_metric_scored_result_json"]
    )
    assert artifact_hashes["failure_attribution_jsonl_sha256"] == _sha256_file(
        tmp_path / paths["failure_attribution_jsonl"]
    )
    status_rows = _read_jsonl(tmp_path / "ai/eval/reports/rag-ingestion/status.jsonl")
    latest = status_rows[-1]
    assert latest["short_run_id"] == V5_6_SHORT_RUN_ID
    assert latest["status"] == V5_6_STATUS
    assert latest["backend_unavailable"] is True
    assert latest["official_metric_input_rows"] == 29
    assert latest["scored_answer_rows"] == 0
    assert latest["failure_category_counts"] == {"backend_unavailable": 29}

    for rel_path in (
        "ai/eval/reports/rag-ingestion/runs/v5_6/report.json",
        "ai/eval/reports/rag-ingestion/runs/v5_6/official_metric_scored_result.json",
        "ai/eval/reports/rag-ingestion/runs/v5_6/failure_attribution.jsonl",
        "ai/eval/reports/rag-ingestion/status.jsonl",
    ):
        assert subprocess.run(["git", "check-ignore", "-q", rel_path], cwd=ROOT).returncode == 0, rel_path


def test_v560_write_path_validates_report_before_writing_and_uses_v550_source(monkeypatch) -> None:
    import ai.scripts.rag_eval as runner
    from ai.eval import rag_v560_official_metric_scored_execution_and_failure_attribution_nonprod as v560

    source_report = {
        "short_run_id": V5_5_SHORT_RUN_ID,
        "canonical_long_run_id": V5_5_LONG_RUN_ID,
        "status": V5_5_STATUS,
        "sentinel_from_check_run": True,
    }
    observed: dict[str, object] = {}
    call_order: list[str] = []

    def fake_check_run(key: str) -> dict[str, object]:
        assert key == "v5_5"
        return source_report

    def fake_build_report(*, root: Path, source_report: dict[str, object] | None = None, **_: object) -> dict[str, object]:
        assert root == ROOT
        assert source_report is not None
        assert source_report["sentinel_from_check_run"] is True
        observed["used_source_report"] = True
        call_order.append("build")
        return {
            "status": V5_6_STATUS,
            "artifact_paths": {"report_json": "ai/eval/reports/rag-ingestion/runs/v5_6/report.json"},
            "official_metric_input_rows": 29,
            "official_metric_scored_result": {"scored_answer_rows": 0},
            "counters": {},
        }

    def fake_check_report(report: dict[str, object], **_: object) -> None:
        call_order.append("check")

    def fake_write_report_bundle(root: Path, report: dict[str, object]) -> tuple[dict[str, object], dict[str, str]]:
        assert "check" in call_order
        call_order.append("write")
        return report, {"report_json_sha256": "0" * 64}

    monkeypatch.setattr(runner, "check_run", fake_check_run)
    monkeypatch.setattr(v560, "build_report", fake_build_report)
    monkeypatch.setattr(v560, "write_report_bundle", fake_write_report_bundle)
    monkeypatch.setattr(v560, "check_report", fake_check_report)
    monkeypatch.setattr(v560, "update_docs", lambda root, report: None)
    monkeypatch.setattr(v560, "append_status", lambda root, report, *, artifact_hashes: None)

    assert runner.main(["v5_6", "--write"]) == 0
    assert observed["used_source_report"] is True
    assert call_order == ["build", "check", "write", "check"]


def test_v562_env_gate_disabled_records_execution_gate_disabled_without_fake_metrics() -> None:
    from ai.eval import rag_v562_official_metric_backend_enabled_preflight_scored_rerun_nonprod as v562

    report = v562.build_report(root=ROOT, execute=True, env={}, generated_at="2026-06-04T00:00:00Z")
    v562.check_report(report)
    result = report["official_metric_scored_result"]

    assert report["logical_run_key"] == V5_6_2_SHORT_KEY
    assert report["short_run_id"] == V5_6_2_SHORT_RUN_ID
    assert report["canonical_long_run_id"] == V5_6_2_LONG_RUN_ID
    assert report["status"] == V5_6_2_STATUS
    assert report["source_run_id"] == V5_5_SHORT_RUN_ID
    assert report["current_resolves_to"] == V5_6_SHORT_KEY
    assert report["approval_scope"] == {
        "source_run_id": V5_5_SHORT_RUN_ID,
        "source_artifact_path": "ai/eval/reports/rag-ingestion/runs/v5_5/official_metric_input.jsonl",
        "row_count": 29,
        "scope_policy": "exact_v5_5_official_metric_input_rows_only",
        "excluded_scopes": [
            "silver_rows",
            "v5_2_or_v5_3_residual_rows",
            "overlay_90_rows",
            "xlsx_candidate_state_buckets",
            "pdf_text_residual_taxonomy_denominators",
        ],
    }
    assert report["official_metric_input_rows"] == 29
    assert report["official_metric_input_rows_consumed"] == 29
    assert report["row_count_by_track"] == {
        "text_namu_v2_1": 6,
        "xlsx_business_structured": 19,
        "pdf_business_ocr_mm": 4,
    }
    assert report["source_artifact_validation"]["path_matches_v5_5_report"] is True
    assert report["source_artifact_validation"]["sha256_matches_v5_5_report"] is True
    assert report["source_artifact_validation"]["sha256_matches_v5_6_report"] is True
    assert report["official_metric_input_row_ref_count"] == 29
    assert len(report["official_metric_input_row_refs"]) == 29
    assert "official_metric_input_rows_payload" not in report
    assert report["backend_preflight"]["status"] == "EXECUTION_GATE_DISABLED_FAIL_CLOSED"
    assert report["backend_preflight"]["failure_category"] == "execution_gate_disabled"
    assert report["backend_preflight"]["available"] is False
    assert report["backend_preflight"]["env_enabled"] is False
    assert report["backend_preflight"]["enabled_env_var"] == "RAG_V5_6_2_ENABLE_OFFICIAL_METRIC_EXECUTION"
    assert result["status"] == "preflight_failed"
    assert result["backend_unavailable"] is True
    assert result["official_metric_input_rows"] == 29
    assert result["scored_answer_rows"] == 0
    assert result["answer_quality_metric_computed"] is False
    assert result["official_metric_finalized"] is False
    assert result["pass_count"] == 0
    assert result["fail_count"] == 29
    assert result["pass_fail_counts_interpretable_as_quality_metric"] is False
    assert result["failure_category_counts"] == {"execution_gate_disabled": 29}
    assert {row["failure_category"] for row in report["failure_attribution_rows"]} == {"execution_gate_disabled"}
    assert report["duplicate_supporting_evidence_policy"] == (
        "recorded_for_locator_precision_audit; row-level citation_locator remains authoritative"
    )
    assert report["duplicate_supporting_evidence_id_count"] == 1
    assert report["protected_namespaces_touched"] == []
    for key in v562.CLOSED_FALSE_KEYS:
        assert report[key] is False, key
    generated_text = json.dumps(report, ensure_ascii=False)
    for forbidden in (
        '"raw_prompt_payload":',
        '"raw_response_payload":',
        '"raw_llm_response":',
        '"official_metric_input_rows_payload":',
        '"expected_answer_ko":',
        '"supporting_evidence_note":',
    ):
        assert forbidden not in generated_text


def test_v562_env_enabled_answer_backend_unreachable_is_not_execution_gate_disabled() -> None:
    from ai.eval import rag_v562_official_metric_backend_enabled_preflight_scored_rerun_nonprod as v562

    report = v562.build_report(
        root=ROOT,
        execute=True,
        env={
            v562.ENABLE_ENV_VAR: "1",
            v562.BASE_URL_ENV_VAR: "http://127.0.0.1:9",
        },
        backend_timeout_seconds=1,
        generated_at="2026-06-04T00:00:00Z",
    )
    v562.check_report(report)
    result = report["official_metric_scored_result"]

    assert report["backend_preflight"]["env_enabled"] is True
    assert report["backend_preflight"]["failure_category"] == "answer_generation_backend_unreachable"
    assert "execution_gate_disabled" not in result["failure_category_counts"]
    assert result["failure_category_counts"] == {"answer_generation_backend_unreachable": 29}
    assert result["scored_answer_rows"] == 0
    assert result["answer_quality_metric_computed"] is False


def test_v562_answer_generation_model_unavailable_is_separate_from_backend_unreachable(monkeypatch) -> None:
    from ai.eval import rag_v562_official_metric_backend_enabled_preflight_scored_rerun_nonprod as v562

    def fake_request_json(url: str, *, payload: dict[str, object] | None, timeout_seconds: int) -> object:
        assert payload is None
        assert url.endswith("/models")
        return {"data": [{"id": "different-model"}]}

    monkeypatch.setattr(v562.local_llm, "request_json", fake_request_json)
    report = v562.build_report(
        root=ROOT,
        execute=True,
        env={
            v562.ENABLE_ENV_VAR: "1",
            v562.MODEL_ENV_VAR: "missing-model",
        },
        backend_timeout_seconds=1,
        generated_at="2026-06-04T00:00:00Z",
    )
    v562.check_report(report)

    assert report["backend_preflight"]["failure_category"] == "answer_generation_model_unavailable"
    assert report["official_metric_scored_result"]["failure_category_counts"] == {
        "answer_generation_model_unavailable": 29
    }


def test_v562_scorer_preflight_failures_are_separate_after_answer_backend_probe() -> None:
    from ai.eval import rag_v562_official_metric_backend_enabled_preflight_scored_rerun_nonprod as v562

    calls = {"answer": 0, "scorer": 0}

    def answer_client(prompt: str) -> str:
        calls["answer"] += 1
        payload = json.loads(prompt)
        assert payload["task"] == "v5_6_2_non_gold_answer_generation_probe"
        return json.dumps({"final_answer": "probe answer", "citations": [], "abstain": False}, sort_keys=True)

    unreachable = v562.build_report(
        root=ROOT,
        execute=True,
        env={
            v562.ENABLE_ENV_VAR: "1",
            v562.SCORER_BASE_URL_ENV_VAR: "http://127.0.0.1:9",
        },
        answer_client=answer_client,
        backend_timeout_seconds=1,
        generated_at="2026-06-04T00:00:00Z",
    )
    v562.check_report(unreachable)
    assert calls == {"answer": 1, "scorer": 0}
    assert unreachable["backend_preflight"]["failure_category"] == "scorer_backend_unreachable"
    assert unreachable["official_metric_scored_result"]["failure_category_counts"] == {"scorer_backend_unreachable": 29}

    def scorer_contract_client(prompt: str) -> str:
        calls["scorer"] += 1
        payload = json.loads(prompt)
        assert payload["task"] == "v5_6_2_non_gold_scorer_contract_probe"
        return json.dumps({"unexpected": True}, sort_keys=True)

    contract = v562.build_report(
        root=ROOT,
        execute=True,
        env={v562.ENABLE_ENV_VAR: "1"},
        answer_client=answer_client,
        scorer_client=scorer_contract_client,
        generated_at="2026-06-04T00:00:00Z",
    )
    v562.check_report(contract)
    assert calls == {"answer": 2, "scorer": 1}
    assert contract["backend_preflight"]["failure_category"] == "scorer_contract_unavailable"
    assert contract["official_metric_scored_result"]["failure_category_counts"] == {
        "scorer_contract_unavailable": 29
    }


def test_v562_injected_answer_and_scorer_backends_score_all_29_rows_after_non_gold_probes() -> None:
    from ai.eval import rag_v562_official_metric_backend_enabled_preflight_scored_rerun_nonprod as v562

    calls = {"answer": 0, "scorer": 0}

    def answer_client(prompt: str) -> str:
        calls["answer"] += 1
        payload = json.loads(prompt)
        if payload["task"] == "v5_6_2_non_gold_answer_generation_probe":
            return json.dumps({"final_answer": "probe answer", "citations": [], "abstain": False}, sort_keys=True)
        prompt_text = json.dumps(payload, ensure_ascii=False)
        for forbidden in (
            "query_id",
            "expected_answer_ko",
            "supporting_evidence_note",
            "supporting_evidence_ids",
            "citation_locator",
        ):
            assert forbidden not in prompt_text
        context = payload["official_metric_row"]
        return json.dumps(
            {
                "final_answer": f"synthetic answer for {context['track']}",
                "citations": ["runtime-citation"],
                "abstain": False,
            },
            ensure_ascii=False,
            sort_keys=True,
        )

    def scorer_client(prompt: str) -> str:
        calls["scorer"] += 1
        payload = json.loads(prompt)
        if payload["task"] == "v5_6_2_non_gold_scorer_contract_probe":
            return json.dumps(
                {
                    "passed": True,
                    "answer_score": 1.0,
                    "citation_support_score": 1.0,
                    "failure_category": "pass",
                    "failure_detail": "probe contract ok",
                },
                sort_keys=True,
            )
        cited = bool(payload["generated_citations"])
        passed = cited and bool(payload["generated_answer"])
        return json.dumps(
            {
                "passed": passed,
                "answer_score": 1.0,
                "citation_support_score": 1.0,
                "failure_category": "pass" if passed else "answer_wrong",
                "failure_detail": "generated answer uses approved evidence id",
            },
            ensure_ascii=False,
            sort_keys=True,
        )

    report = v562.build_report(
        root=ROOT,
        execute=True,
        env={v562.ENABLE_ENV_VAR: "1"},
        answer_client=answer_client,
        scorer_client=scorer_client,
        generated_at="2026-06-04T00:00:00Z",
    )
    v562.check_report(report)
    result = report["official_metric_scored_result"]

    assert calls == {"answer": 30, "scorer": 30}
    assert report["status"] == V5_6_2_SCORED_STATUS
    assert report["backend_preflight"]["status"] == "BACKEND_AVAILABLE_PREFLIGHT_PASSED_NONPROD"
    assert report["backend_preflight"]["failure_category"] == ""
    assert result["status"] == "scored"
    assert result["backend_unavailable"] is False
    assert result["answer_quality_metric_computed"] is True
    assert result["official_metric_finalized"] is True
    assert result["scored_answer_rows"] == 29
    assert result["pass_count"] == 29
    assert result["fail_count"] == 0
    assert result["pass_fail_counts_interpretable_as_quality_metric"] is True
    assert result["failure_category_counts"] == {"pass": 29}
    assert len(result["row_results"]) == 29
    assert all(row["prompt_sha256"] for row in result["row_results"])
    assert all(row["scorer_prompt_sha256"] for row in result["row_results"])
    assert all(row["raw_answer_response_sha256"] for row in result["row_results"])
    assert all(row["raw_scorer_response_sha256"] for row in result["row_results"])
    generated_text = json.dumps(report, ensure_ascii=False)
    for forbidden in ('"raw_prompt_payload":', '"raw_response_payload":', '"raw_llm_response":'):
        assert forbidden not in generated_text


def test_v562_check_report_rejects_scope_expansion_fake_quality_metrics_and_v56_hash_drift() -> None:
    from ai.eval import rag_v562_official_metric_backend_enabled_preflight_scored_rerun_nonprod as v562

    report = v562.build_report(root=ROOT, execute=True, env={}, generated_at="2026-06-04T00:00:00Z")
    v562.check_report(report)

    for path, value, message in (
        (("source_run_id",), V5_4_SHORT_RUN_ID, "source run"),
        (("current_resolves_to",), V5_6_2_SHORT_KEY, "current"),
        (("approval_scope", "row_count"), 30, "approval scope"),
        (("official_metric_input_rows",), 30, "official metric input"),
        (("official_metric_input_rows_consumed",), 30, "official metric input"),
        (("official_metric_input_row_ref_count",), 30, "row ref"),
        (("row_count_by_track", "text_namu_v2_1"), 7, "row count"),
        (("source_artifact_validation", "sha256_matches_v5_6_report"), False, "v5_6"),
        (("official_metric_scored_result", "scored_answer_rows"), 29, "scored rows"),
        (("official_metric_scored_result", "answer_quality_metric_computed"), True, "fake metric"),
        (("official_metric_scored_result", "official_metric_finalized"), True, "fake metric"),
        (
            ("official_metric_scored_result", "pass_fail_counts_interpretable_as_quality_metric"),
            True,
            "quality metric",
        ),
        (("official_metric_scored_result", "failure_category_counts", "execution_gate_disabled"), 28, "failure"),
        (("failure_attribution_rows", 0, "failure_category"), "backend_unavailable", "failure attribution"),
        (("backend_preflight", "noop_backend_used"), True, "noop"),
        (("protected_namespaces_touched",), ["ai/eval/eval_queries"], "protected"),
        (("duplicate_supporting_evidence_policy",), "collapse duplicate evidence ids", "duplicate"),
    ):
        mutated = json.loads(json.dumps(report))
        cursor = mutated
        for key in path[:-1]:
            cursor = cursor[key]
        cursor[path[-1]] = value
        try:
            v562.check_report(mutated)
        except ValueError as exc:
            assert message in str(exc)
        else:
            raise AssertionError(f"v5_6_2 accepted drift at {path}")


def test_v562_write_path_writes_preflight_status_without_measurements_when_unscored(tmp_path: Path) -> None:
    from ai.eval import rag_v562_official_metric_backend_enabled_preflight_scored_rerun_nonprod as v562

    report = v562.build_report(root=ROOT, execute=True, env={}, generated_at="2026-06-04T00:00:00Z")
    written, artifact_hashes = v562.write_report_bundle(tmp_path, report)
    v562.check_report(written, root=tmp_path)
    v562.append_status(tmp_path, written, artifact_hashes=artifact_hashes)

    docs_dir = tmp_path / "docs"
    docs_dir.mkdir(parents=True)
    progress_path = docs_dir / "rag-ingestion-progress.md"
    measurements_path = docs_dir / "rag-ingestion-measurements.md"
    progress_path.write_text("Last updated: 2026-06-04 KST.\n\n## Current Status\n\nOld\n\n## Short History\n\nOld\n", encoding="utf-8")
    measurements_before = "Last updated: 2026-06-04 KST.\n\n## Measurements\n\nsentinel\n"
    measurements_path.write_text(measurements_before, encoding="utf-8")
    v562.update_docs(tmp_path, written)

    paths = written["artifact_paths"]
    assert (tmp_path / paths["report_json"]).exists()
    assert (tmp_path / paths["backend_preflight_result_json"]).exists()
    assert (tmp_path / paths["official_metric_scored_result_json"]).exists()
    assert (tmp_path / paths["failure_attribution_jsonl"]).exists()
    assert _read_json(tmp_path / paths["backend_preflight_result_json"]) == written["backend_preflight"]
    assert _read_json(tmp_path / paths["official_metric_scored_result_json"]) == written["official_metric_scored_result"]
    assert _read_jsonl(tmp_path / paths["failure_attribution_jsonl"]) == written["failure_attribution_rows"]
    assert measurements_path.read_text(encoding="utf-8") == measurements_before

    status_rows = _read_jsonl(tmp_path / "ai/eval/reports/rag-ingestion/status.jsonl")
    latest = status_rows[-1]
    assert latest["short_run_id"] == V5_6_2_SHORT_RUN_ID
    assert latest["status"] == V5_6_2_STATUS
    assert latest["backend_preflight_failure_category"] == "execution_gate_disabled"
    assert latest["answer_quality_metric_computed"] is False
    assert latest["scored_answer_rows"] == 0
    assert latest["failure_category_counts"] == {"execution_gate_disabled": 29}

    progress_text = progress_path.read_text(encoding="utf-8")
    assert V5_6_2_SHORT_RUN_ID in progress_text
    assert "execution_gate_disabled" in progress_text
    assert "v5_6 artifacts remain immutable fail-closed baseline" in progress_text
    assert "pass_count=0/fail_count=29 is not an answer-quality metric" in progress_text

    for rel_path in (
        "ai/eval/reports/rag-ingestion/runs/v5_6_2/report.json",
        "ai/eval/reports/rag-ingestion/runs/v5_6_2/backend_preflight_result.json",
        "ai/eval/reports/rag-ingestion/runs/v5_6_2/official_metric_scored_result.json",
        "ai/eval/reports/rag-ingestion/runs/v5_6_2/failure_attribution.jsonl",
        "ai/eval/reports/rag-ingestion/status.jsonl",
    ):
        assert subprocess.run(["git", "check-ignore", "-q", rel_path], cwd=ROOT).returncode == 0, rel_path


def test_v562_write_path_validates_report_before_writing_and_uses_v550_source(monkeypatch) -> None:
    import ai.scripts.rag_eval as runner
    from ai.eval import rag_v562_official_metric_backend_enabled_preflight_scored_rerun_nonprod as v562

    source_report = {
        "short_run_id": V5_5_SHORT_RUN_ID,
        "canonical_long_run_id": V5_5_LONG_RUN_ID,
        "status": V5_5_STATUS,
        "sentinel_from_check_run": True,
    }
    observed: dict[str, object] = {}
    call_order: list[str] = []

    def fake_check_run(key: str) -> dict[str, object]:
        assert key == "v5_5"
        return source_report

    def fake_build_report(*, root: Path, source_report: dict[str, object] | None = None, **_: object) -> dict[str, object]:
        assert root == ROOT
        assert source_report is not None
        assert source_report["sentinel_from_check_run"] is True
        observed["used_source_report"] = True
        call_order.append("build")
        return {
            "status": V5_6_2_STATUS,
            "artifact_paths": {"report_json": "ai/eval/reports/rag-ingestion/runs/v5_6_2/report.json"},
            "official_metric_input_rows": 29,
            "official_metric_scored_result": {"scored_answer_rows": 0},
            "counters": {},
        }

    def fake_check_report(report: dict[str, object], **_: object) -> None:
        call_order.append("check")

    def fake_write_report_bundle(root: Path, report: dict[str, object]) -> tuple[dict[str, object], dict[str, str]]:
        assert "check" in call_order
        call_order.append("write")
        return report, {"report_json_sha256": "0" * 64}

    monkeypatch.setattr(runner, "check_run", fake_check_run)
    monkeypatch.setattr(v562, "build_report", fake_build_report)
    monkeypatch.setattr(v562, "write_report_bundle", fake_write_report_bundle)
    monkeypatch.setattr(v562, "check_report", fake_check_report)
    monkeypatch.setattr(v562, "update_docs", lambda root, report: None)
    monkeypatch.setattr(v562, "append_status", lambda root, report, *, artifact_hashes: None)

    assert runner.main(["v5_6_2", "--write"]) == 0
    assert observed["used_source_report"] is True
    assert call_order == ["build", "check", "write", "check"]


def test_v563_env_gate_disabled_records_execution_gate_disabled_without_fake_metrics() -> None:
    from ai.eval import rag_v563_official_metric_backend_probe_and_scored_execution_nonprod as v563

    report = v563.build_report(root=ROOT, execute=True, env={}, generated_at="2026-06-05T00:00:00Z")
    v563.check_report(report)
    result = report["official_metric_scored_result"]

    assert report["logical_run_key"] == V5_6_3_SHORT_KEY
    assert report["short_run_id"] == V5_6_3_SHORT_RUN_ID
    assert report["canonical_long_run_id"] == V5_6_3_LONG_RUN_ID
    assert report["status"] == V5_6_3_STATUS
    assert report["source_run_id"] == V5_5_SHORT_RUN_ID
    assert report["current_resolves_to"] == V5_6_SHORT_KEY
    assert report["v5_6_baseline_run_id"] == V5_6_SHORT_RUN_ID
    assert report["v5_6_2_preflight_run_id"] == V5_6_2_SHORT_RUN_ID
    assert report["approval_scope"] == {
        "source_run_id": V5_5_SHORT_RUN_ID,
        "source_artifact_path": "ai/eval/reports/rag-ingestion/runs/v5_5/official_metric_input.jsonl",
        "row_count": 29,
        "scope_policy": "exact_v5_5_official_metric_input_rows_only",
        "excluded_scopes": [
            "silver_rows",
            "v5_2_or_v5_3_residual_rows",
            "overlay_90_rows",
            "xlsx_candidate_state_buckets",
            "pdf_text_residual_taxonomy_denominators",
        ],
    }
    assert report["official_metric_input_rows"] == 29
    assert report["official_metric_input_rows_consumed"] == 29
    assert report["row_count_by_track"] == {
        "text_namu_v2_1": 6,
        "xlsx_business_structured": 19,
        "pdf_business_ocr_mm": 4,
    }
    validation = report["source_artifact_validation"]
    assert validation["source_official_metric_input_path"] == (
        "ai/eval/reports/rag-ingestion/runs/v5_5/official_metric_input.jsonl"
    )
    assert validation["path_matches_v5_5_report"] is True
    assert validation["sha256_matches_v5_5_report"] is True
    assert validation["sha256_matches_v5_6_report"] is True
    assert validation["sha256_matches_v5_6_2_report"] is True
    assert validation["row_count_matches_v5_5_report"] is True
    assert validation["row_count_matches_v5_6_report"] is True
    assert validation["row_count_matches_v5_6_2_report"] is True
    assert report["official_metric_input_row_ref_count"] == 29
    assert len(report["official_metric_input_row_refs"]) == 29
    assert "official_metric_input_rows_payload" not in report

    assert report["backend_preflight"]["status"] == "EXECUTION_GATE_DISABLED_FAIL_CLOSED"
    assert report["backend_preflight"]["failure_category"] == "execution_gate_disabled"
    assert report["backend_preflight"]["available"] is False
    assert report["backend_preflight"]["env_enabled"] is False
    assert report["backend_preflight"]["enabled_env_var"] == "RAG_V5_6_3_ENABLE_OFFICIAL_METRIC_EXECUTION"
    assert result["status"] == "fail_closed"
    assert result["backend_unavailable"] is True
    assert result["official_metric_input_rows"] == 29
    assert result["scored_answer_rows"] == 0
    assert result["answer_quality_metric_computed"] is False
    assert result["official_metric_finalized"] is False
    assert result["pass_count"] == 0
    assert result["fail_count"] == 29
    assert result["pass_fail_counts_interpretable_as_quality_metric"] is False
    assert result["quality_metric_blocked_reason"] == "execution_gate_disabled"
    assert result["failure_category_counts"] == {"execution_gate_disabled": 29}
    assert {row["failure_category"] for row in report["failure_attribution_rows"]} == {"execution_gate_disabled"}
    assert all(row["scoring_attempted"] is False for row in report["failure_attribution_rows"])
    assert report["duplicate_supporting_evidence_policy"] == (
        "recorded_for_locator_precision_audit; row-level citation_locator remains authoritative"
    )
    assert report["duplicate_supporting_evidence_id_count"] == 1
    assert report["protected_namespaces_touched"] == []
    assert set(v563.PREFLIGHT_FAILURE_CATEGORIES) == {
        "execution_gate_disabled",
        "answer_generation_backend_unreachable",
        "answer_generation_model_unavailable",
        "answer_generation_probe_failed",
        "scorer_backend_unreachable",
        "scorer_contract_unavailable",
        "scorer_contract_probe_failed",
        "scoring_runtime_failed",
        "unknown_fail_closed",
    }
    for key in v563.CLOSED_FALSE_KEYS:
        assert report[key] is False, key
    generated_text = json.dumps(report, ensure_ascii=False)
    for forbidden in (
        '"raw_prompt_payload":',
        '"raw_response_payload":',
        '"raw_llm_response":',
        '"official_metric_input_rows_payload":',
        '"expected_answer_ko":',
        '"supporting_evidence_note":',
    ):
        assert forbidden not in generated_text


def test_v563_env_enabled_preflight_failure_categories_are_precise(monkeypatch) -> None:
    from ai.eval import rag_v563_official_metric_backend_probe_and_scored_execution_nonprod as v563

    unreachable = v563.build_report(
        root=ROOT,
        execute=True,
        env={
            v563.ENABLE_ENV_VAR: "1",
            v563.BASE_URL_ENV_VAR: "http://127.0.0.1:9",
        },
        backend_timeout_seconds=1,
        generated_at="2026-06-05T00:00:00Z",
    )
    v563.check_report(unreachable)
    assert unreachable["backend_preflight"]["env_enabled"] is True
    assert unreachable["backend_preflight"]["failure_category"] == "answer_generation_backend_unreachable"
    assert unreachable["official_metric_scored_result"]["failure_category_counts"] == {
        "answer_generation_backend_unreachable": 29
    }
    assert "execution_gate_disabled" not in unreachable["official_metric_scored_result"]["failure_category_counts"]

    def fake_answer_catalog(url: str, *, payload: dict[str, object] | None, timeout_seconds: int) -> object:
        assert payload is None
        assert url.endswith("/models")
        return {"data": [{"id": "different-model"}]}

    monkeypatch.setattr(v563.local_llm, "request_json", fake_answer_catalog)
    model_unavailable = v563.build_report(
        root=ROOT,
        execute=True,
        env={
            v563.ENABLE_ENV_VAR: "1",
            v563.MODEL_ENV_VAR: "missing-model",
        },
        backend_timeout_seconds=1,
        generated_at="2026-06-05T00:00:00Z",
    )
    v563.check_report(model_unavailable)
    assert model_unavailable["backend_preflight"]["failure_category"] == "answer_generation_model_unavailable"
    assert model_unavailable["official_metric_scored_result"]["failure_category_counts"] == {
        "answer_generation_model_unavailable": 29
    }
    monkeypatch.undo()

    calls = {"answer": 0, "scorer": 0}

    def invalid_answer_probe(prompt: str) -> str:
        calls["answer"] += 1
        payload = json.loads(prompt)
        assert payload["task"] == "v5_6_3_non_gold_answer_generation_probe"
        return json.dumps({"unexpected": True}, sort_keys=True)

    answer_probe_failed = v563.build_report(
        root=ROOT,
        execute=True,
        env={v563.ENABLE_ENV_VAR: "1"},
        answer_client=invalid_answer_probe,
        generated_at="2026-06-05T00:00:00Z",
    )
    v563.check_report(answer_probe_failed)
    assert calls == {"answer": 1, "scorer": 0}
    assert answer_probe_failed["backend_preflight"]["failure_category"] == "answer_generation_probe_failed"
    assert answer_probe_failed["official_metric_scored_result"]["failure_category_counts"] == {
        "answer_generation_probe_failed": 29
    }

    def valid_answer_probe(prompt: str) -> str:
        calls["answer"] += 1
        payload = json.loads(prompt)
        assert payload["task"] == "v5_6_3_non_gold_answer_generation_probe"
        return json.dumps({"final_answer": "probe answer", "citations": [], "abstain": False}, sort_keys=True)

    scorer_unreachable = v563.build_report(
        root=ROOT,
        execute=True,
        env={
            v563.ENABLE_ENV_VAR: "1",
            v563.SCORER_BASE_URL_ENV_VAR: "http://127.0.0.1:9",
        },
        answer_client=valid_answer_probe,
        backend_timeout_seconds=1,
        generated_at="2026-06-05T00:00:00Z",
    )
    v563.check_report(scorer_unreachable)
    assert calls == {"answer": 2, "scorer": 0}
    assert scorer_unreachable["backend_preflight"]["failure_category"] == "scorer_backend_unreachable"
    assert scorer_unreachable["official_metric_scored_result"]["failure_category_counts"] == {
        "scorer_backend_unreachable": 29
    }

    def scorer_contract_unavailable_catalog(
        url: str,
        *,
        payload: dict[str, object] | None,
        timeout_seconds: int,
    ) -> object:
        assert payload is None
        assert url.endswith("/models")
        return {"data": [{"id": "answer-model"}]}

    monkeypatch.setattr(v563.local_llm, "request_json", scorer_contract_unavailable_catalog)
    scorer_contract_unavailable = v563.build_report(
        root=ROOT,
        execute=True,
        env={
            v563.ENABLE_ENV_VAR: "1",
            v563.MODEL_ENV_VAR: "answer-model",
            v563.SCORER_MODEL_ENV_VAR: "missing-scorer-contract",
        },
        answer_client=valid_answer_probe,
        backend_timeout_seconds=1,
        generated_at="2026-06-05T00:00:00Z",
    )
    v563.check_report(scorer_contract_unavailable)
    assert scorer_contract_unavailable["backend_preflight"]["failure_category"] == "scorer_contract_unavailable"
    assert scorer_contract_unavailable["official_metric_scored_result"]["failure_category_counts"] == {
        "scorer_contract_unavailable": 29
    }

    def invalid_scorer_probe(prompt: str) -> str:
        calls["scorer"] += 1
        payload = json.loads(prompt)
        assert payload["task"] == "v5_6_3_non_gold_scorer_contract_probe"
        return json.dumps({"unexpected": True}, sort_keys=True)

    scorer_probe_failed = v563.build_report(
        root=ROOT,
        execute=True,
        env={v563.ENABLE_ENV_VAR: "1"},
        answer_client=valid_answer_probe,
        scorer_client=invalid_scorer_probe,
        generated_at="2026-06-05T00:00:00Z",
    )
    v563.check_report(scorer_probe_failed)
    assert calls["scorer"] == 1
    assert scorer_probe_failed["backend_preflight"]["failure_category"] == "scorer_contract_probe_failed"
    assert scorer_probe_failed["official_metric_scored_result"]["failure_category_counts"] == {
        "scorer_contract_probe_failed": 29
    }


def test_v563_injected_answer_and_scorer_backends_score_all_29_rows_after_non_gold_probes() -> None:
    from ai.eval import rag_v563_official_metric_backend_probe_and_scored_execution_nonprod as v563

    calls = {"answer": 0, "scorer": 0}

    def answer_client(prompt: str) -> str:
        calls["answer"] += 1
        payload = json.loads(prompt)
        if payload["task"] == "v5_6_3_non_gold_answer_generation_probe":
            return json.dumps({"final_answer": "probe answer", "citations": [], "abstain": False}, sort_keys=True)
        prompt_text = json.dumps(payload, ensure_ascii=False)
        for forbidden in (
            "query_id",
            "expected_answer_ko",
            "supporting_evidence_note",
            "supporting_evidence_ids",
            "citation_locator",
            "target_locator",
            "gold_locator",
        ):
            assert forbidden not in prompt_text
        context = payload["official_metric_row"]
        return json.dumps(
            {
                "final_answer": f"synthetic answer for {context['track']}",
                "citations": ["runtime-citation"],
                "abstain": False,
            },
            ensure_ascii=False,
            sort_keys=True,
        )

    def scorer_client(prompt: str) -> str:
        calls["scorer"] += 1
        payload = json.loads(prompt)
        if payload["task"] == "v5_6_3_non_gold_scorer_contract_probe":
            return json.dumps(
                {
                    "passed": True,
                    "answer_score": 1.0,
                    "citation_support_score": 1.0,
                    "failure_category": "pass",
                    "failure_detail": "probe contract ok",
                },
                sort_keys=True,
            )
        assert payload["expected_answer_ko"]
        assert payload["supporting_evidence_note"]
        cited = bool(payload["generated_citations"])
        passed = cited and bool(payload["generated_answer"])
        return json.dumps(
            {
                "passed": passed,
                "answer_score": 1.0,
                "citation_support_score": 1.0,
                "failure_category": "pass" if passed else "unknown_fail_closed",
                "failure_detail": "scorer used approved scoring-only expected evidence",
            },
            ensure_ascii=False,
            sort_keys=True,
        )

    report = v563.build_report(
        root=ROOT,
        execute=True,
        env={v563.ENABLE_ENV_VAR: "1"},
        answer_client=answer_client,
        scorer_client=scorer_client,
        generated_at="2026-06-05T00:00:00Z",
    )
    v563.check_report(report)
    result = report["official_metric_scored_result"]

    assert calls == {"answer": 30, "scorer": 30}
    assert report["status"] == V5_6_3_SCORED_STATUS
    assert report["backend_preflight"]["status"] == "BACKEND_PROBE_PASSED_NONPROD"
    assert report["backend_preflight"]["failure_category"] == ""
    assert result["status"] == "scored"
    assert result["backend_unavailable"] is False
    assert result["answer_quality_metric_computed"] is True
    assert result["official_metric_finalized"] is True
    assert result["scored_answer_rows"] == 29
    assert result["pass_count"] == 29
    assert result["fail_count"] == 0
    assert result["pass_fail_counts_interpretable_as_quality_metric"] is True
    assert result["failure_category_counts"] == {"pass": 29}
    assert len(result["row_results"]) == 29
    assert all(row["prompt_sha256"] for row in result["row_results"])
    assert all(row["scorer_prompt_sha256"] for row in result["row_results"])
    assert all(row["raw_answer_response_sha256"] for row in result["row_results"])
    assert all(row["raw_scorer_response_sha256"] for row in result["row_results"])
    generated_text = json.dumps(report, ensure_ascii=False)
    for forbidden in ('"raw_prompt_payload":', '"raw_response_payload":', '"raw_llm_response":'):
        assert forbidden not in generated_text


def test_v563_scoring_runtime_failure_fails_closed_without_partial_quality_metric() -> None:
    from ai.eval import rag_v563_official_metric_backend_probe_and_scored_execution_nonprod as v563

    calls = {"answer": 0, "scorer": 0}

    def answer_client(prompt: str) -> str:
        calls["answer"] += 1
        return json.dumps({"final_answer": "probe or row answer", "citations": [], "abstain": False}, sort_keys=True)

    def scorer_client(prompt: str) -> str:
        calls["scorer"] += 1
        payload = json.loads(prompt)
        if payload["task"] == "v5_6_3_non_gold_scorer_contract_probe":
            return json.dumps(
                {
                    "passed": True,
                    "answer_score": 1.0,
                    "citation_support_score": 1.0,
                    "failure_category": "pass",
                },
                sort_keys=True,
            )
        raise RuntimeError("runtime scorer outage")

    report = v563.build_report(
        root=ROOT,
        execute=True,
        env={v563.ENABLE_ENV_VAR: "1"},
        answer_client=answer_client,
        scorer_client=scorer_client,
        generated_at="2026-06-05T00:00:00Z",
    )
    v563.check_report(report)
    result = report["official_metric_scored_result"]

    assert report["backend_preflight"]["status"] == "BACKEND_PROBE_PASSED_NONPROD"
    assert result["status"] == "fail_closed"
    assert result["failure_category_counts"] == {"scoring_runtime_failed": 29}
    assert result["scored_answer_rows"] == 0
    assert result["answer_quality_metric_computed"] is False
    assert result["official_metric_finalized"] is False
    assert result["pass_fail_counts_interpretable_as_quality_metric"] is False
    assert result["quality_metric_blocked_reason"] == "scoring_runtime_failed"
    assert {row["failure_category"] for row in report["failure_attribution_rows"]} == {"scoring_runtime_failed"}
    assert calls == {"answer": 2, "scorer": 2}


def test_v563_check_report_rejects_scope_expansion_fake_quality_metrics_and_prior_hash_drift() -> None:
    from ai.eval import rag_v563_official_metric_backend_probe_and_scored_execution_nonprod as v563

    report = v563.build_report(root=ROOT, execute=True, env={}, generated_at="2026-06-05T00:00:00Z")
    v563.check_report(report)

    for path, value, message in (
        (("source_run_id",), V5_4_SHORT_RUN_ID, "source run"),
        (("current_resolves_to",), V5_6_3_SHORT_KEY, "current"),
        (("approval_scope", "row_count"), 30, "approval scope"),
        (("approval_scope", "excluded_scopes"), ["silver_rows"], "approval scope"),
        (("official_metric_input_rows",), 30, "official metric input"),
        (("official_metric_input_rows_consumed",), 30, "official metric input"),
        (("official_metric_input_row_ref_count",), 30, "row ref"),
        (("row_count_by_track", "text_namu_v2_1"), 7, "row count"),
        (("source_artifact_validation", "sha256_matches_v5_6_report"), False, "v5_6"),
        (("source_artifact_validation", "sha256_matches_v5_6_2_report"), False, "v5_6_2"),
        (("official_metric_scored_result", "scored_answer_rows"), 29, "scored rows"),
        (("official_metric_scored_result", "answer_quality_metric_computed"), True, "fake metric"),
        (("official_metric_scored_result", "official_metric_finalized"), True, "fake metric"),
        (
            ("official_metric_scored_result", "pass_fail_counts_interpretable_as_quality_metric"),
            True,
            "quality metric",
        ),
        (("official_metric_scored_result", "failure_category_counts", "execution_gate_disabled"), 28, "failure"),
        (("failure_attribution_rows", 0, "failure_category"), "unknown_fail_closed", "failure attribution"),
        (("backend_preflight", "noop_backend_used"), True, "noop"),
        (("protected_namespaces_touched",), ["ai/eval/eval_queries"], "protected"),
        (("duplicate_supporting_evidence_policy",), "collapse duplicate evidence ids", "duplicate"),
    ):
        mutated = json.loads(json.dumps(report))
        cursor = mutated
        for key in path[:-1]:
            cursor = cursor[key]
        cursor[path[-1]] = value
        try:
            v563.check_report(mutated)
        except ValueError as exc:
            assert message in str(exc)
        else:
            raise AssertionError(f"v5_6_3 accepted drift at {path}")


def test_v563_write_path_writes_preflight_status_without_measurements_when_unscored(tmp_path: Path) -> None:
    from ai.eval import rag_v563_official_metric_backend_probe_and_scored_execution_nonprod as v563

    report = v563.build_report(root=ROOT, execute=True, env={}, generated_at="2026-06-05T00:00:00Z")
    written, artifact_hashes = v563.write_report_bundle(tmp_path, report)
    v563.check_report(written, root=tmp_path)
    v563.append_status(tmp_path, written, artifact_hashes=artifact_hashes)

    docs_dir = tmp_path / "docs"
    docs_dir.mkdir(parents=True)
    progress_path = docs_dir / "rag-ingestion-progress.md"
    measurements_path = docs_dir / "rag-ingestion-measurements.md"
    progress_path.write_text("Last updated: 2026-06-04 KST.\n\n## Current Status\n\nOld\n\n## Short History\n\nOld\n", encoding="utf-8")
    measurements_before = "Last updated: 2026-06-04 KST.\n\n## Measurements\n\nsentinel\n"
    measurements_path.write_text(measurements_before, encoding="utf-8")
    v563.update_docs(tmp_path, written)

    paths = written["artifact_paths"]
    assert (tmp_path / paths["report_json"]).exists()
    assert (tmp_path / paths["backend_preflight_result_json"]).exists()
    assert (tmp_path / paths["official_metric_scored_result_json"]).exists()
    assert (tmp_path / paths["failure_attribution_jsonl"]).exists()
    assert _read_json(tmp_path / paths["backend_preflight_result_json"]) == written["backend_preflight"]
    assert _read_json(tmp_path / paths["official_metric_scored_result_json"]) == written["official_metric_scored_result"]
    assert _read_jsonl(tmp_path / paths["failure_attribution_jsonl"]) == written["failure_attribution_rows"]
    assert measurements_path.read_text(encoding="utf-8") == measurements_before

    status_rows = _read_jsonl(tmp_path / "ai/eval/reports/rag-ingestion/status.jsonl")
    latest = status_rows[-1]
    assert latest["short_run_id"] == V5_6_3_SHORT_RUN_ID
    assert latest["status"] == V5_6_3_STATUS
    assert latest["backend_preflight_failure_category"] == "execution_gate_disabled"
    assert latest["answer_quality_metric_computed"] is False
    assert latest["scored_answer_rows"] == 0
    assert latest["failure_category_counts"] == {"execution_gate_disabled": 29}

    progress_text = progress_path.read_text(encoding="utf-8")
    assert V5_6_3_SHORT_RUN_ID in progress_text
    assert "execution_gate_disabled" in progress_text
    assert "v5_6 and v5_6_2 artifacts remain immutable" in progress_text
    assert "pass_count=0/fail_count=29 is not an answer-quality metric" in progress_text
    assert "RAG_V5_6_3_ENABLE_OFFICIAL_METRIC_EXECUTION" in progress_text

    for rel_path in (
        "ai/eval/reports/rag-ingestion/runs/v5_6_3/report.json",
        "ai/eval/reports/rag-ingestion/runs/v5_6_3/backend_preflight_result.json",
        "ai/eval/reports/rag-ingestion/runs/v5_6_3/official_metric_scored_result.json",
        "ai/eval/reports/rag-ingestion/runs/v5_6_3/failure_attribution.jsonl",
        "ai/eval/reports/rag-ingestion/status.jsonl",
    ):
        assert subprocess.run(["git", "check-ignore", "-q", rel_path], cwd=ROOT).returncode == 0, rel_path


def test_v563_write_path_validates_report_before_writing_and_uses_v550_source(monkeypatch) -> None:
    import ai.scripts.rag_eval as runner
    from ai.eval import rag_v563_official_metric_backend_probe_and_scored_execution_nonprod as v563

    source_report = {
        "short_run_id": V5_5_SHORT_RUN_ID,
        "canonical_long_run_id": V5_5_LONG_RUN_ID,
        "status": V5_5_STATUS,
        "sentinel_from_check_run": True,
    }
    observed: dict[str, object] = {}
    call_order: list[str] = []

    def fake_check_run(key: str) -> dict[str, object]:
        assert key == "v5_5"
        return source_report

    def fake_build_report(*, root: Path, source_report: dict[str, object] | None = None, **_: object) -> dict[str, object]:
        assert root == ROOT
        assert source_report is not None
        assert source_report["sentinel_from_check_run"] is True
        observed["used_source_report"] = True
        call_order.append("build")
        return {
            "status": V5_6_3_STATUS,
            "artifact_paths": {"report_json": "ai/eval/reports/rag-ingestion/runs/v5_6_3/report.json"},
            "official_metric_input_rows": 29,
            "official_metric_scored_result": {"scored_answer_rows": 0},
            "counters": {},
        }

    def fake_check_report(report: dict[str, object], **_: object) -> None:
        call_order.append("check")

    def fake_write_report_bundle(root: Path, report: dict[str, object]) -> tuple[dict[str, object], dict[str, str]]:
        assert "check" in call_order
        call_order.append("write")
        return report, {"report_json_sha256": "0" * 64}

    monkeypatch.setattr(runner, "check_run", fake_check_run)
    monkeypatch.setattr(v563, "build_report", fake_build_report)
    monkeypatch.setattr(v563, "write_report_bundle", fake_write_report_bundle)
    monkeypatch.setattr(v563, "check_report", fake_check_report)
    monkeypatch.setattr(v563, "update_docs", lambda root, report: None)
    monkeypatch.setattr(v563, "append_status", lambda root, report, *, artifact_hashes: None)

    assert runner.main(["v5_6_3", "--write"]) == 0
    assert observed["used_source_report"] is True
    assert call_order == ["build", "check", "write", "check"]


def test_v56_refactor_comparison_packet_separates_route_replay_from_quality_subsets() -> None:
    from ai.eval import rag_v56_refactor_route_comparison_packet_diagnostic_nonprod as comparison

    report = comparison.build_report(root=ROOT, generated_at="2026-06-05T00:00:00Z")
    comparison.check_report(report)

    assert report["short_run_id"] == V5_6_REFACTOR_COMPARISON_SHORT_RUN_ID
    assert report["canonical_long_run_id"] == V5_6_REFACTOR_COMPARISON_LONG_RUN_ID
    assert report["status"] == V5_6_REFACTOR_COMPARISON_STATUS
    assert report["current_resolves_to"] == V5_6_SHORT_KEY
    assert report["diagnostic_only"] is True
    assert report["official_metric"] is False
    assert report["official_metric_input_rows"] == 0
    assert report["official_metric_input_rows_consumed"] == 0
    assert report["quality_delta_claim_supported"] is False
    assert report["retrieval_quality_delta_computed"] is False
    assert report["answer_quality_delta_computed"] is False
    assert report["protected_namespaces_touched"] == []
    assert report["comparator_scope"]["fixture_row_count"] >= 6
    assert set(report["comparator_scope"]["families"]) == {
        "TEXT",
        "XLSX",
        "PDF",
    }
    assert report["comparator_scope"]["scope_counts"]["metadata_scoped"] >= 3
    assert report["comparator_scope"]["scope_counts"]["query_only"] >= 3

    route_rows = report["route_comparison_rows"]
    assert len(route_rows) == report["comparator_scope"]["fixture_row_count"]
    required_route_fields = {
        "query_id",
        "family",
        "scope_type",
        "before_route",
        "after_route",
        "route_changed",
        "after_blocked_reason",
        "llm_required",
        "metadata_scoped",
    }
    assert all(required_route_fields <= set(row) for row in route_rows)
    assert any(row["metadata_scoped"] is True and row["route_changed"] is True for row in route_rows)
    assert any(row["metadata_scoped"] is False and row["llm_required"] is True for row in route_rows)
    assert report["route_change_summary"]["route_changed_count"] == sum(
        1 for row in route_rows if row["route_changed"]
    )
    assert report["route_change_summary"]["by_scope_type"]["metadata_scoped"]["row_count"] >= 3
    assert report["route_change_summary"]["by_scope_type"]["query_only"]["row_count"] >= 3

    retrieval_rows = report["retrieval_quality_rows"]
    computed_retrieval = [row for row in retrieval_rows if row["retrieval_quality_delta_computed"]]
    blocked_retrieval = [row for row in retrieval_rows if not row["retrieval_quality_delta_computed"]]
    assert computed_retrieval
    assert blocked_retrieval
    assert all(row["existing_qrels_or_locator_valid"] is True for row in computed_retrieval)
    assert all(row["retrieval_blocked_reason"] for row in blocked_retrieval)
    assert all(
        key in computed_retrieval[0]["before_retrieval"]
        for key in ("hit_at_1", "hit_at_k", "target_in_candidates", "candidate_count", "evidence_sufficiency_status")
    )
    assert report["retrieval_quality_subset"]["computed_row_count"] == len(computed_retrieval)
    assert report["retrieval_quality_subset"]["global_delta_claim_supported"] is False

    answer_rows = report["answer_quality_rows"]
    computed_answer = [row for row in answer_rows if row["answer_quality_delta_computed"]]
    unavailable_answer = [row for row in answer_rows if row["answer_execution_status"] == "execution_unavailable"]
    assert computed_answer
    assert unavailable_answer
    assert all(
        row["before_answer"]["execution_status"] == "executed"
        and row["after_answer"]["execution_status"] == "executed"
        for row in computed_answer
    )
    assert all(row["answer_quality_blocked_reason"] for row in unavailable_answer)
    assert report["answer_quality_subset"]["computed_row_count"] == len(computed_answer)
    assert report["answer_quality_subset"]["execution_unavailable_row_count"] == len(unavailable_answer)
    assert report["answer_quality_subset"]["global_delta_claim_supported"] is False


def test_v56_refactor_comparison_check_report_rejects_metric_gate_and_subset_drift() -> None:
    from ai.eval import rag_v56_refactor_route_comparison_packet_diagnostic_nonprod as comparison

    report = comparison.build_report(root=ROOT, generated_at="2026-06-05T00:00:00Z")
    comparison.check_report(report)

    for path, value, message in (
        (("current_resolves_to",), V5_6_3_SHORT_KEY, "current"),
        (("diagnostic_only",), False, "diagnostic"),
        (("official_metric",), True, "official"),
        (("official_metric_input_rows",), 29, "official metric input"),
        (("official_metric_input_rows_consumed",), 29, "official metric input"),
        (("quality_delta_claim_supported",), True, "quality delta"),
        (("retrieval_quality_delta_computed",), True, "retrieval quality"),
        (("answer_quality_delta_computed",), True, "answer quality"),
        (("protected_namespaces_touched",), ["ai/eval/eval_queries"], "protected"),
        (("route_comparison_rows", 0, "before_route"), "", "route row"),
        (("retrieval_quality_rows", 0, "existing_qrels_or_locator_valid"), False, "retrieval subset"),
        (("answer_quality_rows", 0, "after_answer", "execution_status"), "execution_unavailable", "answer subset"),
    ):
        mutated = json.loads(json.dumps(report))
        cursor = mutated
        for key in path[:-1]:
            cursor = cursor[key]
        cursor[path[-1]] = value
        try:
            comparison.check_report(mutated)
        except ValueError as exc:
            assert message in str(exc)
        else:
            raise AssertionError(f"v5_6 refactor comparison accepted drift at {path}")

    for key in (
        "gold_mutation",
        "qrels_mutation",
        "label_mutation",
        "expected_answer_mutation",
        "supporting_evidence_mutation",
        "denominator_mutation",
        "training_dataset_created",
        "fine_tuning",
        "ft_a_execution",
        "promotion_evidence",
        "product_success_evidence_allowed",
        "live_db_index_cache_readiness",
        "production_db_mutated",
        "source_registry_mutated",
        "silver_mutation",
        "index_rebuilt",
        "cache_mutated",
    ):
        mutated = json.loads(json.dumps(report))
        mutated[key] = True
        try:
            comparison.check_report(mutated)
        except ValueError as exc:
            assert key in str(exc) or "closed" in str(exc)
        else:
            raise AssertionError(f"v5_6 refactor comparison accepted {key}=True")


def test_v56_refactor_comparison_write_status_and_runner_keep_current_v56(tmp_path: Path) -> None:
    import ai.scripts.rag_eval as runner
    from ai.eval import rag_v56_refactor_route_comparison_packet_diagnostic_nonprod as comparison

    report = comparison.build_report(root=ROOT, generated_at="2026-06-05T00:00:00Z")
    written, artifact_hashes = comparison.write_report_bundle(tmp_path, report)
    comparison.check_report(written, root=tmp_path)
    comparison.append_status(tmp_path, written, artifact_hashes=artifact_hashes)

    paths = written["artifact_paths"]
    assert paths == {
        "report_json": "ai/eval/reports/rag-ingestion/runs/v5_6_refactor_comparison/report.json",
        "status_jsonl": "ai/eval/reports/rag-ingestion/status.jsonl",
    }
    assert (tmp_path / paths["report_json"]).exists()
    assert artifact_hashes["report_json_sha256"] == _sha256_file(tmp_path / paths["report_json"])
    status_rows = _read_jsonl(tmp_path / "ai/eval/reports/rag-ingestion/status.jsonl")
    latest = status_rows[-1]
    assert latest["short_run_id"] == V5_6_REFACTOR_COMPARISON_SHORT_RUN_ID
    assert latest["current_resolves_to"] == V5_6_SHORT_KEY
    assert latest["diagnostic_only"] is True
    assert latest["official_metric_input_rows"] == 0
    assert latest["quality_delta_claim_supported"] is False
    assert latest["retrieval_quality_delta_computed"] is False
    assert latest["answer_quality_delta_computed"] is False

    checked = runner.check_run(V5_6_REFACTOR_COMPARISON_SHORT_KEY)
    assert checked["short_run_id"] == V5_6_REFACTOR_COMPARISON_SHORT_RUN_ID
    assert runner.check_run("current")["short_run_id"] == V5_6_SHORT_RUN_ID

    assert subprocess.run(["git", "check-ignore", "-q", paths["report_json"]], cwd=ROOT).returncode == 0
    assert subprocess.run(["git", "check-ignore", "-q", paths["status_jsonl"]], cwd=ROOT).returncode == 0


def test_v56_refactor_comparison_write_path_validates_report_before_writing(monkeypatch) -> None:
    import ai.scripts.rag_eval as runner
    from ai.eval import rag_v56_refactor_route_comparison_packet_diagnostic_nonprod as comparison

    call_order: list[str] = []

    def fake_build_report(*, root: Path, generated_at: str | None = None, check: bool = True) -> dict[str, object]:
        del generated_at, check
        call_order.append("build")
        return {
            "short_run_id": V5_6_REFACTOR_COMPARISON_SHORT_RUN_ID,
            "status": V5_6_REFACTOR_COMPARISON_STATUS,
            "current_resolves_to": V5_6_SHORT_KEY,
            "artifact_paths": {"report_json": V5_6_REFACTOR_COMPARISON_REPORT.relative_to(ROOT).as_posix()},
            "route_change_summary": {},
            "retrieval_quality_subset": {},
            "answer_quality_subset": {},
        }

    def fake_check_report(report: dict[str, object], **_: object) -> None:
        del report
        call_order.append("check")

    def fake_write_report_bundle(root: Path, report: dict[str, object]) -> tuple[dict[str, object], dict[str, str]]:
        assert "check" in call_order
        call_order.append("write")
        return report, {"report_json_sha256": "0" * 64}

    monkeypatch.setattr(comparison, "build_report", fake_build_report)
    monkeypatch.setattr(comparison, "check_report", fake_check_report)
    monkeypatch.setattr(comparison, "write_report_bundle", fake_write_report_bundle)
    monkeypatch.setattr(comparison, "append_status", lambda root, report, *, artifact_hashes: None)
    monkeypatch.setattr(comparison, "update_docs", lambda root, report: None)

    assert runner.main([V5_6_REFACTOR_COMPARISON_SHORT_KEY, "--write"]) == 0
    assert call_order == ["build", "check", "write", "check"]


def test_v4711_injected_local_llm_replays_v4710_candidates_and_records_answer_audits() -> None:
    from ai.eval import rag_v4711_actual_llm_answer_replay_and_silver_diagnostic_smoke as v4711

    prompts: list[str] = []

    def fake_client(prompt: str) -> str:
        prompts.append(prompt)
        return _fake_strict_korean_answer(prompt)

    report = v4711.build_report(
        root=ROOT,
        execute=True,
        sync_surfaces=False,
        env={"RAG_V4_7_11_ENABLE_LOCAL_LLM_REPLAY": "1"},
        llm_client=fake_client,
        generated_at="2026-05-30T00:00:00Z",
    )
    v4711.check_report(report)
    counters = report["counters"]

    assert report["short_run_id"] == V4_7_11_SHORT_RUN_ID
    assert report["source_run_id"] == V4_7_10_SHORT_RUN_ID
    assert report["diagnostic_only"] is True
    assert report["non_production"] is True
    assert report["official_metric"] is False
    assert report["official_metric_input_rows"] == 0
    assert report["promotion_evidence"] is False
    assert counters["v4_7_10_answer_ready_evidence_bundle_count"] == 57
    assert counters["v4_7_10_answer_replay_candidate_count"] == 9
    assert counters["v4_7_10_replayed_candidate_count"] == 9
    assert counters["v4_7_10_skipped_weak_residual_count"] == 1
    assert counters["local_llm_replay_env_enabled"] is True
    assert counters["local_llm_available"] is True
    assert counters["llm_invoked_count"] == 9
    assert counters["generated_response_count"] == 9
    assert counters["raw_llm_response_present_count"] == 0
    assert counters["parsed_final_answer_present_count"] == 9
    assert counters["citation_rendered_count"] == 9
    assert counters["citation_grounded_to_evidence_count"] == 9
    assert counters["korean_final_answer_count"] == 9
    assert counters["non_korean_answer_flag_count"] == 0
    assert counters["claim_support_verifier_pass_count"] + counters["claim_support_verifier_fail_count"] == 9
    assert counters["prompt_leakage_flag_count"] == 0
    assert counters["response_leakage_flag_count"] == 0
    assert counters["path_leakage_flag_count"] == 0
    assert counters["evidence_truth_violation_count"] == 0
    assert counters["vector_payload_evidence_truth_violation_count"] == 0
    assert len(prompts) == 9
    assert all(
        not re.search(r"\b[A-Z]:[\\/]|gold|qrels|expected|supporting|source_file_title", prompt, re.I)
        for prompt in prompts
    )

    rows = report["pdf_answer_replay_rows"]
    assert len(rows) == 9
    assert all(row["llm_invoked"] is True for row in rows)
    assert all(row["answer_replay_audit"]["status"] == "LOCAL_LLM_GENERATED_DIAGNOSTIC_ONLY" for row in rows)
    assert all(row["answer_replay_audit"]["parsed_final_answer_present"] is True for row in rows)
    assert all(row["answer_replay_audit"]["citation_grounded_to_evidence"] is True for row in rows)


def test_v4711_unavailable_or_disabled_local_llm_fails_closed_without_fake_answers() -> None:
    from ai.eval import rag_v4711_actual_llm_answer_replay_and_silver_diagnostic_smoke as v4711

    disabled = v4711.build_report(
        root=ROOT,
        execute=False,
        sync_surfaces=False,
        env={},
        generated_at="2026-05-30T00:00:00Z",
    )
    unavailable = v4711.build_report(
        root=ROOT,
        execute=True,
        sync_surfaces=False,
        env={
            "RAG_V4_7_11_ENABLE_LOCAL_LLM_REPLAY": "1",
            "RAG_V4_7_11_LOCAL_LLM_BASE_URL": "http://127.0.0.1:9/v1",
        },
        llm_timeout_seconds=1,
        generated_at="2026-05-30T00:00:00Z",
    )

    for report, expected_status, expected_counter in (
        (disabled, "LOCAL_LLM_REPLAY_DISABLED_FAIL_CLOSED", "local_llm_replay_disabled_fail_closed_count"),
        (unavailable, "LOCAL_LLM_UNAVAILABLE_FAIL_CLOSED", "local_llm_unavailable_fail_closed_count"),
    ):
        v4711.check_report(report)
        counters = report["counters"]
        assert counters["llm_invoked_count"] == 0
        assert counters["generated_response_count"] == 0
        assert counters["parsed_final_answer_present_count"] == 0
        assert counters["raw_llm_response_present_count"] == 0
        assert counters[expected_counter] == counters["v4_7_10_answer_replay_candidate_count"]
        assert counters["noop_or_extractive_generator_used"] is False
        assert all(row["answer_replay_audit"]["status"] == expected_status for row in report["pdf_answer_replay_rows"])
        assert all(row["answer_replay_audit"]["llm_invoked"] is False for row in report["pdf_answer_replay_rows"])
        assert all(row["answer_replay_audit"]["generated_response_created"] is False for row in report["pdf_answer_replay_rows"])
        assert all(row["answer_replay_audit"]["raw_prompt_created"] is False for row in report["pdf_answer_replay_rows"])
        assert all(row["answer_replay_audit"]["raw_llm_response_created"] is False for row in report["pdf_answer_replay_rows"])
        assert all(row["answer_replay_audit"]["parsed_final_answer_present"] is False for row in report["pdf_answer_replay_rows"])
        assert all(not row.get("final_answer") for row in report["pdf_answer_replay_rows"])

    mutated = json.loads(json.dumps(disabled))
    mutated["pdf_answer_replay_rows"][0]["final_answer"] = "old extractive fallback answer"
    mutated["pdf_answer_replay_rows"][0]["rendered_citations"] = ["evidence_1"]
    try:
        v4711.check_report(mutated)
    except ValueError as exc:
        assert "fail-closed row carried answer payload" in str(exc)
    else:
        raise AssertionError("v4_7_11 check_report accepted a fail-closed answer payload")


def test_v4711_response_leakage_flags_are_hard_failures() -> None:
    from ai.eval import rag_v4711_actual_llm_answer_replay_and_silver_diagnostic_smoke as v4711

    def leaking_client(prompt: str) -> str:
        payload = json.loads(prompt)
        evidence = str(payload.get("bounded_evidence_excerpt") or "근거")
        return json.dumps(
            {
                "final_answer": f"{evidence} expected supporting source_file_title",
                "abstain": False,
                "citations": ["evidence_1"],
                "answer_plan": "leakage regression fixture",
                "unsupported_claim_risk": False,
                "evidence_underuse_flag": False,
                "context_understanding_miss": False,
                "over_abstain_candidate": False,
            },
            ensure_ascii=False,
            sort_keys=True,
        )

    report = v4711.build_report(
        root=ROOT,
        execute=True,
        sync_surfaces=False,
        env={"RAG_V4_7_11_ENABLE_LOCAL_LLM_REPLAY": "1"},
        llm_client=leaking_client,
        generated_at="2026-05-30T00:00:00Z",
        check=False,
    )
    assert report["counters"]["response_leakage_flag_count"] == 9
    try:
        v4711.check_report(report)
    except ValueError as exc:
        assert "response_leakage_flag_count" in str(exc)
    else:
        raise AssertionError("v4_7_11 check_report accepted response leakage")


def test_v4711_answer_review_packet_is_ignored_compact_and_status_has_no_raw_payloads() -> None:
    from ai.eval import rag_v4711_actual_llm_answer_replay_and_silver_diagnostic_smoke as v4711

    report = v4711.build_report(
        root=ROOT,
        execute=True,
        sync_surfaces=False,
        env={"RAG_V4_7_11_ENABLE_LOCAL_LLM_REPLAY": "1"},
        llm_client=_fake_strict_korean_answer,
        generated_at="2026-05-30T00:00:00Z",
    )
    paths = report["artifact_paths"]
    answer_packet = ROOT / paths["answer_review_packet_jsonl"]
    assert paths["report_json"] == "ai/eval/reports/rag-ingestion/runs/v4_7_11/report.json"
    assert paths["answer_review_packet_jsonl"] == (
        "ai/eval/reports/rag-ingestion/runs/v4_7_11/answer_review_packet_ko.jsonl"
    )
    for rel_path in (paths["report_json"], paths["answer_review_packet_jsonl"]):
        result = subprocess.run(["git", "check-ignore", "-q", rel_path], cwd=ROOT)
        assert result.returncode == 0, rel_path
    assert answer_packet.parent == V4_7_11_REPORT.parent

    event = v4711.status_event(report, report_sha256="0" * 64, answer_packet_sha256="1" * 64)
    event_text = json.dumps(event, ensure_ascii=False)
    assert event["answer_review_packet_row_count"] == 9
    assert event["artifact_paths"]["answer_review_packet_jsonl"] == paths["answer_review_packet_jsonl"]
    assert event["raw_llm_response_present_count"] == 0
    for forbidden in ("raw_prompt_payload", "raw_response_payload", "prompt_payload", '"final_answer":'):
        assert forbidden not in event_text


def test_v4711_silver_diagnostic_smoke_is_bounded_or_fail_closed_plan_only() -> None:
    from ai.eval import rag_v4711_actual_llm_answer_replay_and_silver_diagnostic_smoke as v4711

    report = v4711.build_report(
        root=ROOT,
        execute=True,
        sync_surfaces=False,
        env={"RAG_V4_7_11_ENABLE_LOCAL_LLM_REPLAY": "1"},
        llm_client=_fake_strict_korean_answer,
        generated_at="2026-05-30T00:00:00Z",
    )
    v4711.check_report(report)
    smoke = report["silver_diagnostic_smoke"]
    counters = report["counters"]

    assert smoke["diagnostic_silver_only"] is True
    assert smoke["official_metric_input_rows"] == 0
    assert smoke["silver_promoted_to_gold_count"] == 0
    assert counters["silver_official_metric_input_rows"] == 0
    assert counters["silver_promoted_to_gold_count"] == 0
    if smoke["executed"]:
        assert counters["silver_smoke_sample_count"] <= smoke["target_sample_count"]
        assert counters["silver_smoke_text_count"] <= 10
        assert counters["silver_smoke_pdf_count"] <= 10
        assert counters["silver_smoke_xlsx_count"] <= 10
        assert counters["silver_llm_invoked_count"] == counters["silver_smoke_sample_count"]
    else:
        assert smoke["status"].endswith("_FAIL_CLOSED")
        assert smoke["blocked_reason"]
        assert smoke["plan"]["target_sample_count"] == 30
        assert counters["silver_llm_invoked_count"] == 0
        assert counters["silver_generated_response_count"] == 0


def test_v477_manifest_classifies_every_v3_legacy_artifact_with_action_or_hold_reason() -> None:
    from ai.eval import rag_eval_registry as registry

    report = registry.load_report("v4_7_7", root=ROOT)
    manifest_path = ROOT / str(report["artifact_paths"]["v3_legacy_manifest_jsonl"])
    manifest_rows = _read_jsonl(manifest_path)

    assert report["archive_aware_eval_surface"] is True
    assert report["v3_legacy_artifact_count"] == len(manifest_rows)
    assert report["v3_legacy_unclassified_count"] == 0
    assert report["v3_legacy_artifact_count"] == (
        report["v3_legacy_archived_or_removed_count"]
        + report["v3_legacy_deleted_count"]
        + report["v3_legacy_manual_hold_count"]
    )
    assert report["archive_copy_failed_count"] == 0
    assert report["hash_verification_failed_count"] == 0
    assert report["v3_legacy_archived_or_removed_count"] > 0
    assert report["v3_legacy_manual_hold_count"] > 0
    assert sum(report["v3_legacy_hold_counts_by_classification"].values()) == report["v3_legacy_manual_hold_count"]
    assert report["v3_legacy_hold_counts_by_classification"]["EXPLICIT_HOLD_CURRENT_TEST_OR_DOC_CONTRACT"] > 0
    assert report["v3_legacy_hold_counts_by_classification"]["EXPLICIT_HOLD_AMBIGUOUS_GENERATED_SURFACE"] > 0

    seen: set[tuple[str, str]] = set()
    classifications = {row["classification"] for row in manifest_rows}
    assert "EXTERNALLY_ARCHIVED_REMOVED" in classifications
    assert "EXPLICIT_HOLD_CURRENT_TEST_OR_DOC_CONTRACT" in classifications
    for row in manifest_rows:
        key = (str(row["original_relative_path"]), str(row["classification"]))
        assert key not in seen
        seen.add(key)
        assert str(row["original_relative_path"]).startswith("ai/eval/reports/rag-ingestion/")
        assert ".." not in Path(str(row["original_relative_path"])).parts
        assert "D:\\" not in json.dumps(row, ensure_ascii=False)
        if row["classification"].startswith("EXPLICIT_HOLD"):
            assert row["hold_reason"]
        if row["classification"] in {"EXTERNALLY_ARCHIVED_REMOVED", "ARCHIVE_THEN_REMOVE"}:
            assert row["sha256"]


def test_v477_status_docs_and_script_consolidation_stay_closed_and_compact() -> None:
    report = _read_json(V4_7_7_REPORT)
    latest = next(
        event
        for event in reversed(_read_jsonl(STATUS_JSONL))
        if event.get("short_run_id") == V4_7_7_SHORT_RUN_ID
    )
    progress = PROGRESS_DOC.read_text(encoding="utf-8")
    measurements = MEASUREMENTS_DOC.read_text(encoding="utf-8")
    triage = TRIAGE_DOC.read_text(encoding="utf-8")

    assert report["short_run_id"] == V4_7_7_SHORT_RUN_ID
    assert report["status"] == V4_7_7_STATUS
    assert report["diagnostic_only"] is True
    assert report["non_production"] is True
    assert report["cleanup_only"] is True
    for key in (
        "official_metric",
        "gold_mutation",
        "qrels_mutation",
        "label_mutation",
        "expected_answer_mutation",
        "supporting_evidence_mutation",
        "denominator_mutation",
        "training_dataset_created",
        "ft_a_execution",
        "fine_tuning",
        "promotion_evidence",
        "product_success_evidence_allowed",
        "live_db_index_cache_readiness",
    ):
        assert report[key] is False, key
    assert report["official_metric_input_rows"] == 0
    assert report["protected_namespaces_touched"] == []
    assert report["script_consolidation"]["stable_runner"] == "ai/scripts/rag_eval.py"
    assert "v3_22" in report["script_consolidation"]["safe_check_aliases"]
    assert "v3_16" in report["script_consolidation"]["held_legacy_entrypoints"]
    assert not list((ROOT / "ai" / "scripts").glob("rag_v4_7_7*.py"))

    doc_text = "\n".join((progress, measurements, triage))
    if V4_7_7_SHORT_RUN_ID in doc_text:
        assert "official_metric=false" in doc_text or "official metric" in doc_text

    generated_text = "\n".join(
        [
            json.dumps(report, ensure_ascii=False),
            json.dumps(latest, ensure_ascii=False),
            _optional_doc_section(measurements, "### v4_7_7"),
            _optional_doc_section(triage, "### v4_7_7"),
        ]
    )
    for pattern in (
        r"\b[A-Z]:[\\/]",
        r"source_identity_key",
        r"prompt_payload",
        r"raw_response_payload",
        r"fine[-_ ]?tuned",
        r"promotion-ready",
    ):
        assert not re.search(pattern, generated_text), pattern


def test_v478_report_manifest_status_docs_and_alias_expansion_stay_closed() -> None:
    report = _read_json(V4_7_8_REPORT)
    latest = next(
        event
        for event in reversed(_read_jsonl(STATUS_JSONL))
        if event.get("short_run_id") == V4_7_8_SHORT_RUN_ID
    )
    manifest_rows = _read_jsonl(ROOT / report["artifact_paths"]["v3_legacy_hold_reduction_manifest_jsonl"])
    progress = PROGRESS_DOC.read_text(encoding="utf-8")
    measurements = MEASUREMENTS_DOC.read_text(encoding="utf-8")
    triage = TRIAGE_DOC.read_text(encoding="utf-8")

    assert report["short_run_id"] == V4_7_8_SHORT_RUN_ID
    assert report["canonical_long_run_id"] == V4_7_8_LONG_RUN_ID
    assert report["status"] == V4_7_8_STATUS
    assert report["diagnostic_only"] is True
    assert report["cleanup_only"] is True
    for key in (
        "official_metric",
        "gold_mutation",
        "qrels_mutation",
        "label_mutation",
        "expected_answer_mutation",
        "supporting_evidence_mutation",
        "denominator_mutation",
        "training_dataset_created",
        "ft_a_execution",
        "fine_tuning",
        "promotion_evidence",
        "product_success_evidence_allowed",
        "live_db_index_cache_readiness",
    ):
        assert report[key] is False, key
    assert report["official_metric_input_rows"] == 0
    assert report["protected_namespaces_touched"] == []

    assert report["v3_legacy_artifact_count"] == len(manifest_rows)
    assert report["unclassified_count"] == 0
    assert report["archive_copy_failed_count"] == 0
    assert report["hash_verification_failed_count"] == 0
    assert report["after_hold_counts_by_classification"]["EXPLICIT_HOLD_CURRENT_TEST_OR_DOC_CONTRACT"] <= 80
    assert report["after_hold_counts_by_classification"].get("EXPLICIT_HOLD_AMBIGUOUS_GENERATED_SURFACE", 0) <= 20
    assert report["v3_legacy_manual_hold_count"] <= 120
    assert report["documented_review_packet_hold_count_before"] == 16
    assert report["documented_review_packet_hold_count_after"] == 16
    assert report["resolved_current_test_or_doc_contract_count"] > 0
    assert report["resolved_ambiguous_generated_surface_count"] > 0
    assert report["archived_count"] == report["removed_count"]

    aliases = set(report["script_consolidation"]["safe_check_aliases"])
    assert {"v3_18", "v3_19", "v3_20", "v3_21", "v3_22"}.issubset(aliases)
    assert "v3_16" in report["script_consolidation"]["held_legacy_entrypoints"]
    assert "v3_17" in report["script_consolidation"]["held_legacy_entrypoints"]
    assert report["safe_runner_check_alias_count_before"] == 2
    assert report["safe_runner_check_alias_count_after"] >= 5
    assert not list((ROOT / "ai" / "scripts").glob("rag_v4_7_8*.py"))

    assert latest["short_run_id"] == V4_7_8_SHORT_RUN_ID
    assert latest["artifact_paths"]["report_json"] == "ai/eval/reports/rag-ingestion/runs/v4_7_8/report.json"
    assert latest["artifact_sha256"]["report_json_sha256"] == _sha256_file(V4_7_8_REPORT)
    assert latest["official_metric"] is False
    assert latest["official_metric_input_rows"] == 0

    doc_text = "\n".join((progress, measurements, triage))
    if V4_7_8_SHORT_RUN_ID in doc_text:
        assert "official_metric=false" in doc_text or "official metric" in doc_text

    generated_text = "\n".join(
        [
            json.dumps(report, ensure_ascii=False),
            json.dumps(latest, ensure_ascii=False),
            _optional_doc_section(measurements, "### v4_7_8"),
            _optional_doc_section(triage, "### v4_7_8"),
        ]
    )
    for pattern in (
        r"\b[A-Z]:[\\/]",
        r"prompt_payload",
        r"raw_response_payload",
        r"promotion-ready",
    ):
        assert not re.search(pattern, generated_text), pattern


def test_v479_pdf_residual_replay_repairs_only_residual_rows_and_fails_closed_without_llm() -> None:
    report = _read_json(V4_7_9_REPORT)
    latest = next(
        row
        for row in reversed(_read_jsonl(STATUS_JSONL))
        if row.get("short_run_id") == V4_7_9_SHORT_RUN_ID
    )
    progress = PROGRESS_DOC.read_text(encoding="utf-8")
    measurements = MEASUREMENTS_DOC.read_text(encoding="utf-8")
    triage = TRIAGE_DOC.read_text(encoding="utf-8")

    assert report["short_run_id"] == V4_7_9_SHORT_RUN_ID
    assert report["canonical_long_run_id"] == V4_7_9_LONG_RUN_ID
    assert report["status"] == V4_7_9_STATUS
    assert report["source_run_id"] == "v4_7_5_pdf_evidence_repair_eval_compaction"
    assert report["prior_replay_run_id"] == "official_answer_citation_agentic_loop_run_v4_7_4_pdf_survivor_retrieval_evidence_answer_quality_replay_nonprod"
    assert report["diagnostic_only"] is True
    assert report["non_production"] is True
    for key in (
        "official_metric",
        "gold_mutation",
        "qrels_mutation",
        "label_mutation",
        "expected_answer_mutation",
        "supporting_evidence_mutation",
        "denominator_mutation",
        "training_dataset_created",
        "ft_a_execution",
        "fine_tuning",
        "promotion_evidence",
        "product_success_evidence_allowed",
        "live_db_index_cache_readiness",
    ):
        assert report[key] is False, key
    assert report["official_metric_input_rows"] == 0
    assert report["protected_namespaces_touched"] == []

    counters = report["counters"]
    assert counters["pdf_survivor_row_count"] == 58
    assert counters["prior_answer_ready_evidence_bundle_count"] == 48
    assert counters["residual_weak_evidence_window_count_before"] == 10
    assert counters["residual_weak_evidence_window_count_after"] == 3
    assert counters["missing_neighbor_context_count_before"] == 10
    assert counters["missing_neighbor_context_count_after"] == counters["residual_weak_evidence_window_count_after"]
    assert counters["repaired_evidence_bundle_count"] == 7
    assert counters["answer_replay_candidate_count"] == 7
    assert counters["repaired_evidence_bundle_count"] == counters["answer_replay_candidate_count"]
    assert counters["llm_invoked_count"] == 0
    assert counters["local_llm_unavailable_fail_closed_count"] == counters["answer_replay_candidate_count"]
    assert counters["generated_response_count"] == 0
    assert counters["parsed_final_answer_present_count"] == 0
    assert counters["citation_rendered_count"] == 0
    assert counters["claim_support_verifier_fail_count"] == 0
    assert counters["unsupported_claim_risk_count"] == 0
    assert counters["regression_count_for_prior_answer_ready_rows"] == 0
    assert counters["official_metric_input_rows"] == 0
    assert counters["protected_namespaces_touched"] == []

    rows = report["pdf_residual_replay_rows"]
    assert len(rows) == 58
    prior_ready_rows = [row for row in rows if row["prior_answer_ready_evidence_bundle"]]
    residual_rows = [row for row in rows if row["v4_7_5_residual_weak_evidence_window"]]
    repaired_rows = [row for row in rows if row["v4_7_9_repair_applied"]]
    dropped_residual_rows = [row for row in residual_rows if not row["v4_7_9_repair_applied"]]
    assert len(prior_ready_rows) == 48
    assert len(residual_rows) == 10
    assert len(repaired_rows) == counters["repaired_evidence_bundle_count"]
    assert [row["query_id"] for row in dropped_residual_rows] == [
        "v4_7_pdf_query_04_03",
        "v4_7_pdf_query_04_04",
        "v4_7_pdf_query_04_05",
    ]
    assert all(row["repair_audit"]["decision"] == "dropped" for row in dropped_residual_rows)
    assert all(row["answer_replay_audit"]["answer_replay_candidate"] is False for row in dropped_residual_rows)
    assert all(row["answer_ready_evidence_bundle"] for row in prior_ready_rows)
    assert all(row["SourceAtom_EvidenceBundle_role"] == "evidence_truth" for row in rows)
    assert all(row["SearchView_vector_payload_role"] == "candidate_only" for row in rows)
    assert all(row["raw_pdf_query_time_parsing"] is False for row in rows)
    assert all(row["hidden_target_locator_used"] is False for row in rows)
    assert all(row["expected_or_supporting_gold_text_used"] is False for row in rows)
    assert all(row["source_file_title_shortcut_used"] is False for row in rows)
    assert all(row["direct_answer_value_matching_used"] is False for row in rows)
    assert all(row["full_page_dump_used"] is False for row in rows)
    assert all(row["preserved_locator_metadata"]["page_candidate"] == row["page_candidate"] for row in rows)
    assert all(row["repair_audit"]["decision"] in {"protected_no_regression", "repaired", "dropped"} for row in rows)
    assert all(
        row["answer_replay_audit"]["status"] == "LOCAL_LLM_UNAVAILABLE_FAIL_CLOSED"
        for row in repaired_rows
    )
    assert all(row["answer_replay_audit"]["raw_llm_response_created"] is False for row in repaired_rows)
    assert all(row["answer_replay_audit"]["claim_support_verifier_status"] == "not_run_local_llm_unavailable" for row in repaired_rows)

    assert latest["short_run_id"] == V4_7_9_SHORT_RUN_ID
    assert latest["artifact_paths"]["report_json"] == "ai/eval/reports/rag-ingestion/runs/v4_7_9/report.json"
    assert latest["artifact_sha256"]["report_json_sha256"] == _sha256_file(V4_7_9_REPORT)
    assert latest["official_metric"] is False
    assert latest["official_metric_input_rows"] == 0
    assert latest["local_llm_available"] is False
    assert latest["local_llm_unavailable_fail_closed_count"] == counters["local_llm_unavailable_fail_closed_count"]

    doc_text = "\n".join((progress, measurements, triage))
    if V4_7_9_SHORT_RUN_ID in doc_text:
        assert "official_metric=false" in doc_text or "official metric" in doc_text
    if "### v4_7_9" in measurements:
        assert "LOCAL_LLM_UNAVAILABLE_FAIL_CLOSED" in measurements
    if "### v4_7_9" in triage:
        assert "SourceAtom/EvidenceBundle remains evidence truth" in triage

    generated_text = "\n".join(
        [
            json.dumps(report, ensure_ascii=False),
            json.dumps(latest, ensure_ascii=False),
            _optional_doc_section(measurements, "### v4_7_9"),
            _optional_doc_section(triage, "### v4_7_9"),
        ]
    )
    for pattern in (
        r"\b[A-Z]:[\\/]",
        r"prompt_payload",
        r"raw_response_payload",
        r"promotion-ready",
    ):
        assert not re.search(pattern, generated_text), pattern


def test_v4710_korean_evidence_normalization_repairs_compacted_sourceatom_spans_without_gold() -> None:
    from ai.eval import rag_v4710_pdf_korean_evidence_normalization_and_answer_replay_readiness as v4710

    decision = v4710.korean_normalized_repair_decision(
        query_text="조달사업의 환경 변화에 따른 조달특별회계의 중장기적 운영 개선을 위한 수수료 체계 확립에 대해 어떤 내용이 제시되어 있습니까?",
        evidence_text="조달사업의환경변화에따른조달특별회계중장기적운영개선을위한수수료체계확립및",
        inherited_overlap=0,
    )

    assert decision["decision"] == "repaired"
    assert decision["reason"] == "spacing_insensitive_korean_query_evidence_overlap"
    assert decision["normalization_scope"] == "query_text_and_existing_sourceatom_span_only"
    assert decision["query_evidence_token_overlap_count"] >= 4
    assert decision["spacing_insensitive_korean_overlap_count"] >= 4
    assert decision["normalization_applied"] is True
    assert decision["source_text_added_from_raw_pdf"] is False
    assert decision["expected_or_supporting_gold_text_used"] is False
    assert decision["direct_answer_value_matching_used"] is False


def test_v4710_korean_evidence_normalization_drops_empty_short_or_unrelated_existing_spans() -> None:
    from ai.eval import rag_v4710_pdf_korean_evidence_normalization_and_answer_replay_readiness as v4710

    for evidence_text in (None, "", "조달사업", "친환경비료공장운영개선정책참고자료로활용"):
        decision = v4710.korean_normalized_repair_decision(
            query_text="조달특별회계 수수료 운영개선 방안은 무엇인가요?",
            evidence_text="" if evidence_text is None else evidence_text,
            inherited_overlap=0,
        )

        assert decision["decision"] == "dropped", evidence_text
        assert decision["raw_pdf_query_time_parsing"] is False
        assert decision["expected_or_supporting_gold_text_used"] is False
        assert decision["direct_answer_value_matching_used"] is False

    numeric_only = v4710.korean_normalized_repair_decision(
        query_text="2020년 조달특별회계 수수료 운영개선 방안은 무엇인가요?",
        evidence_text="2020년 친환경비료 공장 운영 개선 정책 참고자료로 활용되었습니다.",
        inherited_overlap=0,
    )

    assert numeric_only["decision"] == "dropped"
    assert numeric_only["numeric_overlap_count"] == 1
    assert numeric_only["spacing_insensitive_korean_overlap_count"] < 4


def test_v4710_pdf_korean_evidence_normalization_reduces_v479_residuals_and_records_replay_readiness() -> None:
    report = _load_v4710_report()
    status_rows = _read_jsonl(STATUS_JSONL) if STATUS_JSONL.exists() else []
    latest = next(
        (
            row
            for row in reversed(status_rows)
            if row.get("short_run_id") == V4_7_10_SHORT_RUN_ID
        ),
        None,
    )
    progress = PROGRESS_DOC.read_text(encoding="utf-8")
    measurements = MEASUREMENTS_DOC.read_text(encoding="utf-8")
    triage = TRIAGE_DOC.read_text(encoding="utf-8")

    assert report["short_run_id"] == V4_7_10_SHORT_RUN_ID
    assert report["canonical_long_run_id"] == V4_7_10_LONG_RUN_ID
    assert report["status"] == V4_7_10_STATUS
    assert report["source_run_id"] == V4_7_9_SHORT_RUN_ID
    assert report["source_report_json"] == "ai/eval/reports/rag-ingestion/runs/v4_7_9/report.json"
    assert report["diagnostic_only"] is True
    assert report["non_production"] is True
    for key in (
        "official_metric",
        "gold_mutation",
        "qrels_mutation",
        "label_mutation",
        "expected_answer_mutation",
        "supporting_evidence_mutation",
        "denominator_mutation",
        "training_dataset_created",
        "ft_a_execution",
        "fine_tuning",
        "promotion_evidence",
        "product_success_evidence_allowed",
        "live_db_index_cache_readiness",
    ):
        assert report[key] is False, key
    assert report["official_metric_input_rows"] == 0
    assert report["protected_namespaces_touched"] == []

    counters = report["counters"]
    assert counters["pdf_survivor_row_count"] == 58
    assert counters["answer_ready_evidence_bundle_count_before"] == 55
    assert counters["answer_ready_evidence_bundle_count"] == 57
    assert counters["v4_7_9_residual_weak_evidence_window_count_before"] == 3
    assert counters["residual_weak_evidence_window_count_before"] == 3
    assert counters["residual_weak_evidence_window_count_after"] == 1
    assert counters["missing_neighbor_context_count_after"] == counters["residual_weak_evidence_window_count_after"]
    assert counters["korean_normalization_repair_count"] == 2
    assert counters["korean_normalized_evidence_repair_count"] == 2
    assert counters["newly_repaired_evidence_bundle_count"] == counters["korean_normalization_repair_count"]
    assert counters["new_answer_replay_ready_count"] == 2
    assert counters["answer_replay_ready_count"] == 9
    assert counters["regression_count_for_v4_7_9_answer_ready_rows"] == 0
    assert counters["llm_invoked_count"] == 0
    assert counters["generated_response_count"] == 0
    assert counters["local_llm_unavailable_fail_closed_count"] == counters["answer_replay_candidate_count"]
    assert counters["official_metric_input_rows"] == 0
    assert counters["protected_namespaces_touched"] == []

    rows = report["pdf_residual_replay_rows"]
    assert len(rows) == 58
    targeted_rows = [row for row in rows if row["v4_7_10_repair_targeted"]]
    repaired_rows = [row for row in rows if row["v4_7_10_repair_applied"]]
    remaining_rows = [row for row in rows if row["weak_evidence_window"]]
    assert len(targeted_rows) == 3
    assert [row["query_id"] for row in targeted_rows] == [
        "v4_7_pdf_query_04_03",
        "v4_7_pdf_query_04_04",
        "v4_7_pdf_query_04_05",
    ]
    assert [row["query_id"] for row in repaired_rows] == [
        "v4_7_pdf_query_04_04",
        "v4_7_pdf_query_04_05",
    ]
    assert [row["query_id"] for row in remaining_rows] == ["v4_7_pdf_query_04_03"]
    assert len(repaired_rows) == counters["korean_normalization_repair_count"]
    assert len(remaining_rows) == counters["residual_weak_evidence_window_count_after"]
    assert all(row["SourceAtom_EvidenceBundle_role"] == "evidence_truth" for row in rows)
    assert all(row["SearchView_vector_payload_role"] == "candidate_only" for row in rows)
    assert all(row["raw_pdf_query_time_parsing"] is False for row in rows)
    assert all(row["hidden_target_locator_used"] is False for row in rows)
    assert all(row["expected_or_supporting_gold_text_used"] is False for row in rows)
    assert all(row["source_file_title_shortcut_used"] is False for row in rows)
    assert all(row["direct_answer_value_matching_used"] is False for row in rows)
    assert all(row["full_page_dump_used"] is False for row in rows)
    assert all(
        row["repair_audit"]["reason"] == "spacing_insensitive_korean_query_evidence_overlap"
        for row in repaired_rows
    )
    assert all(
        row["repair_audit"]["spacing_insensitive_korean_overlap_count"] >= 2
        for row in repaired_rows
    )
    assert all(
        row["answer_replay_audit"]["status"] == "LOCAL_LLM_UNAVAILABLE_FAIL_CLOSED"
        for row in rows
        if row["answer_replay_audit"]["answer_replay_candidate"]
    )
    assert report["remaining_residual_rows"] == [
        {
            "row_index_1based": row["row_index_1based"],
            "candidate_id_hash": row["candidate_id_hash"],
            "query_id_hash": row["query_id_hash"],
            "reason": row["repair_audit"]["reason"],
        }
        for row in remaining_rows
    ]

    if latest is not None:
        assert latest["short_run_id"] == V4_7_10_SHORT_RUN_ID
        assert latest["artifact_paths"]["report_json"] == "ai/eval/reports/rag-ingestion/runs/v4_7_10/report.json"
        if V4_7_10_REPORT.exists():
            assert latest["artifact_sha256"]["report_json_sha256"] == _sha256_file(V4_7_10_REPORT)
        assert latest["official_metric"] is False
        assert latest["official_metric_input_rows"] == 0

    before_after = (
        f"weak evidence/window {counters['residual_weak_evidence_window_count_before']} -> "
        f"{counters['residual_weak_evidence_window_count_after']}"
    )
    doc_text = "\n".join((progress, measurements, triage))
    if V4_7_10_SHORT_RUN_ID in doc_text:
        assert before_after in doc_text
        assert "official_metric=false" in doc_text or "official metric" in doc_text
    if "### v4_7_10" in triage:
        assert "spacing-insensitive Korean evidence normalization" in triage
    if "### v4_7_10" in measurements:
        assert "LOCAL_LLM_UNAVAILABLE_FAIL_CLOSED" in measurements

    generated_text = "\n".join(
        [
            json.dumps(report, ensure_ascii=False),
            json.dumps(latest or {}, ensure_ascii=False),
            _optional_doc_section(measurements, "### v4_7_10"),
            _optional_doc_section(triage, "### v4_7_10"),
        ]
    )
    for pattern in (
        r"\b[A-Z]:[\\/]",
        r"prompt_payload",
        r"raw_response_payload",
        r"promotion-ready",
    ):
        assert not re.search(pattern, generated_text), pattern


def test_v4711_docs_do_not_describe_current_as_v4710() -> None:
    progress = PROGRESS_DOC.read_text(encoding="utf-8")

    assert "use resolver key `current` for v4_7_10" not in progress
    assert "current` for v4_7_10" not in progress
    assert "`current` resolves to `v4_7_10`" not in progress


def test_v477_stable_runner_executes_safe_legacy_check_aliases() -> None:
    for alias, expected_status in (
        ("v3_18", "DIAGNOSTIC_V3_18_AGENT_RUNTIME_TOOL_INVOCATION_CONTRACT_NONPROD_READY"),
        ("v3_19", "DIAGNOSTIC_V3_19_LOCATOR_AMBIGUITY_DEICTIC_RESPONSE_POLICY_NONPROD_READY"),
        ("v3_20", "DIAGNOSTIC_V3_20_LIVE_RUNTIME_LIKE_DB_INDEX_CACHE_SMOKE_NONPROD_READY"),
        ("v3_21", "DIAGNOSTIC_V3_21_AGENT_RUNTIME_LLM_IO_OBSERVABILITY_PACKET_NONPROD_READY"),
        ("v3_22", "DIAGNOSTIC_V3_22_XLSX_VALUE_FORMATTING_AND_CELL_RANGE_ANSWER_RENDERING_NONPROD_READY"),
    ):
        result = subprocess.run(
            [sys.executable, "-X", "utf8", "ai/scripts/rag_eval.py", alias, "--check"],
            cwd=ROOT,
            text=True,
            capture_output=True,
            check=True,
            timeout=180,
        )
        payload = json.loads(result.stdout)
        assert payload["run_key"] == alias
        assert payload["status"] == expected_status
        assert payload["safe_legacy_alias"] == alias
        assert payload["write_supported"] is False
        assert payload["official_metric_input_rows"] == 0
